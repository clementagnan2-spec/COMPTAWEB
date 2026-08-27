"""
main.py — Application de comptabilité SYSCOHADA autonome (Tkinter).

Navigation par menu (SAISIE, COMMERCE, PRODUCTION, ENGAGEMENTS-PROJETS,
ÉTATS ET RAPPORTS) : un seul panneau de contenu, qui change selon le menu
choisi. Les données sont stockées localement dans un fichier SQLite
(%LOCALAPPDATA%\\SaisieComptable\\comptabilite.db sous Windows).
"""
import tkinter as tk
import os
import sys
import subprocess
from tkinter import ttk, messagebox, filedialog, simpledialog
from datetime import date, datetime

import core


def _ouvrir_fichier(path):
    """Ouvre un fichier avec l'application par défaut du système
    d'exploitation (Excel pour un .xls/.xlsx, en général) — multiplateforme."""
    try:
        if os.name == "nt":
            os.startfile(path)
        else:
            import subprocess
            import sys
            opener = "open" if sys.platform == "darwin" else "xdg-open"
            subprocess.Popen([opener, path])
        return True
    except Exception:
        return False


def fmt_cfa(v):
    """Formate un montant façon rapport financier SYSCOHADA (espace comme
    séparateur de milliers, pas de décimales — les francs CFA n'ont pas de
    subdivision usuelle en comptabilité) : 10100000000 -> « 10 100 000 000 »."""
    if v is None:
        return ""
    return f"{v:,.0f}".replace(",", " ")


def export_etat_gabarit(parent, conn, etat_id, nom_fichier, titre):
    """Exporte un état financier (Bilan/Résultat/Situation/TFT) dans son
    gabarit officiel (voir core.generate_etat_xlsx) — commun aux 4 onglets
    du menu ÉTATS ET RAPPORTS concernés."""
    path = filedialog.asksaveasfilename(
        defaultextension=".xlsx", filetypes=[("Classeur Excel", "*.xlsx")],
        initialfile=nom_fichier, title=f"Exporter {titre} (gabarit officiel)",
    )
    if not path:
        return
    try:
        rapport = core.generate_etat_xlsx(conn, etat_id, path)
    except Exception as exc:
        messagebox.showerror("Erreur", f"Échec de l'export : {exc}", parent=parent)
        return
    msg = f"{titre} exporté dans le gabarit officiel :\n{path}\n\n{rapport['cells_ok']} cellule(s) calculée(s)."
    if rapport["cells_error"]:
        msg += f"\n⚠ {len(rapport['cells_error'])} cellule(s) en erreur (affichées « #ERREUR » dans le fichier)."
    messagebox.showinfo("Export terminé", msg, parent=parent)


class LoginDialog(tk.Toplevel):
    """Écran de connexion obligatoire au démarrage de l'application, DÈS
    QU'AU MOINS UN UTILISATEUR EXISTE dans la base (menu ADMIN >
    Utilisateurs) — voir App.__init__. Tant qu'aucun utilisateur n'a
    encore été créé, l'application démarre normalement sans connexion
    (mode amorçage), pour ne jamais bloquer l'accès à sa propre
    installation."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.utilisateur = None  # rempli si connexion réussie
        self.title("Connexion")
        self.geometry("380x260")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        self.protocol("WM_DELETE_WINDOW", self._annuler)

        frame = ttk.Frame(self, padding=20)
        frame.pack(fill="both", expand=True)
        ttk.Label(frame, text="Connexion requise", font=("Segoe UI", 13, "bold")).pack(anchor="w", pady=(0, 12))

        ttk.Label(frame, text="Identifiant :").pack(anchor="w")
        self.user_var = tk.StringVar()
        user_entry = ttk.Entry(frame, textvariable=self.user_var, width=30)
        user_entry.pack(fill="x", pady=(2, 10))
        user_entry.focus_set()

        ttk.Label(frame, text="Mot de passe :").pack(anchor="w")
        self.pwd_var = tk.StringVar()
        pwd_entry = ttk.Entry(frame, textvariable=self.pwd_var, width=30, show="•")
        pwd_entry.pack(fill="x", pady=(2, 10))
        pwd_entry.bind("<Return>", lambda e: self._connecter())

        self.status_var = tk.StringVar()
        ttk.Label(frame, textvariable=self.status_var, foreground="#B00020", wraplength=330).pack(
            anchor="w", pady=(0, 8))

        ttk.Button(frame, text="Se connecter", command=self._connecter).pack(anchor="w")

        self.wait_window(self)

    def _connecter(self):
        nom_utilisateur = self.user_var.get().strip()
        mot_de_passe = self.pwd_var.get()
        if not nom_utilisateur or not mot_de_passe:
            self.status_var.set("Identifiant et mot de passe obligatoires.")
            return
        utilisateur = core.verify_password(self.conn, nom_utilisateur, mot_de_passe)
        if not utilisateur:
            self.status_var.set("Identifiant ou mot de passe incorrect.")
            return
        self.utilisateur = utilisateur
        self.destroy()

    def _annuler(self):
        self.utilisateur = None
        self.destroy()


class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("PLATEFORME INTEGREE DE GESTION")
        try:
            icon_path = core.get_app_icon_path()
            self.iconbitmap(icon_path)
        except Exception:
            pass  # icône facultative — ne doit jamais empêcher le démarrage de l'application
        self.geometry("1400x820")
        try:
            self.state("zoomed")  # démarre maximisée (Windows) — plus de place pour les Bilan/Balance denses en chiffres
        except tk.TclError:
            pass
        self.conn = core.get_connection()

        # ---- Connexion obligatoire DÈS QU'AU MOINS UN UTILISATEUR EXISTE
        # (menu ADMIN > Utilisateurs) — sinon démarrage libre (amorçage),
        # pour ne jamais verrouiller l'accès à une installation neuve. ----
        self.utilisateur_connecte = None
        self.niveau_acces_connecte = "Administrateur"  # par défaut tant qu'aucune connexion n'est requise
        if core.list_utilisateurs(self.conn):
            login = LoginDialog(self, self.conn)
            if not login.utilisateur:
                self.destroy()
                return
            self.utilisateur_connecte = login.utilisateur["nom_utilisateur"]
            self.niveau_acces_connecte = login.utilisateur["niveau_acces"]

        # ---- Barre d'exercice comptable (toujours visible, en haut) ----
        top_bar = ttk.Frame(self, relief="raised", padding=4)
        top_bar.pack(fill="x", side="top")
        ttk.Label(top_bar, text="EXERCICE COMPTABLE :", font=("Segoe UI", 9, "bold")).pack(side="left", padx=(6, 4))
        self.exercice_var = tk.StringVar(value=core.get_current_exercice(self.conn))
        self.exercice_combo = ttk.Combobox(top_bar, textvariable=self.exercice_var, width=10, state="readonly")
        self.exercice_combo.pack(side="left", padx=4)
        self.exercice_combo.bind("<<ComboboxSelected>>", self._on_exercice_changed)
        ttk.Button(top_bar, text="+ Nouvel exercice", command=self._new_exercice).pack(side="left", padx=8)
        self.exercice_status_var = tk.StringVar()
        ttk.Label(top_bar, textvariable=self.exercice_status_var, foreground="#B00020",
                  font=("Segoe UI", 9, "bold")).pack(side="left", padx=12)
        if self.utilisateur_connecte:
            ttk.Label(top_bar, text=f"Connecté : {self.utilisateur_connecte} ({self.niveau_acces_connecte})",
                      foreground="#1F4E78", font=("Segoe UI", 9, "bold")).pack(side="right", padx=12)
        self._refresh_exercice_list()

        self.content = ttk.Frame(self)
        self.content.pack(fill="both", expand=True)
        self.content.grid_rowconfigure(0, weight=1)
        self.content.grid_columnconfigure(0, weight=1)

        self.pages = {}

        def register(key, cls, *args):
            w = cls(self.content, self.conn, *args)
            w.grid(row=0, column=0, sticky="nsew")
            self.pages[key] = w
            return w

        # ---- Instanciation de toutes les pages (une seule fois) ----
        register("saisie", SaisieTab)
        register("ouverture", OpeningBalancesTab)
        register("exercices", ExercicesTab, self)
        register("plan_comptable", PlanComptableTab)
        register("plan_analytique", PlanAnalytiqueTab)
        register("taux_tva", TauxTVATab)
        register("taux_retenue", TauxRetenueTab)
        register("admin_factures", AdminFacturesTab)
        register("admin_modele_bon_commande", AdminModeleBonCommandeTab)
        register("synchronisation", SynchronisationTab)
        register("niveaux_acces", NiveauxAccesTab)
        register("utilisateurs", UtilisateursTab)
        register("reinitialisation", ReinitialisationTab)
        register("plan_budgetaire", PlanBudgetaireTab)
        register("plan_bailleur", PlanBailleurTab)
        register("stocks", StocksTab)
        register("production", ProductionTab)
        register("grand_livre", GrandLivreTab)
        register("balance", BalanceTab)
        register("bilan_syscohada", BilanSyscohadaTab)
        register("compte_resultat_sig", EtatFormuleTab, "Compte de résultat (SIG)", core._cr_template_path)
        register("tft", EtatFormuleTab, "Tableau des flux de trésorerie (TFT)", core._tft_template_path)
        register("situation_financiere", EtatFormuleTab, "Situation financière (FR-BFR-TN)",
                 core._situation_template_path)
        register("arrete_comptes", ArreteComptesTab)
        register("grh_personnel", PersonnelTab)
        register("grh_time_sheet", TimeSheetTab)
        register("grh_kpi", KpiTab)
        register("grh_tableau_bord", TableauBordGrhTab)
        register("grh_hs", HsTab)
        register("grh_paie", PaieTab)
        register("tresorerie", TresorerieTab)
        register("transport", ParcAutoTab)
        register("missions", MissionsTab)
        register("pieces_rechange", PiecesRechangeTab)
        register("reparations", ReparationsTab)
        register("immobilisations", ImmobilisationsTab)
        register("amortissements", AmortissementsTab)
        register("rapports_technique", PlaceholderTab, "Rapports technique",
                 "À définir — dites-moi quels rapports techniques vous voulez ici et je construis l'écran.")
        register("clients", ClientsTab)
        register("recouvrement", RecouvrementTab)
        register("facturation", FacturationTab)
        register("marges", MargesTab)
        register("fournisseurs", FournisseursTab)
        register("contrats", ContratsTab)
        register("expression_besoin", ExpressionBesoinTab)
        register("ep_bon_commande", BonCommandeEPTab)
        register("bordereau_livraison", BordereauLivraisonTab)
        register("reglements", ReglementTab)
        register("energie", AnalytiquePeriodeTab,
                 "Énergie", "Coûts d'énergie (eau, électricité, essence, gasoil, gaz...) par code analytique, "
                            "sur une période choisie — alimentés par les écritures de Saisie taguées avec un "
                            "code « ENERGIE- » et par les lignes « Énergie » des recettes de Fabrication.",
                 core.PREFIX_ENERGIE, core.SUGGESTIONS_ENERGIE)
        register("maintenance", AnalytiquePeriodeTab,
                 "Maintenance", "Coûts de maintenance (véhicules, bâtiments, machines, informatique...) par "
                                "code analytique, sur une période choisie — alimentés par les écritures de "
                                "Saisie taguées avec un code « MAINT- » et par les lignes « Autre charge » "
                                "des recettes de Fabrication qui leur sont associées.",
                 core.PREFIX_MAINTENANCE, core.SUGGESTIONS_MAINTENANCE)

        # ---- Barre de menu — filtrée selon le niveau d'accès connecté
        # (core.get_menus_autorises) : un sous-menu non autorisé n'apparaît
        # tout simplement pas ; si AUCUN sous-menu d'un menu de premier
        # niveau n'est autorisé, ce menu entier est masqué. ----
        menubar = tk.Menu(self)
        bold = ("Segoe UI", 9, "bold")
        menus_autorises = core.get_menus_autorises(self.conn, self.niveau_acces_connecte)

        def add_top_menu(label, items):
            items_autorises = [(item_label, key) for item_label, key in items if key in menus_autorises]
            if not items_autorises:
                return
            m = tk.Menu(menubar, tearoff=0)
            for item_label, key in items_autorises:
                m.add_command(label=item_label, command=lambda k=key: self.show(k))
            menubar.add_cascade(label=label, menu=m)
            menubar.entryconfig(menubar.index("end"), font=bold)

        add_top_menu("SAISIE", [
            ("Saisie des écritures", "saisie"),
            ("Soldes d'ouverture", "ouverture"),
        ])
        add_top_menu("COMMERCIAL", [
            ("Clients", "clients"),
            ("Recouvrement", "recouvrement"),
            ("Facturation", "facturation"),
            ("Stocks", "stocks"),
            ("Marges bénéficiaires", "marges"),
        ])
        add_top_menu("PRODUCTION", [
            ("Matières premières", "stocks"),
            ("Fabrication", "production"),
            ("Produits finis", "stocks"),
        ])
        add_top_menu("RAPPORTS FINANCIERS", [
            ("Grand livre", "grand_livre"),
            ("Balance", "balance"),
            ("Bilan SYSCOHADA", "bilan_syscohada"),
            ("Compte de résultat (SIG)", "compte_resultat_sig"),
            ("TFT", "tft"),
            ("Situation financière", "situation_financiere"),
            ("Arrêté de comptes", "arrete_comptes"),
        ])
        add_top_menu("ENGAGEMENTS-PROJETS", [
            ("Fournisseurs", "fournisseurs"),
            ("Contrats", "contrats"),
            ("Expression de besoin", "expression_besoin"),
            ("Bon de commande", "ep_bon_commande"),
            ("Bordereau de livraison", "bordereau_livraison"),
            ("Règlements", "reglements"),
        ])
        add_top_menu("GRH", [
            ("Liste du personnel", "grh_personnel"),
            ("Time sheet", "grh_time_sheet"),
            ("KPI", "grh_kpi"),
            ("Tableau de bord GRH", "grh_tableau_bord"),
            ("HS (hygiène santé)", "grh_hs"),
            ("Paie", "grh_paie"),
        ])
        add_top_menu("TRESORERIE", [
            ("Trésorerie", "tresorerie"),
        ])
        add_top_menu("TRANSPORT", [
            ("Parc auto", "transport"),
            ("Missions", "missions"),
            ("Pièces de rechange", "pieces_rechange"),
            ("Réparations", "reparations"),
        ])
        add_top_menu("IMMOBILISATIONS", [
            ("Immobilisations", "immobilisations"),
            ("Amortissements", "amortissements"),
        ])
        add_top_menu("RAPPORTS TECHNIQUES", [
            ("Rapports technique", "rapports_technique"),
        ])
        add_top_menu("MAINTENANCE-QUALITÉ", [
            ("Énergie", "energie"),
            ("Maintenance", "maintenance"),
            ("Pièces de rechange", "pieces_rechange"),
        ])
        add_top_menu("PARAMÈTRES", [
            ("Exercices comptables (clôture)", "exercices"),
            ("Plan comptable", "plan_comptable"),
            ("Plan analytique", "plan_analytique"),
            ("Plan budgétaire", "plan_budgetaire"),
            ("Plan bailleurs de fonds", "plan_bailleur"),
            ("Synchronisation", "synchronisation"),
        ])
        add_top_menu("ADMIN", [
            ("Taux de TVA", "taux_tva"),
            ("Taux de retenue à la source", "taux_retenue"),
            ("Modification des factures", "admin_factures"),
            ("Modèle de bon de commande", "admin_modele_bon_commande"),
            ("Niveaux d'accès", "niveaux_acces"),
            ("Utilisateurs", "utilisateurs"),
            ("Réinitialisation des données", "reinitialisation"),
        ])
        self.config(menu=menubar)

        self.show("saisie")

    def _refresh_exercice_list(self):
        exercices = core.list_exercices(self.conn)
        values = [e["exercice"] + (" (clôturé)" if e["cloture"] else "") for e in exercices]
        self.exercice_combo["values"] = values
        current = core.get_current_exercice(self.conn)
        match = next((v for v in values if v.startswith(current)), current)
        self.exercice_var.set(match)
        if core.is_exercice_cloture(self.conn, current):
            self.exercice_status_var.set("⚠ Cet exercice est clôturé (lecture seule).")
        else:
            self.exercice_status_var.set("")

    def _on_exercice_changed(self, event=None):
        raw = self.exercice_var.get().split(" ")[0]
        core.set_current_exercice(self.conn, raw)
        self._refresh_exercice_list()
        self.refresh_current_page()

    def _new_exercice(self):
        current = core.get_current_exercice(self.conn)
        suggestion = str(int(current) + 1)

        if not core.is_exercice_cloture(self.conn, current):
            bilan = core.compute_bilan(self.conn, exercice=current)
            if abs(bilan["ecart"]) >= 1:
                messagebox.showwarning(
                    "Bilan déséquilibré",
                    f"L'exercice {current} n'est pas équilibré (écart de {fmt_cfa(bilan['ecart'])}) — "
                    f"corrigez-le avant de le clôturer (voir l'onglet Bilan pour le diagnostic). "
                    f"Un nouvel exercice sera créé, mais SANS solde d'ouverture reporté pour l'instant.",
                )
            elif messagebox.askyesno(
                "Clôturer l'exercice en cours ?",
                f"L'exercice {current} n'est pas encore clôturé : ses soldes de clôture ne seraient donc "
                f"pas reportés comme soldes d'ouverture du nouvel exercice — c'est le cas le plus courant "
                f"d'un Bilan qui semble « vide » au démarrage d'une nouvelle année.\n\n"
                f"Clôturer {current} maintenant et reporter ses soldes sur le nouvel exercice ?"
            ):
                try:
                    suggestion = core.close_exercice(self.conn, current)
                except ValueError as exc:
                    messagebox.showerror("Erreur", str(exc))
                    return

        new_ex = simpledialog.askstring("Nouvel exercice", "Année de l'exercice (AAAA) :",
                                         initialvalue=suggestion, parent=self)
        if not new_ex:
            return
        core.set_current_exercice(self.conn, new_ex.strip())
        self._refresh_exercice_list()
        self.refresh_current_page()

    def refresh_current_page(self):
        for page in self.pages.values():
            if hasattr(page, "refresh"):
                try:
                    page.refresh()
                except Exception:
                    pass

    def show(self, key):
        page = self.pages[key]
        page.tkraise()
        if hasattr(page, "refresh"):
            page.refresh()


class MultiLigneDialog(tk.Toplevel):
    """Fenêtre de saisie multi-lignes : un nombre libre de comptes au DÉBIT
    ET un nombre libre de comptes au CRÉDIT dans la même écriture (ex.
    plusieurs charges de classe 6 réglées par plusieurs comptes de
    trésorerie) — une seule grille, chaque ligne renseignée au débit OU au
    crédit, le tout devant s'équilibrer avant l'enregistrement.
    N'entraîne pas les mouvements de stock automatiques que fait la saisie
    standard pour un achat 601x/602x avec quantité — pour un achat de
    stock, utilisez le formulaire habituel."""

    def __init__(self, parent, conn, on_saved):
        super().__init__(parent)
        self.conn = conn
        self.on_saved = on_saved
        self.lignes = []
        self.title("Saisie multi-lignes (plusieurs comptes au débit et au crédit)")
        self.geometry("1080x680")
        self.minsize(900, 500)
        self.resizable(True, True)
        self.transient(parent)
        self.grab_set()

        header = ttk.LabelFrame(self, text="Informations communes à l'écriture")
        header.pack(fill="x", padx=10, pady=8, side="top")
        ttk.Label(header, text="Date (JJ/MM/AAAA) :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(header, textvariable=self.date_var, width=14).grid(row=0, column=1, padx=4)
        ttk.Label(header, text="N° Pièce :").grid(row=0, column=2, sticky="w", padx=(16, 4))
        self.piece_var = tk.StringVar()
        ttk.Entry(header, textvariable=self.piece_var, width=16).grid(row=0, column=3, padx=4)
        ttk.Label(header, text="Journal :").grid(row=0, column=4, sticky="w", padx=(16, 4))
        self.journal_var = tk.StringVar(value="OD")
        journal_combo = ttk.Combobox(header, textvariable=self.journal_var, width=8, state="readonly",
                                      values=["AC", "VE", "OD", "BQ", "CA"])
        journal_combo.grid(row=0, column=5, padx=4)
        journal_combo.bind("<Button-1>", self._open_dropdown)
        # (Le champ « Tiers » générique a été retiré : le tiers se choisit
        # désormais LIGNE PAR LIGNE, via Fournisseur/Client, dès qu'un compte
        # de la racine 40 ou 41 est utilisé — plus fiable qu'un champ global.)

        # ---- Boutons et section stock : ancrés en bas EN PREMIER, pour
        # qu'ils restent toujours visibles même sur un petit écran — le
        # tableau de lignes (au milieu) se réduit/scrolle si besoin, jamais
        # les boutons d'enregistrement. ----
        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=10, side="bottom")
        ttk.Button(btns, text="Enregistrer l'écriture", command=self.save).pack(side="left", padx=4)
        ttk.Button(btns, text="Annuler", command=self.destroy).pack(side="left", padx=4)

        stock_frame = ttk.LabelFrame(self, text=(
            "Compte stock (optionnel) — pour une facture globale d'achat (matière + transport/douane) "
            "ou une vente groupée à plusieurs clients : regroupe le mouvement de stock en un seul"))
        stock_frame.pack(fill="x", padx=10, pady=(0, 6), side="bottom")
        ttk.Label(stock_frame, text="Compte stock :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.stock_compte_var = tk.StringVar()
        self.stock_compte_combo = ttk.Combobox(stock_frame, textvariable=self.stock_compte_var, width=28)
        self.stock_compte_combo.grid(row=0, column=1, padx=4)
        self.stock_compte_combo.bind("<KeyRelease>", self._on_stock_compte_keyrelease)
        self.stock_compte_combo.bind("<Button-1>", self._open_dropdown)
        self.stock_compte_combo["values"] = [
            f"{s['code']} — {s['label']}" for s in core.compute_stocks_detail(self.conn, prefixes=["31", "32", "33", "36"])]
        ttk.Label(stock_frame, text="Sens :").grid(row=0, column=2, sticky="w", padx=(16, 4))
        self.stock_sens_var = tk.StringVar(value="Entrée (achat)")
        stock_sens_combo = ttk.Combobox(stock_frame, textvariable=self.stock_sens_var, width=16, state="readonly",
                                         values=["Entrée (achat)", "Sortie (vente)"])
        stock_sens_combo.grid(row=0, column=3, padx=4)
        stock_sens_combo.bind("<Button-1>", self._open_dropdown)
        stock_sens_combo.bind("<<ComboboxSelected>>", self._on_stock_sens_changed)
        self.stock_qte_label = ttk.Label(stock_frame, text="Quantité reçue :")
        self.stock_qte_label.grid(row=0, column=4, sticky="w", padx=(16, 4))
        self.stock_qte_var = tk.StringVar()
        ttk.Entry(stock_frame, textvariable=self.stock_qte_var, width=12).grid(row=0, column=5, padx=4)
        self.stock_info_var = tk.StringVar(value=(
            "Entrée : le coût du stock = somme des lignes au débit (matière + frais accessoires). "
            "Laissez « Compte stock » vide pour revenir au comportement ligne par ligne."))
        ttk.Label(stock_frame, textvariable=self.stock_info_var,
                  foreground="#595959", wraplength=1000).grid(row=1, column=0, columnspan=6, sticky="w", padx=4, pady=(2, 4))

        # ---- Lignes de l'écriture : occupe tout l'espace restant, au milieu ----
        lignes_frame = ttk.LabelFrame(self, text=(
            "Lignes de l'écriture — autant de comptes que nécessaire au débit ET au crédit "
            "(renseignez Débit OU Crédit par ligne, pas les deux)"))
        lignes_frame.pack(fill="both", padx=10, pady=6, side="top")

        form = ttk.Frame(lignes_frame)
        form.pack(fill="x", padx=6, pady=4)
        ttk.Label(form, text="Compte :").grid(row=0, column=0, sticky="w")
        self.ligne_compte_var = tk.StringVar()
        self.ligne_compte_combo = ttk.Combobox(form, textvariable=self.ligne_compte_var, width=24)
        self.ligne_compte_combo.grid(row=0, column=1, padx=4)
        self.ligne_compte_combo.bind("<KeyRelease>", self._on_compte_keyrelease)
        self.ligne_compte_combo.bind("<Button-1>", self._open_dropdown)
        self.ligne_compte_combo.bind("<<ComboboxSelected>>", self._on_compte_selected)
        self._refresh_compte_values(self.ligne_compte_combo)

        ttk.Label(form, text="Libellé :").grid(row=0, column=2, sticky="w", padx=(12, 0))
        self.ligne_libelle_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_libelle_var, width=20).grid(row=0, column=3, padx=4)

        ttk.Label(form, text="Débit :").grid(row=0, column=4, sticky="w", padx=(12, 0))
        self.ligne_debit_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_debit_var, width=11).grid(row=0, column=5, padx=4)

        ttk.Label(form, text="Crédit :").grid(row=0, column=6, sticky="w", padx=(12, 0))
        self.ligne_credit_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_credit_var, width=11).grid(row=0, column=7, padx=4)

        self.ligne_qte_label = ttk.Label(form, text="Quantité :")
        self.ligne_qte_label.grid(row=1, column=0, sticky="w", pady=(4, 0))
        self.ligne_qte_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_qte_var, width=10).grid(
            row=1, column=1, padx=4, sticky="w", pady=(4, 0))

        ttk.Label(form, text="Code analytique :").grid(row=1, column=2, sticky="w", padx=(12, 0), pady=(4, 0))
        self.ligne_analytic_var = tk.StringVar()
        self.ligne_analytic_combo = ttk.Combobox(form, textvariable=self.ligne_analytic_var, width=20)
        self.ligne_analytic_combo.grid(row=1, column=3, padx=4, pady=(4, 0))
        self.ligne_analytic_combo["values"] = [f"{c['code']} — {c['label']}" for c in core.list_analytic_codes(conn)]
        self.ligne_analytic_combo.bind("<<ComboboxSelected>>", self._on_analytic_changed)
        self.ligne_analytic_combo.bind("<Button-1>", self._open_dropdown)

        self.ligne_tiers_label = ttk.Label(form, text="")
        self.ligne_tiers_label.grid(row=2, column=0, sticky="w", pady=(4, 0))
        self.ligne_tiers_var = tk.StringVar()
        self.ligne_tiers_combo = ttk.Combobox(form, textvariable=self.ligne_tiers_var, width=30)
        self.ligne_tiers_combo.grid(row=2, column=1, columnspan=3, padx=4, pady=(4, 0), sticky="w")
        self.ligne_tiers_combo.bind("<KeyRelease>", self._on_tiers_keyrelease)
        self.ligne_tiers_combo.bind("<Button-1>", self._open_dropdown)
        self._ligne_tiers_kind = None  # None / "fournisseur" / "client", selon le compte choisi

        ttk.Button(form, text="Ajouter la ligne", command=self.add_ligne).grid(
            row=1, column=6, columnspan=2, padx=4, pady=(4, 0), sticky="e")

        cols = ("compte", "libelle", "debit", "credit", "quantite", "analytique", "tiers")
        self.tree = ttk.Treeview(lignes_frame, columns=cols, show="headings", height=10)
        headers = ["Compte", "Libellé", "Débit", "Crédit", "Quantité", "Code analytique", "Tiers"]
        widths = [90, 200, 100, 100, 80, 150, 110]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        tree_scroll = ttk.Scrollbar(lignes_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll.set)
        self.tree.pack(side="left", fill="both", padx=(6, 0), pady=6)
        tree_scroll.pack(side="left", fill="y", padx=(0, 6), pady=6)
        ttk.Button(lignes_frame, text="Supprimer la ligne sélectionnée", command=self.delete_ligne).pack(
            anchor="w", padx=6, pady=(0, 6))

        self.total_var = tk.StringVar(value="Total Débit : 0     Total Crédit : 0     Écart : 0")
        self.total_label = ttk.Label(lignes_frame, textvariable=self.total_var, font=("Segoe UI", 10, "bold"))
        self.total_label.pack(anchor="w", padx=6, pady=(0, 6))

    @staticmethod
    def _open_dropdown(event=None):
        """Ouvre automatiquement la liste déroulante d'un Combobox au clic."""
        widget = event.widget if event else None
        if widget is not None:
            widget.event_generate("<Down>")

    def _refresh_compte_values(self, combo):
        combo["values"] = [f"{a['code']} — {a['label']}" for a in core.search_accounts(self.conn, "", limit=200)]

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def _on_compte_keyrelease(self, event=None):
        query = self.ligne_compte_var.get().strip()
        matches = core.search_accounts(self.conn, query, limit=50)
        self.ligne_compte_combo["values"] = [f"{m['code']} — {m['label']}" for m in matches]
        self._update_tiers_field()

    def _on_compte_selected(self, event=None):
        self._update_tiers_field()

    def _update_tiers_field(self):
        """Détecte si le compte choisi relève des Fournisseurs (racine 40)
        ou des Clients (racine 41) et impose alors de choisir le tiers
        auxiliaire correspondant avant de pouvoir ajouter la ligne."""
        compte = self._extract_code(self.ligne_compte_var.get())
        racine = core.account_racine(compte) if compte else None
        self.ligne_tiers_var.set("")
        if racine == core.RACINE_FOURNISSEURS:
            self._ligne_tiers_kind = "fournisseur"
            self.ligne_tiers_label.configure(text="Fournisseur (obligatoire) :")
            self.ligne_tiers_combo["values"] = [
                f"{f['code']} — {f['raison_sociale']}" for f in core.list_fournisseurs(self.conn)]
        elif racine == core.RACINE_CLIENTS:
            self._ligne_tiers_kind = "client"
            self.ligne_tiers_label.configure(text="Client (obligatoire) :")
            self.ligne_tiers_combo["values"] = [
                f"{c['code']} — {c['raison_sociale']}" for c in core.list_clients(self.conn)]
        else:
            self._ligne_tiers_kind = None
            self.ligne_tiers_label.configure(text="")
            self.ligne_tiers_combo["values"] = []

    def _on_tiers_keyrelease(self, event=None):
        query = self._extract_code(self.ligne_tiers_var.get())
        if self._ligne_tiers_kind == "fournisseur":
            items = core.list_fournisseurs(self.conn, query)
            self.ligne_tiers_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]
        elif self._ligne_tiers_kind == "client":
            items = core.list_clients(self.conn, query)
            self.ligne_tiers_combo["values"] = [f"{c['code']} — {c['raison_sociale']}" for c in items]

    def _on_stock_compte_keyrelease(self, event=None):
        query = self.stock_compte_var.get().strip()
        matches = [a for a in core.search_accounts(self.conn, query, limit=50) if a["code"][:1] == "3"]
        self.stock_compte_combo["values"] = [f"{m['code']} — {m['label']}" for m in matches]

    def _on_stock_sens_changed(self, event=None):
        if self.stock_sens_var.get().startswith("Sortie"):
            self.stock_qte_label.configure(text="Quantité vendue :")
            self.stock_info_var.set(
                "Sortie : le coût du stock = quantité × coût unitaire moyen ACTUEL du stock (pas le montant "
                "des lignes débit, qui sont ici des créances clients, pas un coût). Laissez « Compte stock » "
                "vide pour revenir au comportement ligne par ligne.")
        else:
            self.stock_qte_label.configure(text="Quantité reçue :")
            self.stock_info_var.set(
                "Entrée : le coût du stock = somme des lignes au débit (matière + frais accessoires). "
                "Laissez « Compte stock » vide pour revenir au comportement ligne par ligne.")

    def _on_analytic_changed(self, event=None):
        code = self._extract_code(self.ligne_analytic_var.get())
        unite = core.get_analytic_code_unite(self.conn, code) if code else None
        self.ligne_qte_label.configure(text=f"Quantité ({unite}) :" if unite else "Quantité :")

    def add_ligne(self):
        compte = self._extract_code(self.ligne_compte_var.get())
        if not compte:
            messagebox.showwarning("Champ manquant", "Choisissez un compte.", parent=self)
            return
        if not core.account_exists(self.conn, compte):
            messagebox.showerror("Compte invalide", f"Le compte « {compte} » n'existe pas dans le Plan comptable.",
                                  parent=self)
            return

        racine = core.account_racine(compte)
        fournisseur_code = client_code = None
        tiers_label = ""
        if racine == core.RACINE_FOURNISSEURS:
            fournisseur_code = self._extract_code(self.ligne_tiers_var.get())
            if not fournisseur_code:
                messagebox.showwarning(
                    "Fournisseur obligatoire",
                    f"Le compte « {compte} » relève des Fournisseurs (racine 40) : "
                    f"choisissez le fournisseur concerné avant d'ajouter cette ligne.",
                    parent=self)
                self.ligne_tiers_combo.focus_set()
                return
            if not core.fournisseur_exists(self.conn, fournisseur_code):
                messagebox.showerror("Fournisseur introuvable",
                                      f"Le fournisseur « {fournisseur_code} » n'existe pas (créez-le d'abord dans "
                                      f"l'onglet Fournisseurs).", parent=self)
                return
            tiers_label = fournisseur_code
        elif racine == core.RACINE_CLIENTS:
            client_code = self._extract_code(self.ligne_tiers_var.get())
            if not client_code:
                messagebox.showwarning(
                    "Client obligatoire",
                    f"Le compte « {compte} » relève des Clients (racine 41) : "
                    f"choisissez le client concerné avant d'ajouter cette ligne.",
                    parent=self)
                self.ligne_tiers_combo.focus_set()
                return
            if not core.client_exists(self.conn, client_code):
                messagebox.showerror("Client introuvable",
                                      f"Le client « {client_code} » n'existe pas (créez-le d'abord dans "
                                      f"l'onglet Clients).", parent=self)
                return
            tiers_label = client_code

        def parse_montant(raw):
            raw = (raw or "").strip()
            if not raw:
                return 0.0
            try:
                return float(raw)
            except ValueError:
                return None

        debit = parse_montant(self.ligne_debit_var.get())
        credit = parse_montant(self.ligne_credit_var.get())
        if debit is None or credit is None:
            messagebox.showerror("Erreur", "Débit et Crédit doivent être des nombres.", parent=self)
            return
        if debit and credit:
            messagebox.showerror("Erreur", "Renseignez Débit OU Crédit sur cette ligne, pas les deux.", parent=self)
            return
        if not debit and not credit:
            messagebox.showerror("Erreur", "Renseignez un montant au débit ou au crédit.", parent=self)
            return
        qte = 0.0
        if self.ligne_qte_var.get().strip():
            try:
                qte = float(self.ligne_qte_var.get())
            except ValueError:
                messagebox.showerror("Erreur", "La quantité doit être un nombre.", parent=self)
                return
        analytic_code = self._extract_code(self.ligne_analytic_var.get()) or None
        libelle = self.ligne_libelle_var.get().strip()
        self.lignes.append({"compte": compte, "libelle": libelle, "debit": debit, "credit": credit,
                             "quantite": qte, "analytic_code": analytic_code,
                             "fournisseur_code": fournisseur_code, "client_code": client_code})
        self.tree.insert("", "end", values=(
            compte, libelle, f"{fmt_cfa(debit)}" if debit else "", f"{fmt_cfa(credit)}" if credit else "",
            f"{qte:g}" if qte else "", analytic_code or "", tiers_label))
        self.ligne_compte_var.set("")
        self.ligne_libelle_var.set("")
        self.ligne_debit_var.set("")
        self.ligne_credit_var.set("")
        self.ligne_qte_var.set("")
        self.ligne_analytic_var.set("")
        self.ligne_tiers_var.set("")
        self.ligne_qte_label.configure(text="Quantité :")
        self._update_tiers_field()
        self._update_totaux()

    def delete_ligne(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne dans le tableau.", parent=self)
            return
        idx = self.tree.index(sel[0])
        del self.lignes[idx]
        self.tree.delete(sel[0])
        self._update_totaux()

    def _update_totaux(self):
        total_debit = sum(l["debit"] for l in self.lignes)
        total_credit = sum(l["credit"] for l in self.lignes)
        ecart = total_debit - total_credit
        self.total_var.set(f"Total Débit : {fmt_cfa(total_debit)}     Total Crédit : {fmt_cfa(total_credit)}     "
                            f"Écart : {fmt_cfa(ecart)}")
        self.total_label.configure(foreground="#1F7A1F" if abs(ecart) < 0.01 and total_debit > 0 else "#B00020")

    def save(self):
        if len(self.lignes) < 2:
            messagebox.showwarning("Lignes insuffisantes",
                                    "Ajoutez au moins une ligne au débit et une ligne au crédit.", parent=self)
            return
        date_str = core.to_iso_date(self.date_var.get())
        try:
            datetime.strptime(date_str, "%Y-%m-%d")
        except ValueError:
            messagebox.showerror("Erreur", "Date invalide (format JJ/MM/AAAA).", parent=self)
            return
        piece = self.piece_var.get().strip()
        if not piece:
            messagebox.showwarning("Champ manquant", "Le N° de pièce est obligatoire.", parent=self)
            return
        journal = self.journal_var.get().strip() or "OD"
        compte_stock_global = self._extract_code(self.stock_compte_var.get()) or None
        if compte_stock_global and not core.account_exists(self.conn, compte_stock_global):
            messagebox.showerror("Compte invalide", f"Le compte stock « {compte_stock_global} » n'existe pas.",
                                  parent=self)
            return
        quantite_stock_global = 0.0
        if compte_stock_global:
            if not self.stock_qte_var.get().strip():
                messagebox.showwarning(
                    "Quantité obligatoire",
                    "Vous avez choisi un compte stock : la quantité réellement reçue est obligatoire "
                    "(sinon le stock serait mis à jour avec une quantité de 0).", parent=self)
                return
            try:
                quantite_stock_global = float(self.stock_qte_var.get())
            except ValueError:
                messagebox.showerror("Erreur", "La quantité réellement reçue doit être un nombre.", parent=self)
                return
            if quantite_stock_global <= 0:
                messagebox.showerror("Erreur", "La quantité réellement reçue doit être strictement positive.",
                                      parent=self)
                return
        try:
            sens_stock_global = "sortie" if self.stock_sens_var.get().startswith("Sortie") else "entree"
            core.add_ecriture_multi_lignes(self.conn, date_str, piece, journal, self.lignes, tiers="",
                                            compte_stock_global=compte_stock_global,
                                            quantite_stock_global=quantite_stock_global,
                                            sens_stock_global=sens_stock_global)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc), parent=self)
            return
        messagebox.showinfo(
            "Écriture enregistrée",
            f"{len(self.lignes)} ligne(s) enregistrées (pièce {piece}).",
            parent=self,
        )
        self.on_saved()
        self.destroy()


class SaisieTab(ttk.Frame):
    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.selected_id = None
        self.pending_piece = None
        self._build()
        self.refresh()

    def _default_date(self):
        """Aujourd'hui si son année correspond à l'exercice courant, sinon le
        1er janvier de l'exercice courant."""
        exercice = core.get_current_exercice(self.conn)
        today = date.today()
        if str(today.year) == exercice:
            return today.strftime("%d/%m/%Y")
        return f"01/01/{exercice}"

    def _open_dropdown(self, event=None):
        """Ouvre automatiquement la liste déroulante d'un Combobox au clic,
        pour permettre de faire défiler et choisir sans avoir à taper."""
        widget = event.widget if event else None
        if widget is not None:
            widget.event_generate("<Down>")

    def _select_all(self, event=None):
        """Ctrl+A dans le tableau : sélectionne toutes les lignes visibles
        (pour une suppression groupée par exemple)."""
        self.tree.selection_set(self.tree.get_children())
        return "break"  # empêche le comportement par défaut (sélection de texte)

    def _build(self):
        form = ttk.LabelFrame(self, text="Écriture (partie double : compte débiteur ET compte créditeur obligatoires)")
        form.pack(fill="x", padx=8, pady=8)

        labels = ["Date (JJ/MM/AAAA)", "N° Pièce", "Journal",
                  "Compte débiteur", "Compte créditeur", "Montant",
                  "Tiers", "Libellé", "Fournisseur",
                  "Code analytique (ex: AN-FAB)", "Code budgétaire", "Code bailleur", "Quantité", "Client"]
        self.vars = {k: tk.StringVar() for k in labels}
        self.vars["Date (JJ/MM/AAAA)"].set(self._default_date())
        self.field_labels = {}

        for i, lbl in enumerate(labels):
            r, c = divmod(i, 3)
            lbl_widget = ttk.Label(form, text=lbl)
            lbl_widget.grid(row=r * 2, column=c, sticky="w", padx=4, pady=(4, 0))
            self.field_labels[lbl] = lbl_widget
            if lbl == "Compte débiteur":
                widget = ttk.Combobox(form, textvariable=self.vars[lbl], width=22)
                widget.grid(row=r * 2 + 1, column=c, sticky="we", padx=4, pady=(0, 4))
                widget.bind("<KeyRelease>", lambda e: self._on_compte_keyrelease("Compte débiteur"))
                widget.bind("<<ComboboxSelected>>", lambda e: (
                    self._show_account_labels(), self._validate_compte_field("Compte débiteur")))
                widget.bind("<FocusOut>", lambda e: self._validate_compte_field("Compte débiteur"))
                widget.bind("<Return>", lambda e: self._validate_compte_field("Compte débiteur"))
                widget.bind("<Tab>", lambda e: self._validate_compte_field("Compte débiteur"))
                widget.bind("<Button-1>", self._open_dropdown)
                self.compte_debit_combo = widget
            elif lbl == "Compte créditeur":
                widget = ttk.Combobox(form, textvariable=self.vars[lbl], width=22)
                widget.grid(row=r * 2 + 1, column=c, sticky="we", padx=4, pady=(0, 4))
                widget.bind("<KeyRelease>", lambda e: self._on_compte_keyrelease("Compte créditeur"))
                widget.bind("<<ComboboxSelected>>", lambda e: (
                    self._show_account_labels(), self._validate_compte_field("Compte créditeur")))
                widget.bind("<FocusOut>", lambda e: self._validate_compte_field("Compte créditeur"))
                widget.bind("<Return>", lambda e: self._validate_compte_field("Compte créditeur"))
                widget.bind("<Tab>", lambda e: self._validate_compte_field("Compte créditeur"))
                widget.bind("<Button-1>", self._open_dropdown)
                self.compte_credit_combo = widget
            elif lbl == "Journal":
                widget = ttk.Combobox(form, textvariable=self.vars[lbl], width=22,
                                       values=["AC", "VE", "OD", "BQ", "CA"])
                widget.grid(row=r * 2 + 1, column=c, sticky="we", padx=4, pady=(0, 4))
                widget.bind("<Button-1>", self._open_dropdown)
            elif lbl == "Fournisseur":
                widget = ttk.Combobox(form, textvariable=self.vars[lbl], width=22)
                widget.grid(row=r * 2 + 1, column=c, sticky="we", padx=4, pady=(0, 4))
                widget.bind("<KeyRelease>", self._on_fournisseur_keyrelease)
                widget.bind("<FocusOut>", lambda e: self._validate_fournisseur_field())
                widget.bind("<Button-1>", self._open_dropdown)
                self.fournisseur_combo = widget
                self._refresh_fournisseur_values()
            elif lbl == "Client":
                widget = ttk.Combobox(form, textvariable=self.vars[lbl], width=22)
                widget.grid(row=r * 2 + 1, column=c, sticky="we", padx=4, pady=(0, 4))
                widget.bind("<KeyRelease>", self._on_client_keyrelease)
                widget.bind("<FocusOut>", lambda e: self._validate_client_field())
                widget.bind("<Button-1>", self._open_dropdown)
                self.client_combo = widget
                self._refresh_client_values()
            elif lbl == "Code analytique (ex: AN-FAB)":
                widget = ttk.Combobox(form, textvariable=self.vars[lbl], width=22)
                widget.grid(row=r * 2 + 1, column=c, sticky="we", padx=4, pady=(0, 4))
                widget.bind("<FocusOut>", lambda e: (self._validate_plan_field(
                    "Code analytique (ex: AN-FAB)", "analytique"), self._update_quantite_label()))
                widget.bind("<<ComboboxSelected>>", lambda e: self._update_quantite_label())
                widget.bind("<Button-1>", self._open_dropdown)
                self.analytique_combo = widget
                self._refresh_plan_values("analytique")
            elif lbl == "Code budgétaire":
                widget = ttk.Combobox(form, textvariable=self.vars[lbl], width=22)
                widget.grid(row=r * 2 + 1, column=c, sticky="we", padx=4, pady=(0, 4))
                widget.bind("<FocusOut>", lambda e: self._validate_plan_field(
                    "Code budgétaire", "budgetaire"))
                widget.bind("<Button-1>", self._open_dropdown)
                self.budgetaire_combo = widget
                self._refresh_plan_values("budgetaire")
            elif lbl == "Code bailleur":
                widget = ttk.Combobox(form, textvariable=self.vars[lbl], width=22)
                widget.grid(row=r * 2 + 1, column=c, sticky="we", padx=4, pady=(0, 4))
                widget.bind("<FocusOut>", lambda e: self._validate_plan_field(
                    "Code bailleur", "bailleur"))
                widget.bind("<Button-1>", self._open_dropdown)
                self.bailleur_combo = widget
                self._refresh_plan_values("bailleur")
            else:
                widget = ttk.Entry(form, textvariable=self.vars[lbl], width=24)
                widget.grid(row=r * 2 + 1, column=c, sticky="we", padx=4, pady=(0, 4))

        self.account_label_var = tk.StringVar()
        ttk.Label(form, textvariable=self.account_label_var, foreground="#1F4E78").grid(
            row=10, column=0, columnspan=3, sticky="w", padx=4)

        self.balance_var = tk.StringVar()
        self.balance_label = ttk.Label(form, textvariable=self.balance_var, foreground="#B00020",
                                        font=("Segoe UI", 9, "bold"), wraplength=1000)
        self.balance_label.grid(row=10, column=1, columnspan=2, sticky="w", padx=4)

        btns = ttk.Frame(form)
        btns.grid(row=11, column=0, columnspan=3, sticky="w", pady=6, padx=4)
        ttk.Button(btns, text="Ajouter (écriture équilibrée)", command=self.add_entry).pack(side="left", padx=2)
        ttk.Button(btns, text="Enregistrer modification", command=self.update_entry).pack(side="left", padx=2)
        ttk.Button(btns, text="Supprimer (sélection multiple possible)", command=self.delete_entry).pack(side="left", padx=2)
        ttk.Button(btns, text="Effacer le formulaire", command=self.clear_form).pack(side="left", padx=2)
        ttk.Button(btns, text="Saisie multi-lignes (plusieurs comptes au débit, un seul au crédit)",
                   command=self.open_multi_ligne).pack(side="left", padx=(16, 2))

        import_bar = ttk.Frame(self)
        import_bar.pack(fill="x", padx=8, pady=(0, 4))
        ttk.Button(import_bar, text="Importer des écritures (.xlsx)", command=self.import_xlsx).pack(side="left", padx=2)
        ttk.Button(import_bar, text="Télécharger un modèle (.xlsx)", command=self.download_template).pack(side="left", padx=2)
        ttk.Label(import_bar, text=(
            "Pour les volumes importants : préparez un fichier avec les colonnes Date, N° Pièce, "
            "Journal, N° Compte, Tiers, Libellé, Débit, Crédit, Quantité, Code analytique, Code "
            "budgétaire, Code bailleur (l'ordre n'a pas d'importance), puis importez-le d'un coup. "
            "(L'import accepte un compte par ligne comme avant ; c'est le formulaire ci-dessus qui "
            "impose désormais la paire débit/crédit.)"
        ), foreground="#595959", wraplength=850).pack(side="left", padx=10)

        cols = ("id", "date", "piece", "journal", "compte", "libelle_compte",
                "tiers", "libelle", "debit", "credit", "quantite", "analytique", "budget", "bailleur",
                "fournisseur", "client")
        tree_frame = ttk.Frame(self)
        tree_frame.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        tree_frame.columnconfigure(0, weight=1)
        tree_frame.rowconfigure(0, weight=1)
        self.tree = ttk.Treeview(tree_frame, columns=cols, show="headings", height=15, selectmode="extended")
        headers = ["ID", "Date", "Pièce", "Journal", "Compte", "Libellé du compte",
                   "Tiers", "Libellé écriture", "Débit", "Crédit", "Qté", "Analytique", "Budget", "Bailleur",
                   "Fournisseur", "Client"]
        widths = [40, 90, 80, 60, 70, 160, 85, 140, 70, 70, 50, 75, 75, 75, 95, 95]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        tree_scroll_y = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=tree_scroll_y.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        tree_scroll_y.grid(row=0, column=1, sticky="ns")
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.tree.bind("<Control-a>", self._select_all)
        self.tree.bind("<Control-A>", self._select_all)

        totals = ttk.Frame(self)
        totals.pack(fill="x", padx=8, pady=(0, 8))
        self.totals_var = tk.StringVar()
        ttk.Label(totals, textvariable=self.totals_var, font=("Segoe UI", 10, "bold")).pack(side="left")

        self._refresh_compte_values()

    def _refresh_compte_values(self):
        """Peuple les listes déroulantes Compte débiteur/créditeur avec un
        premier lot de comptes, pour qu'un simple clic affiche déjà une
        liste à faire défiler (sans avoir à taper au clavier)."""
        accounts = core.search_accounts(self.conn, "", limit=300)
        values = [f"{a['code']} — {a['label']}" for a in accounts]
        self.compte_debit_combo["values"] = values
        self.compte_credit_combo["values"] = values

    def _extract_code(self, raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def _on_compte_keyrelease(self, field, event=None):
        combo = self.compte_debit_combo if field == "Compte débiteur" else self.compte_credit_combo
        query = self._extract_code(self.vars[field].get())
        if query:
            matches = core.search_accounts(self.conn, query, limit=30)
            combo["values"] = [f"{m['code']} — {m['label']}" for m in matches]
        self._show_account_labels()

    def _validate_compte_field(self, field):
        """Force un compte valide : propose de créer le compte ou de choisir dans la liste.
        Pour TOUT compte commençant par la racine 40 (Fournisseurs) ou 41 (Clients) —
        qu'il s'agisse de la racine elle-même (40, 41) ou d'un compte de détail
        (401000, 411000, 412000...) — impose de choisir le tiers auxiliaire concerné."""
        code = self._extract_code(self.vars[field].get())
        if not code:
            return
        if code in (core.RACINE_FOURNISSEURS, core.RACINE_CLIENTS):
            # Racine seule (40 ou 41) : pas un compte de détail postable, on
            # bascule sur le compte usuel avant de forcer le choix du tiers.
            kind = "fournisseur" if code == core.RACINE_FOURNISSEURS else "client"
            default_compte = "401000" if kind == "fournisseur" else "411000"
            if not core.account_exists(self.conn, default_compte):
                default_compte = code
            self.vars[field].set(default_compte)
            code = default_compte
        elif not core.account_exists(self.conn, code):
            if messagebox.askyesno(
                "Compte introuvable",
                f"Le compte « {code} » n'existe pas dans le Plan comptable.\n\n"
                f"Voulez-vous le créer maintenant ? (Non pour effacer et choisir un compte existant)"
            ):
                label = simpledialog.askstring("Nouveau compte", f"Libellé du compte « {code} » :", parent=self)
                if not label:
                    self.vars[field].set("")
                    return
                core.add_account(self.conn, code, label)
            else:
                self.vars[field].set("")
                return

        racine = core.account_racine(code)
        if racine == core.RACINE_FOURNISSEURS:
            self._force_tiers_selection(field, "fournisseur", code)
        elif racine == core.RACINE_CLIENTS:
            self._force_tiers_selection(field, "client", code)
        self._show_account_labels()

    def _force_tiers_selection(self, field, kind, code):
        """kind = 'fournisseur' ou 'client'. Impose de choisir le tiers auxiliaire
        pour tout compte de la racine 40/41 (pas seulement la racine elle-même)."""
        tiers_var_key = "Fournisseur" if kind == "fournisseur" else "Client"
        tiers_var = self.vars[tiers_var_key]
        if self._extract_code(tiers_var.get()):
            return  # déjà renseigné, rien à faire
        messagebox.showwarning(
            "Sélection du tiers obligatoire",
            f"Le compte « {code} » relève de la racine "
            f"{core.RACINE_FOURNISSEURS if kind == 'fournisseur' else core.RACINE_CLIENTS} "
            f"({'Fournisseurs' if kind == 'fournisseur' else 'Clients'}).\n\n"
            f"Choisissez le {kind} concerné dans le champ « {tiers_var_key} » ci-dessous "
            f"(il doit déjà exister dans le plan auxiliaire, ex. CL0001 — sinon créez-le d'abord "
            f"dans l'onglet {'Fournisseurs' if kind == 'fournisseur' else 'Clients'}). "
            f"C'est obligatoire pour valider l'écriture."
        )
        combo = self.fournisseur_combo if kind == "fournisseur" else self.client_combo
        combo.focus_set()

    def _show_account_labels(self, event=None):
        d = self._extract_code(self.vars["Compte débiteur"].get())
        c = self._extract_code(self.vars["Compte créditeur"].get())
        parts = []
        if d:
            parts.append(f"Débit {d} : {core.get_account_label(self.conn, d)}")
        if c:
            parts.append(f"Crédit {c} : {core.get_account_label(self.conn, c)}")
        self.account_label_var.set("   |   ".join(parts))

    def _refresh_fournisseur_values(self):
        items = core.list_fournisseurs(self.conn)
        self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _on_fournisseur_keyrelease(self, event=None):
        query = self._extract_code(self.vars["Fournisseur"].get())
        if query:
            items = core.list_fournisseurs(self.conn, query)
            self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _validate_fournisseur_field(self):
        code = self._extract_code(self.vars["Fournisseur"].get())
        if not code or core.fournisseur_exists(self.conn, code):
            return
        if messagebox.askyesno(
            "Fournisseur introuvable",
            f"Le fournisseur « {code} » n'existe pas.\n\n"
            f"Voulez-vous le créer maintenant ? (Non pour effacer et choisir dans la liste existante)"
        ):
            raison = simpledialog.askstring("Nouveau fournisseur", f"Raison sociale pour « {code} » :", parent=self)
            if not raison:
                self.vars["Fournisseur"].set("")
                return
            core.add_fournisseur(self.conn, code, raison)
            self._refresh_fournisseur_values()
        else:
            self.vars["Fournisseur"].set("")

    def _refresh_client_values(self):
        items = core.list_clients(self.conn)
        self.client_combo["values"] = [f"{c['code']} — {c['raison_sociale']}" for c in items]

    def _on_client_keyrelease(self, event=None):
        query = self._extract_code(self.vars["Client"].get())
        if query:
            items = core.list_clients(self.conn, query)
            self.client_combo["values"] = [f"{c['code']} — {c['raison_sociale']}" for c in items]

    def _validate_client_field(self):
        code = self._extract_code(self.vars["Client"].get())
        if not code or core.client_exists(self.conn, code):
            return
        if messagebox.askyesno(
            "Client introuvable",
            f"Le client « {code} » n'existe pas.\n\n"
            f"Voulez-vous le créer maintenant ? (Non pour effacer et choisir dans la liste existante)"
        ):
            raison = simpledialog.askstring("Nouveau client", f"Raison sociale pour « {code} » :", parent=self)
            if not raison:
                self.vars["Client"].set("")
                return
            core.add_client(self.conn, code, raison)
            self._refresh_client_values()
        else:
            self.vars["Client"].set("")

    def _update_quantite_label(self):
        """Met à jour le libellé du champ Quantité avec l'unité du code
        analytique choisi (ex. « Quantité (L) » pour l'eau, « Quantité (H) »
        pour une heure de maintenance) — pour rappeler dans quelle unité
        saisir la quantité, indispensable au calcul du coût unitaire moyen
        pondéré analytique (menu MAINTENANCE-ÉNERGIE, Fabrication)."""
        raw = self.vars["Code analytique (ex: AN-FAB)"].get().strip()
        code = raw.split(" — ", 1)[0].strip() if " — " in raw else raw
        unite = core.get_analytic_code_unite(self.conn, code) if code else None
        label_widget = self.field_labels.get("Quantité")
        if label_widget:
            label_widget.configure(text=f"Quantité ({unite})" if unite else "Quantité")

    def _refresh_plan_values(self, plan):
        if plan == "analytique":
            items = core.list_analytic_codes(self.conn)
            self.analytique_combo["values"] = [f"{i['code']} — {i['label']}" for i in items]
        elif plan == "budgetaire":
            items = core.list_budget_codes(self.conn)
            self.budgetaire_combo["values"] = [f"{i['code']} — {i['label']}" for i in items]
        elif plan == "bailleur":
            items = core.list_donor_codes(self.conn)
            self.bailleur_combo["values"] = [f"{i['code']} — {i['label']}" for i in items]

    def _validate_plan_field(self, var_key, plan):
        raw = self.vars[var_key].get().strip()
        code = raw.split(" — ", 1)[0].strip() if " — " in raw else raw
        if not code:
            return
        exists_fn = {"analytique": core.analytic_code_exists,
                     "budgetaire": core.budget_code_exists,
                     "bailleur": core.donor_code_exists}[plan]
        if exists_fn(self.conn, code):
            return
        plan_name = {"analytique": "Plan analytique", "budgetaire": "Plan budgétaire",
                     "bailleur": "Plan bailleurs de fonds"}[plan]
        if messagebox.askyesno(
            "Code introuvable",
            f"Le code « {code} » n'existe pas dans le {plan_name}.\n\n"
            f"Voulez-vous le créer maintenant ? (Non pour effacer et choisir dans la liste existante)"
        ):
            label = simpledialog.askstring("Nouveau code", f"Libellé pour « {code} » :", parent=self)
            if not label:
                self.vars[var_key].set("")
                return
            if plan == "analytique":
                core.add_analytic_code(self.conn, code, label)
            elif plan == "budgetaire":
                core.add_budget_code(self.conn, code, label)
            elif plan == "bailleur":
                core.add_donor_code(self.conn, code, label)
            self.vars[var_key].set(code)
            self._refresh_plan_values(plan)
        else:
            self.vars[var_key].set("")

    def _get_form(self):
        try:
            montant = float(self.vars["Montant"].get() or 0)
            quantite = float(self.vars["Quantité"].get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Montant et Quantité doivent être des nombres.")
            return None
        return dict(
            date_str=core.to_iso_date(self.vars["Date (JJ/MM/AAAA)"].get().strip()),
            piece=self.vars["N° Pièce"].get().strip(),
            journal=self.vars["Journal"].get().strip(),
            compte_debit=self._extract_code(self.vars["Compte débiteur"].get()),
            compte_credit=self._extract_code(self.vars["Compte créditeur"].get()),
            montant=montant,
            tiers=self.vars["Tiers"].get().strip(),
            libelle=self.vars["Libellé"].get().strip(),
            analytic_code=self._extract_code(self.vars["Code analytique (ex: AN-FAB)"].get()),
            budget_code=self._extract_code(self.vars["Code budgétaire"].get()),
            donor_code=self._extract_code(self.vars["Code bailleur"].get()),
            fournisseur_code=self._extract_code(self.vars["Fournisseur"].get()),
            client_code=self._extract_code(self.vars["Client"].get()),
            quantite=quantite,
        )

    def add_entry(self):
        data = self._get_form()
        if not data:
            return
        missing = []
        if not data["date_str"]:
            missing.append("Date")
        if not data["piece"]:
            missing.append("N° Pièce")
        if not data["compte_debit"]:
            missing.append("Compte débiteur")
        if not data["compte_credit"]:
            missing.append("Compte créditeur")
        if not data["montant"] or data["montant"] <= 0:
            missing.append("Montant (> 0)")
        if missing:
            messagebox.showwarning(
                "Champs manquants",
                "Le principe de la partie double impose de renseigner ensemble le compte "
                "débiteur ET le compte créditeur pour un même montant.\n\n"
                "Champs manquants : " + ", ".join(missing)
            )
            return
        if not core.account_exists(self.conn, data["compte_debit"]):
            messagebox.showerror("Compte invalide", f"Le compte débiteur « {data['compte_debit']} » "
                                                      f"n'existe pas dans le Plan comptable. Créez-le d'abord "
                                                      f"(quittez le champ pour être invité à le créer).")
            return
        if not core.account_exists(self.conn, data["compte_credit"]):
            messagebox.showerror("Compte invalide", f"Le compte créditeur « {data['compte_credit']} » "
                                                      f"n'existe pas dans le Plan comptable. Créez-le d'abord "
                                                      f"(quittez le champ pour être invité à le créer).")
            return
        for cote, code in (("débiteur", data["compte_debit"]), ("créditeur", data["compte_credit"])):
            if core.account_racine(code) == core.RACINE_FOURNISSEURS and not data["fournisseur_code"]:
                messagebox.showwarning(
                    "Fournisseur obligatoire",
                    f"Le compte {cote} « {code} » relève de la racine 40 (Fournisseurs) : "
                    f"le champ « Fournisseur » est obligatoire pour cette écriture."
                )
                self.fournisseur_combo.focus_set()
                return
            if core.account_racine(code) == core.RACINE_CLIENTS and not data["client_code"]:
                messagebox.showwarning(
                    "Client obligatoire",
                    f"Le compte {cote} « {code} » relève de la racine 41 (Clients) : "
                    f"le champ « Client » est obligatoire pour cette écriture."
                )
                self.client_combo.focus_set()
                return
        try:
            core.add_balanced_entry(
                self.conn, data["date_str"], data["piece"], data["journal"],
                data["compte_debit"], data["compte_credit"], data["montant"],
                data["tiers"], data["libelle"],
                analytic_code=data["analytic_code"], budget_code=data["budget_code"],
                donor_code=data["donor_code"], quantite=data["quantite"],
                fournisseur_code=data["fournisseur_code"], client_code=data["client_code"],
            )
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc))
            return
        self.refresh()
        self.balance_var.set("")
        piece = self.vars["N° Pièce"].get().strip()
        for k in ("Compte débiteur", "Compte créditeur", "Montant", "Tiers", "Libellé", "Fournisseur", "Client",
                  "Code analytique (ex: AN-FAB)", "Code budgétaire", "Code bailleur", "Quantité"):
            self.vars[k].set("")
        self.vars["N° Pièce"].set(piece)  # facilite l'ajout d'autres paires sur la même pièce
        self.account_label_var.set("")
        self.selected_id = None
        self.compte_debit_combo.focus_set()

    def update_entry(self):
        if self.selected_id is None:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne dans le tableau.")
            return
        debit_code = self._extract_code(self.vars["Compte débiteur"].get())
        credit_code = self._extract_code(self.vars["Compte créditeur"].get())
        if debit_code and credit_code:
            messagebox.showwarning(
                "Une seule ligne à la fois",
                "Pour modifier une écriture existante, ne renseignez que le compte du côté "
                "concerné (Débit OU Crédit), pas les deux — chaque ligne du tableau est une "
                "moitié d'une écriture en partie double."
            )
            return
        try:
            montant = float(self.vars["Montant"].get() or 0)
            quantite = float(self.vars["Quantité"].get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Montant et Quantité doivent être des nombres.")
            return
        if debit_code:
            if not core.account_exists(self.conn, debit_code):
                messagebox.showerror("Compte invalide", f"Le compte « {debit_code} » n'existe pas.")
                return
            fields = dict(compte=debit_code, debit=montant, credit=0)
        elif credit_code:
            if not core.account_exists(self.conn, credit_code):
                messagebox.showerror("Compte invalide", f"Le compte « {credit_code} » n'existe pas.")
                return
            fields = dict(compte=credit_code, debit=0, credit=montant)
        else:
            messagebox.showwarning("Champ manquant", "Renseignez le compte (débiteur ou créditeur) de cette ligne.")
            return
        fields.update(
            date=core.to_iso_date(self.vars["Date (JJ/MM/AAAA)"].get().strip()),
            piece=self.vars["N° Pièce"].get().strip(),
            journal=self.vars["Journal"].get().strip(),
            tiers=self.vars["Tiers"].get().strip(),
            libelle=self.vars["Libellé"].get().strip(),
            analytic_code=self._extract_code(self.vars["Code analytique (ex: AN-FAB)"].get()),
            budget_code=self._extract_code(self.vars["Code budgétaire"].get()),
            donor_code=self._extract_code(self.vars["Code bailleur"].get()),
            fournisseur_code=self._extract_code(self.vars["Fournisseur"].get()),
            client_code=self._extract_code(self.vars["Client"].get()),
            quantite=quantite,
        )
        try:
            core.update_entry(self.conn, self.selected_id, **fields)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc))
            return
        self.clear_form()
        self.refresh()

    def delete_entry(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ou plusieurs lignes dans le tableau "
                                         "(Ctrl+clic ou Maj+clic pour en sélectionner plusieurs).")
            return
        ids = [int(self.tree.item(item, "values")[0]) for item in sel]
        n = len(ids)
        question = "Supprimer cette écriture ?" if n == 1 else f"Supprimer ces {n} écritures sélectionnées ?"
        if not messagebox.askyesno("Confirmer", question):
            return
        deleted, errors = core.delete_entries_bulk(self.conn, ids)
        self.clear_form()
        self.refresh()
        if errors:
            msg = f"{deleted} écriture(s) supprimée(s) sur {n}."
            msg += "\n\nNon supprimées :\n" + "\n".join(errors[:20])
            messagebox.showwarning("Suppression partielle", msg)
        elif n > 1:
            messagebox.showinfo("Suppression terminée", f"{deleted} écriture(s) supprimée(s).")

    def clear_form(self):
        self.selected_id = None
        self.pending_piece = None
        self.balance_var.set("")
        for k, v in self.vars.items():
            v.set("" if k != "Date (JJ/MM/AAAA)" else self._default_date())
        self.account_label_var.set("")

    def open_multi_ligne(self):
        MultiLigneDialog(self, self.conn, on_saved=self.refresh)

    def download_template(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Modele_import_ecritures.xlsx",
            title="Enregistrer le modèle d'import",
        )
        if not path:
            return
        try:
            core.export_import_template(path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de la création du modèle : {exc}")
            return
        messagebox.showinfo("Modèle créé", f"Modèle enregistré :\n{path}")

    def import_xlsx(self):
        path = filedialog.askopenfilename(
            filetypes=[("Classeur Excel", "*.xlsx")],
            title="Importer des écritures",
        )
        if not path:
            return
        try:
            imported, warnings = core.import_entries_from_xlsx(self.conn, path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de l'import : {exc}")
            return
        self.refresh()
        if warnings:
            preview = "\n".join(warnings[:25])
            more = f"\n... et {len(warnings) - 25} autre(s)." if len(warnings) > 25 else ""
            messagebox.showwarning(
                "Import terminé avec avertissements",
                f"{imported} écriture(s) importée(s).\n\nAvertissements :\n{preview}{more}",
            )
        else:
            messagebox.showinfo("Import terminé", f"{imported} écriture(s) importée(s) avec succès.")

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        values = self.tree.item(sel[0], "values")
        self.selected_id = int(values[0])
        self.vars["Date (JJ/MM/AAAA)"].set(values[1])
        self.vars["N° Pièce"].set(values[2])
        self.vars["Journal"].set(values[3])
        compte = values[4]
        debit_val = values[8]
        credit_val = values[9]
        self.vars["Compte débiteur"].set("")
        self.vars["Compte créditeur"].set("")
        if debit_val:
            self.vars["Compte débiteur"].set(compte)
            self.vars["Montant"].set(debit_val)
        else:
            self.vars["Compte créditeur"].set(compte)
            self.vars["Montant"].set(credit_val)
        self.vars["Tiers"].set(values[6])
        self.vars["Libellé"].set(values[7])
        self.vars["Quantité"].set(values[10])
        self.vars["Code analytique (ex: AN-FAB)"].set(values[11])
        self.vars["Code budgétaire"].set(values[12])
        self.vars["Code bailleur"].set(values[13])
        self.vars["Fournisseur"].set(values[14])
        self.vars["Client"].set(values[15])
        self._show_account_labels()
        self._update_quantite_label()

    def refresh(self):
        self._refresh_compte_values()
        self._refresh_plan_values("analytique")
        self._refresh_plan_values("budgetaire")
        self._refresh_plan_values("bailleur")
        self._refresh_fournisseur_values()
        self._refresh_client_values()
        self._update_quantite_label()
        for row in self.tree.get_children():
            self.tree.delete(row)
        entries = core.list_entries(self.conn, exercice=core.get_current_exercice(self.conn))
        total_d = total_c = 0.0
        for e in entries:
            label = core.get_account_label(self.conn, e["compte"])
            self.tree.insert("", "end", values=(
                e["id"], core.to_display_date(e["date"]), e["piece"] or "", e["journal"] or "", e["compte"], label,
                e["tiers"] or "", e["libelle"] or "",
                fmt_cfa(e["debit"]) if e["debit"] else "",
                fmt_cfa(e["credit"]) if e["credit"] else "",
                f"{e['quantite']:g}" if e["quantite"] else "",
                e["analytic_code"] or "",
                e["budget_code"] or "",
                e["donor_code"] or "",
                e["fournisseur_code"] or "",
                e["client_code"] or "",
            ))
            total_d += e["debit"]
            total_c += e["credit"]
        equilibre = "Équilibré ✓" if abs(total_d - total_c) < 0.01 else "NON ÉQUILIBRÉ ✗"
        self.totals_var.set(f"TOTAUX — Débit : {fmt_cfa(total_d)}   Crédit : {fmt_cfa(total_c)}   {equilibre}")


class BalanceTab(ttk.Frame):
    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        cols = ("compte", "libelle", "ouv_debit", "ouv_credit", "cumul_debit", "cumul_credit",
                "solde_debit", "solde_credit")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        headers = ["N° Compte", "Libellé du compte", "Ouverture Débit", "Ouverture Crédit",
                   "Mouvement Débit", "Mouvement Crédit", "Clôture Débit", "Clôture Crédit"]
        widths = [90, 220, 100, 100, 100, 100, 100, 100]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.tag_configure("classe_total", background="#DCE6F1", font=("Segoe UI", 9, "bold"))
        self.tree.tag_configure("grand_total", background="#1F4E78", foreground="white", font=("Segoe UI", 10, "bold"))
        self.tree.pack(fill="both", padx=8, pady=8)
        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill="x", pady=(0, 4))
        ttk.Button(btn_bar, text="Actualiser", command=self.refresh).pack(side="left", padx=8)
        ttk.Button(btn_bar, text="Exporter (.xlsx)", command=self.export_xlsx).pack(side="left", padx=2)
        self.ecart_var = tk.StringVar()
        ttk.Label(self, textvariable=self.ecart_var, font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=8, pady=(0, 8))
        self.refresh()

    def export_xlsx(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Balance.xlsx", title="Exporter la Balance",
        )
        if not path:
            return
        core.export_balance_xlsx(self.conn, path)
        messagebox.showinfo("Export terminé", f"Balance exportée :\n{path}")

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        data = core.compute_balance_detaillee(self.conn)

        def f(v):
            return f"{fmt_cfa(v)}" if v else ""

        for c in data["classes"]:
            for l in c["lignes"]:
                self.tree.insert("", "end", values=(
                    l["code"], l["label"], f(l["ouverture_debit"]), f(l["ouverture_credit"]),
                    f(l["cumul_debit"]), f(l["cumul_credit"]), f(l["solde_debit"]), f(l["solde_credit"]),
                ))
            st = c["sous_total"]
            self.tree.insert("", "end", tags=("classe_total",), values=(
                "", f"TOTAL CLASSE {c['classe']}", f(st["ouverture_debit"]), f(st["ouverture_credit"]),
                f(st["cumul_debit"]), f(st["cumul_credit"]), f(st["solde_debit"]), f(st["solde_credit"]),
            ))
        gt = data["grand_total"]
        self.tree.insert("", "end", tags=("grand_total",), values=(
            "", "TOTAL BALANCE", f(gt["ouverture_debit"]), f(gt["ouverture_credit"]),
            f(gt["cumul_debit"]), f(gt["cumul_credit"]), f(gt["solde_debit"]), f(gt["solde_credit"]),
        ))
        ecart_ouv = gt["ouverture_debit"] - gt["ouverture_credit"]
        ecart_cumul = gt["cumul_debit"] - gt["cumul_credit"]
        ecart_solde = gt["solde_debit"] - gt["solde_credit"]
        if abs(ecart_ouv) < 1 and abs(ecart_cumul) < 1 and abs(ecart_solde) < 1:
            self.ecart_var.set("✓ Balance équilibrée sur les 3 paires de colonnes (Ouverture, Mouvement, Clôture).")
        else:
            msg = "⚠ Balance déséquilibrée — "
            parts = []
            if abs(ecart_ouv) >= 1:
                parts.append(f"écart Ouverture Débit/Crédit de {fmt_cfa(ecart_ouv)} (soldes d'ouverture "
                              f"incomplets — voir l'onglet Soldes d'ouverture)")
            if abs(ecart_cumul) >= 1:
                parts.append(f"écart Mouvement Débit/Crédit de {fmt_cfa(ecart_cumul)} (des écritures de la "
                              f"période ne sont pas équilibrées — vérifiez un éventuel import massif)")
            if abs(ecart_solde) >= 1:
                parts.append(f"écart Clôture Débit/Crédit de {fmt_cfa(ecart_solde)}")
            self.ecart_var.set(msg + " ; ".join(parts))


class CompteResultatTab(ttk.Frame):
    """Compte de résultat selon les Soldes Intermédiaires de Gestion (SIG),
    présenté selon le modèle officiel, avec une couleur par section.
    Calculé à partir de compute_liasse_resultat() — la même fonction que la
    Liasse fiscale et la Situation financière — donc toujours cohérent avec
    la Balance, le Bilan, le TFT et la Situation financière."""

    SECTIONS = {
        "commerciale": "#D9EAD3",   # vert clair — activité commerciale
        "ca": "#CFE2F3",            # bleu clair — chiffre d'affaires
        "va": "#FFF2CC",            # jaune clair — valeur ajoutée
        "ebe": "#D9D2E9",           # violet clair — EBE / résultat exploitation
        "financier": "#FCE5CD",     # orange clair — résultat financier
        "hao": "#F4CCCC",           # rouge/rose clair — HAO / résultat net
        "total": "#1F4E78",
    }

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        cols = ("libelle", "montant")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=30)
        self.tree.heading("libelle", text="Rubrique")
        self.tree.heading("montant", text="Montant")
        self.tree.column("libelle", width=480, anchor="w")
        self.tree.column("montant", width=160, anchor="e")
        for key, color in self.SECTIONS.items():
            fg = "white" if key == "total" else "black"
            tree_font = ("Segoe UI", 9, "bold") if key == "total" else ("Segoe UI", 9)
            self.tree.tag_configure(key, background=color, foreground=fg, font=tree_font)
            self.tree.tag_configure(key + "_header", background=color, foreground=fg, font=("Segoe UI", 9, "bold"))
        self.tree.pack(fill="both", padx=8, pady=8)
        ttk.Label(self, text=(
            "Calculé à partir de la même fonction que la Liasse fiscale, la Situation financière et "
            "le TFT (compute_liasse_resultat) — toujours cohérent avec la Balance et le Bilan."
        ), foreground="#595959").pack(anchor="w", padx=8)
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(pady=8)
        ttk.Button(self, text="Exporter (gabarit officiel .xlsx)",
                   command=lambda: export_etat_gabarit(
                       self, self.conn, "resultat", "Compte_de_Resultat.xlsx", "le Compte de résultat")
                   ).pack(pady=(0, 8))
        self.refresh()

    def _row(self, tag, label, val):
        self.tree.insert("", "end", tags=(tag,), values=(f"  {label}", f"{fmt_cfa(val)}"))

    def _header(self, tag, titre):
        self.tree.insert("", "end", tags=(tag + "_header",), values=(titre, ""))

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        cr = core.compute_liasse_resultat(self.conn)

        self._header("commerciale", "ACTIVITÉ COMMERCIALE")
        self._row("commerciale", "+ Vente de marchandises (A)", cr["TA"])
        self._row("commerciale", "- Coût d'achat des marchandises vendues", cr["RA"])
        self._row("commerciale", "- Variation de stocks de marchandises", cr["RA_STOCK"])
        self._row("commerciale", "MARGE COMMERCIALE", cr["XA"])

        self._header("ca", "CHIFFRE D'AFFAIRES")
        self._row("ca", "+ Vente de produits fabriqués (B)", cr["TB"])
        self._row("ca", "+ Travaux, services vendus (C)", cr["TC"])
        self._row("ca", "+ Produits accessoires (D)", cr["TD"])
        self._row("ca", "CHIFFRE D'AFFAIRES (A+B+C+D)", cr["XB"])

        self._header("va", "VALEUR AJOUTÉE")
        self._row("va", "+ Production stockée", cr["TE"])
        self._row("va", "+ Subvention d'exploitation", cr["TG"])
        self._row("va", "+ Autres produits", cr["TH"])
        self._row("va", "- Achats de matières premières (+ variation de stocks)", cr["RC"])
        self._row("va", "- Autres achats (+ variation de stocks)", cr["RE"])
        self._row("va", "- Transport", cr["RG"])
        self._row("va", "- Services extérieurs", cr["RH"])
        self._row("va", "- Impôts et taxes", cr["RI"])
        self._row("va", "- Autres charges", cr["RJ"])
        self._row("va", "VALEUR AJOUTÉE", cr["XC"])

        self._header("ebe", "EXCÉDENT BRUT D'EXPLOITATION ET RÉSULTAT D'EXPLOITATION")
        self._row("ebe", "- Charges de personnel", cr["RK"])
        self._row("ebe", "EXCÉDENT BRUT D'EXPLOITATION (EBE)", cr["XD"])
        self._row("ebe", "- Dotations aux amortissements et provisions", cr["RL"])
        self._row("ebe", "RÉSULTAT D'EXPLOITATION", cr["XE"])

        self._header("financier", "RÉSULTAT FINANCIER")
        self._row("financier", "+ Produits financiers", cr["TK"])
        self._row("financier", "- Frais financiers et charges assimilées", cr["RM"])
        self._row("financier", "RÉSULTAT FINANCIER", cr["XF"])
        self._row("financier", "RÉSULTAT DES ACTIVITÉS ORDINAIRES", cr["XG"])

        self._header("hao", "RÉSULTAT HORS ACTIVITÉS ORDINAIRES ET RÉSULTAT NET")
        self._row("hao", "RÉSULTAT HAO", cr["XH"])
        self._row("hao", "- Participation des salariés", cr["RQ"])
        self._row("hao", "- Impôts sur les bénéfices", cr["RS"])
        self.tree.insert("", "end", tags=("total",), values=("RÉSULTAT NET COMPTABLE", f"{fmt_cfa(cr['XI'])}"))


class EtatFormuleTab(ttk.Frame):
    """Écran générique pour un état financier basé sur un gabarit à
    formules CtaCptSolde... (Compte de résultat SIG, TFT, Situation
    financière) — réutilise core.compute_etat_formule_generique() pour ne
    pas dupliquer la logique entre ces 3 états, qui partagent tous la même
    structure « RUBRIQUE | N (| N-1 | %) »."""

    def __init__(self, parent, conn, titre, template_path_getter):
        super().__init__(parent)
        self.conn = conn
        self.titre = titre
        self.template_path_getter = template_path_getter

        ttk.Label(self, text=titre, font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=8, pady=(8, 0))
        ttk.Label(self, text=(
            "Calculé avec les formules CtaCptSolde/CtaCptSoldeDébit/CtaCptSoldeCrédit (et leurs variantes "
            "…Nm1 pour l'exercice N-1) du gabarit officiel — entièrement autonome, aucun fichier externe requis."
        ), foreground="#595959", wraplength=1300, justify="left").pack(anchor="w", padx=8, pady=(0, 4))

        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill="x", padx=8, pady=(0, 4))
        ttk.Button(btn_bar, text="Actualiser", command=self.refresh).pack(side="left")
        ttk.Button(btn_bar, text="Modifier les formules du template",
                   command=self.modifier_template).pack(side="left", padx=8)
        ttk.Button(btn_bar, text="Exporter (.xls)", command=self.export_xls).pack(side="left", padx=8)

        self.cols = ("libelle", "n", "n1", "pct")
        self.tree = ttk.Treeview(self, columns=self.cols, show="headings", height=32)
        self.tree.heading("libelle", text="Rubrique")
        self.tree.column("libelle", width=460, anchor="w", stretch=True)
        for c in ("n", "n1", "pct"):
            self.tree.column(c, width=140, anchor="e")
        self.tree.tag_configure("total", background="#FFCC00", foreground="black", font=("Segoe UI", 10, "bold"))
        self.tree.pack(fill="both", expand=True, padx=8, pady=(0, 8))

        self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        try:
            d = core.compute_etat_formule_generique(self.conn, self.template_path_getter)
        except Exception as exc:
            messagebox.showerror(
                "Erreur",
                f"Impossible de calculer « {self.titre} » :\n\n{exc}\n\n"
                f"Cet écran ne dépend d'aucun fichier externe — réessayez après avoir relancé "
                f"l'application ; si l'erreur persiste, contactez le support.",
            )
            return

        headers = {"N": "Exercice N", "N-1": "Exercice N-1", "%": "%"}
        display_cols = [c for c in ("n", "n1", "pct")]
        for col_key, label in zip(display_cols, d["colonnes"] + [""] * 3):
            self.tree.heading(col_key, text=headers.get(label, label) if label else "")

        def fmt(v):
            return fmt_cfa(v) if isinstance(v, (int, float)) else ""

        for l in d["lignes"]:
            valeurs = [l.get(c, None) for c in d["colonnes"]]
            while len(valeurs) < 3:
                valeurs.append(None)
            tag = ("total",) if any(mot in l["libelle"].upper() for mot in
                                     ("TOTAL", "RESULTAT NET", "MARGE COMMERCIALE", "CAFG",
                                      "TRESORERIE NETTE", "VARIATION DE LA TRÉSORERIE",
                                      "VARIATION DE LA TRESORERIE")) else ()
            self.tree.insert("", "end", tags=tag, values=(l["libelle"], fmt(valeurs[0]), fmt(valeurs[1]),
                                                           fmt(valeurs[2])))

        if d["errors"]:
            detail = "\n".join(f"• Cellule {coord} : {msg}" for coord, _formula, msg in d["errors"][:10])
            messagebox.showwarning(
                "Formules en erreur",
                f"{len(d['errors'])} formule(s) du gabarit n'ont pas pu être évaluées. Les autres lignes "
                f"restent correctes.\n\n{detail}",
            )

    def modifier_template(self):
        path = self.template_path_getter()
        if not messagebox.askyesno(
            "Modifier les formules du template",
            "Le gabarit va s'ouvrir dans Excel. Toute formule que vous modifiez ET ENREGISTREZ (Ctrl+S, "
            "en gardant le même format) sera directement utilisée par l'application au prochain calcul.\n\n"
            "Continuer ?",
        ):
            return
        if not _ouvrir_fichier(path):
            messagebox.showinfo("Ouverture impossible",
                                 f"Ouvrez-le manuellement dans Excel :\n{path}")

    def export_xls(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xls", filetypes=[("Classeur Excel", "*.xls")],
            initialfile=f"{self.titre.replace(' ', '_')}.xls", title=f"Exporter {self.titre}",
        )
        if not path:
            return
        try:
            core.export_etat_formule_xls(self.conn, self.template_path_getter, path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de l'export : {exc}")
            return
        messagebox.showinfo("Export terminé", f"{self.titre} exporté :\n{path}")


class BilanSyscohadaTab(ttk.Frame):
    """Bilan SYSCOHADA (menu RAPPORTS FINANCIERS) — MONTÉ SUR LE SOLDE DE
    CLÔTURE HABITUEL (colonne N = solde d'ouverture + cumul des opérations
    de la période) ; la colonne N-1 contient le solde d'ouverture de
    l'exercice, qui correspond mathématiquement au solde de clôture de
    l'exercice précédent (une fois la clôture d'exercice effectuée) — voir
    core.compute_bilan_detaille(). Chaque colonne est équilibrée
    indépendamment (Actif = Passif en N, Actif = Passif en N-1), la partie
    double garantissant que la somme des soldes de clôture comme celle des
    soldes d'ouverture est nulle. Entièrement autonome — ne dépend d'aucun
    fichier de gabarit externe."""

    RACINE_COLORS = {
        "40": ("#FF6600", "white"), "41": ("#3366FF", "white"), "42": ("#FFFF00", "black"),
        "43": ("#FF99CC", "black"), "44_45": ("#999999", "white"),
        "46": ("#00FFFF", "black"), "47_49": ("#00FFFF", "black"),
    }
    STOCK_COLOR = "#99CCFF"
    TRESO_COLOR = "#00FF00"

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn

        ttk.Label(self, text="BILAN SYSCOHADA", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=8, pady=(8, 0))
        ttk.Label(self, text=(
            "Colonnes « Brut / Amort. / Net » : solde de clôture habituel (solde d'ouverture + cumul des "
            "opérations de la période). Colonne « Solde d'ouverture » : uniquement le solde de début "
            "d'exercice (report à nouveau du 1er janvier), qui correspond au solde de clôture de "
            "l'exercice précédent. Chaque colonne est équilibrée indépendamment (Actif = Passif), sauf "
            "soldes d'ouverture incomplets — voir le diagnostic ci-dessous le cas échéant."
        ), foreground="#595959", wraplength=1300, justify="left").pack(anchor="w", padx=8, pady=(0, 4))

        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill="x", padx=8, pady=(0, 4))
        ttk.Button(btn_bar, text="Actualiser", command=self.refresh).pack(side="left")
        ttk.Button(btn_bar, text="Modifier les formules du template",
                   command=self.modifier_template).pack(side="left", padx=8)
        ttk.Button(btn_bar, text="Visionner le Bilan selon mon template",
                   command=self.visionner_gabarit).pack(side="left", padx=8)
        ttk.Button(btn_bar, text="Exporter selon le gabarit officiel (formules CtaCptSolde exactes)",
                   command=self.export_gabarit).pack(side="left", padx=8)
        self.ecart_var = tk.StringVar()
        self.ecart_label = ttk.Label(btn_bar, textvariable=self.ecart_var, font=("Segoe UI", 10, "bold"))
        self.ecart_label.pack(side="left", padx=16)

        columns_frame = ttk.Frame(self)
        columns_frame.pack(fill="both", expand=True, padx=8, pady=(0, 4))
        columns_frame.columnconfigure(0, weight=1)
        columns_frame.columnconfigure(1, weight=1)
        columns_frame.rowconfigure(0, weight=1)

        actif_cols = ("libelle", "brut", "amort", "net", "net_n1")
        self.tree_actif = ttk.Treeview(columns_frame, columns=actif_cols, show="headings", height=30)
        headers_a = ["Libellé (ACTIF)", "Brut", "Amort.", "Net", "Net N-1 (ouverture)"]
        for c, h, w in zip(actif_cols, headers_a, [260, 100, 100, 110, 130]):
            self.tree_actif.heading(c, text=h)
            self.tree_actif.column(c, width=w, anchor="w" if c == "libelle" else "e", stretch=(c == "libelle"))
        self.tree_actif.grid(row=0, column=0, sticky="nsew", padx=(0, 4))

        passif_cols = ("libelle", "montant", "montant_n1")
        self.tree_passif = ttk.Treeview(columns_frame, columns=passif_cols, show="headings", height=30)
        headers_p = ["Libellé (PASSIF)", "Montant", "Montant N-1 (ouverture)"]
        for c, h, w in zip(passif_cols, headers_p, [280, 140, 140]):
            self.tree_passif.heading(c, text=h)
            self.tree_passif.column(c, width=w, anchor="w" if c == "libelle" else "e", stretch=(c == "libelle"))
        self.tree_passif.grid(row=0, column=1, sticky="nsew", padx=(4, 0))

        style = ttk.Style()
        style.configure("BilanS.Treeview", rowheight=22, font=("Segoe UI", 10))
        for tree in (self.tree_actif, self.tree_passif):
            tree.configure(style="BilanS.Treeview")
            tree.tag_configure("header", background="#1F4E78", foreground="white", font=("Segoe UI", 10, "bold"))
            tree.tag_configure("soustotal", background="#99CCFF", foreground="black", font=("Segoe UI", 10, "bold"))
            tree.tag_configure("total", background="#FFCC00", foreground="black", font=("Segoe UI", 10, "bold"))
            tree.tag_configure("stock", background=self.STOCK_COLOR, foreground="black")
            tree.tag_configure("treso", background=self.TRESO_COLOR, foreground="black")
            for racine, (bg, fg) in self.RACINE_COLORS.items():
                tree.tag_configure(f"racine_{racine}", background=bg, foreground=fg)

        self.diag_var = tk.StringVar()
        self.diag_label = ttk.Label(self, textvariable=self.diag_var, foreground="#B00020", wraplength=1300,
                                     justify="left")
        self.diag_label.pack(anchor="w", padx=8, pady=(4, 8))

        self.refresh()

    def _tag_for(self, key):
        if key in self.RACINE_COLORS:
            return f"racine_{key}"
        if key in ("31", "32", "33", "34", "35", "36", "37", "38", "39"):
            return "stock"
        return ""

    def _add_actif_group(self, titre, lignes, total_label, total_val, total_val_n1, detail=False, tag=""):
        if not lignes and not total_val:
            return
        self.tree_actif.insert("", "end", tags=("header",), values=(titre, "", "", "", ""))
        for l in lignes:
            row_tag = tag or self._tag_for(l.get("key"))
            if detail:
                self.tree_actif.insert("", "end", tags=(row_tag,), values=(
                    f"  {l['label']}", fmt_cfa(l["brut"]) if l["brut"] else "",
                    fmt_cfa(l["amort"]) if l["amort"] else "", fmt_cfa(l["net"]), fmt_cfa(l.get("net_n1", 0))))
                for compte in l.get("comptes", []):
                    self.tree_actif.insert("", "end", values=(
                        f"      • {compte['label']}", fmt_cfa(compte["montant"]), "", "", ""))
            else:
                montant = l.get("sous_total", 0)
                montant_n1 = l.get("sous_total_n1", 0)
                if not montant and not montant_n1:
                    continue
                self.tree_actif.insert("", "end", tags=(row_tag,), values=(
                    f"  {l['label']}", "", "", fmt_cfa(montant), fmt_cfa(montant_n1)))
        self.tree_actif.insert("", "end", tags=("soustotal",), values=(
            f"  {total_label}", "", "", fmt_cfa(total_val), fmt_cfa(total_val_n1)))

    def _add_passif_group(self, titre, lignes, total_label, total_val, total_val_n1, tag=""):
        if not lignes and not total_val:
            return
        self.tree_passif.insert("", "end", tags=("header",), values=(titre, "", ""))
        for l in lignes:
            montant = l.get("sous_total", 0)
            montant_n1 = l.get("sous_total_n1", 0)
            if not montant and not montant_n1:
                continue
            row_tag = tag or self._tag_for(l.get("key"))
            self.tree_passif.insert("", "end", tags=(row_tag,), values=(
                f"  {l['label']}", fmt_cfa(montant), fmt_cfa(montant_n1)))
        self.tree_passif.insert("", "end", tags=("soustotal",), values=(
            f"  {total_label}", fmt_cfa(total_val), fmt_cfa(total_val_n1)))

    def refresh(self):
        for tree in (self.tree_actif, self.tree_passif):
            for row in tree.get_children():
                tree.delete(row)

        d = core.compute_bilan_detaille(self.conn)
        a, p = d["actif"], d["passif"]

        self._add_actif_group("IMMOBILISATIONS", a["immobilisations"], "Total immobilisations nettes",
                               a["total_immo_net"], a["total_immo_net_n1"], detail=True)
        self._add_actif_group("STOCKS", a["stocks"], "Total stocks", a["total_stocks"], a["total_stocks_n1"],
                               tag="stock")
        self._add_actif_group("CRÉANCES", a["creances"], "Total créances", a["total_creances"],
                               a["total_creances_n1"])
        self._add_actif_group("TRÉSORERIE ACTIF", a["tresorerie"], "Total trésorerie actif",
                               a["total_tresorerie"], a["total_tresorerie_n1"], tag="treso")
        self.tree_actif.insert("", "end", tags=("total",), values=(
            "TOTAL ACTIF", "", "", fmt_cfa(d["total_actif"]), fmt_cfa(d["total_actif_n1"])))

        self._add_passif_group("CAPITAUX PROPRES ET RESSOURCES DURABLES", p["capitaux_propres"],
                                "Total capitaux propres", p["total_capitaux_propres"],
                                p["total_capitaux_propres_n1"])
        self._add_passif_group("DETTES CIRCULANTES", p["dettes"], "Total dettes circulantes",
                                p["total_dettes"], p["total_dettes_n1"])
        self._add_passif_group("TRÉSORERIE PASSIF", p["tresorerie"], "Total trésorerie passif",
                                p["total_tresorerie"], p["total_tresorerie_n1"], tag="treso")
        self.tree_passif.insert("", "end", tags=("total",), values=(
            "TOTAL PASSIF", fmt_cfa(d["total_passif"]), fmt_cfa(d["total_passif_n1"])))

        ecart, ecart_n1 = d["ecart"], d["ecart_n1"]
        if abs(ecart) < 1 and abs(ecart_n1) < 1:
            self.ecart_var.set(f"✓ Actif = Passif ({fmt_cfa(d['total_actif'])})   —   Exercice {d['exercice']} "
                                f"/ N-1 (ouverture) = {fmt_cfa(d['total_actif_n1'])}")
            self.ecart_label.configure(foreground="#1F7A1F")
            self.diag_var.set("")
        else:
            msgs = []
            if abs(ecart) >= 1:
                msgs.append(f"exercice N : {fmt_cfa(ecart)}")
            if abs(ecart_n1) >= 1:
                msgs.append(f"solde d'ouverture (N-1) : {fmt_cfa(ecart_n1)}")
            self.ecart_var.set("⚠ Écart Actif - Passif — " + " ; ".join(msgs))
            self.ecart_label.configure(foreground="#B00020")
            diag = core.compute_ecart_diagnostic(self.conn)
            parts = []
            if abs(diag["ecart_soldes_ouverture"]) >= 1:
                parts.append(f"• Soldes d'ouverture non nuls : {fmt_cfa(diag['ecart_soldes_ouverture'])} "
                              f"(voir l'onglet « Soldes d'ouverture »)")
            if abs(diag["ecart_ecritures_periode"]) >= 1:
                parts.append(f"• Écritures de la période Débit ≠ Crédit : {fmt_cfa(diag['ecart_ecritures_periode'])}")
            self.diag_var.set("\n".join(parts))

        self.tree_actif.xview_moveto(0)
        self.tree_passif.xview_moveto(0)

    def telecharger_template(self):
        """Télécharge le gabarit VIERGE (avec ses formules CtaCptSolde
        textuelles non évaluées) tel quel — pour que l'utilisateur puisse
        le consulter, le modifier ou le réutiliser ailleurs."""
        path = filedialog.asksaveasfilename(
            defaultextension=".xls", filetypes=[("Classeur Excel", "*.xls")],
            initialfile="Template_Bilan_avec_formules.xls",
            title="Télécharger le template du Bilan (avec formules)",
        )
        if not path:
            return
        try:
            import shutil
            shutil.copyfile(core.BILAN_TEMPLATE_PATH or core._bilan_template_path(), path)
        except Exception as exc:
            messagebox.showerror(
                "Erreur",
                f"Impossible de récupérer le template :\n\n{exc}\n\n"
                f"Le fichier templates/bilan_template.xls doit être présent dans l'installation.",
            )
            return
        messagebox.showinfo("Téléchargement terminé", f"Template enregistré :\n{path}")

    def modifier_template(self):
        """Ouvre directement le gabarit ACTIF (le même fichier utilisé par
        l'application pour « Visionner » et « Exporter ») avec Excel — le
        template reste figé du point de vue du logiciel (aucun mécanisme
        d'import/remplacement dans l'appli), mais comme c'est le MÊME
        fichier physique, toute formule modifiée et enregistrée dans Excel
        est directement prise en compte au prochain calcul, sans étape
        supplémentaire."""
        path = core.BILAN_TEMPLATE_PATH or core._bilan_template_path()
        if not messagebox.askyesno(
            "Modifier les formules du template",
            "Le gabarit va s'ouvrir dans Excel. Toute formule que vous modifiez ET ENREGISTREZ (Ctrl+S, "
            "en gardant le même format) sera directement utilisée par l'application au prochain calcul "
            "(« Visionner », « Exporter »).\n\nContinuer ?",
        ):
            return
        if not _ouvrir_fichier(path):
            messagebox.showinfo(
                "Ouverture impossible",
                f"Le template n'a pas pu s'ouvrir automatiquement — ouvrez-le manuellement dans Excel :"
                f"\n{path}",
            )

    def visionner_gabarit(self):
        """Génère le Bilan avec les formules exactes du gabarit officiel
        dans un fichier temporaire, puis l'OUVRE directement avec
        l'application par défaut (Excel) pour un visionnage immédiat —
        sans boîte de dialogue « Enregistrer sous »."""
        import tempfile
        tmp_dir = tempfile.gettempdir()
        path = os.path.join(tmp_dir, f"Bilan_visionnage_{core.get_current_exercice(self.conn)}.xls")
        try:
            core.export_bilan_gabarit_xlsx(self.conn, path)
        except Exception as exc:
            messagebox.showerror(
                "Erreur",
                f"Impossible de générer le Bilan selon le template :\n\n{exc}\n\n"
                f"Le fichier templates/bilan_template.xls doit être présent dans l'installation.",
            )
            return
        if not _ouvrir_fichier(path):
            messagebox.showinfo(
                "Bilan généré",
                f"Le Bilan a été généré mais n'a pas pu s'ouvrir automatiquement — ouvrez-le manuellement :"
                f"\n{path}",
            )

    def export_gabarit(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xls", filetypes=[("Classeur Excel", "*.xls")],
            initialfile="Bilan_gabarit_officiel.xls",
            title="Exporter le Bilan selon le gabarit officiel (formules exactes)",
        )
        if not path:
            return
        try:
            core.export_bilan_gabarit_xlsx(self.conn, path)
        except Exception as exc:
            messagebox.showerror(
                "Erreur",
                f"Échec de l'export selon le gabarit officiel :\n\n{exc}\n\n"
                f"Cet export nécessite le fichier templates/bilan_template.xls dans l'installation. "
                f"L'écran Bilan SYSCOHADA ci-dessus reste utilisable normalement, il n'en dépend pas.",
            )
            return
        messagebox.showinfo(
            "Export terminé",
            f"Bilan exporté avec les formules CtaCptSolde exactes du gabarit officiel (solde de clôture — "
            f"note : les formules N/N-1 de ce gabarit utilisent le solde de clôture, contrairement à "
            f"l'écran ci-dessus qui isole les opérations de la période et le solde d'ouverture) :\n{path}",
        )


class PiecesNonEquilibreesTab(ttk.Frame):
    """Diagnostic : liste toutes les pièces (regroupement Pièce + Journal)
    dont le total Débit ne correspond pas au total Crédit — la cause la plus
    fréquente d'un Bilan qui ne s'équilibre pas, typiquement issue d'un
    import en masse d'écritures qui n'a pas respecté la partie double.
    Chaque pièce listée ici doit être corrigée dans l'onglet Saisie."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="ÉCRITURES NON ÉQUILIBRÉES", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Chaque pièce comptable doit avoir un total Débit strictement égal à son total Crédit "
            "(partie double). Les pièces listées ci-dessous ne le sont pas — corrigez-les dans "
            "l'onglet Saisie (filtrez par numéro de pièce) pour rééquilibrer le Bilan."
        ), foreground="#595959", wraplength=1100).pack(anchor="w", padx=16, pady=(0, 8))

        filt = ttk.Frame(self)
        filt.pack(fill="x", padx=16, pady=4)
        self.toutes_dates_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(filt, text="Chercher sur toutes les dates (pas seulement l'exercice en cours)",
                        variable=self.toutes_dates_var, command=self.refresh).pack(side="left")
        ttk.Button(filt, text="Actualiser", command=self.refresh).pack(side="left", padx=12)

        cols = ("piece", "journal", "date_min", "date_max", "nb", "debit", "credit", "ecart")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=22)
        headers = ["Pièce", "Journal", "Date min", "Date max", "Nb lignes", "Total Débit", "Total Crédit", "Écart"]
        widths = [130, 80, 90, 90, 80, 140, 140, 140]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w" if c in ("piece", "journal") else "e")
        self.tree.tag_configure("total", background="#1F4E78", foreground="white", font=("Segoe UI", 10, "bold"))
        self.tree.pack(fill="both", padx=16, pady=8)

        self.total_var = tk.StringVar()
        ttk.Label(self, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=16, pady=(0, 16))
        self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        pieces = core.compute_pieces_non_equilibrees(self.conn, toutes_dates=self.toutes_dates_var.get())
        total_ecart = 0.0
        for p in pieces:
            self.tree.insert("", "end", values=(
                p["piece"] or "(sans n° de pièce)", p["journal"] or "", core.to_display_date(p["date_min"]),
                core.to_display_date(p["date_max"]), p["nb"], fmt_cfa(p["d"]), fmt_cfa(p["c"]), fmt_cfa(p["ecart"]),
            ))
            total_ecart += p["ecart"]
        if pieces:
            self.tree.insert("", "end", tags=("total",), values=(
                "", "", "", "", "", "", "TOTAL ÉCART", fmt_cfa(total_ecart),
            ))
            self.total_var.set(f"{len(pieces)} pièce(s) non équilibrée(s) trouvée(s) — "
                                f"écart cumulé : {fmt_cfa(total_ecart)}")
        else:
            self.total_var.set("✓ Aucune pièce déséquilibrée trouvée — toutes les écritures respectent "
                                "Débit = Crédit sur le périmètre recherché.")


class GrandLivreTab(ttk.Frame):
    """Grand livre complet : affiche TOUS les comptes ayant un mouvement (ou
    un solde d'ouverture) sur l'exercice, groupés par compte puis par classe,
    avec des bandes de couleur (bleu = compte / sous-total de compte, orange
    = total de classe) — comme un grand livre papier classique. Un filtre
    optionnel permet de se recentrer sur un compte ou un tiers précis."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn

        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=8, pady=8)
        ttk.Label(bar, text="Filtrer sur un compte (optionnel) :").pack(side="left")
        self.compte_var = tk.StringVar()
        self.compte_combo = ttk.Combobox(bar, textvariable=self.compte_var, width=30)
        self.compte_combo.pack(side="left", padx=4)
        self.compte_combo.bind("<KeyRelease>", self._on_compte_keyrelease)
        self.compte_combo.bind("<<ComboboxSelected>>", lambda e: self.refresh())
        self.compte_combo.bind("<Button-1>", self._open_dropdown)
        self.compte_combo.bind("<Return>", lambda e: self.refresh())
        self._refresh_compte_values()
        ttk.Label(bar, text="Tiers (optionnel) :").pack(side="left", padx=(12, 0))
        self.tiers_var = tk.StringVar()
        ttk.Entry(bar, textvariable=self.tiers_var, width=18).pack(side="left", padx=4)
        ttk.Button(bar, text="Filtrer", command=self.refresh).pack(side="left", padx=8)
        ttk.Button(bar, text="Réinitialiser (tous les comptes)", command=self._reset).pack(side="left", padx=2)
        ttk.Label(bar, text="Par défaut, tous les comptes de l'exercice sont affichés.",
                  foreground="#595959").pack(side="left", padx=10)

        cols = ("date", "piece", "journal", "libelle", "ouv_debit", "ouv_credit",
                "mvt_debit", "mvt_credit", "sold_debit", "sold_credit")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        headers = ["Date", "Pièce", "Journal", "Libellé", "Ouverture Débit", "Ouverture Crédit",
                   "Mouvement Débit", "Mouvement Crédit", "Clôture Débit", "Clôture Crédit"]
        widths = [85, 70, 55, 220, 100, 100, 100, 100, 100, 100]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.tag_configure("compte_header", background="#B4C6E7", font=("Segoe UI", 9, "bold"))
        self.tree.tag_configure("compte_total", background="#B4C6E7", font=("Segoe UI", 9, "bold"))
        self.tree.tag_configure("classe_total", background="#F4B183", font=("Segoe UI", 10, "bold"))
        self.tree.pack(fill="both", expand=True, padx=8, pady=(0, 8))

    def _extract_compte_code(self):
        raw = self.compte_var.get().strip()
        if " — " in raw:
            return raw.split(" — ", 1)[0].strip()
        return raw

    def _open_dropdown(self, event=None):
        widget = event.widget if event else None
        if widget is not None:
            widget.event_generate("<Down>")

    def _refresh_compte_values(self):
        accounts = core.search_accounts(self.conn, "", limit=300)
        self.compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in accounts]

    def _on_compte_keyrelease(self, event=None):
        if event is not None and event.keysym in ("Up", "Down", "Return", "Tab"):
            return
        query = self._extract_compte_code()
        if query:
            matches = core.search_accounts(self.conn, query, limit=30)
            self.compte_combo["values"] = [f"{m['code']} — {m['label']}" for m in matches]

    def _reset(self):
        self.compte_var.set("")
        self.tiers_var.set("")
        self.refresh()

    def refresh(self):
        self._refresh_compte_values()
        for row in self.tree.get_children():
            self.tree.delete(row)
        compte_prefix = self._extract_compte_code() or None
        tiers = self.tiers_var.get().strip() or None
        classes = core.compute_grand_livre_complet(self.conn, compte_prefix=compte_prefix, tiers=tiers)

        def f(v):
            return f"{fmt_cfa(v)}" if v else ""

        for c in classes:
            classe_ouv_debit = classe_ouv_credit = classe_sold_debit = classe_sold_credit = 0.0
            for compte in c["comptes"]:
                self.tree.insert("", "end", tags=("compte_header",), values=(
                    "", "", "", f"{compte['code']} — {compte['label']}", "", "", "", "", "", "",
                ))
                for l in compte["lignes"]:
                    self.tree.insert("", "end", values=(
                        core.to_display_date(l["date"]), l["piece"] or "", l["journal"] or "",
                        l["libelle"] or "", "", "",
                        f(l["debit"]), f(l["credit"]), "", "",
                    ))
                ouv = compte["solde_ouverture"]
                ouv_debit = ouv if ouv > 0 else 0.0
                ouv_credit = -ouv if ouv < 0 else 0.0
                sold = compte["solde_final"]
                sold_debit = sold if sold > 0 else 0.0
                sold_credit = -sold if sold < 0 else 0.0
                self.tree.insert("", "end", tags=("compte_total",), values=(
                    "", "", "", f"TOTAL COMPTE {compte['code']} — Solde {compte['sens']}",
                    f(ouv_debit), f(ouv_credit), f(compte["total_debit"]), f(compte["total_credit"]),
                    f(sold_debit), f(sold_credit),
                ))
                classe_ouv_debit += ouv_debit
                classe_ouv_credit += ouv_credit
                classe_sold_debit += sold_debit
                classe_sold_credit += sold_credit
            self.tree.insert("", "end", tags=("classe_total",), values=(
                "", "", "", f"TOTAL CLASSE {c['classe']}",
                f(classe_ouv_debit), f(classe_ouv_credit), f(c["total_debit"]), f(c["total_credit"]),
                f(classe_sold_debit), f(classe_sold_credit),
            ))


class OpeningBalancesTab(ttk.Frame):
    """Soldes d'ouverture (report à nouveau) : un solde signé par compte, saisi
    une fois en début d'exercice. Débiteur = positif, créditeur = négatif."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn

        ttk.Label(self, text=(
            "Saisissez ici le solde de report à nouveau de chaque compte de bilan au 1er jour de "
            "l'exercice (= solde de clôture de l'exercice précédent). Convention : solde débiteur = "
            "positif, solde créditeur = négatif (ex. Capital social créditeur de 5 000 000 → -5000000). "
            "La « Balance de clôture » (onglet Balance) et le Bilan intègrent automatiquement ces soldes."
        ), foreground="#595959", wraplength=1000).pack(anchor="w", padx=8, pady=(8, 4))

        import_bar = ttk.Frame(self)
        import_bar.pack(fill="x", padx=8, pady=(0, 4))
        ttk.Button(import_bar, text="Télécharger un modèle (.xlsx)", command=self.download_template).pack(side="left", padx=2)
        ttk.Button(import_bar, text="Importer la balance N-1 (.xlsx) — ÉCRASE la balance actuelle",
                   command=self.import_xlsx).pack(side="left", padx=2)
        ttk.Button(import_bar, text="Exporter la balance N-1 (.xlsx)", command=self.export_xlsx).pack(side="left", padx=2)
        ttk.Label(import_bar, text="(Colonnes attendues : N° Compte, Libellé, Solde — l'écrasement "
                                    "ne concerne que l'exercice comptable actuellement sélectionné.)",
                  foreground="#595959").pack(side="left", padx=10)

        form = ttk.Frame(self)
        form.pack(fill="x", padx=8, pady=4)
        ttk.Label(form, text="N° Compte :").pack(side="left")
        self.compte_var = tk.StringVar()
        self.compte_combo = ttk.Combobox(form, textvariable=self.compte_var, width=34)
        self.compte_combo.pack(side="left", padx=4)
        self.compte_combo.bind("<KeyRelease>", self._on_compte_keyrelease)
        ttk.Label(form, text="Solde d'ouverture :").pack(side="left", padx=(12, 0))
        self.solde_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.solde_var, width=16).pack(side="left", padx=4)
        ttk.Button(form, text="Enregistrer", command=self.save).pack(side="left", padx=6)

        cols = ("code", "label", "debit", "credit")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        headers = ["N° Compte", "Libellé", "Débit", "Crédit"]
        widths = [90, 380, 130, 130]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w" if c in ("code", "label") else "e")
        self.tree.tag_configure("total", background="#1F4E78", foreground="white", font=("Segoe UI", 10, "bold"))
        self.tree.pack(fill="both", padx=8, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        bottom = ttk.Frame(self)
        bottom.pack(fill="x", padx=8, pady=(0, 8))
        self.total_var = tk.StringVar()
        ttk.Label(bottom, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(side="left")
        self.refresh()

    def _extract_compte_code(self):
        raw = self.compte_var.get().strip()
        if " — " in raw:
            return raw.split(" — ", 1)[0].strip()
        return raw

    def _on_compte_keyrelease(self, event=None):
        if event is not None and event.keysym in ("Up", "Down", "Return", "Tab"):
            return
        query = self._extract_compte_code()
        if query:
            matches = core.search_accounts(self.conn, query, limit=30)
            self.compte_combo["values"] = [f"{m['code']} — {m['label']}" for m in matches]

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        values = self.tree.item(sel[0], "values")
        if values[0] == "":  # ligne de total, pas un compte
            return
        self.compte_var.set(values[0])
        try:
            debit = float(values[2].replace(" ", "").replace(",", "")) if values[2] else 0.0
            credit = float(values[3].replace(" ", "").replace(",", "")) if values[3] else 0.0
        except ValueError:
            debit = credit = 0.0
        self.solde_var.set(str(debit - credit))

    def save(self):
        code = self._extract_compte_code()
        if not code:
            messagebox.showinfo("Info", "Choisissez d'abord un compte.")
            return
        try:
            value = float(self.solde_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Le solde d'ouverture doit être un nombre.")
            return
        core.set_opening_balance(self.conn, code, value)
        self.compte_var.set("")
        self.solde_var.set("")
        self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        total_debit = total_credit = 0.0
        for b in core.list_opening_balances(self.conn):
            solde = b["solde"]
            debit = solde if solde > 0 else 0.0
            credit = -solde if solde < 0 else 0.0
            self.tree.insert("", "end", values=(
                b["code"], b["label"], f"{fmt_cfa(debit)}" if debit else "", f"{fmt_cfa(credit)}" if credit else ""))
            total_debit += debit
            total_credit += credit
        self.tree.insert("", "end", tags=("total",), values=(
            "", "TOTAL", f"{fmt_cfa(total_debit)}", f"{fmt_cfa(total_credit)}"))
        ecart = total_debit - total_credit
        equilibre = "Équilibré ✓" if abs(ecart) < 0.01 else "NON ÉQUILIBRÉ ✗ (Débit doit être égal à Crédit)"
        self.total_var.set(f"Total Débit : {fmt_cfa(total_debit)}   —   Total Crédit : {fmt_cfa(total_credit)}   "
                            f"—   {equilibre}")

    def download_template(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Modele_balance_N-1.xlsx", title="Enregistrer le modèle de balance d'ouverture",
        )
        if not path:
            return
        core.export_opening_balances_template(path)
        messagebox.showinfo("Modèle créé", f"Modèle enregistré :\n{path}\n\n"
                                            f"Remplissez-le (une ligne par compte), puis utilisez "
                                            f"« Importer la balance N-1 ».")

    def export_xlsx(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Balance_ouverture.xlsx", title="Exporter la balance d'ouverture",
        )
        if not path:
            return
        core.export_opening_balances_xlsx(self.conn, path)
        messagebox.showinfo("Export terminé", f"Balance d'ouverture exportée :\n{path}")

    def import_xlsx(self):
        path = filedialog.askopenfilename(filetypes=[("Classeur Excel", "*.xlsx")],
                                           title="Importer une balance d'ouverture (N-1)")
        if not path:
            return
        exercice = core.get_current_exercice(self.conn)
        if not messagebox.askyesno(
            "Confirmer l'écrasement",
            f"Importer ce fichier va ÉCRASER complètement les soldes d'ouverture de l'exercice "
            f"{exercice} actuellement sélectionné. Cette action est irréversible. Continuer ?"
        ):
            return
        try:
            n, warnings = core.import_opening_balances_xlsx(self.conn, path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de l'import : {exc}")
            return
        self.refresh()
        msg = f"{n} solde(s) importé(s) pour l'exercice {exercice}. La balance précédente a été remplacée."
        if warnings:
            msg += "\n\nAvertissements :\n" + "\n".join(warnings[:20])
        messagebox.showinfo("Import terminé", msg)


class StocksTab(ttk.Frame):
    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        inner = ttk.Notebook(self)
        inner.pack(fill="both", expand=True)
        self.synthese_tab = StocksSyntheseTab(inner, conn)
        self.mouvements_tab = StocksMouvementsTab(inner, conn)
        inner.add(self.synthese_tab, text="Synthèse par compte")
        inner.add(self.mouvements_tab, text="Mouvements comptables (classe 3)")

    def refresh(self):
        self.synthese_tab.refresh()
        self.mouvements_tab.refresh()


class StocksSyntheseTab(ttk.Frame):
    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text=(
            "Détail RÉEL de chaque compte de stock utilisé (pas seulement les comptes centralisateurs "
            "310000/320000/331000/360000) : tout sous-compte 31x/32x/33x/36x ayant un mouvement ou un "
            "stock initial apparaît ici (ex. 321001 CLINKER). Cliquez une ligne, modifiez la valeur puis "
            "« Enregistrer ». La quantité de mouvement provient du champ « Quantité » saisi sur chaque "
            "écriture (onglet Saisie) — elle permet de calculer un coût unitaire moyen réel."
        ), foreground="#595959", wraplength=1050).pack(anchor="w", padx=8, pady=(8, 0))

        filt = ttk.Frame(self)
        filt.pack(fill="x", padx=8, pady=4)
        ttk.Label(filt, text="Catégorie :").pack(side="left")
        self.categorie_var = tk.StringVar(value="Toutes")
        ttk.Combobox(filt, textvariable=self.categorie_var, width=28, state="readonly", values=[
            "Toutes", "31 — Marchandises", "32 — Matières premières",
            "33 — Autres approvisionnements", "36 — Produits finis",
        ]).pack(side="left", padx=4)
        ttk.Button(filt, text="Filtrer", command=self.refresh).pack(side="left", padx=8)

        ttk.Label(filt, text="Marge de valorisation des produits finis par défaut (%) :").pack(side="left", padx=(24, 4))
        self.marge_defaut_var = tk.StringVar(value=str(core.get_setting(conn, "marge_production_defaut", 30.0)))
        ttk.Entry(filt, textvariable=self.marge_defaut_var, width=6).pack(side="left", padx=2)
        ttk.Button(filt, text="Enregistrer la marge", command=self.save_marge_defaut).pack(side="left", padx=4)

        edit_bar = ttk.Frame(self)
        edit_bar.pack(fill="x", padx=8, pady=4)
        ttk.Label(edit_bar, text="Stock initial (valeur) du compte sélectionné :").pack(side="left")
        self.initial_var = tk.StringVar()
        ttk.Entry(edit_bar, textvariable=self.initial_var, width=14).pack(side="left", padx=4)
        ttk.Button(edit_bar, text="Enregistrer la valeur", command=self.save_initial).pack(side="left", padx=4)
        ttk.Label(edit_bar, text="Quantité initiale :").pack(side="left", padx=(16, 0))
        self.qte_initial_var = tk.StringVar()
        ttk.Entry(edit_bar, textvariable=self.qte_initial_var, width=14).pack(side="left", padx=4)
        ttk.Button(edit_bar, text="Enregistrer la quantité", command=self.save_qte_initial).pack(side="left", padx=4)
        ttk.Label(edit_bar, text="(pour un nouveau compte : tapez son n° ci-dessus dans le champ, puis "
                                  "enregistrez — il apparaîtra dans la liste)", foreground="#595959").pack(side="left", padx=8)

        cols = ("code", "label", "initial", "entrees", "sorties", "final",
                "qte_initiale", "qte_entrees", "qte_sorties", "qte_finale", "cump")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        headers = ["N° Compte", "Libellé", "Stock initial", "Entrées (Débit)", "Sorties (Crédit)", "Stock final",
                   "Qté initiale", "Qté entrées", "Qté sorties", "Qté finale", "Coût unit. moyen"]
        widths = [90, 190, 100, 100, 100, 100, 80, 80, 80, 80, 110]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=8, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.selected_code = None
        self.refresh()

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        values = self.tree.item(sel[0], "values")
        self.selected_code = values[0]
        self.initial_var.set(values[2])
        self.qte_initial_var.set(values[6])

    def save_marge_defaut(self):
        try:
            value = float(self.marge_defaut_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "La marge doit être un nombre.")
            return
        core.set_setting(self.conn, "marge_production_defaut", value)
        messagebox.showinfo("Enregistré", "Marge de valorisation par défaut enregistrée.")

    def save_initial(self):
        if not self.selected_code:
            messagebox.showinfo("Info", "Sélectionnez d'abord un compte de stock dans le tableau "
                                         "(ou saisissez son code dans le champ ci-dessus après l'avoir tapé).")
            return
        try:
            value = float(self.initial_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Le stock initial doit être un nombre.")
            return
        core.set_stock_initial(self.conn, self.selected_code, value)
        self.refresh()

    def save_qte_initial(self):
        if not self.selected_code:
            messagebox.showinfo("Info", "Sélectionnez d'abord un compte de stock dans le tableau.")
            return
        try:
            value = float(self.qte_initial_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "La quantité initiale doit être un nombre.")
            return
        core.set_stock_qte_initiale(self.conn, self.selected_code, value)
        self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        cat = self.categorie_var.get()
        prefixes = None
        if cat != "Toutes":
            prefixes = [cat.split(" — ")[0].strip()]
        for s in core.compute_stocks_detail(self.conn, prefixes=prefixes):
            cump = f"{fmt_cfa(s['cout_unitaire_moyen'])}" if s["cout_unitaire_moyen"] is not None else "—"
            self.tree.insert("", "end", values=(
                s["code"], s["label"], f"{fmt_cfa(s['stock_initial'])}",
                f"{fmt_cfa(s['entrees'])}", f"{fmt_cfa(s['sorties'])}", f"{fmt_cfa(s['stock_final'])}",
                f"{s['qte_initiale']:g}", f"{s['qte_entrees']:g}", f"{s['qte_sorties']:g}",
                f"{s['qte_finale']:g}", cump,
            ))


class StocksMouvementsTab(ttk.Frame):
    """Détail de toutes les écritures comptables sur les comptes de stock
    (classe 3), avec leur origine : générées automatiquement par la
    Facturation (ventes) ou les Factures frs (achats), ou saisies
    manuellement dans l'onglet Saisie."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text=(
            "Tous les mouvements comptables des comptes de stock (310000, 320000, 331000, 360000) "
            "de l'exercice en cours, y compris ceux générés automatiquement par la validation d'une "
            "facture de vente (Commerce → Facturation) ou d'une facture d'achat (Engagements-projets "
            "→ Factures frs)."
        ), foreground="#595959", wraplength=1050).pack(anchor="w", padx=8, pady=(8, 4))

        filt = ttk.Frame(self)
        filt.pack(fill="x", padx=8, pady=4)
        ttk.Label(filt, text="Filtrer par origine :").pack(side="left")
        self.origine_var = tk.StringVar(value="Toutes")
        ttk.Combobox(filt, textvariable=self.origine_var, width=18, state="readonly",
                     values=["Toutes", "Facturation", "Facture frs", "Saisie directe (auto)", "Saisie manuelle"]).pack(side="left", padx=4)
        ttk.Button(filt, text="Filtrer", command=self.refresh).pack(side="left", padx=8)

        cols = ("date", "piece", "compte", "compte_label", "libelle", "debit", "credit", "quantite",
                "qte_cumulee", "valeur_cumulee", "origine")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        headers = ["Date", "Pièce", "Compte", "Libellé du compte", "Libellé écriture",
                   "Débit (valeur)", "Crédit (valeur)", "Qté mvt", "Qté cumulée", "Valeur cumulée", "Origine"]
        widths = [90, 80, 80, 150, 190, 90, 90, 70, 90, 100, 110]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.tag_configure("auto", foreground="#1F4E78")
        self.tree.pack(fill="both", padx=8, pady=8)

        self.totals_var = tk.StringVar()
        ttk.Label(self, textvariable=self.totals_var, font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=8, pady=(0, 8))
        self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        mouvements = core.compute_mouvements_stocks(self.conn)
        filtre = self.origine_var.get()
        total_d = total_c = 0.0
        for m in mouvements:
            if filtre != "Toutes" and m["origine"] != filtre:
                continue
            tags = ("auto",) if m["origine"] != "Saisie manuelle" else ()
            self.tree.insert("", "end", tags=tags, values=(
                core.to_display_date(m["date"]), m["piece"] or "", m["compte"], m["compte_label"],
                m["libelle"] or "", f"{fmt_cfa(m['debit'])}" if m["debit"] else "",
                f"{fmt_cfa(m['credit'])}" if m["credit"] else "", f"{m['quantite']:g}" if m["quantite"] else "",
                f"{m['qte_cumulee']:g}", f"{fmt_cfa(m['valeur_cumulee'])}",
                m["origine"],
            ))
            total_d += m["debit"]
            total_c += m["credit"]
        self.totals_var.set(f"TOTAL — Débit : {fmt_cfa(total_d)}   Crédit : {fmt_cfa(total_c)}")


class CoutsFabricationPeriodeTab(ttk.Frame):
    """Coûts de fabrication réels de la période (écritures taguées AN-FAB)."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text=(
            "Pour qu'une charge remonte ici, saisissez le code analytique « AN-FAB » "
            "sur la ligne correspondante dans l'onglet Saisie."
        ), foreground="#595959").pack(anchor="w", padx=8, pady=(8, 0))
        self.text = tk.Text(self, font=("Consolas", 11), wrap="none")
        self.text.pack(fill="both", padx=8, pady=8)
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(pady=(0, 8))
        self.refresh()

    def refresh(self):
        p = core.compute_production(self.conn)
        lines = ["PRODUCTION DE L'EXERCICE", "=" * 60,
                 f"  {'Ventes (produits finis, travaux, services)':<50} {p['ventes']:>12,.2f}",
                 f"  {'Production stockée (variation stock 360000)':<50} {p['production_stockee']:>12,.2f}",
                 f"  {'VALEUR DE LA PRODUCTION':<50} {p['valeur_production']:>12,.2f}",
                 "", "COÛTS DE FABRICATION (axe AN-FAB)", "=" * 60]
        for poste in p["postes_cout"]:
            lines.append(f"  {poste['label']:<50} {poste['montant']:>12,.2f}")
        lines += [f"  {'COÛT DE PRODUCTION':<50} {p['cout_production']:>12,.2f}", "",
                  f"MARGE SUR COÛT DE PRODUCTION{'':<34}{p['marge']:>12,.2f}"]
        self.text.delete("1.0", "end")
        self.text.insert("1.0", "\n".join(lines))


class RecetteFabricationTab(ttk.Frame):
    """Nomenclature de fabrication (BOM) : combine matières premières (coût
    réel des stocks), main-d'œuvre et énergie pour calculer le coût de
    production d'un produit fini, puis le prix de vente suggéré (+ marge)."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.selected_produit = None

        top = ttk.Frame(self)
        top.pack(fill="x", padx=12, pady=8)
        ttk.Label(top, text="Produit fini :").pack(side="left")
        self.produit_var = tk.StringVar()
        self.produit_combo = ttk.Combobox(top, textvariable=self.produit_var, width=30, state="readonly")
        self.produit_combo.pack(side="left", padx=4)
        self.produit_combo.bind("<<ComboboxSelected>>", self._on_produit_selected)
        ttk.Button(top, text="Nouveau produit fini", command=self._new_produit).pack(side="left", padx=8)
        ttk.Button(top, text="Supprimer ce produit", command=self._delete_produit).pack(side="left", padx=2)

        params = ttk.Frame(self)
        params.pack(fill="x", padx=12, pady=(0, 8))
        ttk.Label(params, text="Quantité produite par recette :").pack(side="left")
        self.qte_produite_var = tk.StringVar()
        ttk.Entry(params, textvariable=self.qte_produite_var, width=8).pack(side="left", padx=4)
        ttk.Label(params, text="Marge (%) :").pack(side="left", padx=(16, 0))
        self.marge_var = tk.StringVar()
        ttk.Entry(params, textvariable=self.marge_var, width=8).pack(side="left", padx=4)
        ttk.Label(params, text="Compte stock produit fini (classe 36) :").pack(side="left", padx=(16, 0))
        self.compte_stock_pf_var = tk.StringVar()
        self.compte_stock_pf_combo = ttk.Combobox(params, textvariable=self.compte_stock_pf_var, width=26)
        self.compte_stock_pf_combo.pack(side="left", padx=4)
        self.compte_stock_pf_combo.bind("<KeyRelease>", self._on_compte_pf_keyrelease)
        self._refresh_compte_pf_values()
        ttk.Button(params, text="Enregistrer ces paramètres", command=self._save_params).pack(side="left", padx=8)

        form = ttk.LabelFrame(self, text="Ajouter un composant à la recette")
        form.pack(fill="x", padx=12, pady=4)
        ttk.Label(form, text="Type :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.type_var = tk.StringVar(value="matiere")
        type_combo = ttk.Combobox(form, textvariable=self.type_var, width=22, state="readonly",
                                   values=list(core.LIGNE_TYPES.values()))
        type_combo.set(core.LIGNE_TYPES["matiere"])
        type_combo.grid(row=0, column=1, padx=4)
        type_combo.bind("<<ComboboxSelected>>", self._on_type_changed)
        self.type_combo = type_combo

        ttk.Label(form, text="Libellé :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.libelle_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.libelle_var, width=22).grid(row=0, column=3, padx=4)

        self.compte_label = ttk.Label(form, text="Compte de stock :")
        self.compte_label.grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.compte_var = tk.StringVar()
        self.compte_combo = ttk.Combobox(form, textvariable=self.compte_var, width=26)
        self.compte_combo.grid(row=0, column=5, padx=4)
        self.compte_combo.bind("<KeyRelease>", self._on_compte_keyrelease)
        self.compte_combo.bind("<<ComboboxSelected>>", self._on_compte_changed)
        self.compte_combo.bind("<FocusOut>", self._on_compte_changed)
        self._refresh_stock_accounts()

        self.ligne_qte_label = ttk.Label(form, text="Quantité :")
        self.ligne_qte_label.grid(row=1, column=0, sticky="w", padx=4, pady=4)
        self.ligne_qte_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_qte_var, width=10).grid(row=1, column=1, padx=4, sticky="w")

        self.cout_label = ttk.Label(form, text="Coût unitaire (si pas de compte de stock) :")
        self.cout_label.grid(row=1, column=2, sticky="w", padx=(12, 4))
        self.ligne_cout_var = tk.StringVar()
        self.cout_entry = ttk.Entry(form, textvariable=self.ligne_cout_var, width=12)
        self.cout_entry.grid(row=1, column=3, padx=4, sticky="w")

        self.compte_apercu_var = tk.StringVar()
        ttk.Label(form, textvariable=self.compte_apercu_var, foreground="#1F7A1F", wraplength=480,
                  justify="left").grid(row=2, column=4, columnspan=2, sticky="w", padx=(12, 4))

        self.analytic_label = ttk.Label(form, text="Code analytique (Énergie/Maintenance...) :")
        self.analytic_label.grid(row=1, column=4, sticky="w", padx=(12, 4))
        self.ligne_analytic_var = tk.StringVar()
        self.analytic_combo = ttk.Combobox(form, textvariable=self.ligne_analytic_var, width=26)
        self.analytic_combo.grid(row=1, column=5, padx=4, sticky="w")
        self.analytic_combo.bind("<<ComboboxSelected>>", self._on_analytic_changed)
        self.analytic_combo.bind("<FocusOut>", self._on_analytic_changed)
        self._refresh_analytic_values()

        self.analytic_apercu_var = tk.StringVar()
        ttk.Label(form, textvariable=self.analytic_apercu_var, foreground="#1F7A1F", wraplength=480,
                  justify="left").grid(row=3, column=4, columnspan=2, sticky="w", padx=(12, 4))

        ttk.Button(form, text="Ajouter le composant", command=self.add_ligne).grid(row=4, column=5, padx=4, pady=4)

        cols = ("id", "type", "libelle", "compte", "quantite", "cout_unitaire", "analytique", "source", "montant")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=8)
        headers = ["ID", "Type", "Libellé", "Compte", "Quantité", "Coût unitaire", "Code analytique",
                   "Origine du coût", "Montant"]
        widths = [40, 90, 180, 90, 80, 110, 130, 170, 110]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", padx=12, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_ligne_select)
        ttk.Button(self, text="Supprimer le composant sélectionné", command=self.delete_ligne).pack(
            anchor="w", padx=12)

        self.result_text = tk.Text(self, font=("Consolas", 11), height=8, wrap="none")
        self.result_text.pack(fill="x", padx=12, pady=8)

        ttk.Button(self, text="Valider la fabrication (comptabiliser)", command=self.valider_fabrication).pack(
            anchor="w", padx=12, pady=(0, 8))

        self._on_type_changed()
        self.refresh_produits()

    def _refresh_stock_accounts(self):
        if self.type_var.get() == core.LIGNE_TYPES["amortissement"]:
            immos = core.compute_immobilisations_liste(self.conn)
            self.compte_combo["values"] = [f"{i['compte']} — {i['libelle']}" for i in immos]
        else:
            stocks = core.compute_stocks_detail(self.conn, prefixes=["31", "32", "33", "36"])
            self.compte_combo["values"] = [f"{s['code']} — {s['label']}" for s in stocks]

    def _on_compte_keyrelease(self, event=None):
        query = self._extract_code(self.compte_var.get())
        if not query:
            return
        if self.type_var.get() == core.LIGNE_TYPES["amortissement"]:
            items = [a for a in core.search_accounts(self.conn, query, limit=30) if a["classe"] == "2"]
        else:
            items = core.search_accounts(self.conn, query, limit=30)
        self.compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _on_compte_changed(self, event=None):
        """Aperçu automatique du coût unitaire dès qu'un compte est choisi :
        coût moyen du stock (matière première), ou coût d'amortissement par
        unité d'usage — tonne, heure... (amortissement d'équipement)."""
        code = self._extract_code(self.compte_var.get())
        if not code:
            self.compte_apercu_var.set("")
            return
        if self.type_var.get() == core.LIGNE_TYPES["amortissement"]:
            fiche = core.get_immobilisation_fiche(self.conn, code)
            base = fiche.get("base_repartition_quantite")
            unite = fiche.get("base_repartition_unite") or "unité"
            if not base:
                self.compte_apercu_var.set(
                    f"Base de répartition non renseignée pour ce compte — allez dans IMMOBILISATIONS, "
                    f"sélectionnez « {code} » et indiquez sa quantité annuelle de référence (ex. "
                    f"5000 tonnes/an ou 2000 heures/an).")
                return
            cu = core.compute_cout_amortissement_unitaire(self.conn, code)
            if cu is not None:
                self.compte_apercu_var.set(
                    f"Coût d'amortissement : {fmt_cfa(cu)} F CFA / {unite} "
                    f"(amortissement de la période ÷ {base:g} {unite}/an — sera utilisé automatiquement)")
            else:
                self.compte_apercu_var.set(
                    "Aucun amortissement comptabilisé pour cet équipement pour l'instant — saisissez un "
                    "coût unitaire manuel en attendant.")
        else:
            stocks_by_code = {s["code"]: s for s in core.compute_stocks_detail(self.conn)}
            stock = stocks_by_code.get(code)
            if stock and stock["cout_unitaire_moyen"] is not None:
                self.compte_apercu_var.set(
                    f"Coût unitaire moyen en stock : {fmt_cfa(stock['cout_unitaire_moyen'])} F CFA — "
                    f"{stock['qte_finale']:g} unité(s) disponible(s) (sera utilisé automatiquement)")
            else:
                self.compte_apercu_var.set(
                    "Aucune quantité en stock pour ce compte pour l'instant — saisissez un coût unitaire "
                    "manuel en attendant, ou renseignez d'abord un stock initial (onglet Stocks).")

    def _refresh_analytic_values(self):
        codes = core.list_analytic_codes(self.conn)
        self.analytic_combo["values"] = [f"{c['code']} — {c['label']}" for c in codes]

    def _on_analytic_changed(self, event=None):
        """Aperçu du coût unitaire moyen pondéré du code analytique choisi
        (ex. F CFA par heure de maintenance, par litre d'eau) — calculé à
        partir de l'ensemble des charges déjà comptabilisées sous ce code,
        comme le sera réellement le coût utilisé dans la recette."""
        code = self._extract_code(self.ligne_analytic_var.get())
        if not code:
            self.analytic_apercu_var.set("")
            self.ligne_qte_label.configure(text="Quantité :")
            return
        unite = core.get_analytic_code_unite(self.conn, code)
        cu = core.compute_cout_unitaire_moyen_analytique(self.conn, code, toutes_dates=True)
        if cu is not None:
            self.analytic_apercu_var.set(
                f"Coût moyen pondéré constaté : {fmt_cfa(cu)} F CFA / {unite or 'unité'} "
                f"(sera utilisé automatiquement)")
        else:
            self.analytic_apercu_var.set(
                f"Aucune quantité comptabilisée sous ce code pour l'instant — "
                f"saisissez un coût unitaire manuel en attendant.")
        self.ligne_qte_label.configure(text=f"Quantité ({unite}) :" if unite else "Quantité :")

    def _refresh_compte_pf_values(self):
        stocks = core.compute_stocks_detail(self.conn, prefixes=["36"])
        values = [f"{s['code']} — {s['label']}" for s in stocks]
        if "360000 — PRODUITS FINIS" not in values and core.account_exists(self.conn, "360000"):
            values.insert(0, f"360000 — {core.get_account_label(self.conn, '360000')}")
        self.compte_stock_pf_combo["values"] = values

    def _on_compte_pf_keyrelease(self, event=None):
        query = self._extract_code(self.compte_stock_pf_var.get())
        if query:
            items = [a for a in core.search_accounts(self.conn, query, limit=50)
                     if a["code"].startswith("36")]
            self.compte_stock_pf_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _on_type_changed(self, event=None):
        type_key = self.type_var.get()
        actif = type_key in (core.LIGNE_TYPES["matiere"], core.LIGNE_TYPES["amortissement"])
        self.compte_combo.configure(state="normal" if actif else "disabled")
        if actif:
            self.compte_label.configure(
                text="Compte d'immobilisation :" if type_key == core.LIGNE_TYPES["amortissement"]
                else "Compte de stock :")
            self._refresh_stock_accounts()
        else:
            self.compte_var.set("")
            self.compte_apercu_var.set("")

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def _type_key(self):
        label = self.type_combo.get()
        for key, val in core.LIGNE_TYPES.items():
            if val == label:
                return key
        return "autre"

    def refresh_produits(self):
        produits = core.list_produits_finis(self.conn)
        self.produit_combo["values"] = [f"{p['code']} — {p['nom']}" for p in produits]
        if produits and not self.selected_produit:
            self.selected_produit = produits[0]["code"]
            self.produit_var.set(f"{produits[0]['code']} — {produits[0]['nom']}")
        self.refresh()

    def _on_produit_selected(self, event=None):
        self.selected_produit = self._extract_code(self.produit_var.get())
        self.refresh()

    def _new_produit(self):
        code = simpledialog.askstring("Nouveau produit fini", "Code du produit :", parent=self)
        if not code:
            return
        nom = simpledialog.askstring("Nouveau produit fini", "Nom du produit :", parent=self)
        if not nom:
            return
        marge_defaut = core.get_setting(self.conn, "marge_production_defaut", 30.0)
        core.add_produit_fini(self.conn, code.strip(), nom.strip(), marge_pourcentage=marge_defaut)
        self.selected_produit = code.strip()
        self.refresh_produits()

    def _delete_produit(self):
        if not self.selected_produit:
            return
        if messagebox.askyesno("Confirmer", f"Supprimer le produit « {self.selected_produit} » et sa recette ?"):
            core.delete_produit_fini(self.conn, self.selected_produit)
            self.selected_produit = None
            self.refresh_produits()

    def _save_params(self):
        if not self.selected_produit:
            return
        produit = core.get_produit_fini(self.conn, self.selected_produit)
        try:
            qte = float(self.qte_produite_var.get() or 1)
            marge = float(self.marge_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Quantité produite et marge doivent être des nombres.")
            return
        compte_stock = self._extract_code(self.compte_stock_pf_var.get()) or "360000"
        if not core.account_exists(self.conn, compte_stock):
            messagebox.showerror("Compte invalide", f"Le compte « {compte_stock} » n'existe pas.")
            return
        core.add_produit_fini(self.conn, self.selected_produit, produit["nom"], produit["description"] or "",
                               qte, marge, compte_stock)
        self.refresh()

    def _on_ligne_select(self, event=None):
        pass

    def add_ligne(self):
        if not self.selected_produit:
            messagebox.showinfo("Info", "Créez ou sélectionnez d'abord un produit fini.")
            return
        libelle = self.libelle_var.get().strip()
        if not libelle:
            messagebox.showwarning("Champ manquant", "Le libellé du composant est obligatoire.")
            return
        try:
            qte = float(self.ligne_qte_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "La quantité doit être un nombre.")
            return
        type_key = self._type_key()
        compte = self._extract_code(self.compte_var.get()) if type_key in ("matiere", "amortissement") else None
        cout_unitaire = None
        if self.ligne_cout_var.get().strip():
            try:
                cout_unitaire = float(self.ligne_cout_var.get())
            except ValueError:
                messagebox.showerror("Erreur", "Le coût unitaire doit être un nombre.")
                return
        if type_key in ("matiere", "amortissement") and not compte and cout_unitaire is None:
            messagebox.showwarning("Champ manquant",
                                    "Choisissez un compte (stock ou immobilisation) ou saisissez un coût "
                                    "unitaire manuel.")
            return
        analytic_code = self._extract_code(self.ligne_analytic_var.get()) or None
        core.add_recette_ligne(self.conn, self.selected_produit, type_key, libelle, qte, compte, cout_unitaire,
                                analytic_code=analytic_code)
        self.libelle_var.set("")
        self.ligne_qte_var.set("")
        self.ligne_cout_var.set("")
        self.ligne_analytic_var.set("")
        self.refresh()

    def delete_ligne(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez d'abord un composant dans le tableau.")
            return
        ligne_id = int(self.tree.item(sel[0], "values")[0])
        core.delete_recette_ligne(self.conn, ligne_id)
        self.refresh()

    def refresh(self):
        self._refresh_stock_accounts()
        self._refresh_analytic_values()
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.result_text.delete("1.0", "end")
        if not self.selected_produit or not core.get_produit_fini(self.conn, self.selected_produit):
            return
        produit = core.get_produit_fini(self.conn, self.selected_produit)
        self.qte_produite_var.set(str(produit["quantite_produite"]))
        self.marge_var.set(str(produit["marge_pourcentage"]))
        self._refresh_compte_pf_values()
        label_pf = core.get_account_label(self.conn, produit["compte_stock"])
        self.compte_stock_pf_var.set(f"{produit['compte_stock']} — {label_pf}")

        resultat = core.compute_cout_production(self.conn, self.selected_produit)
        for l in resultat["lignes"]:
            self.tree.insert("", "end", values=(
                l["id"], core.LIGNE_TYPES.get(l["type_ligne"], l["type_ligne"]), l["libelle"],
                l["compte"] or "", f"{l['quantite']:g}", f"{fmt_cfa(l['cout_unitaire_utilise'])}",
                l.get("analytic_code") or "",
                l["source_cout"], f"{fmt_cfa(l['montant'])}",
            ))
        lines = [
            f"COÛT DE PRODUCTION — {produit['nom']} ({self.selected_produit})", "=" * 70,
            f"  {'Coût de production total (recette)':<45} {resultat['cout_production_total']:>15,.2f}",
            f"  {'Quantité produite':<45} {resultat['quantite_produite']:>15,g}",
            f"  {'COÛT DE PRODUCTION UNITAIRE':<45} {resultat['cout_unitaire_produit']:>15,.2f}", "",
            f"  {'Marge appliquée':<45} {resultat['marge_pourcentage']:>14,g} %",
            f"  {'PRIX DE VENTE UNITAIRE SUGGÉRÉ':<45} {resultat['prix_vente_unitaire']:>15,.2f}",
            f"  {'dont marge unitaire':<45} {resultat['marge_unitaire']:>15,.2f}",
        ]
        self.result_text.insert("1.0", "\n".join(lines))

    def valider_fabrication(self):
        if not self.selected_produit:
            messagebox.showinfo("Info", "Sélectionnez d'abord un produit fini.")
            return
        self._save_params()
        resultat = core.compute_cout_production(self.conn, self.selected_produit)
        if not resultat["lignes"]:
            messagebox.showwarning("Recette vide", "Ajoutez au moins un composant à la recette avant de valider.")
            return
        if not messagebox.askyesno(
            "Confirmer la validation de la fabrication",
            f"Valider la fabrication de « {resultat['produit']['nom']} » ?\n\n"
            f"Coût de production : {fmt_cfa(resultat['cout_production_total'])}\n"
            f"Quantité produite : {resultat['quantite_produite']:g}\n"
            f"Valeur du produit fini mis en stock (coût + marge {resultat['marge_pourcentage']:g}%) : "
            f"{fmt_cfa(resultat['prix_vente_total'])}\n\n"
            f"Cette action va DIMINUER les matières premières consommées (quantité et valeur) et "
            f"AUGMENTER le stock de produit fini, avec envoi des écritures dans le menu SAISIE. "
            f"Cette action est définitive."
        ):
            return
        try:
            _, warnings = core.valider_fabrication(self.conn, self.selected_produit)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc))
            return
        msg = "Fabrication validée. Les matières premières ont été décrémentées et le produit fini mis en stock."
        if warnings:
            msg += "\n\nAvertissements :\n" + "\n".join(warnings)
        messagebox.showinfo("Validation terminée", msg)
        self.refresh()


class ProductionTab(ttk.Frame):
    """Regroupe la nomenclature de fabrication (coût de production, prix de
    vente) et le suivi des coûts réels de fabrication de la période."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        inner = ttk.Notebook(self)
        inner.pack(fill="both", expand=True)
        self.recette_tab = RecetteFabricationTab(inner, conn)
        self.periode_tab = CoutsFabricationPeriodeTab(inner, conn)
        inner.add(self.recette_tab, text="Recettes / Coût de production")
        inner.add(self.periode_tab, text="Coûts de fabrication (période)")

    def refresh(self):
        self.recette_tab.refresh_produits()
        self.periode_tab.refresh()


class TftIndirectTab(ttk.Frame):
    """TFT selon la méthode indirecte SYSCOHADA (avec CAFG), présenté selon le
    modèle officiel avec une couleur par section. Calculé à partir de
    compute_balance() et compute_liasse_resultat() — donc toujours cohérent
    avec la Balance et le Bilan. La ligne CONTRÔLE compare la trésorerie
    calculée à la trésorerie réelle de la Balance : un écart signale un
    mouvement mal classé."""

    SECTIONS = {
        "ouverture": "#D9D2E9",   # violet clair — trésorerie
        "cafg": "#D9EAD3",        # vert clair — CAFG / exploitation
        "invest": "#FCE5CD",      # orange clair — investissement
        "finance": "#CFE2F3",     # bleu clair — financement
        "controle": "#F4CCCC",    # rouge/rose clair — contrôle
        "total": "#1F4E78",       # bandeau total
    }

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        cols = ("libelle", "montant")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=28)
        self.tree.heading("libelle", text="Rubrique")
        self.tree.heading("montant", text="Montant")
        self.tree.column("libelle", width=480, anchor="w")
        self.tree.column("montant", width=160, anchor="e")
        for key, color in self.SECTIONS.items():
            fg = "white" if key == "total" else "black"
            tree_font = ("Segoe UI", 9, "bold") if key == "total" else ("Segoe UI", 9)
            self.tree.tag_configure(key, background=color, foreground=fg, font=tree_font)
            self.tree.tag_configure(key + "_header", background=color, foreground=fg, font=("Segoe UI", 9, "bold"))
        self.tree.pack(fill="both", padx=8, pady=8)
        self.ecart_var = tk.StringVar()
        ttk.Label(self, textvariable=self.ecart_var, font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=8)
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(pady=8)
        ttk.Button(self, text="Exporter (gabarit officiel .xlsx)",
                   command=lambda: export_etat_gabarit(
                       self, self.conn, "flux", "Flux_de_Tresorerie.xlsx", "le TFT")
                   ).pack(pady=(0, 8))
        self.refresh()

    def _row(self, tag, label, val):
        self.tree.insert("", "end", tags=(tag,), values=(f"  {label}", fmt_cfa(val)))

    def _header(self, tag, titre):
        self.tree.insert("", "end", tags=(tag + "_header",), values=(titre, ""))

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        t = core.compute_tft_indirect(self.conn)
        bfr_variation_totale = (t["variation_actif_circulant_hao"] + t["variation_stocks"]
                                 + t["variation_creances"] + t["variation_dettes_circulantes"])
        cessions_immo_incorp_corp = t["cessions_incorp"] + t["cessions_corp"]

        self._header("ouverture", "A — TRÉSORERIE NETTE AU 1ER JANVIER")
        self._row("ouverture", "Trésorerie nette au 1er janvier", t["treso_ouverture"])

        self._header("cafg", "DÉTERMINATION DE LA CAPACITÉ D'AUTOFINANCEMENT")
        self._row("cafg", "EBE", t["ebe"])
        self._row("cafg", "- Valeurs comptables de cession courantes d'immobilisations (654)",
                   t["valeurs_comptables_cessions_courantes"])
        self._row("cafg", "+ Produits de cession courantes d'immobilisations (754)", t["produits_cessions_courantes"])
        self._row("cafg", "+ Transfert de charges d'exploitation (781)", t["transferts_charges_exploitation"])
        self._row("cafg", "CAPACITÉ D'AUTOFINANCEMENT D'EXPLOITATION", t["caf_exploitation"])
        self._row("cafg", "+ Revenus financiers", t["revenus_financiers"])
        self._row("cafg", "- Frais financiers", t["frais_financiers"])
        self._row("cafg", "CAPACITÉ D'AUTOFINANCEMENT GLOBAL (CAFG)", t["cafg"])
        self._row("cafg", "- Variation d'actif circulant HAO (racines 46-49)", t["variation_actif_circulant_hao"])
        self._row("cafg", "- Variation des stocks", t["variation_stocks"])
        self._row("cafg", "- Variation des créances (racines 40-45)", t["variation_creances"])
        self._row("cafg", "+ Variation du passif circulant (racines 40-49)", t["variation_dettes_circulantes"])
        self._row("cafg", "Flux de trésorerie provenant des activités opérationnelles (Somme FA à FE)",
                   bfr_variation_totale)
        self._row("cafg", "B — Flux de trésorerie provenant des activités opérationnelles + CAFG",
                   t["flux_operationnel"])

        self._header("invest", "FLUX DE TRÉSORERIE PROVENANT DES ACTIVITÉS D'INVESTISSEMENT")
        self._row("invest", "- Décaissements liés aux acquisitions d'immobilisations incorporelles",
                   t["acquisitions_incorp"])
        self._row("invest", "- Décaissements liés aux acquisitions d'immobilisations corporelles",
                   t["acquisitions_corp"])
        self._row("invest", "- Décaissements liés aux acquisitions d'immobilisations financières",
                   t["acquisitions_fin"])
        self._row("invest", "+ Encaissements liés aux cessions d'immobilisations incorporelles et corporelles",
                   cessions_immo_incorp_corp)
        self._row("invest", "+ Encaissements liés aux cessions d'immobilisations financières", t["cessions_fin"])
        self._row("invest", "C — Flux de trésorerie provenant des activités d'investissement (somme FF à FJ)",
                   t["flux_investissement"])

        self._header("finance", "FLUX DE TRÉSORERIE PROVENANT DES CAPITAUX")
        self._row("finance", "+ Augmentations de capital par apports nouveaux", t["augmentation_capital"])
        self._row("finance", "+ Subventions d'investissement reçues", t["subventions_recues"])
        self._row("finance", "- Prélèvements sur le capital", t["prelevements_capital"])
        self._row("finance", "- Dividendes versés", t["dividendes_verses"])
        self._row("finance", "Flux de trésorerie provenant des capitaux propres (somme FK à FN)",
                   t["flux_capitaux_propres"])
        self._row("finance", "+ Emprunts", t["emprunts_nouveaux"])
        self._row("finance", "- Remboursements des emprunts et autres dettes financières",
                   t["remboursements_emprunts"])
        self._row("finance", "Flux de trésorerie provenant des capitaux étrangers (somme FO à FQ)",
                   t["flux_capitaux_etrangers"])
        self._row("finance", "D — Flux de trésorerie provenant des capitaux", t["flux_financement"])

        self._header("controle", "VARIATION ET CONTRÔLE")
        self._row("controle", "VARIATION DE LA TRÉSORERIE NETTE DE LA PÉRIODE (B+C+D)", t["variation_treso_nette"])
        self._row("controle", "TRÉSORERIE NETTE AU 31/12/N (B+C+D)+A", t["treso_cloture_calculee"])
        self._row("controle", "CONTRÔLE TRÉSORERIE NETTE AU 31/12/N (Balance, classe 5)", t["treso_cloture_reelle"])
        self._row("controle", "ÉCART", t["ecart"])

        if abs(t["ecart"]) < 1:
            self.ecart_var.set("✓ La trésorerie calculée correspond exactement à la trésorerie de la Balance.")
        else:
            self.ecart_var.set(
                f"⚠ Écart de {fmt_cfa(t['ecart'])} — un mouvement de trésorerie n'est peut-être pas "
                f"correctement classé (comptes d'immobilisations, capital ou emprunts)."
            )


class TftDirectTab(ttk.Frame):
    """Ancienne méthode (directe, par code flux EXP/INV/FIN) — conservée pour
    référence ; la méthode indirecte (CAFG) est désormais la vue principale."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn

        bar = ttk.Frame(self)
        bar.pack(fill="x", padx=8, pady=8)
        ttk.Label(bar, text="Trésorerie d'ouverture (auto., ou forcez une valeur) :").pack(side="left")
        self.ouverture_var = tk.StringVar()
        ttk.Entry(bar, textvariable=self.ouverture_var, width=14).pack(side="left", padx=4)
        ttk.Button(bar, text="Forcer cette valeur", command=self.save_and_refresh).pack(side="left", padx=4)
        ttk.Button(bar, text="Revenir à l'automatique", command=self.reset_auto).pack(side="left", padx=4)
        ttk.Label(bar, text=(
            "Par défaut = somme des soldes d'ouverture des comptes de trésorerie (onglet « Soldes "
            "d'ouverture »). Les mouvements se classent par nature via le code flux EXP/INV/FIN saisi "
            "dans l'onglet Saisie."
        ), foreground="#595959", wraplength=550).pack(side="left", padx=12)

        self.text = tk.Text(self, font=("Consolas", 11), wrap="none")
        self.text.pack(fill="both", expand=True, padx=8, pady=8)
        self.refresh()

    def save_and_refresh(self):
        try:
            value = float(self.ouverture_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "La trésorerie d'ouverture doit être un nombre.")
            return
        core.set_setting(self.conn, "treso_ouverture_override", value)
        core.set_setting(self.conn, "treso_ouverture_use_override", 1)
        self.refresh()

    def reset_auto(self):
        core.set_setting(self.conn, "treso_ouverture_use_override", 0)
        self.refresh()

    def refresh(self):
        use_override = core.get_setting(self.conn, "treso_ouverture_use_override", 0.0)
        ouverture_override = core.get_setting(self.conn, "treso_ouverture_override", 0.0) if use_override else None
        t = core.compute_tft(self.conn, treso_ouverture=ouverture_override)
        self.ouverture_var.set(str(t["ouverture"]))
        label_ouv = "Trésorerie d'ouverture"
        label_inv = "Flux liés aux activités d'investissement (INV)"
        label_clot = "TRÉSORERIE DE CLÔTURE"
        lines = [
            "TABLEAU DES FLUX DE TRÉSORERIE (méthode directe)", "=" * 60,
            f"  {label_ouv:<50} {t['ouverture']:>12,.2f}", "",
            f"  {'Flux liés aux activités opérationnelles (EXP)':<50} {t['exploitation']:>12,.2f}",
            f"  {label_inv:<50} {t['investissement']:>12,.2f}",
            f"  {'Flux liés aux activités de financement (FIN)':<50} {t['financement']:>12,.2f}",
            f"  {'Flux non classés (à coder)':<50} {t['non_classes']:>12,.2f}",
            f"  {'VARIATION NETTE DE TRÉSORERIE':<50} {t['variation']:>12,.2f}", "",
            f"{label_clot:<52} {t['cloture']:>12,.2f}",
        ]
        self.text.delete("1.0", "end")
        self.text.insert("1.0", "\n".join(lines))


class SituationFinanciereTab(ttk.Frame):
    """Situation financière (FR - BFR - TN), présentée selon le modèle
    officiel, avec une couleur par section. Entièrement recalculée à partir
    de compute_bilan(), compute_liasse_resultat() et compute_tft_indirect()
    — donc toujours cohérente avec la Balance, le Bilan et le TFT."""

    SECTIONS = {
        "cafg": "#D9EAD3",       # vert clair — CAFG / rentabilité
        "fr": "#CFE2F3",         # bleu clair — Fonds de roulement
        "bfr": "#FFF2CC",        # jaune clair — Besoin en fonds de roulement
        "tn": "#D9D2E9",         # violet clair — Trésorerie nette
        "flux": "#FCE5CD",       # orange clair — Flux de la période
        "endettement": "#F4CCCC",  # rouge/rose clair — Endettement
        "total": "#1F4E78",
    }

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        cols = ("libelle", "montant")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=30)
        self.tree.heading("libelle", text="Rubrique")
        self.tree.heading("montant", text="Montant")
        self.tree.column("libelle", width=480, anchor="w")
        self.tree.column("montant", width=160, anchor="e")
        for key, color in self.SECTIONS.items():
            fg = "white" if key == "total" else "black"
            tree_font = ("Segoe UI", 9, "bold") if key == "total" else ("Segoe UI", 9)
            self.tree.tag_configure(key, background=color, foreground=fg, font=tree_font)
            self.tree.tag_configure(key + "_header", background=color, foreground=fg, font=("Segoe UI", 9, "bold"))
        self.tree.pack(fill="both", padx=8, pady=8)
        self.ecart_var = tk.StringVar()
        ttk.Label(self, textvariable=self.ecart_var, font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=8)
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(pady=8)
        ttk.Button(self, text="Exporter (gabarit officiel .xlsx)",
                   command=lambda: export_etat_gabarit(
                       self, self.conn, "situation", "Situation_Financiere.xlsx", "la Situation financière")
                   ).pack(pady=(0, 8))
        self.refresh()

    def _row(self, tag, label, val, pct=False):
        suffix = " %" if pct else ""
        display = f"{fmt_cfa(val)}{suffix}" if pct else fmt_cfa(val)
        self.tree.insert("", "end", tags=(tag,), values=(f"  {label}", display))

    def _header(self, tag, titre):
        self.tree.insert("", "end", tags=(tag + "_header",), values=(titre, ""))

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        s = core.compute_situation_financiere(self.conn)

        self._row("cafg", "RÉSULTAT NET COMPTABLE", s["resultat_net_comptable"])

        self._header("cafg", "DÉTERMINATION DE LA CAPACITÉ D'AUTOFINANCEMENT")
        self._row("cafg", "EBE", s["ebe"])
        self._row("cafg", "- Valeurs comptables de cession courantes d'immobilisations (654)",
                   s["valeurs_comptables_cessions_courantes"])
        self._row("cafg", "+ Produits de cession courantes d'immobilisations (754)", s["produits_cessions_courantes"])
        self._row("cafg", "+ Transfert de charges d'exploitation (781)", s["transferts_charges_exploitation"])
        self._row("cafg", "CAPACITÉ D'AUTOFINANCEMENT D'EXPLOITATION", s["caf_exploitation"])
        self._row("cafg", "+ Revenus financiers", s["revenus_financiers"])
        self._row("cafg", "- Frais financiers", s["frais_financiers"])
        self._row("cafg", "CAPACITÉ D'AUTOFINANCEMENT GLOBAL (CAFG)", s["cafg"])
        self._row("cafg", "- Distribution de dividendes opérées durant l'exercice", s["dividendes_verses"])
        self._row("cafg", "AUTOFINANCEMENT", s["autofinancement"])
        self._row("cafg", "Rentabilité économique = Résultat exploitation / Capitaux propres",
                   s["rentabilite_economique"], pct=True)
        self._row("cafg", "Rentabilité financière = Résultat net / Capitaux propres",
                   s["rentabilite_financiere"], pct=True)

        self._header("fr", "ANALYSE DE LA SITUATION FINANCIÈRE")
        self._row("fr", "Capitaux propres et ressources assimilées", s["capitaux_propres_ressources"])
        self._row("fr", "+ Dettes financières", s["dettes_financieres"])
        self._row("fr", "= RESSOURCES STABLES", s["ressources_stables"])
        self._row("fr", "- Actifs immobilisés", -s["actifs_immobilises"])
        self._row("fr", "= FONDS DE ROULEMENT -1", s["fonds_de_roulement"])

        self._header("bfr", "BESOIN EN FONDS DE ROULEMENT (BFR)")
        self._row("bfr", "+ Actif circulant d'exploitation", s["actif_circulant_exploitation"])
        self._row("bfr", "- Passif circulant d'exploitation", s["passif_circulant_exploitation"])
        self._row("bfr", "= BESOIN DE FINANCEMENT D'EXPLOITATION -2", s["besoin_financement_exploitation"])
        self._row("bfr", "+ Actif circulant HAO", s["actif_circulant_hao"])
        self._row("bfr", "- Passif circulant HAO", s["passif_circulant_hao"])
        self._row("bfr", "= BESOIN DE FINANCEMENT HAO -3", s["besoin_financement_hao"])
        self._row("bfr", "BESOIN DE FINANCEMENT GLOBAL -4 = 2+3", s["besoin_financement_global"])

        self._header("tn", "TRÉSORERIE NETTE")
        self._row("tn", "TRÉSORERIE NETTE -5 = 1-4", s["tresorerie_nette"])
        self._row("tn", "CONTRÔLE TRÉSORERIE NETTE (Balance, classe 5)", s["controle_treso_reelle"])
        self._row("tn", "ÉCART", s["controle_ecart"])

        self._header("flux", "FLUX DE TRÉSORERIE DE LA PÉRIODE (cf. onglet TFT)")
        self._row("flux", "+ Flux de la trésorerie des activités opérationnelles", s["flux_operationnel"])
        self._row("flux", "- Flux de la trésorerie des activités d'investissement", s["flux_investissement"])
        self._row("flux", "+ Flux de la trésorerie des activités de financement", s["flux_financement"])
        self._row("flux", "VARIATION DE LA TRÉSORERIE NETTE DE LA PÉRIODE", s["variation_treso_nette"])

        self._header("endettement", "ENDETTEMENT FINANCIER")
        self._row("endettement", "Endettement financier brut (dettes fin. + trésorerie passif)",
                   s["endettement_financier_brut"])
        self._row("endettement", "- Trésorerie actif", s["treso_actif"])
        self._row("endettement", "= ENDETTEMENT FINANCIER NET", s["endettement_financier_net"])

        self.tree.insert("", "end", tags=("total",), values=(
            "TRÉSORERIE NETTE", f"{fmt_cfa(s['controle_treso_reelle'])}"))

        if abs(s["controle_ecart"]) < 1:
            self.ecart_var.set("✓ La trésorerie nette (FR - BFR) correspond exactement à la Balance.")
        else:
            self.ecart_var.set(
                f"⚠ Écart de {fmt_cfa(s['controle_ecart'])} — vérifiez que les soldes d'ouverture de tous "
                f"les comptes (onglet Soldes d'ouverture) sont complets et s'équilibrent à zéro."
            )


class TftTab(ttk.Frame):
    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        inner = ttk.Notebook(self)
        inner.pack(fill="both", expand=True)
        self.indirect_tab = TftIndirectTab(inner, conn)
        self.direct_tab = TftDirectTab(inner, conn)
        inner.add(self.indirect_tab, text="TFT (méthode indirecte — CAFG)")
        inner.add(self.direct_tab, text="TFT (méthode directe — ancien)")

    def refresh(self):
        self.indirect_tab.refresh()
        self.direct_tab.refresh()


class LiasseFiscaleTab(ttk.Frame):
    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn

        info = ttk.LabelFrame(self, text="Identification de l'entité (SYSCOHADA / DGI)")
        info.pack(fill="x", padx=8, pady=8)

        self.vars = {}
        for i, (key, label) in enumerate(core.COMPANY_FIELDS.items()):
            r, c = divmod(i, 2)
            ttk.Label(info, text=label + " :").grid(row=r * 2, column=c, sticky="w", padx=4, pady=(4, 0))
            var = tk.StringVar(value=core.get_company_value(conn, key))
            ttk.Entry(info, textvariable=var, width=40).grid(row=r * 2 + 1, column=c, sticky="we", padx=4, pady=(0, 6))
            self.vars[key] = var
        ttk.Button(info, text="Enregistrer les informations", command=self.save_info).grid(
            row=(len(core.COMPANY_FIELDS) + 1) // 2 * 2, column=0, sticky="w", padx=4, pady=6)

        params = ttk.LabelFrame(self, text="Paramètres d'export")
        params.pack(fill="x", padx=8, pady=(0, 8))
        ttk.Label(params, text="Stock initial total (cf. onglet Stocks) :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.stock_initial_var = tk.StringVar(value="0")
        ttk.Entry(params, textvariable=self.stock_initial_var, width=16).grid(row=0, column=1, sticky="w", padx=4)
        ttk.Label(params, text="(complément optionnel — utilisez plutôt l'onglet « Soldes d'ouverture »)",
                  foreground="#595959").grid(row=0, column=2, sticky="w", padx=(10, 4))

        note = ttk.Label(self, wraplength=900, foreground="#595959", text=(
            "Génère un classeur .xlsx COMPLET reprenant les 92 pages du modèle SYSCOHADA système "
            "normal (mêmes dimensions, mêmes codes officiels) : COUVERTURE, BILAN, RESULTAT, TFT, "
            "39 notes annexes, ~20 tableaux fiscaux DGI. BILAN et RESULTAT sont calculés automatiquement "
            "depuis vos écritures (soldes de clôture = solde d'ouverture + mouvements de l'exercice, "
            "cf. onglet « Soldes d'ouverture »). Le TFT officiel (méthode indirecte, CAFG) est laissé "
            "vierge — un onglet « TFT (simplifie) » calculé en méthode directe est ajouté à titre "
            "indicatif. Toutes les autres pages gardent leur mise en page et leurs dimensions exactes, "
            "mais leurs valeurs sont vidées (ce ne sont pas vos chiffres) pour être complétées "
            "manuellement — le détail des lignes du Bilan (AE à AN, CA à CM, DA à DM) est une "
            "répartition indicative par plage de comptes. À faire vérifier par un expert-comptable "
            "avant tout dépôt officiel auprès de la DGI."
        ))
        note.pack(fill="x", padx=8, pady=(0, 8))

        ttk.Button(self, text="Exporter la liasse fiscale complète (.xlsx)", command=self.export).pack(padx=8, pady=8, anchor="w")
        self.status_var = tk.StringVar()
        ttk.Label(self, textvariable=self.status_var, foreground="#1F4E78").pack(padx=8, anchor="w")

    def save_info(self):
        for key, var in self.vars.items():
            core.set_company_value(self.conn, key, var.get().strip())
        self.status_var.set("Informations enregistrées.")

    def export(self):
        self.save_info()
        try:
            stock_initial = float(self.stock_initial_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Le complément de stock initial doit être un nombre.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Liasse_fiscale.xlsx",
            title="Enregistrer la liasse fiscale",
        )
        if not path:
            return
        try:
            core.export_liasse_fiscale_complete(self.conn, path, stock_initial=stock_initial)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de l'export : {exc}")
            return
        self.status_var.set(f"Export réussi : {path}")
        messagebox.showinfo("Export terminé", f"Liasse fiscale enregistrée :\n{path}")


class PersonnelTab(ttk.Frame):
    """Liste du personnel (menu GRH) — sans lien avec la comptabilité."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.selected_id = None
        ttk.Label(self, text="LISTE DU PERSONNEL", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))

        form = ttk.LabelFrame(self, text="Employé")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Matricule :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.matricule_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.matricule_var, width=14).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Nom :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.nom_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.nom_var, width=16).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Prénom :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.prenom_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.prenom_var, width=16).grid(row=0, column=5, padx=4)
        ttk.Label(form, text="Poste :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.poste_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.poste_var, width=16).grid(row=1, column=1, padx=4, pady=(4, 0))
        ttk.Label(form, text="Service :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.service_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.service_var, width=16).grid(row=1, column=3, padx=4, pady=(4, 0))
        ttk.Label(form, text="Date d'embauche :").grid(row=1, column=4, sticky="w", padx=(12, 4), pady=(4, 0))
        self.date_embauche_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.date_embauche_var, width=12).grid(row=1, column=5, padx=4, pady=(4, 0))
        ttk.Label(form, text="Téléphone :").grid(row=2, column=0, sticky="w", padx=4, pady=(4, 0))
        self.telephone_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.telephone_var, width=16).grid(row=2, column=1, padx=4, pady=(4, 0))
        ttk.Label(form, text="Email :").grid(row=2, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.email_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.email_var, width=20).grid(row=2, column=3, padx=4, pady=(4, 0))
        ttk.Label(form, text="Statut :").grid(row=2, column=4, sticky="w", padx=(12, 4), pady=(4, 0))
        self.statut_var = tk.StringVar(value="actif")
        ttk.Combobox(form, textvariable=self.statut_var, width=13, state="readonly",
                     values=["actif", "congé", "suspendu", "parti"]).grid(row=2, column=5, padx=4, pady=(4, 0))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=6)
        ttk.Button(btns, text="Ajouter", command=self.add).pack(side="left")
        ttk.Button(btns, text="Mettre à jour la sélection", command=self.update_sel).pack(side="left", padx=8)
        ttk.Button(btns, text="Supprimer la sélection", command=self.delete_sel).pack(side="left")
        ttk.Button(btns, text="Effacer le formulaire", command=self.clear_form).pack(side="left", padx=8)

        import_bar = ttk.Frame(self)
        import_bar.pack(fill="x", padx=16, pady=(0, 6))
        ttk.Button(import_bar, text="Télécharger le modèle d'import (.xlsx)",
                   command=self.telecharger_modele).pack(side="left")
        ttk.Button(import_bar, text="Importer (.xlsx)", command=self.importer).pack(side="left", padx=8)

        cols = ("id", "matricule", "nom", "prenom", "poste", "service", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=18)
        for c, h, w in zip(cols, ["ID", "Matricule", "Nom", "Prénom", "Poste", "Service", "Statut"],
                           [40, 100, 130, 130, 150, 130, 90]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_id = v[0]
        self.matricule_var.set(v[1]); self.nom_var.set(v[2]); self.prenom_var.set(v[3])
        self.poste_var.set(v[4]); self.service_var.set(v[5]); self.statut_var.set(v[6])
        p = core.get_personnel(self.conn, self.selected_id)
        if p:
            self.date_embauche_var.set(core.to_display_date(p["date_embauche"] or ""))
            self.telephone_var.set(p["telephone"] or "")
            self.email_var.set(p["email"] or "")

    def clear_form(self):
        self.selected_id = None
        for var in (self.matricule_var, self.nom_var, self.prenom_var, self.poste_var, self.service_var,
                    self.date_embauche_var, self.telephone_var, self.email_var):
            var.set("")
        self.statut_var.set("actif")

    def add(self):
        if not self.nom_var.get().strip():
            messagebox.showwarning("Champ manquant", "Le nom est obligatoire.")
            return
        try:
            core.add_personnel(
                self.conn, self.nom_var.get(), matricule=self.matricule_var.get(), prenom=self.prenom_var.get(),
                poste=self.poste_var.get(), service=self.service_var.get(),
                date_embauche=core.to_iso_date(self.date_embauche_var.get().strip()),
                telephone=self.telephone_var.get(), email=self.email_var.get(), statut=self.statut_var.get())
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc))
            return
        self.clear_form()
        self.refresh()

    def update_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un employé.")
            return
        core.update_personnel(
            self.conn, self.selected_id, matricule=self.matricule_var.get().strip(),
            nom=self.nom_var.get().strip(), prenom=self.prenom_var.get().strip(),
            poste=self.poste_var.get().strip(), service=self.service_var.get().strip(),
            date_embauche=core.to_iso_date(self.date_embauche_var.get().strip()),
            telephone=self.telephone_var.get().strip(), email=self.email_var.get().strip(),
            statut=self.statut_var.get())
        self.clear_form()
        self.refresh()

    def delete_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un employé.")
            return
        if messagebox.askyesno("Confirmer", "Supprimer cet employé ?"):
            core.delete_personnel(self.conn, self.selected_id)
            self.clear_form()
            self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for p in core.list_personnel(self.conn):
            self.tree.insert("", "end", values=(
                p["id"], p["matricule"] or "", p["nom"], p["prenom"] or "", p["poste"] or "",
                p["service"] or "", p["statut"]))

    def telecharger_modele(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Modele_Import_Personnel.xlsx", title="Télécharger le modèle d'import — Personnel",
        )
        if not path:
            return
        core.export_personnel_template_xlsx(path)
        messagebox.showinfo("Modèle téléchargé", f"Modèle enregistré :\n{path}")

    def importer(self):
        path = filedialog.askopenfilename(filetypes=[("Classeur Excel", "*.xlsx")],
                                           title="Importer la liste du personnel")
        if not path:
            return
        try:
            rapport = core.import_personnel_xlsx(self.conn, path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de l'import : {exc}")
            return
        self.refresh()
        msg = f"{rapport['crees']} créé(s), {rapport['mis_a_jour']} mis à jour."
        if rapport["erreurs"]:
            msg += "\n\n⚠ " + "\n".join(rapport["erreurs"][:10])
            if len(rapport["erreurs"]) > 10:
                msg += f"\n... et {len(rapport['erreurs']) - 10} autre(s) erreur(s)."
        messagebox.showinfo("Import terminé", msg)


class TimeSheetTab(ttk.Frame):
    """Time sheet (pointage des heures) — menu GRH, sans lien avec la comptabilité."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="TIME SHEET", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))

        form = ttk.LabelFrame(self, text="Nouveau pointage")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Employé :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.personnel_var = tk.StringVar()
        self.personnel_combo = ttk.Combobox(form, textvariable=self.personnel_var, width=26, state="readonly")
        self.personnel_combo.grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Date (JJ/MM/AAAA) :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(form, textvariable=self.date_var, width=12).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Heures :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.heures_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.heures_var, width=8).grid(row=0, column=5, padx=4)
        ttk.Label(form, text="Activité :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.activite_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.activite_var, width=40).grid(row=1, column=1, columnspan=3, padx=4, pady=(4, 0), sticky="we")
        ttk.Button(form, text="Ajouter le pointage", command=self.add).grid(row=1, column=5, padx=4, pady=(4, 0))

        import_bar = ttk.Frame(self)
        import_bar.pack(fill="x", padx=16, pady=(0, 6))
        ttk.Button(import_bar, text="Télécharger le modèle d'import (.xlsx)",
                   command=self.telecharger_modele).pack(side="left")
        ttk.Button(import_bar, text="Importer (.xlsx)", command=self.importer).pack(side="left", padx=8)

        cols = ("id", "employe", "date", "heures", "activite")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=18)
        for c, h, w in zip(cols, ["ID", "Employé", "Date", "Heures", "Activité"], [40, 180, 100, 80, 350]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", padx=16, pady=8)
        ttk.Button(self, text="Supprimer la ligne sélectionnée", command=self.delete_sel).pack(anchor="w", padx=16, pady=(0, 12))
        self.refresh()

    def _refresh_personnel_values(self):
        self.personnel_list = core.list_personnel(self.conn, actifs_only=True)
        self.personnel_combo["values"] = [f"{p['id']} — {p['prenom'] or ''} {p['nom']}".strip() for p in self.personnel_list]

    def add(self):
        raw = self.personnel_var.get()
        if not raw:
            messagebox.showwarning("Champ manquant", "Choisissez un employé.")
            return
        personnel_id = int(raw.split(" — ", 1)[0])
        date_str = core.to_iso_date(self.date_var.get().strip())
        if not date_str:
            messagebox.showwarning("Champ manquant", "La date est obligatoire.")
            return
        try:
            heures = float(self.heures_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Les heures doivent être un nombre.")
            return
        try:
            core.add_time_sheet(self.conn, personnel_id, date_str, heures, activite=self.activite_var.get())
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc))
            return
        self.heures_var.set(""); self.activite_var.set("")
        self.refresh()

    def delete_sel(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne.")
            return
        ts_id = self.tree.item(sel[0], "values")[0]
        core.delete_time_sheet(self.conn, ts_id)
        self.refresh()

    def refresh(self):
        self._refresh_personnel_values()
        for row in self.tree.get_children():
            self.tree.delete(row)
        for t in core.list_time_sheet(self.conn):
            self.tree.insert("", "end", values=(
                t["id"], t["employe"], core.to_display_date(t["date_pointage"]), f"{t['heures']:g}",
                t["activite"] or ""))

    def telecharger_modele(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Modele_Import_Time_Sheet.xlsx", title="Télécharger le modèle d'import — Time sheet",
        )
        if not path:
            return
        core.export_time_sheet_template_xlsx(path)
        messagebox.showinfo("Modèle téléchargé", f"Modèle enregistré :\n{path}")

    def importer(self):
        path = filedialog.askopenfilename(filetypes=[("Classeur Excel", "*.xlsx")],
                                           title="Importer des pointages Time sheet")
        if not path:
            return
        try:
            rapport = core.import_time_sheet_xlsx(self.conn, path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de l'import : {exc}")
            return
        self.refresh()
        msg = f"{rapport['crees']} pointage(s) créé(s)."
        if rapport["erreurs"]:
            msg += "\n\n⚠ " + "\n".join(rapport["erreurs"][:10])
            if len(rapport["erreurs"]) > 10:
                msg += f"\n... et {len(rapport['erreurs']) - 10} autre(s) erreur(s)."
        messagebox.showinfo("Import terminé", msg)


class KpiTab(ttk.Frame):
    """KPI (indicateurs de performance) — menu GRH, sans lien avec la comptabilité."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.selected_id = None
        ttk.Label(self, text="KPI — INDICATEURS DE PERFORMANCE", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 8))

        form = ttk.LabelFrame(self, text="Indicateur")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Indicateur :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.indicateur_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.indicateur_var, width=28).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Employé (optionnel) :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.personnel_var = tk.StringVar()
        self.personnel_combo = ttk.Combobox(form, textvariable=self.personnel_var, width=22, state="readonly")
        self.personnel_combo.grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Service :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.service_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.service_var, width=14).grid(row=0, column=5, padx=4)
        ttk.Label(form, text="Période :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.periode_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.periode_var, width=14).grid(row=1, column=1, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(form, text="Valeur cible :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.cible_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.cible_var, width=10).grid(row=1, column=3, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(form, text="Valeur réalisée :").grid(row=1, column=4, sticky="w", padx=(12, 4), pady=(4, 0))
        self.realisee_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.realisee_var, width=10).grid(row=1, column=5, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(form, text="Unité :").grid(row=2, column=0, sticky="w", padx=4, pady=(4, 0))
        self.unite_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.unite_var, width=10).grid(row=2, column=1, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(form, text="Statut :").grid(row=2, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.statut_var = tk.StringVar(value="en_cours")
        ttk.Combobox(form, textvariable=self.statut_var, width=13, state="readonly",
                     values=["en_cours", "atteint", "non_atteint"]).grid(row=2, column=3, padx=4, pady=(4, 0))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=6)
        ttk.Button(btns, text="Ajouter", command=self.add).pack(side="left")
        ttk.Button(btns, text="Mettre à jour la sélection", command=self.update_sel).pack(side="left", padx=8)
        ttk.Button(btns, text="Supprimer la sélection", command=self.delete_sel).pack(side="left")
        ttk.Button(btns, text="Effacer le formulaire", command=self.clear_form).pack(side="left", padx=8)

        cols = ("id", "indicateur", "employe", "service", "periode", "cible", "realisee", "taux", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=16)
        headers = ["ID", "Indicateur", "Employé", "Service", "Période", "Cible", "Réalisée", "Taux %", "Statut"]
        widths = [40, 200, 150, 100, 90, 80, 80, 70, 100]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.tag_configure("atteint", foreground="#1F7A1F")
        self.tree.tag_configure("non_atteint", foreground="#B00020")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _refresh_personnel_values(self):
        self.personnel_list = core.list_personnel(self.conn)
        self.personnel_combo["values"] = ["(aucun)"] + [f"{p['id']} — {p['prenom'] or ''} {p['nom']}".strip()
                                                          for p in self.personnel_list]

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_id = v[0]
        self.indicateur_var.set(v[1])
        self.service_var.set(v[3]); self.periode_var.set(v[4])
        self.cible_var.set(v[5]); self.realisee_var.set(v[6])
        self.statut_var.set(v[8])
        k = next((x for x in core.list_kpi(self.conn) if str(x["id"]) == str(self.selected_id)), None)
        if k:
            self.unite_var.set(k["unite"] or "")
            self.personnel_var.set(f"{k['personnel_id']} — {k['employe']}" if k["personnel_id"] else "(aucun)")

    def clear_form(self):
        self.selected_id = None
        for var in (self.indicateur_var, self.service_var, self.periode_var, self.cible_var,
                    self.realisee_var, self.unite_var, self.personnel_var):
            var.set("")
        self.statut_var.set("en_cours")

    def _parse_personnel_id(self):
        raw = self.personnel_var.get()
        if not raw or raw == "(aucun)":
            return None
        return int(raw.split(" — ", 1)[0])

    def add(self):
        if not self.indicateur_var.get().strip():
            messagebox.showwarning("Champ manquant", "Le nom de l'indicateur est obligatoire.")
            return
        try:
            cible = float(self.cible_var.get() or 0)
            realisee = float(self.realisee_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Cible et Réalisée doivent être des nombres.")
            return
        core.add_kpi(self.conn, self.indicateur_var.get(), personnel_id=self._parse_personnel_id(),
                     service=self.service_var.get(), periode=self.periode_var.get(), valeur_cible=cible,
                     valeur_realisee=realisee, unite=self.unite_var.get(), statut=self.statut_var.get())
        self.clear_form()
        self.refresh()

    def update_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un indicateur.")
            return
        try:
            cible = float(self.cible_var.get() or 0)
            realisee = float(self.realisee_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Cible et Réalisée doivent être des nombres.")
            return
        core.update_kpi(self.conn, self.selected_id, indicateur=self.indicateur_var.get().strip(),
                         personnel_id=self._parse_personnel_id(), service=self.service_var.get().strip(),
                         periode=self.periode_var.get().strip(), valeur_cible=cible, valeur_realisee=realisee,
                         unite=self.unite_var.get().strip(), statut=self.statut_var.get())
        self.clear_form()
        self.refresh()

    def delete_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un indicateur.")
            return
        if messagebox.askyesno("Confirmer", "Supprimer cet indicateur ?"):
            core.delete_kpi(self.conn, self.selected_id)
            self.clear_form()
            self.refresh()

    def refresh(self):
        self._refresh_personnel_values()
        for row in self.tree.get_children():
            self.tree.delete(row)
        for k in core.list_kpi(self.conn):
            taux = f"{k['taux_realisation']:.0f}" if k["taux_realisation"] is not None else ""
            tag = ()
            if k["taux_realisation"] is not None:
                tag = ("atteint",) if k["taux_realisation"] >= 100 else ("non_atteint",)
            self.tree.insert("", "end", tags=tag, values=(
                k["id"], k["indicateur"], k["employe"] or "", k["service"] or "", k["periode"] or "",
                f"{k['valeur_cible']:g}", f"{k['valeur_realisee']:g}", taux, k["statut"]))


class TableauBordGrhTab(ttk.Frame):
    """Tableau de bord GRH — synthèse en lecture seule des autres écrans
    (Personnel, Time sheet, KPI, HS)."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="TABLEAU DE BORD GRH", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 8))
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(anchor="w", padx=16)
        self.cards_frame = ttk.Frame(self)
        self.cards_frame.pack(fill="x", padx=16, pady=16)
        self.hs_frame = ttk.LabelFrame(self, text="Incidents HS ouverts, par gravité")
        self.hs_frame.pack(fill="x", padx=16, pady=8)
        self.refresh()

    def _card(self, parent, titre, valeur, col, couleur="#1F4E78"):
        f = ttk.Frame(parent, relief="solid", borderwidth=1)
        f.grid(row=0, column=col, padx=8, sticky="nsew")
        parent.columnconfigure(col, weight=1)
        tk.Label(f, text=titre, font=("Segoe UI", 9), bg="white", fg="#595959").pack(fill="x", padx=12, pady=(10, 0))
        tk.Label(f, text=str(valeur), font=("Segoe UI", 20, "bold"), bg="white", fg=couleur).pack(
            fill="x", padx=12, pady=(0, 10))

    def refresh(self):
        for w in self.cards_frame.winfo_children():
            w.destroy()
        for w in self.hs_frame.winfo_children():
            w.destroy()
        d = core.compute_tableau_bord_grh(self.conn)
        self._card(self.cards_frame, "Personnel actif", f"{d['nb_personnel_actif']} / {d['nb_personnel_total']}", 0)
        self._card(self.cards_frame, "Heures pointées (30j)", f"{d['total_heures_30j']:g} h", 1)
        self._card(self.cards_frame, "KPI en cours", d["nb_kpi_en_cours"], 2)
        self._card(self.cards_frame, "KPI atteints", d["nb_kpi_atteints"], 3, couleur="#1F7A1F")
        self._card(self.cards_frame, "KPI non atteints", d["nb_kpi_non_atteints"], 4, couleur="#B00020")
        self._card(self.cards_frame, "Incidents HS ouverts", d["nb_hs_ouverts"], 5,
                   couleur="#B00020" if d["nb_hs_ouverts"] else "#1F7A1F")
        if not d["hs_par_gravite"]:
            ttk.Label(self.hs_frame, text="Aucun incident ouvert.", foreground="#1F7A1F").pack(
                anchor="w", padx=12, pady=8)
        else:
            for gravite, nb in d["hs_par_gravite"].items():
                ttk.Label(self.hs_frame, text=f"• {gravite} : {nb}").pack(anchor="w", padx=12, pady=2)


class HsTab(ttk.Frame):
    """HS (Hygiène Santé) — menu GRH, sans lien avec la comptabilité."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.selected_id = None
        ttk.Label(self, text="HS — HYGIÈNE SANTÉ", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 8))

        form = ttk.LabelFrame(self, text="Événement")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Date (JJ/MM/AAAA) :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(form, textvariable=self.date_var, width=12).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Type :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.type_var = tk.StringVar(value="incident")
        ttk.Combobox(form, textvariable=self.type_var, width=17, state="readonly",
                     values=["incident", "visite_medicale", "formation_securite", "distribution_epi"]).grid(
            row=0, column=3, padx=4)
        ttk.Label(form, text="Employé (optionnel) :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.personnel_var = tk.StringVar()
        self.personnel_combo = ttk.Combobox(form, textvariable=self.personnel_var, width=22, state="readonly")
        self.personnel_combo.grid(row=0, column=5, padx=4)
        ttk.Label(form, text="Gravité :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.gravite_var = tk.StringVar()
        ttk.Combobox(form, textvariable=self.gravite_var, width=17, state="readonly",
                     values=["", "Mineure", "Modérée", "Grave"]).grid(row=1, column=1, padx=4, pady=(4, 0))
        ttk.Label(form, text="Statut :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.statut_var = tk.StringVar(value="ouvert")
        ttk.Combobox(form, textvariable=self.statut_var, width=13, state="readonly",
                     values=["ouvert", "clos"]).grid(row=1, column=3, padx=4, pady=(4, 0))
        ttk.Label(form, text="Description :").grid(row=2, column=0, sticky="w", padx=4, pady=(4, 0))
        self.description_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.description_var, width=60).grid(
            row=2, column=1, columnspan=5, padx=4, pady=(4, 0), sticky="we")

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=6)
        ttk.Button(btns, text="Ajouter", command=self.add).pack(side="left")
        ttk.Button(btns, text="Mettre à jour la sélection", command=self.update_sel).pack(side="left", padx=8)
        ttk.Button(btns, text="Supprimer la sélection", command=self.delete_sel).pack(side="left")
        ttk.Button(btns, text="Effacer le formulaire", command=self.clear_form).pack(side="left", padx=8)

        cols = ("id", "date", "type", "employe", "gravite", "statut", "description")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=16)
        headers = ["ID", "Date", "Type", "Employé", "Gravité", "Statut", "Description"]
        widths = [40, 90, 140, 150, 90, 80, 320]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.tag_configure("ouvert", foreground="#B00020")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _refresh_personnel_values(self):
        self.personnel_list = core.list_personnel(self.conn)
        self.personnel_combo["values"] = ["(aucun)"] + [f"{p['id']} — {p['prenom'] or ''} {p['nom']}".strip()
                                                          for p in self.personnel_list]

    def _parse_personnel_id(self):
        raw = self.personnel_var.get()
        if not raw or raw == "(aucun)":
            return None
        return int(raw.split(" — ", 1)[0])

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_id = v[0]
        self.date_var.set(v[1]); self.type_var.set(v[2])
        self.gravite_var.set(v[4]); self.statut_var.set(v[5]); self.description_var.set(v[6])
        h = next((x for x in core.list_hs(self.conn) if str(x["id"]) == str(self.selected_id)), None)
        if h:
            self.personnel_var.set(f"{h['personnel_id']} — {h['employe']}" if h["personnel_id"] else "(aucun)")

    def clear_form(self):
        self.selected_id = None
        self.date_var.set(date.today().strftime("%d/%m/%Y"))
        self.type_var.set("incident"); self.gravite_var.set(""); self.statut_var.set("ouvert")
        self.description_var.set(""); self.personnel_var.set("")

    def add(self):
        date_str = core.to_iso_date(self.date_var.get().strip())
        if not date_str:
            messagebox.showwarning("Champ manquant", "La date est obligatoire.")
            return
        core.add_hs(self.conn, date_str, type_evenement=self.type_var.get(),
                    personnel_id=self._parse_personnel_id(), description=self.description_var.get(),
                    gravite=self.gravite_var.get(), statut=self.statut_var.get())
        self.clear_form()
        self.refresh()

    def update_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un événement.")
            return
        date_str = core.to_iso_date(self.date_var.get().strip())
        core.update_hs(self.conn, self.selected_id, date_evenement=date_str, type_evenement=self.type_var.get(),
                       personnel_id=self._parse_personnel_id(), description=self.description_var.get().strip(),
                       gravite=self.gravite_var.get(), statut=self.statut_var.get())
        self.clear_form()
        self.refresh()

    def delete_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un événement.")
            return
        if messagebox.askyesno("Confirmer", "Supprimer cet événement ?"):
            core.delete_hs(self.conn, self.selected_id)
            self.clear_form()
            self.refresh()

    def refresh(self):
        self._refresh_personnel_values()
        for row in self.tree.get_children():
            self.tree.delete(row)
        for h in core.list_hs(self.conn):
            tag = ("ouvert",) if h["statut"] == "ouvert" else ()
            self.tree.insert("", "end", tags=tag, values=(
                h["id"], core.to_display_date(h["date_evenement"]), h["type_evenement"], h["employe"] or "",
                h["gravite"] or "", h["statut"], h["description"] or ""))


class PaieBulletinsTab(ttk.Frame):
    """Saisie des éléments de gain (salaire de base, primes, indemnités...)
    de chaque employé pour une période de paie — équivalent de
    EmployeesTab dans Paie Burkina, mais réutilisant les employés déjà
    saisis dans GRH > Personnel plutôt que d'en tenir une liste séparée."""

    CHAMPS = [
        ("classification", "Classification", "combo"),
        ("salaire_base", "Salaire de base", "num"),
        ("prime_anciennete", "Prime d'ancienneté", "num"),
        ("heures_sup", "Heures supplémentaires", "num"),
        ("sursalaire", "Sursalaire", "num"),
        ("gratification", "Gratification", "num"),
        ("indemnite_caisse", "Indemnité Caisse", "num"),
        ("indemnite_logement", "Indemnité Logement", "num"),
        ("indemnite_fonction", "Indemnité Fonction", "num"),
        ("indemnite_transport", "Indemnité Transport", "num"),
        ("personnes_a_charge", "Personnes à charge", "num"),
        ("retenue_pret", "Retenue prêt/avance", "num"),
    ]

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.selected_bulletin_id = None
        self.selected_personnel_id = None

        top = ttk.Frame(self)
        top.pack(fill="x", padx=12, pady=8)
        ttk.Label(top, text="Période (AAAA-MM) :").pack(side="left")
        self.periode_var = tk.StringVar(value=date.today().strftime("%Y-%m"))
        ttk.Entry(top, textvariable=self.periode_var, width=10).pack(side="left", padx=4)
        ttk.Button(top, text="Actualiser", command=self.refresh).pack(side="left", padx=8)
        ttk.Button(top, text="Dupliquer vers une autre période...", command=self.dupliquer).pack(
            side="left", padx=8)

        import_bar = ttk.Frame(self)
        import_bar.pack(fill="x", padx=12, pady=(0, 4))
        ttk.Button(import_bar, text="Importer des bulletins (.xlsx)", command=self.import_xlsx).pack(
            side="left", padx=2)
        ttk.Button(import_bar, text="Télécharger un modèle (.xlsx)", command=self.download_template).pack(
            side="left", padx=2)

        body = ttk.Frame(self)
        body.pack(fill="both", expand=True, padx=12, pady=(0, 8))

        right = ttk.Frame(body, width=320)
        right.pack(side="right", fill="y")
        right.pack_propagate(False)

        left = ttk.Frame(body)
        left.pack(side="left", fill="both", expand=True, padx=(0, 8))

        cols = ("matricule", "nom", "prenom", "classification", "salaire_base", "net_percu")
        self.tree = ttk.Treeview(left, columns=cols, show="headings")
        headers = ["Matricule", "Nom", "Prénom", "Classification", "Salaire base", "Net perçu (calculé)"]
        widths = [90, 150, 130, 110, 110, 130]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        ttk.Label(right, text="Bulletin employé", font=("Segoe UI", 11, "bold")).pack(pady=(0, 8))
        ttk.Label(right, text="Employé :").pack(anchor="w")
        self.employe_var = tk.StringVar()
        self.employe_combo = ttk.Combobox(right, textvariable=self.employe_var, width=32, state="readonly")
        self.employe_combo.pack(anchor="w", pady=(0, 8))
        self._refresh_employes()

        self.form_vars = {}
        form = ttk.Frame(right)
        form.pack(fill="x")
        for i, (key, label, kind) in enumerate(self.CHAMPS):
            ttk.Label(form, text=label).grid(row=i, column=0, sticky="w", pady=2)
            var = tk.StringVar()
            if kind == "combo":
                w = ttk.Combobox(form, textvariable=var, values=["CADRE", "AUTRE"], state="readonly", width=16)
                var.set("AUTRE")
            else:
                w = ttk.Entry(form, textvariable=var, width=18)
                var.set("0")
            w.grid(row=i, column=1, pady=2, sticky="w")
            self.form_vars[key] = var

        btns = ttk.Frame(right)
        btns.pack(pady=12)
        ttk.Button(btns, text="Enregistrer le bulletin", command=self.save_bulletin).grid(row=0, column=0, padx=2)
        ttk.Button(btns, text="Supprimer", command=self.delete_bulletin).grid(row=0, column=1, padx=2)
        ttk.Button(right, text="Vider le formulaire", command=self.clear_form).pack()

        self.refresh()

    def _refresh_employes(self):
        self.personnel_by_label = {}
        items = core.list_personnel(self.conn, actifs_only=True)
        values = []
        for p in items:
            label = f"{p['matricule'] or p['id']} — {p['nom']} {p['prenom'] or ''}".strip()
            values.append(label)
            self.personnel_by_label[label] = p
        self.employe_combo["values"] = values

    def clear_form(self):
        self.selected_bulletin_id = None
        self.selected_personnel_id = None
        self.employe_var.set("")
        for key, _, kind in self.CHAMPS:
            self.form_vars[key].set("AUTRE" if kind == "combo" else "0")

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        personnel_id = int(sel[0])
        self.selected_personnel_id = personnel_id
        periode = self.periode_var.get().strip()
        b = core.get_bulletin_paie(self.conn, personnel_id, periode)
        if not b:
            return
        self.selected_bulletin_id = b["id"]
        p = core.get_personnel(self.conn, personnel_id)
        label = f"{p['matricule'] or p['id']} — {p['nom']} {p['prenom'] or ''}".strip()
        self.employe_var.set(label)
        for key, _, _ in self.CHAMPS:
            self.form_vars[key].set(str(b.get(key, 0)))

    def save_bulletin(self):
        label = self.employe_var.get().strip()
        p = self.personnel_by_label.get(label)
        if not p:
            messagebox.showwarning("Champ manquant", "Choisissez un employé dans la liste.")
            return
        periode = self.periode_var.get().strip()
        if len(periode) != 7 or periode[4] != "-":
            messagebox.showerror("Erreur", "Période invalide — format attendu : AAAA-MM (ex. 2026-08).")
            return
        try:
            champs = {}
            for key, _, kind in self.CHAMPS:
                if kind == "combo":
                    champs[key] = self.form_vars[key].get()
                elif key == "personnes_a_charge":
                    champs[key] = int(float(self.form_vars[key].get() or 0))
                else:
                    champs[key] = float(self.form_vars[key].get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Merci de vérifier les valeurs numériques saisies.")
            return
        try:
            core.set_bulletin_paie(self.conn, p["id"], periode, **champs)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc))
            return
        self.refresh()
        messagebox.showinfo("Enregistré", "Bulletin de paie enregistré.")

    def delete_bulletin(self):
        if not self.selected_bulletin_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un bulletin dans le tableau.")
            return
        if messagebox.askyesno("Confirmer", "Supprimer ce bulletin de paie ?"):
            core.delete_bulletin_paie(self.conn, self.selected_bulletin_id)
            self.clear_form()
            self.refresh()

    def dupliquer(self):
        cible = simpledialog.askstring(
            "Dupliquer vers une autre période",
            "Copier les bulletins de la période affichée vers quelle période (AAAA-MM) ?",
            initialvalue=self.periode_var.get().strip())
        if not cible:
            return
        try:
            n = core.dupliquer_bulletins_periode(self.conn, self.periode_var.get().strip(), cible.strip())
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc))
            return
        messagebox.showinfo("Terminé", f"{n} bulletin(s) dupliqué(s) vers {cible.strip()}.")

    def download_template(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Modele_bulletins_paie.xlsx", title="Enregistrer le modèle")
        if not path:
            return
        core.export_paie_bulletins_template(path)
        messagebox.showinfo("Modèle créé", f"Modèle enregistré :\n{path}")

    def import_xlsx(self):
        path = filedialog.askopenfilename(filetypes=[("Classeur Excel", "*.xlsx")],
                                           title="Importer des bulletins de paie")
        if not path:
            return
        try:
            rows = core.parse_paie_bulletins_xlsx(path)
            imported, warnings = core.apply_paie_bulletins_rows(self.conn, rows)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de l'import : {exc}")
            return
        self.refresh()
        msg = f"{imported} bulletin(s) importé(s)/mis à jour."
        if warnings:
            msg += "\n\nAvertissements :\n" + "\n".join(warnings[:20])
        messagebox.showinfo("Import terminé", msg)

    def refresh(self):
        self._refresh_employes()
        for row in self.tree.get_children():
            self.tree.delete(row)
        periode = self.periode_var.get().strip()
        if len(periode) != 7:
            return
        etat = core.compute_paie_periode(self.conn, periode)
        for l in etat["lignes"]:
            self.tree.insert("", "end", iid=str(l["personnel_id"]), values=(
                l["matricule"], l["nom"], l["prenom"] or "", l["classification"],
                f"{l['salaire_base']:,.0f}".replace(",", " "), f"{l['net_percu']:,.0f}".replace(",", " ")))


class PaieEtatTab(ttk.Frame):
    """État de paie calculé pour une période — équivalent de PayrollTab
    dans Paie Burkina : tableau des montants calculés (CNSS, IUTS, net
    perçu, coût employeur...) avec totaux, et impression des bulletins."""

    RESULT_COLS = [
        ("matricule", "Matricule", 80), ("nom", "Nom", 120), ("prenom", "Prénom", 100),
        ("remuneration_totale", "Rém. Totale", 100), ("cnss_salariale", "CNSS", 90),
        ("salaire_brut", "Sal. Brut", 100), ("base_imposable", "Base Imp.", 100),
        ("iuts_net", "IUTS", 90), ("salaire_net", "Salaire Net", 100),
        ("retenue_obligatoire", "Ret. Oblig.", 90), ("retenue_pret", "Ret. Prêt", 90),
        ("net_percu", "Net Perçu", 110), ("cout_total_employeur", "Coût Employeur", 120),
    ]

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.last_etat = None

        top = ttk.Frame(self)
        top.pack(fill="x", padx=12, pady=8)
        ttk.Label(top, text="Période (AAAA-MM) :").pack(side="left")
        self.periode_var = tk.StringVar(value=date.today().strftime("%Y-%m"))
        ttk.Entry(top, textvariable=self.periode_var, width=10).pack(side="left", padx=4)
        ttk.Button(top, text="Calculer la paie", command=self.calculer).pack(side="left", padx=12)
        ttk.Button(top, text="Exporter vers Excel", command=self.export_excel).pack(side="left", padx=4)
        ttk.Button(top, text="Valider la paie (comptabiliser)", command=self.valider).pack(side="left", padx=12)

        self.statut_var = tk.StringVar()
        ttk.Label(self, textvariable=self.statut_var, foreground="#B00020", font=("Segoe UI", 9, "bold")).pack(
            anchor="w", padx=12)

        cols = [c[0] for c in self.RESULT_COLS]
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        for key, label, width in self.RESULT_COLS:
            self.tree.heading(key, text=label)
            self.tree.column(key, width=width, anchor="w" if key in ("matricule", "nom", "prenom") else "e")
        self.tree.pack(fill="both", expand=True, padx=12, pady=(0, 4))
        self.tree.bind("<Double-1>", self._on_double_click)
        ttk.Label(self, text="Double-cliquez une ligne pour l'aperçu avant impression du bulletin.",
                  foreground="#595959").pack(anchor="w", padx=12)

        self.totaux_var = tk.StringVar()
        ttk.Label(self, textvariable=self.totaux_var, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=12, pady=8)

        self.calculer()

    def calculer(self):
        periode = self.periode_var.get().strip()
        if len(periode) != 7 or periode[4] != "-":
            messagebox.showerror("Erreur", "Période invalide — format attendu : AAAA-MM (ex. 2026-08).")
            return
        etat = core.compute_paie_periode(self.conn, periode)
        self.last_etat = etat
        for row in self.tree.get_children():
            self.tree.delete(row)
        for l in etat["lignes"]:
            values = [f"{l[k]:,.0f}".replace(",", " ") if k not in ("matricule", "nom", "prenom") else l[k]
                      for k, _, _ in self.RESULT_COLS]
            self.tree.insert("", "end", iid=str(l["bulletin_id"]), values=values)
        t = etat["totaux"]
        self.totaux_var.set(
            f"Total Net Perçu : {t['net_percu']:,.0f}  |  Total CNSS : {t['cnss_total']:,.0f}  |  "
            f"Total IUTS : {t['iuts_net']:,.0f}  |  Total Ret. Oblig. : {t['retenue_obligatoire']:,.0f}  |  "
            f"Coût total employeur : {t['cout_total_employeur']:,.0f}  F CFA".replace(",", " "))
        if core.est_periode_paie_validee(self.conn, periode):
            self.statut_var.set(
                f"✓ Paie de {periode} déjà VALIDÉE (comptabilisée) — les bulletins ne sont plus modifiables.")
        else:
            self.statut_var.set("")
        if not etat["lignes"]:
            messagebox.showinfo("Info", "Aucun bulletin saisi pour cette période — utilisez l'onglet Bulletins.")

    def valider(self):
        periode = self.periode_var.get().strip()
        if len(periode) != 7 or periode[4] != "-":
            messagebox.showerror("Erreur", "Période invalide — format attendu : AAAA-MM (ex. 2026-08).")
            return
        if core.est_periode_paie_validee(self.conn, periode):
            messagebox.showinfo("Info", f"La paie de {periode} est déjà validée.")
            return
        etat = core.compute_paie_periode(self.conn, periode)
        if not etat["lignes"]:
            messagebox.showwarning("Rien à valider", "Aucun bulletin saisi pour cette période.")
            return
        t = etat["totaux"]
        if not messagebox.askyesno(
            "Confirmer la validation de la paie",
            f"Valider la paie de {periode} pour {len(etat['lignes'])} employé(s) ?\n\n"
            f"Total Net à payer : {t['net_percu']:,.0f} F CFA\n"
            f"Total CNSS (salariale + patronale) : {t['cnss_total']:,.0f} F CFA\n"
            f"Total IUTS : {t['iuts_net']:,.0f} F CFA\n"
            f"Coût total employeur : {t['cout_total_employeur']:,.0f} F CFA\n\n"
            f"Cette action envoie les écritures comptables dans le menu SAISIE (débit charges de "
            f"personnel, crédit CNSS/IUTS/rémunérations dues) et VERROUILLE les bulletins de cette "
            f"période — ils ne pourront plus être modifiés. Cette action est définitive."
        ):
            return
        try:
            _, piece = core.valider_paie_periode(self.conn, periode)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc))
            return
        messagebox.showinfo("Validation terminée",
                             f"Paie de {periode} comptabilisée (pièce {piece}). Les écritures sont visibles "
                             f"dans le menu SAISIE.")
        self.calculer()

    def _on_double_click(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        bulletin_id = int(sel[0])
        html = core.render_bulletin_paie_html(self.conn, bulletin_id)
        import tempfile, webbrowser, os as _os
        path = _os.path.join(tempfile.gettempdir(), f"bulletin_paie_{bulletin_id}.html")
        with open(path, "w", encoding="utf-8") as f:
            f.write(html)
        webbrowser.open(f"file://{path}")

    def export_excel(self):
        if not self.last_etat or not self.last_etat["lignes"]:
            messagebox.showinfo("Info", "Rien à exporter — calculez d'abord la paie de la période.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile=f"Etat_paie_{self.periode_var.get().strip()}.xlsx", title="Exporter l'état de paie")
        if not path:
            return
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Etat de paie"
        header_font = Font(bold=True, color="FFFFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E78")
        for i, (_, label, _) in enumerate(self.RESULT_COLS, start=1):
            c = ws.cell(row=1, column=i, value=label)
            c.font = header_font
            c.fill = header_fill
        for r, l in enumerate(self.last_etat["lignes"], start=2):
            for i, (key, _, _) in enumerate(self.RESULT_COLS, start=1):
                ws.cell(row=r, column=i, value=l[key])
        wb.save(path)
        messagebox.showinfo("Export terminé", f"État de paie exporté :\n{path}")


class PaieParametresTab(ttk.Frame):
    """Paramètres de paie (taux CNSS, plafonds, abattements, exonérations)
    — modifiables par un administrateur. Le barème IUTS et la table de
    réduction pour charges de famille restent fixes (mêmes valeurs que la
    réglementation en vigueur), modifiables uniquement en base si besoin."""

    CHAMPS = [
        ("taux_cnss_salarie", "Taux CNSS salariale", True),
        ("plafond_cnss", "Plafond rémunération CNSS", False),
        ("cnss_salariale_plafonnee", "CNSS salariale plafonnée", False),
        ("taux_cnss_patronale", "Taux CNSS patronale", True),
        ("taux_tpa", "Taux TPA (patronale)", True),
        ("taux_retenue_obligatoire", "Taux retenue obligatoire", True),
        ("abattement_cadre", "Abattement CADRE", True),
        ("abattement_autre", "Abattement AUTRE", True),
        ("taux_plafond_fiscal", "Taux plafond fiscal", True),
    ]

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="PARAMÈTRES DE PAIE", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Taux et plafonds utilisés pour tous les calculs de paie (CNSS, TPA, abattements). "
            "Les indemnités exonérées (Logement/Fonction/Transport) et le barème IUTS restent aux valeurs "
            "réglementaires par défaut."
        ), foreground="#595959", wraplength=1000).pack(anchor="w", padx=16, pady=(0, 12))

        form = ttk.Frame(self)
        form.pack(anchor="w", padx=16)
        self.vars = {}
        for i, (key, label, is_pct) in enumerate(self.CHAMPS):
            ttk.Label(form, text=label + (" (%)" if is_pct else "") + " :").grid(
                row=i, column=0, sticky="w", padx=4, pady=4)
            var = tk.StringVar()
            ttk.Entry(form, textvariable=var, width=14).grid(row=i, column=1, padx=4, pady=4)
            self.vars[key] = (var, is_pct)
        ttk.Button(self, text="Enregistrer les paramètres", command=self.save).pack(
            anchor="w", padx=16, pady=12)

        self.refresh()

    def refresh(self):
        params = core.get_paie_parametres(self.conn)
        for key, (var, is_pct) in self.vars.items():
            val = params.get(key, 0)
            var.set(str(val * 100 if is_pct else val))

    def save(self):
        params = core.get_paie_parametres(self.conn)
        try:
            for key, (var, is_pct) in self.vars.items():
                val = float(var.get())
                params[key] = val / 100 if is_pct else val
        except ValueError:
            messagebox.showerror("Erreur", "Toutes les valeurs doivent être des nombres.")
            return
        core.set_paie_parametres(self.conn, params)
        messagebox.showinfo("Enregistré", "Paramètres de paie enregistrés.")


class PaieTab(ttk.Frame):
    """Regroupe la saisie des bulletins, l'état de paie calculé et les
    paramètres — équivalent complet du module Paie Burkina, intégré ici
    aux employés déjà saisis dans GRH > Personnel."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        inner = ttk.Notebook(self)
        inner.pack(fill="both", expand=True)
        self.bulletins_tab = PaieBulletinsTab(inner, conn)
        self.etat_tab = PaieEtatTab(inner, conn)
        self.params_tab = PaieParametresTab(inner, conn)
        inner.add(self.bulletins_tab, text="Bulletins")
        inner.add(self.etat_tab, text="État de paie")
        inner.add(self.params_tab, text="Paramètres de paie")

    def refresh(self):
        self.bulletins_tab.refresh()
        self.etat_tab.calculer()
        self.params_tab.refresh()


class ArreteComptesTab(ttk.Frame):
    """Tableau de vérification avant clôture (« arrêté de comptes ») —
    rassemble les contrôles habituels de fin de période sur un seul
    écran : comptes fournisseurs, factures non parvenues, comptes
    clients, rapprochements bancaires, impôts, engagements, paie."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.last_resultat = None

        top = ttk.Frame(self)
        top.pack(fill="x", padx=16, pady=(16, 4))
        ttk.Label(top, text="ARRÊTÉ DE COMPTES", font=("Segoe UI", 14, "bold")).pack(side="left")
        ttk.Label(top, text="   Date d'arrêté (JJ/MM/AAAA) :").pack(side="left", padx=(20, 4))
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(top, textvariable=self.date_var, width=12).pack(side="left")
        ttk.Button(top, text="Calculer", command=self.calculer).pack(side="left", padx=12)
        ttk.Label(self, text=(
            "Photographie des points à vérifier avant de clôturer une période — ne modifie rien. "
            "Cliquez sur un onglet pour le détail de chaque contrôle."
        ), foreground="#595959", wraplength=1100).pack(anchor="w", padx=16, pady=(0, 8))

        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=16, pady=(0, 16))

        self.tab_fournisseurs = self._make_tab_tree(
            "Fournisseurs",
            "Comptes 40x avec mouvement sur l'exercice. En rouge : solde DÉBITEUR (anomalie possible — "
            "un fournisseur ne devrait normalement pas nous devoir de l'argent, sauf avance versée).",
            ("compte", "libelle", "solde"), ["Compte", "Libellé", "Solde"], [90, 320, 130])

        self.tab_fnp = self._make_tab_tree(
            "Factures non parvenues",
            "Commandes fournisseurs LIVRÉES mais pour lesquelles aucune facture d'achat n'a été saisie "
            "depuis — à vérifier manuellement (détection approximative, pas de lien direct commande / "
            "facture dans le logiciel).",
            ("piece", "fournisseur", "montant", "date_commande", "date_livraison"),
            ["Pièce", "Fournisseur", "Montant", "Date commande", "Date livraison"], [90, 220, 110, 110, 110])

        self.tab_clients = self._make_tab_tree(
            "Clients",
            "Balance âgée des impayés clients (voir COMMERCIAL > Recouvrement pour le détail par tranche).",
            ("client", "total"), ["Client", "Total impayé"], [300, 150])

        self.tab_banques = self._make_tab_tree(
            "Rapprochements bancaires",
            "Solde comptable de chaque compte banque à la date d'arrêté — le pointage détaillé "
            "mouvement par mouvement se fait dans TRESORERIE > Rapprochement bancaire.",
            ("compte", "libelle", "solde"), ["Compte", "Libellé", "Solde comptable"], [90, 320, 150])

        self.tab_impots = self._make_tab_tree(
            "Impôts & charges sociales",
            "Soldes des comptes TVA (443/444), IUTS/retenues (447) et CNSS (43) à la date d'arrêté.",
            ("compte", "libelle", "solde"), ["Compte", "Libellé", "Solde"], [90, 320, 130])

        self.tab_engagements = self._make_tab_tree(
            "Engagements en retard",
            "Commandes fournisseurs et factures clients en dépassement de délai (livraison ou paiement).",
            ("type", "piece", "tiers", "montant", "statut"),
            ["Type", "Pièce", "Tiers", "Montant", "Statut"], [90, 90, 220, 110, 180])

        self.tab_paie = self._make_tab_tree(
            "Paie",
            "Statut des 3 dernières périodes de paie saisies (validée = déjà comptabilisée et verrouillée).",
            ("periode", "statut"), ["Période", "Statut"], [120, 250])

        self.calculer()

    def _make_tab_tree(self, titre, description, cols, headers, widths):
        frame = ttk.Frame(self.notebook)
        self.notebook.add(frame, text=titre)
        ttk.Label(frame, text=description, foreground="#595959", wraplength=1050).pack(
            anchor="w", padx=8, pady=8)
        tree = ttk.Treeview(frame, columns=cols, show="headings")
        for c, h, w in zip(cols, headers, widths):
            tree.heading(c, text=h)
            tree.column(c, width=w, anchor="w")
        tree.tag_configure("alerte", foreground="#B00020")
        tree.pack(fill="x", padx=8, pady=(0, 8))
        return tree

    def calculer(self):
        date_str = core.to_iso_date(self.date_var.get().strip())
        if not date_str:
            messagebox.showerror("Erreur", "Date invalide — format attendu : JJ/MM/AAAA.")
            return
        resultat = core.compute_arrete_comptes(self.conn, date_arrete=date_str)
        self.last_resultat = resultat

        for tree in (self.tab_fournisseurs, self.tab_fnp, self.tab_clients, self.tab_banques,
                     self.tab_impots, self.tab_engagements, self.tab_paie):
            for row in tree.get_children():
                tree.delete(row)

        anomalies_comptes = {b["code"] for b in resultat["fournisseurs"]["anomalies"]}
        for b in resultat["fournisseurs"]["comptes"]:
            tag = ("alerte",) if b["code"] in anomalies_comptes else ()
            self.tab_fournisseurs.insert("", "end", tags=tag, values=(
                b["code"], b["label"], fmt_cfa(b["solde_cloture"])))

        for c in resultat["factures_non_parvenues"]:
            self.tab_fnp.insert("", "end", tags=("alerte",), values=(
                c["piece"] or "", c["raison_sociale"], fmt_cfa(c["montant"]),
                core.to_display_date(c["date_commande"]), core.to_display_date(c["date_livraison_reelle"])))

        for cl in resultat["clients_balance_agee"]:
            self.tab_clients.insert("", "end", tags=("alerte",), values=(
                cl["raison_sociale"], fmt_cfa(cl["total"])))

        for b in resultat["rapprochements_bancaires"]:
            self.tab_banques.insert("", "end", values=(b["compte"], b["libelle"], fmt_cfa(b["solde_comptable"])))

        for c in resultat["impots"] + resultat["charges_sociales"]:
            if c["solde_fin_periode"]:
                self.tab_impots.insert("", "end", values=(c["code"], c["label"], fmt_cfa(c["solde_fin_periode"])))

        for c in resultat["engagements_fournisseurs_retard"]:
            self.tab_engagements.insert("", "end", tags=("alerte",), values=(
                "Fournisseur", c["piece"] or "", c["raison_sociale"], fmt_cfa(c["montant"]),
                c["statut_livraison"] if c["depassement_livraison"] else c["statut_paiement"]))
        for f in resultat["factures_clients_retard"]:
            self.tab_engagements.insert("", "end", tags=("alerte",), values=(
                "Client", f["piece"] or "", f["raison_sociale"], fmt_cfa(f["montant"]), f["statut_paiement"]))

        for p in resultat["paie_statuts"]:
            tag = () if p["validee"] else ("alerte",)
            self.tab_paie.insert("", "end", tags=tag, values=(
                p["periode"], "✓ Validée (comptabilisée)" if p["validee"] else "En attente de validation"))

    def refresh(self):
        self.calculer()


class TresorerieTab(ttk.Frame):
    """Trésorerie (menu TRESORERIE) — banques alignées horizontalement avec
    Entrées/Sorties de la période, et liste des engagements (Règlements
    validés non encore payés) pour évaluer la capacité à y faire face."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="TRÉSORERIE", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))

        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True, padx=16, pady=8)
        tab_banques = ttk.Frame(notebook)
        tab_engagements = ttk.Frame(notebook)
        notebook.add(tab_banques, text="Banques (Entrées / Sorties)")
        notebook.add(tab_engagements, text="Engagements à payer")

        self._build_banques(tab_banques)
        self._build_engagements(tab_engagements)
        self.refresh()

    def _build_banques(self, parent):
        ttk.Label(parent, text=(
            "Chaque compte de trésorerie (banque, caisse) sur une ligne, avec le solde de début de période, "
            "les Entrées (débit) et Sorties (crédit) de la période, et le solde de fin — exercice courant "
            "par défaut."
        ), foreground="#595959", wraplength=1050).pack(anchor="w", padx=8, pady=(8, 8))
        ttk.Button(parent, text="Actualiser", command=self.refresh).pack(anchor="w", padx=8)

        cols = ("compte", "libelle", "debut", "entrees", "sorties", "fin")
        self.tree_banques = ttk.Treeview(parent, columns=cols, show="headings", height=16)
        headers = ["Compte", "Libellé", "Solde début", "Entrées", "Sorties", "Solde fin"]
        widths = [90, 220, 130, 130, 130, 140]
        for c, h, w in zip(cols, headers, widths):
            self.tree_banques.heading(c, text=h)
            self.tree_banques.column(c, width=w, anchor="w" if c in ("compte", "libelle") else "e")
        self.tree_banques.tag_configure("total", background="#1F4E78", foreground="white", font=("Segoe UI", 9, "bold"))
        self.tree_banques.pack(fill="both", expand=True, padx=8, pady=8)

    def _build_engagements(self, parent):
        ttk.Label(parent, text=(
            "Règlements déjà validés (charge comptabilisée) dont le paiement bancaire n'a pas encore été "
            "enregistré (voir menu ENGAGEMENTS-PROJETS > Règlements) — ce que l'entreprise doit encore "
            "décaisser, comparé à la trésorerie disponible."
        ), foreground="#595959", wraplength=1050).pack(anchor="w", padx=8, pady=(8, 8))
        ttk.Button(parent, text="Actualiser", command=self.refresh).pack(anchor="w", padx=8)

        self.synthese_var = tk.StringVar()
        self.synthese_label = ttk.Label(parent, textvariable=self.synthese_var, font=("Segoe UI", 11, "bold"))
        self.synthese_label.pack(anchor="w", padx=8, pady=(8, 8))

        cols = ("numero", "date", "fournisseur", "montant")
        self.tree_engagements = ttk.Treeview(parent, columns=cols, show="headings", height=14)
        headers = ["N° Règlement", "Date", "Fournisseur", "Montant net à payer"]
        widths = [130, 100, 260, 160]
        for c, h, w in zip(cols, headers, widths):
            self.tree_engagements.heading(c, text=h)
            self.tree_engagements.column(c, width=w, anchor="w" if c != "montant" else "e")
        self.tree_engagements.tag_configure("total", background="#1F4E78", foreground="white", font=("Segoe UI", 9, "bold"))
        self.tree_engagements.pack(fill="both", expand=True, padx=8, pady=8)

    def refresh(self):
        for row in self.tree_banques.get_children():
            self.tree_banques.delete(row)
        lignes, total = core.compute_tresorerie_banques_horizontal(self.conn)
        for l in lignes:
            self.tree_banques.insert("", "end", values=(
                l["code"], l["label"], fmt_cfa(l["solde_debut_periode"]), fmt_cfa(l["debit_periode"]),
                fmt_cfa(l["credit_periode"]), fmt_cfa(l["solde_fin_periode"])))
        self.tree_banques.insert("", "end", tags=("total",), values=(
            "TOTAL", "", fmt_cfa(total["solde_debut_periode"]), fmt_cfa(total["debit_periode"]),
            fmt_cfa(total["credit_periode"]), fmt_cfa(total["solde_fin_periode"])))

        for row in self.tree_engagements.get_children():
            self.tree_engagements.delete(row)
        d = core.compute_engagements_a_payer(self.conn)
        for e in d["engagements"]:
            self.tree_engagements.insert("", "end", values=(
                e["numero"], core.to_display_date(e["date_reglement"]), e["raison_sociale"],
                fmt_cfa(e["net_a_payer"])))
        self.tree_engagements.insert("", "end", tags=("total",), values=(
            "TOTAL ENGAGEMENTS", "", "", fmt_cfa(d["total_engagements"])))

        if d["peut_faire_face"]:
            self.synthese_var.set(
                f"✓ Trésorerie disponible : {fmt_cfa(d['treso_disponible'])}  —  Engagements : "
                f"{fmt_cfa(d['total_engagements'])}  —  Solde après engagements : "
                f"{fmt_cfa(d['solde_apres_engagements'])} — l'entreprise peut faire face à ses engagements.")
            self.synthese_label.configure(foreground="#1F7A1F")
        else:
            self.synthese_var.set(
                f"⚠ Trésorerie disponible : {fmt_cfa(d['treso_disponible'])}  —  Engagements : "
                f"{fmt_cfa(d['total_engagements'])}  —  Solde après engagements : "
                f"{fmt_cfa(d['solde_apres_engagements'])} — insuffisant pour faire face à tous les engagements.")
            self.synthese_label.configure(foreground="#B00020")


class ParcAutoTab(ttk.Frame):
    """Parc automobile — liste des véhicules, sans lien avec la comptabilité."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.selected_id = None
        ttk.Label(self, text="PARC AUTO", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))

        form = ttk.LabelFrame(self, text="Véhicule")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Immatriculation :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.immat_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.immat_var, width=16).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Marque :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.marque_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.marque_var, width=14).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Modèle :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.modele_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.modele_var, width=14).grid(row=0, column=5, padx=4)
        ttk.Label(form, text="Type :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.type_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.type_var, width=16).grid(row=1, column=1, padx=4, pady=(4, 0))
        ttk.Label(form, text="Chauffeur affecté :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.chauffeur_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.chauffeur_var, width=16).grid(row=1, column=3, padx=4, pady=(4, 0))
        ttk.Label(form, text="Statut :").grid(row=1, column=4, sticky="w", padx=(12, 4), pady=(4, 0))
        self.statut_var = tk.StringVar(value="actif")
        ttk.Combobox(form, textvariable=self.statut_var, width=13, state="readonly",
                     values=["actif", "en panne", "en réparation", "vendu"]).grid(row=1, column=5, padx=4, pady=(4, 0))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=6)
        ttk.Button(btns, text="Ajouter", command=self.add).pack(side="left")
        ttk.Button(btns, text="Mettre à jour la sélection", command=self.update_sel).pack(side="left", padx=8)
        ttk.Button(btns, text="Supprimer la sélection", command=self.delete_sel).pack(side="left")
        ttk.Button(btns, text="Effacer le formulaire", command=self.clear_form).pack(side="left", padx=8)

        cols = ("id", "immat", "marque", "modele", "type", "chauffeur", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=20)
        for c, h, w in zip(cols, ["ID", "Immatriculation", "Marque", "Modèle", "Type", "Chauffeur", "Statut"],
                           [40, 140, 120, 120, 120, 150, 110]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_id = v[0]
        self.immat_var.set(v[1]); self.marque_var.set(v[2]); self.modele_var.set(v[3])
        self.type_var.set(v[4]); self.chauffeur_var.set(v[5]); self.statut_var.set(v[6])

    def clear_form(self):
        self.selected_id = None
        for var in (self.immat_var, self.marque_var, self.modele_var, self.type_var, self.chauffeur_var):
            var.set("")
        self.statut_var.set("actif")

    def add(self):
        if not self.immat_var.get().strip():
            messagebox.showwarning("Champ manquant", "L'immatriculation est obligatoire.")
            return
        core.add_vehicule(self.conn, self.immat_var.get().strip(), marque=self.marque_var.get().strip(),
                           modele=self.modele_var.get().strip(), type_vehicule=self.type_var.get().strip(),
                           chauffeur_affecte=self.chauffeur_var.get().strip(), statut=self.statut_var.get())
        self.clear_form()
        self.refresh()

    def update_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un véhicule.")
            return
        core.update_vehicule(self.conn, self.selected_id, immatriculation=self.immat_var.get().strip(),
                              marque=self.marque_var.get().strip(), modele=self.modele_var.get().strip(),
                              type_vehicule=self.type_var.get().strip(),
                              chauffeur_affecte=self.chauffeur_var.get().strip(), statut=self.statut_var.get())
        self.clear_form()
        self.refresh()

    def delete_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un véhicule.")
            return
        if messagebox.askyesno("Confirmer", "Supprimer ce véhicule ?"):
            core.delete_vehicule(self.conn, self.selected_id)
            self.clear_form()
            self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for v in core.list_vehicules(self.conn):
            self.tree.insert("", "end", values=(v["id"], v["immatriculation"], v["marque"] or "", v["modele"] or "",
                                                  v["type_vehicule"] or "", v["chauffeur_affecte"] or "", v["statut"]))


class MissionsTab(ttk.Frame):
    """Missions des véhicules du parc auto — sans lien avec la comptabilité."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="MISSIONS", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))

        form = ttk.LabelFrame(self, text="Nouvelle mission")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Véhicule :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.vehicule_var = tk.StringVar()
        self.vehicule_combo = ttk.Combobox(form, textvariable=self.vehicule_var, width=22, state="readonly")
        self.vehicule_combo.grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Chauffeur :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.chauffeur_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.chauffeur_var, width=16).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Destination :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.destination_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.destination_var, width=18).grid(row=0, column=5, padx=4)
        ttk.Label(form, text="Motif :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.motif_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.motif_var, width=22).grid(row=1, column=1, padx=4, pady=(4, 0))
        ttk.Label(form, text="Date départ :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.date_depart_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(form, textvariable=self.date_depart_var, width=12).grid(row=1, column=3, padx=4, pady=(4, 0))
        ttk.Label(form, text="Date retour :").grid(row=1, column=4, sticky="w", padx=(12, 4), pady=(4, 0))
        self.date_retour_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.date_retour_var, width=12).grid(row=1, column=5, padx=4, pady=(4, 0))
        ttk.Label(form, text="Km départ :").grid(row=2, column=0, sticky="w", padx=4, pady=(4, 0))
        self.km_depart_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.km_depart_var, width=10).grid(row=2, column=1, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(form, text="Km retour :").grid(row=2, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.km_retour_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.km_retour_var, width=10).grid(row=2, column=3, padx=4, pady=(4, 0), sticky="w")
        ttk.Button(form, text="Ajouter la mission", command=self.add).grid(row=2, column=5, padx=4, pady=(4, 0))

        cols = ("id", "vehicule", "chauffeur", "destination", "depart", "retour", "km", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=18)
        headers = ["ID", "Véhicule", "Chauffeur", "Destination", "Départ", "Retour", "Km parcourus", "Statut"]
        widths = [40, 130, 130, 160, 90, 90, 110, 100]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.refresh()

    def _refresh_vehicule_values(self):
        self.vehicules = core.list_vehicules(self.conn)
        self.vehicule_combo["values"] = [f"{v['id']} — {v['immatriculation']}" for v in self.vehicules]

    def add(self):
        if not self.destination_var.get().strip():
            messagebox.showwarning("Champ manquant", "La destination est obligatoire.")
            return
        vehicule_id = None
        raw = self.vehicule_var.get()
        if raw:
            vehicule_id = int(raw.split(" — ", 1)[0])
        try:
            km_depart = float(self.km_depart_var.get()) if self.km_depart_var.get().strip() else None
            km_retour = float(self.km_retour_var.get()) if self.km_retour_var.get().strip() else None
        except ValueError:
            messagebox.showerror("Erreur", "Les km doivent être des nombres.")
            return
        core.add_mission(
            self.conn, self.destination_var.get().strip(), vehicule_id=vehicule_id,
            chauffeur=self.chauffeur_var.get().strip(), motif=self.motif_var.get().strip(),
            date_depart=core.to_iso_date(self.date_depart_var.get().strip()),
            date_retour=core.to_iso_date(self.date_retour_var.get().strip()),
            km_depart=km_depart, km_retour=km_retour,
        )
        self.destination_var.set(""); self.motif_var.set(""); self.chauffeur_var.set("")
        self.km_depart_var.set(""); self.km_retour_var.set("")
        self.refresh()

    def refresh(self):
        self._refresh_vehicule_values()
        for row in self.tree.get_children():
            self.tree.delete(row)
        for m in core.list_missions(self.conn):
            km = ""
            if m["km_depart"] is not None and m["km_retour"] is not None:
                km = f"{m['km_retour'] - m['km_depart']:g}"
            self.tree.insert("", "end", values=(
                m["id"], m["immatriculation"], m["chauffeur"] or "", m["destination"],
                core.to_display_date(m["date_depart"] or ""), core.to_display_date(m["date_retour"] or ""),
                km, m["statut"]))


class PiecesRechangeTab(ttk.Frame):
    """Stock de pièces de rechange — PARTAGÉ entre le menu TRANSPORT
    (réparations de véhicules) et le menu MAINTENANCE-ÉNERGIE (maintenance
    générale). Sans lien avec la comptabilité."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.selected_id = None
        ttk.Label(self, text="PIÈCES DE RECHANGE", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Stock partagé, utilisé aussi bien pour les réparations de véhicules (menu TRANSPORT) que pour "
            "la maintenance générale (menu MAINTENANCE-ÉNERGIE)."
        ), foreground="#595959").pack(anchor="w", padx=16, pady=(0, 8))

        form = ttk.LabelFrame(self, text="Pièce")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Code :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.code_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.code_var, width=12).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Désignation :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.designation_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.designation_var, width=28).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Quantité en stock :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.qte_var = tk.StringVar(value="0")
        ttk.Entry(form, textvariable=self.qte_var, width=10).grid(row=1, column=1, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(form, text="Unité :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.unite_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.unite_var, width=10).grid(row=1, column=3, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(form, text="Coût unitaire :").grid(row=2, column=0, sticky="w", padx=4, pady=(4, 0))
        self.cout_var = tk.StringVar(value="0")
        ttk.Entry(form, textvariable=self.cout_var, width=12).grid(row=2, column=1, padx=4, pady=(4, 0), sticky="w")

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=6)
        ttk.Button(btns, text="Ajouter", command=self.add).pack(side="left")
        ttk.Button(btns, text="Mettre à jour la sélection", command=self.update_sel).pack(side="left", padx=8)
        ttk.Button(btns, text="Supprimer la sélection", command=self.delete_sel).pack(side="left")
        ttk.Button(btns, text="Effacer le formulaire", command=self.clear_form).pack(side="left", padx=8)

        cols = ("id", "code", "designation", "qte", "unite", "cout", "valeur")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=18)
        headers = ["ID", "Code", "Désignation", "Qté en stock", "Unité", "Coût unitaire", "Valeur du stock"]
        widths = [40, 90, 260, 100, 80, 110, 130]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_id = v[0]
        self.code_var.set(v[1]); self.designation_var.set(v[2]); self.qte_var.set(v[3])
        self.unite_var.set(v[4]); self.cout_var.set(v[5])

    def clear_form(self):
        self.selected_id = None
        self.code_var.set(""); self.designation_var.set(""); self.qte_var.set("0")
        self.unite_var.set(""); self.cout_var.set("0")

    def _parse(self):
        if not self.designation_var.get().strip():
            messagebox.showwarning("Champ manquant", "La désignation est obligatoire.")
            return None
        try:
            qte = float(self.qte_var.get() or 0)
            cout = float(self.cout_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Quantité et coût unitaire doivent être des nombres.")
            return None
        return qte, cout

    def add(self):
        parsed = self._parse()
        if not parsed:
            return
        qte, cout = parsed
        core.add_piece_rechange(self.conn, self.designation_var.get().strip(), code=self.code_var.get().strip(),
                                 quantite_stock=qte, unite=self.unite_var.get().strip(), cout_unitaire=cout)
        self.clear_form()
        self.refresh()

    def update_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord une pièce.")
            return
        parsed = self._parse()
        if not parsed:
            return
        qte, cout = parsed
        core.update_piece_rechange(self.conn, self.selected_id, code=self.code_var.get().strip(),
                                    designation=self.designation_var.get().strip(), quantite_stock=qte,
                                    unite=self.unite_var.get().strip(), cout_unitaire=cout)
        self.clear_form()
        self.refresh()

    def delete_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord une pièce.")
            return
        if messagebox.askyesno("Confirmer", "Supprimer cette pièce ?"):
            core.delete_piece_rechange(self.conn, self.selected_id)
            self.clear_form()
            self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for p in core.list_pieces_rechange(self.conn):
            valeur = p["quantite_stock"] * p["cout_unitaire"]
            self.tree.insert("", "end", values=(
                p["id"], p["code"] or "", p["designation"], f"{p['quantite_stock']:g}", p["unite"] or "",
                f"{fmt_cfa(p['cout_unitaire'])}", f"{fmt_cfa(valeur)}"))


class ReparationDialog(tk.Toplevel):
    """Détail d'une réparation (double-clic) — pièces utilisées (décrémente
    automatiquement le stock de Pièces de rechange) + main d'œuvre."""

    def __init__(self, parent, conn, reparation_id, on_saved):
        super().__init__(parent)
        self.conn = conn
        self.reparation_id = reparation_id
        self.on_saved = on_saved
        self.title("Réparation")
        self.geometry("850x560")
        self.transient(parent)
        self.grab_set()

        rep = core.get_reparation(conn, reparation_id)
        header = ttk.LabelFrame(self, text="Informations")
        header.pack(fill="x", padx=10, pady=8)
        ttk.Label(header, text="Description :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.description_var = tk.StringVar(value=rep["description"])
        ttk.Entry(header, textvariable=self.description_var, width=40).grid(row=0, column=1, padx=4)
        ttk.Label(header, text="Garage :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.garage_var = tk.StringVar(value=rep["garage"] or "")
        ttk.Entry(header, textvariable=self.garage_var, width=20).grid(row=0, column=3, padx=4)
        ttk.Label(header, text="Main d'œuvre :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.mo_var = tk.StringVar(value=str(rep["cout_main_oeuvre"]))
        ttk.Entry(header, textvariable=self.mo_var, width=14).grid(row=1, column=1, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(header, text="Statut :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.statut_var = tk.StringVar(value=rep["statut"])
        ttk.Combobox(header, textvariable=self.statut_var, width=17, state="readonly",
                     values=["en_cours", "terminee"]).grid(row=1, column=3, padx=4, pady=(4, 0))

        lignes_frame = ttk.LabelFrame(self, text="Pièces utilisées (décrémente le stock)")
        lignes_frame.pack(fill="both", padx=10, pady=6)
        form = ttk.Frame(lignes_frame)
        form.pack(fill="x", padx=6, pady=4)
        ttk.Label(form, text="Pièce :").grid(row=0, column=0, sticky="w")
        self.piece_var = tk.StringVar()
        self.piece_combo = ttk.Combobox(form, textvariable=self.piece_var, width=32, state="readonly")
        self.piece_combo.grid(row=0, column=1, padx=4)
        self._refresh_pieces()
        ttk.Label(form, text="Quantité :").grid(row=0, column=2, sticky="w", padx=(12, 0))
        self.qte_var = tk.StringVar(value="1")
        ttk.Entry(form, textvariable=self.qte_var, width=8).grid(row=0, column=3, padx=4)
        ttk.Button(form, text="Ajouter", command=self.add_ligne).grid(row=0, column=4, padx=12)

        cols = ("id", "designation", "qte", "cout_unit", "montant")
        self.tree = ttk.Treeview(lignes_frame, columns=cols, show="headings", height=10)
        for c, h, w in zip(cols, ["ID", "Pièce", "Quantité", "Coût unit.", "Montant"], [40, 300, 90, 100, 110]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", padx=6, pady=6)
        ttk.Button(lignes_frame, text="Supprimer la ligne sélectionnée (restitue le stock)",
                   command=self.delete_ligne).pack(anchor="w", padx=6, pady=(0, 6))
        self.total_var = tk.StringVar()
        ttk.Label(lignes_frame, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=6, pady=(0, 6))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=10)
        ttk.Button(btns, text="Enregistrer", command=self.save).pack(side="left", padx=4)
        ttk.Button(btns, text="Fermer", command=self.destroy).pack(side="left", padx=4)

        self.refresh_lignes()

    def _refresh_pieces(self):
        self.pieces = core.list_pieces_rechange(self.conn)
        self.piece_combo["values"] = [f"{p['id']} — {p['designation']} ({p['quantite_stock']:g} en stock)"
                                       for p in self.pieces]

    def add_ligne(self):
        raw = self.piece_var.get()
        if not raw:
            messagebox.showwarning("Champ manquant", "Choisissez une pièce.", parent=self)
            return
        piece_id = int(raw.split(" — ", 1)[0])
        try:
            qte = float(self.qte_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "La quantité doit être un nombre.", parent=self)
            return
        try:
            core.add_ligne_reparation(self.conn, self.reparation_id, piece_id, quantite=qte)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc), parent=self)
            return
        self.qte_var.set("1")
        self._refresh_pieces()
        self.refresh_lignes()

    def delete_ligne(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne.", parent=self)
            return
        ligne_id = self.tree.item(sel[0], "values")[0]
        core.delete_ligne_reparation(self.conn, ligne_id)
        self._refresh_pieces()
        self.refresh_lignes()

    def refresh_lignes(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for l in core.list_lignes_reparation(self.conn, self.reparation_id):
            self.tree.insert("", "end", values=(
                l["id"], l["designation"], f"{l['quantite']:g}", f"{fmt_cfa(l['cout_unitaire'])}", f"{fmt_cfa(l['montant'])}"))
        cout_total = core.compute_cout_total_reparation(self.conn, self.reparation_id)
        self.total_var.set(f"Coût total (pièces + main d'œuvre) : {fmt_cfa(cout_total)}")

    def save(self):
        try:
            mo = float(self.mo_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "La main d'œuvre doit être un nombre.", parent=self)
            return
        core.update_reparation(self.conn, self.reparation_id, description=self.description_var.get().strip(),
                                garage=self.garage_var.get().strip(), cout_main_oeuvre=mo,
                                statut=self.statut_var.get())
        self.refresh_lignes()
        messagebox.showinfo("Enregistré", "Réparation enregistrée.", parent=self)
        self.on_saved()


class ReparationsTab(ttk.Frame):
    """Réparations de véhicules — double-clic pour ajouter les pièces
    utilisées (décrémente le stock de Pièces de rechange)."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="RÉPARATIONS", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 8))

        form = ttk.Frame(self)
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Véhicule :").grid(row=0, column=0, sticky="w")
        self.vehicule_var = tk.StringVar()
        self.vehicule_combo = ttk.Combobox(form, textvariable=self.vehicule_var, width=22, state="readonly")
        self.vehicule_combo.grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Description :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.description_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.description_var, width=30).grid(row=0, column=3, padx=4)
        ttk.Button(form, text="Nouvelle réparation", command=self.new_reparation).grid(row=0, column=4, padx=12)

        cols = ("id", "vehicule", "date", "description", "garage", "cout_mo", "cout_total", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=20)
        headers = ["ID", "Véhicule", "Date", "Description", "Garage", "Main d'œuvre", "Coût total", "Statut"]
        widths = [40, 130, 90, 220, 130, 100, 100, 100]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<Double-1>", self._on_double_click)
        self._by_iid = {}
        self.refresh()

    def _refresh_vehicule_values(self):
        self.vehicules = core.list_vehicules(self.conn)
        self.vehicule_combo["values"] = [f"{v['id']} — {v['immatriculation']}" for v in self.vehicules]

    def new_reparation(self):
        if not self.description_var.get().strip():
            messagebox.showwarning("Champ manquant", "La description est obligatoire.")
            return
        vehicule_id = None
        raw = self.vehicule_var.get()
        if raw:
            vehicule_id = int(raw.split(" — ", 1)[0])
        rid = core.create_reparation(self.conn, self.description_var.get().strip(), vehicule_id=vehicule_id)
        self.description_var.set("")
        self.refresh()
        ReparationDialog(self, self.conn, rid, on_saved=self.refresh)

    def _on_double_click(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        rid = self._by_iid.get(sel[0])
        if rid:
            ReparationDialog(self, self.conn, rid, on_saved=self.refresh)

    def refresh(self):
        self._refresh_vehicule_values()
        for row in self.tree.get_children():
            self.tree.delete(row)
        self._by_iid = {}
        for r in core.list_reparations(self.conn):
            cout_total = core.compute_cout_total_reparation(self.conn, r["id"])
            iid = self.tree.insert("", "end", values=(
                r["id"], r["immatriculation"] or "", core.to_display_date(r["date_reparation"]), r["description"],
                r["garage"] or "", f"{fmt_cfa(r['cout_main_oeuvre'])}", f"{fmt_cfa(cout_total)}", r["statut"]))
            self._by_iid[iid] = r["id"]


class ImmobilisationsTab(ttk.Frame):
    """Liste des comptes de classe 2 (immobilisations), avec fournisseur,
    prix d'achat, catégorie/taux d'amortissement, et Valeur Brute /
    Amortissement / Valeur Nette (calculés depuis la Balance, même méthode
    exacte que le Bilan)."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.selected_compte = None
        ttk.Label(self, text="IMMOBILISATIONS", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Comptes de classe 2 ayant un solde dans la Balance. Sélectionnez une ligne pour renseigner son "
            "fournisseur et son prix d'achat. Le taux d'amortissement par catégorie se paramètre dans le "
            "sous-menu « Amortissements »."
        ), foreground="#595959", wraplength=1100).pack(anchor="w", padx=16, pady=(0, 8))

        import_bar = ttk.Frame(self)
        import_bar.pack(fill="x", padx=16, pady=(0, 4))
        ttk.Button(import_bar, text="Importer des immobilisations (.xlsx)", command=self.import_xlsx).pack(
            side="left", padx=2)
        ttk.Button(import_bar, text="Télécharger un modèle (.xlsx)", command=self.download_template).pack(
            side="left", padx=2)

        form = ttk.LabelFrame(self, text="Fiche du compte sélectionné")
        form.pack(fill="x", padx=16, pady=4)
        self.compte_label_var = tk.StringVar(value="(sélectionnez une ligne dans le tableau ci-dessous)")
        ttk.Label(form, textvariable=self.compte_label_var, font=("Segoe UI", 9, "bold")).grid(
            row=0, column=0, columnspan=4, sticky="w", padx=4, pady=4)
        ttk.Label(form, text="Fournisseur :").grid(row=1, column=0, sticky="w", padx=4)
        self.fournisseur_var = tk.StringVar()
        self.fournisseur_combo = ttk.Combobox(form, textvariable=self.fournisseur_var, width=28)
        self.fournisseur_combo.grid(row=1, column=1, padx=4, pady=4)
        self.fournisseur_combo.bind("<KeyRelease>", self._on_fournisseur_keyrelease)
        self._refresh_fournisseur_values()
        ttk.Label(form, text="Prix d'achat :").grid(row=1, column=2, sticky="w", padx=(12, 4))
        self.prix_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.prix_var, width=16).grid(row=1, column=3, padx=4)
        ttk.Label(form, text="Date d'acquisition :").grid(row=1, column=4, sticky="w", padx=(12, 4))
        self.date_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.date_var, width=12).grid(row=1, column=5, padx=4)
        ttk.Label(form, text="Base de répartition (quantité annuelle) :").grid(
            row=2, column=0, sticky="w", padx=4, pady=(4, 0))
        self.base_qte_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.base_qte_var, width=12).grid(row=2, column=1, padx=4, pady=(4, 0))
        ttk.Label(form, text="Unité (tonnes, heures...) :").grid(row=2, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.base_unite_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.base_unite_var, width=16).grid(row=2, column=3, padx=4, pady=(4, 0))
        ttk.Label(form, text="Amortissement annuel (si pas comptabilisé) :").grid(
            row=2, column=4, sticky="w", padx=(12, 4), pady=(4, 0))
        self.amort_manuel_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.amort_manuel_var, width=16).grid(row=2, column=5, padx=4, pady=(4, 0))
        ttk.Label(form, text=(
            "Pour utiliser cet équipement dans une recette de fabrication (composant « Amortissement "
            "d'équipement ») : indiquez sa capacité annuelle normale (ex. 5000 tonnes/an ou 2000 heures/an). "
            "Le coût unitaire = amortissement RÉELLEMENT comptabilisé (dotations 68x/28x déjà saisies) ÷ "
            "cette capacité ; si aucune dotation n'est encore comptabilisée pour cet équipement, le montant "
            "« Amortissement annuel » saisi ci-dessus est utilisé à la place, en attendant."
        ), foreground="#595959", wraplength=1050).grid(row=3, column=0, columnspan=6, sticky="w", padx=4, pady=(2, 0))
        ttk.Button(form, text="Enregistrer la fiche", command=self.save_fiche).grid(row=1, column=6, padx=12)
        ttk.Button(form, text="Modifier la fiche", command=self.save_fiche).grid(row=2, column=6, padx=12, pady=(4, 0))

        cols = ("compte", "libelle", "categorie", "fournisseur", "prix_achat", "taux", "brut", "amort", "net")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=18)
        headers = ["Compte", "Libellé", "Catégorie", "Fournisseur", "Prix d'achat", "Taux %",
                    "Valeur brute", "Amortissement", "Valeur nette"]
        widths = [80, 200, 220, 160, 110, 60, 110, 110, 110]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        ttk.Button(self, text="Actualiser", command=self.refresh).pack(anchor="w", padx=16, pady=(0, 12))
        self.refresh()

    def _refresh_fournisseur_values(self):
        self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}"
                                             for f in core.list_fournisseurs(self.conn)]

    def _on_fournisseur_keyrelease(self, event=None):
        query = self.fournisseur_var.get().strip()
        items = core.list_fournisseurs(self.conn, query)
        self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_compte = v[0]
        self.compte_label_var.set(f"{v[0]} — {v[1]}")
        fiche = core.get_immobilisation_fiche(self.conn, v[0])
        fournisseur = core.get_fournisseur(self.conn, fiche["fournisseur_code"]) if fiche["fournisseur_code"] else None
        self.fournisseur_var.set(f"{fiche['fournisseur_code']} — {fournisseur['raison_sociale']}"
                                  if fournisseur else (fiche["fournisseur_code"] or ""))
        self.prix_var.set(str(fiche["prix_achat"]) if fiche["prix_achat"] else "")
        self.date_var.set(core.to_display_date(fiche["date_acquisition"] or ""))
        self.base_qte_var.set(str(fiche["base_repartition_quantite"]) if fiche.get("base_repartition_quantite") else "")
        self.base_unite_var.set(fiche.get("base_repartition_unite") or "")
        self.amort_manuel_var.set(
            str(fiche["amortissement_annuel_manuel"]) if fiche.get("amortissement_annuel_manuel") else "")

    def save_fiche(self):
        if not self.selected_compte:
            messagebox.showinfo("Info", "Sélectionnez d'abord un compte dans le tableau.")
            return
        raw = self.fournisseur_var.get().strip()
        fournisseur_code = raw.split(" — ", 1)[0].strip() if " — " in raw else raw
        try:
            prix = float(self.prix_var.get()) if self.prix_var.get().strip() else 0
        except ValueError:
            messagebox.showerror("Erreur", "Le prix d'achat doit être un nombre.")
            return
        base_qte = None
        if self.base_qte_var.get().strip():
            try:
                base_qte = float(self.base_qte_var.get())
            except ValueError:
                messagebox.showerror("Erreur", "La base de répartition doit être un nombre.")
                return
        amort_manuel = None
        if self.amort_manuel_var.get().strip():
            try:
                amort_manuel = float(self.amort_manuel_var.get())
            except ValueError:
                messagebox.showerror("Erreur", "L'amortissement annuel doit être un nombre.")
                return
        core.set_immobilisation_fiche(self.conn, self.selected_compte, fournisseur_code=fournisseur_code or None,
                                       prix_achat=prix, date_acquisition=core.to_iso_date(self.date_var.get().strip()),
                                       base_repartition_quantite=base_qte,
                                       base_repartition_unite=self.base_unite_var.get().strip() or None,
                                       amortissement_annuel_manuel=amort_manuel)
        self.refresh()
        messagebox.showinfo("Enregistré", "Fiche d'immobilisation enregistrée.")

    def download_template(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Modele_immobilisations.xlsx", title="Enregistrer le modèle",
        )
        if not path:
            return
        core.export_immobilisations_template(path)
        messagebox.showinfo("Modèle créé", f"Modèle enregistré :\n{path}")

    def import_xlsx(self):
        path = filedialog.askopenfilename(filetypes=[("Classeur Excel", "*.xlsx")],
                                           title="Importer des immobilisations")
        if not path:
            return
        try:
            imported, warnings = core.import_immobilisations_from_xlsx(self.conn, path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de l'import : {exc}")
            return
        self.refresh()
        msg = f"{imported} fiche(s) d'immobilisation importée(s)/mise(s) à jour."
        if warnings:
            msg += "\n\nAvertissements :\n" + "\n".join(warnings[:20])
        messagebox.showinfo("Import terminé", msg)

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for l in core.compute_immobilisations_liste(self.conn):
            self.tree.insert("", "end", values=(
                l["compte"], l["libelle"], l["categorie"], l["fournisseur_nom"] or l["fournisseur_code"] or "",
                f"{fmt_cfa(l['prix_achat'])}" if l["prix_achat"] else "", f"{l['taux_pct']:g}",
                f"{fmt_cfa(l['valeur_brute'])}", f"{fmt_cfa(l['amortissement'])}", f"{fmt_cfa(l['valeur_nette'])}"))


class AmortissementsTab(ttk.Frame):
    """Taux d'amortissement par catégorie d'immobilisation (même
    catégorisation que le Bilan) — utilisés à titre indicatif dans le
    sous-menu Immobilisations."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="TAUX D'AMORTISSEMENT PAR CATÉGORIE", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 8))
        cols = ("categorie", "taux")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=12)
        self.tree.heading("categorie", text="Catégorie d'immobilisation")
        self.tree.heading("taux", text="Taux annuel (%)")
        self.tree.column("categorie", width=560, anchor="w")
        self.tree.column("taux", width=120, anchor="e")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<Double-1>", self._on_double_click)
        ttk.Label(self, text="Double-cliquez sur une catégorie pour modifier son taux.",
                  foreground="#595959").pack(anchor="w", padx=16)
        self.refresh()

    def _on_double_click(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        values = self.tree.item(sel[0], "values")
        categorie, taux_actuel = values[0], values[1]
        nouveau = simpledialog.askfloat("Taux d'amortissement", f"Taux annuel (%) pour « {categorie} » :",
                                         initialvalue=float(taux_actuel), parent=self, minvalue=0, maxvalue=100)
        if nouveau is None:
            return
        core.set_taux_amortissement(self.conn, categorie, nouveau)
        self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for t in core.list_taux_amortissement(self.conn):
            self.tree.insert("", "end", values=(t["categorie"], f"{t['taux_pct']:g}"))


class PlaceholderTab(ttk.Frame):
    """Page pas encore développée : structure de menu en place, contenu à venir."""

    def __init__(self, parent, conn, title, description):
        super().__init__(parent)
        ttk.Label(self, text=title, font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=24, pady=(24, 8))
        ttk.Label(self, text=description, wraplength=900, foreground="#595959").pack(anchor="w", padx=24)
        ttk.Label(self, text="Fonctionnalité pas encore développée — dites-moi si vous voulez que je "
                              "la construise en priorité.", foreground="#B00020").pack(anchor="w", padx=24, pady=(16, 0))


class AnalytiquePeriodeTab(ttk.Frame):
    """Coûts d'une catégorie de codes analytiques (Énergie ou Maintenance),
    par code (eau, électricité, essence... / véhicules, bâtiments...), sur
    une période librement choisie — alimenté par toute écriture de Saisie
    taguée avec ce code analytique (champ « Code analytique »), ainsi que
    par les lignes de recette de Fabrication qui lui sont associées (menu
    PRODUCTION > Fabrication)."""

    def __init__(self, parent, conn, title, description, prefix, suggestions):
        super().__init__(parent)
        self.conn = conn
        self.prefix = prefix
        self.suggestions = suggestions
        ttk.Label(self, text=title, font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=description, foreground="#595959", wraplength=1050).pack(anchor="w", padx=16, pady=(0, 8))

        filt = ttk.Frame(self)
        filt.pack(fill="x", padx=16, pady=4)
        ttk.Label(filt, text="Du (JJ/MM/AAAA) :").pack(side="left")
        self.date_from_var = tk.StringVar()
        ttk.Entry(filt, textvariable=self.date_from_var, width=12).pack(side="left", padx=4)
        ttk.Label(filt, text="Au (JJ/MM/AAAA) :").pack(side="left", padx=(12, 0))
        self.date_to_var = tk.StringVar()
        ttk.Entry(filt, textvariable=self.date_to_var, width=12).pack(side="left", padx=4)
        ttk.Button(filt, text="Afficher", command=self.refresh).pack(side="left", padx=8)
        ttk.Button(filt, text="Exercice entier", command=self._reset_filter).pack(side="left", padx=2)
        ttk.Button(filt, text="Ajouter les codes courants", command=self._add_suggestions).pack(side="left", padx=12)

        cols = ("code", "libelle", "debut", "debit", "credit", "fin")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        headers = ["Code analytique", "Libellé", "Charge début période", "Débit période",
                   "Crédit période (avoir)", "Charge cumulée fin période"]
        widths = [130, 280, 140, 120, 140, 160]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.tag_configure("total", background="#1F4E78", foreground="white", font=("Segoe UI", 10, "bold"))
        self.tree.pack(fill="both", padx=16, pady=8)

        ttk.Label(self, text=(
            "Pour qu'une charge apparaisse ici : dans l'onglet Saisie, renseignez le champ "
            "« Code analytique » avec l'un des codes ci-dessous, sur la ligne du compte de charge "
            "(classe 6) — ex. 605100 Eau, électricité... Créez ou gérez ces codes dans PARAMÈTRES > "
            "Plan analytique."
        ), foreground="#595959", wraplength=1050).pack(anchor="w", padx=16)

        self.total_var = tk.StringVar()
        ttk.Label(self, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=16, pady=(4, 12))
        self.refresh()

    def _reset_filter(self):
        self.date_from_var.set("")
        self.date_to_var.set("")
        self.refresh()

    def _add_suggestions(self):
        n = core.ajouter_codes_analytiques_suggeres(self.conn, self.suggestions)
        if n:
            messagebox.showinfo("Codes ajoutés", f"{n} code(s) analytique(s) ajouté(s).")
        else:
            messagebox.showinfo("Rien à ajouter", "Tous les codes courants existent déjà.")
        self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        date_from = core.to_iso_date(self.date_from_var.get()) if self.date_from_var.get().strip() else None
        date_to = core.to_iso_date(self.date_to_var.get()) if self.date_to_var.get().strip() else None
        codes = core.compute_couts_analytiques_categorie(self.conn, self.prefix, date_from=date_from, date_to=date_to)
        total_debut = total_debit = total_credit = total_fin = 0.0
        for c in codes:
            self.tree.insert("", "end", values=(
                c["code"], c["label"], f"{fmt_cfa(c['solde_debut_periode'])}",
                f"{fmt_cfa(c['debit_periode'])}", f"{fmt_cfa(c['credit_periode'])}", f"{fmt_cfa(c['solde_fin_periode'])}",
            ))
            total_debut += c["solde_debut_periode"]
            total_debit += c["debit_periode"]
            total_credit += c["credit_periode"]
            total_fin += c["solde_fin_periode"]
        self.tree.insert("", "end", tags=("total",), values=(
            "", "TOTAL", f"{fmt_cfa(total_debut)}", f"{fmt_cfa(total_debit)}", f"{fmt_cfa(total_credit)}", f"{fmt_cfa(total_fin)}",
        ))
        periode = f"du {self.date_from_var.get()} au {self.date_to_var.get()}" if date_from or date_to else \
            f"exercice {core.get_current_exercice(self.conn)} entier"
        self.total_var.set(f"{len(codes)} code(s) avec charge — période : {periode}.")


class ClassePeriodeTab(ttk.Frame):
    """Tous les comptes d'une classe donnée (ex. 44 Impôts, 43 Organismes
    sociaux), en Solde début de période / Mouvements Débit-Crédit / Solde fin
    de période, sur une période librement choisie (par défaut l'exercice
    comptable entier). Calculé à partir de la même Balance générale que les
    autres écrans — toujours cohérent avec elle."""

    def __init__(self, parent, conn, title, description, prefix):
        super().__init__(parent)
        self.conn = conn
        self.prefix = prefix
        ttk.Label(self, text=title, font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=description, foreground="#595959", wraplength=1050).pack(anchor="w", padx=16, pady=(0, 8))

        filt = ttk.Frame(self)
        filt.pack(fill="x", padx=16, pady=4)
        ttk.Label(filt, text="Du (JJ/MM/AAAA) :").pack(side="left")
        self.date_from_var = tk.StringVar()
        ttk.Entry(filt, textvariable=self.date_from_var, width=12).pack(side="left", padx=4)
        ttk.Label(filt, text="Au (JJ/MM/AAAA) :").pack(side="left", padx=(12, 0))
        self.date_to_var = tk.StringVar()
        ttk.Entry(filt, textvariable=self.date_to_var, width=12).pack(side="left", padx=4)
        ttk.Button(filt, text="Afficher", command=self.refresh).pack(side="left", padx=8)
        ttk.Button(filt, text="Exercice entier", command=self._reset_filter).pack(side="left", padx=2)
        ttk.Label(filt, text="(par défaut : exercice comptable en cours entier)",
                  foreground="#595959").pack(side="left", padx=10)

        cols = ("compte", "libelle", "debut", "debit", "credit", "fin")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        headers = ["Compte", "Libellé", "Solde début période", "Débit période", "Crédit période", "Solde fin période"]
        widths = [90, 320, 140, 120, 120, 140]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.tag_configure("total", background="#1F4E78", foreground="white", font=("Segoe UI", 10, "bold"))
        self.tree.pack(fill="both", padx=16, pady=8)
        self.total_var = tk.StringVar()
        ttk.Label(self, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=16, pady=(0, 12))
        self.refresh()

    def _reset_filter(self):
        self.date_from_var.set("")
        self.date_to_var.set("")
        self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        date_from = core.to_iso_date(self.date_from_var.get()) if self.date_from_var.get().strip() else None
        date_to = core.to_iso_date(self.date_to_var.get()) if self.date_to_var.get().strip() else None
        comptes = core.compute_comptes_prefixe_periode(self.conn, self.prefix, date_from=date_from, date_to=date_to)
        total_debut = total_debit = total_credit = total_fin = 0.0
        for c in comptes:
            self.tree.insert("", "end", values=(
                c["code"], c["label"], f"{fmt_cfa(c['solde_debut_periode'])}",
                f"{fmt_cfa(c['debit_periode'])}", f"{fmt_cfa(c['credit_periode'])}", f"{fmt_cfa(c['solde_fin_periode'])}",
            ))
            total_debut += c["solde_debut_periode"]
            total_debit += c["debit_periode"]
            total_credit += c["credit_periode"]
            total_fin += c["solde_fin_periode"]
        self.tree.insert("", "end", tags=("total",), values=(
            "", f"TOTAL CLASSE {self.prefix}", f"{fmt_cfa(total_debut)}",
            f"{fmt_cfa(total_debit)}", f"{fmt_cfa(total_credit)}", f"{fmt_cfa(total_fin)}",
        ))
        periode = f"du {self.date_from_var.get()} au {self.date_to_var.get()}" if date_from or date_to else \
            f"exercice {core.get_current_exercice(self.conn)} entier"
        self.total_var.set(f"{len(comptes)} compte(s) avec solde ou mouvement — période : {periode}.")


class RapprochementBancaireTab(ttk.Frame):
    """Rapprochement bancaire : tous les comptes de banque (racine 52,
    détaillés compte par compte à 6 chiffres), mouvement par mouvement sur
    la période choisie, avec une case à cocher par mouvement pour le pointer
    comme retrouvé dans le relevé bancaire papier — le pointage est enregistré
    et reste visible à la prochaine ouverture."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self._row_entry_ids = {}
        ttk.Label(self, text="RAPPROCHEMENT BANCAIRE", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Tous les comptes de banque (52xxxx), détaillés compte par compte, avec chaque mouvement de la "
            "période choisie. Cliquez sur la colonne « Pointé » pour cocher/décocher un mouvement retrouvé "
            "dans le relevé bancaire papier — le pointage est mémorisé."
        ), foreground="#595959", wraplength=1100).pack(anchor="w", padx=16, pady=(0, 8))

        filt = ttk.Frame(self)
        filt.pack(fill="x", padx=16, pady=4)
        ttk.Label(filt, text="Du (JJ/MM/AAAA) :").pack(side="left")
        self.date_from_var = tk.StringVar()
        ttk.Entry(filt, textvariable=self.date_from_var, width=12).pack(side="left", padx=4)
        ttk.Label(filt, text="Au (JJ/MM/AAAA) :").pack(side="left", padx=(12, 0))
        self.date_to_var = tk.StringVar()
        ttk.Entry(filt, textvariable=self.date_to_var, width=12).pack(side="left", padx=4)
        ttk.Button(filt, text="Afficher", command=self.refresh).pack(side="left", padx=8)
        ttk.Button(filt, text="Exercice entier", command=self._reset_filter).pack(side="left", padx=2)

        cols = ("pointe", "date", "piece", "libelle", "debit", "credit", "solde")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=24)
        headers = ["Pointé", "Date", "Pièce", "Libellé", "Débit", "Crédit", "Solde cumulé"]
        widths = [60, 85, 70, 320, 100, 100, 120]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            anchor = "center" if c in ("pointe", "debit", "credit", "solde") else "w"
            self.tree.column(c, width=w, anchor=anchor)
        self.tree.tag_configure("compte_header", background="#B4C6E7", font=("Segoe UI", 9, "bold"))
        self.tree.tag_configure("compte_footer", background="#DCE6F1", font=("Segoe UI", 9, "bold"))
        self.tree.tag_configure("pointe", background="#D9EAD3")
        self.tree.pack(fill="both", padx=16, pady=8)
        self.tree.bind("<Button-1>", self._on_click)

        self.ecart_var = tk.StringVar()
        ttk.Label(self, textvariable=self.ecart_var, font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=16, pady=(0, 12))
        self.refresh()

    def _reset_filter(self):
        self.date_from_var.set("")
        self.date_to_var.set("")
        self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self._row_entry_ids = {}
        date_from = core.to_iso_date(self.date_from_var.get()) if self.date_from_var.get().strip() else None
        date_to = core.to_iso_date(self.date_to_var.get()) if self.date_to_var.get().strip() else None
        comptes = core.compute_mouvements_prefixe_periode(self.conn, "52", date_from=date_from, date_to=date_to)

        total_periode = 0.0
        total_pointe = 0.0
        for c in comptes:
            self.tree.insert("", "end", tags=("compte_header",), values=(
                "", "", "", f"{c['code']} {c['label']} — solde début de période : {fmt_cfa(c['solde_debut_periode'])}",
                "", "", "",
            ))
            for m in c["mouvements"]:
                iid = self.tree.insert("", "end", tags=("pointe",) if m["pointe"] else (), values=(
                    "☑" if m["pointe"] else "☐", core.to_display_date(m["date"]), m["piece"] or "",
                    m["libelle"] or "", f"{fmt_cfa(m['debit'])}" if m["debit"] else "",
                    f"{fmt_cfa(m['credit'])}" if m["credit"] else "", f"{fmt_cfa(m['solde_cumule'])}",
                ))
                self._row_entry_ids[iid] = m["id"]
                total_periode += m["debit"] - m["credit"]
            self.tree.insert("", "end", tags=("compte_footer",), values=(
                "", "", "", f"  Solde fin de période — {c['code']}", "", "", f"{fmt_cfa(c['solde_fin_periode'])}",
            ))
            total_pointe += c["total_pointe"]

        self.ecart_var.set(
            f"Total pointé (retrouvé dans le relevé) : {fmt_cfa(total_pointe)}    "
            f"Total des mouvements de la période : {fmt_cfa(total_periode)}    "
            f"Écart non pointé : {fmt_cfa(total_periode - total_pointe)}"
        )

    def _on_click(self, event):
        if self.tree.identify_region(event.x, event.y) != "cell":
            return
        if self.tree.identify_column(event.x) != "#1":
            return
        row = self.tree.identify_row(event.y)
        if row not in self._row_entry_ids:
            return
        entry_id = self._row_entry_ids[row]
        deja_pointe = self.tree.set(row, "pointe") == "☑"
        core.set_pointage_bancaire(self.conn, entry_id, not deja_pointe)
        self.refresh()


class VentesTab(ttk.Frame):
    """Soldes des opérations avec chaque client, total par client,
    avec filtre sur une plage de dates."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="VENTES — SOLDES PAR CLIENT", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Solde = Débit − Crédit sur les comptes clients (411xxx) taggés à chaque client dans "
            "la Saisie. Positif = montant restant dû par le client (à recouvrer)."
        ), foreground="#595959", wraplength=1000).pack(anchor="w", padx=16, pady=(0, 8))

        filt = ttk.Frame(self)
        filt.pack(fill="x", padx=16, pady=4)
        ttk.Label(filt, text="Du (JJ/MM/AAAA) :").pack(side="left")
        self.date_from_var = tk.StringVar()
        ttk.Entry(filt, textvariable=self.date_from_var, width=12).pack(side="left", padx=4)
        ttk.Label(filt, text="Au (JJ/MM/AAAA) :").pack(side="left", padx=(12, 0))
        self.date_to_var = tk.StringVar()
        ttk.Entry(filt, textvariable=self.date_to_var, width=12).pack(side="left", padx=4)
        ttk.Button(filt, text="Filtrer", command=self.refresh).pack(side="left", padx=8)
        ttk.Button(filt, text="Réinitialiser", command=self._reset_filter).pack(side="left", padx=2)

        cols = ("code", "raison_sociale", "debit", "credit", "solde")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        headers = ["Code", "Client", "Total Débit", "Total Crédit", "Solde (dû si positif)"]
        widths = [90, 320, 120, 120, 160]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", padx=16, pady=8)
        self.total_var = tk.StringVar()
        ttk.Label(self, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=16, pady=(0, 12))
        self.refresh()

    def _reset_filter(self):
        self.date_from_var.set("")
        self.date_to_var.set("")
        self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        date_from = core.to_iso_date(self.date_from_var.get()) if self.date_from_var.get().strip() else None
        date_to = core.to_iso_date(self.date_to_var.get()) if self.date_to_var.get().strip() else None
        ventes, total_debit, total_credit = core.compute_ventes_par_client(
            self.conn, date_from=date_from, date_to=date_to)
        for v in ventes:
            self.tree.insert("", "end", values=(
                v["code"], v["raison_sociale"], f"{fmt_cfa(v['debit'])}", f"{fmt_cfa(v['credit'])}", f"{fmt_cfa(v['solde'])}"
            ))
        self.total_var.set(
            f"TOTAL — Débit : {fmt_cfa(total_debit)}   Crédit : {fmt_cfa(total_credit)}   "
            f"Solde global à recouvrer : {fmt_cfa(total_debit - total_credit)}"
        )


class AchatsTab(ttk.Frame):
    """Soldes des opérations avec chaque fournisseur, total par fournisseur,
    avec filtre sur une plage de dates."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="ACHATS — SOLDES PAR FOURNISSEUR", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Solde = Débit − Crédit sur les comptes fournisseurs (401xxx/408xxx) taggés à chaque "
            "fournisseur dans la Saisie. Négatif = montant restant dû au fournisseur."
        ), foreground="#595959", wraplength=1000).pack(anchor="w", padx=16, pady=(0, 8))

        filt = ttk.Frame(self)
        filt.pack(fill="x", padx=16, pady=4)
        ttk.Label(filt, text="Du (JJ/MM/AAAA) :").pack(side="left")
        self.date_from_var = tk.StringVar()
        ttk.Entry(filt, textvariable=self.date_from_var, width=12).pack(side="left", padx=4)
        ttk.Label(filt, text="Au (JJ/MM/AAAA) :").pack(side="left", padx=(12, 0))
        self.date_to_var = tk.StringVar()
        ttk.Entry(filt, textvariable=self.date_to_var, width=12).pack(side="left", padx=4)
        ttk.Button(filt, text="Filtrer", command=self.refresh).pack(side="left", padx=8)
        ttk.Button(filt, text="Réinitialiser", command=self._reset_filter).pack(side="left", padx=2)

        cols = ("code", "raison_sociale", "debit", "credit", "solde")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        headers = ["Code", "Fournisseur", "Total Débit", "Total Crédit", "Solde (dû si négatif)"]
        widths = [90, 320, 120, 120, 160]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", padx=16, pady=8)
        self.total_var = tk.StringVar()
        ttk.Label(self, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=16, pady=(0, 12))
        self.refresh()

    def _reset_filter(self):
        self.date_from_var.set("")
        self.date_to_var.set("")
        self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        date_from = core.to_iso_date(self.date_from_var.get()) if self.date_from_var.get().strip() else None
        date_to = core.to_iso_date(self.date_to_var.get()) if self.date_to_var.get().strip() else None
        achats, total_debit, total_credit = core.compute_achats_par_fournisseur(
            self.conn, date_from=date_from, date_to=date_to)
        for a in achats:
            self.tree.insert("", "end", values=(
                a["code"], a["raison_sociale"], f"{fmt_cfa(a['debit'])}", f"{fmt_cfa(a['credit'])}", f"{fmt_cfa(a['solde'])}"
            ))
        self.total_var.set(
            f"TOTAL — Débit : {fmt_cfa(total_debit)}   Crédit : {fmt_cfa(total_credit)}   "
            f"Solde global : {fmt_cfa(total_debit - total_credit)}"
        )


class MargesTab(ttk.Frame):
    """Marge commerciale et valeur ajoutée, calculées comme dans la Liasse fiscale."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.text = tk.Text(self, font=("Consolas", 11), wrap="none")
        self.text.pack(fill="both", expand=True, padx=16, pady=16)
        self.refresh()

    def refresh(self):
        cr = core.compute_liasse_resultat(self.conn)
        label_ca = "Chiffre d'affaires (XB)"
        label_re = "Résultat d'exploitation (XE)"
        lines = [
            "MARGES BÉNÉFICIAIRES", "=" * 60, "",
            f"  {'Ventes de marchandises (TA)':<45} {cr['TA']:>14,.2f}",
            f"  {'Achats de marchandises (RA)':<45} {-cr['RA']:>14,.2f}",
            f"  {'MARGE COMMERCIALE (XA)':<45} {cr['XA']:>14,.2f}", "",
            f"  {label_ca:<45} {cr['XB']:>14,.2f}",
            f"  {'VALEUR AJOUTÉE (XC)':<45} {cr['XC']:>14,.2f}",
            f"  {label_re:<45} {cr['XE']:>14,.2f}",
            f"  {'RÉSULTAT NET (XI)':<45} {cr['XI']:>14,.2f}",
        ]
        self.text.delete("1.0", "end")
        self.text.insert("1.0", "\n".join(lines))


class ClientsTab(ttk.Frame):
    """Liste auxiliaire des clients : créer / modifier / importer."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="CLIENTS (LISTE AUXILIAIRE)", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 8))
        ttk.Label(self, text=(
            "Ces fiches sont rattachées à la racine 41 (Clients et comptes rattachés) du Plan "
            "comptable — les écritures qui les taguent doivent utiliser un compte 41xxxx."
        ), foreground="#595959").pack(anchor="w", padx=16, pady=(0, 4))

        form = ttk.Frame(self)
        form.pack(fill="x", padx=16, pady=4)
        labels = ["Code", "Raison sociale", "Contact", "Téléphone", "Adresse", "Délai paiement (jours)"]
        self.vars = {k: tk.StringVar() for k in labels}
        for i, lbl in enumerate(labels):
            r, c = divmod(i, 4)
            ttk.Label(form, text=lbl + " :").grid(row=r * 2, column=c, sticky="w", padx=4, pady=(4, 0))
            ttk.Entry(form, textvariable=self.vars[lbl], width=22).grid(
                row=r * 2 + 1, column=c, sticky="we", padx=4, pady=(0, 4))
        btns = ttk.Frame(form)
        btns.grid(row=2, column=0, columnspan=4, sticky="w", pady=6)
        ttk.Button(btns, text="Créer / Modifier", command=self.save).pack(side="left", padx=2)
        ttk.Button(btns, text="Supprimer", command=self.delete).pack(side="left", padx=2)
        ttk.Button(btns, text="Effacer le formulaire", command=self.clear_form).pack(side="left", padx=2)

        import_bar = ttk.Frame(self)
        import_bar.pack(fill="x", padx=16, pady=(4, 4))
        ttk.Button(import_bar, text="Importer des clients (.xlsx)", command=self.import_xlsx).pack(side="left", padx=2)
        ttk.Button(import_bar, text="Télécharger un modèle (.xlsx)", command=self.download_template).pack(side="left", padx=2)

        search_bar = ttk.Frame(self)
        search_bar.pack(fill="x", padx=16, pady=4)
        ttk.Label(search_bar, text="Rechercher :").pack(side="left")
        self.search_var = tk.StringVar()
        se = ttk.Entry(search_bar, textvariable=self.search_var, width=30)
        se.pack(side="left", padx=6)
        se.bind("<KeyRelease>", lambda e: self.refresh())

        web_bar = ttk.LabelFrame(self, text="Trouver de nouveaux clients / prospects sur Internet")
        web_bar.pack(fill="x", padx=16, pady=(4, 4))
        ttk.Label(web_bar, text="Produit / service vendu :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.web_produit_var = tk.StringVar()
        ttk.Entry(web_bar, textvariable=self.web_produit_var, width=28).grid(row=0, column=1, padx=4)
        ttk.Label(web_bar, text="Ville :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.web_ville_var = tk.StringVar(value="Ouagadougou")
        ttk.Entry(web_bar, textvariable=self.web_ville_var, width=18).grid(row=0, column=3, padx=4)
        ttk.Button(web_bar, text="Rechercher sur Internet", command=self.rechercher_internet).grid(
            row=0, column=4, padx=12)
        ttk.Label(web_bar, text=(
            "Ouvre votre navigateur avec une recherche Google déjà remplie (entreprises susceptibles "
            "d'acheter ce produit/service dans cette ville). Copiez ensuite les coordonnées du prospect "
            "choisi dans le formulaire ci-dessus pour l'enregistrer."
        ), foreground="#595959", wraplength=1050).grid(row=1, column=0, columnspan=5, sticky="w", padx=4, pady=(2, 4))

        cols = ("code", "raison_sociale", "contact", "telephone", "adresse", "dp")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        headers = ["Code", "Raison sociale", "Contact", "Téléphone", "Adresse", "Délai paiement (j)"]
        widths = [90, 220, 130, 110, 220, 130]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.vars["Code"].set(v[0])
        self.vars["Raison sociale"].set(v[1])
        self.vars["Contact"].set(v[2])
        self.vars["Téléphone"].set(v[3])
        self.vars["Adresse"].set(v[4])
        self.vars["Délai paiement (jours)"].set(v[5])

    def rechercher_internet(self):
        produit = self.web_produit_var.get().strip()
        if not produit:
            messagebox.showwarning("Champ manquant", "Indiquez le produit ou service vendu.")
            return
        ville = self.web_ville_var.get().strip()
        requete = f"entreprises acheteurs {produit} {ville}".strip()
        import webbrowser
        from urllib.parse import quote_plus
        webbrowser.open(f"https://www.google.com/search?q={quote_plus(requete)}")

    def clear_form(self):
        for v in self.vars.values():
            v.set("")

    def save(self):
        code = self.vars["Code"].get().strip()
        raison = self.vars["Raison sociale"].get().strip()
        if not code or not raison:
            messagebox.showwarning("Champs manquants", "Code et Raison sociale sont obligatoires.")
            return
        try:
            dp = int(self.vars["Délai paiement (jours)"].get() or 30)
        except ValueError:
            messagebox.showerror("Erreur", "Le délai de paiement doit être un nombre entier de jours.")
            return
        core.add_client(self.conn, code, raison, self.vars["Contact"].get().strip(),
                         self.vars["Téléphone"].get().strip(), self.vars["Adresse"].get().strip(), dp)
        self.refresh()

    def delete(self):
        code = self.vars["Code"].get().strip()
        if not code:
            messagebox.showinfo("Info", "Sélectionnez d'abord un client.")
            return
        if messagebox.askyesno("Confirmer", f"Supprimer le client {code} ?"):
            core.delete_client(self.conn, code)
            self.clear_form()
            self.refresh()

    def download_template(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Modele_clients.xlsx", title="Enregistrer le modèle",
        )
        if not path:
            return
        core.export_clients_template(path)
        messagebox.showinfo("Modèle créé", f"Modèle enregistré :\n{path}")

    def import_xlsx(self):
        path = filedialog.askopenfilename(filetypes=[("Classeur Excel", "*.xlsx")],
                                           title="Importer des clients")
        if not path:
            return
        try:
            imported, warnings = core.import_clients_from_xlsx(self.conn, path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de l'import : {exc}")
            return
        self.refresh()
        msg = f"{imported} client(s) importé(s)."
        if warnings:
            msg += "\n\nAvertissements :\n" + "\n".join(warnings[:20])
        messagebox.showinfo("Import terminé", msg)

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for c in core.list_clients(self.conn, self.search_var.get().strip() or None):
            self.tree.insert("", "end", values=(
                c["code"], c["raison_sociale"], c["contact"] or "", c["telephone"] or "",
                c["adresse"] or "", c["delai_paiement_jours"],
            ))


class FournisseursTab(ttk.Frame):
    """Liste auxiliaire des fournisseurs : créer / modifier / importer."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="FOURNISSEURS (LISTE AUXILIAIRE)", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 8))
        ttk.Label(self, text=(
            "Ces fiches sont rattachées à la racine 40 (Fournisseurs et comptes rattachés) du Plan "
            "comptable — les écritures qui les taguent doivent utiliser un compte 40xxxx."
        ), foreground="#595959").pack(anchor="w", padx=16, pady=(0, 4))

        form = ttk.Frame(self)
        form.pack(fill="x", padx=16, pady=4)
        labels = ["Code", "Raison sociale", "Contact", "Téléphone", "Adresse",
                  "Délai paiement (jours)", "Délai livraison (jours)"]
        self.vars = {k: tk.StringVar() for k in labels}
        for i, lbl in enumerate(labels):
            r, c = divmod(i, 4)
            ttk.Label(form, text=lbl + " :").grid(row=r * 2, column=c, sticky="w", padx=4, pady=(4, 0))
            ttk.Entry(form, textvariable=self.vars[lbl], width=22).grid(
                row=r * 2 + 1, column=c, sticky="we", padx=4, pady=(0, 4))
        btns = ttk.Frame(form)
        btns.grid(row=4, column=0, columnspan=4, sticky="w", pady=6)
        ttk.Button(btns, text="Créer / Modifier", command=self.save).pack(side="left", padx=2)
        ttk.Button(btns, text="Supprimer", command=self.delete).pack(side="left", padx=2)
        ttk.Button(btns, text="Effacer le formulaire", command=self.clear_form).pack(side="left", padx=2)

        import_bar = ttk.Frame(self)
        import_bar.pack(fill="x", padx=16, pady=(4, 4))
        ttk.Button(import_bar, text="Importer des fournisseurs (.xlsx)", command=self.import_xlsx).pack(side="left", padx=2)
        ttk.Button(import_bar, text="Télécharger un modèle (.xlsx)", command=self.download_template).pack(side="left", padx=2)

        search_bar = ttk.Frame(self)
        search_bar.pack(fill="x", padx=16, pady=4)
        ttk.Label(search_bar, text="Rechercher :").pack(side="left")
        self.search_var = tk.StringVar()
        se = ttk.Entry(search_bar, textvariable=self.search_var, width=30)
        se.pack(side="left", padx=6)
        se.bind("<KeyRelease>", lambda e: self.refresh())

        web_bar = ttk.LabelFrame(self, text="Trouver de nouveaux fournisseurs sur Internet")
        web_bar.pack(fill="x", padx=16, pady=(4, 4))
        ttk.Label(web_bar, text="Produit recherché :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.web_produit_var = tk.StringVar()
        ttk.Entry(web_bar, textvariable=self.web_produit_var, width=28).grid(row=0, column=1, padx=4)
        ttk.Label(web_bar, text="Ville :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.web_ville_var = tk.StringVar(value="Ouagadougou")
        ttk.Entry(web_bar, textvariable=self.web_ville_var, width=18).grid(row=0, column=3, padx=4)
        ttk.Button(web_bar, text="Rechercher sur Internet", command=self.rechercher_internet).grid(
            row=0, column=4, padx=12)
        ttk.Label(web_bar, text=(
            "Ouvre votre navigateur avec une recherche Google déjà remplie. Les résultats s'affichent "
            "dans le navigateur, pas ici — copiez ensuite les coordonnées du fournisseur choisi dans le "
            "formulaire ci-dessus pour l'enregistrer."
        ), foreground="#595959", wraplength=1050).grid(row=1, column=0, columnspan=5, sticky="w", padx=4, pady=(2, 4))

        cols = ("code", "raison_sociale", "contact", "telephone", "adresse", "dp", "dl")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        headers = ["Code", "Raison sociale", "Contact", "Téléphone", "Adresse",
                   "Délai paiement (j)", "Délai livraison (j)"]
        widths = [90, 220, 130, 110, 200, 110, 110]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.vars["Code"].set(v[0])
        self.vars["Raison sociale"].set(v[1])
        self.vars["Contact"].set(v[2])
        self.vars["Téléphone"].set(v[3])
        self.vars["Adresse"].set(v[4])
        self.vars["Délai paiement (jours)"].set(v[5])
        self.vars["Délai livraison (jours)"].set(v[6])

    def rechercher_internet(self):
        produit = self.web_produit_var.get().strip()
        if not produit:
            messagebox.showwarning("Champ manquant", "Indiquez le produit recherché.")
            return
        ville = self.web_ville_var.get().strip()
        requete = f"fournisseur {produit} {ville}".strip()
        import webbrowser
        from urllib.parse import quote_plus
        webbrowser.open(f"https://www.google.com/search?q={quote_plus(requete)}")

    def clear_form(self):
        for v in self.vars.values():
            v.set("")

    def save(self):
        code = self.vars["Code"].get().strip()
        raison = self.vars["Raison sociale"].get().strip()
        if not code or not raison:
            messagebox.showwarning("Champs manquants", "Code et Raison sociale sont obligatoires.")
            return
        try:
            dp = int(self.vars["Délai paiement (jours)"].get() or 30)
            dl = int(self.vars["Délai livraison (jours)"].get() or 15)
        except ValueError:
            messagebox.showerror("Erreur", "Les délais doivent être des nombres entiers de jours.")
            return
        core.add_fournisseur(self.conn, code, raison, self.vars["Contact"].get().strip(),
                              self.vars["Téléphone"].get().strip(), self.vars["Adresse"].get().strip(),
                              dp, dl)
        self.refresh()

    def delete(self):
        code = self.vars["Code"].get().strip()
        if not code:
            messagebox.showinfo("Info", "Sélectionnez d'abord un fournisseur.")
            return
        if messagebox.askyesno("Confirmer", f"Supprimer le fournisseur {code} ?"):
            core.delete_fournisseur(self.conn, code)
            self.clear_form()
            self.refresh()

    def download_template(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Modele_fournisseurs.xlsx", title="Enregistrer le modèle",
        )
        if not path:
            return
        core.export_fournisseurs_template(path)
        messagebox.showinfo("Modèle créé", f"Modèle enregistré :\n{path}")

    def import_xlsx(self):
        path = filedialog.askopenfilename(filetypes=[("Classeur Excel", "*.xlsx")],
                                           title="Importer des fournisseurs")
        if not path:
            return
        try:
            imported, warnings = core.import_fournisseurs_from_xlsx(self.conn, path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de l'import : {exc}")
            return
        self.refresh()
        msg = f"{imported} fournisseur(s) importé(s)."
        if warnings:
            msg += "\n\nAvertissements :\n" + "\n".join(warnings[:20])
        messagebox.showinfo("Import terminé", msg)

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for f in core.list_fournisseurs(self.conn, self.search_var.get().strip() or None):
            self.tree.insert("", "end", values=(
                f["code"], f["raison_sociale"], f["contact"] or "", f["telephone"] or "",
                f["adresse"] or "", f["delai_paiement_jours"], f["delai_livraison_jours"],
            ))


class FacturationTab(ttk.Frame):
    """Facturation clients : présente directement une facture (entête, lignes de
    vente liées à un compte 70x, TVA paramétrable, pied de page), et sa
    validation envoie les écritures comptables en Saisie — avec sortie de stock
    automatique pour les lignes liées aux marchandises (31) ou produits finis (36)."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.current_facture_id = None

        # ---- Barre du haut : liste des factures ----
        top = ttk.Frame(self)
        top.pack(fill="x", padx=12, pady=8)
        ttk.Label(top, text="Facture n° :").pack(side="left")
        self.facture_combo = ttk.Combobox(top, width=40, state="readonly")
        self.facture_combo.pack(side="left", padx=4)
        self.facture_combo.bind("<<ComboboxSelected>>", self._on_facture_selected)
        ttk.Button(top, text="Nouvelle facture", command=self.new_facture).pack(side="left", padx=8)
        ttk.Button(top, text="Supprimer cette facture", command=self.delete_facture).pack(side="left", padx=2)
        self.corriger_btn = ttk.Button(top, text="Corriger cette facture (erreur sur les chiffres)",
                                        command=self.corriger_facture)
        self.corriger_btn.pack(side="left", padx=2)
        ttk.Button(top, text="Aperçu avant impression", command=self.imprimer_facture).pack(side="left", padx=2)
        self.statut_var = tk.StringVar()
        ttk.Label(top, textvariable=self.statut_var, font=("Segoe UI", 10, "bold")).pack(side="left", padx=16)

        # ---- Entête modifiable ----
        ttk.Label(self, text="En-tête de la facture (modifiable) :").pack(anchor="w", padx=12)
        self.entete_text = tk.Text(self, height=3, font=("Segoe UI", 10))
        self.entete_text.pack(fill="x", padx=12, pady=(0, 8))

        # ---- Champs d'en-tête structurés ----
        info = ttk.Frame(self)
        info.pack(fill="x", padx=12, pady=4)
        ttk.Label(info, text="N° Facture :").grid(row=0, column=0, sticky="w", padx=4)
        self.numero_var = tk.StringVar()
        ttk.Entry(info, textvariable=self.numero_var, width=16).grid(row=0, column=1, padx=4)
        ttk.Label(info, text="Date (JJ/MM/AAAA) :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(info, textvariable=self.date_var, width=14).grid(row=0, column=3, padx=4)
        ttk.Label(info, text="Client (compte 41) :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.client_var = tk.StringVar()
        self.client_combo = ttk.Combobox(info, textvariable=self.client_var, width=26)
        self.client_combo.grid(row=0, column=5, padx=4)
        self.client_combo.bind("<KeyRelease>", self._on_client_keyrelease)
        self.client_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_client_values()
        ttk.Label(info, text="TVA % (compte 44) :").grid(row=0, column=6, sticky="w", padx=(12, 4))
        self.tva_var = tk.StringVar(value=str(core.get_setting(conn, "tva_taux_defaut", core.TVA_TAUX_DEFAUT)))
        ttk.Entry(info, textvariable=self.tva_var, width=6).grid(row=0, column=7, padx=4)
        ttk.Label(info, text="Préréglage (ADMIN) :").grid(row=1, column=6, sticky="w", padx=(12, 4), pady=(4, 0))
        self.tva_preset_var = tk.StringVar()
        self.tva_preset_combo = ttk.Combobox(info, textvariable=self.tva_preset_var, width=18, state="readonly")
        self.tva_preset_combo.grid(row=1, column=7, padx=4, pady=(4, 0))
        self.tva_preset_combo.bind("<<ComboboxSelected>>", self._on_tva_preset_selected)
        self.tva_preset_combo.bind("<Button-1>", self._open_dropdown)
        self.tva_compte_var = tk.StringVar(value=core.get_text_setting(conn, "tva_compte_defaut", core.COMPTE_TVA_VENTES))
        self._refresh_tva_presets()

        # ---- Lignes ----
        form = ttk.LabelFrame(self, text="Ajouter une ligne (produit/service vendu — compte 70x)")
        form.pack(fill="x", padx=12, pady=6)
        ttk.Label(form, text="Compte de vente :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.ligne_compte_var = tk.StringVar()
        self.ligne_compte_combo = ttk.Combobox(form, textvariable=self.ligne_compte_var, width=34)
        self.ligne_compte_combo.grid(row=0, column=1, padx=4)
        self.ligne_compte_combo.bind("<KeyRelease>", self._on_ligne_compte_keyrelease)
        self.ligne_compte_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_ligne_compte_values()
        ttk.Label(form, text="Libellé :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.ligne_libelle_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_libelle_var, width=26).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Quantité :").grid(row=1, column=0, sticky="w", padx=4, pady=4)
        self.ligne_qte_var = tk.StringVar(value="1")
        ttk.Entry(form, textvariable=self.ligne_qte_var, width=10).grid(row=1, column=1, sticky="w", padx=4)
        ttk.Label(form, text="Prix unitaire :").grid(row=1, column=2, sticky="w", padx=(12, 4))
        self.ligne_prix_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_prix_var, width=14).grid(row=1, column=3, sticky="w", padx=4)
        ttk.Label(form, text="Code analytique :").grid(row=1, column=4, sticky="w", padx=(12, 4))
        self.ligne_analytic_var = tk.StringVar()
        self.ligne_analytic_combo = ttk.Combobox(form, textvariable=self.ligne_analytic_var, width=20)
        self.ligne_analytic_combo.grid(row=1, column=5, padx=4, sticky="w")
        self.ligne_analytic_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_ligne_analytic_values()
        ttk.Button(form, text="Ajouter la ligne", command=self.add_ligne).grid(row=1, column=6, padx=12)

        cols = ("id", "compte", "libelle", "type_stock", "qte", "prix", "analytique", "montant")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=6)
        headers = ["ID", "Compte", "Libellé", "Impact stock", "Qté", "Prix unit.", "Analytique", "Montant HT"]
        widths = [40, 90, 200, 110, 60, 90, 130, 110]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", padx=12, pady=6)
        ttk.Button(self, text="Supprimer la ligne sélectionnée", command=self.delete_ligne).pack(anchor="w", padx=12)

        self.totals_var = tk.StringVar()
        ttk.Label(self, textvariable=self.totals_var, font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=12, pady=(8, 0))

        # ---- Pied de page modifiable ----
        ttk.Label(self, text="Pied de page de la facture (modifiable) :").pack(anchor="w", padx=12, pady=(8, 0))
        self.pied_text = tk.Text(self, height=3, font=("Segoe UI", 10))
        self.pied_text.pack(fill="x", padx=12, pady=(0, 8))

        # ---- Validation ----
        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=12, pady=8)
        ttk.Button(btns, text="Enregistrer (brouillon)", command=self.save_facture).pack(side="left", padx=2)
        ttk.Button(btns, text="Valider et envoyer en Saisie", command=self.valider).pack(side="left", padx=2)

        self.refresh_factures_list()

    # -- Client --
    def _refresh_client_values(self):
        items = core.list_clients(self.conn)
        self.client_combo["values"] = [f"{c['code']} — {c['raison_sociale']}" for c in items]

    def _on_client_keyrelease(self, event=None):
        query = self._extract_code(self.client_var.get())
        if query:
            items = core.list_clients(self.conn, query)
            self.client_combo["values"] = [f"{c['code']} — {c['raison_sociale']}" for c in items]

    # -- Compte de vente --
    def _refresh_ligne_compte_values(self):
        items = core.search_accounts(self.conn, "7", limit=100)
        items = [a for a in items if a["classe"] == "7"]
        self.ligne_compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _on_ligne_compte_keyrelease(self, event=None):
        query = self._extract_code(self.ligne_compte_var.get())
        if query:
            items = [a for a in core.search_accounts(self.conn, query, limit=50) if a["classe"] == "7"]
            self.ligne_compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _refresh_ligne_analytic_values(self):
        codes = core.list_analytic_codes(self.conn)
        self.ligne_analytic_combo["values"] = [f"{c['code']} — {c['label']}" for c in codes]

    @staticmethod
    def _open_dropdown(event=None):
        if event is not None:
            event.widget.event_generate("<Down>")

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    # -- Gestion des factures --
    def refresh_factures_list(self):
        factures = core.list_factures_vente(self.conn)
        values = [f"{f['numero']} — {f['raison_sociale']} — {f['statut']}" for f in factures]
        self.facture_combo["values"] = values
        self._factures_cache = factures
        if self.current_facture_id is None and factures:
            self.current_facture_id = factures[0]["id"]
            self.facture_combo.current(0)
        self.load_facture()

    def new_facture(self):
        numero = simpledialog.askstring("Nouvelle facture", "N° de facture :", parent=self)
        if not numero:
            return
        client_code = self._extract_code(self.client_var.get())
        if not client_code or not core.client_exists(self.conn, client_code):
            messagebox.showinfo("Client requis", "Choisissez d'abord un client existant dans le champ Client.")
            return
        date_str = core.to_iso_date(self.date_var.get().strip()) or date.today().strftime("%Y-%m-%d")
        fid = core.create_facture_vente(self.conn, numero, date_str, client_code,
                                         tva_compte=self.tva_compte_var.get().strip() or None)
        self.current_facture_id = fid
        self.refresh_factures_list()

    def _on_facture_selected(self, event=None):
        idx = self.facture_combo.current()
        if 0 <= idx < len(self._factures_cache):
            self.current_facture_id = self._factures_cache[idx]["id"]
        self.load_facture()

    def load_facture(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.entete_text.delete("1.0", "end")
        self.pied_text.delete("1.0", "end")
        if not self.current_facture_id:
            self.statut_var.set("Aucune facture — créez-en une nouvelle.")
            self.totals_var.set("")
            self.corriger_btn.configure(state="disabled")
            return
        f = core.get_facture_vente(self.conn, self.current_facture_id)
        if not f:
            self.current_facture_id = None
            self.statut_var.set("")
            self.corriger_btn.configure(state="disabled")
            return
        self.numero_var.set(f["numero"])
        self.date_var.set(core.to_display_date(f["date_facture"]))
        client = core.get_client(self.conn, f["client_code"])
        self.client_var.set(f"{f['client_code']} — {client['raison_sociale']}" if client else f["client_code"])
        self.tva_var.set(str(f["tva_taux"]))
        self.tva_compte_var.set(f.get("tva_compte") or core.COMPTE_TVA_VENTES)
        self.entete_text.insert("1.0", f["entete"] or "")
        self.pied_text.insert("1.0", f["pied_page"] or "")
        statut_label = "VALIDÉE (écritures envoyées en Saisie)" if f["statut"] == "validee" else "Brouillon"
        self.statut_var.set(f"Statut : {statut_label}")
        self.corriger_btn.configure(state="normal" if f["statut"] == "validee" else "disabled")

        editable = f["statut"] != "validee"
        state = "normal" if editable else "disabled"
        for w in (self.entete_text, self.pied_text):
            w.configure(state="normal")
        if not editable:
            self.entete_text.configure(state="disabled")
            self.pied_text.configure(state="disabled")

        lignes = core.list_lignes_facture_vente(self.conn, self.current_facture_id)
        for l in lignes:
            impact = {"marchandise": "Stock marchandises (31)", "produit_fini": "Stock produits finis (36)"}.get(
                l["type_stock"], "Aucun (service)")
            self.tree.insert("", "end", values=(
                l["id"], l["compte_vente"], l["libelle"], impact,
                f"{l['quantite']:g}", f"{fmt_cfa(l['prix_unitaire'])}", l.get("analytic_code") or "",
                f"{fmt_cfa(l['montant_ht'])}",
            ))
        totals = core.compute_facture_totals(self.conn, self.current_facture_id)
        self.totals_var.set(
            f"TOTAL HT : {fmt_cfa(totals['total_ht'])}    TVA ({totals['tva_taux']:g}%) : "
            f"{fmt_cfa(totals['tva_montant'])}    TOTAL TTC : {fmt_cfa(totals['total_ttc'])}"
        )

    def _ensure_facture(self):
        if not self.current_facture_id:
            messagebox.showinfo("Info", "Créez d'abord une nouvelle facture.")
            return None
        f = core.get_facture_vente(self.conn, self.current_facture_id)
        if f and f["statut"] == "validee":
            messagebox.showwarning("Facture validée", "Cette facture est déjà validée et ne peut plus être modifiée.")
            return None
        return f

    def add_ligne(self):
        f = self._ensure_facture()
        if not f:
            return
        compte = self._extract_code(self.ligne_compte_var.get())
        if not compte:
            messagebox.showwarning("Champ manquant", "Choisissez un compte de vente (classe 70).")
            return
        if not core.account_exists(self.conn, compte) or core.account_racine(compte) != "7":
            messagebox.showerror("Compte invalide", "Le compte de vente doit être un compte existant de la classe 7.")
            return
        libelle = self.ligne_libelle_var.get().strip()
        if not libelle:
            messagebox.showwarning("Champ manquant", "Le libellé de la ligne est obligatoire.")
            return
        try:
            qte = float(self.ligne_qte_var.get() or 0)
            prix = float(self.ligne_prix_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Quantité et Prix unitaire doivent être des nombres.")
            return
        analytic_code = self._extract_code(self.ligne_analytic_var.get()) or None
        core.add_ligne_facture_vente(self.conn, self.current_facture_id, compte, libelle, qte, prix,
                                      analytic_code=analytic_code)
        self.ligne_libelle_var.set("")
        self.ligne_qte_var.set("1")
        self.ligne_prix_var.set("")
        self.ligne_analytic_var.set("")
        self.load_facture()

    def delete_ligne(self):
        f = self._ensure_facture()
        if not f:
            return
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne.")
            return
        ligne_id = int(self.tree.item(sel[0], "values")[0])
        core.delete_ligne_facture_vente(self.conn, ligne_id)
        self.load_facture()

    def save_facture(self):
        f = self._ensure_facture()
        if not f:
            return
        client_code = self._extract_code(self.client_var.get())
        if not client_code or not core.client_exists(self.conn, client_code):
            messagebox.showerror("Client invalide", "Choisissez un client existant.")
            return
        date_str = core.to_iso_date(self.date_var.get().strip())
        try:
            tva = float(self.tva_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Le taux de TVA doit être un nombre.")
            return
        core.update_facture_vente(
            self.conn, self.current_facture_id,
            numero=self.numero_var.get().strip(), date_facture=date_str, client_code=client_code,
            entete=self.entete_text.get("1.0", "end").strip(),
            pied_page=self.pied_text.get("1.0", "end").strip(),
            tva_taux=tva, tva_compte=self.tva_compte_var.get().strip() or core.COMPTE_TVA_VENTES,
        )
        core.set_setting(self.conn, "tva_taux_defaut", tva)
        core.set_text_setting(self.conn, "tva_compte_defaut", self.tva_compte_var.get().strip() or core.COMPTE_TVA_VENTES)
        messagebox.showinfo("Enregistré", "Facture enregistrée (brouillon).")
        self.refresh_factures_list()

    def valider(self):
        f = self._ensure_facture()
        if not f:
            return
        self.save_facture()
        if messagebox.askyesno(
            "Confirmer la validation",
            "Valider cette facture ? Les écritures comptables seront envoyées dans le menu SAISIE "
            "(débit client, crédit ventes, TVA, et sortie de stock automatique pour les lignes "
            "marchandises/produits finis). Cette action est définitive."
        ):
            try:
                warnings = core.valider_facture_vente(self.conn, self.current_facture_id)
            except ValueError as exc:
                messagebox.showerror("Erreur", str(exc))
                return
            msg = "Facture validée et écritures envoyées en Saisie."
            if warnings:
                msg += "\n\nAvertissements :\n" + "\n".join(warnings)
            messagebox.showinfo("Validation terminée", msg)
            self.refresh_factures_list()

    def delete_facture(self):
        if not self.current_facture_id:
            return
        if messagebox.askyesno("Confirmer", "Supprimer cette facture ?"):
            try:
                core.delete_facture_vente(self.conn, self.current_facture_id)
            except ValueError as exc:
                messagebox.showerror("Erreur", str(exc))
                return
            self.current_facture_id = None
            self.refresh_factures_list()

    def corriger_facture(self):
        """Repasse une facture déjà validée en brouillon modifiable, en
        supprimant les écritures comptables qu'elle avait générées — pour
        corriger une erreur sur les chiffres, puis revalider ensuite."""
        if not self.current_facture_id:
            return
        if not messagebox.askyesno(
            "Corriger cette facture",
            "Cette facture est déjà validée : ses écritures comptables (débit client, "
            "crédit ventes, TVA, sortie de stock) vont être RETIRÉES de la Saisie et la "
            "facture repassera en brouillon modifiable.\n\n"
            "Vous pourrez alors corriger les chiffres puis la revalider.\n\n"
            "Continuer ?"
        ):
            return
        try:
            core.devalider_facture_vente(self.conn, self.current_facture_id)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc))
            return
        messagebox.showinfo("Facture repassée en brouillon",
                             "La facture est de nouveau modifiable. Corrigez les chiffres puis "
                             "cliquez sur « Valider et envoyer en Saisie ».")
        self.refresh_factures_list()

    def _refresh_tva_presets(self):
        presets = core.list_taux_tva(self.conn)
        self.tva_preset_combo["values"] = [f"{p['label']} ({p['montant']:g}%)" for p in presets]
        self._tva_presets = presets

    def _on_tva_preset_selected(self, event=None):
        idx = self.tva_preset_combo.current()
        if idx is not None and 0 <= idx < len(getattr(self, "_tva_presets", [])):
            preset = self._tva_presets[idx]
            self.tva_var.set(str(preset["montant"]))
            if preset.get("compte"):
                self.tva_compte_var.set(preset["compte"])

    def imprimer_facture(self):
        """Génère la facture en HTML imprimable (bouton « Imprimer » intégré,
        ou Ctrl+P depuis le navigateur qui s'ouvre) et l'ouvre directement."""
        if not self.current_facture_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord une facture.")
            return
        import tempfile
        import webbrowser
        path = os.path.join(tempfile.gettempdir(), f"facture_vente_{self.current_facture_id}.html")
        try:
            core.export_facture_vente_html(self.conn, self.current_facture_id, path)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc))
            return
        webbrowser.open(f"file://{path}")

    def refresh(self):
        self._refresh_client_values()
        self._refresh_ligne_compte_values()
        self._refresh_ligne_analytic_values()
        self._refresh_tva_presets()
        self.refresh_factures_list()


class RecouvrementTab(ttk.Frame):
    """Journal des factures clients : suivi des retards de paiement
    (recouvrement) + balance âgée des créances (onglet séparé)."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn

        notebook = ttk.Notebook(self)
        notebook.pack(fill="both", expand=True)
        tab_factures = ttk.Frame(notebook)
        tab_agee = ttk.Frame(notebook)
        notebook.add(tab_factures, text="Factures")
        notebook.add(tab_agee, text="Balance âgée")

        self._build_factures(tab_factures)
        self._build_balance_agee(tab_agee)
        self.selected_id = None
        self.refresh()

    def _build_factures(self, parent):
        ttk.Label(parent, text="RECOUVREMENT — SUIVI DES RETARDS DE PAIEMENT CLIENTS",
                  font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Label(parent, text=(
            "Enregistrez ici chaque facture émise à un client. L'échéance de paiement est calculée "
            "automatiquement à partir du délai par défaut du client (modifiable dans l'onglet "
            "Clients), à la date de facture. Renseignez ensuite la date réelle de paiement au fur "
            "et à mesure des encaissements — les retards sont signalés automatiquement."
        ), foreground="#595959", wraplength=1050).pack(anchor="w", padx=16, pady=(0, 8))

        form = ttk.LabelFrame(parent, text="Nouvelle facture")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Client :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.client_var = tk.StringVar()
        self.client_combo = ttk.Combobox(form, textvariable=self.client_var, width=28)
        self.client_combo.grid(row=0, column=1, padx=4)
        self.client_combo.bind("<KeyRelease>", self._on_client_keyrelease)
        self._refresh_client_values()

        ttk.Label(form, text="N° Pièce :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.piece_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.piece_var, width=14).grid(row=0, column=3, padx=4)

        ttk.Label(form, text="Libellé :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.libelle_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.libelle_var, width=26).grid(row=0, column=5, padx=4)

        ttk.Label(form, text="Montant :").grid(row=1, column=0, sticky="w", padx=4, pady=4)
        self.montant_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.montant_var, width=16).grid(row=1, column=1, padx=4)

        ttk.Label(form, text="Date facture (JJ/MM/AAAA) :").grid(row=1, column=2, sticky="w", padx=(12, 4))
        self.date_facture_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(form, textvariable=self.date_facture_var, width=14).grid(row=1, column=3, padx=4)

        ttk.Button(form, text="Créer la facture (échéance auto)", command=self.add_facture).grid(
            row=1, column=4, columnspan=2, sticky="w", padx=12, pady=4)

        update_frame = ttk.LabelFrame(parent, text="Enregistrer le règlement de la facture sélectionnée")
        update_frame.pack(fill="x", padx=16, pady=(8, 4))
        ttk.Label(update_frame, text="Date paiement réel (JJ/MM/AAAA) :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.paiement_reel_var = tk.StringVar()
        ttk.Entry(update_frame, textvariable=self.paiement_reel_var, width=14).grid(row=0, column=1, padx=4)
        ttk.Label(update_frame, text="Compte banque/caisse :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.compte_reglement_var = tk.StringVar()
        self.compte_reglement_combo = ttk.Combobox(update_frame, textvariable=self.compte_reglement_var, width=26)
        self.compte_reglement_combo.grid(row=0, column=3, padx=4)
        self.compte_reglement_combo.bind("<KeyRelease>", self._on_compte_reglement_keyrelease)
        self.compte_reglement_combo.bind("<Button-1>", lambda e: e.widget.event_generate("<Down>"))
        self._refresh_compte_reglement_values()
        ttk.Button(update_frame, text="Enregistrer le paiement (comptabilise)", command=self.save_paiement).grid(
            row=0, column=4, padx=8)
        ttk.Button(update_frame, text="Supprimer la facture sélectionnée", command=self.delete_facture).grid(
            row=1, column=4, padx=8, pady=(4, 0))
        ttk.Label(update_frame, text=(
            "Comptabilise automatiquement le règlement (Débit banque/caisse choisi, Crédit compte client "
            "411000) — une seule fois par facture."
        ), foreground="#595959").grid(row=1, column=0, columnspan=4, sticky="w", padx=4, pady=(4, 4))

        cols = ("id", "client", "piece", "libelle", "montant", "date_facture",
                "echeance_paiement", "statut_paiement")
        self.tree = ttk.Treeview(parent, columns=cols, show="headings")
        headers = ["ID", "Client", "Pièce", "Libellé", "Montant", "Date facture",
                   "Échéance paiement", "Statut paiement"]
        widths = [40, 180, 90, 200, 110, 110, 130, 160]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.tag_configure("depasse", foreground="#B00020")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

    def _build_balance_agee(self, parent):
        ttk.Label(parent, text="BALANCE ÂGÉE DES CRÉANCES CLIENTS", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(parent, text=(
            "Répartit le montant des factures NON PAYÉES de chaque client par ancienneté (jours écoulés "
            "depuis la date de facture). Choisissez les seuils des tranches ci-dessous, puis double-cliquez "
            "sur un client pour voir le détail de ses factures impayées."
        ), foreground="#595959", wraplength=1050).pack(anchor="w", padx=16, pady=(0, 8))

        seuils_bar = ttk.Frame(parent)
        seuils_bar.pack(fill="x", padx=16, pady=4)
        ttk.Label(seuils_bar, text="Seuils des tranches (jours) :").pack(side="left")
        self.seuil1_var = tk.StringVar(value="30")
        self.seuil2_var = tk.StringVar(value="60")
        self.seuil3_var = tk.StringVar(value="90")
        ttk.Entry(seuils_bar, textvariable=self.seuil1_var, width=6).pack(side="left", padx=4)
        ttk.Label(seuils_bar, text="/").pack(side="left")
        ttk.Entry(seuils_bar, textvariable=self.seuil2_var, width=6).pack(side="left", padx=4)
        ttk.Label(seuils_bar, text="/").pack(side="left")
        ttk.Entry(seuils_bar, textvariable=self.seuil3_var, width=6).pack(side="left", padx=4)
        ttk.Button(seuils_bar, text="Appliquer", command=self.refresh_balance_agee).pack(side="left", padx=12)
        ttk.Label(seuils_bar, text="Préréglages :").pack(side="left", padx=(20, 4))
        ttk.Button(seuils_bar, text="30/60/90", command=lambda: self._preset_seuils(30, 60, 90)).pack(side="left", padx=2)
        ttk.Button(seuils_bar, text="15/30/60", command=lambda: self._preset_seuils(15, 30, 60)).pack(side="left", padx=2)
        ttk.Button(seuils_bar, text="30/60/120", command=lambda: self._preset_seuils(30, 60, 120)).pack(side="left", padx=2)

        cols = ("client", "t0", "t1", "t2", "t3", "total")
        self.tree_agee = ttk.Treeview(parent, columns=cols, show="headings", height=18)
        self.tree_agee.heading("client", text="Client")
        self.tree_agee.column("client", width=260, anchor="w")
        self.tree_agee.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree_agee.tag_configure("total", background="#1F4E78", foreground="white", font=("Segoe UI", 9, "bold"))
        self.tree_agee.bind("<Double-1>", self._on_double_click_agee)
        self._by_iid_agee = {}

        ttk.Label(parent, text="Double-cliquez sur un client pour voir le détail de ses factures impayées.",
                  foreground="#595959").pack(anchor="w", padx=16, pady=(0, 12))

    def _preset_seuils(self, s1, s2, s3):
        self.seuil1_var.set(str(s1))
        self.seuil2_var.set(str(s2))
        self.seuil3_var.set(str(s3))
        self.refresh_balance_agee()

    def _get_seuils(self):
        try:
            s1 = int(self.seuil1_var.get())
            s2 = int(self.seuil2_var.get())
            s3 = int(self.seuil3_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Les seuils doivent être des nombres entiers (jours).")
            return None
        if not (s1 < s2 < s3):
            messagebox.showerror("Erreur", "Les seuils doivent être croissants (ex. 30 < 60 < 90).")
            return None
        return (s1, s2, s3)

    def _on_double_click_agee(self, event=None):
        sel = self.tree_agee.selection()
        if not sel:
            return
        client = self._by_iid_agee.get(sel[0])
        if client:
            BalanceAgeeDetailDialog(self, client)

    def refresh_balance_agee(self):
        seuils = self._get_seuils()
        if not seuils:
            return
        s1, s2, s3 = seuils
        headers = ["Client", f"0-{s1} j", f"{s1+1}-{s2} j", f"{s2+1}-{s3} j", f">{s3} j", "Total"]
        widths = [260, 110, 110, 110, 110, 120]
        cols = ("client", "t0", "t1", "t2", "t3", "total")
        for c, h, w in zip(cols, headers, widths):
            self.tree_agee.heading(c, text=h)
            self.tree_agee.column(c, width=w, anchor="w" if c == "client" else "e")

        for row in self.tree_agee.get_children():
            self.tree_agee.delete(row)
        self._by_iid_agee = {}
        self._clients_agee = core.compute_balance_agee(self.conn, seuils=seuils)
        totaux = [0.0, 0.0, 0.0, 0.0]
        for c in self._clients_agee:
            iid = self.tree_agee.insert("", "end", values=(
                c["raison_sociale"], fmt_cfa(c["tranches"][0]), fmt_cfa(c["tranches"][1]),
                fmt_cfa(c["tranches"][2]), fmt_cfa(c["tranches"][3]), fmt_cfa(c["total"])))
            self._by_iid_agee[iid] = c
            for i in range(4):
                totaux[i] += c["tranches"][i]
        self.tree_agee.insert("", "end", tags=("total",), values=(
            "TOTAL", fmt_cfa(totaux[0]), fmt_cfa(totaux[1]), fmt_cfa(totaux[2]), fmt_cfa(totaux[3]),
            fmt_cfa(sum(totaux))))

    def _refresh_client_values(self):
        items = core.list_clients(self.conn)
        self.client_combo["values"] = [f"{c['code']} — {c['raison_sociale']}" for c in items]

    def _on_client_keyrelease(self, event=None):
        query = self._extract_code(self.client_var.get())
        if query:
            items = core.list_clients(self.conn, query)
            self.client_combo["values"] = [f"{c['code']} — {c['raison_sociale']}" for c in items]

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        values = self.tree.item(sel[0], "values")
        self.selected_id = int(values[0])

    def add_facture(self):
        code = self._extract_code(self.client_var.get())
        if not code:
            messagebox.showwarning("Champ manquant", "Choisissez un client.")
            return
        if not core.client_exists(self.conn, code):
            messagebox.showerror("Client invalide", f"Le client « {code} » n'existe pas. "
                                                      f"Créez-le d'abord dans l'onglet Clients.")
            return
        if not self.montant_var.get().strip():
            messagebox.showwarning("Champ manquant", "Le montant est obligatoire.")
            return
        try:
            montant = float(self.montant_var.get())
        except ValueError:
            messagebox.showerror("Erreur", "Le montant doit être un nombre.")
            return
        if montant <= 0:
            messagebox.showerror("Erreur", "Le montant doit être strictement positif.")
            return
        date_facture = core.to_iso_date(self.date_facture_var.get().strip())
        if not date_facture:
            messagebox.showwarning("Champ manquant", "La date de facture est obligatoire.")
            return
        try:
            core.add_facture(self.conn, code, self.piece_var.get().strip(), self.libelle_var.get().strip(),
                              montant, date_facture)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc))
            return
        self.piece_var.set("")
        self.libelle_var.set("")
        self.montant_var.set("")
        self.refresh()

    def _refresh_compte_reglement_values(self):
        items = [a for a in core.search_accounts(self.conn, "", limit=200) if a["classe"] == "5"]
        self.compte_reglement_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _on_compte_reglement_keyrelease(self, event=None):
        query = self._extract_code(self.compte_reglement_var.get())
        items = [a for a in core.search_accounts(self.conn, query, limit=50) if a["classe"] == "5"]
        self.compte_reglement_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def save_paiement(self):
        if self.selected_id is None:
            messagebox.showinfo("Info", "Sélectionnez d'abord une facture dans le tableau.")
            return
        d = core.to_iso_date(self.paiement_reel_var.get().strip())
        if not d:
            messagebox.showwarning("Champ manquant", "Saisissez la date de paiement réel.")
            return
        compte = self._extract_code(self.compte_reglement_var.get())
        if not compte:
            messagebox.showwarning("Champ manquant",
                                    "Choisissez le compte banque ou caisse ayant reçu le règlement.")
            return
        try:
            core.enregistrer_paiement_facture(self.conn, self.selected_id, d, compte)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc))
            return
        self.paiement_reel_var.set("")
        self.compte_reglement_var.set("")
        self.refresh()
        messagebox.showinfo("Paiement enregistré",
                             "Le règlement a été comptabilisé (Débit banque/caisse, Crédit client).")

    def delete_facture(self):
        if self.selected_id is None:
            messagebox.showinfo("Info", "Sélectionnez d'abord une facture.")
            return
        if messagebox.askyesno("Confirmer", "Supprimer cette facture ?"):
            core.delete_facture(self.conn, self.selected_id)
            self.selected_id = None
            self.refresh()

    def refresh(self):
        self._refresh_client_values()
        self._refresh_compte_reglement_values()
        for row in self.tree.get_children():
            self.tree.delete(row)
        for f in core.list_factures(self.conn):
            tags = ("depasse",) if f["depassement"] else ()
            self.tree.insert("", "end", tags=tags, values=(
                f["id"], f["raison_sociale"], f["piece"] or "", f["libelle"] or "",
                f"{fmt_cfa(f['montant'])}", core.to_display_date(f["date_facture"]),
                core.to_display_date(f["date_echeance_paiement"]), f["statut_paiement"],
            ))
        self.refresh_balance_agee()


class BalanceAgeeDetailDialog(tk.Toplevel):
    """Détail des factures impayées d'un client (double-clic depuis la
    Balance âgée) — liste chaque facture avec son ancienneté en jours."""

    def __init__(self, parent, client):
        super().__init__(parent)
        self.title(f"Détail — {client['raison_sociale']}")
        self.geometry("700x420")
        self.transient(parent)
        self.grab_set()

        ttk.Label(self, text=f"Factures impayées — {client['raison_sociale']}",
                  font=("Segoe UI", 12, "bold")).pack(anchor="w", padx=12, pady=(12, 4))
        ttk.Label(self, text=f"Total dû : {fmt_cfa(client['total'])}",
                  font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=12, pady=(0, 8))

        cols = ("piece", "libelle", "date", "age", "montant")
        tree = ttk.Treeview(self, columns=cols, show="headings", height=14)
        headers = ["Pièce", "Libellé", "Date facture", "Ancienneté (j)", "Montant"]
        widths = [90, 240, 100, 110, 120]
        for c, h, w in zip(cols, headers, widths):
            tree.heading(c, text=h)
            tree.column(c, width=w, anchor="w" if c in ("piece", "libelle") else "e")
        tree.pack(fill="both", padx=12, pady=8)
        for f in sorted(client["factures"], key=lambda x: -x["age_jours"]):
            tree.insert("", "end", values=(
                f["piece"], f["libelle"], core.to_display_date(f["date_facture"]), f["age_jours"],
                fmt_cfa(f["montant"])))

        ttk.Button(self, text="Fermer", command=self.destroy).pack(pady=(0, 12))


class FacturesFrsTab(ttk.Frame):
    """Factures fournisseurs (achats) : présente directement une facture (entête,
    lignes d'achat liées à un compte 6x, retenue fiscale à la source paramétrable,
    pied de page), et sa validation envoie les écritures comptables en Saisie —
    avec entrée de stock automatique pour les lignes liées aux marchandises (31)
    ou matières premières (32)."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.current_facture_id = None

        top = ttk.Frame(self)
        top.pack(fill="x", padx=12, pady=8)
        ttk.Label(top, text="Facture n° :").pack(side="left")
        self.facture_combo = ttk.Combobox(top, width=40, state="readonly")
        self.facture_combo.pack(side="left", padx=4)
        self.facture_combo.bind("<<ComboboxSelected>>", self._on_facture_selected)
        ttk.Button(top, text="Nouvelle facture", command=self.new_facture).pack(side="left", padx=8)
        ttk.Button(top, text="Supprimer cette facture", command=self.delete_facture).pack(side="left", padx=2)
        self.corriger_btn = ttk.Button(top, text="Corriger cette facture (erreur sur les chiffres)",
                                        command=self.corriger_facture)
        self.corriger_btn.pack(side="left", padx=2)
        self.statut_var = tk.StringVar()
        ttk.Label(top, textvariable=self.statut_var, font=("Segoe UI", 10, "bold")).pack(side="left", padx=16)

        ttk.Label(self, text="En-tête de la facture (modifiable) :").pack(anchor="w", padx=12)
        self.entete_text = tk.Text(self, height=3, font=("Segoe UI", 10))
        self.entete_text.pack(fill="x", padx=12, pady=(0, 8))

        info = ttk.Frame(self)
        info.pack(fill="x", padx=12, pady=4)
        ttk.Label(info, text="N° Facture :").grid(row=0, column=0, sticky="w", padx=4)
        self.numero_var = tk.StringVar()
        ttk.Entry(info, textvariable=self.numero_var, width=16).grid(row=0, column=1, padx=4)
        ttk.Label(info, text="Date (JJ/MM/AAAA) :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.date_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(info, textvariable=self.date_var, width=14).grid(row=0, column=3, padx=4)
        ttk.Label(info, text="Fournisseur (compte 40) :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.fournisseur_var = tk.StringVar()
        self.fournisseur_combo = ttk.Combobox(info, textvariable=self.fournisseur_var, width=26)
        self.fournisseur_combo.grid(row=0, column=5, padx=4)
        self.fournisseur_combo.bind("<KeyRelease>", self._on_fournisseur_keyrelease)
        self.fournisseur_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_fournisseur_values()

        ttk.Label(info, text="Retenue % :").grid(row=1, column=0, sticky="w", padx=4, pady=(6, 0))
        self.retenue_taux_var = tk.StringVar(
            value=str(core.get_setting(conn, "retenue_taux_defaut", core.RETENUE_TAUX_DEFAUT)))
        ttk.Entry(info, textvariable=self.retenue_taux_var, width=6).grid(row=1, column=1, sticky="w", padx=4, pady=(6, 0))
        ttk.Label(info, text="Compte retenue (classe 44) :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(6, 0))
        self.retenue_compte_var = tk.StringVar(
            value=core.get_text_setting(conn, "retenue_compte_defaut", core.COMPTE_RETENUE_DEFAUT))
        self.retenue_compte_combo = ttk.Combobox(info, textvariable=self.retenue_compte_var, width=30)
        self.retenue_compte_combo.grid(row=1, column=3, columnspan=2, sticky="w", padx=4, pady=(6, 0))
        self.retenue_compte_combo.bind("<KeyRelease>", self._on_retenue_compte_keyrelease)
        self.retenue_compte_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_retenue_compte_values()
        ttk.Label(info, text="Préréglage (ADMIN) :").grid(row=2, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        self.retenue_preset_var = tk.StringVar()
        self.retenue_preset_combo = ttk.Combobox(info, textvariable=self.retenue_preset_var, width=22, state="readonly")
        self.retenue_preset_combo.grid(row=2, column=3, columnspan=2, sticky="w", padx=4, pady=(4, 0))
        self.retenue_preset_combo.bind("<<ComboboxSelected>>", self._on_retenue_preset_selected)
        self.retenue_preset_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_retenue_presets()

        form = ttk.LabelFrame(self, text="Ajouter une ligne (produit/service acheté — compte 6x)")
        form.pack(fill="x", padx=12, pady=6)
        ttk.Label(form, text="Compte d'achat :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.ligne_compte_var = tk.StringVar()
        self.ligne_compte_combo = ttk.Combobox(form, textvariable=self.ligne_compte_var, width=34)
        self.ligne_compte_combo.grid(row=0, column=1, padx=4)
        self.ligne_compte_combo.bind("<KeyRelease>", self._on_ligne_compte_keyrelease)
        self.ligne_compte_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_ligne_compte_values()
        ttk.Label(form, text="Libellé :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.ligne_libelle_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_libelle_var, width=26).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Quantité :").grid(row=1, column=0, sticky="w", padx=4, pady=4)
        self.ligne_qte_var = tk.StringVar(value="1")
        ttk.Entry(form, textvariable=self.ligne_qte_var, width=10).grid(row=1, column=1, sticky="w", padx=4)
        ttk.Label(form, text="Prix unitaire :").grid(row=1, column=2, sticky="w", padx=(12, 4))
        self.ligne_prix_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_prix_var, width=14).grid(row=1, column=3, sticky="w", padx=4)
        ttk.Label(form, text="Code analytique :").grid(row=1, column=4, sticky="w", padx=(12, 4))
        self.ligne_analytic_var = tk.StringVar()
        self.ligne_analytic_combo = ttk.Combobox(form, textvariable=self.ligne_analytic_var, width=20)
        self.ligne_analytic_combo.grid(row=1, column=5, padx=4, sticky="w")
        self.ligne_analytic_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_ligne_analytic_values()
        ttk.Button(form, text="Ajouter la ligne", command=self.add_ligne).grid(row=1, column=6, padx=12)

        cols = ("id", "compte", "libelle", "type_stock", "qte", "prix", "analytique", "montant")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=6)
        headers = ["ID", "Compte", "Libellé", "Impact stock", "Qté", "Prix unit.", "Montant HT"]
        widths = [40, 90, 220, 110, 70, 100, 110]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", padx=12, pady=6)
        ttk.Button(self, text="Supprimer la ligne sélectionnée", command=self.delete_ligne).pack(anchor="w", padx=12)

        self.totals_var = tk.StringVar()
        ttk.Label(self, textvariable=self.totals_var, font=("Segoe UI", 11, "bold")).pack(anchor="w", padx=12, pady=(8, 0))

        ttk.Label(self, text="Pied de page de la facture (modifiable) :").pack(anchor="w", padx=12, pady=(8, 0))
        self.pied_text = tk.Text(self, height=3, font=("Segoe UI", 10))
        self.pied_text.pack(fill="x", padx=12, pady=(0, 8))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=12, pady=8)
        ttk.Button(btns, text="Enregistrer BON DE COMMANDE", command=self.save_facture).pack(side="left", padx=2)
        ttk.Button(btns, text="Aperçu avant impression (bon de commande)", command=self.imprimer_facture).pack(side="left", padx=2)
        ttk.Button(btns, text="Valider et envoyer en Saisie", command=self.valider).pack(side="left", padx=2)

        self.refresh_factures_list()

    # -- Fournisseur --
    def _refresh_fournisseur_values(self):
        items = core.list_fournisseurs(self.conn)
        self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _on_fournisseur_keyrelease(self, event=None):
        query = self._extract_code(self.fournisseur_var.get())
        if query:
            items = core.list_fournisseurs(self.conn, query)
            self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    # -- Compte de retenue (classe 44) --
    def _refresh_retenue_compte_values(self):
        items = [a for a in core.search_accounts(self.conn, "44", limit=100) if core.account_racine(a["code"]) == "44"]
        self.retenue_compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _on_retenue_compte_keyrelease(self, event=None):
        query = self._extract_code(self.retenue_compte_var.get())
        if query:
            items = [a for a in core.search_accounts(self.conn, query, limit=50)
                     if core.account_racine(a["code"]) == "44"]
            self.retenue_compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    # -- Compte d'achat (classe 6) --
    def _refresh_ligne_compte_values(self):
        items = [a for a in core.search_accounts(self.conn, "6", limit=100) if a["classe"] == "6"]
        self.ligne_compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _on_ligne_compte_keyrelease(self, event=None):
        query = self._extract_code(self.ligne_compte_var.get())
        if query:
            items = [a for a in core.search_accounts(self.conn, query, limit=50) if a["classe"] == "6"]
            self.ligne_compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _refresh_ligne_analytic_values(self):
        codes = core.list_analytic_codes(self.conn)
        self.ligne_analytic_combo["values"] = [f"{c['code']} — {c['label']}" for c in codes]

    @staticmethod
    def _open_dropdown(event=None):
        if event is not None:
            event.widget.event_generate("<Down>")

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    # -- Gestion des factures --
    def refresh_factures_list(self):
        factures = core.list_factures_achat(self.conn)
        values = [f"{f['numero']} — {f['raison_sociale']} — {f['statut']}" for f in factures]
        self.facture_combo["values"] = values
        self._factures_cache = factures
        if self.current_facture_id is None and factures:
            self.current_facture_id = factures[0]["id"]
            self.facture_combo.current(0)
        self.load_facture()

    def new_facture(self):
        numero = simpledialog.askstring("Nouvelle facture", "N° de facture :", parent=self)
        if not numero:
            return
        fournisseur_code = self._extract_code(self.fournisseur_var.get())
        if not fournisseur_code or not core.fournisseur_exists(self.conn, fournisseur_code):
            messagebox.showinfo("Fournisseur requis", "Choisissez d'abord un fournisseur existant dans le champ Fournisseur.")
            return
        date_str = core.to_iso_date(self.date_var.get().strip()) or date.today().strftime("%Y-%m-%d")
        try:
            retenue_taux = float(self.retenue_taux_var.get() or 0)
        except ValueError:
            retenue_taux = 0
        retenue_compte = self._extract_code(self.retenue_compte_var.get()) or core.COMPTE_RETENUE_DEFAUT
        fid = core.create_facture_achat(self.conn, numero, date_str, fournisseur_code,
                                         retenue_taux=retenue_taux, retenue_compte=retenue_compte)
        self.current_facture_id = fid
        self.refresh_factures_list()

    def _on_facture_selected(self, event=None):
        idx = self.facture_combo.current()
        if 0 <= idx < len(self._factures_cache):
            self.current_facture_id = self._factures_cache[idx]["id"]
        self.load_facture()

    def load_facture(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.entete_text.delete("1.0", "end")
        self.pied_text.delete("1.0", "end")
        if not self.current_facture_id:
            self.statut_var.set("Aucune facture — créez-en une nouvelle.")
            self.totals_var.set("")
            self.corriger_btn.configure(state="disabled")
            return
        f = core.get_facture_achat(self.conn, self.current_facture_id)
        if not f:
            self.current_facture_id = None
            self.statut_var.set("")
            self.corriger_btn.configure(state="disabled")
            return
        self.numero_var.set(f["numero"])
        self.date_var.set(core.to_display_date(f["date_facture"]))
        fournisseur = core.get_fournisseur(self.conn, f["fournisseur_code"])
        self.fournisseur_var.set(
            f"{f['fournisseur_code']} — {fournisseur['raison_sociale']}" if fournisseur else f["fournisseur_code"])
        self.retenue_taux_var.set(str(f["retenue_taux"]))
        self.retenue_compte_var.set(f["retenue_compte"])
        self.entete_text.insert("1.0", f["entete"] or "")
        self.pied_text.insert("1.0", f["pied_page"] or "")
        statut_label = "VALIDÉE (écritures envoyées en Saisie)" if f["statut"] == "validee" else "Brouillon"
        self.statut_var.set(f"Statut : {statut_label}")
        self.corriger_btn.configure(state="normal" if f["statut"] == "validee" else "disabled")

        lignes = core.list_lignes_facture_achat(self.conn, self.current_facture_id)
        for l in lignes:
            impact = {"marchandise": "Stock marchandises (31)", "matiere_premiere": "Stock matières (32)"}.get(
                l["type_stock"], "Aucun (service)")
            self.tree.insert("", "end", values=(
                l["id"], l["compte_achat"], l["libelle"], impact,
                f"{l['quantite']:g}", f"{fmt_cfa(l['prix_unitaire'])}", l.get("analytic_code") or "",
                f"{fmt_cfa(l['montant_ht'])}",
            ))
        totals = core.compute_facture_achat_totals(self.conn, self.current_facture_id)
        self.totals_var.set(
            f"TOTAL HT : {fmt_cfa(totals['total_ht'])}    Retenue ({totals['retenue_taux']:g}%) : "
            f"{fmt_cfa(totals['retenue_montant'])}    NET À PAYER : {fmt_cfa(totals['net_a_payer'])}"
        )

    def _ensure_facture(self):
        if not self.current_facture_id:
            messagebox.showinfo("Info", "Créez d'abord une nouvelle facture.")
            return None
        f = core.get_facture_achat(self.conn, self.current_facture_id)
        if f and f["statut"] == "validee":
            messagebox.showwarning("Facture validée", "Cette facture est déjà validée et ne peut plus être modifiée.")
            return None
        return f

    def add_ligne(self):
        f = self._ensure_facture()
        if not f:
            return
        compte = self._extract_code(self.ligne_compte_var.get())
        if not compte:
            messagebox.showwarning("Champ manquant", "Choisissez un compte d'achat (classe 6).")
            return
        if not core.account_exists(self.conn, compte) or core.account_racine(compte) != "6":
            messagebox.showerror("Compte invalide", "Le compte d'achat doit être un compte existant de la classe 6.")
            return
        libelle = self.ligne_libelle_var.get().strip()
        if not libelle:
            messagebox.showwarning("Champ manquant", "Le libellé de la ligne est obligatoire.")
            return
        try:
            qte = float(self.ligne_qte_var.get() or 0)
            prix = float(self.ligne_prix_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Quantité et Prix unitaire doivent être des nombres.")
            return
        analytic_code = self._extract_code(self.ligne_analytic_var.get()) or None
        core.add_ligne_facture_achat(self.conn, self.current_facture_id, compte, libelle, qte, prix,
                                      analytic_code=analytic_code)
        self.ligne_libelle_var.set("")
        self.ligne_qte_var.set("1")
        self.ligne_prix_var.set("")
        self.ligne_analytic_var.set("")
        self.load_facture()

    def delete_ligne(self):
        f = self._ensure_facture()
        if not f:
            return
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne.")
            return
        ligne_id = int(self.tree.item(sel[0], "values")[0])
        core.delete_ligne_facture_achat(self.conn, ligne_id)
        self.load_facture()

    def save_facture(self):
        f = self._ensure_facture()
        if not f:
            return
        fournisseur_code = self._extract_code(self.fournisseur_var.get())
        if not fournisseur_code or not core.fournisseur_exists(self.conn, fournisseur_code):
            messagebox.showerror("Fournisseur invalide", "Choisissez un fournisseur existant.")
            return
        date_str = core.to_iso_date(self.date_var.get().strip())
        try:
            retenue_taux = float(self.retenue_taux_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Le taux de retenue doit être un nombre.")
            return
        retenue_compte = self._extract_code(self.retenue_compte_var.get()) or core.COMPTE_RETENUE_DEFAUT
        core.update_facture_achat(
            self.conn, self.current_facture_id,
            numero=self.numero_var.get().strip(), date_facture=date_str, fournisseur_code=fournisseur_code,
            entete=self.entete_text.get("1.0", "end").strip(),
            pied_page=self.pied_text.get("1.0", "end").strip(),
            retenue_taux=retenue_taux, retenue_compte=retenue_compte,
        )
        core.set_setting(self.conn, "retenue_taux_defaut", retenue_taux)
        core.set_setting(self.conn, "retenue_compte_defaut", retenue_compte)
        messagebox.showinfo("Enregistré", "Facture enregistrée (brouillon).")
        self.refresh_factures_list()

    def valider(self):
        f = self._ensure_facture()
        if not f:
            return
        self.save_facture()
        if messagebox.askyesno(
            "Confirmer la validation",
            "Valider cette facture ? Les écritures comptables seront envoyées dans le menu SAISIE "
            "(débit achats, crédit fournisseur, retenue à la source, et entrée de stock automatique "
            "pour les lignes marchandises/matières premières). Cette action est définitive."
        ):
            try:
                warnings = core.valider_facture_achat(self.conn, self.current_facture_id)
            except ValueError as exc:
                messagebox.showerror("Erreur", str(exc))
                return
            msg = "Facture validée et écritures envoyées en Saisie."
            if warnings:
                msg += "\n\nAvertissements :\n" + "\n".join(warnings)
            messagebox.showinfo("Validation terminée", msg)
            self.refresh_factures_list()

    def delete_facture(self):
        if not self.current_facture_id:
            return
        if messagebox.askyesno("Confirmer", "Supprimer cette facture ?"):
            try:
                core.delete_facture_achat(self.conn, self.current_facture_id)
            except ValueError as exc:
                messagebox.showerror("Erreur", str(exc))
                return
            self.current_facture_id = None
            self.refresh_factures_list()

    def corriger_facture(self):
        """Repasse une facture d'achat déjà validée en brouillon modifiable,
        en supprimant les écritures comptables qu'elle avait générées — pour
        corriger une erreur sur les chiffres, puis revalider ensuite."""
        if not self.current_facture_id:
            return
        if not messagebox.askyesno(
            "Corriger cette facture",
            "Cette facture est déjà validée : ses écritures comptables (débit achats, "
            "crédit fournisseur, retenue à la source, entrée de stock) vont être RETIRÉES "
            "de la Saisie et la facture repassera en brouillon modifiable.\n\n"
            "Vous pourrez alors corriger les chiffres puis la revalider.\n\n"
            "Continuer ?"
        ):
            return
        try:
            core.devalider_facture_achat(self.conn, self.current_facture_id)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc))
            return
        messagebox.showinfo("Facture repassée en brouillon",
                             "La facture est de nouveau modifiable. Corrigez les chiffres puis "
                             "cliquez sur « Valider et envoyer en Saisie ».")
        self.refresh_factures_list()

    def _refresh_retenue_presets(self):
        presets = core.list_taux_retenue(self.conn)
        self.retenue_preset_combo["values"] = [f"{p['label']} ({p['montant']:g}%)" for p in presets]
        self._retenue_presets = presets

    def _on_retenue_preset_selected(self, event=None):
        idx = self.retenue_preset_combo.current()
        if idx is not None and 0 <= idx < len(getattr(self, "_retenue_presets", [])):
            preset = self._retenue_presets[idx]
            self.retenue_taux_var.set(str(preset["montant"]))
            if preset.get("compte"):
                self.retenue_compte_var.set(preset["compte"])

    def imprimer_facture(self):
        """Génère le document imprimable et l'ouvre directement (bouton
        « Imprimer » intégré, ou Ctrl+P depuis le navigateur) : Bon de
        commande tant que la facture est en brouillon (utilise le modèle
        par défaut ADMIN si l'en-tête/pied de page n'est pas rempli),
        Facture d'achat une fois validée."""
        if not self.current_facture_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord une facture.")
            return
        f = core.get_facture_achat(self.conn, self.current_facture_id)
        import tempfile
        import webbrowser
        if f and f["statut"] == "validee":
            path = os.path.join(tempfile.gettempdir(), f"facture_achat_{self.current_facture_id}.html")
            try:
                core.export_facture_achat_html(self.conn, self.current_facture_id, path)
            except ValueError as exc:
                messagebox.showerror("Erreur", str(exc))
                return
        else:
            path = os.path.join(tempfile.gettempdir(), f"bon_commande_{self.current_facture_id}.html")
            try:
                core.export_bon_commande_html(self.conn, self.current_facture_id, path)
            except ValueError as exc:
                messagebox.showerror("Erreur", str(exc))
                return
        webbrowser.open(f"file://{path}")

    def refresh(self):
        self._refresh_fournisseur_values()
        self._refresh_ligne_compte_values()
        self._refresh_retenue_compte_values()
        self._refresh_retenue_presets()
        self._refresh_ligne_analytic_values()
        self.refresh_factures_list()


class ExpressionBesoinDialog(tk.Toplevel):
    """Détail d'une Expression de besoin (double-clic depuis la liste) —
    aucun lien avec la comptabilité. La validation la fait basculer en Bon
    de commande (menu ENGAGEMENTS-PROJETS > Bon de commande)."""

    def __init__(self, parent, conn, expression_id, on_saved):
        super().__init__(parent)
        self.conn = conn
        self.expression_id = expression_id
        self.on_saved = on_saved
        self.title("Expression de besoin")
        self.geometry("900x600")
        self.minsize(700, 450)
        self.transient(parent)
        self.grab_set()

        exp = core.get_expression_besoin(conn, expression_id)
        self.validee = exp["statut"] == "validee"

        header = ttk.LabelFrame(self, text="Informations")
        header.pack(fill="x", padx=10, pady=8)
        ttk.Label(header, text="N° :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.numero_var = tk.StringVar(value=exp["numero"])
        ttk.Entry(header, textvariable=self.numero_var, width=16).grid(row=0, column=1, padx=4)
        ttk.Label(header, text="Date (JJ/MM/AAAA) :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.date_var = tk.StringVar(value=core.to_display_date(exp["date_demande"]))
        ttk.Entry(header, textvariable=self.date_var, width=14).grid(row=0, column=3, padx=4)
        ttk.Label(header, text="Demandeur :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.demandeur_var = tk.StringVar(value=exp["demandeur"] or "")
        ttk.Entry(header, textvariable=self.demandeur_var, width=20).grid(row=0, column=5, padx=4)
        ttk.Label(header, text="Service :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.service_var = tk.StringVar(value=exp["service"] or "")
        ttk.Entry(header, textvariable=self.service_var, width=20).grid(row=1, column=1, columnspan=2, padx=4,
                                                                          pady=(4, 0), sticky="w")
        self.statut_var = tk.StringVar(value=f"Statut : {'VALIDÉE (bon de commande créé)' if self.validee else 'Brouillon'}")
        ttk.Label(header, textvariable=self.statut_var, font=("Segoe UI", 10, "bold")).grid(
            row=1, column=3, columnspan=3, sticky="w", padx=(12, 4), pady=(4, 0))

        lignes_frame = ttk.LabelFrame(self, text="Lignes du besoin")
        lignes_frame.pack(fill="both", padx=10, pady=6)
        form = ttk.Frame(lignes_frame)
        form.pack(fill="x", padx=6, pady=4)
        ttk.Label(form, text="Libellé :").grid(row=0, column=0, sticky="w")
        self.ligne_libelle_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_libelle_var, width=32).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Quantité :").grid(row=0, column=2, sticky="w", padx=(12, 0))
        self.ligne_qte_var = tk.StringVar(value="1")
        ttk.Entry(form, textvariable=self.ligne_qte_var, width=10).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Unité :").grid(row=0, column=4, sticky="w", padx=(12, 0))
        self.ligne_unite_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_unite_var, width=10).grid(row=0, column=5, padx=4)
        self.add_ligne_btn = ttk.Button(form, text="Ajouter la ligne", command=self.add_ligne)
        self.add_ligne_btn.grid(row=0, column=6, padx=12)

        cols = ("id", "libelle", "quantite", "unite")
        self.tree = ttk.Treeview(lignes_frame, columns=cols, show="headings", height=12)
        for c, h, w in zip(cols, ["ID", "Libellé", "Quantité", "Unité"], [40, 400, 100, 100]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", padx=6, pady=6)
        self.delete_ligne_btn = ttk.Button(lignes_frame, text="Supprimer la ligne sélectionnée",
                                            command=self.delete_ligne)
        self.delete_ligne_btn.pack(anchor="w", padx=6, pady=(0, 6))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=10)
        self.save_btn = ttk.Button(btns, text="Enregistrer", command=self.save)
        self.save_btn.pack(side="left", padx=4)
        self.valider_btn = ttk.Button(btns, text="Valider → crée le Bon de commande", command=self.valider)
        self.valider_btn.pack(side="left", padx=4)
        ttk.Button(btns, text="Fermer", command=self.destroy).pack(side="left", padx=4)

        self.refresh_lignes()
        self._apply_lock()

    def _apply_lock(self):
        state = "disabled" if self.validee else "normal"
        self.add_ligne_btn.configure(state=state)
        self.delete_ligne_btn.configure(state=state)
        self.save_btn.configure(state=state)
        self.valider_btn.configure(state=state)

    def refresh_lignes(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for l in core.list_lignes_expression_besoin(self.conn, self.expression_id):
            self.tree.insert("", "end", values=(l["id"], l["libelle"], f"{l['quantite']:g}", l["unite"] or ""))

    def add_ligne(self):
        libelle = self.ligne_libelle_var.get().strip()
        if not libelle:
            messagebox.showwarning("Champ manquant", "Le libellé est obligatoire.", parent=self)
            return
        try:
            qte = float(self.ligne_qte_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "La quantité doit être un nombre.", parent=self)
            return
        core.add_ligne_expression_besoin(self.conn, self.expression_id, libelle, qte,
                                          unite=self.ligne_unite_var.get().strip() or None)
        self.ligne_libelle_var.set("")
        self.ligne_qte_var.set("1")
        self.ligne_unite_var.set("")
        self.refresh_lignes()

    def delete_ligne(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne.", parent=self)
            return
        ligne_id = self.tree.item(sel[0], "values")[0]
        core.delete_ligne_expression_besoin(self.conn, ligne_id)
        self.refresh_lignes()

    def save(self):
        date_str = core.to_iso_date(self.date_var.get().strip())
        core.update_expression_besoin(
            self.conn, self.expression_id,
            numero=self.numero_var.get().strip(), date_demande=date_str,
            demandeur=self.demandeur_var.get().strip(), service=self.service_var.get().strip(),
        )
        messagebox.showinfo("Enregistré", "Expression de besoin enregistrée.", parent=self)
        self.on_saved()

    def valider(self):
        if not messagebox.askyesno(
            "Valider cette expression de besoin",
            "Cette expression va être verrouillée et un Bon de commande va être créé automatiquement "
            "avec les mêmes lignes (menu ENGAGEMENTS-PROJETS > Bon de commande).\n\n"
            "Aucune écriture comptable n'est générée à cette étape.\n\nContinuer ?",
            parent=self,
        ):
            return
        self.save()
        try:
            core.valider_expression_besoin(self.conn, self.expression_id)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc), parent=self)
            return
        messagebox.showinfo("Validée", "Expression de besoin validée. Le Bon de commande a été créé.", parent=self)
        self.on_saved()
        self.destroy()


class ExpressionBesoinTab(ttk.Frame):
    """Liste des Expressions de besoin (menu ENGAGEMENTS-PROJETS) — double-clic
    sur une ligne pour l'ouvrir en grand et la valider (bascule en Bon de
    commande). Circuit interne, sans lien avec la comptabilité."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="EXPRESSION DE BESOIN", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Circuit interne d'approbation d'achat — sans lien avec la comptabilité. Double-cliquez sur une "
            "ligne pour l'ouvrir, ajouter des lignes de besoin, et la valider : elle bascule alors "
            "automatiquement dans le sous-menu « Bon de commande »."
        ), foreground="#595959", wraplength=1100).pack(anchor="w", padx=16, pady=(0, 8))

        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill="x", padx=16, pady=4)
        ttk.Button(btn_bar, text="Nouvelle expression de besoin", command=self.new_expression).pack(side="left")
        ttk.Button(btn_bar, text="Actualiser", command=self.refresh).pack(side="left", padx=8)

        cols = ("numero", "date", "demandeur", "service", "nb_lignes", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=22)
        headers = ["N°", "Date", "Demandeur", "Service", "Lignes", "Statut"]
        widths = [110, 100, 200, 200, 70, 140]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<Double-1>", self._on_double_click)
        self._by_iid = {}
        self.refresh()

    def new_expression(self):
        numero = simpledialog.askstring("Nouvelle expression de besoin", "N° :", parent=self)
        if not numero:
            return
        eid = core.create_expression_besoin(self.conn, numero, date.today().strftime("%Y-%m-%d"))
        self.refresh()
        ExpressionBesoinDialog(self, self.conn, eid, on_saved=self.refresh)

    def _on_double_click(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        eid = self._by_iid.get(sel[0])
        if eid:
            ExpressionBesoinDialog(self, self.conn, eid, on_saved=self.refresh)

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self._by_iid = {}
        for exp in core.list_expressions_besoin(self.conn):
            nb = len(core.list_lignes_expression_besoin(self.conn, exp["id"]))
            statut = "Validée" if exp["statut"] == "validee" else "Brouillon"
            iid = self.tree.insert("", "end", values=(
                exp["numero"], core.to_display_date(exp["date_demande"]), exp["demandeur"] or "",
                exp["service"] or "", nb, statut,
            ))
            self._by_iid[iid] = exp["id"]


class BonCommandeEPDialog(tk.Toplevel):
    """Détail d'un Bon de commande (double-clic depuis la liste). La
    validation COMPTABILISE DIRECTEMENT l'achat (chaque ligne doit avoir un
    compte de charge choisi) et fait basculer le document en Bordereau de
    livraison — voir core.valider_ep_bon_commande()."""

    def __init__(self, parent, conn, bon_id, on_saved):
        super().__init__(parent)
        self.conn = conn
        self.bon_id = bon_id
        self.on_saved = on_saved
        self.selected_ligne_id = None
        self.title("Bon de commande")
        self.geometry("1080x680")
        self.minsize(850, 500)
        self.transient(parent)
        self.grab_set()

        bon = core.get_ep_bon_commande(conn, bon_id)
        self.validee = bon["statut"] == "validee"

        header = ttk.LabelFrame(self, text="Informations")
        header.pack(fill="x", padx=10, pady=8)
        ttk.Label(header, text="N° :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.numero_var = tk.StringVar(value=bon["numero"])
        ttk.Entry(header, textvariable=self.numero_var, width=14).grid(row=0, column=1, padx=4)
        ttk.Label(header, text="Date (JJ/MM/AAAA) :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.date_var = tk.StringVar(value=core.to_display_date(bon["date_commande"]))
        ttk.Entry(header, textvariable=self.date_var, width=12).grid(row=0, column=3, padx=4)
        ttk.Label(header, text="Fournisseur :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.fournisseur_var = tk.StringVar(value=bon["fournisseur_code"] or "")
        self.fournisseur_combo = ttk.Combobox(header, textvariable=self.fournisseur_var, width=24)
        self.fournisseur_combo.grid(row=0, column=5, padx=4)
        self.fournisseur_combo.bind("<KeyRelease>", self._on_fournisseur_keyrelease)
        self.fournisseur_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_fournisseur_values()

        ttk.Label(header, text="Date de facture :").grid(row=1, column=0, sticky="w", padx=4, pady=(6, 0))
        self.date_facture_var = tk.StringVar(value=core.to_display_date(bon.get("date_facture") or ""))
        ttk.Entry(header, textvariable=self.date_facture_var, width=12).grid(row=1, column=1, padx=4, pady=(6, 0), sticky="w")
        ttk.Label(header, text="Date de saisie :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(6, 0))
        self.date_saisie_var = tk.StringVar(value=core.to_display_date(bon.get("date_saisie") or ""))
        ttk.Entry(header, textvariable=self.date_saisie_var, width=12).grid(row=1, column=3, padx=4, pady=(6, 0), sticky="w")
        ttk.Label(header, text="Paiement attendu :").grid(row=1, column=4, sticky="w", padx=(12, 4), pady=(6, 0))
        self.date_paiement_var = tk.StringVar(value=core.to_display_date(bon.get("date_paiement_attendu") or ""))
        ttk.Entry(header, textvariable=self.date_paiement_var, width=12).grid(row=1, column=5, padx=4, pady=(6, 0), sticky="w")

        ttk.Label(header, text="Retenue % :").grid(row=2, column=0, sticky="w", padx=4, pady=(6, 0))
        self.retenue_taux_var = tk.StringVar(value=str(bon.get("retenue_taux") or 0))
        ttk.Entry(header, textvariable=self.retenue_taux_var, width=8).grid(row=2, column=1, padx=4, pady=(6, 0), sticky="w")
        ttk.Label(header, text="Compte retenue :").grid(row=2, column=2, sticky="w", padx=(12, 4), pady=(6, 0))
        self.retenue_compte_var = tk.StringVar(value=bon.get("retenue_compte") or "447800")
        self.retenue_compte_combo = ttk.Combobox(header, textvariable=self.retenue_compte_var, width=18)
        self.retenue_compte_combo.grid(row=2, column=3, padx=4, pady=(6, 0), sticky="w")
        self.retenue_compte_combo.bind("<KeyRelease>", self._on_retenue_compte_keyrelease)
        self.retenue_compte_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_retenue_compte_values()
        ttk.Label(header, text="Préréglage :").grid(row=2, column=4, sticky="w", padx=(12, 4), pady=(6, 0))
        self.retenue_preset_var = tk.StringVar()
        self.retenue_preset_combo = ttk.Combobox(header, textvariable=self.retenue_preset_var, width=20, state="readonly")
        self.retenue_preset_combo.grid(row=2, column=5, padx=4, pady=(6, 0))
        self.retenue_preset_combo.bind("<<ComboboxSelected>>", self._on_retenue_preset_selected)
        self.retenue_preset_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_retenue_presets()

        origine = "" if not bon["expression_id"] else f"Issu de l'expression de besoin n° {bon['expression_id']}"
        self.statut_var = tk.StringVar(
            value=f"Statut : {'VALIDÉ (comptabilisé + bordereau créé)' if self.validee else 'Brouillon'}   {origine}")
        ttk.Label(header, textvariable=self.statut_var, font=("Segoe UI", 10, "bold")).grid(
            row=3, column=0, columnspan=4, sticky="w", padx=4, pady=(4, 0))
        self.retard_var = tk.StringVar()
        self.retard_label = ttk.Label(header, textvariable=self.retard_var, font=("Segoe UI", 10, "bold"))
        self.retard_label.grid(row=3, column=4, columnspan=2, sticky="w", padx=4, pady=(4, 0))
        self._refresh_retard()

        lignes_frame = ttk.LabelFrame(self, text=(
            "Lignes — un compte débiteur (charge ou immobilisation) est OBLIGATOIRE sur chaque ligne "
            "pour pouvoir valider"))
        lignes_frame.pack(fill="both", padx=10, pady=6)
        form = ttk.Frame(lignes_frame)
        form.pack(fill="x", padx=6, pady=4)
        ttk.Label(form, text="Compte débiteur (charge ou immobilisation) :").grid(row=0, column=0, sticky="w")
        self.ligne_compte_var = tk.StringVar()
        self.ligne_compte_combo = ttk.Combobox(form, textvariable=self.ligne_compte_var, width=26)
        self.ligne_compte_combo.grid(row=0, column=1, padx=4)
        self.ligne_compte_combo.bind("<KeyRelease>", self._on_ligne_compte_keyrelease)
        self.ligne_compte_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_ligne_compte_values()
        ttk.Label(form, text="Code analytique :").grid(row=0, column=2, sticky="w", padx=(12, 0))
        self.ligne_analytic_var = tk.StringVar()
        self.ligne_analytic_combo = ttk.Combobox(form, textvariable=self.ligne_analytic_var, width=18)
        self.ligne_analytic_combo.grid(row=0, column=3, padx=4)
        self.ligne_analytic_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_ligne_analytic_values()

        ttk.Label(form, text="Libellé :").grid(row=1, column=0, sticky="w", pady=(4, 0))
        self.ligne_libelle_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_libelle_var, width=26).grid(row=1, column=1, padx=4, pady=(4, 0))
        ttk.Label(form, text="Quantité :").grid(row=1, column=2, sticky="w", padx=(12, 0), pady=(4, 0))
        self.ligne_qte_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_qte_var, width=8).grid(row=1, column=3, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(form, text="Prix unitaire :").grid(row=1, column=4, sticky="w", padx=(12, 0), pady=(4, 0))
        self.ligne_prix_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_prix_var, width=10).grid(row=1, column=5, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(form, text="Unité :").grid(row=1, column=6, sticky="w", padx=(12, 0), pady=(4, 0))
        self.ligne_unite_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_unite_var, width=8).grid(row=1, column=7, padx=4, pady=(4, 0))

        btn_row = ttk.Frame(lignes_frame)
        btn_row.pack(fill="x", padx=6)
        self.update_ligne_btn = ttk.Button(btn_row, text="Mettre à jour la ligne sélectionnée",
                                            command=self.update_ligne)
        self.update_ligne_btn.pack(side="left")
        self.add_ligne_btn = ttk.Button(btn_row, text="Ajouter une nouvelle ligne", command=self.add_ligne)
        self.add_ligne_btn.pack(side="left", padx=8)
        self.delete_ligne_btn = ttk.Button(btn_row, text="Supprimer la ligne sélectionnée", command=self.delete_ligne)
        self.delete_ligne_btn.pack(side="left", padx=8)

        cols = ("id", "compte", "libelle", "quantite", "prix", "montant", "unite", "analytique")
        self.tree = ttk.Treeview(lignes_frame, columns=cols, show="headings", height=10)
        headers = ["ID", "Compte", "Libellé", "Qté", "Prix unit.", "Montant estimé", "Unité", "Analytique"]
        widths = [40, 90, 220, 70, 90, 110, 70, 130]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", padx=6, pady=6)
        self.tree.bind("<<TreeviewSelect>>", self._on_select_ligne)
        self.total_var = tk.StringVar()
        ttk.Label(lignes_frame, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=6, pady=(0, 6))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=10)
        self.save_btn = ttk.Button(btns, text="Enregistrer", command=self.save)
        self.save_btn.pack(side="left", padx=4)
        self.valider_btn = ttk.Button(btns, text="Valider (comptabilise + crée le Bordereau)", command=self.valider)
        self.valider_btn.pack(side="left", padx=4)
        self.corriger_btn = ttk.Button(btns, text="Corriger (repasser en brouillon)", command=self.corriger)
        self.corriger_btn.pack(side="left", padx=4)
        ttk.Button(btns, text="Fermer", command=self.destroy).pack(side="left", padx=4)

        self.refresh_lignes()
        self._apply_lock()

    @staticmethod
    def _open_dropdown(event=None):
        if event is not None:
            event.widget.event_generate("<Down>")

    def _refresh_fournisseur_values(self):
        items = core.list_fournisseurs(self.conn)
        self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _on_fournisseur_keyrelease(self, event=None):
        query = self._extract_code(self.fournisseur_var.get())
        items = core.list_fournisseurs(self.conn, query)
        self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _refresh_retenue_compte_values(self):
        items = [a for a in core.search_accounts(self.conn, "44", limit=100) if core.account_racine(a["code"]) == "44"]
        self.retenue_compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _on_retenue_compte_keyrelease(self, event=None):
        query = self._extract_code(self.retenue_compte_var.get())
        items = [a for a in core.search_accounts(self.conn, query, limit=100) if core.account_racine(a["code"]) == "44"]
        self.retenue_compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _refresh_retenue_presets(self):
        presets = core.list_taux_retenue(self.conn)
        self.retenue_preset_combo["values"] = [f"{p['label']} ({p['montant']:g}%)" for p in presets]
        self._retenue_presets = presets

    def _on_retenue_preset_selected(self, event=None):
        idx = self.retenue_preset_combo.current()
        if idx is not None and 0 <= idx < len(getattr(self, "_retenue_presets", [])):
            preset = self._retenue_presets[idx]
            self.retenue_taux_var.set(str(preset["montant"]))
            if preset.get("compte"):
                self.retenue_compte_var.set(preset["compte"])

    def _refresh_ligne_compte_values(self):
        items = [a for a in core.search_accounts(self.conn, "", limit=150) if a["classe"] in ("2", "6")]
        self.ligne_compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _on_ligne_compte_keyrelease(self, event=None):
        query = self._extract_code(self.ligne_compte_var.get())
        items = [a for a in core.search_accounts(self.conn, query, limit=50) if a["classe"] in ("2", "6")]
        self.ligne_compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _refresh_ligne_analytic_values(self):
        self.ligne_analytic_combo["values"] = [f"{c['code']} — {c['label']}" for c in core.list_analytic_codes(self.conn)]

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def _apply_lock(self):
        state = "disabled" if self.validee else "normal"
        for w in (self.update_ligne_btn, self.add_ligne_btn, self.delete_ligne_btn, self.save_btn, self.valider_btn):
            w.configure(state=state)
        self.corriger_btn.configure(state="normal" if self.validee else "disabled")

    def _on_select_ligne(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_ligne_id = v[0]
        self.ligne_compte_var.set(v[1])
        self.ligne_libelle_var.set(v[2])
        self.ligne_qte_var.set(v[3])
        self.ligne_prix_var.set(v[4])
        self.ligne_unite_var.set(v[6])
        self.ligne_analytic_var.set(v[7])

    def _clear_ligne_form(self):
        self.selected_ligne_id = None
        self.ligne_compte_var.set(""); self.ligne_libelle_var.set("")
        self.ligne_qte_var.set(""); self.ligne_prix_var.set("")
        self.ligne_unite_var.set(""); self.ligne_analytic_var.set("")

    def refresh_lignes(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        total = 0.0
        for l in core.list_lignes_ep_bon_commande(self.conn, self.bon_id):
            total += l["montant_ht"]
            self.tree.insert("", "end", values=(
                l["id"], l["compte_charge"] or "⚠ à choisir", l["libelle"], f"{l['quantite']:g}",
                f"{fmt_cfa(l['prix_unitaire'])}", f"{fmt_cfa(l['montant_ht'])}", l["unite"] or "",
                l["analytic_code"] or ""))
        self.total_var.set(f"Total HT : {fmt_cfa(total)}")

    def add_ligne(self):
        libelle = self.ligne_libelle_var.get().strip()
        if not libelle:
            messagebox.showwarning("Champ manquant", "Le libellé est obligatoire.", parent=self)
            return
        try:
            qte = float(self.ligne_qte_var.get() or 0)
            prix = float(self.ligne_prix_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Quantité et Prix unitaire doivent être des nombres.", parent=self)
            return
        compte = self._extract_code(self.ligne_compte_var.get()) or None
        analytic = self._extract_code(self.ligne_analytic_var.get()) or None
        core.add_ligne_ep_bon_commande(self.conn, self.bon_id, libelle, qte, prix_unitaire=prix,
                                        unite=self.ligne_unite_var.get().strip() or None,
                                        compte_charge=compte, analytic_code=analytic)
        self._clear_ligne_form()
        self.refresh_lignes()

    def update_ligne(self):
        if not self.selected_ligne_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne dans le tableau.", parent=self)
            return
        libelle = self.ligne_libelle_var.get().strip()
        if not libelle:
            messagebox.showwarning("Champ manquant", "Le libellé est obligatoire.", parent=self)
            return
        try:
            qte = float(self.ligne_qte_var.get() or 0)
            prix = float(self.ligne_prix_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Quantité et Prix unitaire doivent être des nombres.", parent=self)
            return
        compte = self._extract_code(self.ligne_compte_var.get()) or None
        analytic = self._extract_code(self.ligne_analytic_var.get()) or None
        core.update_ligne_ep_bon_commande(
            self.conn, self.selected_ligne_id, libelle=libelle, quantite=qte, prix_unitaire=prix,
            unite=self.ligne_unite_var.get().strip() or None, compte_charge=compte, analytic_code=analytic)
        self._clear_ligne_form()
        self.refresh_lignes()

    def delete_ligne(self):
        if not self.selected_ligne_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne.", parent=self)
            return
        core.delete_ligne_ep_bon_commande(self.conn, self.selected_ligne_id)
        self._clear_ligne_form()
        self.refresh_lignes()

    def _refresh_retard(self):
        bons = core.list_ep_bons_commande(self.conn)
        match = next((b for b in bons if b["id"] == self.bon_id), None)
        if not match or not match.get("statut_paiement"):
            self.retard_var.set("")
            return
        self.retard_var.set(f"Paiement : {match['statut_paiement']}")
        self.retard_label.configure(foreground="#B00020" if match["depassement_paiement"] else "#1F7A1F")

    def save(self):
        date_str = core.to_iso_date(self.date_var.get().strip())
        try:
            retenue_taux = float(self.retenue_taux_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Le taux de retenue doit être un nombre.", parent=self)
            return
        core.update_ep_bon_commande(
            self.conn, self.bon_id,
            numero=self.numero_var.get().strip(), date_commande=date_str,
            fournisseur_code=self._extract_code(self.fournisseur_var.get()),
            date_facture=core.to_iso_date(self.date_facture_var.get().strip()) or None,
            date_saisie=core.to_iso_date(self.date_saisie_var.get().strip()) or None,
            date_paiement_attendu=core.to_iso_date(self.date_paiement_var.get().strip()) or None,
            retenue_taux=retenue_taux, retenue_compte=self._extract_code(self.retenue_compte_var.get()) or "447800",
        )
        self._refresh_retard()
        messagebox.showinfo("Enregistré", "Bon de commande enregistré.", parent=self)
        self.on_saved()

    def valider(self):
        if not messagebox.askyesno(
            "Valider ce bon de commande",
            "Ce bon de commande va être verrouillé et COMPTABILISÉ DIRECTEMENT (débit des comptes de "
            "charge choisis, crédit fournisseur, retenue fiscale éventuelle, entrée de stock automatique "
            "si applicable) — l'écriture sera envoyée en Saisie. Un Bordereau de livraison sera aussi "
            "créé pour le suivi de la réception.\n\nContinuer ?",
            parent=self,
        ):
            return
        self.save()
        try:
            core.valider_ep_bon_commande(self.conn, self.bon_id)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc), parent=self)
            return
        messagebox.showinfo(
            "Validé",
            "Bon de commande validé et comptabilisé — écriture envoyée en Saisie. Le Bordereau de "
            "livraison a été créé.",
            parent=self,
        )
        self.on_saved()
        self.destroy()

    def corriger(self):
        if not messagebox.askyesno(
            "Corriger ce bon de commande",
            "Ce bon de commande est déjà validé : ses écritures comptables vont être RETIRÉES de la "
            "Saisie, et il repassera en brouillon modifiable (le Règlement lié aussi). Le Bordereau de "
            "livraison déjà créé n'est pas supprimé.\n\n"
            "Vous pourrez alors corriger les lignes (compte, prix...) puis revalider.\n\nContinuer ?",
            parent=self,
        ):
            return
        try:
            core.devalider_ep_bon_commande(self.conn, self.bon_id)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc), parent=self)
            return
        self.validee = False
        bon = core.get_ep_bon_commande(self.conn, self.bon_id)
        origine = "" if not bon["expression_id"] else f"Issu de l'expression de besoin n° {bon['expression_id']}"
        self.statut_var.set(f"Statut : Brouillon   {origine}")
        self._apply_lock()
        self.refresh_lignes()
        messagebox.showinfo("Repassé en brouillon",
                             "Le bon de commande est de nouveau modifiable. Corrigez les lignes puis "
                             "cliquez sur « Valider ».", parent=self)
        self.on_saved()


class BonCommandeEPTab(ttk.Frame):
    """Liste des Bons de commande du circuit interne (menu ENGAGEMENTS-PROJETS)
    — double-clic pour ouvrir et valider (bascule en Bordereau de livraison).
    Sans lien avec la comptabilité."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="BON DE COMMANDE", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Circuit interne d'approbation d'achat — sans lien avec la comptabilité (créé automatiquement "
            "en validant une Expression de besoin, ou directement ici). Double-cliquez sur une ligne pour "
            "l'ouvrir et la valider : elle bascule alors dans le sous-menu « Bordereau de livraison »."
        ), foreground="#595959", wraplength=1100).pack(anchor="w", padx=16, pady=(0, 8))

        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill="x", padx=16, pady=4)
        ttk.Button(btn_bar, text="Nouveau bon de commande", command=self.new_bon).pack(side="left")
        ttk.Button(btn_bar, text="Actualiser", command=self.refresh).pack(side="left", padx=8)

        cols = ("numero", "date", "fournisseur", "nb_lignes", "statut", "paiement_attendu", "statut_paiement")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=22)
        headers = ["N°", "Date", "Fournisseur", "Lignes", "Statut", "Paiement attendu", "Statut paiement"]
        widths = [110, 100, 220, 60, 130, 130, 160]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.tag_configure("retard", foreground="#B00020", font=("Segoe UI", 9, "bold"))
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<Double-1>", self._on_double_click)
        self._by_iid = {}
        self.refresh()

    def new_bon(self):
        numero = simpledialog.askstring("Nouveau bon de commande", "N° :", parent=self)
        if not numero:
            return
        bid = core.create_ep_bon_commande(self.conn, numero, date.today().strftime("%Y-%m-%d"))
        self.refresh()
        BonCommandeEPDialog(self, self.conn, bid, on_saved=self.refresh)

    def _on_double_click(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        bid = self._by_iid.get(sel[0])
        if bid:
            BonCommandeEPDialog(self, self.conn, bid, on_saved=self.refresh)

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self._by_iid = {}
        for bon in core.list_ep_bons_commande(self.conn):
            nb = len(core.list_lignes_ep_bon_commande(self.conn, bon["id"]))
            statut = "Validé" if bon["statut"] == "validee" else "Brouillon"
            fournisseur = ""
            if bon["fournisseur_code"]:
                f = core.get_fournisseur(self.conn, bon["fournisseur_code"])
                fournisseur = f"{bon['fournisseur_code']} — {f['raison_sociale']}" if f else bon["fournisseur_code"]
            tags = ("retard",) if bon.get("depassement_paiement") else ()
            iid = self.tree.insert("", "end", tags=tags, values=(
                bon["numero"], core.to_display_date(bon["date_commande"]), fournisseur, nb, statut,
                core.to_display_date(bon.get("date_paiement_attendu") or ""), bon.get("statut_paiement") or "",
            ))
            self._by_iid[iid] = bon["id"]


class BordereauLivraisonDialog(tk.Toplevel):
    """Détail d'un Bordereau de livraison (double-clic depuis la liste) —
    dernière étape du circuit interne, sans lien avec la comptabilité. La
    validation marque simplement la réception comme confirmée."""

    def __init__(self, parent, conn, bordereau_id, on_saved):
        super().__init__(parent)
        self.conn = conn
        self.bordereau_id = bordereau_id
        self.on_saved = on_saved
        self.title("Bordereau de livraison")
        self.geometry("950x600")
        self.minsize(750, 450)
        self.transient(parent)
        self.grab_set()

        bl = core.get_bordereau_livraison(conn, bordereau_id)
        self.validee = bl["statut"] == "validee"

        header = ttk.LabelFrame(self, text="Informations")
        header.pack(fill="x", padx=10, pady=8)
        ttk.Label(header, text="N° :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.numero_var = tk.StringVar(value=bl["numero"])
        ttk.Entry(header, textvariable=self.numero_var, width=16).grid(row=0, column=1, padx=4)
        ttk.Label(header, text="Date (JJ/MM/AAAA) :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.date_var = tk.StringVar(value=core.to_display_date(bl["date_livraison"]))
        ttk.Entry(header, textvariable=self.date_var, width=14).grid(row=0, column=3, padx=4)
        origine = "" if not bl["bon_commande_id"] else f"Issu du bon de commande n° {bl['bon_commande_id']}"
        self.statut_var = tk.StringVar(
            value=f"Statut : {'VALIDÉ (réception confirmée)' if self.validee else 'Brouillon'}   {origine}")
        ttk.Label(header, textvariable=self.statut_var, font=("Segoe UI", 10, "bold")).grid(
            row=1, column=0, columnspan=4, sticky="w", padx=4, pady=(4, 0))

        lignes_frame = ttk.LabelFrame(self, text="Lignes livrées (quantité livrée modifiable)")
        lignes_frame.pack(fill="both", padx=10, pady=6)

        cols = ("id", "libelle", "qte_cmd", "qte_liv", "unite")
        self.tree = ttk.Treeview(lignes_frame, columns=cols, show="headings", height=14)
        headers = ["ID", "Libellé", "Quantité commandée", "Quantité livrée", "Unité"]
        widths = [40, 380, 150, 140, 80]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", padx=6, pady=6)
        self.tree.bind("<Double-1>", self._on_edit_qte_livree)

        ttk.Label(lignes_frame, text=(
            "Double-cliquez sur une ligne pour corriger la quantité réellement livrée."
        ), foreground="#595959").pack(anchor="w", padx=6, pady=(0, 6))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=10)
        self.save_btn = ttk.Button(btns, text="Enregistrer", command=self.save)
        self.save_btn.pack(side="left", padx=4)
        self.valider_btn = ttk.Button(btns, text="Valider (réception confirmée)", command=self.valider)
        self.valider_btn.pack(side="left", padx=4)
        ttk.Button(btns, text="Fermer", command=self.destroy).pack(side="left", padx=4)

        self.refresh_lignes()
        self._apply_lock()

    def _apply_lock(self):
        state = "disabled" if self.validee else "normal"
        self.save_btn.configure(state=state)
        self.valider_btn.configure(state=state)

    def refresh_lignes(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for l in core.list_lignes_bordereau_livraison(self.conn, self.bordereau_id):
            self.tree.insert("", "end", values=(
                l["id"], l["libelle"], f"{l['quantite_commandee']:g}", f"{l['quantite_livree']:g}", l["unite"] or ""))

    def _on_edit_qte_livree(self, event=None):
        if self.validee:
            return
        sel = self.tree.selection()
        if not sel:
            return
        values = self.tree.item(sel[0], "values")
        ligne_id, libelle, qte_cmd, qte_liv_actuelle = values[0], values[1], values[2], values[3]
        nouvelle = simpledialog.askfloat(
            "Quantité livrée", f"Quantité réellement livrée pour « {libelle} » (commandé : {qte_cmd}) :",
            initialvalue=float(qte_liv_actuelle), parent=self,
        )
        if nouvelle is None:
            return
        core.update_ligne_bordereau_livraison(self.conn, ligne_id, nouvelle)
        self.refresh_lignes()

    def save(self):
        date_str = core.to_iso_date(self.date_var.get().strip())
        core.update_bordereau_livraison(
            self.conn, self.bordereau_id,
            numero=self.numero_var.get().strip(), date_livraison=date_str,
        )
        messagebox.showinfo("Enregistré", "Bordereau de livraison enregistré.", parent=self)
        self.on_saved()

    def valider(self):
        if not messagebox.askyesno(
            "Valider ce bordereau",
            "Ce bordereau de livraison va être marqué comme reçu et verrouillé.\n\n"
            "Aucune écriture comptable n'est générée — c'est la dernière étape du circuit interne.\n\n"
            "Continuer ?",
            parent=self,
        ):
            return
        self.save()
        try:
            core.valider_bordereau_livraison(self.conn, self.bordereau_id)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc), parent=self)
            return
        messagebox.showinfo("Validé", "Bordereau de livraison validé.", parent=self)
        self.on_saved()
        self.destroy()


class BordereauLivraisonTab(ttk.Frame):
    """Liste des Bordereaux de livraison (menu ENGAGEMENTS-PROJETS) —
    dernière étape du circuit interne, sans lien avec la comptabilité.
    Double-clic pour ouvrir, corriger les quantités livrées et valider."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="BORDEREAU DE LIVRAISON", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Circuit interne d'approbation d'achat — sans lien avec la comptabilité (créé automatiquement "
            "en validant un Bon de commande). Double-cliquez sur une ligne pour l'ouvrir, corriger les "
            "quantités réellement livrées et confirmer la réception."
        ), foreground="#595959", wraplength=1100).pack(anchor="w", padx=16, pady=(0, 8))

        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill="x", padx=16, pady=4)
        ttk.Button(btn_bar, text="Actualiser", command=self.refresh).pack(side="left")

        cols = ("numero", "date", "nb_lignes", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=22)
        headers = ["N°", "Date", "Lignes", "Statut"]
        widths = [110, 100, 70, 160]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<Double-1>", self._on_double_click)
        self._by_iid = {}
        self.refresh()

    def _on_double_click(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        bid = self._by_iid.get(sel[0])
        if bid:
            BordereauLivraisonDialog(self, self.conn, bid, on_saved=self.refresh)

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self._by_iid = {}
        for bl in core.list_bordereaux_livraison(self.conn):
            nb = len(core.list_lignes_bordereau_livraison(self.conn, bl["id"]))
            statut = "Validé" if bl["statut"] == "validee" else "Brouillon"
            iid = self.tree.insert("", "end", values=(
                bl["numero"], core.to_display_date(bl["date_livraison"]), nb, statut,
            ))
            self._by_iid[iid] = bl["id"]


class ReglementDialog(tk.Toplevel):
    """Détail d'un Règlement (double-clic depuis la liste) — créé
    automatiquement en validant un Bon de commande du circuit interne
    (lignes recopiées SANS compte de charge). C'est ICI que la
    comptabilisation a lieu : chaque ligne doit recevoir un compte de
    charge (classe 6), un code analytique optionnel, et une retenue
    fiscale optionnelle, avant de pouvoir valider (envoie l'écriture en
    Saisie, avec entrée de stock automatique si applicable)."""

    def __init__(self, parent, conn, reglement_id, on_saved):
        super().__init__(parent)
        self.conn = conn
        self.reglement_id = reglement_id
        self.on_saved = on_saved
        self.selected_ligne_id = None
        self.title("Règlement")
        self.geometry("1080x680")
        self.minsize(850, 500)
        self.transient(parent)
        self.grab_set()

        reg = core.get_reglement(conn, reglement_id)
        self.validee = reg["statut"] == "validee"

        header = ttk.LabelFrame(self, text="Informations")
        header.pack(fill="x", padx=10, pady=8)
        ttk.Label(header, text="N° :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.numero_var = tk.StringVar(value=reg["numero"])
        ttk.Entry(header, textvariable=self.numero_var, width=14).grid(row=0, column=1, padx=4)
        ttk.Label(header, text="Date (JJ/MM/AAAA) :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.date_var = tk.StringVar(value=core.to_display_date(reg["date_reglement"]))
        ttk.Entry(header, textvariable=self.date_var, width=12).grid(row=0, column=3, padx=4)
        ttk.Label(header, text="Fournisseur :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.fournisseur_var = tk.StringVar()
        f0 = core.get_fournisseur(conn, reg["fournisseur_code"]) if reg["fournisseur_code"] else None
        self.fournisseur_var.set(f"{reg['fournisseur_code']} — {f0['raison_sociale']}" if f0
                                  else (reg["fournisseur_code"] or ""))
        self.fournisseur_combo = ttk.Combobox(header, textvariable=self.fournisseur_var, width=26)
        self.fournisseur_combo.grid(row=0, column=5, padx=4)
        self.fournisseur_combo.bind("<KeyRelease>", self._on_fournisseur_keyrelease)
        self.fournisseur_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_fournisseur_values()

        ttk.Label(header, text="Retenue % :").grid(row=1, column=0, sticky="w", padx=4, pady=(6, 0))
        self.retenue_taux_var = tk.StringVar(value=str(reg["retenue_taux"]))
        ttk.Entry(header, textvariable=self.retenue_taux_var, width=8).grid(row=1, column=1, sticky="w", padx=4, pady=(6, 0))
        ttk.Label(header, text="Compte retenue :").grid(row=1, column=2, sticky="w", padx=(12, 4), pady=(6, 0))
        self.retenue_compte_var = tk.StringVar(value=reg["retenue_compte"])
        self.retenue_compte_combo = ttk.Combobox(header, textvariable=self.retenue_compte_var, width=22)
        self.retenue_compte_combo.grid(row=1, column=3, padx=4, pady=(6, 0), sticky="w")
        self.retenue_compte_combo.bind("<KeyRelease>", self._on_retenue_compte_keyrelease)
        self.retenue_compte_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_retenue_compte_values()
        ttk.Label(header, text="Préréglage (ADMIN) :").grid(row=1, column=4, sticky="w", padx=(12, 4), pady=(6, 0))
        self.retenue_preset_var = tk.StringVar()
        self.retenue_preset_combo = ttk.Combobox(header, textvariable=self.retenue_preset_var, width=22, state="readonly")
        self.retenue_preset_combo.grid(row=1, column=5, padx=4, pady=(6, 0))
        self.retenue_preset_combo.bind("<<ComboboxSelected>>", self._on_retenue_preset_selected)
        self.retenue_preset_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_retenue_presets()

        paiement_frame = ttk.LabelFrame(header, text="Paiement bancaire/caisse (après validation de la charge)")
        paiement_frame.grid(row=2, column=0, columnspan=6, sticky="we", padx=4, pady=(8, 0))
        ttk.Label(paiement_frame, text="Date de paiement :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.date_paiement_var = tk.StringVar(value=core.to_display_date(reg.get("date_paiement") or "")
                                               or date.today().strftime("%d/%m/%Y"))
        ttk.Entry(paiement_frame, textvariable=self.date_paiement_var, width=12).grid(row=0, column=1, padx=4)
        ttk.Label(paiement_frame, text="Compte banque/caisse :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.compte_paiement_var = tk.StringVar(value=reg.get("compte_paiement") or "")
        self.compte_paiement_combo = ttk.Combobox(paiement_frame, textvariable=self.compte_paiement_var, width=26)
        self.compte_paiement_combo.grid(row=0, column=3, padx=4)
        self.compte_paiement_combo.bind("<KeyRelease>", self._on_compte_paiement_keyrelease)
        self.compte_paiement_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_compte_paiement_values()
        self.paiement_statut_var = tk.StringVar(
            value="✓ Paiement déjà comptabilisé" if reg.get("paiement_comptabilise") else "Paiement non encore comptabilisé")
        ttk.Label(paiement_frame, textvariable=self.paiement_statut_var).grid(
            row=0, column=4, sticky="w", padx=(12, 4))
        self.paiement_btn = ttk.Button(paiement_frame, text="Enregistrer le paiement (comptabilise)",
                                        command=self.enregistrer_paiement)
        self.paiement_btn.grid(row=1, column=0, columnspan=2, sticky="w", padx=4, pady=(4, 6))
        self.annuler_paiement_btn = ttk.Button(paiement_frame, text="Annuler le paiement comptabilisé",
                                                command=self.annuler_paiement)
        self.annuler_paiement_btn.grid(row=1, column=2, columnspan=2, sticky="w", padx=4, pady=(4, 6))

        origine = "" if not reg["bon_commande_id"] else f"Issu du bon de commande n° {reg['bon_commande_id']}"
        self.statut_var = tk.StringVar(
            value=f"Statut : {'VALIDÉ (comptabilisé)' if self.validee else 'Brouillon — à compléter'}   {origine}")
        ttk.Label(header, textvariable=self.statut_var, font=("Segoe UI", 10, "bold")).grid(
            row=3, column=0, columnspan=6, sticky="w", padx=4, pady=(4, 0))

        lignes_frame = ttk.LabelFrame(self, text=(
            "Lignes — sélectionnez une ligne pour lui affecter un compte de charge et un code analytique"))
        lignes_frame.pack(fill="both", padx=10, pady=6)

        form = ttk.Frame(lignes_frame)
        form.pack(fill="x", padx=6, pady=4)
        ttk.Label(form, text="Compte débiteur (charge ou immobilisation) :").grid(row=0, column=0, sticky="w")
        self.ligne_compte_var = tk.StringVar()
        self.ligne_compte_combo = ttk.Combobox(form, textvariable=self.ligne_compte_var, width=30)
        self.ligne_compte_combo.grid(row=0, column=1, padx=4)
        self.ligne_compte_combo.bind("<KeyRelease>", self._on_ligne_compte_keyrelease)
        self.ligne_compte_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_ligne_compte_values()
        ttk.Label(form, text="Code analytique :").grid(row=0, column=2, sticky="w", padx=(12, 0))
        self.ligne_analytic_var = tk.StringVar()
        self.ligne_analytic_combo = ttk.Combobox(form, textvariable=self.ligne_analytic_var, width=22)
        self.ligne_analytic_combo.grid(row=0, column=3, padx=4)
        self.ligne_analytic_combo.bind("<Button-1>", self._open_dropdown)
        self._refresh_ligne_analytic_values()

        ttk.Label(form, text="Libellé :").grid(row=1, column=0, sticky="w", pady=(4, 0))
        self.ligne_libelle_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_libelle_var, width=32).grid(row=1, column=1, padx=4, pady=(4, 0))
        ttk.Label(form, text="Quantité :").grid(row=1, column=2, sticky="w", padx=(12, 0), pady=(4, 0))
        self.ligne_qte_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_qte_var, width=10).grid(row=1, column=3, padx=4, pady=(4, 0), sticky="w")
        ttk.Label(form, text="Prix unitaire :").grid(row=1, column=4, sticky="w", padx=(12, 0), pady=(4, 0))
        self.ligne_prix_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.ligne_prix_var, width=12).grid(row=1, column=5, padx=4, pady=(4, 0), sticky="w")

        btn_row = ttk.Frame(lignes_frame)
        btn_row.pack(fill="x", padx=6)
        self.update_ligne_btn = ttk.Button(btn_row, text="Mettre à jour la ligne sélectionnée",
                                            command=self.update_ligne)
        self.update_ligne_btn.pack(side="left")
        self.add_ligne_btn = ttk.Button(btn_row, text="Ajouter une nouvelle ligne", command=self.add_ligne)
        self.add_ligne_btn.pack(side="left", padx=8)
        self.delete_ligne_btn = ttk.Button(btn_row, text="Supprimer la ligne sélectionnée", command=self.delete_ligne)
        self.delete_ligne_btn.pack(side="left", padx=8)

        cols = ("id", "compte", "libelle", "qte", "prix", "montant", "analytique")
        self.tree = ttk.Treeview(lignes_frame, columns=cols, show="headings", height=10)
        headers = ["ID", "Compte", "Libellé", "Qté", "Prix unit.", "Montant HT", "Code analytique"]
        widths = [40, 90, 260, 70, 100, 110, 150]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", padx=6, pady=6)
        self.tree.bind("<<TreeviewSelect>>", self._on_select_ligne)

        self.total_var = tk.StringVar()
        ttk.Label(lignes_frame, textvariable=self.total_var, font=("Segoe UI", 10, "bold")).pack(
            anchor="w", padx=6, pady=(0, 6))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=10, pady=10)
        self.save_btn = ttk.Button(btns, text="Enregistrer", command=self.save)
        self.save_btn.pack(side="left", padx=4)
        self.valider_btn = ttk.Button(btns, text="Valider (comptabiliser)", command=self.valider)
        self.valider_btn.pack(side="left", padx=4)
        self.corriger_btn = ttk.Button(btns, text="Corriger (repasser en brouillon)", command=self.corriger)
        self.corriger_btn.pack(side="left", padx=4)
        ttk.Button(btns, text="Fermer", command=self.destroy).pack(side="left", padx=4)

        self.refresh_lignes()
        self._apply_lock()

    @staticmethod
    def _open_dropdown(event=None):
        if event is not None:
            event.widget.event_generate("<Down>")

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def _refresh_fournisseur_values(self):
        self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in core.list_fournisseurs(self.conn)]

    def _on_fournisseur_keyrelease(self, event=None):
        items = core.list_fournisseurs(self.conn, self._extract_code(self.fournisseur_var.get()))
        self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _refresh_retenue_compte_values(self):
        items = [a for a in core.search_accounts(self.conn, "44", limit=100) if core.account_racine(a["code"]) == "44"]
        self.retenue_compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _on_retenue_compte_keyrelease(self, event=None):
        query = self._extract_code(self.retenue_compte_var.get())
        items = [a for a in core.search_accounts(self.conn, query, limit=100) if core.account_racine(a["code"]) == "44"]
        self.retenue_compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _refresh_retenue_presets(self):
        presets = core.list_taux_retenue(self.conn)
        self.retenue_preset_combo["values"] = [f"{p['label']} ({p['montant']:g}%)" for p in presets]
        self._retenue_presets = presets

    def _on_retenue_preset_selected(self, event=None):
        idx = self.retenue_preset_combo.current()
        if idx is not None and 0 <= idx < len(getattr(self, "_retenue_presets", [])):
            preset = self._retenue_presets[idx]
            self.retenue_taux_var.set(str(preset["montant"]))
            if preset.get("compte"):
                self.retenue_compte_var.set(preset["compte"])

    def _refresh_ligne_compte_values(self):
        items = [a for a in core.search_accounts(self.conn, "", limit=150) if a["classe"] in ("2", "6")]
        self.ligne_compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _on_ligne_compte_keyrelease(self, event=None):
        query = self._extract_code(self.ligne_compte_var.get())
        items = [a for a in core.search_accounts(self.conn, query, limit=50) if a["classe"] in ("2", "6")]
        self.ligne_compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _refresh_ligne_analytic_values(self):
        self.ligne_analytic_combo["values"] = [f"{c['code']} — {c['label']}" for c in core.list_analytic_codes(self.conn)]

    def _apply_lock(self):
        state = "disabled" if self.validee else "normal"
        for w in (self.update_ligne_btn, self.add_ligne_btn, self.delete_ligne_btn, self.save_btn, self.valider_btn):
            w.configure(state=state)
        self.corriger_btn.configure(state="normal" if self.validee else "disabled")
        reg = core.get_reglement(self.conn, self.reglement_id)
        deja_paye = bool(reg["paiement_comptabilise"])
        self.paiement_btn.configure(state="normal" if (self.validee and not deja_paye) else "disabled")
        self.annuler_paiement_btn.configure(state="normal" if deja_paye else "disabled")

    def _refresh_compte_paiement_values(self):
        items = [a for a in core.search_accounts(self.conn, "", limit=200) if a["classe"] == "5"]
        self.compte_paiement_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _on_compte_paiement_keyrelease(self, event=None):
        query = self._extract_code(self.compte_paiement_var.get())
        items = [a for a in core.search_accounts(self.conn, query, limit=50) if a["classe"] == "5"]
        self.compte_paiement_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def enregistrer_paiement(self):
        date_paiement = core.to_iso_date(self.date_paiement_var.get().strip())
        if not date_paiement:
            messagebox.showwarning("Champ manquant", "Saisissez la date de paiement.", parent=self)
            return
        compte = self._extract_code(self.compte_paiement_var.get())
        if not compte:
            messagebox.showwarning("Champ manquant", "Choisissez le compte banque ou caisse.", parent=self)
            return
        try:
            montant = core.enregistrer_paiement_reglement(self.conn, self.reglement_id, date_paiement, compte)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc), parent=self)
            return
        self.paiement_statut_var.set("✓ Paiement déjà comptabilisé")
        self._apply_lock()
        messagebox.showinfo("Paiement comptabilisé",
                             f"Paiement de {fmt_cfa(montant)} comptabilisé (Débit fournisseur, Crédit banque/caisse).",
                             parent=self)
        self.on_saved()

    def annuler_paiement(self):
        if not messagebox.askyesno("Annuler ce paiement",
                                    "Le paiement déjà comptabilisé va être retiré de la Saisie. Continuer ?",
                                    parent=self):
            return
        try:
            core.devalider_paiement_reglement(self.conn, self.reglement_id)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc), parent=self)
            return
        self.paiement_statut_var.set("Paiement non encore comptabilisé")
        self._apply_lock()
        messagebox.showinfo("Paiement annulé", "Le paiement a été retiré de la Saisie.", parent=self)
        self.on_saved()

    def _on_select_ligne(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        values = self.tree.item(sel[0], "values")
        self.selected_ligne_id = values[0]
        self.ligne_compte_var.set(values[1])
        self.ligne_libelle_var.set(values[2])
        self.ligne_qte_var.set(values[3])
        self.ligne_prix_var.set(values[4])
        self.ligne_analytic_var.set(values[6])

    def refresh_lignes(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        lignes = core.list_lignes_reglement(self.conn, self.reglement_id)
        total = sum(l["montant_ht"] for l in lignes)
        for l in lignes:
            self.tree.insert("", "end", values=(
                l["id"], l["compte_charge"] or "⚠ à choisir", l["libelle"], f"{l['quantite']:g}",
                f"{fmt_cfa(l['prix_unitaire'])}", f"{fmt_cfa(l['montant_ht'])}", l["analytic_code"] or ""))
        self.total_var.set(f"Total HT : {fmt_cfa(total)}")

    def add_ligne(self):
        libelle = self.ligne_libelle_var.get().strip()
        if not libelle:
            messagebox.showwarning("Champ manquant", "Le libellé est obligatoire.", parent=self)
            return
        try:
            qte = float(self.ligne_qte_var.get() or 0)
            prix = float(self.ligne_prix_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Quantité et Prix unitaire doivent être des nombres.", parent=self)
            return
        compte = self._extract_code(self.ligne_compte_var.get()) or None
        analytic = self._extract_code(self.ligne_analytic_var.get()) or None
        core.add_ligne_reglement(self.conn, self.reglement_id, compte, libelle, qte, prix_unitaire=prix,
                                  analytic_code=analytic)
        self._clear_ligne_form()
        self.refresh_lignes()

    def update_ligne(self):
        if not self.selected_ligne_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne dans le tableau.", parent=self)
            return
        libelle = self.ligne_libelle_var.get().strip()
        if not libelle:
            messagebox.showwarning("Champ manquant", "Le libellé est obligatoire.", parent=self)
            return
        try:
            qte = float(self.ligne_qte_var.get() or 0)
            prix = float(self.ligne_prix_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Quantité et Prix unitaire doivent être des nombres.", parent=self)
            return
        compte = self._extract_code(self.ligne_compte_var.get()) or None
        analytic = self._extract_code(self.ligne_analytic_var.get()) or None
        core.update_ligne_reglement(self.conn, self.selected_ligne_id, compte_charge=compte, libelle=libelle,
                                     quantite=qte, prix_unitaire=prix, analytic_code=analytic)
        self._clear_ligne_form()
        self.refresh_lignes()

    def delete_ligne(self):
        if not self.selected_ligne_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne dans le tableau.", parent=self)
            return
        core.delete_ligne_reglement(self.conn, self.selected_ligne_id)
        self._clear_ligne_form()
        self.refresh_lignes()

    def _clear_ligne_form(self):
        self.selected_ligne_id = None
        self.ligne_compte_var.set("")
        self.ligne_libelle_var.set("")
        self.ligne_qte_var.set("")
        self.ligne_prix_var.set("")
        self.ligne_analytic_var.set("")

    def save(self):
        date_str = core.to_iso_date(self.date_var.get().strip())
        try:
            retenue_taux = float(self.retenue_taux_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Le taux de retenue doit être un nombre.", parent=self)
            return
        core.update_reglement(
            self.conn, self.reglement_id,
            numero=self.numero_var.get().strip(), date_reglement=date_str,
            fournisseur_code=self._extract_code(self.fournisseur_var.get()),
            retenue_taux=retenue_taux, retenue_compte=self.retenue_compte_var.get().strip(),
        )
        messagebox.showinfo("Enregistré", "Règlement enregistré (brouillon).", parent=self)
        self.on_saved()

    def valider(self):
        self.save()
        try:
            core.valider_reglement(self.conn, self.reglement_id)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc), parent=self)
            return
        messagebox.showinfo("Comptabilisé", "Règlement validé et envoyé en Saisie.", parent=self)
        self.on_saved()
        self.destroy()

    def corriger(self):
        if not messagebox.askyesno(
            "Corriger ce règlement",
            "Ce règlement est déjà validé : ses écritures comptables vont être RETIRÉES de la Saisie et "
            "il repassera en brouillon modifiable.\n\nContinuer ?",
            parent=self,
        ):
            return
        try:
            core.devalider_reglement(self.conn, self.reglement_id)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc), parent=self)
            return
        self.validee = False
        self._apply_lock()
        self.refresh_lignes()
        messagebox.showinfo("Repassé en brouillon", "Le règlement est de nouveau modifiable.", parent=self)
        self.on_saved()


class ReglementTab(ttk.Frame):
    """Liste des Règlements (menu ENGAGEMENTS-PROJETS) — créés automatiquement
    en validant un Bon de commande du circuit interne. Double-clic pour
    ouvrir, choisir un compte de charge et un code analytique par ligne,
    une retenue fiscale, puis valider : C'EST ICI que la comptabilisation
    a réellement lieu (écriture envoyée en Saisie)."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="RÈGLEMENTS", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Créés automatiquement en validant un Bon de commande (circuit interne). Double-cliquez sur "
            "une ligne pour l'ouvrir : choisissez un compte de charge et un code analytique pour chaque "
            "ligne, une retenue fiscale si applicable, puis validez — c'est cette étape, et elle seule, "
            "qui envoie l'écriture comptable en Saisie."
        ), foreground="#595959", wraplength=1100).pack(anchor="w", padx=16, pady=(0, 8))

        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill="x", padx=16, pady=4)
        ttk.Button(btn_bar, text="Actualiser", command=self.refresh).pack(side="left")

        cols = ("numero", "date", "fournisseur", "nb_lignes", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=22)
        headers = ["N°", "Date", "Fournisseur", "Lignes", "Statut"]
        widths = [110, 100, 260, 70, 160]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<Double-1>", self._on_double_click)
        self._by_iid = {}
        self.refresh()

    def _on_double_click(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        rid = self._by_iid.get(sel[0])
        if rid:
            ReglementDialog(self, self.conn, rid, on_saved=self.refresh)

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self._by_iid = {}
        for reg in core.list_reglements(self.conn):
            nb = len(core.list_lignes_reglement(self.conn, reg["id"]))
            statut = "Validé" if reg["statut"] == "validee" else "Brouillon — à compléter"
            iid = self.tree.insert("", "end", values=(
                reg["numero"], core.to_display_date(reg["date_reglement"]), reg["raison_sociale"], nb, statut,
            ))
            self._by_iid[iid] = reg["id"]


class ContratsTab(ttk.Frame):
    """Journal des commandes/contrats fournisseurs : délais de paiement et de
    livraison, avec détection des dépassements."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="CONTRATS FOURNISSEURS — SUIVI DES DÉLAIS",
                  font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Enregistrez ici chaque commande/contrat passé avec un fournisseur. Les échéances de "
            "livraison et de paiement sont calculées automatiquement à partir des délais par défaut "
            "du fournisseur (modifiables dans l'onglet Fournisseurs), à la date de commande. "
            "Renseignez ensuite les dates réelles de livraison/paiement au fur et à mesure — les "
            "dépassements sont signalés automatiquement."
        ), foreground="#595959", wraplength=1050).pack(anchor="w", padx=16, pady=(0, 8))

        form = ttk.LabelFrame(self, text="Nouvelle commande / contrat")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Fournisseur :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.fournisseur_var = tk.StringVar()
        self.fournisseur_combo = ttk.Combobox(form, textvariable=self.fournisseur_var, width=28)
        self.fournisseur_combo.grid(row=0, column=1, padx=4)
        self.fournisseur_combo.bind("<KeyRelease>", self._on_fournisseur_keyrelease)
        self._refresh_fournisseur_values()

        ttk.Label(form, text="N° Pièce :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.piece_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.piece_var, width=14).grid(row=0, column=3, padx=4)

        ttk.Label(form, text="Libellé :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.libelle_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.libelle_var, width=26).grid(row=0, column=5, padx=4)

        ttk.Label(form, text="Montant :").grid(row=1, column=0, sticky="w", padx=4, pady=4)
        self.montant_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.montant_var, width=16).grid(row=1, column=1, padx=4)

        ttk.Label(form, text="Date commande (JJ/MM/AAAA) :").grid(row=1, column=2, sticky="w", padx=(12, 4))
        self.date_commande_var = tk.StringVar(value=date.today().strftime("%d/%m/%Y"))
        ttk.Entry(form, textvariable=self.date_commande_var, width=14).grid(row=1, column=3, padx=4)

        ttk.Button(form, text="Créer la commande (échéances auto)", command=self.add_commande).grid(
            row=1, column=4, columnspan=2, sticky="w", padx=12, pady=4)

        update_frame = ttk.LabelFrame(self, text="Mettre à jour la commande sélectionnée (dates réelles)")
        update_frame.pack(fill="x", padx=16, pady=(8, 4))
        ttk.Label(update_frame, text="Date livraison réelle (JJ/MM/AAAA) :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.livraison_reelle_var = tk.StringVar()
        ttk.Entry(update_frame, textvariable=self.livraison_reelle_var, width=14).grid(row=0, column=1, padx=4)
        ttk.Button(update_frame, text="Enregistrer la livraison", command=self.save_livraison).grid(
            row=0, column=2, padx=8)
        ttk.Label(update_frame, text="Date paiement réel (JJ/MM/AAAA) :").grid(row=0, column=3, sticky="w", padx=(20, 4))
        self.paiement_reel_var = tk.StringVar()
        ttk.Entry(update_frame, textvariable=self.paiement_reel_var, width=14).grid(row=0, column=4, padx=4)
        ttk.Button(update_frame, text="Enregistrer le paiement", command=self.save_paiement).grid(
            row=0, column=5, padx=8)
        ttk.Button(update_frame, text="Supprimer la commande sélectionnée", command=self.delete_commande).grid(
            row=0, column=6, padx=20)

        cols = ("id", "fournisseur", "piece", "libelle", "montant", "date_commande",
                "livraison_prevue", "statut_livraison", "echeance_paiement", "statut_paiement")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        headers = ["ID", "Fournisseur", "Pièce", "Libellé", "Montant", "Date commande",
                   "Livraison prévue", "Statut livraison", "Échéance paiement", "Statut paiement"]
        widths = [40, 160, 80, 160, 100, 100, 100, 140, 110, 140]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.tag_configure("depasse", foreground="#B00020")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.selected_id = None
        self.refresh()

    def _refresh_fournisseur_values(self):
        items = core.list_fournisseurs(self.conn)
        self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    def _on_fournisseur_keyrelease(self, event=None):
        query = self._extract_code(self.fournisseur_var.get())
        if query:
            items = core.list_fournisseurs(self.conn, query)
            self.fournisseur_combo["values"] = [f"{f['code']} — {f['raison_sociale']}" for f in items]

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        values = self.tree.item(sel[0], "values")
        self.selected_id = int(values[0])

    def add_commande(self):
        code = self._extract_code(self.fournisseur_var.get())
        if not code:
            messagebox.showwarning("Champ manquant", "Choisissez un fournisseur.")
            return
        if not core.fournisseur_exists(self.conn, code):
            messagebox.showerror("Fournisseur invalide", f"Le fournisseur « {code} » n'existe pas. "
                                                           f"Créez-le d'abord dans l'onglet Fournisseurs.")
            return
        try:
            montant = float(self.montant_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Le montant doit être un nombre.")
            return
        date_commande = core.to_iso_date(self.date_commande_var.get().strip())
        if not date_commande:
            messagebox.showwarning("Champ manquant", "La date de commande est obligatoire.")
            return
        core.add_commande(self.conn, code, self.piece_var.get().strip(), self.libelle_var.get().strip(),
                           montant, date_commande)
        self.piece_var.set("")
        self.libelle_var.set("")
        self.montant_var.set("")
        self.refresh()

    def save_livraison(self):
        if self.selected_id is None:
            messagebox.showinfo("Info", "Sélectionnez d'abord une commande dans le tableau.")
            return
        d = core.to_iso_date(self.livraison_reelle_var.get().strip())
        if not d:
            messagebox.showwarning("Champ manquant", "Saisissez la date de livraison réelle.")
            return
        core.update_commande(self.conn, self.selected_id, date_livraison_reelle=d)
        self.livraison_reelle_var.set("")
        self.refresh()

    def save_paiement(self):
        if self.selected_id is None:
            messagebox.showinfo("Info", "Sélectionnez d'abord une commande dans le tableau.")
            return
        d = core.to_iso_date(self.paiement_reel_var.get().strip())
        if not d:
            messagebox.showwarning("Champ manquant", "Saisissez la date de paiement réel.")
            return
        core.update_commande(self.conn, self.selected_id, date_paiement_reel=d)
        self.paiement_reel_var.set("")
        self.refresh()

    def delete_commande(self):
        if self.selected_id is None:
            messagebox.showinfo("Info", "Sélectionnez d'abord une commande.")
            return
        if messagebox.askyesno("Confirmer", "Supprimer cette commande ?"):
            core.delete_commande(self.conn, self.selected_id)
            self.selected_id = None
            self.refresh()

    def refresh(self):
        self._refresh_fournisseur_values()
        for row in self.tree.get_children():
            self.tree.delete(row)
        for c in core.list_commandes(self.conn):
            tags = ("depasse",) if (c["depassement_livraison"] or c["depassement_paiement"]) else ()
            self.tree.insert("", "end", tags=tags, values=(
                c["id"], c["raison_sociale"], c["piece"] or "", c["libelle"] or "",
                f"{fmt_cfa(c['montant'])}", core.to_display_date(c["date_commande"]),
                core.to_display_date(c["date_livraison_prevue"]), c["statut_livraison"],
                core.to_display_date(c["date_echeance_paiement"]), c["statut_paiement"],
            ))


class ExercicesTab(ttk.Frame):
    """Liste des exercices comptables et clôture annuelle."""

    def __init__(self, parent, conn, app):
        super().__init__(parent)
        self.conn = conn
        self.app = app
        ttk.Label(self, text="EXERCICES COMPTABLES", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "La clôture calcule le solde de clôture de chaque compte de bilan (classes 1 à 5) de "
            "l'exercice sélectionné, l'intègre au résultat net dans le compte 121000 (Report à "
            "nouveau créditeur), et reporte le tout comme solde d'ouverture de l'exercice suivant. "
            "Un exercice clôturé passe en lecture seule : impossible d'y ajouter, modifier ou "
            "supprimer une écriture."
        ), foreground="#595959", wraplength=1000).pack(anchor="w", padx=16, pady=(0, 8))

        cols = ("exercice", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=10)
        self.tree.heading("exercice", text="Exercice")
        self.tree.heading("statut", text="Statut")
        self.tree.column("exercice", width=100, anchor="w")
        self.tree.column("statut", width=150, anchor="w")
        self.tree.pack(fill="x", padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        self.selected_exercice = None
        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=4)
        ttk.Button(btns, text="Basculer sur cet exercice", command=self._switch).pack(side="left", padx=2)
        ttk.Button(btns, text="Clôturer l'exercice sélectionné", command=self._close).pack(side="left", padx=2)

        self.refresh()

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if sel:
            self.selected_exercice = self.tree.item(sel[0], "values")[0]

    def _switch(self):
        if not self.selected_exercice:
            messagebox.showinfo("Info", "Sélectionnez d'abord un exercice.")
            return
        core.set_current_exercice(self.conn, self.selected_exercice)
        self.app._refresh_exercice_list()
        self.app.refresh_current_page()

    def _close(self):
        if not self.selected_exercice:
            messagebox.showinfo("Info", "Sélectionnez d'abord un exercice.")
            return
        ex = self.selected_exercice
        if core.is_exercice_cloture(self.conn, ex):
            messagebox.showinfo("Info", f"L'exercice {ex} est déjà clôturé.")
            return
        bilan = core.compute_bilan(self.conn, exercice=ex)
        if abs(bilan["ecart"]) >= 1:
            if not messagebox.askyesno(
                "Bilan non équilibré",
                f"Le Bilan de l'exercice {ex} n'est pas équilibré (écart de {fmt_cfa(bilan['ecart'])}). "
                f"Clôturer quand même ?"
            ):
                return
        resultat_net = bilan['passif']["Résultat net de l'exercice"]
        if not messagebox.askyesno(
            "Confirmer la clôture",
            f"Clôturer définitivement l'exercice {ex} ?\n\n"
            f"Résultat net : {fmt_cfa(resultat_net)}\n"
            f"Cette action reporte les soldes de clôture comme soldes d'ouverture de l'exercice "
            f"suivant et verrouille l'exercice {ex} en lecture seule."
        ):
            return
        try:
            next_ex = core.close_exercice(self.conn, ex)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc))
            return
        messagebox.showinfo("Clôture effectuée",
                             f"Exercice {ex} clôturé. Les soldes d'ouverture de {next_ex} ont été calculés.")
        core.set_current_exercice(self.conn, next_ex)
        self.app._refresh_exercice_list()
        self.app.refresh_current_page()
        self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for e in core.list_exercices(self.conn):
            statut = "Clôturé" if e["cloture"] else "Ouvert"
            self.tree.insert("", "end", values=(e["exercice"], statut))


class PlanComptableTab(ttk.Frame):
    """Créer / modifier / supprimer des comptes du Plan comptable."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="PLAN COMPTABLE", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Chaque compte est rattaché à une racine : 1 chiffre pour les classes 1, 2, 3, 5, 6, 7, 8, 9 ; "
            "2 chiffres pour la classe 4 (comptes de tiers), qui se subdivise en 40 (Fournisseurs), "
            "41 (Clients), 42 (Personnel), 43 (Organismes sociaux), 44 (État), 45 (Organismes "
            "internationaux), 46 (Associés/Groupe), 47 (Débiteurs/créditeurs divers), 48 "
            "(Régularisations), 49 (Dépréciations sur tiers). Les fiches auxiliaires créées dans "
            "l'onglet Fournisseurs sont rattachées à la racine 40, celles de l'onglet Clients à la "
            "racine 41 — c'est ce qui permet au Bilan de classer correctement les créances et les dettes."
        ), foreground="#595959", wraplength=1050).pack(anchor="w", padx=16, pady=(0, 8))

        search_bar = ttk.Frame(self)
        search_bar.pack(fill="x", padx=16, pady=4)
        ttk.Label(search_bar, text="Rechercher (code ou libellé) :").pack(side="left")
        self.search_var = tk.StringVar()
        search_entry = ttk.Entry(search_bar, textvariable=self.search_var, width=30)
        search_entry.pack(side="left", padx=6)
        search_entry.bind("<KeyRelease>", lambda e: self.refresh())

        import_bar = ttk.Frame(self)
        import_bar.pack(fill="x", padx=16, pady=4)
        ttk.Button(import_bar, text="Importer un plan (.xlsx) — ÉCRASE le plan actuel",
                   command=self.import_xlsx).pack(side="left", padx=2)
        ttk.Button(import_bar, text="Exporter le plan actuel (.xlsx)", command=self.export_xlsx).pack(side="left", padx=2)

        form = ttk.Frame(self)
        form.pack(fill="x", padx=16, pady=6)
        ttk.Label(form, text="N° Compte :").grid(row=0, column=0, sticky="w")
        self.code_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.code_var, width=16).grid(row=0, column=1, padx=6)
        ttk.Label(form, text="Libellé :").grid(row=0, column=2, sticky="w", padx=(16, 0))
        self.label_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.label_var, width=45).grid(row=0, column=3, padx=6)
        ttk.Button(form, text="Créer / Modifier", command=self.save).grid(row=0, column=4, padx=6)
        ttk.Button(form, text="Supprimer le compte sélectionné", command=self.delete).grid(row=0, column=5, padx=6)

        cols = ("code", "label", "classe", "racine", "racine_label")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        headers = ["N° Compte", "Libellé", "Classe", "Racine", "Libellé de la racine"]
        widths = [100, 420, 60, 70, 260]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        values = self.tree.item(sel[0], "values")
        self.code_var.set(values[0])
        self.label_var.set(values[1])

    def save(self):
        code = self.code_var.get().strip()
        label = self.label_var.get().strip()
        if not code or not label:
            messagebox.showwarning("Champs manquants", "N° Compte et Libellé sont obligatoires.")
            return
        core.add_account(self.conn, code, label)
        self.refresh()

    def delete(self):
        code = self.code_var.get().strip()
        if not code:
            messagebox.showinfo("Info", "Sélectionnez d'abord un compte.")
            return
        if messagebox.askyesno("Confirmer", f"Supprimer le compte {code} ?"):
            core.delete_account(self.conn, code)
            self.code_var.set("")
            self.label_var.set("")
            self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for a in core.search_accounts(self.conn, self.search_var.get(), limit=200):
            racine = core.account_racine(a["code"])
            racine_label = core.RACINE_LABELS.get(racine, "")
            self.tree.insert("", "end", values=(a["code"], a["label"], a["classe"], racine, racine_label))

    def export_xlsx(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Plan_comptable.xlsx", title="Exporter le Plan comptable",
        )
        if not path:
            return
        core.export_plan_comptable_xlsx(self.conn, path)
        messagebox.showinfo("Export terminé", f"Plan comptable exporté :\n{path}")

    def import_xlsx(self):
        path = filedialog.askopenfilename(filetypes=[("Classeur Excel", "*.xlsx")],
                                           title="Importer un Plan comptable")
        if not path:
            return
        if not messagebox.askyesno(
            "Confirmer l'écrasement",
            "Importer ce fichier va ÉCRASER complètement le Plan comptable actuel (tous les comptes "
            "existants seront supprimés et remplacés par ceux du fichier). Cette action est "
            "irréversible. Continuer ?"
        ):
            return
        try:
            n = core.import_plan_comptable_xlsx(self.conn, path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de l'import : {exc}")
            return
        self.refresh()
        messagebox.showinfo("Import terminé", f"{n} compte(s) importé(s). Le plan précédent a été remplacé.")


class _SimplePlanTab(ttk.Frame):
    """Base pour les plans Code + Libellé (analytique, bailleurs, taux)."""
    TITLE = ""
    CODE_LABEL = "Code"
    HAS_UNITE = False  # PlanAnalytiqueTab l'active pour gérer L / Kw / H...
    HAS_TAUX = False    # TauxTVATab/TauxRetenueTab l'activent pour gérer un taux (%)
    SUGGESTIONS_FN = None     # core.ajouter_xxx_suggeres(conn) -> nb ajoutés, si applicable
    SUGGESTIONS_LABEL = None  # texte du bouton, si SUGGESTIONS_FN est défini

    def list_fn(self, conn):
        raise NotImplementedError

    def add_fn(self, conn, code, label):
        raise NotImplementedError

    def delete_fn(self, conn, code):
        raise NotImplementedError

    def export_fn(self, conn, path):
        raise NotImplementedError

    def import_fn(self, conn, path):
        raise NotImplementedError

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text=self.TITLE, font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))

        import_bar = ttk.Frame(self)
        import_bar.pack(fill="x", padx=16, pady=(0, 4))
        ttk.Button(import_bar, text="Importer (.xlsx) — ÉCRASE le plan actuel",
                   command=self.import_xlsx).pack(side="left", padx=2)
        ttk.Button(import_bar, text="Exporter (.xlsx)", command=self.export_xlsx).pack(side="left", padx=2)
        if self.SUGGESTIONS_FN is not None:
            ttk.Button(import_bar, text=self.SUGGESTIONS_LABEL or "Ajouter les catégories courantes",
                       command=self._add_suggestions).pack(side="left", padx=8)

        form = ttk.Frame(self)
        form.pack(fill="x", padx=16, pady=6)
        ttk.Label(form, text=self.CODE_LABEL + " :").grid(row=0, column=0, sticky="w")
        self.code_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.code_var, width=20).grid(row=0, column=1, padx=6)
        ttk.Label(form, text="Libellé :").grid(row=0, column=2, sticky="w", padx=(16, 0))
        self.label_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.label_var, width=45).grid(row=0, column=3, padx=6)
        next_col = 4
        if self.HAS_UNITE:
            ttk.Label(form, text="Unité (L, Kw, H...) :").grid(row=0, column=4, sticky="w", padx=(16, 0))
            self.unite_var = tk.StringVar()
            ttk.Entry(form, textvariable=self.unite_var, width=10).grid(row=0, column=5, padx=6)
            next_col = 6
        if self.HAS_TAUX:
            ttk.Label(form, text="Taux (%) :").grid(row=0, column=4, sticky="w", padx=(16, 0))
            self.taux_var = tk.StringVar()
            ttk.Entry(form, textvariable=self.taux_var, width=10).grid(row=0, column=5, padx=6)
            ttk.Label(form, text="Compte fiscal (classe 44) :").grid(row=1, column=4, sticky="w", padx=(16, 0), pady=(4, 0))
            self.compte_var = tk.StringVar()
            self.compte_combo = ttk.Combobox(form, textvariable=self.compte_var, width=30)
            self.compte_combo.grid(row=1, column=5, padx=6, pady=(4, 0), sticky="w")
            self.compte_combo.bind("<KeyRelease>", self._on_compte_keyrelease)
            self.compte_combo.bind("<Button-1>", self._open_dropdown)
            self._refresh_compte_values()
            next_col = 6
        ttk.Button(form, text="Créer / Modifier", command=self.save).grid(row=0, column=next_col, padx=6)
        ttk.Button(form, text="Supprimer", command=self.delete).grid(row=0, column=next_col + 1, padx=6)

        cols = ("code", "label", "unite") if self.HAS_UNITE else (
            ("code", "label", "taux", "compte") if self.HAS_TAUX else ("code", "label"))
        headers = ([self.CODE_LABEL, "Libellé", "Unité"] if self.HAS_UNITE else
                   ([self.CODE_LABEL, "Libellé", "Taux (%)", "Compte fiscal"] if self.HAS_TAUX
                    else [self.CODE_LABEL, "Libellé"]))
        widths = [140, 460, 80] if self.HAS_UNITE else ([140, 380, 80, 220] if self.HAS_TAUX else [140, 500])
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    @staticmethod
    def _open_dropdown(event=None):
        if event is not None:
            event.widget.event_generate("<Down>")

    @staticmethod
    def _extract_code(raw):
        raw = (raw or "").strip()
        return raw.split(" — ", 1)[0].strip() if " — " in raw else raw

    def _refresh_compte_values(self):
        items = [a for a in core.search_accounts(self.conn, "44", limit=100) if core.account_racine(a["code"]) == "44"]
        self.compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _on_compte_keyrelease(self, event=None):
        query = self._extract_code(self.compte_var.get())
        items = [a for a in core.search_accounts(self.conn, query, limit=100) if core.account_racine(a["code"]) == "44"]
        self.compte_combo["values"] = [f"{a['code']} — {a['label']}" for a in items]

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        values = self.tree.item(sel[0], "values")
        self.code_var.set(values[0])
        self.label_var.set(values[1])
        if self.HAS_UNITE:
            self.unite_var.set(values[2] if len(values) > 2 else "")
        if self.HAS_TAUX:
            self.taux_var.set(values[2] if len(values) > 2 else "")
            self.compte_var.set(values[3] if len(values) > 3 else "")

    def save(self):
        code = self.code_var.get().strip()
        label = self.label_var.get().strip()
        if not code or not label:
            messagebox.showwarning("Champs manquants", f"{self.CODE_LABEL} et Libellé sont obligatoires.")
            return
        if self.HAS_UNITE:
            self.add_fn(self.conn, code, label, unite=self.unite_var.get().strip() or None)
        elif self.HAS_TAUX:
            try:
                taux = float(self.taux_var.get().strip() or 0)
            except ValueError:
                messagebox.showerror("Erreur", "Le taux doit être un nombre.")
                return
            compte = self._extract_code(self.compte_var.get()) or None
            self.add_fn(self.conn, code, label, montant=taux, compte=compte)
        else:
            self.add_fn(self.conn, code, label)
        self.refresh()

    def delete(self):
        code = self.code_var.get().strip()
        if not code:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne.")
            return
        if messagebox.askyesno("Confirmer", f"Supprimer « {code} » ?"):
            self.delete_fn(self.conn, code)
            self.code_var.set("")
            self.label_var.set("")
            if self.HAS_UNITE:
                self.unite_var.set("")
            if self.HAS_TAUX:
                self.taux_var.set("")
                self.compte_var.set("")
            self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for item in self.list_fn(self.conn):
            if self.HAS_UNITE:
                self.tree.insert("", "end", values=(item["code"], item["label"], item.get("unite") or ""))
            elif self.HAS_TAUX:
                self.tree.insert("", "end", values=(item["code"], item["label"], item.get("montant") or 0,
                                                      item.get("compte") or ""))
            else:
                self.tree.insert("", "end", values=(item["code"], item["label"]))

    def export_xlsx(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile=f"{self.TITLE.title().replace(' ', '_')}.xlsx", title=f"Exporter {self.TITLE}",
        )
        if not path:
            return
        self.export_fn(self.conn, path)
        messagebox.showinfo("Export terminé", f"Plan exporté :\n{path}")

    def import_xlsx(self):
        path = filedialog.askopenfilename(filetypes=[("Classeur Excel", "*.xlsx")],
                                           title=f"Importer {self.TITLE}")
        if not path:
            return
        if not messagebox.askyesno(
            "Confirmer l'écrasement",
            f"Importer ce fichier va ÉCRASER complètement le {self.TITLE.lower()} actuel. "
            f"Cette action est irréversible. Continuer ?"
        ):
            return
        try:
            n = self.import_fn(self.conn, path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de l'import : {exc}")
            return
        self.refresh()
        messagebox.showinfo("Import terminé", f"{n} ligne(s) importée(s). Le plan précédent a été remplacé.")

    def _add_suggestions(self):
        n = self.SUGGESTIONS_FN(self.conn)
        if n:
            messagebox.showinfo(
                "Catégories ajoutées",
                f"{n} catégorie(s) ajoutée(s), à 0% et sans compte — complétez le taux et le compte "
                f"fiscal exacts pour chacune avant utilisation."
            )
        else:
            messagebox.showinfo("Rien à ajouter", "Toutes les catégories courantes existent déjà.")
        self.refresh()


class PlanAnalytiqueTab(_SimplePlanTab):
    TITLE = "PLAN ANALYTIQUE"
    CODE_LABEL = "Code analytique"
    HAS_UNITE = True

    def list_fn(self, conn):
        return core.list_analytic_codes(conn)

    def add_fn(self, conn, code, label, unite=None):
        core.add_analytic_code(conn, code, label, unite=unite)

    def delete_fn(self, conn, code):
        core.delete_analytic_code(conn, code)

    def export_fn(self, conn, path):
        core.export_analytic_codes_xlsx(conn, path)

    def import_fn(self, conn, path):
        return core.import_analytic_codes_xlsx(conn, path)


class PlanBailleurTab(_SimplePlanTab):
    TITLE = "PLAN BAILLEURS DE FONDS"
    CODE_LABEL = "Code bailleur"

    def list_fn(self, conn):
        return core.list_donor_codes(conn)

    def add_fn(self, conn, code, label):
        core.add_donor_code(conn, code, label)

    def delete_fn(self, conn, code):
        core.delete_donor_code(conn, code)

    def export_fn(self, conn, path):
        core.export_donor_codes_xlsx(conn, path)

    def import_fn(self, conn, path):
        return core.import_donor_codes_xlsx(conn, path)


class TauxTVATab(_SimplePlanTab):
    """Taux de TVA paramétrables (menu ADMIN) — utilisés en préréglage dans
    l'onglet Facturation (COMMERCE), à la place d'un taux tapé à la main."""
    TITLE = "TAUX DE TVA"
    CODE_LABEL = "Code"
    HAS_TAUX = True

    def list_fn(self, conn):
        return core.list_taux_tva(conn)

    def add_fn(self, conn, code, label, montant=0, compte=None):
        core.add_taux_tva(conn, code, label, montant=montant, compte=compte)

    def delete_fn(self, conn, code):
        core.delete_taux_tva(conn, code)

    def export_fn(self, conn, path):
        core.export_taux_tva_xlsx(conn, path)

    def import_fn(self, conn, path):
        return core.import_taux_tva_xlsx(conn, path)


class TauxRetenueTab(_SimplePlanTab):
    """Taux de retenue à la source paramétrables (menu ADMIN) — utilisés en
    préréglage dans l'onglet Factures frs (ENGAGEMENTS-PROJETS), en plus du
    choix du compte de retenue (classe 44)."""
    TITLE = "TAUX DE RETENUE À LA SOURCE"
    CODE_LABEL = "Code"
    HAS_TAUX = True
    SUGGESTIONS_FN = staticmethod(core.ajouter_taux_retenue_suggeres)
    SUGGESTIONS_LABEL = "Ajouter les catégories courantes (BIC, IS, TVA...)"

    def list_fn(self, conn):
        return core.list_taux_retenue(conn)

    def add_fn(self, conn, code, label, montant=0, compte=None):
        core.add_taux_retenue(conn, code, label, montant=montant, compte=compte)

    def delete_fn(self, conn, code):
        core.delete_taux_retenue(conn, code)

    def export_fn(self, conn, path):
        core.export_taux_retenue_xlsx(conn, path)

    def import_fn(self, conn, path):
        return core.import_taux_retenue_xlsx(conn, path)


class SynchronisationTab(ttk.Frame):
    """Synchronisation de la base (menu PARAMÈTRES) — revérifie/répare la
    structure de toutes les tables (utile après une mise à jour du
    logiciel). Ne touche jamais aux données existantes."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="SYNCHRONISATION", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Revérifie et met à jour la structure de toutes les tables de la base (utile après avoir "
            "installé une nouvelle version du logiciel) — crée toute table ou colonne manquante, sans "
            "jamais modifier ou supprimer vos données existantes."
        ), foreground="#595959", wraplength=1000).pack(anchor="w", padx=16, pady=(0, 12))
        ttk.Button(self, text="Synchroniser maintenant", command=self.synchroniser).pack(anchor="w", padx=16)
        self.result_var = tk.StringVar()
        ttk.Label(self, textvariable=self.result_var, font=("Segoe UI", 10)).pack(
            anchor="w", padx=16, pady=(12, 0))

    def synchroniser(self):
        try:
            rapport = core.synchroniser_base(self.conn)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de la synchronisation : {exc}")
            return
        ecart = rapport["ecart_bilan"]
        ecart_txt = (f"écart Actif/Passif : {fmt_cfa(ecart)}" if ecart is not None
                     else "Bilan non calculable (base vide ?)")
        self.result_var.set(
            f"✓ Synchronisation terminée — {rapport['nb_tables']} tables vérifiées, "
            f"exercice courant : {rapport['exercice']}, {ecart_txt}."
        )
        messagebox.showinfo("Synchronisation terminée",
                             f"{rapport['nb_tables']} tables vérifiées et à jour.\nAucune donnée n'a été modifiée.")


class NiveauxAccesTab(ttk.Frame):
    """Niveaux d'accès paramétrables (menu ADMIN) — utilisés lors de la
    création d'un utilisateur, ET désormais pour restreindre les menus
    (et sous-menus) réellement accessibles à chaque niveau, via une
    case à cocher par sous-menu (voir core.MENU_STRUCTURE)."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.selected_niveau = None
        self.check_vars = {}  # {menu_key: tk.BooleanVar}

        ttk.Label(self, text="NIVEAUX D'ACCÈS", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Chaque niveau détermine les sous-menus visibles pour les utilisateurs qui lui sont rattachés — "
            "sélectionnez un niveau ci-dessous pour cocher/décocher ses autorisations. Le niveau "
            "« Administrateur » a toujours accès à tout, quelle que soit la configuration."
        ), foreground="#595959", wraplength=1200, justify="left").pack(anchor="w", padx=16, pady=(0, 8))

        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill="x", padx=16, pady=(0, 4))
        ttk.Button(btn_bar, text="Ajouter les niveaux courants (Administrateur, Comptable...)",
                   command=self._add_suggestions).pack(side="left")

        form = ttk.Frame(self)
        form.pack(fill="x", padx=16, pady=6)
        ttk.Label(form, text="Nom du niveau :").grid(row=0, column=0, sticky="w")
        self.code_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.code_var, width=24).grid(row=0, column=1, padx=6)
        ttk.Label(form, text="Description :").grid(row=0, column=2, sticky="w", padx=(16, 0))
        self.label_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.label_var, width=60).grid(row=0, column=3, padx=6)
        ttk.Button(form, text="Ajouter", command=self.add).grid(row=0, column=4, padx=8)
        ttk.Button(form, text="Mettre à jour la description", command=self.update_sel).grid(row=0, column=5, padx=4)
        ttk.Button(form, text="Supprimer", command=self.delete_sel).grid(row=0, column=6, padx=4)

        body = ttk.Frame(self)
        body.pack(fill="both", expand=True, padx=16, pady=(4, 16))
        body.columnconfigure(0, weight=1)
        body.columnconfigure(1, weight=2)
        body.rowconfigure(0, weight=1)

        list_frame = ttk.Frame(body)
        list_frame.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        cols = ("nom", "description")
        self.tree = ttk.Treeview(list_frame, columns=cols, show="headings", height=20)
        for c, h, w in zip(cols, ["Niveau", "Description"], [140, 260]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)

        self.menus_frame = ttk.LabelFrame(body, text="Sous-menus autorisés — sélectionnez un niveau à gauche")
        self.menus_frame.grid(row=0, column=1, sticky="nsew")
        self.menus_canvas = tk.Canvas(self.menus_frame, highlightthickness=0)
        menus_scroll = ttk.Scrollbar(self.menus_frame, orient="vertical", command=self.menus_canvas.yview)
        self.menus_inner = ttk.Frame(self.menus_canvas)
        self.menus_inner.bind("<Configure>", lambda e: self.menus_canvas.configure(
            scrollregion=self.menus_canvas.bbox("all")))
        self.menus_canvas.create_window((0, 0), window=self.menus_inner, anchor="nw")
        self.menus_canvas.configure(yscrollcommand=menus_scroll.set)
        self.menus_canvas.pack(side="left", fill="both", expand=True)
        menus_scroll.pack(side="right", fill="y")

        self.save_menus_btn = ttk.Button(self, text="Enregistrer les autorisations de ce niveau",
                                          command=self.save_menus, state="disabled")
        self.save_menus_btn.pack(anchor="w", padx=16, pady=(0, 16))

        self.refresh()

    def _add_suggestions(self):
        core.ajouter_niveaux_acces_suggeres(self.conn)
        core.ajouter_niveaux_acces_suggeres_menus(self.conn)
        self.refresh()
        messagebox.showinfo("Ajouté", "Niveaux courants et leurs autorisations par défaut ajoutés.", parent=self)

    def add(self):
        nom = self.code_var.get().strip()
        if not nom:
            messagebox.showwarning("Champ manquant", "Le nom du niveau est obligatoire.", parent=self)
            return
        core.add_niveau_acces(self.conn, nom, self.label_var.get().strip())
        self.code_var.set(""); self.label_var.set("")
        self.refresh()

    def update_sel(self):
        if not self.selected_niveau:
            messagebox.showinfo("Info", "Sélectionnez d'abord un niveau.", parent=self)
            return
        core.add_niveau_acces(self.conn, self.selected_niveau, self.label_var.get().strip())
        self.refresh()

    def delete_sel(self):
        if not self.selected_niveau:
            messagebox.showinfo("Info", "Sélectionnez d'abord un niveau.", parent=self)
            return
        if messagebox.askyesno("Confirmer", f"Supprimer le niveau « {self.selected_niveau} » ?", parent=self):
            core.delete_niveau_acces(self.conn, self.selected_niveau)
            self.selected_niveau = None
            self.code_var.set(""); self.label_var.set("")
            self.refresh()

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        values = self.tree.item(sel[0], "values")
        self.selected_niveau = values[0]
        self.code_var.set(values[0])
        self.label_var.set(values[1])
        self._build_menu_checklist()

    def _build_menu_checklist(self):
        for w in self.menus_inner.winfo_children():
            w.destroy()
        self.check_vars = {}

        if self.selected_niveau == "Administrateur":
            self.menus_frame.configure(text="Sous-menus autorisés — Administrateur a toujours accès à tout")
            ttk.Label(self.menus_inner, text="✓ Ce niveau a systématiquement accès à la totalité des menus — "
                                              "rien à configurer.", foreground="#1F7A1F").pack(anchor="w", padx=8, pady=8)
            self.save_menus_btn.configure(state="disabled")
            return

        self.menus_frame.configure(text=f"Sous-menus autorisés — {self.selected_niveau}")
        autorises = core.get_menus_autorises(self.conn, self.selected_niveau)
        for titre, items in core.MENU_STRUCTURE:
            ttk.Label(self.menus_inner, text=titre, font=("Segoe UI", 9, "bold")).pack(
                anchor="w", padx=8, pady=(8, 2))
            for label, key in items:
                if key in self.check_vars:
                    continue  # certaines clés sont partagées entre menus (ex. "stocks") — une seule case suffit
                var = tk.BooleanVar(value=(key in autorises))
                self.check_vars[key] = var
                ttk.Checkbutton(self.menus_inner, text=label, variable=var).pack(anchor="w", padx=24)
        self.save_menus_btn.configure(state="normal")

    def save_menus(self):
        if not self.selected_niveau or self.selected_niveau == "Administrateur":
            return
        menu_keys = [key for key, var in self.check_vars.items() if var.get()]
        core.set_menus_autorises(self.conn, self.selected_niveau, menu_keys)
        messagebox.showinfo("Enregistré", f"Autorisations mises à jour pour « {self.selected_niveau} ».",
                             parent=self)

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for n in core.list_niveaux_acces(self.conn):
            self.tree.insert("", "end", values=(n["nom"], n["description"] or ""))


class ReinitialisationTab(ttk.Frame):
    """Réinitialisation ciblée des données (menu ADMIN) — outil DESTRUCTIF
    et explicite. La Synchronisation (menu PARAMÈTRES) ne touche jamais aux
    données ; c'est ici, et seulement ici, qu'on peut vider des catégories
    de données choisies. Chaque catégorie est indépendante : supprimer les
    écritures comptables ne vide PAS automatiquement les soldes d'ouverture
    (table séparée — d'où le Stock/les Immobilisations qui semblent
    persister après une suppression des seules écritures)."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="RÉINITIALISATION DES DONNÉES", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))

        dossier_frame = ttk.Frame(self)
        dossier_frame.pack(fill="x", padx=16, pady=(0, 8))
        ttk.Button(dossier_frame, text="📁 Ouvrir le dossier de la base de données",
                   command=self._ouvrir_dossier_donnees).pack(side="left")
        ttk.Label(dossier_frame, text=core.default_db_path(), foreground="#595959").pack(
            side="left", padx=12)
        ttk.Separator(self).pack(fill="x", padx=16, pady=(0, 8))

        ttk.Label(self, text=(
            "⚠ Action destructive et irréversible. La Synchronisation (menu PARAMÈTRES) ne supprime "
            "JAMAIS de données — elle ne fait que réparer la structure des tables. Ici, chaque catégorie "
            "cochée sera VIDÉE définitivement. Les catégories sont indépendantes : par exemple, vider les "
            "« Écritures comptables » ne vide PAS les « Soldes d'ouverture » (table séparée) — c'est pour "
            "cela que le Stock et les Immobilisations peuvent sembler ne pas se vider après avoir "
            "supprimé uniquement les écritures : cochez aussi « Soldes d'ouverture »."
        ), foreground="#B00020", wraplength=1050).pack(anchor="w", padx=16, pady=(0, 12))

        exercice_bar = ttk.Frame(self)
        exercice_bar.pack(fill="x", padx=16, pady=4)
        ttk.Label(exercice_bar, text="Portée :").pack(side="left")
        self.portee_var = tk.StringVar(value="tout")
        ttk.Radiobutton(exercice_bar, text="Toutes les années", variable=self.portee_var,
                        value="tout").pack(side="left", padx=8)
        ttk.Radiobutton(exercice_bar, text="Exercice courant seulement :", variable=self.portee_var,
                        value="exercice").pack(side="left", padx=(16, 4))
        self.exercice_var = tk.StringVar(value=core.get_current_exercice(conn))
        ttk.Entry(exercice_bar, textvariable=self.exercice_var, width=8).pack(side="left")
        ttk.Label(exercice_bar, text="(s'applique aux écritures et soldes d'ouverture uniquement)",
                  foreground="#595959").pack(side="left", padx=8)

        self.vars = {}
        cats_frame = ttk.LabelFrame(self, text="Catégories à vider")
        cats_frame.pack(fill="x", padx=16, pady=8)
        for key, label in core.REINIT_CATEGORIES.items():
            var = tk.BooleanVar(value=False)
            self.vars[key] = var
            ttk.Checkbutton(cats_frame, text=label, variable=var).pack(anchor="w", padx=8, pady=2)

        confirm_frame = ttk.LabelFrame(self, text="Confirmation")
        confirm_frame.pack(fill="x", padx=16, pady=8)
        ttk.Label(confirm_frame, text='Tapez SUPPRIMER en majuscules pour activer le bouton :').pack(
            anchor="w", padx=8, pady=(6, 2))
        self.confirm_var = tk.StringVar()
        self.confirm_entry = ttk.Entry(confirm_frame, textvariable=self.confirm_var, width=20)
        self.confirm_entry.pack(anchor="w", padx=8, pady=(0, 8))

        self.reset_btn = ttk.Button(self, text="Réinitialiser les catégories cochées",
                                     command=self.reinitialiser, state="disabled")
        self.reset_btn.pack(anchor="w", padx=16, pady=(0, 8))
        self.confirm_var.trace_add("write", self._on_confirm_change)

        self.result_var = tk.StringVar()
        ttk.Label(self, textvariable=self.result_var, font=("Segoe UI", 10)).pack(
            anchor="w", padx=16, pady=(4, 16))

    def _on_confirm_change(self, *_):
        self.reset_btn.configure(state="normal" if self.confirm_var.get() == "SUPPRIMER" else "disabled")

    def _ouvrir_dossier_donnees(self):
        dossier = os.path.dirname(core.default_db_path())
        try:
            if sys.platform.startswith("win"):
                os.startfile(dossier)
            elif sys.platform == "darwin":
                subprocess.run(["open", dossier], check=False)
            else:
                subprocess.run(["xdg-open", dossier], check=False)
        except Exception as exc:
            messagebox.showerror(
                "Impossible d'ouvrir le dossier",
                f"Le dossier n'a pas pu être ouvert automatiquement :\n{exc}\n\nChemin exact :\n{dossier}",
                parent=self,
            )

    def reinitialiser(self):
        categories = {k for k, v in self.vars.items() if v.get()}
        if not categories:
            messagebox.showwarning("Aucune catégorie", "Cochez au moins une catégorie à vider.")
            return
        labels = "\n".join(f"• {core.REINIT_CATEGORIES[k]}" for k in categories)
        if not messagebox.askyesno(
            "Confirmation finale",
            f"Vous allez vider DÉFINITIVEMENT :\n\n{labels}\n\n"
            f"Cette action est IRRÉVERSIBLE. Continuer ?",
            icon="warning",
        ):
            return
        exercice = self.exercice_var.get().strip() if self.portee_var.get() == "exercice" else None
        rapport = core.reinitialiser_donnees(self.conn, categories, exercice=exercice)
        total = sum(rapport.values())
        detail = "\n".join(f"• {core.REINIT_CATEGORIES[k]} : {n} ligne(s)" for k, n in rapport.items())
        self.result_var.set(f"✓ Réinitialisation terminée — {total} ligne(s) supprimée(s) au total.")
        self.confirm_var.set("")
        messagebox.showinfo("Réinitialisation terminée", f"{detail}\n\nTotal : {total} ligne(s) supprimée(s).")


class UtilisateursTab(ttk.Frame):
    """Comptes utilisateurs et niveau d'accès (menu ADMIN) — pose la base
    (comptes, mots de passe hachés, niveaux) en vue d'un futur contrôle
    d'accès par écran/action ; l'application ne demande pas encore de
    connexion au démarrage."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        self.selected_id = None
        ttk.Label(self, text="UTILISATEURS", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Les niveaux d'accès se paramètrent dans le sous-menu « Niveaux d'accès » (modules "
            "autorisés par niveau). Dès qu'au moins un utilisateur existe ici, l'application demande "
            "une connexion au démarrage, et les menus affichés sont filtrés selon son niveau d'accès."
        ), foreground="#595959", wraplength=1000).pack(anchor="w", padx=16, pady=(0, 8))

        form = ttk.LabelFrame(self, text="Utilisateur")
        form.pack(fill="x", padx=16, pady=4)
        ttk.Label(form, text="Nom d'utilisateur :").grid(row=0, column=0, sticky="w", padx=4, pady=4)
        self.login_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.login_var, width=18).grid(row=0, column=1, padx=4)
        ttk.Label(form, text="Nom complet :").grid(row=0, column=2, sticky="w", padx=(12, 4))
        self.nom_complet_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.nom_complet_var, width=24).grid(row=0, column=3, padx=4)
        ttk.Label(form, text="Niveau d'accès :").grid(row=0, column=4, sticky="w", padx=(12, 4))
        self.niveau_var = tk.StringVar()
        self.niveau_combo = ttk.Combobox(form, textvariable=self.niveau_var, width=18, state="readonly")
        self.niveau_combo.grid(row=0, column=5, padx=4)
        self.niveau_combo.bind("<<ComboboxSelected>>", lambda e: self._refresh_apercu_modules())
        self._refresh_niveaux()
        ttk.Label(form, text="Mot de passe :").grid(row=1, column=0, sticky="w", padx=4, pady=(4, 0))
        self.password_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.password_var, width=18, show="•").grid(
            row=1, column=1, padx=4, pady=(4, 0))
        self.actif_var = tk.BooleanVar(value=True)
        ttk.Checkbutton(form, text="Actif", variable=self.actif_var).grid(
            row=1, column=2, sticky="w", padx=(12, 4), pady=(4, 0))
        ttk.Label(form, text="(laisser le mot de passe vide pour ne pas le changer, lors d'une mise à jour)",
                  foreground="#595959").grid(row=1, column=3, columnspan=3, sticky="w", padx=(12, 4), pady=(4, 0))

        apercu_frame = ttk.LabelFrame(self, text="Modules autorisés pour le niveau sélectionné")
        apercu_frame.pack(fill="x", padx=16, pady=(0, 4))
        self.apercu_var = tk.StringVar(value="Choisissez un niveau d'accès ci-dessus pour voir ses modules autorisés.")
        ttk.Label(apercu_frame, textvariable=self.apercu_var, foreground="#1F4E78", wraplength=1300,
                  justify="left").pack(anchor="w", padx=8, pady=6)
        ttk.Button(apercu_frame, text="Modifier les modules de ce niveau (ouvre Niveaux d'accès)",
                   command=self._ouvrir_niveaux_acces).pack(anchor="w", padx=8, pady=(0, 6))

        btns = ttk.Frame(self)
        btns.pack(fill="x", padx=16, pady=6)
        ttk.Button(btns, text="Créer l'utilisateur", command=self.add).pack(side="left")
        ttk.Button(btns, text="Mettre à jour la sélection", command=self.update_sel).pack(side="left", padx=8)
        ttk.Button(btns, text="Supprimer la sélection", command=self.delete_sel).pack(side="left")
        ttk.Button(btns, text="Effacer le formulaire", command=self.clear_form).pack(side="left", padx=8)

        cols = ("id", "login", "nom_complet", "niveau", "actif", "date_creation")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=16)
        headers = ["ID", "Utilisateur", "Nom complet", "Niveau d'accès", "Actif", "Créé le"]
        widths = [40, 140, 220, 160, 60, 100]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _refresh_niveaux(self):
        self.niveau_combo["values"] = [n["nom"] for n in core.list_niveaux_acces(self.conn)]

    def _refresh_apercu_modules(self):
        niveau = self.niveau_var.get().strip()
        if not niveau:
            self.apercu_var.set("Choisissez un niveau d'accès ci-dessus pour voir ses modules autorisés.")
            return
        menus = core.get_menus_autorises(self.conn, niveau)
        if niveau == "Administrateur":
            self.apercu_var.set("✓ Ce niveau a accès à la totalité des modules de l'application (48/48).")
            return
        if not menus:
            self.apercu_var.set(f"⚠ Aucun module autorisé pour « {niveau} » — cet utilisateur ne verra AUCUN "
                                 f"menu. Configurez-le via le bouton ci-dessous.")
            return
        labels = []
        for _titre, items in core.MENU_STRUCTURE:
            for label, key in items:
                if key in menus:
                    labels.append(label)
        self.apercu_var.set(f"{len(menus)} module(s) autorisé(s) : " + ", ".join(labels))

    def _ouvrir_niveaux_acces(self):
        app = self.winfo_toplevel()
        if hasattr(app, "show"):
            app.show("niveaux_acces")
        else:
            messagebox.showinfo("Niveaux d'accès",
                                 "Ouvrez le menu ADMIN > Niveaux d'accès pour configurer les modules.",
                                 parent=self)

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        v = self.tree.item(sel[0], "values")
        self.selected_id = v[0]
        self.login_var.set(v[1]); self.nom_complet_var.set(v[2]); self.niveau_var.set(v[3])
        self.actif_var.set(v[4] == "Oui")
        self.password_var.set("")
        self._refresh_apercu_modules()

    def clear_form(self):
        self.selected_id = None
        self.login_var.set(""); self.nom_complet_var.set(""); self.niveau_var.set("")
        self.password_var.set(""); self.actif_var.set(True)

    def add(self):
        if not self.login_var.get().strip():
            messagebox.showwarning("Champ manquant", "Le nom d'utilisateur est obligatoire.")
            return
        if not self.password_var.get():
            messagebox.showwarning("Champ manquant", "Le mot de passe est obligatoire à la création.")
            return
        try:
            core.add_utilisateur(self.conn, self.login_var.get().strip(), self.password_var.get(),
                                  nom_complet=self.nom_complet_var.get().strip(),
                                  niveau_acces=self.niveau_var.get() or "Lecture seule",
                                  actif=self.actif_var.get())
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc))
            return
        self.clear_form()
        self.refresh()

    def update_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un utilisateur.")
            return
        core.update_utilisateur(
            self.conn, self.selected_id,
            nouveau_mot_de_passe=self.password_var.get() or None,
            nom_utilisateur=self.login_var.get().strip(), nom_complet=self.nom_complet_var.get().strip(),
            niveau_acces=self.niveau_var.get() or "Lecture seule", actif=1 if self.actif_var.get() else 0,
        )
        self.clear_form()
        self.refresh()

    def delete_sel(self):
        if not self.selected_id:
            messagebox.showinfo("Info", "Sélectionnez d'abord un utilisateur.")
            return
        if messagebox.askyesno("Confirmer", "Supprimer cet utilisateur ?"):
            core.delete_utilisateur(self.conn, self.selected_id)
            self.clear_form()
            self.refresh()

    def refresh(self):
        self._refresh_niveaux()
        for row in self.tree.get_children():
            self.tree.delete(row)
        for u in core.list_utilisateurs(self.conn):
            self.tree.insert("", "end", values=(
                u["id"], u["nom_utilisateur"], u["nom_complet"] or "", u["niveau_acces"],
                "Oui" if u["actif"] else "Non", u["date_creation"] or ""))


class AdminModeleBonCommandeTab(ttk.Frame):
    """Modèle par défaut (menu ADMIN) du Bon de commande imprimé depuis
    l'onglet Factures frs : en-tête et pied de page appliqués à tout bon de
    commande dont l'en-tête/pied de page propre est vide — permet de définir
    une seule fois l'entête de la société (logo texte, coordonnées, mentions
    légales) plutôt que de le retaper à chaque commande."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="MODÈLE DE BON DE COMMANDE (PAR DÉFAUT)", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Appliqué automatiquement à tout bon de commande (onglet Factures frs) dont l'en-tête ou le "
            "pied de page n'a pas été rempli spécifiquement pour cette commande — pratique pour ne définir "
            "qu'une seule fois les coordonnées de votre société, mentions légales, etc."
        ), foreground="#595959", wraplength=1000).pack(anchor="w", padx=16, pady=(0, 8))

        ttk.Label(self, text="En-tête par défaut :").pack(anchor="w", padx=16)
        self.entete_text = tk.Text(self, height=8, wrap="word")
        self.entete_text.pack(fill="x", padx=16, pady=(2, 10))

        ttk.Label(self, text="Pied de page par défaut :").pack(anchor="w", padx=16)
        self.pied_text = tk.Text(self, height=6, wrap="word")
        self.pied_text.pack(fill="x", padx=16, pady=(2, 10))

        ttk.Button(self, text="Enregistrer le modèle", command=self.save).pack(anchor="w", padx=16, pady=(0, 16))
        self.refresh()

    def refresh(self):
        self.entete_text.delete("1.0", "end")
        self.entete_text.insert("1.0", core.get_text_setting(self.conn, "bon_commande_entete_defaut", ""))
        self.pied_text.delete("1.0", "end")
        self.pied_text.insert("1.0", core.get_text_setting(self.conn, "bon_commande_pied_defaut", ""))

    def save(self):
        core.set_text_setting(self.conn, "bon_commande_entete_defaut", self.entete_text.get("1.0", "end").strip())
        core.set_text_setting(self.conn, "bon_commande_pied_defaut", self.pied_text.get("1.0", "end").strip())
        messagebox.showinfo("Enregistré", "Modèle de bon de commande enregistré.")


class AdminFacturesTab(ttk.Frame):
    """Vue consolidée (menu ADMIN) de toutes les factures VALIDÉES — vente
    et achat — avec la possibilité de les repasser en brouillon modifiable
    en cas d'erreur sur les chiffres (dévalidation : retire les écritures
    comptables générées, sans rien supprimer d'autre). Complète, sans le
    remplacer, le bouton de correction déjà présent dans chaque onglet de
    facturation."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="ADMIN — MODIFICATION DES FACTURES", font=("Segoe UI", 14, "bold")).pack(
            anchor="w", padx=16, pady=(16, 4))
        ttk.Label(self, text=(
            "Liste de toutes les factures déjà validées (vente et achat). Sélectionnez-en une et cliquez sur "
            "« Dévalider » pour la repasser en brouillon modifiable : ses écritures comptables générées sont "
            "retirées de la Saisie, puis vous pouvez corriger les chiffres dans l'onglet Facturation ou "
            "Factures frs et la revalider."
        ), foreground="#595959", wraplength=1100).pack(anchor="w", padx=16, pady=(0, 8))

        btn_bar = ttk.Frame(self)
        btn_bar.pack(fill="x", padx=16, pady=4)
        ttk.Button(btn_bar, text="Actualiser", command=self.refresh).pack(side="left")
        ttk.Button(btn_bar, text="Dévalider la facture sélectionnée", command=self.devalider).pack(
            side="left", padx=8)

        cols = ("type", "numero", "date", "tiers", "statut")
        self.tree = ttk.Treeview(self, columns=cols, show="headings", height=20)
        headers = ["Type", "N° Pièce", "Date", "Client / Fournisseur", "Statut"]
        widths = [90, 140, 100, 320, 100]
        for c, h, w in zip(cols, headers, widths):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self._by_iid = {}
        self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self._by_iid = {}
        for f in core.list_factures_vente(self.conn):
            if f["statut"] != "validee":
                continue
            iid = self.tree.insert("", "end", values=(
                "Vente", f["numero"], core.to_display_date(f["date_facture"]), f["raison_sociale"], "Validée"))
            self._by_iid[iid] = ("vente", f["id"])
        for f in core.list_factures_achat(self.conn):
            if f["statut"] != "validee":
                continue
            iid = self.tree.insert("", "end", values=(
                "Achat", f["numero"], core.to_display_date(f["date_facture"]), f["raison_sociale"], "Validée"))
            self._by_iid[iid] = ("achat", f["id"])

    def devalider(self):
        sel = self.tree.selection()
        if not sel:
            messagebox.showinfo("Info", "Sélectionnez d'abord une facture dans le tableau.")
            return
        type_facture, facture_id = self._by_iid[sel[0]]
        if not messagebox.askyesno(
            "Dévalider cette facture",
            "Cette facture va être repassée en brouillon modifiable : ses écritures comptables générées "
            "vont être RETIRÉES de la Saisie.\n\nVous pourrez ensuite la corriger dans l'onglet correspondant "
            "puis la revalider.\n\nContinuer ?"
        ):
            return
        try:
            if type_facture == "vente":
                core.devalider_facture_vente(self.conn, facture_id)
            else:
                core.devalider_facture_achat(self.conn, facture_id)
        except ValueError as exc:
            messagebox.showerror("Erreur", str(exc))
            return
        messagebox.showinfo("Facture repassée en brouillon",
                             "La facture est de nouveau modifiable dans son onglet d'origine.")
        self.refresh()


class PlanBudgetaireTab(ttk.Frame):
    """Plan budgétaire : Code + Libellé + Montant prévu."""

    def __init__(self, parent, conn):
        super().__init__(parent)
        self.conn = conn
        ttk.Label(self, text="PLAN BUDGÉTAIRE", font=("Segoe UI", 14, "bold")).pack(anchor="w", padx=16, pady=(16, 4))

        import_bar = ttk.Frame(self)
        import_bar.pack(fill="x", padx=16, pady=(0, 4))
        ttk.Button(import_bar, text="Importer (.xlsx) — ÉCRASE le plan actuel",
                   command=self.import_xlsx).pack(side="left", padx=2)
        ttk.Button(import_bar, text="Exporter (.xlsx)", command=self.export_xlsx).pack(side="left", padx=2)

        form = ttk.Frame(self)
        form.pack(fill="x", padx=16, pady=6)
        ttk.Label(form, text="Code budgétaire :").grid(row=0, column=0, sticky="w")
        self.code_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.code_var, width=16).grid(row=0, column=1, padx=6)
        ttk.Label(form, text="Libellé :").grid(row=0, column=2, sticky="w", padx=(16, 0))
        self.label_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.label_var, width=35).grid(row=0, column=3, padx=6)
        ttk.Label(form, text="Montant prévu :").grid(row=0, column=4, sticky="w", padx=(16, 0))
        self.montant_var = tk.StringVar()
        ttk.Entry(form, textvariable=self.montant_var, width=16).grid(row=0, column=5, padx=6)
        ttk.Button(form, text="Créer / Modifier", command=self.save).grid(row=0, column=6, padx=6)
        ttk.Button(form, text="Supprimer", command=self.delete).grid(row=0, column=7, padx=6)

        cols = ("code", "label", "montant")
        self.tree = ttk.Treeview(self, columns=cols, show="headings")
        for c, h, w in zip(cols, ["Code budgétaire", "Libellé", "Montant prévu"], [120, 400, 130]):
            self.tree.heading(c, text=h)
            self.tree.column(c, width=w, anchor="w")
        self.tree.pack(fill="both", expand=True, padx=16, pady=8)
        self.tree.bind("<<TreeviewSelect>>", self._on_select)
        self.refresh()

    def _on_select(self, event=None):
        sel = self.tree.selection()
        if not sel:
            return
        values = self.tree.item(sel[0], "values")
        self.code_var.set(values[0])
        self.label_var.set(values[1])
        self.montant_var.set(values[2])

    def save(self):
        code = self.code_var.get().strip()
        label = self.label_var.get().strip()
        try:
            montant = float(self.montant_var.get() or 0)
        except ValueError:
            messagebox.showerror("Erreur", "Le montant prévu doit être un nombre.")
            return
        if not code or not label:
            messagebox.showwarning("Champs manquants", "Code budgétaire et Libellé sont obligatoires.")
            return
        core.add_budget_code(self.conn, code, label, montant)
        self.refresh()

    def delete(self):
        code = self.code_var.get().strip()
        if not code:
            messagebox.showinfo("Info", "Sélectionnez d'abord une ligne.")
            return
        if messagebox.askyesno("Confirmer", f"Supprimer « {code} » ?"):
            core.delete_budget_code(self.conn, code)
            self.code_var.set("")
            self.label_var.set("")
            self.montant_var.set("")
            self.refresh()

    def refresh(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        for item in core.list_budget_codes(self.conn):
            self.tree.insert("", "end", values=(item["code"], item["label"], f"{fmt_cfa(item['montant'])}"))

    def export_xlsx(self):
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Classeur Excel", "*.xlsx")],
            initialfile="Plan_budgetaire.xlsx", title="Exporter le Plan budgétaire",
        )
        if not path:
            return
        core.export_budget_codes_xlsx(self.conn, path)
        messagebox.showinfo("Export terminé", f"Plan budgétaire exporté :\n{path}")

    def import_xlsx(self):
        path = filedialog.askopenfilename(filetypes=[("Classeur Excel", "*.xlsx")],
                                           title="Importer un Plan budgétaire")
        if not path:
            return
        if not messagebox.askyesno(
            "Confirmer l'écrasement",
            "Importer ce fichier va ÉCRASER complètement le Plan budgétaire actuel. "
            "Cette action est irréversible. Continuer ?"
        ):
            return
        try:
            n = core.import_budget_codes_xlsx(self.conn, path)
        except Exception as exc:
            messagebox.showerror("Erreur", f"Échec de l'import : {exc}")
            return
        self.refresh()
        messagebox.showinfo("Import terminé", f"{n} ligne(s) importée(s). Le plan précédent a été remplacé.")


if __name__ == "__main__":
    app = App()
    if app.winfo_exists():
        app.mainloop()
