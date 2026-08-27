"""
core.py — Moteur comptable (sans interface graphique).

Toute la logique métier vit ici, indépendamment de Tkinter, pour rester
testable en ligne de commande. main.py ne fait qu'appeler ces fonctions.
"""
import hashlib
import json
import math
import os
import copy
import re
import secrets
import sys
import sqlite3
from datetime import date, datetime, timedelta


def _resource_dir():
    """Dossier des ressources bundlées : gère le cas exécutable PyInstaller."""
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        return sys._MEIPASS
    return os.path.dirname(os.path.abspath(__file__))

# ---------------------------------------------------------------------------
# Comptes SYSCOHADA "spéciaux" utilisés par les calculs automatiques
# (repris de la maquette Excel d'origine).
# ---------------------------------------------------------------------------
COMPTES_STOCK = ["310000", "320000", "331000", "360000"]
# Préfixes (3 chiffres) des mêmes comptes, utilisés uniquement pour agréger le
# total des stocks au Bilan (capture aussi d'éventuels sous-comptes de stock
# détaillés) — le suivi détaillé (onglet Stocks) reste lui scopé aux 4 comptes
# maîtres ci-dessus.
COMPTES_STOCK_PREFIXES = ["31", "32", "33", "36"]


def account_racine(code):
    """Racine (compte de rattachement) d'un compte : 1 chiffre pour les classes
    1,2,3,5,6,7,8,9 ; 2 chiffres pour la classe 4 (40 à 49), qui se subdivise
    par nature de tiers (40=Fournisseurs, 41=Clients, 42=Personnel,
    43=Organismes sociaux, 44=État, 45=Organismes internationaux,
    46=Associés/Groupe, 47=Débiteurs/créditeurs divers, 48=Régularisations,
    49=Dépréciations)."""
    code = str(code)
    if not code:
        return ""
    return code[:2] if code[0] == "4" else code[:1]


RACINE_FOURNISSEURS = "40"
RACINE_CLIENTS = "41"

# ---------------------------------------------------------------------------
# Facturation — mapping compte de vente (classe 70) -> impact sur les stocks.
# Un compte de vente lié à des marchandises (classe 31) ou des produits finis
# (classe 36) déclenche une sortie de stock automatique à la validation de la
# facture ; un compte de service (ex. 706000) n'impacte aucun stock.
# Le rattachement se fait par PRÉFIXE (3 chiffres) pour couvrir aussi les
# sous-comptes détaillés (ex. 701100, 701900... tous rattachés au préfixe 701).
# ---------------------------------------------------------------------------
VENTE_STOCK_MAPPING = {
    "701": ("marchandise", "310000", "603100"),   # Ventes marchandises -> stock 31, coût 603100
    "702": ("produit_fini", "360000", "736000"),  # Ventes produits finis -> stock 36, coût 736000
}
COMPTE_TVA_VENTES = "443100"  # État, T.V.A. facturée sur ventes
TVA_TAUX_DEFAUT = 18.0

# Achats (classe 6) -> impact sur les stocks : un achat de marchandises ou de
# matières premières augmente le stock correspondant (par préfixe, ex. 602101
# "Achat clinker" est bien rattaché au préfixe 602) ; un achat de service
# (ex. 622000) n'impacte aucun stock.
ACHAT_STOCK_MAPPING = {
    "601": ("marchandise", "310000", "603100"),        # Achats marchandises -> stock 31
    "602": ("matiere_premiere", "320000", "603200"),   # Achats matières premières -> stock 32
}
RETENUE_TAUX_DEFAUT = 0.0
COMPTE_RETENUE_DEFAUT = "447800"  # État, autres impôts et contributions (retenues à la source)


def _match_stock_mapping(compte, mapping):
    """Retrouve le mapping stock applicable à un compte, par préfixe de 3
    chiffres (ex. 602101 correspond au préfixe 602)."""
    if not compte or len(compte) < 3:
        return None
    return mapping.get(compte[:3])

RACINE_LABELS = {
    "1": "Comptes de ressources durables",
    "2": "Comptes d'actif immobilisé",
    "3": "Comptes de stocks",
    "40": "Fournisseurs et comptes rattachés",
    "41": "Clients et comptes rattachés",
    "42": "Personnel",
    "43": "Organismes sociaux",
    "44": "État et collectivités publiques",
    "45": "Organismes internationaux",
    "46": "Associés et groupe",
    "47": "Débiteurs et créditeurs divers",
    "48": "Comptes de régularisation",
    "49": "Dépréciations et provisions sur tiers",
    "5": "Comptes de trésorerie",
    "6": "Comptes de charges",
    "7": "Comptes de produits",
    "8": "Comptes des autres charges et produits (HAO)",
    "9": "Comptes analytiques/engagements",
}

COMPTES_TRESORERIE = ["521", "531", "570", "585"]
COMPTES_CAPITAL = ["101", "118", "121"]
COMPTE_SUBVENTIONS = "141"
COMPTE_PROVISIONS = "191"
COMPTES_DETTES_FIN = ["162", "165"]
COMPTES_PRODUITS_EXPL = ["701", "702", "705", "706", "736"]
COMPTE_SUBV_EXPL = "710"
COMPTE_AUTRES_PRODUITS = "758"
COMPTES_ACHATS = ["601", "602", "604", "605", "603"]
COMPTES_TRANSPORT = ["610", "614"]
COMPTES_SERVICES_EXT = ["622", "624", "625", "626", "627", "628",
                         "631", "632", "633"]
COMPTES_IMPOTS = ["641", "645"]
COMPTE_AUTRES_CHARGES = "651"
COMPTES_PERSONNEL = ["661", "663", "664"]
COMPTES_DOTATIONS = ["681", "691"]
COMPTES_PRODUITS_FIN = ["771", "776"]
COMPTES_CHARGES_FIN = ["671", "676"]

# ---------------------------------------------------------------------------
# Liasse fiscale — codes SYSCOHADA "système normal" (BILAN / RESULTAT)
# NB : les totaux (AD, AI, AZ, BK, BT, BZ, CP, DD, DP, DT, DZ) sont fiables
# (dérivés directement de la partie double). Le détail par ligne (AE..AN,
# CA/CH/CJ, DA/DJ/DK/DM/DR) est une répartition indicative par plage de
# comptes — à vérifier avec votre expert-comptable avant tout dépôt officiel.
# ---------------------------------------------------------------------------
RANGES_INCORP = {"AE": (211000, 211999), "AF": (212000, 214999),
                  "AG": (215000, 216999), "AH": (217000, 219999)}
RANGE_AMORT_INCORP = (281000, 281999)
RANGES_CORP = {"AJ": [(220000, 229999)], "AK": [(230000, 233999)],
               "AL": [(234000, 239999)],
               "AM": [(240000, 244999), (246000, 249999)],
               "AN": [(245000, 245999)]}
RANGE_AMORT_CORP = (282000, 297999)
RANGE_AVANCES_IMMO = (250000, 252999)
RANGE_TITRES_PARTICIPATION = (260000, 268999)
RANGE_AUTRES_IMMO_FIN = (270000, 278999)

RANGES_CAPITAUX = {"CA": [(101000, 104999)], "CD": [(105000, 105999)],
                    "CF_CG": [(110000, 118999)], "CH": [(120000, 129999)],
                    "CL": [(140000, 148999)], "CM": [(150000, 158999)]}
RANGE_DETTES_FIN = (160000, 168999)
RANGE_DETTES_LOCATION = (170000, 178999)
RANGE_PROVISIONS_RC = (190000, 198999)

RANGE_STOCKS = (300000, 399999)
RANGE_AVANCES_FOURN = (409000, 409999)
RANGE_CLIENTS = (411000, 419999)
RANGE_FOURNISSEURS = (401000, 408999)
RANGE_DETTES_FISC_SOC = (420000, 449999)
RANGE_AUTRES_DETTES = (450000, 499999)


def _in_ranges(code_int, ranges):
    if isinstance(ranges, tuple):
        ranges = [ranges]
    return any(lo <= code_int <= hi for lo, hi in ranges)


def _sum_range(balance, ranges, classe=None):
    total = 0.0
    for b in balance:
        code_int = int(b["code"])
        if classe and b["classe"] != classe:
            continue
        if _in_ranges(code_int, ranges):
            total += b["solde_cloture"]
    return total


def compute_liasse_bilan(conn, stock_initial=0.0, exercice=None):
    """Bilan au format SYSCOHADA système normal (codes officiels)."""
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    bilan_simple = compute_bilan(conn, stock_initial=stock_initial, exercice=exercice)

    # --- Détail indicatif Immobilisations incorporelles ---
    incorp_brut = {k: _sum_range(balance, [rng]) for k, rng in RANGES_INCORP.items()}
    total_incorp_brut = sum(incorp_brut.values())
    amort_incorp_total = -_sum_range(balance, [RANGE_AMORT_INCORP])  # positif
    incorp_net = {}
    for k, brut in incorp_brut.items():
        part = (brut / total_incorp_brut * amort_incorp_total) if total_incorp_brut else 0
        incorp_net[k] = brut - part

    # --- Détail indicatif Immobilisations corporelles ---
    corp_brut = {k: _sum_range(balance, rngs) for k, rngs in RANGES_CORP.items()}
    total_corp_brut = sum(corp_brut.values())
    amort_corp_total = -_sum_range(balance, [RANGE_AMORT_CORP])
    corp_net = {}
    for k, brut in corp_brut.items():
        part = (brut / total_corp_brut * amort_corp_total) if total_corp_brut else 0
        corp_net[k] = brut - part

    avances_immo = _sum_range(balance, [RANGE_AVANCES_IMMO])
    titres_participation = _sum_range(balance, [RANGE_TITRES_PARTICIPATION])
    autres_immo_fin = _sum_range(balance, [RANGE_AUTRES_IMMO_FIN])

    # --- Détail indicatif Capitaux propres ---
    capitaux_detail = {k: -_sum_range(balance, rngs) for k, rngs in RANGES_CAPITAUX.items()}
    dettes_financieres = -_sum_range(balance, [RANGE_DETTES_FIN])
    dettes_location = -_sum_range(balance, [RANGE_DETTES_LOCATION])
    provisions_rc = -_sum_range(balance, [RANGE_PROVISIONS_RC])

    # --- Détail indicatif Passif circulant ---
    fournisseurs = -_sum_range(balance, [RANGE_FOURNISSEURS])
    avances_fourn = -_sum_range(balance, [RANGE_AVANCES_FOURN])
    dettes_fisc_soc = -_sum_range(balance, [RANGE_DETTES_FISC_SOC])
    autres_dettes = -_sum_range(balance, [RANGE_AUTRES_DETTES])

    # --- Détail indicatif Actif circulant ---
    avances_versees = _sum_range(balance, [RANGE_AVANCES_FOURN])
    clients = _sum_range(balance, [RANGE_CLIENTS])

    return {
        "totaux": bilan_simple,
        "actif_detail": {
            **{k: {"brut": incorp_brut[k], "net": incorp_net[k]} for k in incorp_brut},
            **{k: {"brut": corp_brut[k], "net": corp_net[k]} for k in corp_brut},
            "AP": {"brut": avances_immo, "net": avances_immo},
            "AR": {"brut": titres_participation, "net": titres_participation},
            "AS": {"brut": autres_immo_fin, "net": autres_immo_fin},
        },
        "actif_circulant_detail": {
            "BH": avances_versees, "BI": clients,
        },
        "passif_detail": {
            **capitaux_detail,
            "DA": dettes_financieres, "DB": dettes_location, "DC": provisions_rc,
            "DJ": fournisseurs, "DH_avances": avances_fourn,
            "DK": dettes_fisc_soc, "DM": autres_dettes,
        },
    }


def compute_liasse_resultat(conn, exercice=None):
    """Compte de résultat au format SYSCOHADA système normal (codes officiels)."""
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)

    def net_produit(codes):
        d, c = _sum_accounts(balance, codes)
        return c - d

    def net_charge(codes):
        d, c = _sum_accounts(balance, codes)
        return d - c

    ta = net_produit(["701"])
    ra = net_charge(["601"])
    ra_stock = net_charge(["603100"])  # variation de stock de marchandises (préfixe précis pour ne pas doubler avec 603200)
    xa = ta - ra - ra_stock  # marge commerciale

    tb = net_produit(["702"])
    tc = net_produit(["705", "706"])
    td = 0.0
    xb = ta + tb + tc + td

    stock_d, stock_c = _sum_accounts(balance, ["360"])
    te = stock_d - stock_c
    th = net_produit(["758"])
    tg = net_produit(["710"])

    rc = net_charge(["602", "603200"])
    re = net_charge(["604", "605"])
    rg = net_charge(["610", "614"])
    rh = net_charge(["622", "624", "625", "626", "627", "628",
                      "631", "632", "633"])
    ri = net_charge(["641", "645"])
    rj = net_charge(["651"])
    xc = xb + (-ra) + (-ra_stock) + te + tg + th + (-rc) + (-re) + (-rg) + (-rh) + (-ri) + (-rj)

    rk = net_charge(["661", "663", "664"])
    xd = xc - rk

    rl = net_charge(["681", "691"])
    xe = xd - rl

    tk = net_produit(["771", "776"])
    rm = net_charge(["671", "676"])
    xf = tk - rm
    xg = xe + xf

    # Résultat HAO : intègre TOUTE la classe 8 (comptes 81 à 89 — cessions
    # d'immobilisations, charges/produits HAO, participation des travailleurs,
    # impôts sur le résultat...), plutôt que de la considérer comme nulle.
    balance8 = compute_balance(conn, only_with_movement=False, exercice=exercice)
    xh = -sum(b["solde_cloture"] for b in balance8 if b["classe"] == "8")
    rq = 0.0  # Participation des travailleurs — comprise dans XH si postée en classe 8
    rs = 0.0  # Impôt sur le résultat — compris dans XH si posté en classe 8
    xi = xg + xh + rq + rs

    # --- Recalage sur le résultat net exhaustif (compute_resultat_net_complet) ---
    # TA..RM ci-dessus reposent sur des listes de comptes usuelles (classe 6/7)
    # qui peuvent ne pas couvrir tout le plan comptable réel (1591 comptes).
    # Le résultat net exhaustif (classes 6+7+8, sans aucune omission) sert de
    # référence : l'écart éventuel est replié dans « Autres produits »/« Autres
    # charges » (TH/RJ), pour que XI corresponde TOUJOURS exactement au
    # résultat net réel — et reste donc cohérent avec le Bilan.
    resultat_complet = compute_resultat_net_complet(conn, exercice=exercice)
    ecart_reclassement = resultat_complet - xi
    if ecart_reclassement >= 0:
        th += ecart_reclassement
    else:
        rj += -ecart_reclassement
    xc += ecart_reclassement
    xd += ecart_reclassement
    xe += ecart_reclassement
    xg += ecart_reclassement
    xi += ecart_reclassement

    return {
        "TA": ta, "RA": ra, "RA_STOCK": ra_stock, "XA": xa,
        "TB": tb, "TC": tc, "TD": td, "XB": xb,
        "TE": te, "TG": tg, "TH": th,
        "RC": rc, "RE": re, "RG": rg, "RH": rh, "RI": ri, "RJ": rj, "XC": xc,
        "RK": rk, "XD": xd,
        "RL": rl, "XE": xe,
        "TK": tk, "RM": rm, "XF": xf, "XG": xg,
        "XH": xh, "RQ": rq, "RS": rs, "XI": xi,
    }


def default_db_path():
    """Emplacement du fichier de données, à côté de l'exécutable."""
    base = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
    folder = os.path.join(base, "SaisieComptable")
    os.makedirs(folder, exist_ok=True)
    return os.path.join(folder, "comptabilite.db")


def get_connection(db_path=None):
    db_path = db_path or default_db_path()
    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    init_db(conn)
    return conn


def init_db(conn):
    conn.execute("""
        CREATE TABLE IF NOT EXISTS accounts (
            code TEXT PRIMARY KEY,
            label TEXT NOT NULL,
            classe TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS entries (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            date TEXT NOT NULL,
            piece TEXT,
            journal TEXT,
            compte TEXT NOT NULL,
            tiers TEXT,
            libelle TEXT,
            debit REAL NOT NULL DEFAULT 0,
            credit REAL NOT NULL DEFAULT 0,
            flux_code TEXT,
            analytic_code TEXT,
            budget_code TEXT,
            donor_code TEXT,
            quantite REAL NOT NULL DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS settings (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS opening_balances (
            code TEXT NOT NULL,
            exercice TEXT NOT NULL,
            solde REAL NOT NULL DEFAULT 0,
            PRIMARY KEY (code, exercice)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS exercices (
            exercice TEXT PRIMARY KEY,
            cloture INTEGER NOT NULL DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS pointages_bancaires (
            entry_id INTEGER PRIMARY KEY REFERENCES entries(id) ON DELETE CASCADE,
            pointe INTEGER NOT NULL DEFAULT 1,
            date_pointage TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS analytic_codes (
            code TEXT PRIMARY KEY,
            label TEXT NOT NULL,
            unite TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS budget_codes (
            code TEXT PRIMARY KEY,
            label TEXT NOT NULL,
            montant REAL NOT NULL DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS taux_tva (
            code TEXT PRIMARY KEY,
            label TEXT NOT NULL,
            montant REAL NOT NULL DEFAULT 0,
            compte TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS taux_retenue (
            code TEXT PRIMARY KEY,
            label TEXT NOT NULL,
            montant REAL NOT NULL DEFAULT 0,
            compte TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS donor_codes (
            code TEXT PRIMARY KEY,
            label TEXT NOT NULL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS fournisseurs (
            code TEXT PRIMARY KEY,
            raison_sociale TEXT NOT NULL,
            contact TEXT,
            telephone TEXT,
            adresse TEXT,
            delai_paiement_jours INTEGER NOT NULL DEFAULT 30,
            delai_livraison_jours INTEGER NOT NULL DEFAULT 15
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS commandes_fournisseurs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            fournisseur_code TEXT NOT NULL,
            piece TEXT,
            libelle TEXT,
            montant REAL NOT NULL DEFAULT 0,
            date_commande TEXT NOT NULL,
            date_livraison_prevue TEXT,
            date_livraison_reelle TEXT,
            date_echeance_paiement TEXT,
            date_paiement_reel TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS clients (
            code TEXT PRIMARY KEY,
            raison_sociale TEXT NOT NULL,
            contact TEXT,
            telephone TEXT,
            adresse TEXT,
            delai_paiement_jours INTEGER NOT NULL DEFAULT 30
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS factures_clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_code TEXT NOT NULL,
            piece TEXT,
            libelle TEXT,
            montant REAL NOT NULL DEFAULT 0,
            date_facture TEXT NOT NULL,
            date_echeance_paiement TEXT,
            date_paiement_reel TEXT,
            compte_reglement TEXT,
            reglement_comptabilise INTEGER NOT NULL DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS produits_finis (
            code TEXT PRIMARY KEY,
            nom TEXT NOT NULL,
            description TEXT,
            quantite_produite REAL NOT NULL DEFAULT 1,
            marge_pourcentage REAL NOT NULL DEFAULT 30,
            compte_stock TEXT NOT NULL DEFAULT '360000'
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS recette_lignes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            produit_code TEXT NOT NULL,
            type_ligne TEXT NOT NULL,
            libelle TEXT NOT NULL,
            compte TEXT,
            quantite REAL NOT NULL DEFAULT 0,
            cout_unitaire REAL,
            analytic_code TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS factures_vente (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero TEXT NOT NULL,
            date_facture TEXT NOT NULL,
            client_code TEXT NOT NULL,
            entete TEXT,
            pied_page TEXT,
            tva_taux REAL NOT NULL DEFAULT 0,
            tva_compte TEXT,
            statut TEXT NOT NULL DEFAULT 'brouillon',
            piece TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS facture_vente_lignes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            facture_id INTEGER NOT NULL,
            compte_vente TEXT NOT NULL,
            libelle TEXT NOT NULL,
            quantite REAL NOT NULL DEFAULT 0,
            prix_unitaire REAL NOT NULL DEFAULT 0,
            analytic_code TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS factures_achat (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero TEXT NOT NULL,
            date_facture TEXT NOT NULL,
            fournisseur_code TEXT NOT NULL,
            entete TEXT,
            pied_page TEXT,
            retenue_taux REAL NOT NULL DEFAULT 0,
            retenue_compte TEXT NOT NULL DEFAULT '447800',
            statut TEXT NOT NULL DEFAULT 'brouillon',
            piece TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS expressions_besoin (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero TEXT NOT NULL,
            date_demande TEXT NOT NULL,
            demandeur TEXT,
            service TEXT,
            entete TEXT,
            pied_page TEXT,
            statut TEXT NOT NULL DEFAULT 'brouillon'
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS expression_besoin_lignes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            expression_id INTEGER NOT NULL,
            libelle TEXT NOT NULL,
            quantite REAL NOT NULL DEFAULT 0,
            unite TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ep_bons_commande (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero TEXT NOT NULL,
            date_commande TEXT NOT NULL,
            expression_id INTEGER,
            fournisseur_code TEXT,
            entete TEXT,
            pied_page TEXT,
            statut TEXT NOT NULL DEFAULT 'brouillon',
            date_facture TEXT,
            date_saisie TEXT,
            date_paiement_attendu TEXT,
            retenue_taux REAL NOT NULL DEFAULT 0,
            retenue_compte TEXT DEFAULT '447800',
            piece TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS ep_bon_commande_lignes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bon_commande_id INTEGER NOT NULL,
            libelle TEXT NOT NULL,
            quantite REAL NOT NULL DEFAULT 0,
            prix_unitaire REAL NOT NULL DEFAULT 0,
            unite TEXT,
            compte_charge TEXT,
            analytic_code TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS bordereaux_livraison (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero TEXT NOT NULL,
            date_livraison TEXT NOT NULL,
            bon_commande_id INTEGER,
            entete TEXT,
            pied_page TEXT,
            statut TEXT NOT NULL DEFAULT 'brouillon'
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS bordereau_livraison_lignes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            bordereau_id INTEGER NOT NULL,
            libelle TEXT NOT NULL,
            quantite_commandee REAL NOT NULL DEFAULT 0,
            quantite_livree REAL NOT NULL DEFAULT 0,
            unite TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS reglements (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            numero TEXT NOT NULL,
            date_reglement TEXT NOT NULL,
            bon_commande_id INTEGER,
            fournisseur_code TEXT,
            entete TEXT,
            pied_page TEXT,
            retenue_taux REAL NOT NULL DEFAULT 0,
            retenue_compte TEXT NOT NULL DEFAULT '447800',
            statut TEXT NOT NULL DEFAULT 'brouillon',
            piece TEXT,
            date_paiement TEXT,
            compte_paiement TEXT,
            paiement_comptabilise INTEGER NOT NULL DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS reglement_lignes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reglement_id INTEGER NOT NULL,
            compte_charge TEXT,
            libelle TEXT NOT NULL,
            quantite REAL NOT NULL DEFAULT 0,
            prix_unitaire REAL NOT NULL DEFAULT 0,
            analytic_code TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS facture_achat_lignes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            facture_id INTEGER NOT NULL,
            compte_achat TEXT NOT NULL,
            libelle TEXT NOT NULL,
            quantite REAL NOT NULL DEFAULT 0,
            prix_unitaire REAL NOT NULL DEFAULT 0,
            analytic_code TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS pieces_rechange (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            code TEXT,
            designation TEXT NOT NULL,
            quantite_stock REAL NOT NULL DEFAULT 0,
            unite TEXT,
            cout_unitaire REAL NOT NULL DEFAULT 0,
            fournisseur_code TEXT,
            notes TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS vehicules (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            immatriculation TEXT NOT NULL,
            marque TEXT,
            modele TEXT,
            type_vehicule TEXT,
            date_acquisition TEXT,
            chauffeur_affecte TEXT,
            statut TEXT NOT NULL DEFAULT 'actif',
            notes TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS missions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vehicule_id INTEGER,
            chauffeur TEXT,
            destination TEXT NOT NULL,
            motif TEXT,
            date_depart TEXT,
            date_retour TEXT,
            km_depart REAL,
            km_retour REAL,
            statut TEXT NOT NULL DEFAULT 'en_cours',
            notes TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS reparations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            vehicule_id INTEGER,
            date_reparation TEXT NOT NULL,
            description TEXT NOT NULL,
            garage TEXT,
            cout_main_oeuvre REAL NOT NULL DEFAULT 0,
            statut TEXT NOT NULL DEFAULT 'en_cours',
            notes TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS reparation_lignes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            reparation_id INTEGER NOT NULL,
            piece_id INTEGER NOT NULL,
            quantite REAL NOT NULL DEFAULT 1
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS immobilisations_fiche (
            compte TEXT PRIMARY KEY,
            fournisseur_code TEXT,
            prix_achat REAL NOT NULL DEFAULT 0,
            date_acquisition TEXT,
            base_repartition_quantite REAL,
            base_repartition_unite TEXT,
            amortissement_annuel_manuel REAL
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS taux_amortissement (
            categorie TEXT PRIMARY KEY,
            taux_pct REAL NOT NULL DEFAULT 0
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS personnel (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            matricule TEXT,
            nom TEXT NOT NULL,
            prenom TEXT,
            poste TEXT,
            service TEXT,
            date_embauche TEXT,
            telephone TEXT,
            email TEXT,
            salaire_base REAL NOT NULL DEFAULT 0,
            statut TEXT NOT NULL DEFAULT 'actif',
            notes TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS grh_time_sheet (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            personnel_id INTEGER NOT NULL,
            date_pointage TEXT NOT NULL,
            heures REAL NOT NULL DEFAULT 0,
            activite TEXT,
            notes TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS paie_bulletins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            personnel_id INTEGER NOT NULL,
            periode TEXT NOT NULL,
            classification TEXT NOT NULL DEFAULT 'AUTRE',
            salaire_base REAL NOT NULL DEFAULT 0,
            prime_anciennete REAL NOT NULL DEFAULT 0,
            heures_sup REAL NOT NULL DEFAULT 0,
            sursalaire REAL NOT NULL DEFAULT 0,
            gratification REAL NOT NULL DEFAULT 0,
            indemnite_caisse REAL NOT NULL DEFAULT 0,
            indemnite_logement REAL NOT NULL DEFAULT 0,
            indemnite_fonction REAL NOT NULL DEFAULT 0,
            indemnite_transport REAL NOT NULL DEFAULT 0,
            personnes_a_charge INTEGER NOT NULL DEFAULT 0,
            retenue_pret REAL NOT NULL DEFAULT 0,
            date_saisie TEXT,
            UNIQUE(personnel_id, periode)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS paie_periodes_validees (
            periode TEXT PRIMARY KEY,
            date_validation TEXT NOT NULL,
            piece TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS grh_kpi (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            indicateur TEXT NOT NULL,
            description TEXT,
            personnel_id INTEGER,
            service TEXT,
            periode TEXT,
            valeur_cible REAL NOT NULL DEFAULT 0,
            valeur_realisee REAL NOT NULL DEFAULT 0,
            unite TEXT,
            statut TEXT NOT NULL DEFAULT 'en_cours'
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS grh_hs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            personnel_id INTEGER,
            date_evenement TEXT NOT NULL,
            type_evenement TEXT NOT NULL DEFAULT 'incident',
            description TEXT,
            gravite TEXT,
            statut TEXT NOT NULL DEFAULT 'ouvert',
            notes TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS niveaux_acces (
            nom TEXT PRIMARY KEY,
            description TEXT
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS niveau_acces_menus (
            niveau_acces TEXT NOT NULL,
            menu_key TEXT NOT NULL,
            PRIMARY KEY (niveau_acces, menu_key)
        )
    """)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS utilisateurs (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nom_utilisateur TEXT NOT NULL UNIQUE,
            nom_complet TEXT,
            mot_de_passe_hash TEXT,
            sel TEXT,
            niveau_acces TEXT NOT NULL DEFAULT 'Lecture seule',
            actif INTEGER NOT NULL DEFAULT 1,
            date_creation TEXT
        )
    """)
    conn.commit()
    _migrate(conn)
    if conn.execute("SELECT COUNT(*) FROM accounts").fetchone()[0] == 0:
        load_plan_comptable(conn)
    ensure_racine_accounts(conn)


def _migrate(conn):
    """Ajoute les colonnes/tables manquantes si la base a été créée par une version antérieure."""
    cols = [r["name"] for r in conn.execute("PRAGMA table_info(entries)")]
    if "analytic_code" not in cols:
        conn.execute("ALTER TABLE entries ADD COLUMN analytic_code TEXT")
    if "budget_code" not in cols:
        conn.execute("ALTER TABLE entries ADD COLUMN budget_code TEXT")
    if "donor_code" not in cols:
        conn.execute("ALTER TABLE entries ADD COLUMN donor_code TEXT")
    if "quantite" not in cols:
        conn.execute("ALTER TABLE entries ADD COLUMN quantite REAL NOT NULL DEFAULT 0")
    if "fournisseur_code" not in cols:
        conn.execute("ALTER TABLE entries ADD COLUMN fournisseur_code TEXT")
    if "client_code" not in cols:
        conn.execute("ALTER TABLE entries ADD COLUMN client_code TEXT")

    pf_cols = [r["name"] for r in conn.execute("PRAGMA table_info(produits_finis)")]
    if pf_cols and "compte_stock" not in pf_cols:
        conn.execute("ALTER TABLE produits_finis ADD COLUMN compte_stock TEXT NOT NULL DEFAULT '360000'")

    rl_cols = [r["name"] for r in conn.execute("PRAGMA table_info(recette_lignes)")]
    if rl_cols and "analytic_code" not in rl_cols:
        conn.execute("ALTER TABLE recette_lignes ADD COLUMN analytic_code TEXT")

    ac_cols = [r["name"] for r in conn.execute("PRAGMA table_info(analytic_codes)")]
    if ac_cols and "unite" not in ac_cols:
        conn.execute("ALTER TABLE analytic_codes ADD COLUMN unite TEXT")

    for table in ("taux_tva", "taux_retenue"):
        t_cols = [r["name"] for r in conn.execute(f"PRAGMA table_info({table})")]
        if t_cols and "compte" not in t_cols:
            conn.execute(f"ALTER TABLE {table} ADD COLUMN compte TEXT")

    fvl_cols = [r["name"] for r in conn.execute("PRAGMA table_info(facture_vente_lignes)")]
    if fvl_cols and "analytic_code" not in fvl_cols:
        conn.execute("ALTER TABLE facture_vente_lignes ADD COLUMN analytic_code TEXT")
    fal_cols = [r["name"] for r in conn.execute("PRAGMA table_info(facture_achat_lignes)")]
    if fal_cols and "analytic_code" not in fal_cols:
        conn.execute("ALTER TABLE facture_achat_lignes ADD COLUMN analytic_code TEXT")
    fv_cols = [r["name"] for r in conn.execute("PRAGMA table_info(factures_vente)")]
    if fv_cols and "tva_compte" not in fv_cols:
        conn.execute("ALTER TABLE factures_vente ADD COLUMN tva_compte TEXT")
    bc_cols = [r["name"] for r in conn.execute("PRAGMA table_info(ep_bons_commande)")]
    for col in ("date_facture", "date_saisie", "date_paiement_attendu"):
        if bc_cols and col not in bc_cols:
            conn.execute(f"ALTER TABLE ep_bons_commande ADD COLUMN {col} TEXT")
    for col, default in (("retenue_taux", "0"), ("retenue_compte", "'447800'"), ("piece", "NULL")):
        if bc_cols and col not in bc_cols:
            conn.execute(f"ALTER TABLE ep_bons_commande ADD COLUMN {col} DEFAULT {default}")
    bcl_cols = [r["name"] for r in conn.execute("PRAGMA table_info(ep_bon_commande_lignes)")]
    if bcl_cols and "compte_charge" not in bcl_cols:
        conn.execute("ALTER TABLE ep_bon_commande_lignes ADD COLUMN compte_charge TEXT")
    if bcl_cols and "analytic_code" not in bcl_cols:
        conn.execute("ALTER TABLE ep_bon_commande_lignes ADD COLUMN analytic_code TEXT")
    fc_cols = [r["name"] for r in conn.execute("PRAGMA table_info(factures_clients)")]
    if fc_cols and "compte_reglement" not in fc_cols:
        conn.execute("ALTER TABLE factures_clients ADD COLUMN compte_reglement TEXT")
    if fc_cols and "reglement_comptabilise" not in fc_cols:
        conn.execute("ALTER TABLE factures_clients ADD COLUMN reglement_comptabilise INTEGER NOT NULL DEFAULT 0")
    r_cols = [r["name"] for r in conn.execute("PRAGMA table_info(reglements)")]
    if r_cols and "date_paiement" not in r_cols:
        conn.execute("ALTER TABLE reglements ADD COLUMN date_paiement TEXT")
    if r_cols and "compte_paiement" not in r_cols:
        conn.execute("ALTER TABLE reglements ADD COLUMN compte_paiement TEXT")
    if r_cols and "paiement_comptabilise" not in r_cols:
        conn.execute("ALTER TABLE reglements ADD COLUMN paiement_comptabilise INTEGER NOT NULL DEFAULT 0")

    if_cols = [r["name"] for r in conn.execute("PRAGMA table_info(immobilisations_fiche)")]
    if if_cols and "base_repartition_quantite" not in if_cols:
        conn.execute("ALTER TABLE immobilisations_fiche ADD COLUMN base_repartition_quantite REAL")
    if if_cols and "base_repartition_unite" not in if_cols:
        conn.execute("ALTER TABLE immobilisations_fiche ADD COLUMN base_repartition_unite TEXT")
    if if_cols and "amortissement_annuel_manuel" not in if_cols:
        conn.execute("ALTER TABLE immobilisations_fiche ADD COLUMN amortissement_annuel_manuel REAL")

    # Migre l'ancien mécanisme "stock_initial_<compte>" (settings) vers opening_balances
    default_exercice = str(datetime.today().year)
    old_rows = conn.execute("SELECT key, value FROM settings WHERE key LIKE 'stock_initial_%'").fetchall()
    for row in old_rows:
        code = row["key"].replace("stock_initial_", "")
        try:
            val = float(row["value"])
        except (TypeError, ValueError):
            val = 0.0
        if val:
            conn.execute("INSERT OR REPLACE INTO opening_balances (code, exercice, solde) VALUES (?, ?, ?)",
                         (code, default_exercice, val))
        conn.execute("DELETE FROM settings WHERE key = ?", (row["key"],))

    # Migre l'ancienne table opening_balances (sans colonne exercice) vers le nouveau schéma
    ob_cols = [r["name"] for r in conn.execute("PRAGMA table_info(opening_balances)")]
    if "exercice" not in ob_cols:
        conn.execute("ALTER TABLE opening_balances RENAME TO opening_balances_old")
        conn.execute("""
            CREATE TABLE opening_balances (
                code TEXT NOT NULL, exercice TEXT NOT NULL, solde REAL NOT NULL DEFAULT 0,
                PRIMARY KEY (code, exercice)
            )
        """)
        for row in conn.execute("SELECT code, solde FROM opening_balances_old"):
            conn.execute("INSERT OR REPLACE INTO opening_balances (code, exercice, solde) VALUES (?, ?, ?)",
                         (row["code"], default_exercice, row["solde"]))
        conn.execute("DROP TABLE opening_balances_old")

    if conn.execute("SELECT 1 FROM exercices WHERE exercice = ?", (default_exercice,)).fetchone() is None:
        # S'assure qu'au moins l'exercice courant existe dans la table
        conn.execute("INSERT OR IGNORE INTO exercices (exercice, cloture) VALUES (?, 0)", (default_exercice,))
    conn.commit()


def get_setting(conn, key, default=0.0):
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    return float(row["value"]) if row else default


def set_setting(conn, key, value):
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, str(value)))
    conn.commit()


def set_text_setting(conn, key, value):
    """Alias explicite de set_setting() pour les réglages textuels (comptes,
    modèles de document...) — même mécanisme, juste un nom plus clair à
    l'usage que set_setting() pour un contenu qui n'est pas un nombre."""
    set_setting(conn, key, value)


def get_text_setting(conn, key, default=""):
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else default


# ---------------------------------------------------------------------------
# Exercices comptables
# ---------------------------------------------------------------------------
def get_current_exercice(conn):
    ex = get_text_setting(conn, "exercice_courant", "")
    if ex:
        return ex
    ex = str(datetime.today().year)
    set_setting(conn, "exercice_courant", ex)
    conn.execute("INSERT OR IGNORE INTO exercices (exercice, cloture) VALUES (?, 0)", (ex,))
    conn.commit()
    return ex


def set_current_exercice(conn, exercice):
    conn.execute("INSERT OR IGNORE INTO exercices (exercice, cloture) VALUES (?, 0)", (exercice,))
    set_setting(conn, "exercice_courant", exercice)


def list_exercices(conn):
    """Tous les exercices connus : ceux créés explicitement + toutes les années
    présentes dans les écritures, triés."""
    years_in_entries = {r[0][:4] for r in conn.execute("SELECT DISTINCT date FROM entries") if r[0]}
    years_known = {r["exercice"] for r in conn.execute("SELECT exercice FROM exercices")}
    all_years = sorted(years_in_entries | years_known)
    cloture_map = {r["exercice"]: bool(r["cloture"]) for r in conn.execute("SELECT exercice, cloture FROM exercices")}
    return [{"exercice": y, "cloture": cloture_map.get(y, False)} for y in all_years]


def is_exercice_cloture(conn, exercice):
    row = conn.execute("SELECT cloture FROM exercices WHERE exercice = ?", (exercice,)).fetchone()
    return bool(row["cloture"]) if row else False


def _exercice_of_date(date_str):
    return (date_str or "")[:4]


def close_exercice(conn, exercice):
    """Clôture un exercice : calcule les soldes de clôture de tous les comptes de
    bilan (classes 1 à 5), les reporte comme soldes d'ouverture de l'exercice
    suivant, y intègre le résultat net de l'exercice clôturé (compte 121000 —
    report à nouveau), puis marque l'exercice comme clôturé."""
    if is_exercice_cloture(conn, exercice):
        raise ValueError(f"L'exercice {exercice} est déjà clôturé.")
    next_exercice = str(int(exercice) + 1)

    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    resultat_net = compute_resultat_net_complet(conn, exercice=exercice)

    for b in balance:
        if b["classe"] not in ("1", "2", "3", "4", "5"):
            continue
        cloture = b["solde_cloture"]
        if b["code"] == "121000":
            cloture -= resultat_net  # intègre le résultat net dans le report à nouveau
        if cloture:
            conn.execute(
                "INSERT OR REPLACE INTO opening_balances (code, exercice, solde) VALUES (?, ?, ?)",
                (b["code"], next_exercice, cloture),
            )

    conn.execute("INSERT OR REPLACE INTO exercices (exercice, cloture) VALUES (?, 1)", (exercice,))
    conn.execute("INSERT OR IGNORE INTO exercices (exercice, cloture) VALUES (?, 0)", (next_exercice,))
    conn.commit()
    return next_exercice


# ---------------------------------------------------------------------------
# Soldes d'ouverture (report à nouveau) — un solde signé par compte, PAR
# EXERCICE. Balance de clôture = solde d'ouverture de l'exercice + mouvements
# de l'exercice (Débit - Crédit) enregistrés à des dates de cet exercice.
# ---------------------------------------------------------------------------
def get_opening_balance(conn, code, exercice=None):
    exercice = exercice or get_current_exercice(conn)
    row = conn.execute("SELECT solde FROM opening_balances WHERE code = ? AND exercice = ?",
                        (code, exercice)).fetchone()
    return row["solde"] if row else 0.0


def set_opening_balance(conn, code, value, exercice=None):
    exercice = exercice or get_current_exercice(conn)
    conn.execute("INSERT OR REPLACE INTO opening_balances (code, exercice, solde) VALUES (?, ?, ?)",
                 (code, exercice, value))
    conn.commit()


def list_opening_balances(conn, exercice=None):
    """IMPORTANT : jointure LEFT (pas INNER) avec le Plan comptable — un
    compte présent dans les soldes d'ouverture mais absent du Plan
    comptable doit rester visible (avec un libellé de repli), sous peine
    de fausser silencieusement le total affiché à l'écran alors que les
    données réellement enregistrées sont équilibrées."""
    exercice = exercice or get_current_exercice(conn)
    rows = conn.execute("""
        SELECT o.code, COALESCE(a.label, o.code || ' (hors Plan comptable)') AS label,
               COALESCE(a.classe, substr(o.code, 1, 1)) AS classe, o.solde
        FROM opening_balances o LEFT JOIN accounts a ON a.code = o.code
        WHERE o.solde != 0 AND o.exercice = ?
        ORDER BY o.code
    """, (exercice,)).fetchall()
    return [dict(r) for r in rows]


def total_opening_balance(conn, exercice=None):
    exercice = exercice or get_current_exercice(conn)
    row = conn.execute("SELECT COALESCE(SUM(solde), 0) t FROM opening_balances WHERE exercice = ?",
                        (exercice,)).fetchone()
    return row["t"]


def compute_ecart_diagnostic(conn, exercice=None):
    """Diagnostique la cause d'un Bilan non équilibré, en décomposant l'écart
    Actif - Passif en ses deux seules causes possibles (le calcul du Bilan
    lui-même est garanti mathématiquement équilibré — voir compute_bilan) :
    (1) la somme des soldes d'ouverture de l'exercice, qui DOIT être nulle
    par construction de la partie double (sinon, une balance de clôture N-1
    incomplète ou mal reportée) ; (2) le déséquilibre Cumul Débit / Cumul
    Crédit des écritures de la période (des écritures posées débit ≠ crédit
    dans la Saisie, typiquement issues d'un import en masse — voir
    compute_pieces_non_equilibrees() pour les localiser précisément).
    ecart_soldes_ouverture + ecart_ecritures_periode = l'écart du Bilan."""
    exercice = exercice or get_current_exercice(conn)
    ecart_ouverture = total_opening_balance(conn, exercice=exercice)
    row = conn.execute(
        """SELECT COALESCE(SUM(debit), 0) d, COALESCE(SUM(credit), 0) c FROM entries
           WHERE date >= ? AND date <= ?""",
        (f"{exercice}-01-01", f"{exercice}-12-31"),
    ).fetchone()
    ecart_ecritures = row["d"] - row["c"]
    return {
        "ecart_soldes_ouverture": ecart_ouverture,
        "ecart_ecritures_periode": ecart_ecritures,
        "ecart_total": ecart_ouverture + ecart_ecritures,
    }


def compute_pieces_non_equilibrees(conn, exercice=None, toutes_dates=False):
    """Liste les pièces (regroupement Pièce + Journal) dont le total Débit ne
    correspond pas au total Crédit — signe d'une écriture mal saisie en
    Saisie directe, ou d'un import en masse qui n'a pas respecté la partie
    double (chaque pièce doit normalement avoir Débit = Crédit). Triées par
    écart absolu décroissant. `toutes_dates=True` ignore le filtre d'exercice
    (utile si l'écriture fautive porte une date hors exercice courant)."""
    query = """SELECT piece, journal, MIN(date) date_min, MAX(date) date_max, COUNT(*) nb,
                      COALESCE(SUM(debit), 0) d, COALESCE(SUM(credit), 0) c
               FROM entries"""
    params = []
    if not toutes_dates:
        exercice = exercice or get_current_exercice(conn)
        query += " WHERE date >= ? AND date <= ?"
        params += [f"{exercice}-01-01", f"{exercice}-12-31"]
    query += """ GROUP BY piece, journal
                 HAVING ABS(SUM(debit) - SUM(credit)) >= 1
                 ORDER BY ABS(SUM(debit) - SUM(credit)) DESC"""
    result = []
    for r in conn.execute(query, params).fetchall():
        d = dict(r)
        d["ecart"] = d["d"] - d["c"]
        result.append(d)
    return result


def export_opening_balances_xlsx(conn, path, exercice=None):
    """Exporte la balance d'ouverture (soldes d'ouverture) de l'exercice en .xlsx."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    exercice = exercice or get_current_exercice(conn)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Soldes ouverture"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for i, label in enumerate(["N° Compte", "Libellé", "Solde (débit +, crédit -)"], start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    for r, row in enumerate(list_opening_balances(conn, exercice=exercice), start=2):
        ws.cell(row=r, column=1, value=row["code"])
        ws.cell(row=r, column=2, value=row["label"])
        ws.cell(row=r, column=3, value=row["solde"])
    for i, w in enumerate([14, 40, 20], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def export_opening_balances_template(path):
    """Génère un modèle .xlsx vierge (bons en-têtes + exemple équilibré) pour
    préparer une balance d'ouverture (N-1) à importer."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Soldes ouverture"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for i, label in enumerate(["N° Compte", "Libellé", "Solde (débit +, crédit -)"], start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    exemples = [
        ("101000", "CAPITAL SOCIAL", -10000000),
        ("521000", "BANQUES LOCALES", 8000000),
        ("411000", "CLIENTS", 3000000),
        ("401000", "FOURNISSEURS, DETTES EN COMPTE", -1000000),
    ]
    for r, (code, label, solde) in enumerate(exemples, start=2):
        ws.cell(row=r, column=1, value=code)
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=3, value=solde)
    total_row = len(exemples) + 3
    ws.cell(row=total_row, column=2, value="Total (doit être 0) :")
    ws.cell(row=total_row, column=3, value=f"=SUM(C2:C{len(exemples) + 1})")
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    ws.cell(row=total_row, column=3).font = Font(bold=True)
    for i, w in enumerate([14, 40, 22], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def import_opening_balances_xlsx(conn, path, exercice=None):
    """Importe une balance d'ouverture depuis un .xlsx et ÉCRASE les soldes
    d'ouverture existants pour cet exercice (les autres exercices ne sont
    pas affectés). Tolère plusieurs formats d'en-têtes : une colonne
    « Solde » signée, OU deux colonnes séparées « Solde débit »/« Solde
    crédit » (comme un export de Balance générale classique) — dans ce
    second cas, le solde est recalculé comme Débit - Crédit."""
    import openpyxl

    exercice = exercice or get_current_exercice(conn)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    code_aliases = ["n° compte", "n° de compte", "no compte", "numero compte", "num compte",
                    "compte", "code", "code compte", "intitulé du compte", "n° compte "]
    solde_aliases = ["solde (débit +, crédit -)", "solde", "solde d'ouverture", "montant",
                     "solde débit", "solde debit", "cumul débit", "cumul debit", "débit", "debit"]
    credit_aliases = ["solde crédit", "solde credit", "cumul crédit", "cumul credit", "crédit", "credit"]

    def _norm(h):
        return (str(h).strip().lower() if h is not None else "")

    header_row_idx = None
    colmap = {}
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        headers = [_norm(c.value) for c in next(ws.iter_rows(min_row=row_idx, max_row=row_idx))]
        candidate = {}
        for i, h in enumerate(headers):
            if h in code_aliases and "code" not in candidate:
                candidate["code"] = i
            if h in solde_aliases and "solde" not in candidate:
                candidate["solde"] = i
            if h in credit_aliases and "credit" not in candidate:
                candidate["credit"] = i
        if "code" in candidate and ("solde" in candidate or "credit" in candidate):
            header_row_idx = row_idx
            colmap = candidate
            break

    if header_row_idx is None:
        # Aucune ligne d'en-tête reconnue : on donne les en-têtes de la 1re ligne
        # non vide pour aider au diagnostic.
        first_headers = []
        for row_idx in range(1, min(ws.max_row, 10) + 1):
            vals = [c.value for c in next(ws.iter_rows(min_row=row_idx, max_row=row_idx)) if c.value not in (None, "")]
            if vals:
                first_headers = vals
                break
        raise ValueError(
            "Colonnes obligatoires introuvables (une colonne « N° Compte » et une colonne "
            "« Solde » — ou « Solde débit » — sont nécessaires). En-têtes détectés dans le "
            f"fichier : {first_headers if first_headers else 'aucun'}. Utilisez le bouton "
            "« Exporter la balance N-1 » pour obtenir un fichier au bon format, ou renommez "
            "vos colonnes en « N° Compte » et « Solde »."
        )

    rows = []
    warnings = []
    for r_idx, r in enumerate(ws.iter_rows(min_row=header_row_idx + 1), start=header_row_idx + 1):
        values = [c.value for c in r]
        if all(v in (None, "") for v in values):
            continue
        code_idx = colmap["code"]
        if code_idx >= len(values):
            continue
        code = str(values[code_idx] or "").strip()
        if code.endswith(".0") and code.replace(".0", "").isdigit():
            code = code[:-2]
        if not code:
            continue

        def _to_float(idx):
            if idx is None or idx >= len(values) or values[idx] in (None, ""):
                return 0.0
            try:
                return float(values[idx])
            except (TypeError, ValueError):
                return None

        if "credit" in colmap:
            debit = _to_float(colmap.get("solde"))
            credit = _to_float(colmap.get("credit"))
            if debit is None or credit is None:
                warnings.append(f"Ligne {r_idx} : solde invalide pour le compte {code}, ignoré.")
                continue
            solde = (debit or 0) - (credit or 0)
        else:
            solde = _to_float(colmap.get("solde"))
            if solde is None:
                warnings.append(f"Ligne {r_idx} : solde invalide pour le compte {code}, ignoré.")
                continue

        if solde == 0:
            continue
        if not account_exists(conn, code):
            warnings.append(f"Ligne {r_idx} : compte « {code} » introuvable dans le Plan comptable — "
                             f"importé quand même.")
        rows.append((code, exercice, solde))

    if not rows:
        raise ValueError("Le fichier ne contient aucune ligne de solde non nul.")

    conn.execute("DELETE FROM opening_balances WHERE exercice = ?", (exercice,))
    conn.executemany("INSERT INTO opening_balances (code, exercice, solde) VALUES (?, ?, ?)", rows)
    conn.commit()
    return len(rows), warnings


def load_plan_comptable(conn, json_path=None):
    """Charge le plan comptable (bundlé avec l'application) dans la base."""
    if json_path is None:
        json_path = os.path.join(_resource_dir(), "plan_comptable.json")
    with open(json_path, encoding="utf-8") as f:
        accounts = json.load(f)
    conn.executemany(
        "INSERT OR REPLACE INTO accounts (code, label, classe) VALUES (?, ?, ?)",
        [(a["code"], a["label"], a["classe"]) for a in accounts],
    )
    conn.commit()
    ensure_racine_accounts(conn)


def ensure_racine_accounts(conn):
    """Insère les comptes racines (1 chiffre pour les classes 1,2,3,5,6,7,8,9 ;
    2 chiffres 40 à 49 pour la classe 4) s'ils n'existent pas déjà, sans écraser
    un compte que l'utilisateur aurait éventuellement créé avec ce même code.
    Grâce au tri alphabétique des codes (ex. '1' < '101000'), ces racines
    apparaissent en tête de chaque groupe dans les listes de comptes."""
    racines = []
    for c in ("1", "2", "3", "5", "6", "7", "8", "9"):
        racines.append((c, RACINE_LABELS.get(c, f"Classe {c}"), c))
    for r in range(40, 50):
        code = str(r)
        racines.append((code, RACINE_LABELS.get(code, f"Racine {code}"), "4"))
    for code, label, classe in racines:
        exists = conn.execute("SELECT 1 FROM accounts WHERE code = ?", (code,)).fetchone()
        if not exists:
            conn.execute("INSERT INTO accounts (code, label, classe) VALUES (?, ?, ?)",
                         (code, f"— {label} —", classe))
    conn.commit()


# ---------------------------------------------------------------------------
# Comptes
# ---------------------------------------------------------------------------
def search_accounts(conn, query, limit=50):
    query = (query or "").strip()
    if not query:
        rows = conn.execute("SELECT code, label, classe FROM accounts ORDER BY code LIMIT ?", (limit,))
    else:
        like = f"%{query}%"
        rows = conn.execute(
            "SELECT code, label, classe FROM accounts "
            "WHERE code LIKE ? OR label LIKE ? ORDER BY code LIMIT ?",
            (f"{query}%", like, limit),
        )
    return [dict(r) for r in rows]


def to_display_date(iso_str):
    """AAAA-MM-JJ (stockage) -> JJ/MM/AAAA (affichage)."""
    if not iso_str:
        return ""
    s = str(iso_str).strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(s[:19], fmt).strftime("%d/%m/%Y")
        except ValueError:
            continue
    return s  # déjà dans un autre format ou invalide : renvoyé tel quel


def to_iso_date(display_str):
    """JJ/MM/AAAA (saisie) -> AAAA-MM-JJ (stockage). Accepte aussi AAAA-MM-JJ en entrée."""
    s = (display_str or "").strip()
    if not s:
        return ""
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s  # format non reconnu : renvoyé tel quel (l'appelant peut valider)


def get_account_label(conn, code):
    row = conn.execute("SELECT label FROM accounts WHERE code = ?", (code,)).fetchone()
    return row["label"] if row else "Compte introuvable"


def account_exists(conn, code):
    return conn.execute("SELECT 1 FROM accounts WHERE code = ?", (code,)).fetchone() is not None


def add_account(conn, code, label, classe=None):
    code = str(code).strip()
    classe = classe or (code[0] if code else "")
    conn.execute("INSERT OR REPLACE INTO accounts (code, label, classe) VALUES (?, ?, ?)",
                 (code, label.strip(), classe))
    conn.commit()


def delete_account(conn, code):
    conn.execute("DELETE FROM accounts WHERE code = ?", (code,))
    conn.commit()


def export_plan_comptable_xlsx(conn, path):
    """Exporte tout le Plan comptable (code, libellé, classe) en .xlsx."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Plan comptable"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for i, label in enumerate(["N° Compte", "Libellé", "Classe"], start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    for r, a in enumerate(conn.execute("SELECT code, label, classe FROM accounts ORDER BY code"), start=2):
        ws.cell(row=r, column=1, value=a["code"])
        ws.cell(row=r, column=2, value=a["label"])
        ws.cell(row=r, column=3, value=a["classe"])
    for i, w in enumerate([14, 45, 10], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def import_plan_comptable_xlsx(conn, path):
    """Importe un Plan comptable depuis un .xlsx et ÉCRASE l'ancien plan
    (toutes les fiches auxiliaires clients/fournisseurs et les écritures
    existantes ne sont PAS supprimées, mais leurs comptes ne seront plus
    reconnus s'ils ne figurent pas dans le nouveau plan)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in header_cells]
    aliases = {"code": ["n° compte", "code", "compte"], "label": ["libellé", "libelle"],
               "classe": ["classe"]}
    colmap = {}
    for key, alist in aliases.items():
        for i, h in enumerate(headers):
            if h in alist:
                colmap[key] = i
                break
    if "code" not in colmap or "label" not in colmap:
        raise ValueError("Colonnes obligatoires introuvables (« N° Compte » et « Libellé »).")

    rows = []
    for r in ws.iter_rows(min_row=2):
        values = [c.value for c in r]
        if all(v in (None, "") for v in values):
            continue
        code = str(values[colmap["code"]] or "").strip()
        label = str(values[colmap["label"]] or "").strip()
        if not code or not label:
            continue
        classe = None
        if "classe" in colmap and colmap["classe"] < len(values) and values[colmap["classe"]]:
            classe = str(values[colmap["classe"]]).strip()
        if not classe:
            classe = code[:2] if code[:1] == "4" else code[:1]
        rows.append((code, label, classe))

    if not rows:
        raise ValueError("Le fichier ne contient aucun compte valide.")

    conn.execute("DELETE FROM accounts")
    conn.executemany("INSERT INTO accounts (code, label, classe) VALUES (?, ?, ?)", rows)
    conn.commit()
    ensure_racine_accounts(conn)
    return len(rows)


# ---------------------------------------------------------------------------
# Plans auxiliaires : analytique, budgétaire, bailleurs de fonds
# (même logique CRUD simple pour les 3, table dédiée chacun)
# ---------------------------------------------------------------------------
def _plan_list(conn, table, extra_cols=""):
    cols = "code, label" + (f", {extra_cols}" if extra_cols else "")
    rows = conn.execute(f"SELECT {cols} FROM {table} ORDER BY code").fetchall()
    return [dict(r) for r in rows]


def _plan_exists(conn, table, code):
    return conn.execute(f"SELECT 1 FROM {table} WHERE code = ?", (code,)).fetchone() is not None


def _plan_delete(conn, table, code):
    conn.execute(f"DELETE FROM {table} WHERE code = ?", (code,))
    conn.commit()


def list_analytic_codes(conn):
    return _plan_list(conn, "analytic_codes", extra_cols="unite")


def analytic_code_exists(conn, code):
    return _plan_exists(conn, "analytic_codes", code)


def get_analytic_code_unite(conn, code):
    row = conn.execute("SELECT unite FROM analytic_codes WHERE code = ?", (code,)).fetchone()
    return row["unite"] if row else None


def add_analytic_code(conn, code, label, unite=None):
    """`unite` (ex. 'L' pour litre, 'Kw' pour kilowatt, 'H' pour heure) :
    conservée telle quelle si non précisée, pour ne pas écraser une unité
    déjà définie lors d'une simple modification de libellé."""
    code = code.strip()
    if unite is None:
        row = conn.execute("SELECT unite FROM analytic_codes WHERE code = ?", (code,)).fetchone()
        if row:
            unite = row["unite"]
    conn.execute("INSERT OR REPLACE INTO analytic_codes (code, label, unite) VALUES (?, ?, ?)",
                 (code, label.strip(), unite))
    conn.commit()


def delete_analytic_code(conn, code):
    _plan_delete(conn, "analytic_codes", code)


def list_budget_codes(conn):
    return _plan_list(conn, "budget_codes", extra_cols="montant")


def budget_code_exists(conn, code):
    return _plan_exists(conn, "budget_codes", code)


def add_budget_code(conn, code, label, montant=0):
    conn.execute("INSERT OR REPLACE INTO budget_codes (code, label, montant) VALUES (?, ?, ?)",
                 (code.strip(), label.strip(), montant or 0))
    conn.commit()


def list_taux_tva(conn):
    return _plan_list(conn, "taux_tva", extra_cols="montant, compte")


def taux_tva_exists(conn, code):
    return _plan_exists(conn, "taux_tva", code)


def add_taux_tva(conn, code, label, montant=0, compte=None):
    conn.execute("INSERT OR REPLACE INTO taux_tva (code, label, montant, compte) VALUES (?, ?, ?, ?)",
                 (code.strip(), label.strip(), montant or 0, compte or None))
    conn.commit()


def delete_taux_tva(conn, code):
    return _plan_delete(conn, "taux_tva", code)


def list_taux_retenue(conn):
    return _plan_list(conn, "taux_retenue", extra_cols="montant, compte")


def taux_retenue_exists(conn, code):
    return _plan_exists(conn, "taux_retenue", code)


def add_taux_retenue(conn, code, label, montant=0, compte=None):
    conn.execute("INSERT OR REPLACE INTO taux_retenue (code, label, montant, compte) VALUES (?, ?, ?, ?)",
                 (code.strip(), label.strip(), montant or 0, compte or None))
    conn.commit()


def delete_taux_retenue(conn, code):
    return _plan_delete(conn, "taux_retenue", code)


# Catégories de retenues à la source reconnues par la DGI du Burkina Faso
# (voir la liste officielle des formulaires de déclaration sur dgi.bf) — SANS
# taux ni compte pré-remplis : les taux et seuils exacts dépendent du régime
# du fournisseur, du type de prestation et de la loi de finances en vigueur
# (ex. la retenue à la source de la TVA est passée de 20% à 30% en 2026),
# et doivent être vérifiés par l'utilisateur avant usage.
SUGGESTIONS_RETENUE = [
    ("RET-BIC", "Retenue BIC (fournisseurs non attributaires)"),
    ("RET-IS", "Retenue Impôt sur les Sociétés (IS)"),
    ("RET-TVA", "Retenue à la source de la TVA"),
    ("RET-LOYER", "Retenue sur loyers d'immeuble"),
    ("RET-PRESTA-BF", "Retenue sur sommes versées aux prestataires établis au Burkina Faso"),
    ("RET-PRESTA-ETR", "Retenue sur sommes versées aux personnes sans installation professionnelle au Burkina Faso"),
    ("RET-CMD-PUB", "Retenue sur commandes publiques"),
]


def ajouter_taux_retenue_suggeres(conn):
    """Ajoute les catégories de retenue à la source courantes (reconnues par
    la DGI du Burkina Faso) qui n'existent pas encore, à 0% et sans compte —
    à compléter par l'utilisateur avec le taux et le compte fiscal exacts,
    SANS écraser une catégorie déjà personnalisée. Retourne le nombre de
    catégories effectivement ajoutées."""
    ajoutes = 0
    for code, label in SUGGESTIONS_RETENUE:
        if not taux_retenue_exists(conn, code):
            add_taux_retenue(conn, code, label, montant=0, compte=None)
            ajoutes += 1
    return ajoutes


def delete_budget_code(conn, code):
    _plan_delete(conn, "budget_codes", code)


def list_donor_codes(conn):
    return _plan_list(conn, "donor_codes")


def donor_code_exists(conn, code):
    return _plan_exists(conn, "donor_codes", code)


def add_donor_code(conn, code, label):
    conn.execute("INSERT OR REPLACE INTO donor_codes (code, label) VALUES (?, ?)",
                 (code.strip(), label.strip()))
    conn.commit()


def delete_donor_code(conn, code):
    _plan_delete(conn, "donor_codes", code)


def _export_plan_generic_xlsx(conn, path, table, title, has_montant=False, has_unite=False, has_compte=False):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    headers = (["Code", "Libellé"] + (["Montant"] if has_montant else []) + (["Unité"] if has_unite else [])
               + (["Compte"] if has_compte else []))
    for i, label in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    cols = ("code, label" + (", montant" if has_montant else "") + (", unite" if has_unite else "")
            + (", compte" if has_compte else ""))
    for r, row in enumerate(conn.execute(f"SELECT {cols} FROM {table} ORDER BY code"), start=2):
        ws.cell(row=r, column=1, value=row["code"])
        ws.cell(row=r, column=2, value=row["label"])
        col = 3
        if has_montant:
            ws.cell(row=r, column=col, value=row["montant"])
            col += 1
        if has_unite:
            ws.cell(row=r, column=col, value=row["unite"])
            col += 1
        if has_compte:
            ws.cell(row=r, column=col, value=row["compte"])
    widths = [16, 40] + ([16] if has_montant else []) + ([12] if has_unite else []) + ([16] if has_compte else [])
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def _import_plan_generic_xlsx(conn, path, table, has_montant=False, has_unite=False, has_compte=False):
    """Importe un plan (code/libellé[/montant][/unité][/compte]) depuis un
    .xlsx et ÉCRASE l'ancien contenu de la table."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in header_cells]
    aliases = {"code": ["code"], "label": ["libellé", "libelle"], "montant": ["montant"],
               "unite": ["unité", "unite"], "compte": ["compte"]}
    colmap = {}
    for key, alist in aliases.items():
        for i, h in enumerate(headers):
            if h in alist:
                colmap[key] = i
                break
    if "code" not in colmap or "label" not in colmap:
        raise ValueError("Colonnes obligatoires introuvables (« Code » et « Libellé »).")

    rows = []
    for r in ws.iter_rows(min_row=2):
        values = [c.value for c in r]
        if all(v in (None, "") for v in values):
            continue
        code = str(values[colmap["code"]] or "").strip()
        label = str(values[colmap["label"]] or "").strip()
        if not code or not label:
            continue
        row = [code, label]
        if has_montant:
            montant = 0.0
            if "montant" in colmap and colmap["montant"] < len(values):
                try:
                    montant = float(values[colmap["montant"]] or 0)
                except (TypeError, ValueError):
                    montant = 0.0
            row.append(montant)
        if has_unite:
            unite = None
            if "unite" in colmap and colmap["unite"] < len(values):
                unite = str(values[colmap["unite"]] or "").strip() or None
            row.append(unite)
        if has_compte:
            compte = None
            if "compte" in colmap and colmap["compte"] < len(values):
                compte = str(values[colmap["compte"]] or "").strip() or None
            row.append(compte)
        rows.append(tuple(row))

    if not rows:
        raise ValueError("Le fichier ne contient aucune ligne valide.")

    conn.execute(f"DELETE FROM {table}")
    if has_montant and has_compte:
        conn.executemany(f"INSERT INTO {table} (code, label, montant, compte) VALUES (?, ?, ?, ?)", rows)
    elif has_montant:
        conn.executemany(f"INSERT INTO {table} (code, label, montant) VALUES (?, ?, ?)", rows)
    elif has_unite:
        conn.executemany(f"INSERT INTO {table} (code, label, unite) VALUES (?, ?, ?)", rows)
    else:
        conn.executemany(f"INSERT INTO {table} (code, label) VALUES (?, ?)", rows)
    conn.commit()
    return len(rows)


def export_analytic_codes_xlsx(conn, path):
    return _export_plan_generic_xlsx(conn, path, "analytic_codes", "Plan analytique", has_unite=True)


def import_analytic_codes_xlsx(conn, path):
    return _import_plan_generic_xlsx(conn, path, "analytic_codes", has_unite=True)


def export_budget_codes_xlsx(conn, path):
    return _export_plan_generic_xlsx(conn, path, "budget_codes", "Plan budgétaire", has_montant=True)


def import_budget_codes_xlsx(conn, path):
    return _import_plan_generic_xlsx(conn, path, "budget_codes", has_montant=True)


def export_taux_tva_xlsx(conn, path):
    return _export_plan_generic_xlsx(conn, path, "taux_tva", "Taux de TVA", has_montant=True, has_compte=True)


def import_taux_tva_xlsx(conn, path):
    return _import_plan_generic_xlsx(conn, path, "taux_tva", has_montant=True, has_compte=True)


def export_taux_retenue_xlsx(conn, path):
    return _export_plan_generic_xlsx(conn, path, "taux_retenue", "Retenues à la source", has_montant=True, has_compte=True)


def import_taux_retenue_xlsx(conn, path):
    return _import_plan_generic_xlsx(conn, path, "taux_retenue", has_montant=True, has_compte=True)


def export_donor_codes_xlsx(conn, path):
    return _export_plan_generic_xlsx(conn, path, "donor_codes", "Plan bailleurs")


def import_donor_codes_xlsx(conn, path):
    return _import_plan_generic_xlsx(conn, path, "donor_codes")


# ---------------------------------------------------------------------------
# Équilibrage d'une pièce comptable
# ---------------------------------------------------------------------------
def get_piece_balance(conn, piece):
    """Retourne (total_debit, total_credit) pour toutes les lignes d'une pièce donnée."""
    row = conn.execute(
        "SELECT COALESCE(SUM(debit),0) d, COALESCE(SUM(credit),0) c FROM entries WHERE piece = ?",
        (piece,),
    ).fetchone()
    return row["d"], row["c"]


# ---------------------------------------------------------------------------
# Fournisseurs (fiche auxiliaire)
# ---------------------------------------------------------------------------
def list_fournisseurs(conn, query=None):
    if query:
        like = f"%{query}%"
        rows = conn.execute(
            "SELECT * FROM fournisseurs WHERE code LIKE ? OR raison_sociale LIKE ? ORDER BY code",
            (f"{query}%", like),
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM fournisseurs ORDER BY code").fetchall()
    return [dict(r) for r in rows]


def fournisseur_exists(conn, code):
    return conn.execute("SELECT 1 FROM fournisseurs WHERE code = ?", (code,)).fetchone() is not None


def get_fournisseur(conn, code):
    row = conn.execute("SELECT * FROM fournisseurs WHERE code = ?", (code,)).fetchone()
    return dict(row) if row else None


def add_fournisseur(conn, code, raison_sociale, contact="", telephone="", adresse="",
                     delai_paiement_jours=30, delai_livraison_jours=15):
    conn.execute(
        """INSERT OR REPLACE INTO fournisseurs
           (code, raison_sociale, contact, telephone, adresse, delai_paiement_jours, delai_livraison_jours)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (code.strip(), raison_sociale.strip(), contact, telephone, adresse,
         int(delai_paiement_jours or 0), int(delai_livraison_jours or 0)),
    )
    conn.commit()


def delete_fournisseur(conn, code):
    conn.execute("DELETE FROM fournisseurs WHERE code = ?", (code,))
    conn.commit()


FOURNISSEUR_IMPORT_COLUMNS = [
    ("code", "Code fournisseur", ["code", "code fournisseur"]),
    ("raison_sociale", "Raison sociale", ["raison sociale", "nom", "dénomination"]),
    ("contact", "Contact", ["contact"]),
    ("telephone", "Téléphone", ["téléphone", "telephone", "tel"]),
    ("adresse", "Adresse", ["adresse"]),
    ("delai_paiement_jours", "Délai paiement (jours)", ["délai paiement (jours)", "delai paiement", "délai paiement"]),
    ("delai_livraison_jours", "Délai livraison (jours)", ["délai livraison (jours)", "delai livraison", "délai livraison"]),
]


def export_fournisseurs_template(path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fournisseurs"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for i, (_, label, _) in enumerate(FOURNISSEUR_IMPORT_COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    example = ["FRS-0001", "Etablissements Dupont", "M. Dupont", "+226 70 00 00 00",
               "Ouagadougou", 30, 15]
    for i, val in enumerate(example, start=1):
        ws.cell(row=2, column=i, value=val)
    for i, w in enumerate([14, 30, 18, 16, 26, 16, 16], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def import_fournisseurs_from_xlsx(conn, path):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in header_cells]

    colmap = {}
    for key, _, aliases in FOURNISSEUR_IMPORT_COLUMNS:
        for i, h in enumerate(headers):
            if h in aliases:
                colmap[key] = i
                break

    if "code" not in colmap or "raison_sociale" not in colmap:
        raise ValueError(
            "Colonnes obligatoires introuvables (« Code fournisseur » et « Raison sociale »). "
            "Utilisez le bouton « Télécharger un modèle »."
        )

    imported, warnings = 0, []

    def get(values, key, default=None):
        idx = colmap.get(key)
        if idx is None or idx >= len(values):
            return default
        return values[idx]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [c.value for c in row]
        if all(v in (None, "") for v in values):
            continue
        code = str(get(values, "code") or "").strip()
        raison = str(get(values, "raison_sociale") or "").strip()
        if not code or not raison:
            warnings.append(f"Ligne {row_idx} : code ou raison sociale manquant, ligne ignorée.")
            continue
        try:
            dp = int(get(values, "delai_paiement_jours", 30) or 30)
        except (TypeError, ValueError):
            dp = 30
        try:
            dl = int(get(values, "delai_livraison_jours", 15) or 15)
        except (TypeError, ValueError):
            dl = 15
        add_fournisseur(conn, code, raison, str(get(values, "contact") or ""),
                         str(get(values, "telephone") or ""), str(get(values, "adresse") or ""),
                         dp, dl)
        imported += 1
    return imported, warnings


def compute_achats_par_fournisseur(conn, date_from=None, date_to=None):
    """Total Débit/Crédit/Solde par fournisseur, sur les seuls comptes fournisseurs
    (racine 40, tous les comptes 40xxxx) tagués avec le code fournisseur — le solde reflète ce qui reste
    dû (négatif = nous devons au fournisseur), sur une plage de dates optionnelle."""
    query = """
        SELECT e.fournisseur_code AS code,
               COALESCE(f.raison_sociale, e.fournisseur_code) AS raison_sociale,
               COALESCE(SUM(e.debit), 0) AS debit,
               COALESCE(SUM(e.credit), 0) AS credit
        FROM entries e
        LEFT JOIN fournisseurs f ON f.code = e.fournisseur_code
        WHERE e.fournisseur_code IS NOT NULL AND e.fournisseur_code != ''
          AND e.compte LIKE '40%'
    """
    params = []
    if date_from:
        query += " AND e.date >= ?"
        params.append(date_from)
    if date_to:
        query += " AND e.date <= ?"
        params.append(date_to)
    query += " GROUP BY e.fournisseur_code, raison_sociale ORDER BY raison_sociale"
    rows = conn.execute(query, params).fetchall()
    result = []
    total_debit = total_credit = 0.0
    for r in rows:
        solde = r["debit"] - r["credit"]
        result.append({"code": r["code"], "raison_sociale": r["raison_sociale"],
                        "debit": r["debit"], "credit": r["credit"], "solde": solde})
        total_debit += r["debit"]
        total_credit += r["credit"]
    return result, total_debit, total_credit


# ---------------------------------------------------------------------------
# Contrats / commandes fournisseurs — suivi des délais de paiement et de
# livraison, avec détection des dépassements.
# ---------------------------------------------------------------------------
def add_commande(conn, fournisseur_code, piece, libelle, montant, date_commande,
                  date_livraison_prevue=None, date_paiement_prevue_override=None):
    fournisseur = get_fournisseur(conn, fournisseur_code)
    delai_paiement = fournisseur["delai_paiement_jours"] if fournisseur else 30
    delai_livraison = fournisseur["delai_livraison_jours"] if fournisseur else 15
    base = datetime.strptime(date_commande, "%Y-%m-%d")
    if not date_livraison_prevue:
        date_livraison_prevue = (base + timedelta(days=delai_livraison)).strftime("%Y-%m-%d")
    date_echeance_paiement = date_paiement_prevue_override or (
        base + timedelta(days=delai_paiement)).strftime("%Y-%m-%d")
    conn.execute(
        """INSERT INTO commandes_fournisseurs
           (fournisseur_code, piece, libelle, montant, date_commande, date_livraison_prevue,
            date_echeance_paiement)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (fournisseur_code, piece, libelle, montant, date_commande, date_livraison_prevue,
         date_echeance_paiement),
    )
    conn.commit()


def update_commande(conn, commande_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE commandes_fournisseurs SET {cols} WHERE id = ?", (*fields.values(), commande_id))
    conn.commit()


def delete_commande(conn, commande_id):
    conn.execute("DELETE FROM commandes_fournisseurs WHERE id = ?", (commande_id,))
    conn.commit()


def list_commandes(conn, fournisseur_code=None, date_from=None, date_to=None):
    query = """SELECT c.*, COALESCE(f.raison_sociale, c.fournisseur_code) AS raison_sociale
               FROM commandes_fournisseurs c LEFT JOIN fournisseurs f ON f.code = c.fournisseur_code
               WHERE 1=1"""
    params = []
    if fournisseur_code:
        query += " AND c.fournisseur_code = ?"
        params.append(fournisseur_code)
    if date_from:
        query += " AND c.date_commande >= ?"
        params.append(date_from)
    if date_to:
        query += " AND c.date_commande <= ?"
        params.append(date_to)
    query += " ORDER BY c.date_commande DESC, c.id DESC"
    rows = [dict(r) for r in conn.execute(query, params).fetchall()]
    today = date.today().strftime("%Y-%m-%d")
    for r in rows:
        # Statut livraison
        if r["date_livraison_reelle"]:
            retard = (datetime.strptime(r["date_livraison_reelle"], "%Y-%m-%d")
                      - datetime.strptime(r["date_livraison_prevue"], "%Y-%m-%d")).days if r["date_livraison_prevue"] else 0
            r["statut_livraison"] = f"Livré (retard {retard} j)" if retard > 0 else "Livré à temps"
            r["depassement_livraison"] = retard > 0
        elif r["date_livraison_prevue"] and today > r["date_livraison_prevue"]:
            retard = (datetime.strptime(today, "%Y-%m-%d")
                      - datetime.strptime(r["date_livraison_prevue"], "%Y-%m-%d")).days
            r["statut_livraison"] = f"EN RETARD ({retard} j)"
            r["depassement_livraison"] = True
        else:
            r["statut_livraison"] = "En attente"
            r["depassement_livraison"] = False
        # Statut paiement
        if r["date_paiement_reel"]:
            retard = (datetime.strptime(r["date_paiement_reel"], "%Y-%m-%d")
                      - datetime.strptime(r["date_echeance_paiement"], "%Y-%m-%d")).days if r["date_echeance_paiement"] else 0
            r["statut_paiement"] = f"Payé (retard {retard} j)" if retard > 0 else "Payé à temps"
            r["depassement_paiement"] = retard > 0
        elif r["date_echeance_paiement"] and today > r["date_echeance_paiement"]:
            retard = (datetime.strptime(today, "%Y-%m-%d")
                      - datetime.strptime(r["date_echeance_paiement"], "%Y-%m-%d")).days
            r["statut_paiement"] = f"EN RETARD ({retard} j)"
            r["depassement_paiement"] = True
        else:
            r["statut_paiement"] = "En attente"
            r["depassement_paiement"] = False
    return rows


# ---------------------------------------------------------------------------
# Clients (fiche auxiliaire)
# ---------------------------------------------------------------------------
def list_clients(conn, query=None):
    if query:
        like = f"%{query}%"
        rows = conn.execute(
            "SELECT * FROM clients WHERE code LIKE ? OR raison_sociale LIKE ? ORDER BY code",
            (f"{query}%", like),
        ).fetchall()
    else:
        rows = conn.execute("SELECT * FROM clients ORDER BY code").fetchall()
    return [dict(r) for r in rows]


def client_exists(conn, code):
    return conn.execute("SELECT 1 FROM clients WHERE code = ?", (code,)).fetchone() is not None


def get_client(conn, code):
    row = conn.execute("SELECT * FROM clients WHERE code = ?", (code,)).fetchone()
    return dict(row) if row else None


def add_client(conn, code, raison_sociale, contact="", telephone="", adresse="",
                delai_paiement_jours=30):
    conn.execute(
        """INSERT OR REPLACE INTO clients
           (code, raison_sociale, contact, telephone, adresse, delai_paiement_jours)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (code.strip(), raison_sociale.strip(), contact, telephone, adresse, int(delai_paiement_jours or 0)),
    )
    conn.commit()


def delete_client(conn, code):
    conn.execute("DELETE FROM clients WHERE code = ?", (code,))
    conn.commit()


CLIENT_IMPORT_COLUMNS = [
    ("code", "Code client", ["code", "code client"]),
    ("raison_sociale", "Raison sociale", ["raison sociale", "nom", "dénomination"]),
    ("contact", "Contact", ["contact"]),
    ("telephone", "Téléphone", ["téléphone", "telephone", "tel"]),
    ("adresse", "Adresse", ["adresse"]),
    ("delai_paiement_jours", "Délai paiement (jours)", ["délai paiement (jours)", "delai paiement", "délai paiement"]),
]


def export_clients_template(path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Clients"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for i, (_, label, _) in enumerate(CLIENT_IMPORT_COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    example = ["CLI-0001", "Société ABC", "Mme Traoré", "+226 70 11 11 11", "Ouagadougou", 30]
    for i, val in enumerate(example, start=1):
        ws.cell(row=2, column=i, value=val)
    for i, w in enumerate([14, 30, 18, 16, 26, 18], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def import_clients_from_xlsx(conn, path):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in header_cells]

    colmap = {}
    for key, _, aliases in CLIENT_IMPORT_COLUMNS:
        for i, h in enumerate(headers):
            if h in aliases:
                colmap[key] = i
                break

    if "code" not in colmap or "raison_sociale" not in colmap:
        raise ValueError(
            "Colonnes obligatoires introuvables (« Code client » et « Raison sociale »). "
            "Utilisez le bouton « Télécharger un modèle »."
        )

    imported, warnings = 0, []

    def get(values, key, default=None):
        idx = colmap.get(key)
        if idx is None or idx >= len(values):
            return default
        return values[idx]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [c.value for c in row]
        if all(v in (None, "") for v in values):
            continue
        code = str(get(values, "code") or "").strip()
        raison = str(get(values, "raison_sociale") or "").strip()
        if not code or not raison:
            warnings.append(f"Ligne {row_idx} : code ou raison sociale manquant, ligne ignorée.")
            continue
        try:
            dp = int(get(values, "delai_paiement_jours", 30) or 30)
        except (TypeError, ValueError):
            dp = 30
        add_client(conn, code, raison, str(get(values, "contact") or ""),
                   str(get(values, "telephone") or ""), str(get(values, "adresse") or ""), dp)
        imported += 1
    return imported, warnings


def compute_ventes_par_client(conn, date_from=None, date_to=None):
    """Total Débit/Crédit/Solde par client, sur les seuls comptes clients
    (racine 41, tous les comptes 41xxxx)
    tagués avec le code client — solde positif = montant restant dû par le client
    (à recouvrer), sur une plage de dates optionnelle."""
    query = """
        SELECT e.client_code AS code,
               COALESCE(c.raison_sociale, e.client_code) AS raison_sociale,
               COALESCE(SUM(e.debit), 0) AS debit,
               COALESCE(SUM(e.credit), 0) AS credit
        FROM entries e
        LEFT JOIN clients c ON c.code = e.client_code
        WHERE e.client_code IS NOT NULL AND e.client_code != ''
          AND e.compte LIKE '41%'
    """
    params = []
    if date_from:
        query += " AND e.date >= ?"
        params.append(date_from)
    if date_to:
        query += " AND e.date <= ?"
        params.append(date_to)
    query += " GROUP BY e.client_code, raison_sociale ORDER BY raison_sociale"
    rows = conn.execute(query, params).fetchall()
    result = []
    total_debit = total_credit = 0.0
    for r in rows:
        solde = r["debit"] - r["credit"]
        result.append({"code": r["code"], "raison_sociale": r["raison_sociale"],
                        "debit": r["debit"], "credit": r["credit"], "solde": solde})
        total_debit += r["debit"]
        total_credit += r["credit"]
    return result, total_debit, total_credit


# ---------------------------------------------------------------------------
# Recouvrement — factures clients avec échéance et retard de paiement
# ---------------------------------------------------------------------------
def add_facture(conn, client_code, piece, libelle, montant, date_facture,
                 date_echeance_override=None):
    if not montant or montant <= 0:
        raise ValueError("Le montant de la facture doit être strictement positif.")
    client = get_client(conn, client_code)
    delai_paiement = client["delai_paiement_jours"] if client else 30
    base = datetime.strptime(date_facture, "%Y-%m-%d")
    date_echeance = date_echeance_override or (base + timedelta(days=delai_paiement)).strftime("%Y-%m-%d")
    conn.execute(
        """INSERT INTO factures_clients
           (client_code, piece, libelle, montant, date_facture, date_echeance_paiement)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (client_code, piece, libelle, montant, date_facture, date_echeance),
    )
    conn.commit()


def enregistrer_paiement_facture(conn, facture_id, date_paiement_reel, compte_reglement, exercice=None):
    """Enregistre le règlement d'une facture client (Recouvrement) ET
    comptabilise l'encaissement : une écriture équilibrée Débit compte
    banque/caisse choisi / Crédit compte client (411000, même convention que
    la Facturation), pour le montant de la facture. Ne comptabilise
    qu'UNE SEULE FOIS par facture (garde-fou `reglement_comptabilise`) —
    modifier ensuite la date de paiement ne repostera pas une seconde
    écriture."""
    facture = conn.execute("SELECT * FROM factures_clients WHERE id = ?", (facture_id,)).fetchone()
    if not facture:
        raise ValueError("Facture introuvable.")
    if not compte_reglement or not account_exists(conn, compte_reglement):
        raise ValueError(f"Le compte de règlement « {compte_reglement} » n'existe pas.")
    if account_racine(compte_reglement) != "5":
        raise ValueError("Le compte de règlement doit être un compte de trésorerie (classe 5 — banque ou caisse).")

    client = get_client(conn, facture["client_code"])
    tiers_label = client["raison_sociale"] if client else facture["client_code"]
    piece = facture["piece"] or f"REGL-{facture_id}"

    if not facture["reglement_comptabilise"]:
        _check_exercice_editable(conn, date_paiement_reel)
        add_entry(conn, date_paiement_reel, piece, "BQ", compte_reglement, tiers_label,
                  f"Règlement facture {piece}", facture["montant"], 0)
        add_entry(conn, date_paiement_reel, piece, "BQ", "411000", tiers_label,
                  f"Règlement facture {piece}", 0, facture["montant"], client_code=facture["client_code"])

    conn.execute(
        """UPDATE factures_clients SET date_paiement_reel = ?, compte_reglement = ?, reglement_comptabilise = 1
           WHERE id = ?""",
        (date_paiement_reel, compte_reglement, facture_id),
    )
    conn.commit()


def update_facture(conn, facture_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE factures_clients SET {cols} WHERE id = ?", (*fields.values(), facture_id))
    conn.commit()


def delete_facture(conn, facture_id):
    conn.execute("DELETE FROM factures_clients WHERE id = ?", (facture_id,))
    conn.commit()


def list_factures(conn, client_code=None, date_from=None, date_to=None):
    query = """SELECT f.*, COALESCE(c.raison_sociale, f.client_code) AS raison_sociale
               FROM factures_clients f LEFT JOIN clients c ON c.code = f.client_code
               WHERE 1=1"""
    params = []
    if client_code:
        query += " AND f.client_code = ?"
        params.append(client_code)
    if date_from:
        query += " AND f.date_facture >= ?"
        params.append(date_from)
    if date_to:
        query += " AND f.date_facture <= ?"
        params.append(date_to)
    query += " ORDER BY f.date_facture DESC, f.id DESC"
    rows = [dict(r) for r in conn.execute(query, params).fetchall()]
    today = date.today().strftime("%Y-%m-%d")
    for r in rows:
        if r["date_paiement_reel"]:
            retard = (datetime.strptime(r["date_paiement_reel"], "%Y-%m-%d")
                      - datetime.strptime(r["date_echeance_paiement"], "%Y-%m-%d")).days if r["date_echeance_paiement"] else 0
            r["statut_paiement"] = f"Payé (retard {retard} j)" if retard > 0 else "Payé à temps"
            r["depassement"] = retard > 0
        elif r["date_echeance_paiement"] and today > r["date_echeance_paiement"]:
            retard = (datetime.strptime(today, "%Y-%m-%d")
                      - datetime.strptime(r["date_echeance_paiement"], "%Y-%m-%d")).days
            r["statut_paiement"] = f"EN RETARD ({retard} j)"
            r["depassement"] = True
        else:
            r["statut_paiement"] = "En attente"
            r["depassement"] = False
    return rows


# ---------------------------------------------------------------------------
def _check_exercice_editable(conn, date_str):
    exercice = _exercice_of_date(date_str)
    if exercice and is_exercice_cloture(conn, exercice):
        raise ValueError(
            f"L'exercice {exercice} est clôturé : impossible d'ajouter, modifier ou supprimer "
            f"une écriture datée de cet exercice."
        )


def add_entry(conn, date_str, piece, journal, compte, tiers, libelle, debit, credit,
              flux_code="", analytic_code="", budget_code="", donor_code="", quantite=0,
              fournisseur_code="", client_code=""):
    _check_exercice_editable(conn, date_str)
    conn.execute(
        """INSERT INTO entries (date, piece, journal, compte, tiers, libelle, debit, credit,
                                 flux_code, analytic_code, budget_code, donor_code, quantite,
                                 fournisseur_code, client_code)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (date_str, piece, journal, compte, tiers, libelle, debit or 0, credit or 0,
         flux_code, analytic_code, budget_code, donor_code, quantite or 0, fournisseur_code, client_code),
    )
    conn.commit()


def add_ecriture_multi_lignes(conn, date_str, piece, journal, lignes, tiers="",
                               fournisseur_code="", client_code="", budget_code="", donor_code="",
                               compte_stock_global=None, quantite_stock_global=0,
                               sens_stock_global="entree"):
    """Écriture comptable à un nombre libre de lignes, chacune au débit OU
    au crédit (jamais les deux sur la même ligne) — la vraie saisie
    multi-lignes façon journal général : autant de comptes que nécessaire
    des deux côtés (ex. plusieurs charges de classe 6 au débit, réglées par
    plusieurs comptes de trésorerie au crédit). Reste équilibrée par
    construction : le total des débits DOIT être strictement égal au total
    des crédits, sous peine de refus — impossible d'enregistrer un
    déséquilibre.

    `lignes` : liste de dicts {compte, libelle, debit, credit, quantite,
    analytic_code} — chaque ligne renseigne SOIT debit SOIT credit (l'autre
    valant 0 ou absent).

    Deux façons de générer le mouvement de stock automatique :
    - `compte_stock_global` renseigné : TOUTES les lignes au débit de cette
      écriture sont regroupées en UN SEUL mouvement de stock, à la quantité
      `quantite_stock_global` (la quantité réellement concernée, pas une
      quantité par ligne) — le sens dépend de `sens_stock_global` :
        - "entree" (achat — matière première + transport + douane...) :
          le COÛT du mouvement = somme des montants des lignes débit ;
          le stock AUGMENTE.
        - "sortie" (vente à un ou plusieurs clients réglés en une fois...) :
          le COÛT du mouvement = quantité × coût unitaire moyen ACTUEL du
          stock (même logique que pour une vente simple) ; le stock
          DIMINUE. Les lignes débit de l'écriture (ex. comptes clients) ne
          servent alors PAS de base de calcul du coût — c'est le coût de
          revient réel du stock qui est utilisé, pas le prix de vente.
    - `compte_stock_global` absent (None) : comportement ligne par ligne —
      chaque ligne débit qui renseigne sa PROPRE quantité et porte un
      compte d'achat lié à un stock (601x/602x) génère sa propre entrée de
      stock indépendante ; idem au crédit pour une sortie de stock
      (701x/702x). Adapté à des lignes réellement indépendantes (plusieurs
      achats sans lien entre eux dans la même écriture)."""
    if len(lignes) < 2:
        raise ValueError("Une écriture multi-lignes nécessite au moins 2 lignes (au moins une au débit, une au crédit).")
    if fournisseur_code and not fournisseur_exists(conn, fournisseur_code):
        raise ValueError(f"Le fournisseur « {fournisseur_code} » n'existe pas dans la liste des fournisseurs.")
    if client_code and not client_exists(conn, client_code):
        raise ValueError(f"Le client « {client_code} » n'existe pas dans la liste des clients.")
    if budget_code and not budget_code_exists(conn, budget_code):
        raise ValueError(f"Le code budgétaire « {budget_code} » n'existe pas dans le plan budgétaire.")
    if donor_code and not donor_code_exists(conn, donor_code):
        raise ValueError(f"Le code bailleur « {donor_code} » n'existe pas dans le plan des bailleurs.")
    total_debit = total_credit = 0.0
    for l in lignes:
        if not l.get("compte"):
            raise ValueError("Chaque ligne doit avoir un compte.")
        if not account_exists(conn, l["compte"]):
            raise ValueError(
                f"Le compte « {l['compte']} » n'existe pas dans le plan comptable — "
                f"choisissez un compte dans la liste plutôt que de le saisir librement."
            )
        ligne_analytic = l.get("analytic_code")
        if ligne_analytic and not analytic_code_exists(conn, ligne_analytic):
            raise ValueError(f"Le code analytique « {ligne_analytic} » n'existe pas dans le plan analytique.")
        ligne_fournisseur = l.get("fournisseur_code")
        if ligne_fournisseur and not fournisseur_exists(conn, ligne_fournisseur):
            raise ValueError(f"Le fournisseur « {ligne_fournisseur} » n'existe pas dans la liste des fournisseurs.")
        ligne_client = l.get("client_code")
        if ligne_client and not client_exists(conn, ligne_client):
            raise ValueError(f"Le client « {ligne_client} » n'existe pas dans la liste des clients.")
        d, c = l.get("debit") or 0, l.get("credit") or 0
        if d and c:
            raise ValueError(f"La ligne « {l['compte']} » ne peut pas être à la fois au débit et au crédit.")
        if not d and not c:
            raise ValueError(f"La ligne « {l['compte']} » doit avoir un montant au débit ou au crédit.")
        racine = account_racine(l["compte"])
        if racine == RACINE_FOURNISSEURS and not (l.get("fournisseur_code") or fournisseur_code):
            raise ValueError(
                f"La ligne « {l['compte']} » relève des Fournisseurs (racine 40) : "
                f"vous devez choisir le fournisseur concerné."
            )
        if racine == RACINE_CLIENTS and not (l.get("client_code") or client_code):
            raise ValueError(
                f"La ligne « {l['compte']} » relève des Clients (racine 41) : "
                f"vous devez choisir le client concerné."
            )
        total_debit += d
        total_credit += c
    if total_debit <= 0:
        raise ValueError("Le montant total de l'écriture doit être strictement positif.")
    if abs(total_debit - total_credit) >= 0.01:
        raise ValueError(
            f"Écriture déséquilibrée : Total Débit = {total_debit:,.2f}, Total Crédit = {total_credit:,.2f} "
            f"(écart de {total_debit - total_credit:,.2f}) — corrigez avant d'enregistrer."
        )
    if compte_stock_global:
        if not account_exists(conn, compte_stock_global):
            raise ValueError(f"Le compte stock « {compte_stock_global} » n'existe pas.")
        if not quantite_stock_global:
            raise ValueError(
                "La quantité est obligatoire quand un compte stock est choisi "
                "(sinon le stock serait mis à jour avec une quantité de 0)."
            )
        if sens_stock_global not in ("entree", "sortie"):
            raise ValueError("Le sens du mouvement de stock doit être « entree » ou « sortie ».")
    _check_exercice_editable(conn, date_str)
    for l in lignes:
        add_entry(conn, date_str, piece, journal, l["compte"], l.get("tiers") or tiers, l.get("libelle") or "",
                  l.get("debit") or 0, l.get("credit") or 0, analytic_code=l.get("analytic_code") or "",
                  budget_code=budget_code, donor_code=donor_code, quantite=l.get("quantite") or 0,
                  fournisseur_code=l.get("fournisseur_code") or fournisseur_code,
                  client_code=l.get("client_code") or client_code)

    if compte_stock_global:
        contre_compte = _compte_variation_stock(compte_stock_global)
        if sens_stock_global == "entree":
            # ---- Coût global d'achat : toutes les lignes débit forment le
            # coût d'un même lot de stock (matière + frais accessoires...) ----
            cout_total = sum(l.get("debit") or 0 for l in lignes)
            if cout_total > 0:
                add_entry(conn, date_str, piece, journal, compte_stock_global, "",
                          "Entrée stock (auto, coût global — matière + frais accessoires)",
                          cout_total, 0, quantite=quantite_stock_global or 0)
                add_entry(conn, date_str, piece, journal, contre_compte, "",
                          "Entrée stock (auto, coût global — matière + frais accessoires)", 0, cout_total)
        else:
            # ---- Sortie globale de vente : coût = quantité × coût unitaire
            # moyen ACTUEL du stock (les lignes débit — ex. comptes clients —
            # ne sont PAS le coût, ce sont des créances). compute_stocks_detail
            # (pas compute_stocks, limité aux 4 comptes centralisateurs) pour
            # couvrir tout compte de stock réel, y compris les sous-comptes
            # granulaires (ex. 321001 CLINKER).
            stocks_by_code = {s["code"]: s for s in compute_stocks_detail(conn, exercice=_exercice_of_date(date_str))}
            stock = stocks_by_code.get(compte_stock_global)
            cout_unitaire = stock["cout_unitaire_moyen"] if stock else None
            if cout_unitaire is None:
                raise ValueError(
                    f"Coût unitaire moyen inconnu pour le compte stock « {compte_stock_global} » "
                    f"(aucun stock ou aucune quantité en stock) — impossible de calculer la sortie."
                )
            montant_sortie = quantite_stock_global * cout_unitaire
            if montant_sortie > 0:
                add_entry(conn, date_str, piece, journal, contre_compte, "",
                          "Sortie stock (auto, coût global — vente groupée)", montant_sortie, 0)
                add_entry(conn, date_str, piece, journal, compte_stock_global, "",
                          "Sortie stock (auto, coût global — vente groupée)", 0, montant_sortie,
                          quantite=quantite_stock_global)
    else:
        # ---- Mouvements de stock automatiques, ligne par ligne (indépendantes) ----
        for l in lignes:
            qte = l.get("quantite") or 0
            if not qte:
                continue
            compte = l["compte"]
            libelle = l.get("libelle") or ""
            if l.get("debit"):
                achat_map = _match_stock_mapping(compte, ACHAT_STOCK_MAPPING)
                if achat_map and compte not in (achat_map[1], achat_map[2]):
                    _, stock_compte, contre_compte = achat_map
                    montant = l["debit"]
                    add_entry(conn, date_str, piece, journal, stock_compte, "", f"Entrée stock (auto) — {libelle}",
                              montant, 0, quantite=qte)
                    add_entry(conn, date_str, piece, journal, contre_compte, "", f"Entrée stock (auto) — {libelle}",
                              0, montant)
            elif l.get("credit"):
                vente_map = _match_stock_mapping(compte, VENTE_STOCK_MAPPING)
                if vente_map and compte not in (vente_map[1], vente_map[2]):
                    _, stock_compte, cout_compte = vente_map
                    stocks_by_code = {s["code"]: s for s in compute_stocks(conn, exercice=_exercice_of_date(date_str))}
                    stock = stocks_by_code.get(stock_compte)
                    cout_unitaire = stock["cout_unitaire_moyen"] if stock else None
                    if cout_unitaire is not None:
                        montant_sortie = qte * cout_unitaire
                        if montant_sortie > 0:
                            add_entry(conn, date_str, piece, journal, cout_compte, "",
                                      f"Sortie stock (auto) — {libelle}", montant_sortie, 0)
                            add_entry(conn, date_str, piece, journal, stock_compte, "",
                                      f"Sortie stock (auto) — {libelle}", 0, montant_sortie, quantite=qte)

    return total_debit


def add_balanced_entry(conn, date_str, piece, journal, compte_debit, compte_credit, montant,
                        tiers, libelle, analytic_code="", budget_code="", donor_code="", quantite=0,
                        fournisseur_code="", client_code=""):
    """Crée en une seule opération une écriture équilibrée par construction :
    une ligne au débit d'un compte, une ligne au crédit d'un autre, même montant.
    C'est le principe de la partie double — impossible de créer un déséquilibre
    en passant par cette fonction.

    Si une quantité est renseignée et que le compte débiteur est un compte
    d'achat lié à un stock (601x marchandises, 602x matières premières), une
    ENTRÉE de stock est automatiquement comptabilisée à sa suite. De même, si
    le compte créditeur est un compte de vente lié à un stock (701x
    marchandises, 702x produits finis), une SORTIE de stock est automatiquement
    comptabilisée (au coût unitaire moyen réel). Cela s'applique à toute
    écriture saisie directement dans l'onglet Saisie — pas seulement à celles
    créées via Facturation / Factures frs."""
    if montant <= 0:
        raise ValueError("Le montant doit être strictement positif.")
    if compte_debit == compte_credit:
        raise ValueError("Le compte débiteur et le compte créditeur doivent être différents.")
    _check_exercice_editable(conn, date_str)
    add_entry(conn, date_str, piece, journal, compte_debit, tiers, libelle, montant, 0,
              analytic_code=analytic_code, budget_code=budget_code, donor_code=donor_code,
              quantite=quantite, fournisseur_code=fournisseur_code, client_code=client_code)
    add_entry(conn, date_str, piece, journal, compte_credit, tiers, libelle, 0, montant,
              analytic_code=analytic_code, budget_code=budget_code, donor_code=donor_code,
              quantite=quantite, fournisseur_code=fournisseur_code, client_code=client_code)

    if quantite:
        achat_map = _match_stock_mapping(compte_debit, ACHAT_STOCK_MAPPING)
        if achat_map and compte_debit not in (achat_map[1], achat_map[2]):
            _, stock_compte, contre_compte = achat_map
            add_entry(conn, date_str, piece, journal, stock_compte, "", f"Entrée stock (auto) — {libelle}",
                      montant, 0, quantite=quantite)
            add_entry(conn, date_str, piece, journal, contre_compte, "", f"Entrée stock (auto) — {libelle}",
                      0, montant)

        vente_map = _match_stock_mapping(compte_credit, VENTE_STOCK_MAPPING)
        if vente_map and compte_credit not in (vente_map[1], vente_map[2]):
            _, stock_compte, cout_compte = vente_map
            stocks_by_code = {s["code"]: s for s in compute_stocks(conn, exercice=_exercice_of_date(date_str))}
            stock = stocks_by_code.get(stock_compte)
            cout_unitaire = stock["cout_unitaire_moyen"] if stock else None
            if cout_unitaire is not None:
                montant_sortie = quantite * cout_unitaire
                if montant_sortie > 0:
                    add_entry(conn, date_str, piece, journal, cout_compte, "", f"Sortie stock (auto) — {libelle}",
                              montant_sortie, 0)
                    add_entry(conn, date_str, piece, journal, stock_compte, "", f"Sortie stock (auto) — {libelle}",
                              0, montant_sortie, quantite=quantite)


def update_entry(conn, entry_id, **fields):
    if not fields:
        return
    row = conn.execute("SELECT * FROM entries WHERE id = ?", (entry_id,)).fetchone()
    if not row:
        raise ValueError(f"Écriture ID {entry_id} introuvable.")
    _check_exercice_editable(conn, row["date"])
    if "date" in fields:
        _check_exercice_editable(conn, fields["date"])
    if "compte" in fields and fields["compte"] and not account_exists(conn, fields["compte"]):
        raise ValueError(
            f"Le compte « {fields['compte']} » n'existe pas dans le plan comptable — "
            f"choisissez un compte dans la liste plutôt que de le saisir librement."
        )
    if "analytic_code" in fields and fields["analytic_code"] and not analytic_code_exists(conn, fields["analytic_code"]):
        raise ValueError(f"Le code analytique « {fields['analytic_code']} » n'existe pas dans le plan analytique.")
    if "fournisseur_code" in fields and fields["fournisseur_code"] and not fournisseur_exists(conn, fields["fournisseur_code"]):
        raise ValueError(f"Le fournisseur « {fields['fournisseur_code']} » n'existe pas.")
    if "client_code" in fields and fields["client_code"] and not client_exists(conn, fields["client_code"]):
        raise ValueError(f"Le client « {fields['client_code']} » n'existe pas.")
    # Racine du compte APRÈS modification : si elle relève des Fournisseurs (40x)
    # ou des Clients (41x), le tiers auxiliaire correspondant doit être renseigné
    # — soit déjà présent sur la ligne, soit fourni dans cette modification.
    compte_final = fields.get("compte", row["compte"])
    racine = account_racine(compte_final)
    if racine == RACINE_FOURNISSEURS:
        fournisseur_final = fields.get("fournisseur_code", row["fournisseur_code"])
        if not fournisseur_final:
            raise ValueError(
                f"Le compte « {compte_final} » relève des Fournisseurs (racine 40) : "
                f"vous devez choisir le fournisseur concerné."
            )
    if racine == RACINE_CLIENTS:
        client_final = fields.get("client_code", row["client_code"])
        if not client_final:
            raise ValueError(
                f"Le compte « {compte_final} » relève des Clients (racine 41) : "
                f"vous devez choisir le client concerné."
            )
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE entries SET {cols} WHERE id = ?", (*fields.values(), entry_id))
    conn.commit()


def delete_entry(conn, entry_id):
    row = conn.execute("SELECT date FROM entries WHERE id = ?", (entry_id,)).fetchone()
    if row:
        _check_exercice_editable(conn, row["date"])
    conn.execute("DELETE FROM entries WHERE id = ?", (entry_id,))
    conn.commit()


def delete_entries_bulk(conn, entry_ids):
    """Supprime plusieurs écritures en une seule transaction (un seul commit
    à la fin, beaucoup plus rapide qu'un appel à delete_entry par ligne —
    chaque commit implique une écriture synchrone sur disque). Retourne
    (nb_supprimées, liste_erreurs)."""
    deleted = 0
    errors = []
    for entry_id in entry_ids:
        row = conn.execute("SELECT date FROM entries WHERE id = ?", (entry_id,)).fetchone()
        if not row:
            continue
        try:
            _check_exercice_editable(conn, row["date"])
        except ValueError as exc:
            errors.append(f"ID {entry_id} : {exc}")
            continue
        conn.execute("DELETE FROM entries WHERE id = ?", (entry_id,))
        deleted += 1
    conn.commit()
    return deleted, errors


def list_entries(conn, order_by="date", exercice=None):
    if exercice:
        date_from, date_to = f"{exercice}-01-01", f"{exercice}-12-31"
        rows = conn.execute(
            f"SELECT * FROM entries WHERE date >= ? AND date <= ? ORDER BY {order_by}, id",
            (date_from, date_to),
        ).fetchall()
    else:
        rows = conn.execute(f"SELECT * FROM entries ORDER BY {order_by}, id").fetchall()
    return [dict(r) for r in rows]


def totals_debit_credit(conn):
    row = conn.execute("SELECT COALESCE(SUM(debit),0) d, COALESCE(SUM(credit),0) c FROM entries").fetchone()
    return row["d"], row["c"]


# ---------------------------------------------------------------------------
# Import massif d'écritures depuis un fichier .xlsx
# ---------------------------------------------------------------------------
IMPORT_COLUMNS = [
    ("date", "Date", ["date", "date piece", "date pièce"]),
    ("piece", "N° Pièce", ["pièce", "piece", "n° pièce", "n° piece", "numero piece", "num piece"]),
    ("journal", "Journal", ["journal"]),
    ("compte", "N° Compte", ["compte", "n° compte", "numero compte", "num compte"]),
    ("tiers", "Tiers", ["tiers"]),
    ("libelle", "Libellé", ["libellé", "libelle"]),
    ("debit", "Débit", ["débit", "debit"]),
    ("credit", "Crédit", ["crédit", "credit"]),
    ("quantite", "Quantité", ["quantité", "quantite", "qté", "qte"]),
    ("analytic_code", "Code analytique", ["code analytique", "analytique"]),
    ("budget_code", "Code budgétaire", ["code budgétaire", "code budgetaire", "budget"]),
    ("donor_code", "Code bailleur", ["code bailleur", "bailleur"]),
]


def export_import_template(path):
    """Génère un modèle .xlsx vierge (bon en-têtes) pour préparer un import massif."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ecritures"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for i, (_, label, _) in enumerate(IMPORT_COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    example = ["15/01/2024", "FA-0001", "AC", "601000", "Fournisseur X", "Achat marchandises",
               100000, 0, 10, "", "", ""]
    for i, val in enumerate(example, start=1):
        ws.cell(row=2, column=i, value=val)
    example2 = ["15/01/2024", "FA-0001", "AC", "401000", "Fournisseur X", "Facture FA-0001",
                0, 100000, 0, "", "", ""]
    for i, val in enumerate(example2, start=1):
        ws.cell(row=3, column=i, value=val)
    for i, w in enumerate([12, 14, 10, 12, 20, 30, 14, 14, 10, 16, 16, 14], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def import_entries_from_xlsx(conn, path):
    """Importe en masse des écritures depuis un .xlsx. Reconnaît les en-têtes en
    français (Date, N° Compte, Débit, Crédit, etc. — voir IMPORT_COLUMNS) quel que
    soit leur ordre. Retourne (nb_importées, liste_avertissements)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in header_cells]

    colmap = {}
    for key, _, aliases in IMPORT_COLUMNS:
        for i, h in enumerate(headers):
            if h in aliases:
                colmap[key] = i
                break

    if "date" not in colmap or "compte" not in colmap:
        raise ValueError(
            "Colonnes obligatoires introuvables dans le fichier (« Date » et « N° Compte »). "
            "Utilisez le bouton « Télécharger un modèle » pour obtenir les bons en-têtes."
        )

    valid_accounts = {r["code"] for r in conn.execute("SELECT code FROM accounts")}
    imported = 0
    warnings = []
    total_debit = total_credit = 0.0

    def get(values, key):
        idx = colmap.get(key)
        if idx is None or idx >= len(values):
            return None
        return values[idx]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [c.value for c in row]
        if all(v in (None, "") for v in values):
            continue

        date_val = get(values, "date")
        if date_val in (None, ""):
            warnings.append(f"Ligne {row_idx} : date manquante, ligne ignorée.")
            continue
        if isinstance(date_val, datetime):
            date_str = date_val.strftime("%Y-%m-%d")
        elif isinstance(date_val, date):
            date_str = date_val.strftime("%Y-%m-%d")
        else:
            date_str = to_iso_date(str(date_val))

        compte = get(values, "compte")
        compte = "" if compte is None else str(compte).strip()
        if compte.endswith(".0") and compte.replace(".0", "").isdigit():
            compte = compte[:-2]
        if not compte:
            warnings.append(f"Ligne {row_idx} : N° Compte manquant, ligne ignorée.")
            continue
        if compte not in valid_accounts:
            warnings.append(f"Ligne {row_idx} : compte '{compte}' absent du plan comptable (importée quand même).")

        def to_float(v, label):
            if v in (None, ""):
                return 0.0
            try:
                return float(v)
            except (TypeError, ValueError):
                warnings.append(f"Ligne {row_idx} : {label} invalide ('{v}'), remplacé par 0.")
                return 0.0

        debit = to_float(get(values, "debit"), "Débit")
        credit = to_float(get(values, "credit"), "Crédit")
        quantite = to_float(get(values, "quantite"), "Quantité")
        piece = get(values, "piece") or ""
        journal = get(values, "journal") or ""
        tiers = get(values, "tiers") or ""
        libelle = get(values, "libelle") or ""
        analytic_code = get(values, "analytic_code") or ""
        budget_code = get(values, "budget_code") or ""
        donor_code = get(values, "donor_code") or ""

        add_entry(conn, date_str, str(piece), str(journal), compte, str(tiers), str(libelle),
                  debit, credit, "", str(analytic_code), str(budget_code), str(donor_code), quantite)
        imported += 1
        total_debit += debit
        total_credit += credit

    ecart = total_debit - total_credit
    if abs(ecart) >= 1:
        warnings.append(
            f"⚠ ATTENTION : ce fichier n'est PAS équilibré dans son ensemble — Total Débit "
            f"{total_debit:,.2f} ≠ Total Crédit {total_credit:,.2f} (écart de {ecart:,.2f}). "
            f"Toutes les lignes ont été importées telles quelles, mais la Balance affichera un "
            f"écart tant que ce déséquilibre n'est pas corrigé (complétez le fichier avec les "
            f"lignes de contrepartie manquantes, ou ajoutez une écriture de régularisation)."
        )

    return imported, warnings


# ---------------------------------------------------------------------------
# Balance
# ---------------------------------------------------------------------------
def compute_balance(conn, only_with_movement=True, include_zero_opening=True, exercice=None):
    """IMPORTANT : parcourt l'UNION des comptes du Plan comptable, des
    comptes ayant un solde d'ouverture et des comptes ayant reçu au moins
    une écriture — PAS seulement le Plan comptable. Un compte présent dans
    les soldes d'ouverture ou dans les écritures mais ABSENT du Plan
    comptable (ex. plan comptable de l'utilisateur plus détaillé que celui
    bundlé avec l'application) doit être inclus quand même, sous peine de
    faire disparaître silencieusement son solde de TOUS les calculs
    (Bilan, Balance, Grand livre...) et de créer un faux écart Actif/Passif."""
    exercice = exercice or get_current_exercice(conn)
    date_from, date_to = f"{exercice}-01-01", f"{exercice}-12-31"

    mouvements = {}
    for r in conn.execute("""
        SELECT compte,
               COALESCE(SUM(CASE WHEN date BETWEEN ? AND ? THEN debit ELSE 0 END), 0)  AS debit,
               COALESCE(SUM(CASE WHEN date BETWEEN ? AND ? THEN credit ELSE 0 END), 0) AS credit
        FROM entries GROUP BY compte
    """, (date_from, date_to, date_from, date_to)):
        mouvements[r["compte"]] = (r["debit"], r["credit"])

    openings = {r["code"]: r["solde"] for r in conn.execute(
        "SELECT code, solde FROM opening_balances WHERE exercice = ?", (exercice,))}

    accounts_info = {r["code"]: (r["label"], r["classe"]) for r in conn.execute(
        "SELECT code, label, classe FROM accounts")}

    tous_les_codes = set(accounts_info) | set(openings) | set(mouvements)

    result = []
    for code in sorted(tous_les_codes):
        if code in accounts_info:
            label, classe = accounts_info[code]
        else:
            # Compte absent du Plan comptable mais utilisé (solde d'ouverture
            # ou écriture) — inclus quand même, avec un libellé de repli et
            # une classe déduite du premier chiffre du code.
            label, classe = f"{code} (hors Plan comptable)", (code[:1] if code else "")
        debit, credit = mouvements.get(code, (0, 0))
        ouverture = openings.get(code, 0.0)
        if only_with_movement and debit == 0 and credit == 0 and ouverture == 0:
            continue
        solde_mouvement = debit - credit
        result.append({
            "code": code, "label": label, "classe": classe,
            "debit": debit, "credit": credit, "solde": solde_mouvement,
            "solde_ouverture": ouverture,
            "solde_cloture": ouverture + solde_mouvement,
        })
    return result


def compute_balance_detaillee(conn, exercice=None):
    """Balance générale groupée par classe, avec un sous-total par classe et
    un total général — 6 colonnes : Solde Ouverture Débit/Crédit, Mouvement
    (cumul) Débit/Crédit de la période, Solde Clôture Débit/Crédit. Calculée
    à partir de la même compute_balance() que le Bilan, donc garantie
    cohérente avec lui."""
    balance = sorted(compute_balance(conn, only_with_movement=True, exercice=exercice),
                      key=lambda b: b["code"])
    classes = {}
    for b in balance:
        classes.setdefault(b["classe"], []).append(b)

    cols = ("ouverture_debit", "ouverture_credit", "cumul_debit", "cumul_credit", "solde_debit", "solde_credit")
    result_classes = []
    grand = {k: 0.0 for k in cols}
    for classe in sorted(classes.keys()):
        lignes = []
        sous_total = {k: 0.0 for k in cols}
        for b in classes[classe]:
            solde_ouverture = b["solde_ouverture"]
            solde_cloture = b["solde_cloture"]
            ligne = {
                "code": b["code"], "label": b["label"],
                "ouverture_debit": solde_ouverture if solde_ouverture > 0 else 0.0,
                "ouverture_credit": -solde_ouverture if solde_ouverture < 0 else 0.0,
                "cumul_debit": b["debit"], "cumul_credit": b["credit"],
                "solde_debit": solde_cloture if solde_cloture > 0 else 0.0,
                "solde_credit": -solde_cloture if solde_cloture < 0 else 0.0,
                # Conservé pour compatibilité (ancien code affichant un solde d'ouverture signé unique)
                "solde_ouverture": solde_ouverture,
            }
            lignes.append(ligne)
            for k in cols:
                sous_total[k] += ligne[k]
        result_classes.append({"classe": classe, "lignes": lignes, "sous_total": sous_total})
        for k in cols:
            grand[k] += sous_total[k]

    return {"classes": result_classes, "grand_total": grand}


def export_balance_xlsx(conn, path, exercice=None):
    """Exporte la Balance générale (avec sous-totaux par classe et total
    général) en .xlsx — mêmes données que l'onglet Balance."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    exercice = exercice or get_current_exercice(conn)
    data = compute_balance_detaillee(conn, exercice=exercice)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    total_font = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="DCE6F1")
    grand_font = Font(bold=True, color="FFFFFFFF")
    grand_fill = PatternFill("solid", fgColor="1F4E78")

    headers = ["N° Compte", "Libellé du compte", "Ouverture Débit", "Ouverture Crédit",
               "Mouvement Débit", "Mouvement Crédit", "Clôture Débit", "Clôture Crédit"]
    ws.append([f"BALANCE GÉNÉRALE — Exercice {exercice}"])
    ws["A1"].font = Font(bold=True, size=13)
    ws.append([])
    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=c)
        cell.font = header_font
        cell.fill = header_fill

    for c in data["classes"]:
        for l in c["lignes"]:
            ws.append([l["code"], l["label"], l["ouverture_debit"] or None, l["ouverture_credit"] or None,
                       l["cumul_debit"] or None, l["cumul_credit"] or None,
                       l["solde_debit"] or None, l["solde_credit"] or None])
        st = c["sous_total"]
        ws.append(["", f"TOTAL CLASSE {c['classe']}", st["ouverture_debit"], st["ouverture_credit"],
                   st["cumul_debit"], st["cumul_credit"], st["solde_debit"], st["solde_credit"]])
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = total_font
            cell.fill = total_fill

    gt = data["grand_total"]
    ws.append(["", "TOTAL BALANCE", gt["ouverture_debit"], gt["ouverture_credit"],
               gt["cumul_debit"], gt["cumul_credit"], gt["solde_debit"], gt["solde_credit"]])
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=ws.max_row, column=col)
        cell.font = grand_font
        cell.fill = grand_fill

    for col in range(3, len(headers) + 1):
        for row in range(4, ws.max_row + 1):
            ws.cell(row=row, column=col).number_format = "#,##0.00"

    for i, w in enumerate([14, 36, 15, 15, 15, 15, 15, 15], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A4"
    wb.save(path)
    return path


def compute_tresorerie_detail(conn, exercice=None, balance=None):
    """Détail de la trésorerie (classe 5) par compte réel — ex. chaque banque
    séparément (521110 WENDKUNI BANK, 521120 CORIS BANK...) — calculé à
    partir de la même compute_balance() que le Bilan et la Balance.
    `balance` : voir compute_bilan() — permet de réutiliser cette fonction
    avec les mouvements de la période seule ou le solde d'ouverture seul."""
    if balance is None:
        balance = compute_balance(conn, only_with_movement=True, exercice=exercice)
    lignes = [b for b in balance if b["classe"] == "5" and (b["solde_cloture"] or b.get("debit") or b.get("credit"))]
    lignes.sort(key=lambda b: b["code"])
    total = sum(b["solde_cloture"] for b in lignes)
    return lignes, total


# ---------------------------------------------------------------------------
# Comptes d'une classe/racine sur une PÉRIODE LIBRE (pas forcément l'exercice
# entier) — utilisé par les onglets Impôts (classe 44), Déclarations sociales
# (classe 43) et Rapprochements bancaires (racine 52). Pour chaque compte :
# solde en début de période (solde d'ouverture de l'exercice + mouvements
# antérieurs à la période), mouvements Débit/Crédit DANS la période choisie,
# solde en fin de période.
# ---------------------------------------------------------------------------
def compute_comptes_prefixe_periode(conn, prefix, date_from=None, date_to=None, exercice=None):
    """Tous les comptes dont le code commence par `prefix` (ex. '44', '43',
    '52'), avec solde de début de période / mouvements de la période choisie
    / solde de fin de période. `date_from`/`date_to` (AAAA-MM-JJ) — par
    défaut, l'exercice entier. Un compte est inclus dès qu'il a un solde de
    début de période non nul ou un mouvement dans la période."""
    exercice = exercice or get_current_exercice(conn)
    exercice_debut = f"{exercice}-01-01"
    exercice_fin = f"{exercice}-12-31"
    date_from = date_from or exercice_debut
    date_to = date_to or exercice_fin

    comptes = conn.execute(
        "SELECT code, label, classe FROM accounts WHERE code LIKE ? ORDER BY code",
        (f"{prefix}%",),
    ).fetchall()

    result = []
    for c in comptes:
        code = c["code"]
        ouverture = get_opening_balance(conn, code, exercice)
        avant = conn.execute(
            """SELECT COALESCE(SUM(debit), 0) d, COALESCE(SUM(credit), 0) c FROM entries
               WHERE compte = ? AND date >= ? AND date < ?""",
            (code, exercice_debut, date_from),
        ).fetchone()
        solde_debut_periode = ouverture + (avant["d"] - avant["c"])

        periode = conn.execute(
            """SELECT COALESCE(SUM(debit), 0) d, COALESCE(SUM(credit), 0) c FROM entries
               WHERE compte = ? AND date >= ? AND date <= ?""",
            (code, date_from, date_to),
        ).fetchone()
        debit_periode, credit_periode = periode["d"], periode["c"]
        solde_fin_periode = solde_debut_periode + (debit_periode - credit_periode)

        if not (solde_debut_periode or debit_periode or credit_periode or solde_fin_periode):
            continue
        result.append({
            "code": code, "label": c["label"], "classe": c["classe"],
            "solde_debut_periode": solde_debut_periode,
            "debit_periode": debit_periode, "credit_periode": credit_periode,
            "solde_fin_periode": solde_fin_periode,
        })
    return result


def compute_mouvements_prefixe_periode(conn, prefix, date_from=None, date_to=None, exercice=None):
    """Détail écriture par écriture (pas seulement les soldes) de tous les
    comptes dont le code commence par `prefix`, sur la période choisie —
    utilisé par le Rapprochement bancaire pour lister chaque mouvement avec
    sa case de pointage."""
    exercice = exercice or get_current_exercice(conn)
    date_from = date_from or f"{exercice}-01-01"
    date_to = date_to or f"{exercice}-12-31"

    comptes = conn.execute(
        "SELECT code, label FROM accounts WHERE code LIKE ? AND LENGTH(code) = 6 ORDER BY code",
        (f"{prefix}%",),
    ).fetchall()

    result = []
    for c in comptes:
        code = c["code"]
        ouverture = get_opening_balance(conn, code, exercice)
        avant = conn.execute(
            """SELECT COALESCE(SUM(debit), 0) d, COALESCE(SUM(credit), 0) c FROM entries
               WHERE compte = ? AND date >= ? AND date < ?""",
            (code, f"{exercice}-01-01", date_from),
        ).fetchone()
        solde = ouverture + (avant["d"] - avant["c"])

        rows = conn.execute(
            """SELECT e.*, COALESCE(p.pointe, 0) AS pointe
               FROM entries e LEFT JOIN pointages_bancaires p ON p.entry_id = e.id
               WHERE e.compte = ? AND e.date >= ? AND e.date <= ?
               ORDER BY e.date, e.id""",
            (code, date_from, date_to),
        ).fetchall()
        if not rows and not solde:
            continue
        mouvements = []
        for r in rows:
            d = dict(r)
            solde += d["debit"] - d["credit"]
            d["solde_cumule"] = solde
            d["pointe"] = bool(d["pointe"])
            mouvements.append(d)
        result.append({
            "code": code, "label": c["label"], "solde_debut_periode": ouverture + (avant["d"] - avant["c"]),
            "mouvements": mouvements, "solde_fin_periode": solde,
            "total_pointe": sum(d["debit"] - d["credit"] for d in mouvements if d["pointe"]),
        })
    return result


def set_pointage_bancaire(conn, entry_id, pointe):
    """Coche/décoche un mouvement bancaire comme retrouvé dans le relevé
    (rapprochement bancaire)."""
    if pointe:
        conn.execute(
            """INSERT INTO pointages_bancaires (entry_id, pointe, date_pointage)
               VALUES (?, 1, date('now'))
               ON CONFLICT(entry_id) DO UPDATE SET pointe = 1, date_pointage = date('now')""",
            (entry_id,),
        )
    else:
        conn.execute("DELETE FROM pointages_bancaires WHERE entry_id = ?", (entry_id,))
    conn.commit()


def _sum_accounts(balance, codes):
    """Somme Débit/Crédit pour tous les comptes dont le code COMMENCE PAR l'un
    des préfixes donnés (rétro-compatible : un préfixe de 6 chiffres ne
    matche que le compte exact ; un préfixe de 3 chiffres couvre aussi tous
    les sous-comptes détaillés, ex. « 602 » couvre 602000, 602101, 602102...)."""
    debit = credit = 0.0
    for b in balance:
        if any(b["code"].startswith(c) for c in codes):
            debit += b["debit"]
            credit += b["credit"]
    return debit, credit


def _sum_accounts_cloture(balance, codes):
    """Somme des soldes de CLÔTURE (ouverture + mouvements) pour tous les
    comptes dont le code commence par l'un des préfixes donnés."""
    return sum(b["solde_cloture"] for b in balance if any(b["code"].startswith(c) for c in codes))


def _sum_class(balance, classe, sign=None, field="solde_cloture"):
    total = 0
    for b in balance:
        if b["classe"] != classe:
            continue
        v = b[field]
        if sign == "pos" and v <= 0:
            continue
        if sign == "neg" and v >= 0:
            continue
        total += v
    return total


def _sum_racine(balance, racine, sign=None, field="solde_cloture"):
    """Somme les soldes des comptes dont la racine (cf. account_racine) correspond,
    avec un filtre de signe optionnel (utilisé pour les racines « fourre-tout »
    42 à 49, dont la nature actif/passif dépend du solde effectif)."""
    total = 0
    for b in balance:
        if account_racine(b["code"]) != racine:
            continue
        v = b[field]
        if sign == "pos" and v <= 0:
            continue
        if sign == "neg" and v >= 0:
            continue
        total += v
    return total


def compute_resultat_net_complet(conn, exercice=None):
    """Résultat net calculé de façon EXHAUSTIVE, directement à partir de
    TOUTES les classes 6 (charges), 7 (produits) et 8 (autres charges et
    produits HAO) de la Balance — par opposition aux calculs détaillés
    (compute_compte_resultat, compute_liasse_resultat) qui répartissent ce
    résultat ligne par ligne via des listes de comptes pré-définies, lesquelles
    peuvent ne pas couvrir tous les sous-comptes d'un plan comptable réel
    (le plan de l'utilisateur compte 1591 comptes importés de Sage).

    Cette fonction sert de RÉFÉRENCE UNIQUE du résultat net pour tout le
    reste de l'application (Bilan, clôture d'exercice) et garantit que le
    Bilan reste équilibré quel que soit le détail réel du plan comptable :
    aucun compte de classe 6/7/8 n'est jamais oublié. compute_liasse_resultat()
    est recalée sur cette référence (ligne « Autres produits/charges non
    classés ») afin que le Compte de résultat, le TFT, la Situation financière
    et la Liasse fiscale restent eux aussi cohérents avec le Bilan.
    """
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    # Pour un compte de charge (classe 6, nature débitrice) ou de produit
    # (classe 7, nature créditrice) ou HAO (classe 8, mixte), sa contribution
    # au résultat est (crédit - débit) = -solde_cloture (solde_cloture = débit - crédit).
    return -sum(b["solde_cloture"] for b in balance if b["classe"] in ("6", "7", "8"))


# ---------------------------------------------------------------------------
# Compte de résultat
# ---------------------------------------------------------------------------
def compute_compte_resultat(conn, exercice=None):
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)

    def net_produit(codes):
        d, c = _sum_accounts(balance, codes)
        return c - d

    def net_charge(codes):
        d, c = _sum_accounts(balance, codes)
        return d - c

    produits = {
        "Ventes (marchandises, produits finis, travaux, services)": net_produit(COMPTES_PRODUITS_EXPL),
        "Subventions d'exploitation": net_produit([COMPTE_SUBV_EXPL]),
        "Autres produits": net_produit([COMPTE_AUTRES_PRODUITS]),
    }
    total_produits = sum(produits.values())

    charges = {
        "Achats (marchandises et matières)": net_charge(COMPTES_ACHATS),
        "Transports": net_charge(COMPTES_TRANSPORT),
        "Services extérieurs": net_charge(COMPTES_SERVICES_EXT),
        "Impôts et taxes": net_charge(COMPTES_IMPOTS),
        "Autres charges": net_charge([COMPTE_AUTRES_CHARGES]),
        "Charges de personnel": net_charge(COMPTES_PERSONNEL),
        "Dotations aux amortissements et provisions": net_charge(COMPTES_DOTATIONS),
    }
    total_charges = sum(charges.values())

    resultat_exploitation = total_produits - total_charges

    produits_fin = net_produit(COMPTES_PRODUITS_FIN)
    charges_fin = net_charge(COMPTES_CHARGES_FIN)
    resultat_financier = produits_fin - charges_fin

    resultat_net = resultat_exploitation + resultat_financier

    return {
        "produits": produits, "total_produits": total_produits,
        "charges": charges, "total_charges": total_charges,
        "resultat_exploitation": resultat_exploitation,
        "produits_financiers": produits_fin, "charges_financieres": charges_fin,
        "resultat_financier": resultat_financier,
        "resultat_net": resultat_net,
    }


# ---------------------------------------------------------------------------
# Bilan
# ---------------------------------------------------------------------------
def compute_bilan(conn, stock_initial=0.0, exercice=None, balance=None, resultat_net_override=None):
    """stock_initial : conservé pour compatibilité, normalement inutile désormais —
    utilisez la table des soldes d'ouverture (onglet « Soldes d'ouverture »).

    `balance`/`resultat_net_override` : permettent de RÉUTILISER cette même
    logique de classification pour d'autres grandeurs que le solde de
    clôture habituel — ex. les MOUVEMENTS de la période seule, ou les
    SOLDES D'OUVERTURE seuls (voir compute_bilan_mouvement_periode() et
    compute_bilan_solde_ouverture()) — en passant une balance où
    `solde_cloture` représente la grandeur voulue, et le résultat net
    déjà calculé dans cette même logique. Sinon (cas normal), calculés ici
    à partir de la vraie Balance de l'exercice.

    IMPORTANT — équilibre garanti : ce calcul classe CHAQUE compte de la
    Balance (classes 1 à 5, plus le résultat net qui absorbe les classes
    6/7/8) dans une case et une seule de l'Actif ou du Passif, sans jamais
    s'appuyer sur des listes de comptes partielles codées en dur pour les
    totaux (seul le détail affiché à titre indicatif peut être partiel).
    Le Bilan est donc TOUJOURS équilibré (Actif = Passif), à la seule
    condition que la somme des soldes d'ouverture de l'exercice soit nulle
    (partie double) — sinon l'écart affiché pointe vers l'onglet
    « Soldes d'ouverture »."""
    if balance is None:
        balance = compute_balance(conn, only_with_movement=False, exercice=exercice)

    immo_brutes = sum(b["solde_cloture"] for b in balance if b["classe"] == "2" and int(b["code"]) < 280000)
    amortissements = sum(b["solde_cloture"] for b in balance if b["classe"] == "2" and int(b["code"]) >= 280000)
    immo_nettes = immo_brutes + amortissements

    # Stocks : l'intégralité de la classe 3 (pas seulement les 4 comptes
    # maîtres 310000/320000/331000/360000 suivis dans l'onglet Stocks) — tout
    # sous-compte de stock détaillé (33, 37, 38, 39...) doit être inclus au
    # Bilan, sous peine de faire disparaître un solde réel du Total Actif.
    stocks = stock_initial + _sum_class(balance, "3")

    # Comptes de tiers (classe 4), classés PAR COMPTE selon le signe de son
    # solde de clôture — comme l'indiquent les libellés du rapport financier
    # de référence (chaque racine 40 à 49 apparaît potentiellement des DEUX
    # côtés du Bilan : ex. « Frs avances versées *40* » à l'Actif ET
    # « Fournisseurs *40 » au Passif) : un compte débiteur (solde > 0) va en
    # créances, un compte créditeur (solde < 0) va en dettes — quelle que
    # soit sa racine, y compris 40 (Fournisseurs) et 41 (Clients).
    racines_tiers = [str(r) for r in range(40, 50)]
    creances = sum(_sum_racine(balance, r, sign="pos") for r in racines_tiers)
    dettes_circulantes = sum(-_sum_racine(balance, r, sign="neg") for r in racines_tiers)

    treso_actif = _sum_class(balance, "5", sign="pos")
    total_actif = immo_nettes + stocks + creances + treso_actif

    # Ressources durables (classe 1 dans son intégralité : capital, réserves,
    # report à nouveau, subventions, provisions réglementées, emprunts et
    # dettes financières 16-17, comptes de liaison 18, provisions financières
    # pour risques 19...) — plus le résultat net de l'exercice, calculé de
    # façon exhaustive (voir compute_resultat_net_complet), qui n'est pas
    # encore posté sur un compte tant que l'exercice n'est pas clôturé.
    ressources_durables = -_sum_class(balance, "1")
    resultat_net = resultat_net_override if resultat_net_override is not None else compute_resultat_net_complet(
        conn, exercice=exercice)

    # Détail indicatif (catégories usuelles) pour l'affichage — une ligne
    # « Autres postes de ressources durables » absorbe tout compte de classe 1
    # non couvert par ces catégories usuelles, afin que le détail affiché
    # somme TOUJOURS exactement au vrai total (ressources_durables).
    capital = _sum_accounts_cloture(balance, COMPTES_CAPITAL) * -1
    subventions = _sum_accounts_cloture(balance, [COMPTE_SUBVENTIONS]) * -1
    provisions = _sum_accounts_cloture(balance, [COMPTE_PROVISIONS]) * -1
    dettes_financieres = _sum_accounts_cloture(balance, COMPTES_DETTES_FIN) * -1
    autres_ressources = ressources_durables - (capital + subventions + provisions + dettes_financieres)

    treso_passif = -_sum_class(balance, "5", sign="neg")
    total_passif = ressources_durables + resultat_net + dettes_circulantes + treso_passif

    return {
        "actif": {
            "Immobilisations brutes": immo_brutes,
            "Amortissements (à déduire)": amortissements,
            "Immobilisations nettes": immo_nettes,
            "Stocks": stocks,
            "Créances et emplois assimilés": creances,
            "Trésorerie actif": treso_actif,
        },
        "total_actif": total_actif,
        "passif": {
            "Capital et réserves": capital,
            "Subventions d'investissement": subventions,
            "Provisions pour risques et charges": provisions,
            "Autres postes de ressources durables": autres_ressources,
            "Résultat net de l'exercice": resultat_net,
            "Dettes financières": dettes_financieres,
            "Dettes circulantes": dettes_circulantes,
            "Trésorerie passif": treso_passif,
        },
        "total_passif": total_passif,
        "ecart": total_actif - total_passif,
    }


# ---------------------------------------------------------------------------
# Bilan détaillé (« avec détails ») — présentation ligne par ligne, calquée
# sur le modèle de rapport financier fourni par l'utilisateur : Actif en
# Brut / Amortissements / Net, Passif détaillé par nature de compte (chaque
# compte de tiers, chaque banque affichés séparément), plutôt que les seuls
# grands agrégats de compute_bilan(). Construit de façon à ce que la somme
# des lignes affichées corresponde TOUJOURS exactement aux totaux (mêmes
# totaux que compute_bilan(), puisque chaque compte de la Balance est classé
# dans une case et une seule).
# ---------------------------------------------------------------------------
def _detail_racine(balance, racine, sign=None):
    """Comptes d'une racine (ex. '44'), triés par code, avec filtre de signe
    optionnel sur le solde de clôture."""
    lignes = [b for b in balance if account_racine(b["code"]) == racine
              and b["solde_cloture"]
              and (sign is None
                   or (sign == "pos" and b["solde_cloture"] > 0)
                   or (sign == "neg" and b["solde_cloture"] < 0))]
    lignes.sort(key=lambda b: b["code"])
    return lignes


def _detail_prefix2(balance, classe, prefix2, sign=None):
    """Comptes d'une classe donnée dont le code commence par un préfixe à 2
    chiffres (ex. classe '1', préfixe '10' -> comptes 10xxxx), triés, avec
    filtre de signe optionnel."""
    lignes = [b for b in balance if b["classe"] == classe and b["code"][:2] == prefix2
              and b["solde_cloture"]
              and (sign is None
                   or (sign == "pos" and b["solde_cloture"] > 0)
                   or (sign == "neg" and b["solde_cloture"] < 0))]
    lignes.sort(key=lambda b: b["code"])
    return lignes


IMMO_CATEGORIES = [
    # (libellé, plages de comptes BRUT, plages de comptes AMORTISSEMENT/PROVISION)
    # — tirées directement des formules du rapport de référence de l'utilisateur
    # (CtaCptSolde sur les mêmes plages de comptes 28*/29*) : plus de répartition
    # proportionnelle indicative, l'amortissement de chaque catégorie est EXACT.
    ("Charges immobilisées (frais d'établissement, charges à répartir)",
     [(201000, 209999)], [(280000, 280999), (290000, 290999)]),
    ("Brevets, licences, logiciels et immobilisations incorporelles",
     [(210000, 219999)], [(281000, 281999), (291000, 291999)]),
    ("Terrains", [(220000, 229999)], [(282000, 282999), (292000, 292999)]),
    ("Bâtiments", [(230000, 233999)], [(283100, 283399), (293100, 293399)]),
    ("Installations, agencements et aménagements",
     [(234000, 239999)], [(283400, 283999), (293400, 293999)]),
    ("Matériel", [(240000, 244999)], [(284000, 284499), (294000, 294499)]),
    ("Matériel de transport", [(245000, 245999)], [(284500, 284999), (294500, 294999)]),
    ("Avances et acomptes versés sur immobilisations",
     [(250000, 259999)], [(285000, 285999), (295000, 295999)]),
    ("Immobilisations financières (titres, prêts, dépôts...)",
     [(260000, 279999)], [(286000, 286999), (296000, 296999)]),
]


RACINE_LABELS_CREANCES = {
    "40": "Fournisseurs — avances et acomptes versés", "41": "Clients débiteurs",
    "42": "Personnel — débiteurs", "43": "Organismes sociaux (CNSS...) — débiteurs",
    "44_45": "IUTS-TPA-TVA — débiteurs (État, racines 44-45)",
    "47_49": "HAO — débiteurs divers (racines 47-49)", "46": "Débiteurs divers",
}
RACINE_LABELS_DETTES = {
    "40": "Fournisseurs", "41": "Clients créditeurs (avoirs)",
    "42": "Personnel — créditeurs", "43": "Organismes sociaux (CNSS...) — créditeurs",
    "44_45": "IUTS-TPA-TVA — créditeur (État, racines 44-45)",
    "47_49": "HAO — créditeurs divers (racines 47-49)", "46": "Créditeurs divers",
}
TRESO_LABELS = {
    "50_56": "Banques créditrices/débitrices (racines 50-56)",
    "57_59": "Caisse créditrice/débitrice (racines 57-59)",
}


def _cle_racine_bilan(racine):
    """Regroupement exact du rapport financier de référence de
    l'utilisateur : racines 44 et 45 combinées en une seule ligne
    (« IUTS-TPA-TVA »), racines 47 à 49 combinées en une seule ligne
    (« HAO »)."""
    if racine in ("44", "45"):
        return "44_45"
    if racine in ("47", "48", "49"):
        return "47_49"
    return racine


def _cle_prefixe_treso(code_compte):
    """Regroupement Trésorerie exact du rapport de référence : racines 50 à
    56 combinées (banques, CtaCptSolde("50*","56*") — une PLAGE NUMÉRIQUE
    continue 500000-569999, pas une liste de racines choisies une par une,
    pour ne jamais oublier une racine intermédiaire comme 55), racines 57 à
    59 combinées (caisse, plage 570000-599999)."""
    code_int = int(code_compte)
    if 500000 <= code_int <= 569999:
        return "50_56"
    return "57_59"


def _grouper_avec_sous_total(lignes_plates, cle, labels):
    """Regroupe une liste de lignes compte-par-compte {label, montant, <cle>}
    par la valeur de `cle` (racine ou préfixe), avec un sous-total par
    groupe — pour que la comparaison N / N-1 ait un sens (le détail
    compte-par-compte ne s'aligne pas forcément d'un exercice à l'autre :
    un compte peut apparaître une année et pas l'autre)."""
    groupes = {}
    ordre = []
    for l in lignes_plates:
        cle_val = l[cle]
        if cle_val not in groupes:
            groupes[cle_val] = {"key": cle_val, "label": f"{cle_val} — {labels.get(cle_val, 'Autres')}",
                                 "comptes": [], "sous_total": 0.0}
            ordre.append(cle_val)
        groupes[cle_val]["comptes"].append(l)
        groupes[cle_val]["sous_total"] += l["montant"]
    return [groupes[k] for k in ordre if abs(groupes[k]["sous_total"]) >= 1 or groupes[k]["comptes"]]


def _balance_mouvement_periode(conn, exercice):
    """Renvoie compute_balance() avec `solde_cloture` REMPLACÉ par le seul
    mouvement de la période (Débit - Crédit, hors solde d'ouverture) — pour
    calculer un Bilan basé exclusivement sur les opérations de la période
    (voir compute_bilan_mouvement_periode())."""
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    result = []
    for b in balance:
        b2 = dict(b)
        b2["solde_cloture"] = b["debit"] - b["credit"]
        result.append(b2)
    return result


def _balance_solde_ouverture(conn, exercice):
    """Renvoie compute_balance() avec `solde_cloture` REMPLACÉ par le seul
    solde d'ouverture (report à nouveau du début de l'exercice) — pour la
    colonne N-1 du Bilan basé sur les opérations de la période (voir
    compute_bilan_solde_ouverture())."""
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    result = []
    for b in balance:
        b2 = dict(b)
        b2["solde_cloture"] = b["solde_ouverture"]
        result.append(b2)
    return result


def _resultat_net_mouvement_periode(balance_mouvement):
    """Résultat net calculé UNIQUEMENT à partir des mouvements de la
    période (classes 6/7/8, déjà substitués dans `solde_cloture` par
    _balance_mouvement_periode) — même principe exhaustif que
    compute_resultat_net_complet(), mais sans solde d'ouverture."""
    return -_sum_class(balance_mouvement, "7") - _sum_class(balance_mouvement, "6") - _sum_class(balance_mouvement, "8")


def compute_bilan_mouvement_periode(conn, exercice=None):
    """Bilan basé EXCLUSIVEMENT sur les opérations (mouvements Débit/Crédit)
    de la période — PAS le solde d'ouverture. Réutilise exactement la même
    logique de classification que compute_bilan() (garantie d'équilibre
    Actif = Passif identique, car la somme des mouvements Débit - Crédit
    de TOUTES les classes sur la période est toujours nulle, par la partie
    double)."""
    exercice = exercice or get_current_exercice(conn)
    balance = _balance_mouvement_periode(conn, exercice)
    resultat_net = _resultat_net_mouvement_periode(balance)
    return compute_bilan(conn, exercice=exercice, balance=balance, resultat_net_override=resultat_net)


def compute_bilan_solde_ouverture(conn, exercice=None):
    """Bilan basé EXCLUSIVEMENT sur les soldes d'ouverture de l'exercice
    (report à nouveau du 1er janvier) — utilisé comme colonne « N-1 » du
    Bilan basé sur les opérations de la période (voir
    compute_bilan_mouvement_periode()). Même garantie d'équilibre."""
    exercice = exercice or get_current_exercice(conn)
    balance = _balance_solde_ouverture(conn, exercice)
    resultat_net = _resultat_net_mouvement_periode(balance)  # classes 6/7/8 = 0 en ouverture si l'exercice est bien clos
    return compute_bilan(conn, exercice=exercice, balance=balance, resultat_net_override=resultat_net)


def _compute_bilan_groupes(conn, exercice, balance=None, resultat_net_override=None):
    """Calcule les groupes du Bilan détaillé pour UN exercice donné —
    utilisé pour l'exercice courant (avec le détail compte par compte) ET
    pour l'exercice N-1 (uniquement les sous-totaux, pour la comparaison).
    `balance`/`resultat_net_override` : voir compute_bilan() — permettent
    de réutiliser cette même logique pour les mouvements de la période
    seule ou le solde d'ouverture seul (voir compute_bilan_detaille()).
    Retourne un dict de listes de groupes {key, label, sous_total,
    comptes:[...]}, plus les totaux brut/amortissement/net des
    immobilisations et les totaux généraux."""
    if balance is None:
        balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    bilan = compute_bilan(conn, exercice=exercice, balance=balance, resultat_net_override=resultat_net_override)

    # ---- Immobilisations (Brut / Amortissements / Net), par catégorie exacte ----
    total_brut = sum(b["solde_cloture"] for b in balance if b["classe"] == "2" and int(b["code"]) < 280000)
    total_amort = sum(b["solde_cloture"] for b in balance if b["classe"] == "2" and int(b["code"]) >= 280000)
    immobilisations = []
    somme_brut_categories = somme_amort_categories = 0.0
    comptes_brut_classes, comptes_amort_classes = set(), set()
    for label, brut_ranges, amort_ranges in IMMO_CATEGORIES:
        brut = _sum_range(balance, brut_ranges, classe="2")
        amort = _sum_range(balance, amort_ranges, classe="2")
        somme_brut_categories += brut
        somme_amort_categories += amort
        if brut or amort:
            immobilisations.append({"key": label, "label": label, "brut": brut, "amort": amort, "net": brut + amort})
        for b in balance:
            if b["classe"] != "2":
                continue
            code_int = int(b["code"])
            if any(lo <= code_int <= hi for lo, hi in brut_ranges):
                comptes_brut_classes.add(b["code"])
            if any(lo <= code_int <= hi for lo, hi in amort_ranges):
                comptes_amort_classes.add(b["code"])
    autres_brut = total_brut - somme_brut_categories
    autres_amort = total_amort - somme_amort_categories
    if abs(autres_brut) >= 1 or abs(autres_amort) >= 1:
        # Détail des comptes hors de TOUTES les plages IMMO_CATEGORIES — pour
        # identifier précisément quels comptes réels ne correspondent à
        # aucune catégorie attendue (ex. un plan de comptes différent de
        # celui du rapport de référence), au lieu d'un simple montant global.
        comptes_non_classes = []
        for b in balance:
            if b["classe"] != "2" or not b["solde_cloture"]:
                continue
            code_int = int(b["code"])
            est_brut = code_int < 280000
            if est_brut and b["code"] not in comptes_brut_classes:
                comptes_non_classes.append({"label": f"{b['code']} {b['label']}", "montant": b["solde_cloture"]})
            elif not est_brut and b["code"] not in comptes_amort_classes:
                comptes_non_classes.append({"label": f"{b['code']} {b['label']} (amort./provision)",
                                             "montant": b["solde_cloture"]})
        immobilisations.append({"key": "Autres immobilisations non classées",
                                 "label": "Autres immobilisations non classées",
                                 "brut": autres_brut, "amort": autres_amort, "net": autres_brut + autres_amort,
                                 "comptes": comptes_non_classes})
    immobilisations.sort(key=lambda l: -abs(l["brut"]))

    # ---- Stocks (classe 3), par préfixe à 2 chiffres ----
    stocks_labels = {
        "31": "Marchandises", "32": "Matières premières et fournitures liées",
        "33": "Autres approvisionnements (pièces de rechange...)",
        "34": "Produits en cours", "35": "Services en cours",
        "36": "Produits finis", "37": "Produits intermédiaires et résiduels",
        "38": "Stocks en cours de route / en consignation", "39": "Dépréciations des stocks",
    }
    stocks_par_prefixe = {}
    for b in balance:
        if b["classe"] != "3" or not b["solde_cloture"]:
            continue
        prefixe = b["code"][:2]
        stocks_par_prefixe[prefixe] = stocks_par_prefixe.get(prefixe, 0.0) + b["solde_cloture"]
    stocks = [{"key": p, "label": f"Stocks {p} — {stocks_labels.get(p, 'Autres stocks')}",
               "sous_total": v, "comptes": []} for p, v in sorted(stocks_par_prefixe.items()) if v]

    # ---- Créances / Dettes (racines 40-49, compte par compte, groupées selon
    # le rapport de référence : 44+45 combinées, 47-49 combinées) ----
    creances_flat = []
    for racine in [str(r) for r in range(40, 50)]:
        for b in _detail_racine(balance, racine, sign="pos"):
            creances_flat.append({"label": f"{b['code']} {b['label']}", "montant": b["solde_cloture"],
                                   "racine": _cle_racine_bilan(racine)})
    creances = _grouper_avec_sous_total(creances_flat, "racine", RACINE_LABELS_CREANCES)

    dettes_flat = []
    for racine in [str(r) for r in range(40, 50)]:
        for b in _detail_racine(balance, racine, sign="neg"):
            dettes_flat.append({"label": f"{b['code']} {b['label']}", "montant": -b["solde_cloture"],
                                 "racine": _cle_racine_bilan(racine)})
    dettes = _grouper_avec_sous_total(dettes_flat, "racine", RACINE_LABELS_DETTES)

    # ---- Trésorerie (classe 5), groupée Banques (50-56) / Caisse (57-59)
    # comme le rapport de référence ----
    treso_lignes, _ = compute_tresorerie_detail(conn, exercice=exercice, balance=balance)
    treso_actif_flat = [{"label": f"{t['code']} {t['label']}", "montant": t["solde_cloture"],
                          "prefixe": _cle_prefixe_treso(t["code"])}
                        for t in treso_lignes if t["solde_cloture"] > 0]
    treso_actif = _grouper_avec_sous_total(treso_actif_flat, "prefixe", TRESO_LABELS)
    treso_passif_flat = [{"label": f"{t['code']} {t['label']}", "montant": -t["solde_cloture"],
                           "prefixe": _cle_prefixe_treso(t["code"])}
                         for t in treso_lignes if t["solde_cloture"] < 0]
    treso_passif = _grouper_avec_sous_total(treso_passif_flat, "prefixe", TRESO_LABELS)

    # ---- Capitaux propres et ressources durables (classe 1), par préfixe —
    # emprunts 16+17 combinés en une seule ligne « Emprunts bancaires »,
    # comme le rapport de référence ----
    capitaux_labels = {
        "10": "Capital", "11": "Réserves", "12": "Report à nouveau",
        "13_14_15": "Résultats antérieurs (racines 13-15)",
        "16_17": "Emprunts bancaires (racines 16-17)",
        "18": "Comptes de liaison des établissements et sociétés en participation",
        "19": "Provisions financières pour risques et charges",
    }
    capitaux_flat = []
    for prefixe in ("10", "11", "12", "13", "14", "15", "16", "17", "18", "19"):
        if prefixe in ("13", "14", "15"):
            cle = "13_14_15"
        elif prefixe in ("16", "17"):
            cle = "16_17"
        else:
            cle = prefixe
        for b in _detail_prefix2(balance, "1", prefixe):
            capitaux_flat.append({"label": f"{b['code']} {b['label']}", "montant": -b["solde_cloture"], "prefixe": cle})
    capitaux_propres = _grouper_avec_sous_total(capitaux_flat, "prefixe", capitaux_labels)
    somme_classee = sum(g["sous_total"] for g in capitaux_propres)
    ressources_durables_total = -_sum_class(balance, "1")
    autres_ress = ressources_durables_total - somme_classee
    if abs(autres_ress) >= 1:
        capitaux_propres.append({"key": "autres", "label": "Autres postes de ressources durables (non classés)",
                                  "comptes": [], "sous_total": autres_ress})
    resultat_net = bilan["passif"]["Résultat net de l'exercice"]
    capitaux_propres.append({"key": "resultat", "label": "Résultat net de l'exercice",
                              "comptes": [], "sous_total": resultat_net})

    return {
        "immobilisations": immobilisations, "total_immo_brut": total_brut, "total_immo_amort": total_amort,
        "total_immo_net": bilan["actif"]["Immobilisations nettes"],
        "stocks": stocks, "total_stocks": bilan["actif"]["Stocks"],
        "creances": creances, "total_creances": bilan["actif"]["Créances et emplois assimilés"],
        "tresorerie_actif": treso_actif, "total_tresorerie_actif": bilan["actif"]["Trésorerie actif"],
        "capitaux_propres": capitaux_propres, "total_capitaux_propres": ressources_durables_total + resultat_net,
        "dettes": dettes, "total_dettes": bilan["passif"]["Dettes circulantes"],
        "tresorerie_passif": treso_passif, "total_tresorerie_passif": bilan["passif"]["Trésorerie passif"],
        "total_actif": bilan["total_actif"], "total_passif": bilan["total_passif"],
    }


def _merge_bilan_lignes(lignes_n, lignes_n1, key_field, value_fields):
    """Fusion COMPLÈTE (union) de deux listes de lignes/groupes du Bilan par
    clé — contrairement à un simple rattachement N -> N-1, une ligne
    présente UNIQUEMENT en N-1 (ex. un compte SANS mouvement cette
    période, mais avec un solde d'ouverture non nul) est CONSERVÉE, avec
    des valeurs à 0 côté N — sinon elle disparaîtrait purement et
    simplement du Bilan alors qu'elle a un solde d'ouverture réel."""
    par_cle_n = {g[key_field]: g for g in lignes_n}
    par_cle_n1 = {g[key_field]: g for g in lignes_n1}
    toutes_cles = list(dict.fromkeys(list(par_cle_n.keys()) + list(par_cle_n1.keys())))
    result = []
    for cle in toutes_cles:
        gn, gn1 = par_cle_n.get(cle), par_cle_n1.get(cle)
        base = dict(gn) if gn else dict(gn1)
        for vf in value_fields:
            base[vf] = gn[vf] if gn else 0.0
            base[vf + "_n1"] = gn1[vf] if gn1 else 0.0
        result.append(base)
    return result


def _merge_n1(groupes_n, groupes_n1, key_field="key", montant_field="sous_total"):
    """Attache à chaque groupe/ligne de `groupes_n` le montant N-1 du groupe
    correspondant dans `groupes_n1` (même clé) — 0 si l'exercice N-1 n'a
    pas ce groupe (compte inexistant l'an dernier, par ex.)."""
    par_cle_n1 = {g[key_field]: g[montant_field] for g in groupes_n1}
    for g in groupes_n:
        g[montant_field + "_n1"] = par_cle_n1.get(g[key_field], 0.0)
    return groupes_n


def compute_bilan_detaille(conn, exercice=None):
    """Bilan présenté ligne par ligne (« avec détails »), MONTÉ SUR LA BASE
    DU SOLDE DE CLÔTURE HABITUEL (solde d'ouverture + cumul des opérations
    de la période — voir compute_bilan()). La colonne « N-1 » de ce Bilan
    contient le solde d'ouverture de l'exercice (le report à nouveau du
    1er janvier) — qui correspond mathématiquement au solde de clôture de
    l'exercice précédent dès lors que la clôture d'exercice a été utilisée
    normalement (voir close_exercice()), donc une vraie comparaison N-1
    sans avoir à recalculer un second Bilan complet sur un exercice
    antérieur. ACTIF en Brut / Amortissements et provisions / Net
    (immobilisations par catégorie AVEC sous-totaux Brut/Amortissements,
    stocks par nature, créances compte par compte GROUPÉES PAR RACINE avec
    un sous-total par racine) ; PASSIF détaillé de la même façon (capitaux
    propres par racine, dettes groupées par racine, trésorerie créditrice
    groupée Banques/Caisse). S'appuie sur exactement les mêmes fonctions de
    classification que compute_bilan() — donc toujours équilibré (Actif =
    Passif) sur CHACUNE des deux colonnes indépendamment, la partie double
    garantissant que la somme des soldes de clôture comme celle des soldes
    d'ouverture est nulle."""
    exercice = exercice or get_current_exercice(conn)
    exercice_n1 = str(int(exercice) - 1)

    n = _compute_bilan_groupes(conn, exercice)  # solde de clôture standard (ouverture + mouvements)

    balance_ouverture = _balance_solde_ouverture(conn, exercice)
    resultat_ouverture = _resultat_net_mouvement_periode(balance_ouverture)
    n1 = _compute_bilan_groupes(conn, exercice, balance=balance_ouverture, resultat_net_override=resultat_ouverture)

    n["immobilisations"] = _merge_bilan_lignes(n["immobilisations"], n1["immobilisations"], "key",
                                                ["brut", "amort", "net"])
    n["immobilisations"].sort(key=lambda l: -max(abs(l["brut"]), abs(l.get("brut_n1", 0))))
    n["stocks"] = _merge_bilan_lignes(n["stocks"], n1["stocks"], "key", ["sous_total"])
    n["creances"] = _merge_bilan_lignes(n["creances"], n1["creances"], "key", ["sous_total"])
    n["tresorerie_actif"] = _merge_bilan_lignes(n["tresorerie_actif"], n1["tresorerie_actif"], "key", ["sous_total"])
    n["capitaux_propres"] = _merge_bilan_lignes(n["capitaux_propres"], n1["capitaux_propres"], "key", ["sous_total"])
    n["dettes"] = _merge_bilan_lignes(n["dettes"], n1["dettes"], "key", ["sous_total"])
    n["tresorerie_passif"] = _merge_bilan_lignes(n["tresorerie_passif"], n1["tresorerie_passif"], "key",
                                                  ["sous_total"])

    return {
        "exercice": exercice, "exercice_n1": exercice_n1,
        "actif": {
            "immobilisations": n["immobilisations"],
            "total_immo_brut": n["total_immo_brut"], "total_immo_brut_n1": n1["total_immo_brut"],
            "total_immo_amort": n["total_immo_amort"], "total_immo_amort_n1": n1["total_immo_amort"],
            "total_immo_net": n["total_immo_net"], "total_immo_net_n1": n1["total_immo_net"],
            "stocks": n["stocks"], "total_stocks": n["total_stocks"], "total_stocks_n1": n1["total_stocks"],
            "creances": n["creances"], "total_creances": n["total_creances"], "total_creances_n1": n1["total_creances"],
            "tresorerie": n["tresorerie_actif"], "total_tresorerie": n["total_tresorerie_actif"],
            "total_tresorerie_n1": n1["total_tresorerie_actif"],
        },
        "total_actif": n["total_actif"], "total_actif_n1": n1["total_actif"],
        "passif": {
            "capitaux_propres": n["capitaux_propres"],
            "total_capitaux_propres": n["total_capitaux_propres"],
            "total_capitaux_propres_n1": n1["total_capitaux_propres"],
            "dettes": n["dettes"], "total_dettes": n["total_dettes"], "total_dettes_n1": n1["total_dettes"],
            "tresorerie": n["tresorerie_passif"], "total_tresorerie": n["total_tresorerie_passif"],
            "total_tresorerie_n1": n1["total_tresorerie_passif"],
        },
        "total_passif": n["total_passif"], "total_passif_n1": n1["total_passif"],
        "ecart": n["total_actif"] - n["total_passif"],
        "ecart_n1": n1["total_actif"] - n1["total_passif"],
    }


# ---------------------------------------------------------------------------
# Génération des états financiers (Bilan, Compte de Résultat, Situation
# Financière, TFT) à partir des GABARITS EXCEL EXACTS du projet de référence
# de l'utilisateur (templates/modele_*.xlsx + bilan_template.xls) — moteur
# d'évaluation multi-passes des formules CtaCptSolde.../[Rxxx.EtLoc], porté
# tel quel du projet de référence (« bilan-auto ») fourni par l'utilisateur,
# adapté pour lire les soldes directement depuis CETTE application (Balance
# SQLite) au lieu d'exiger l'import de deux fichiers de balance externes.
# ---------------------------------------------------------------------------
def _generer_template_depuis_b64(nom_fichier, b64_data):
    """Régénère un gabarit Excel à la volée depuis des données encodées en
    base64 (déjà chargées par l'appelant via un `import` STATIQUE — voir
    les fonctions ci-dessous) — élimine toute dépendance au bundle
    PyInstaller. IMPORTANT : ne jamais importer le module de données via
    importlib/nom dynamique ici, PyInstaller ne détecterait pas la
    dépendance et ne l'inclurait pas dans l'exécutable (déjà arrivé)."""
    base = os.environ.get("LOCALAPPDATA") or os.path.expanduser("~")
    folder = os.path.join(base, "SaisieComptable")
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, nom_fichier)
    if not os.path.exists(path):
        import base64
        with open(path, "wb") as f:
            f.write(base64.b64decode(b64_data))
    return path


def _bilan_template_path():
    """Emplacement du gabarit Bilan — régénéré à la volée depuis les
    données encodées en base64 dans bilan_template_data.py si le fichier
    n'existe pas encore (ou plus) à cet endroit. Écrit dans le même
    dossier persistant que la base de données (%LOCALAPPDATA%\\SaisieComptable
    sous Windows) plutôt que dans le dossier d'installation/d'extraction
    PyInstaller — élimine TOUTE dépendance au bundle --add-data du build,
    qui s'est révélé peu fiable en pratique (fichier absent de
    l'exécutable compilé malgré une configuration correcte)."""
    import bilan_template_data
    return _generer_template_depuis_b64("bilan_template.xls", bilan_template_data.BILAN_TEMPLATE_B64)


def _cr_template_path():
    """Emplacement du gabarit Compte de résultat (SIG) — même principe que
    _bilan_template_path()."""
    import etats_financiers_data
    return _generer_template_depuis_b64("resultat_template.xls", etats_financiers_data.CR_TEMPLATE_B64)


def _tft_template_path():
    """Emplacement du gabarit TFT (Tableau des flux de trésorerie) — même
    principe que _bilan_template_path()."""
    import etats_financiers_data
    return _generer_template_depuis_b64("tft_template.xls", etats_financiers_data.TFT_TEMPLATE_B64)


def _situation_template_path():
    """Emplacement du gabarit Situation financière (FR-BFR-TN) — même
    principe que _bilan_template_path()."""
    import etats_financiers_data
    return _generer_template_depuis_b64("situation_template.xls", etats_financiers_data.SITUATION_TEMPLATE_B64)


def get_app_icon_path():
    """Emplacement de l'icône de l'application (usine) — régénérée à la
    volée depuis les données encodées en base64 dans factory_icon_data.py,
    même principe que les gabarits d'états financiers (élimine toute
    dépendance au bundle PyInstaller)."""
    import factory_icon_data
    return _generer_template_depuis_b64("factory_icon.ico", factory_icon_data.FACTORY_ICON_B64)


def import_bilan_template(path_source):
    """Remplace le gabarit Bilan ACTIF par un fichier fourni par
    l'utilisateur (ses formules modifiées/corrigées) — validé au préalable
    (doit s'ouvrir correctement et contenir au moins une formule
    CtaCptSolde reconnue) avant d'écraser l'ancien. Toutes les futures
    utilisations (écran « Visionner », export « gabarit officiel »)
    utiliseront ce nouveau fichier."""
    try:
        wb = open_template_workbook(path_source)
    except Exception as exc:
        raise ValueError(f"Ce fichier n'a pas pu être ouvert comme gabarit Excel : {exc}")
    nb_formules = 0
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if is_formula(cell.value):
                    nb_formules += 1
    if nb_formules == 0:
        raise ValueError(
            "Aucune formule CtaCptSolde... reconnue dans ce fichier — vérifiez qu'il s'agit bien d'un "
            "gabarit Bilan valide (même structure que le fichier téléchargé via « Télécharger mon "
            "template »)."
        )
    dest = _bilan_template_path()
    import shutil
    shutil.copyfile(path_source, dest)
    return nb_formules


def restaurer_bilan_template_original():
    """Revient au gabarit Bilan d'ORIGINE (celui encodé dans le code
    source, bilan_template_data.py) — écrase tout template importé/modifié
    par l'utilisateur."""
    import base64
    import bilan_template_data
    dest = _bilan_template_path()
    with open(dest, "wb") as f:
        f.write(base64.b64decode(bilan_template_data.BILAN_TEMPLATE_B64))
    return dest


try:
    BILAN_TEMPLATE_PATH = _bilan_template_path()
except Exception:
    # Ne doit jamais empêcher le reste de l'application de démarrer — les
    # fonctions qui utilisent ce chemin (export_bilan_gabarit_xlsx...)
    # retentent la régénération à chaque appel si besoin (voir plus bas).
    BILAN_TEMPLATE_PATH = None

ETATS_TEMPLATES = {
    "bilan": {"path": os.path.join(_resource_dir(), "templates", "modele_bilan.xlsx"),
              "sheet_hint": "BILAN", "output_name": "Bilan.xlsx"},
    "resultat": {"path": os.path.join(_resource_dir(), "templates", "modele_resultat.xlsx"),
                 "sheet_hint": "Feuil1", "output_name": "Compte_de_Resultat.xlsx"},
    "situation": {"path": os.path.join(_resource_dir(), "templates", "modele_situation.xlsx"),
                  "sheet_hint": "Feuil1", "output_name": "Situation_Financiere.xlsx"},
    "flux": {"path": os.path.join(_resource_dir(), "templates", "modele_flux.xlsx"),
             "sheet_hint": "Feuil1", "output_name": "Flux_de_Tresorerie.xlsx"},
}


def _racine_range(prefixes):
    clean = [p.rstrip("*") for p in prefixes]
    if len(clean) == 1:
        return clean[0], clean[0]
    return clean[0], clean[1]


def _compte_dans_plage(compte, debut, fin):
    if debut == fin:
        return compte.startswith(debut)
    L = max(len(debut), len(fin))
    borne_min = int(debut.ljust(L, "0"))
    borne_max = int(fin.ljust(L, "9"))
    prefix_compte = compte[:L].ljust(L, "0")
    try:
        val = int(prefix_compte)
    except ValueError:
        return False
    return borne_min <= val <= borne_max


def _comptes_matching(soldes, prefixes):
    debut, fin = _racine_range(prefixes)
    for compte in soldes:
        if _compte_dans_plage(compte, debut, fin):
            yield compte


def cta_cpt_solde_debit(soldes, *prefixes):
    return sum(soldes[c] for c in _comptes_matching(soldes, prefixes) if soldes[c] > 0)


def cta_cpt_solde_credit(soldes, *prefixes):
    return sum(-soldes[c] for c in _comptes_matching(soldes, prefixes) if soldes[c] < 0)


def cta_cpt_solde(soldes, *prefixes):
    return sum(soldes[c] for c in _comptes_matching(soldes, prefixes))


_FUNC_TOKENS = [
    ("CtaCptSoldeDébitNm1", "__F_DEBIT_NM1__"), ("CtaCptSoldeDebitNm1", "__F_DEBIT_NM1__"),
    ("CtaCptSoldeCréditNm1", "__F_CREDIT_NM1__"), ("CtaCptSoldeCreditNm1", "__F_CREDIT_NM1__"),
    ("CtaCptSoldeNm1", "__F_SOLDE_NM1__"),
    ("CtaCptSoldeDébit", "__F_DEBIT__"), ("CtaCptSoldeDebit", "__F_DEBIT__"),
    ("CtaCptSoldeCrédit", "__F_CREDIT__"), ("CtaCptSoldeCredit", "__F_CREDIT__"),
    ("CtaCptSolde", "__F_SOLDE__"),
    ("Ratio(", "__F_RATIO__("),
    ("kNum1", "1"),  # simple constante numérique de mise en forme (nombre de décimales) — sans effet sur le calcul
]
_RUBRIQUE_REF_RE = re.compile(r"\[([A-Za-z0-9_]+)\.EtLoc\]")
_RUBRIQUE_ASSIGN_RE = re.compile(r"^\[([A-Za-z0-9_]+)\.EtLoc\]=(.*)$")


class FormulaError(Exception):
    pass


class RubriqueNotReady(Exception):
    """Levée quand une formule référence une rubrique [Rxxx.EtLoc] pas
    encore calculée : on réessaiera lors d'une passe ultérieure (les
    formules d'un gabarit ne sont pas forcément dans l'ordre de dépendance —
    ex. un total en haut de page qui additionne des rubriques définies plus
    bas)."""
    pass


def is_formula(value):
    if not isinstance(value, str):
        return False
    v = value.strip()
    if not v.startswith("="):
        return False
    return ("CtaCptSolde" in v) or bool(_RUBRIQUE_REF_RE.search(v))


def _prepare_expr(formula):
    expr = formula.strip()
    if expr.startswith("="):
        expr = expr[1:]
    rubrique_id = None
    m = _RUBRIQUE_ASSIGN_RE.match(expr)
    if m:
        rubrique_id = m.group(1)
        expr = m.group(2)
    expr = _RUBRIQUE_REF_RE.sub(lambda mo: "__RUB__(%r)" % mo.group(1), expr)
    for name, token in _FUNC_TOKENS:
        expr = expr.replace(name, token)
    if "CtaCptSolde" in expr:
        raise FormulaError("Fonction non reconnue : %s" % formula)
    leftover = re.sub(r"__F_[A-Z0-9_]+__|__RUB__\([^)]*\)", "", expr)
    if re.search(r"[A-Za-zÀ-ÿ]", leftover):
        raise FormulaError("Fonction ou référence non reconnue : %s" % formula)
    return rubrique_id, expr


def _make_namespace(soldes_n, soldes_n1, rubrique_values, defined_ids=None, missing_used=None):
    def RUB(rubrique_id):
        if rubrique_id in rubrique_values:
            return rubrique_values[rubrique_id]
        if defined_ids is not None and rubrique_id not in defined_ids:
            if missing_used is not None:
                missing_used.add(rubrique_id)
            return 0
        raise RubriqueNotReady(rubrique_id)

    return {
        "__F_DEBIT__": lambda *p: cta_cpt_solde_debit(soldes_n, *p),
        "__F_CREDIT__": lambda *p: cta_cpt_solde_credit(soldes_n, *p),
        "__F_SOLDE__": lambda *p: cta_cpt_solde(soldes_n, *p),
        "__F_DEBIT_NM1__": lambda *p: cta_cpt_solde_debit(soldes_n1, *p),
        "__F_CREDIT_NM1__": lambda *p: cta_cpt_solde_credit(soldes_n1, *p),
        "__F_SOLDE_NM1__": lambda *p: cta_cpt_solde(soldes_n1, *p),
        "__F_RATIO__": lambda value, *rest: value,  # Ratio(valeur, décimales, unité...) — mise en forme d'affichage, ignorée ici
        "__RUB__": RUB,
    }


def evaluate_formula(formula, soldes_n, soldes_n1, rubrique_values=None, defined_ids=None, missing_used=None):
    rubrique_values = rubrique_values if rubrique_values is not None else {}
    rubrique_id, py_expr = _prepare_expr(formula)
    namespace = _make_namespace(soldes_n, soldes_n1, rubrique_values, defined_ids, missing_used)
    try:
        value = eval(py_expr, {"__builtins__": {}}, namespace)
    except RubriqueNotReady:
        raise
    except Exception as e:
        raise FormulaError("Erreur d'évaluation (%s) : %s" % (formula, e))
    return value, rubrique_id


def evaluate_sheet_formulas(ws, soldes_n, soldes_n1):
    """Évalue toutes les cellules-formules d'une feuille en plusieurs passes
    (résout les dépendances entre cellules liées par des rubriques
    [xxx.EtLoc], quel que soit leur ordre dans la feuille). Renvoie
    (results, errors, warnings)."""
    pending = []
    for row in ws.iter_rows():
        for cell in row:
            if is_formula(cell.value):
                pending.append((cell, cell.value))

    defined_ids = set()
    for _cell, formula in pending:
        s = formula.strip()
        m = _RUBRIQUE_ASSIGN_RE.match(s[1:] if s.startswith("=") else s)
        if m:
            defined_ids.add(m.group(1))

    rubrique_values, results, errors, warnings = {}, {}, [], []
    missing_used = set()
    max_passes = len(pending) + 2
    for _ in range(max_passes):
        if not pending:
            break
        still_pending, progress = [], False
        for cell, formula in pending:
            try:
                value, rubrique_id = evaluate_formula(formula, soldes_n, soldes_n1, rubrique_values,
                                                        defined_ids, missing_used)
            except RubriqueNotReady:
                still_pending.append((cell, formula))
                continue
            except FormulaError as e:
                errors.append((cell.coordinate, formula, str(e)))
                progress = True
                continue
            results[cell.coordinate] = value
            if rubrique_id:
                rubrique_values[rubrique_id] = value
            if missing_used:
                for rid in missing_used:
                    warnings.append((cell.coordinate, formula,
                                      "Rubrique [%s.EtLoc] jamais définie : traitée comme 0" % rid))
                missing_used.clear()
            progress = True
        pending = still_pending
        if not progress:
            break
    for cell, formula in pending:
        errors.append((cell.coordinate, formula, "Rubrique référencée jamais calculée (dépendance manquante)"))
    return results, errors, warnings


_SS_NS = "urn:schemas-microsoft-com:office:spreadsheet"


def _is_spreadsheetml(path):
    try:
        with open(path, "rb") as f:
            head = f.read(300)
        head_txt = head.decode("utf-8", errors="ignore")
        return "<?xml" in head_txt and ("mso-application" in head_txt or "Workbook" in head_txt)
    except Exception:
        return False


def _spreadsheetml_to_workbook(path):
    """Convertit un vieux fichier XML SpreadsheetML (Excel 2003, souvent
    nommé .xls) en classeur openpyxl équivalent (comme le gabarit de Bilan
    fourni par l'utilisateur — voir templates/bilan_template.xls)."""
    import openpyxl as _openpyxl
    tree_xml = __import__("xml.etree.ElementTree", fromlist=["ElementTree"]).parse(path)
    root = tree_xml.getroot()
    wb_out = _openpyxl.Workbook()
    wb_out.remove(wb_out.active)

    def tag(name):
        return "{%s}%s" % (_SS_NS, name)

    for ws_el in root.findall(tag("Worksheet")):
        sheet_name = ws_el.get(tag("Name")) or "Sheet"
        ws_out = wb_out.create_sheet(title=sheet_name[:31])
        table = ws_el.find(tag("Table"))
        if table is None:
            continue
        row_idx = 0
        for row_el in table.findall(tag("Row")):
            idx_attr = row_el.get(tag("Index"))
            row_idx = int(idx_attr) if idx_attr else row_idx + 1
            col_idx = 0
            for cell_el in row_el.findall(tag("Cell")):
                idx_attr = cell_el.get(tag("Index"))
                col_idx = int(idx_attr) if idx_attr else col_idx + 1
                data_el = cell_el.find(tag("Data"))
                if data_el is not None:
                    val = data_el.text
                    dtype = data_el.get(tag("Type"), "String")
                    if val is not None and dtype == "Number":
                        try:
                            fval = float(val)
                            val = int(fval) if fval.is_integer() else fval
                        except ValueError:
                            pass
                    if val is not None:
                        ws_out.cell(row=row_idx, column=col_idx, value=val)
                span = cell_el.get(tag("MergeAcross"))
                if span:
                    col_idx += int(span)
    return wb_out


def open_template_workbook(template_path):
    """Ouvre un gabarit Excel — détecte automatiquement l'ancien format
    SpreadsheetML (XML, souvent nommé .xls) et le convertit. Auto-réparant :
    si le premier essai échoue (mauvaise détection, encodage inhabituel...),
    retente l'AUTRE méthode avant d'abandonner avec un message clair —
    pour ne jamais planter toute l'application sur un simple souci de
    lecture d'un gabarit."""
    import openpyxl as _openpyxl
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            f"Le gabarit Excel « {template_path} » est introuvable. Réinstallez l'application avec "
            f"une version complète (le dossier « templates » doit être présent à côté de l'exécutable)."
        )
    essai_spreadsheetml = _is_spreadsheetml(template_path)
    try:
        if essai_spreadsheetml:
            return _spreadsheetml_to_workbook(template_path)
        return _openpyxl.load_workbook(template_path, data_only=False)
    except Exception as premiere_erreur:
        try:
            if essai_spreadsheetml:
                return _openpyxl.load_workbook(template_path, data_only=False)
            return _spreadsheetml_to_workbook(template_path)
        except Exception:
            raise ValueError(
                f"Impossible de lire le gabarit Excel « {template_path} » (fichier corrompu ou format "
                f"non reconnu) : {premiere_erreur}"
            )


def _guess_etat_sheet(wb, preferred):
    if preferred in wb.sheetnames:
        return preferred
    best_name, best_count = None, -1
    for name in wb.sheetnames:
        count = sum(1 for row in wb[name].iter_rows() for cell in row if is_formula(cell.value))
        if count > best_count:
            best_name, best_count = name, count
    if best_count <= 0:
        raise ValueError("Aucune feuille du modèle ne contient de formules CtaCptSolde... — vérifiez le fichier.")
    return best_name


def _soldes_dict(conn, exercice):
    """{code_compte: solde_cloture} pour un exercice — alimente le moteur
    de formules CtaCptSolde... directement depuis la Balance de cette
    application (au lieu d'exiger l'import d'un fichier de balance externe,
    comme le fait le projet de référence)."""
    try:
        balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    except Exception:
        return {}
    soldes = {}
    for b in balance:
        soldes[b["code"]] = soldes.get(b["code"], 0.0) + b["solde_cloture"]
    return soldes


def _soldes_ouverture_dict(conn, exercice):
    """{code_compte: solde_ouverture} — alimente les formules …Nm1 (« N-1 »)
    des états financiers (Compte de résultat, TFT, Situation financière)
    avec le solde d'OUVERTURE de l'exercice courant, PAS un second exercice
    antérieur distinct (qui n'existe généralement pas dans la base tant que
    l'utilisateur n'a pas créé et alimenté un exercice N-1 complet — le
    solde d'ouverture de N correspond déjà mathématiquement au solde de
    clôture de N-1). Même principe que _balance_solde_ouverture() pour le
    Bilan."""
    try:
        balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    except Exception:
        return {}
    return {b["code"]: b["solde_ouverture"] for b in balance}


def generate_etat_xlsx(conn, etat_id, output_path, exercice=None):
    """Génère UN état financier (Bilan / Compte de Résultat / Situation
    Financière / TFT) à partir de son gabarit officiel
    (voir ETATS_TEMPLATES) et de la Balance de CETTE application (exercice
    demandé pour N, exercice-1 pour N-1) — même moteur de formules que les
    4 gabarits partagent. Retourne un dict {cells_ok, cells_error,
    cells_warning} pour affichage d'un éventuel diagnostic."""
    info = ETATS_TEMPLATES.get(etat_id)
    if not info:
        raise ValueError(f"État inconnu : {etat_id}")
    exercice = exercice or get_current_exercice(conn)
    exercice_n1 = str(int(exercice) - 1)
    soldes_n = _soldes_dict(conn, exercice)
    soldes_n1 = _soldes_ouverture_dict(conn, exercice)

    wb = open_template_workbook(info["path"])
    actual_sheet = _guess_etat_sheet(wb, preferred=info["sheet_hint"])
    ws = wb[actual_sheet]

    results, errors, warnings = evaluate_sheet_formulas(ws, soldes_n, soldes_n1)
    for coord, value in results.items():
        ws[coord] = value
    for coord, formula, msg in errors:
        ws[coord] = "#ERREUR"

    wb.save(output_path)
    return {"cells_ok": len(results), "cells_error": errors, "cells_warning": warnings}


def compute_cr(conn, exercice=None):
    """Compte de résultat (SIG) — enveloppe RPC-compatible de
    compute_etat_formule_generique() (dont le 2e argument est une fonction
    Python, non transmissible par réseau)."""
    return compute_etat_formule_generique(conn, _cr_template_path, exercice=exercice)


def compute_tft_gabarit(conn, exercice=None):
    """TFT (gabarit officiel à formules) — enveloppe RPC-compatible de
    compute_etat_formule_generique(). Nommée différemment de
    compute_tft() (fonction plus ancienne, méthode de calcul différente)
    pour éviter toute collision de nom dans ce module."""
    return compute_etat_formule_generique(conn, _tft_template_path, exercice=exercice)


def compute_situation_fin(conn, exercice=None):
    """Situation financière (FR-BFR-TN) — enveloppe RPC-compatible de
    compute_etat_formule_generique()."""
    return compute_etat_formule_generique(conn, _situation_template_path, exercice=exercice)


def compute_etat_formule_generique(conn, template_path_getter, exercice=None):
    """Lit un gabarit à une seule colonne de libellés (A) suivie de 1 à 3
    colonnes de valeurs (détectées via la ligne d'en-tête : « N », « N-1 »,
    « % ») et évalue chaque formule avec le moteur CtaCptSolde/…Nm1 — même
    principe que compute_bilan_plat(), généralisé pour être réutilisé par
    le Compte de résultat (SIG), le TFT et la Situation financière, qui
    partagent tous cette même structure « RUBRIQUE | N (| N-1 | %) »."""
    exercice = exercice or get_current_exercice(conn)
    exercice_n1 = str(int(exercice) - 1)
    soldes_n = _soldes_dict(conn, exercice)
    soldes_n1 = _soldes_ouverture_dict(conn, exercice)

    wb = open_template_workbook(template_path_getter())
    ws = wb[wb.sheetnames[0]]

    # Repère la ligne d'en-tête (« RUBRIQUE » en colonne A) pour savoir
    # quelles colonnes de valeurs existent dans CE gabarit précis.
    header_row = None
    for r in range(1, min(ws.max_row, 15) + 1):
        if isinstance(ws.cell(row=r, column=1).value, str) and \
                ws.cell(row=r, column=1).value.strip().upper() == "RUBRIQUE":
            header_row = r
            break
    value_cols = []  # [(lettre_colonne, libelle_colonne), ...]
    if header_row:
        for col_idx in range(2, 6):
            h = ws.cell(row=header_row, column=col_idx).value
            if isinstance(h, str) and h.strip():
                value_cols.append((openpyxl_col_letter(col_idx), h.strip()))
    if not value_cols:
        value_cols = [("B", "N")]

    results, errors, warnings = evaluate_sheet_formulas(ws, soldes_n, soldes_n1)

    def cellval(col, row):
        coord = f"{col}{row}"
        if coord in results:
            return results[coord]
        v = ws[coord].value
        return v if isinstance(v, (int, float)) else None

    lignes = []
    start_row = (header_row + 1) if header_row else 1
    for r in range(start_row, ws.max_row + 1):
        libelle = ws.cell(row=r, column=1).value
        if not (isinstance(libelle, str) and libelle.strip()):
            continue
        valeurs = {label: cellval(col, r) for col, label in value_cols}
        if not any(v not in (None, "") for v in valeurs.values()):
            continue
        lignes.append({"libelle": libelle.strip(), **valeurs})

    return {"exercice": exercice, "exercice_n1": exercice_n1, "colonnes": [label for _, label in value_cols],
            "lignes": lignes, "errors": errors}


def openpyxl_col_letter(idx):
    import openpyxl.utils
    return openpyxl.utils.get_column_letter(idx)


def compute_bilan_plat(conn, exercice=None):
    """Bilan « plat » — relit DIRECTEMENT le gabarit officiel
    (templates/modele_bilan.xlsx, ou à défaut bilan_template.xls) ligne
    par ligne, et évalue CHAQUE formule (CtaCptSolde/CtaCptSoldeDébit/
    CtaCptSoldeCrédit et leurs variantes N-1 …Nm1) avec exactement le même
    moteur que l'export .xlsx officiel (evaluate_sheet_formulas) — donc
    strictement les mêmes valeurs, y compris la colonne N-1, sans aucune
    logique de catégorisation ou de regroupement propre à l'application.
    Colonnes du gabarit : A=Libellé Actif, B=Brut, C=Amortissements,
    D=Net, E=Net N-1, F=Libellé Passif, G=Exercice N, H=Exercice N-1.
    Renvoie {exercice, exercice_n1, actif: [...], passif: [...], errors}."""
    exercice = exercice or get_current_exercice(conn)
    exercice_n1 = str(int(exercice) - 1)
    soldes_n = _soldes_dict(conn, exercice)
    soldes_n1 = _soldes_ouverture_dict(conn, exercice)

    info = ETATS_TEMPLATES["bilan"]
    try:
        if os.path.exists(info["path"]):
            wb = open_template_workbook(info["path"])
            actual_sheet = _guess_etat_sheet(wb, preferred=info["sheet_hint"])
        else:
            raise FileNotFoundError(info["path"])
    except Exception:
        # Repli sur l'ancien gabarit SpreadsheetML si le nouveau modèle .xlsx
        # est absent ou illisible (ex. build antérieure sans templates/modele_bilan.xlsx) —
        # ne remonte l'erreur que si les DEUX échouent.
        wb = open_template_workbook(BILAN_TEMPLATE_PATH or _bilan_template_path())
        actual_sheet = _guess_etat_sheet(wb, "BILAN")
    ws = wb[actual_sheet]

    results, errors, warnings = evaluate_sheet_formulas(ws, soldes_n, soldes_n1)

    def cellval(col, row):
        coord = f"{col}{row}"
        if coord in results:
            return results[coord]
        v = ws[coord].value
        return v if isinstance(v, (int, float)) else None

    rows_actif, rows_passif = [], []
    junk = {"ACTIF", "PASSIF", "BILAN SYNTHETIQUE"}
    for r in range(1, ws.max_row + 1):
        libelle_a = ws.cell(row=r, column=1).value
        if (isinstance(libelle_a, str) and libelle_a.strip()
                and libelle_a.strip() not in junk and not libelle_a.strip().isdigit()
                and not libelle_a.strip().startswith("Edition du")):
            rows_actif.append({
                "libelle": libelle_a.strip(),
                "brut": cellval("B", r), "amort": cellval("C", r),
                "net": cellval("D", r), "net_n1": cellval("E", r),
            })
        libelle_p = ws.cell(row=r, column=6).value
        if isinstance(libelle_p, str) and libelle_p.strip() and libelle_p.strip() not in junk:
            rows_passif.append({
                "libelle": libelle_p.strip(),
                "montant": cellval("G", r), "montant_n1": cellval("H", r),
            })

    return {"exercice": exercice, "exercice_n1": exercice_n1, "actif": rows_actif, "passif": rows_passif,
            "errors": errors}


def export_etat_formule_xls(conn, template_path_getter, path, exercice=None):
    """Exporte un état basé sur un gabarit à formules CtaCptSolde... (voir
    compute_etat_formule_generique()) par SUBSTITUTION DE TEXTE DIRECTE sur
    le XML brut (même mécanisme robuste que export_bilan_gabarit_xlsx) —
    préserve intégralement la mise en forme d'origine."""
    exercice = exercice or get_current_exercice(conn)
    exercice_n1 = str(int(exercice) - 1)
    soldes_n = _soldes_dict(conn, exercice)
    soldes_n1 = _soldes_ouverture_dict(conn, exercice)

    with open(template_path_getter(), encoding="utf-8") as f:
        content = f.read()

    # Pré-scan des rubriques réellement définies quelque part dans le
    # gabarit — comme evaluate_sheet_formulas(), pour distinguer une
    # rubrique référencée AVANT sa définition (à réessayer plus tard,
    # ordre du document) d'une rubrique jamais définie nulle part (0,
    # sans erreur — cas de références orphelines déjà présentes dans le
    # gabarit d'origine).
    raw_formulas = re.findall(r'<Data ss:Type="String"[^>]*>(=[^<]*)</Data>', content)
    defined_ids = set()
    for raw in raw_formulas:
        unescaped = (raw.replace("&quot;", '"').replace("&amp;", "&")
                     .replace("&lt;", "<").replace("&gt;", ">").replace("&apos;", "'"))
        m = _RUBRIQUE_ASSIGN_RE.match(unescaped[1:] if unescaped.startswith("=") else unescaped)
        if m:
            defined_ids.add(m.group(1))

    named_refs = {}
    pattern = re.compile(r'<Data ss:Type="String"[^>]*>(=[^<]*)</Data>')

    def repl(m):
        raw = m.group(1)
        unescaped = (raw.replace("&quot;", '"').replace("&amp;", "&")
                     .replace("&lt;", "<").replace("&gt;", ">").replace("&apos;", "'"))
        try:
            value, defined_name = evaluate_formula(unescaped, soldes_n, soldes_n1, rubrique_values=named_refs,
                                                     defined_ids=defined_ids)
        except RubriqueNotReady:
            return m.group(0)  # sera résolu à une passe ultérieure
        except Exception:
            return m.group(0)
        if defined_name:
            named_refs[defined_name] = value
        return f'<Data ss:Type="Number">{value:.4f}</Data>'

    # Plusieurs passes (comme evaluate_sheet_formulas) : une rubrique peut
    # être référencée avant d'être définie plus bas dans le document.
    new_content = content
    for _ in range(5):
        previous = new_content
        new_content = pattern.sub(repl, previous)
        if new_content == previous:
            break
    with open(path, "w", encoding="utf-8") as f:
        f.write(new_content)
    return path


def export_bilan_gabarit_xlsx(conn, path, exercice=None):
    """Exporte le Bilan dans le gabarit officiel (templates/modele_bilan.xlsx,
    ou à défaut l'ancien gabarit SpreadsheetML templates/bilan_template.xls
    si le premier est absent ou illisible) — voir generate_etat_xlsx().
    Le repli sur l'ancien gabarit se fait par SUBSTITUTION DE TEXTE
    DIRECTE sur le XML brut (pas de passage par openpyxl) — préserve
    intégralement la mise en forme d'origine (couleurs, bordures,
    largeurs de colonnes), contrairement à une conversion openpyxl qui
    perdrait ces styles."""
    try:
        if not os.path.exists(ETATS_TEMPLATES["bilan"]["path"]):
            raise FileNotFoundError(ETATS_TEMPLATES["bilan"]["path"])
        generate_etat_xlsx(conn, "bilan", path, exercice=exercice)
    except Exception:
        exercice = exercice or get_current_exercice(conn)
        exercice_n1 = str(int(exercice) - 1)
        soldes_n = _soldes_dict(conn, exercice)
        soldes_n1 = _soldes_ouverture_dict(conn, exercice)

        with open(BILAN_TEMPLATE_PATH or _bilan_template_path(), encoding="utf-8") as f:
            content = f.read()

        named_refs = {}
        pattern = re.compile(r'<Data ss:Type="String"[^>]*>(=[^<]*)</Data>')

        def repl(m):
            raw = m.group(1)
            unescaped = (raw.replace("&quot;", '"').replace("&amp;", "&")
                         .replace("&lt;", "<").replace("&gt;", ">").replace("&apos;", "'"))
            value, defined_name = _eval_bilan_formula_legacy(unescaped, soldes_n, soldes_n1, named_refs)
            if defined_name:
                named_refs[defined_name] = value
            return f'<Data ss:Type="Number">{value:.4f}</Data>'

        new_content = pattern.sub(repl, content)
        with open(path, "w", encoding="utf-8") as f:
            f.write(new_content)

    return path


def _eval_bilan_formula_legacy(formula, soldes_n, soldes_n1, named_refs):
    """Évalue une formule du gabarit SpreadsheetML historique (même syntaxe
    que evaluate_formula, mais avec des références nommées [Rxxx.EtLoc]
    résolues au fil de l'eau via un dict partagé plutôt que rubrique_values
    — utilisé uniquement par le repli texte-brut de export_bilan_gabarit_xlsx,
    pour préserver la mise en forme d'origine)."""
    value, defined_name = evaluate_formula(formula, soldes_n, soldes_n1, rubrique_values=named_refs,
                                            defined_ids=None, missing_used=None)
    return value, defined_name


def export_bilan_detaille_xlsx(conn, path, exercice=None):
    """Exporte le Bilan détaillé (mêmes données que l'onglet Bilan) en .xlsx,
    dans une mise en page proche du rapport financier de référence : ACTIF
    (Brut / Amortissements / Net) à gauche, PASSIF (Montant) à droite, avec
    la MÊME palette de couleurs par racine de compte que le PDF de référence
    (40 Fournisseurs en orange, 41 Clients en bleu, 42 Personnel en jaune,
    43 CNSS en rose, 44/45 État en gris, 46 à 49 HAO/divers en cyan, classe 5
    Trésorerie en vert, stocks en bleu clair, sous-totaux en bleu clair/or,
    total général en or) — la même racine a la même couleur des deux côtés."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    exercice = exercice or get_current_exercice(conn)
    d = compute_bilan_detaille(conn, exercice=exercice)

    RACINE_COLORS = {
        "40": ("FF6600", "FFFFFF"), "41": ("3366FF", "FFFFFF"),
        "42": ("FFFF00", "000000"), "43": ("FF99CC", "000000"),
        "44": ("999999", "FFFFFF"), "45": ("999999", "FFFFFF"),
        "46": ("00FFFF", "000000"), "47": ("00FFFF", "000000"),
        "48": ("00FFFF", "000000"), "49": ("00FFFF", "000000"),
    }
    STOCK_COLOR = ("99CCFF", "000000")
    TRESO_COLOR = ("00FF00", "000000")
    SOUS_TOTAL_ACTIF = ("99CCFF", "000000")
    SOUS_TOTAL_PASSIF = ("FFCC00", "000000")
    GRAND_TOTAL = ("FFCC00", "000000")

    def fill_font(hexbg, hexfg, bold=True):
        return PatternFill("solid", fgColor=hexbg), Font(bold=bold, color=hexfg)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BILAN"

    title_font = Font(bold=True, size=13)
    section_font = Font(bold=True, color="FFFFFFFF")
    section_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="D9D9D9")

    ws["A1"] = f"BILAN — Exercice {exercice} (comparatif exercice {d['exercice_n1']})"
    ws["A1"].font = title_font
    ws["A2"] = "Calculé à partir de la Balance générale (comptes classés un par un, Actif = Passif garanti)."
    ws.merge_cells("A2:I2")

    row = 4
    ws.cell(row=row, column=1, value="ACTIF").font = section_font
    ws.cell(row=row, column=1).fill = section_fill
    ws.cell(row=row, column=7, value="PASSIF").font = section_font
    ws.cell(row=row, column=7).fill = section_fill
    for c in (2, 3, 4, 5, 6, 8, 9):
        ws.cell(row=row, column=c).fill = section_fill
    row += 1
    ws.cell(row=row, column=1, value="Libellé").font = header_font
    ws.cell(row=row, column=2, value="Brut").font = header_font
    ws.cell(row=row, column=3, value="Amortissements").font = header_font
    ws.cell(row=row, column=4, value="Net").font = header_font
    ws.cell(row=row, column=5, value="Net N-1").font = header_font
    ws.cell(row=row, column=7, value="Libellé").font = header_font
    ws.cell(row=row, column=8, value="Exercice N").font = header_font
    ws.cell(row=row, column=9, value="Exercice N-1").font = header_font
    for c in (1, 2, 3, 4, 5, 7, 8, 9):
        ws.cell(row=row, column=c).fill = header_fill
    header_row = row
    row += 1
    actif_start = row
    passif_start = row

    def row_color(item):
        key = item.get("key")
        if key and key in RACINE_COLORS:
            return RACINE_COLORS[key]
        if key in ("31", "32", "33", "34", "35", "36", "37", "38", "39"):
            return STOCK_COLOR
        return None

    def write_actif_section(ws, row, titre, lignes, total_label, total_val, total_val_n1=0.0, detail=False,
                             soustotal_color=SOUS_TOTAL_ACTIF, fixed_color=None):
        ws.cell(row=row, column=1, value=titre).font = header_font
        row += 1
        for l in lignes:
            color = fixed_color or row_color(l)
            if detail:
                ws.cell(row=row, column=1, value=f"  {l['label']}")
                ws.cell(row=row, column=2, value=l["brut"] or None).number_format = "#,##0"
                ws.cell(row=row, column=3, value=l["amort"] or None).number_format = "#,##0"
                ws.cell(row=row, column=4, value=l["net"]).number_format = "#,##0"
                ws.cell(row=row, column=5, value=l.get("net_n1") or None).number_format = "#,##0"
            else:
                montant = l.get("sous_total", l.get("montant", 0))
                montant_n1 = l.get("sous_total_n1", l.get("montant_n1", 0))
                ws.cell(row=row, column=1, value=f"  {l['label']}")
                ws.cell(row=row, column=4, value=montant).number_format = "#,##0"
                ws.cell(row=row, column=5, value=montant_n1 or None).number_format = "#,##0"
            if color:
                fill, font = fill_font(*color)
                for c in range(1, 6):
                    cell = ws.cell(row=row, column=c)
                    cell.fill = fill
                    cell.font = font
            row += 1
        fill, font = fill_font(*soustotal_color)
        ws.cell(row=row, column=1, value=total_label)
        ws.cell(row=row, column=4, value=total_val).number_format = "#,##0"
        ws.cell(row=row, column=5, value=total_val_n1 or None).number_format = "#,##0"
        for c in range(1, 6):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).font = font
        return row + 2

    def write_passif_section(ws, row, titre, lignes, total_label, total_val, total_val_n1=0.0,
                              soustotal_color=SOUS_TOTAL_PASSIF, fixed_color=None):
        ws.cell(row=row, column=7, value=titre).font = header_font
        row += 1
        for l in lignes:
            color = fixed_color or row_color(l)
            montant = l.get("sous_total", l.get("montant", 0))
            montant_n1 = l.get("sous_total_n1", l.get("montant_n1", 0))
            ws.cell(row=row, column=7, value=f"  {l['label']}")
            ws.cell(row=row, column=8, value=montant).number_format = "#,##0"
            ws.cell(row=row, column=9, value=montant_n1 or None).number_format = "#,##0"
            if color:
                fill, font = fill_font(*color)
                for c in (7, 8, 9):
                    cell = ws.cell(row=row, column=c)
                    cell.fill = fill
                    cell.font = font
            row += 1
        fill, font = fill_font(*soustotal_color)
        ws.cell(row=row, column=7, value=total_label)
        ws.cell(row=row, column=8, value=total_val).number_format = "#,##0"
        ws.cell(row=row, column=9, value=total_val_n1 or None).number_format = "#,##0"
        for c in (7, 8, 9):
            ws.cell(row=row, column=c).fill = fill
            ws.cell(row=row, column=c).font = font
        return row + 2

    a = d["actif"]
    row = actif_start
    row = write_actif_section(ws, row, "IMMOBILISATIONS", a["immobilisations"],
                               "Total immobilisations BRUTES", a["total_immo_brut"], a["total_immo_brut_n1"],
                               detail=True)
    fill, font = fill_font(*SOUS_TOTAL_ACTIF)
    ws.cell(row=row, column=1, value="  Total AMORTISSEMENTS et provisions")
    ws.cell(row=row, column=4, value=a["total_immo_amort"]).number_format = "#,##0"
    ws.cell(row=row, column=5, value=a["total_immo_amort_n1"] or None).number_format = "#,##0"
    for c in range(1, 6):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).font = font
    row += 1
    ws.cell(row=row, column=1, value="  Total immobilisations NETTES")
    ws.cell(row=row, column=4, value=a["total_immo_net"]).number_format = "#,##0"
    ws.cell(row=row, column=5, value=a["total_immo_net_n1"] or None).number_format = "#,##0"
    for c in range(1, 6):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).font = font
    row += 2
    row = write_actif_section(ws, row, "STOCKS", a["stocks"], "Total stocks", a["total_stocks"],
                               a["total_stocks_n1"], fixed_color=STOCK_COLOR)
    row = write_actif_section(ws, row, "CRÉANCES", a["creances"], "Total créances", a["total_creances"],
                               a["total_creances_n1"])
    row = write_actif_section(ws, row, "TRÉSORERIE ACTIF", a["tresorerie"], "Total trésorerie actif",
                               a["total_tresorerie"], a["total_tresorerie_n1"], fixed_color=TRESO_COLOR)
    fill, font = fill_font(*GRAND_TOTAL, bold=True)
    ws.cell(row=row, column=1, value="TOTAL ACTIF")
    ws.cell(row=row, column=4, value=d["total_actif"]).number_format = "#,##0"
    ws.cell(row=row, column=5, value=d["total_actif_n1"] or None).number_format = "#,##0"
    for c in range(1, 6):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).font = font
    actif_end = row

    p = d["passif"]
    row = passif_start
    row = write_passif_section(ws, row, "CAPITAUX PROPRES ET RESSOURCES DURABLES", p["capitaux_propres"],
                                "Total capitaux propres et ressources durables", p["total_capitaux_propres"],
                                p["total_capitaux_propres_n1"])
    row = write_passif_section(ws, row, "DETTES CIRCULANTES", p["dettes"], "Total dettes circulantes",
                                p["total_dettes"], p["total_dettes_n1"])
    row = write_passif_section(ws, row, "TRÉSORERIE PASSIF", p["tresorerie"], "Total trésorerie passif",
                                p["total_tresorerie"], p["total_tresorerie_n1"], fixed_color=TRESO_COLOR)
    fill, font = fill_font(*GRAND_TOTAL, bold=True)
    ws.cell(row=row, column=7, value="TOTAL PASSIF")
    ws.cell(row=row, column=8, value=d["total_passif"]).number_format = "#,##0"
    ws.cell(row=row, column=9, value=d["total_passif_n1"] or None).number_format = "#,##0"
    for c in (7, 8, 9):
        ws.cell(row=row, column=c).fill = fill
        ws.cell(row=row, column=c).font = font
    passif_end = row

    last_row = max(actif_end, passif_end) + 2
    ecart = d["ecart"]
    if abs(ecart) < 1:
        ws.cell(row=last_row, column=1, value=f"Écart Actif - Passif : {ecart:,.0f}  ✓ équilibré").font = Font(bold=True)
    else:
        diag = compute_ecart_diagnostic(conn, exercice=exercice)
        ws.cell(row=last_row, column=1,
                value=f"Écart Actif - Passif : {ecart:,.0f} ⚠ — cause(s) détectée(s) dans les données :"
                ).font = Font(bold=True, color="FFB00020")
        r = last_row + 1
        if abs(diag["ecart_soldes_ouverture"]) >= 1:
            ws.cell(row=r, column=1,
                    value=f"• Soldes d'ouverture non nuls : {diag['ecart_soldes_ouverture']:,.0f} "
                          f"(devrait être 0 — voir l'onglet « Soldes d'ouverture »)")
            r += 1
        if abs(diag["ecart_ecritures_periode"]) >= 1:
            ws.cell(row=r, column=1,
                    value=f"• Écritures de la période Débit ≠ Crédit : {diag['ecart_ecritures_periode']:,.0f} "
                          f"(voir l'onglet « Écritures non équilibrées »)")
            r += 1

    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 3
    ws.column_dimensions["G"].width = 46
    ws.column_dimensions["H"].width = 18
    ws.column_dimensions["I"].width = 18
    ws.freeze_panes = f"A{header_row + 1}"
    wb.save(path)
    return path


def compute_grand_livre(conn, compte, tiers=None, date_from=None, date_to=None, exercice=None):
    """Détail chronologique des écritures d'un compte pour un exercice, avec
    solde cumulé démarrant au solde d'ouverture de l'exercice."""
    exercice = exercice or get_current_exercice(conn)
    if date_from is None:
        date_from = f"{exercice}-01-01"
    if date_to is None:
        date_to = f"{exercice}-12-31"
    query = "SELECT * FROM entries WHERE compte = ? AND date >= ? AND date <= ?"
    params = [compte, date_from, date_to]
    if tiers:
        query += " AND tiers LIKE ?"
        params.append(f"%{tiers}%")
    query += " ORDER BY date, id"
    rows = [dict(r) for r in conn.execute(query, params).fetchall()]
    solde = get_opening_balance(conn, compte, exercice)
    for r in rows:
        solde += r["debit"] - r["credit"]
        r["solde_cumule"] = solde
    return rows


def compute_grand_livre_complet(conn, exercice=None, compte_prefix=None, tiers=None):
    """Grand livre complet : TOUS les comptes ayant un solde d'ouverture ou un
    mouvement sur l'exercice, groupés par compte puis par classe, avec pour
    chaque compte sa ligne « À-nouveaux », le détail chronologique de ses
    écritures, un solde cumulé, et un sous-total de compte (avec le sens du
    solde, débiteur ou créditeur) — puis un total par classe. `compte_prefix`
    filtre optionnellement sur un préfixe de compte (ex. « 60 » ou « 601000 »)."""
    exercice = exercice or get_current_exercice(conn)
    date_from, date_to = f"{exercice}-01-01", f"{exercice}-12-31"
    balance = compute_balance(conn, only_with_movement=True, exercice=exercice)
    balance.sort(key=lambda b: b["code"])

    classes = {}
    for b in balance:
        if compte_prefix and not b["code"].startswith(compte_prefix):
            continue
        classes.setdefault(b["classe"], []).append(b)

    result_classes = []
    for classe in sorted(classes.keys()):
        comptes = []
        classe_total_debit = classe_total_credit = 0.0
        for b in classes[classe]:
            query = "SELECT * FROM entries WHERE compte = ? AND date >= ? AND date <= ?"
            params = [b["code"], date_from, date_to]
            if tiers:
                query += " AND tiers LIKE ?"
                params.append(f"%{tiers}%")
            query += " ORDER BY date, id"
            lignes = [dict(r) for r in conn.execute(query, params).fetchall()]
            solde = b["solde_ouverture"]
            for l in lignes:
                solde += l["debit"] - l["credit"]
                l["solde_cumule"] = solde
            solde_final = solde
            comptes.append({
                "code": b["code"], "label": b["label"],
                "solde_ouverture": b["solde_ouverture"],
                "lignes": lignes,
                "total_debit": b["debit"], "total_credit": b["credit"],
                "solde_final": solde_final,
                "sens": "débiteur" if solde_final >= 0 else "créditeur",
            })
            classe_total_debit += b["debit"]
            classe_total_credit += b["credit"]
        if comptes:
            result_classes.append({
                "classe": classe, "comptes": comptes,
                "total_debit": classe_total_debit, "total_credit": classe_total_credit,
            })
    return result_classes


# ---------------------------------------------------------------------------
# Stocks
# ---------------------------------------------------------------------------
def compute_stocks(conn, exercice=None):
    exercice = exercice or get_current_exercice(conn)
    date_from, date_to = f"{exercice}-01-01", f"{exercice}-12-31"
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    by_code = {b["code"]: b for b in balance}
    result = []
    for code in COMPTES_STOCK:
        b = by_code.get(code, {"label": get_account_label(conn, code), "debit": 0.0, "credit": 0.0,
                                "solde_ouverture": 0.0})
        initial = get_opening_balance(conn, code, exercice)
        entrees, sorties = b["debit"], b["credit"]
        qte_row = conn.execute(
            """SELECT COALESCE(SUM(CASE WHEN debit > 0 THEN quantite ELSE 0 END), 0) AS qte_in,
                      COALESCE(SUM(CASE WHEN credit > 0 THEN quantite ELSE 0 END), 0) AS qte_out
               FROM entries WHERE compte = ? AND date >= ? AND date <= ?""",
            (code, date_from, date_to),
        ).fetchone()
        qte_entrees, qte_sorties = qte_row["qte_in"], qte_row["qte_out"]
        qte_initiale = get_setting(conn, f"stock_qte_initiale_{code}_{exercice}", 0.0)
        qte_finale = qte_initiale + qte_entrees - qte_sorties
        stock_final = initial + entrees - sorties
        cout_unitaire_moyen = (stock_final / qte_finale) if qte_finale else None
        result.append({
            "code": code, "label": b["label"], "stock_initial": initial,
            "entrees": entrees, "sorties": sorties,
            "stock_final": stock_final,
            "qte_initiale": qte_initiale, "qte_entrees": qte_entrees, "qte_sorties": qte_sorties,
            "qte_finale": qte_finale, "cout_unitaire_moyen": cout_unitaire_moyen,
        })
    return result


def compute_stocks_detail(conn, exercice=None, prefixes=None):
    """Détail du stock pour CHAQUE compte réel de la classe 3 (pas seulement les
    4 comptes centralisateurs) : tout compte 3xxxxx ayant un mouvement ou un
    solde d'ouverture sur l'exercice. `prefixes` (ex. ["31"], ["32"], ["36"])
    restreint optionnellement à une catégorie (marchandises/matières/produits
    finis)."""
    exercice = exercice or get_current_exercice(conn)
    date_from, date_to = f"{exercice}-01-01", f"{exercice}-12-31"
    balance = compute_balance(conn, only_with_movement=True, exercice=exercice)
    result = []
    for b in balance:
        if b["classe"] != "3":
            continue
        if prefixes and not any(b["code"].startswith(p) for p in prefixes):
            continue
        code = b["code"]
        initial = b["solde_ouverture"]
        entrees, sorties = b["debit"], b["credit"]
        qte_row = conn.execute(
            """SELECT COALESCE(SUM(CASE WHEN debit > 0 THEN quantite ELSE 0 END), 0) AS qte_in,
                      COALESCE(SUM(CASE WHEN credit > 0 THEN quantite ELSE 0 END), 0) AS qte_out
               FROM entries WHERE compte = ? AND date >= ? AND date <= ?""",
            (code, date_from, date_to),
        ).fetchone()
        qte_entrees, qte_sorties = qte_row["qte_in"], qte_row["qte_out"]
        qte_initiale = get_setting(conn, f"stock_qte_initiale_{code}_{exercice}", 0.0)
        qte_finale = qte_initiale + qte_entrees - qte_sorties
        stock_final = b["solde_cloture"]
        cout_unitaire_moyen = (stock_final / qte_finale) if qte_finale else None
        result.append({
            "code": code, "label": b["label"], "stock_initial": initial,
            "entrees": entrees, "sorties": sorties, "stock_final": stock_final,
            "qte_initiale": qte_initiale, "qte_entrees": qte_entrees, "qte_sorties": qte_sorties,
            "qte_finale": qte_finale, "cout_unitaire_moyen": cout_unitaire_moyen,
        })
    result.sort(key=lambda r: r["code"])
    return result


def set_stock_qte_initiale(conn, code, value, exercice=None):
    exercice = exercice or get_current_exercice(conn)
    set_setting(conn, f"stock_qte_initiale_{code}_{exercice}", value)


def set_stock_initial(conn, code, value, exercice=None):
    set_opening_balance(conn, code, value, exercice=exercice)


def compute_mouvements_stocks(conn, exercice=None):
    """Détail chronologique de toutes les écritures sur les comptes de stock
    (classe 3), quelle que soit leur origine — saisie manuelle, ou générées
    automatiquement par la validation d'une Facture (vente) ou d'une Facture
    frs (achat). Pour chaque compte, les lignes sont triées par date et un
    cumul est tenu à la fois en VALEUR (solde du compte) et en QUANTITÉ
    (comme une fiche de stock), en partant du stock initial de l'exercice."""
    exercice = exercice or get_current_exercice(conn)
    date_from, date_to = f"{exercice}-01-01", f"{exercice}-12-31"
    stocks_synthese = {s["code"]: s for s in compute_stocks(conn, exercice=exercice)}

    result = []
    for code in COMPTES_STOCK:
        rows = conn.execute(
            """SELECT e.*, COALESCE(a.label, e.compte) AS compte_label FROM entries e
               LEFT JOIN accounts a ON a.code = e.compte
               WHERE e.compte = ? AND e.date >= ? AND e.date <= ?
               ORDER BY e.date, e.id""",
            (code, date_from, date_to),
        ).fetchall()
        synth = stocks_synthese.get(code, {})
        valeur_cumulee = synth.get("stock_initial", 0.0)
        qte_cumulee = synth.get("qte_initiale", 0.0)
        for r in rows:
            d = dict(r)
            libelle = d["libelle"] or ""
            if libelle.startswith("Entrée stock (auto) —") or libelle.startswith("Sortie stock (auto) —"):
                d["origine"] = "Saisie directe (auto)"
            elif libelle.startswith("Entrée stock —") or libelle.startswith("Sortie stock —"):
                d["origine"] = "Facturation" if libelle.startswith("Sortie stock —") else "Facture frs"
            else:
                d["origine"] = "Saisie manuelle"
            valeur_cumulee += (d["debit"] or 0) - (d["credit"] or 0)
            qte_cumulee += (d["quantite"] or 0) if (d["debit"] or 0) > 0 else -(d["quantite"] or 0)
            d["valeur_cumulee"] = valeur_cumulee
            d["qte_cumulee"] = qte_cumulee
            result.append(d)
    return result


# ---------------------------------------------------------------------------
# Maintenance & Énergie (menu MAINTENANCE-ÉNERGIE) — suivi par code
# analytique : les codes préfixés « ENERGIE- » (eau, électricité, essence...)
# et « MAINT- » (véhicules, bâtiments, machines...) sont créés dans le Plan
# analytique, puis utilisés en Saisie (champ Code analytique) et dans les
# lignes de recette de Fabrication (main-d'œuvre, énergie, autres charges)
# pour tout retrouver agrégé ici, par code, sur une période choisie.
# ---------------------------------------------------------------------------
PREFIX_ENERGIE = "ENERGIE-"
PREFIX_MAINTENANCE = "MAINT-"

SUGGESTIONS_ENERGIE = [
    ("ENERGIE-EAU", "Eau", "L"),
    ("ENERGIE-ELEC", "Électricité", "Kw"),
    ("ENERGIE-ESSENCE", "Essence / Carburant", "L"),
    ("ENERGIE-GASOIL", "Gasoil", "L"),
    ("ENERGIE-GAZ", "Gaz", "L"),
]
SUGGESTIONS_MAINTENANCE = [
    ("MAINT-VEHIC", "Maintenance véhicules", "H"),
    ("MAINT-BAT", "Maintenance bâtiments", "H"),
    ("MAINT-MACH", "Maintenance machines et équipements", "H"),
    ("MAINT-INFO", "Maintenance informatique", "H"),
]


def ajouter_codes_analytiques_suggeres(conn, suggestions):
    """Ajoute les codes analytiques suggérés (Énergie ou Maintenance) qui
    n'existent pas encore, SANS écraser un code déjà personnalisé par
    l'utilisateur. Retourne le nombre de codes effectivement ajoutés."""
    ajoutes = 0
    for code, label, unite in suggestions:
        if not analytic_code_exists(conn, code):
            add_analytic_code(conn, code, label, unite=unite)
            ajoutes += 1
    return ajoutes


def compute_couts_analytiques_categorie(conn, prefix, date_from=None, date_to=None, exercice=None):
    """Comptes analytiques d'une catégorie (préfixe de code, ex. « ENERGIE- »,
    « MAINT- »), avec le montant de charge de début de période / de la
    période / cumulé en fin de période, sur une période choisie (par défaut
    l'exercice entier) — alimenté par toute écriture de Saisie taguée avec
    ce code analytique, ainsi que par les lignes de recette de Fabrication
    qui lui sont associées. Ne compte QUE le côté charge (classe 6) de
    chaque écriture — la contrepartie (banque, caisse, fournisseur...) porte
    le même code analytique mais ne doit pas être comptée, sous peine
    d'obtenir toujours un solde net nul (comme compute_production/AN-FAB,
    qui applique la même règle). Un code est inclus dès qu'il a une charge
    sur la période ou avant."""
    exercice = exercice or get_current_exercice(conn)
    exercice_debut = f"{exercice}-01-01"
    date_from = date_from or exercice_debut
    date_to = date_to or f"{exercice}-12-31"

    codes = conn.execute(
        "SELECT code, label FROM analytic_codes WHERE code LIKE ? ORDER BY code", (f"{prefix}%",)
    ).fetchall()

    result = []
    for c in codes:
        code = c["code"]
        avant = conn.execute(
            """SELECT COALESCE(SUM(e.debit), 0) d, COALESCE(SUM(e.credit), 0) c
               FROM entries e
               WHERE e.analytic_code = ? AND substr(e.compte, 1, 1) = '6' AND e.date >= ? AND e.date < ?""",
            (code, exercice_debut, date_from),
        ).fetchone()
        charge_avant = avant["d"] - avant["c"]

        periode = conn.execute(
            """SELECT COALESCE(SUM(e.debit), 0) d, COALESCE(SUM(e.credit), 0) c
               FROM entries e
               WHERE e.analytic_code = ? AND substr(e.compte, 1, 1) = '6' AND e.date >= ? AND e.date <= ?""",
            (code, date_from, date_to),
        ).fetchone()
        debit_periode, credit_periode = periode["d"], periode["c"]
        charge_periode = debit_periode - credit_periode
        charge_cumulee = charge_avant + charge_periode

        if not (charge_avant or charge_periode or charge_cumulee):
            continue
        result.append({
            "code": code, "label": c["label"],
            "solde_debut_periode": charge_avant,
            "debit_periode": debit_periode, "credit_periode": credit_periode,
            "solde_fin_periode": charge_cumulee,
        })
    return result


def compute_cout_unitaire_moyen_analytique(conn, analytic_code, exercice=None, toutes_dates=False):
    """Coût unitaire moyen pondéré d'un code analytique — ex. F CFA par
    litre d'eau/gasoil/gaz, par kilowatt d'électricité, par heure de
    main-d'œuvre ou de maintenance. Même principe que le coût unitaire moyen
    des stocks (compute_stocks_detail) : le montant total des charges
    (classe 6) comptabilisées sous ce code, divisé par la quantité totale
    saisie sur ces mêmes lignes (champ Quantité de la Saisie — en litres,
    kilowatts ou heures selon l'unité du code, voir get_analytic_code_unite).
    Se met à jour tout seul après chaque facture saisie avec une quantité.
    Retourne None si aucune quantité n'a encore été renseignée pour ce code
    (coût unitaire pas encore calculable).
    `toutes_dates=True` cumule depuis le début plutôt que sur le seul
    exercice — plus stable quand peu de factures ont été saisies cette année."""
    if not analytic_code:
        return None
    params = [analytic_code]
    date_filter = ""
    if not toutes_dates:
        exercice = exercice or get_current_exercice(conn)
        date_filter = " AND e.date >= ? AND e.date <= ?"
        params += [f"{exercice}-01-01", f"{exercice}-12-31"]
    row = conn.execute(
        f"""SELECT COALESCE(SUM(e.debit), 0) d, COALESCE(SUM(e.credit), 0) c,
                   COALESCE(SUM(e.quantite), 0) q
            FROM entries e
            WHERE e.analytic_code = ? AND substr(e.compte, 1, 1) = '6'{date_filter}""",
        params,
    ).fetchone()
    if not row["q"]:
        return None
    return (row["d"] - row["c"]) / row["q"]


def compute_couts_analytiques_fabrication(conn, prefix):
    """Pour une catégorie de code analytique (Énergie ou Maintenance), les
    lignes de recette de Fabrication (main-d'œuvre, énergie, autres charges)
    qui lui sont associées, tous produits finis confondus — pour vérifier la
    cohérence entre le coût estimé en recette et le coût réel comptabilisé."""
    rows = conn.execute(
        """SELECT rl.*, pf.nom AS produit_nom FROM recette_lignes rl
           JOIN produits_finis pf ON pf.code = rl.produit_code
           WHERE rl.analytic_code LIKE ? ORDER BY rl.analytic_code, rl.produit_code""",
        (f"{prefix}%",),
    ).fetchall()
    return [dict(r) for r in rows]



# ---------------------------------------------------------------------------
# Production / coûts de fabrication (écritures taguées analytic_code = AN-FAB)
# ---------------------------------------------------------------------------
FLUX_FAB = "AN-FAB"
FAB_POSTES = [
    ("Matières premières et fournitures consommées", ["602", "604"]),
    ("Main-d'œuvre directe de production", ["661", "663", "664"]),
    ("Charges indirectes de fabrication", ["624", "625", "681"]),
]


def compute_production(conn, exercice=None):
    exercice = exercice or get_current_exercice(conn)
    date_from, date_to = f"{exercice}-01-01", f"{exercice}-12-31"
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)

    def net_produit(codes):
        d, c = _sum_accounts(balance, codes)
        return c - d

    ventes = net_produit(["702", "705", "706"])
    stock_d, stock_c = _sum_accounts(balance, ["360"])
    production_stockee = stock_d - stock_c
    valeur_production = ventes + production_stockee

    postes = []
    total_cout = 0.0
    for label, codes in FAB_POSTES:
        like_clause = " OR ".join("compte LIKE ?" for _ in codes)
        like_params = [f"{c}%" for c in codes]
        row = conn.execute(
            f"SELECT COALESCE(SUM(debit),0) d, COALESCE(SUM(credit),0) c FROM entries "
            f"WHERE ({like_clause}) AND analytic_code = ? AND date >= ? AND date <= ?",
            (*like_params, FLUX_FAB, date_from, date_to),
        ).fetchone()
        montant = row["d"] - row["c"]
        postes.append({"label": label, "comptes": ", ".join(codes), "montant": montant})
        total_cout += montant

    return {
        "ventes": ventes,
        "production_stockee": production_stockee,
        "valeur_production": valeur_production,
        "postes_cout": postes,
        "cout_production": total_cout,
        "marge": valeur_production - total_cout,
    }


# ---------------------------------------------------------------------------
# Recettes de fabrication (nomenclature / BOM) — combine matières premières
# (coût réel issu des stocks comptables), main-d'œuvre et énergie pour
# calculer un coût de production, puis un prix de vente suggéré (+ marge).
# ---------------------------------------------------------------------------
LIGNE_TYPES = {
    "matiere": "Matière première (depuis un compte de stock)",
    "main_oeuvre": "Main-d'œuvre",
    "energie": "Énergie",
    "amortissement": "Amortissement d'équipement (depuis une immobilisation)",
    "autre": "Autre charge de fabrication",
}


def add_produit_fini(conn, code, nom, description="", quantite_produite=1, marge_pourcentage=30,
                      compte_stock="360000"):
    conn.execute(
        """INSERT OR REPLACE INTO produits_finis
           (code, nom, description, quantite_produite, marge_pourcentage, compte_stock)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (code.strip(), nom.strip(), description, quantite_produite or 1, marge_pourcentage or 0,
         compte_stock or "360000"),
    )
    conn.commit()


def delete_produit_fini(conn, code):
    conn.execute("DELETE FROM recette_lignes WHERE produit_code = ?", (code,))
    conn.execute("DELETE FROM produits_finis WHERE code = ?", (code,))
    conn.commit()


def list_produits_finis(conn):
    return [dict(r) for r in conn.execute("SELECT * FROM produits_finis ORDER BY code").fetchall()]


def get_produit_fini(conn, code):
    row = conn.execute("SELECT * FROM produits_finis WHERE code = ?", (code,)).fetchone()
    return dict(row) if row else None


def add_recette_ligne(conn, produit_code, type_ligne, libelle, quantite, compte=None, cout_unitaire=None,
                       analytic_code=None):
    conn.execute(
        """INSERT INTO recette_lignes (produit_code, type_ligne, libelle, compte, quantite, cout_unitaire,
                                        analytic_code)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (produit_code, type_ligne, libelle, compte, quantite or 0, cout_unitaire, analytic_code or None),
    )
    conn.commit()


def delete_recette_ligne(conn, ligne_id):
    conn.execute("DELETE FROM recette_lignes WHERE id = ?", (ligne_id,))
    conn.commit()


def list_recette_lignes(conn, produit_code):
    return [dict(r) for r in conn.execute(
        "SELECT * FROM recette_lignes WHERE produit_code = ? ORDER BY id", (produit_code,)
    ).fetchall()]


STOCK_VARIATION_PAR_PREFIXE = {
    "31": "603100",  # Variations des stocks de marchandises
    "32": "603200",  # Variations des stocks de matières premières
    "33": "603300",  # Variations des stocks d'autres approvisionnements
    "36": "736000",  # Variations des stocks de produits finis
}


def _compte_variation_stock(compte_stock):
    prefix = (compte_stock or "")[:2]
    return STOCK_VARIATION_PAR_PREFIXE.get(prefix, "603200")


def compute_cout_production(conn, produit_code, exercice=None):
    """Calcule le coût de production d'un produit fini à partir de sa recette :
    - pour chaque ligne « matière première » liée à un compte de stock, le coût
      unitaire réel est repris automatiquement du coût unitaire moyen calculé
      dans l'onglet Stocks (valeur du stock / quantité) ;
    - pour chaque ligne « amortissement d'équipement » liée à un compte
      d'immobilisation (classe 2), le coût unitaire réel est repris
      automatiquement de l'amortissement RÉELLEMENT comptabilisé pour cet
      équipement, divisé par sa base de répartition (tonnes/an, heures/an —
      renseignée une fois dans l'écran Immobilisations), voir
      compute_cout_amortissement_unitaire() ;
    - pour chaque ligne « main-d'œuvre », « énergie » ou « autre » liée à un
      CODE ANALYTIQUE (ex. MAINT-MACH, ENERGIE-EAU), le coût unitaire réel est
      repris automatiquement du coût unitaire moyen pondéré de ce code (total
      des charges comptabilisées sous ce code / quantité totale saisie — en
      litres, kilowatts ou heures selon son unité), voir
      compute_cout_unitaire_moyen_analytique() — se met à jour tout seul
      après chaque facture saisie ;
    - sinon, le coût unitaire saisi manuellement sur la ligne est utilisé."""
    produit = get_produit_fini(conn, produit_code)
    if not produit:
        raise ValueError(f"Produit « {produit_code} » introuvable.")
    lignes = list_recette_lignes(conn, produit_code)
    stocks_by_code = {s["code"]: s for s in compute_stocks_detail(conn, exercice=exercice)}

    detail = []
    total = 0.0
    for l in lignes:
        cu = l["cout_unitaire"]
        source = "manuel"
        if l["type_ligne"] == "matiere" and l["compte"]:
            stock = stocks_by_code.get(l["compte"])
            if stock and stock["cout_unitaire_moyen"] is not None:
                cu = stock["cout_unitaire_moyen"]
                source = "stock (coût unitaire moyen)"
            elif cu is None:
                cu = 0.0
                source = "aucun coût connu — à saisir"
        elif l["type_ligne"] == "amortissement" and l["compte"]:
            cu_amort = compute_cout_amortissement_unitaire(conn, l["compte"], exercice=exercice)
            if cu_amort is not None:
                cu = cu_amort
                fiche = get_immobilisation_fiche(conn, l["compte"])
                unite = fiche.get("base_repartition_unite") or ""
                source = f"amortissement équipement{f' / {unite}' if unite else ''}"
            elif cu is None:
                cu = 0.0
                source = "base de répartition non renseignée (écran Immobilisations) — à saisir"
        elif l["type_ligne"] not in ("matiere", "amortissement") and l["analytic_code"]:
            cu_analytique = compute_cout_unitaire_moyen_analytique(conn, l["analytic_code"], toutes_dates=True)
            if cu_analytique is not None:
                cu = cu_analytique
                unite = get_analytic_code_unite(conn, l["analytic_code"]) or ""
                source = f"analytique (coût moyen pondéré{f' / {unite}' if unite else ''})"
            elif cu is None:
                cu = 0.0
                source = "aucune quantité comptabilisée sous ce code — à saisir"
        elif cu is None:
            cu = 0.0
            source = "à saisir"
        montant = (l["quantite"] or 0) * (cu or 0)
        detail.append({**l, "cout_unitaire_utilise": cu, "source_cout": source, "montant": montant})
        total += montant

    qte_produite = produit["quantite_produite"] or 1
    cout_unitaire_produit = total / qte_produite if qte_produite else 0.0
    marge_pct = produit["marge_pourcentage"] or 0
    prix_vente_unitaire = cout_unitaire_produit * (1 + marge_pct / 100)
    prix_vente_total = prix_vente_unitaire * qte_produite

    return {
        "produit": produit,
        "lignes": detail,
        "cout_production_total": total,
        "quantite_produite": qte_produite,
        "cout_unitaire_produit": cout_unitaire_produit,
        "marge_pourcentage": marge_pct,
        "prix_vente_unitaire": prix_vente_unitaire,
        "prix_vente_total": prix_vente_total,
        "marge_unitaire": prix_vente_unitaire - cout_unitaire_produit,
    }


def valider_fabrication(conn, produit_code, date_str=None, piece=None, exercice=None):
    """Valide une fabrication à partir de sa recette :
    - impute comptablement la consommation de chaque matière première (le
      stock réel utilisé — ex. 321001 CLINKER — diminue en QUANTITÉ et en
      VALEUR, contrepartie en compte de variation de stock 603xxx) ;
    - place le produit fini dans son compte de stock (classe 36) en QUANTITÉ
      et en VALEUR, au coût de production + la marge paramétrée (compte
      736000 en contrepartie).
    Retourne (résultat de compute_cout_production, avertissements)."""
    resultat = compute_cout_production(conn, produit_code, exercice=exercice)
    produit = resultat["produit"]
    exercice = exercice or get_current_exercice(conn)
    if date_str is None:
        today = date.today()
        date_str = today.strftime("%Y-%m-%d") if str(today.year) == exercice else f"{exercice}-01-01"
    piece = piece or f"FAB-{produit_code}-{date_str}"
    warnings = []

    for l in resultat["lignes"]:
        if l["type_ligne"] != "matiere" or not l["compte"]:
            continue
        montant = l["montant"]
        qte = l["quantite"] or 0
        if montant <= 0 or qte <= 0:
            continue
        contre_compte = _compte_variation_stock(l["compte"])
        add_entry(conn, date_str, piece, "OD", contre_compte, "", f"Consommation fabrication — {l['libelle']}",
                  montant, 0)
        add_entry(conn, date_str, piece, "OD", l["compte"], "", f"Consommation fabrication — {l['libelle']}",
                  0, montant, quantite=qte)

    valeur_produit_fini = resultat["prix_vente_total"]
    qte_produite = resultat["quantite_produite"]
    if valeur_produit_fini > 0 and qte_produite > 0:
        compte_stock_pf = produit["compte_stock"]
        contre_compte_pf = _compte_variation_stock(compte_stock_pf)
        add_entry(conn, date_str, piece, "OD", compte_stock_pf, "", f"Production — {produit['nom']}",
                  valeur_produit_fini, 0, quantite=qte_produite)
        add_entry(conn, date_str, piece, "OD", contre_compte_pf, "", f"Production — {produit['nom']}",
                  0, valeur_produit_fini)
    else:
        warnings.append("Valeur ou quantité produite nulle — aucune entrée en stock de produit fini comptabilisée.")

    return resultat, warnings


# ---------------------------------------------------------------------------
# Facturation clients — présente une facture (entête + lignes + pied de page),
# et sa validation envoie automatiquement les écritures comptables dans la
# Saisie : Débit Client (411xxx) pour le TTC, Crédit compte(s) de vente (70x)
# pour le HT de chaque ligne, Crédit TVA (443100) pour la taxe, et pour les
# lignes liées à un stock (marchandises 31 ou produits finis 36), une sortie
# de stock automatique (Débit compte de coût / Crédit compte de stock).
# ---------------------------------------------------------------------------
def create_facture_vente(conn, numero, date_facture, client_code, entete="", pied_page="",
                          tva_taux=None, tva_compte=None):
    if not numero or not str(numero).strip():
        raise ValueError("Le numéro de facture est obligatoire.")
    if not client_code or not str(client_code).strip():
        raise ValueError("Le client est obligatoire pour créer une facture.")
    if not client_exists(conn, client_code):
        raise ValueError(f"Le client « {client_code} » n'existe pas dans la liste des clients.")
    if tva_taux is None:
        tva_taux = get_setting(conn, "tva_taux_defaut", TVA_TAUX_DEFAUT)
    if tva_compte is None:
        tva_compte = get_text_setting(conn, "tva_compte_defaut", COMPTE_TVA_VENTES)
    cur = conn.execute(
        """INSERT INTO factures_vente (numero, date_facture, client_code, entete, pied_page, tva_taux,
                                        tva_compte, statut)
           VALUES (?, ?, ?, ?, ?, ?, ?, 'brouillon')""",
        (numero, date_facture, client_code, entete, pied_page, tva_taux, tva_compte),
    )
    conn.commit()
    return cur.lastrowid


def update_facture_vente(conn, facture_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE factures_vente SET {cols} WHERE id = ?", (*fields.values(), facture_id))
    conn.commit()


def delete_facture_vente(conn, facture_id):
    facture = get_facture_vente(conn, facture_id)
    if facture and facture["statut"] == "validee":
        raise ValueError("Impossible de supprimer une facture déjà validée (écritures envoyées en Saisie).")
    conn.execute("DELETE FROM facture_vente_lignes WHERE facture_id = ?", (facture_id,))
    conn.execute("DELETE FROM factures_vente WHERE id = ?", (facture_id,))
    conn.commit()


def get_facture_vente(conn, facture_id):
    row = conn.execute("SELECT * FROM factures_vente WHERE id = ?", (facture_id,)).fetchone()
    return dict(row) if row else None


def list_factures_vente(conn):
    rows = conn.execute("""
        SELECT f.*, COALESCE(c.raison_sociale, f.client_code) AS raison_sociale
        FROM factures_vente f LEFT JOIN clients c ON c.code = f.client_code
        ORDER BY f.date_facture DESC, f.id DESC
    """).fetchall()
    return [dict(r) for r in rows]


def add_ligne_facture_vente(conn, facture_id, compte_vente, libelle, quantite, prix_unitaire, analytic_code=None):
    facture = get_facture_vente(conn, facture_id)
    if not facture:
        raise ValueError(f"Facture ID {facture_id} introuvable.")
    if facture["statut"] != "brouillon":
        raise ValueError(
            "Impossible d'ajouter une ligne à une facture déjà validée (écritures déjà envoyées en Saisie)."
        )
    if not compte_vente or not account_exists(conn, compte_vente):
        raise ValueError(
            f"Le compte de vente « {compte_vente} » n'existe pas dans le plan comptable — "
            f"choisissez un compte dans la liste plutôt que de le saisir librement."
        )
    if not libelle or not libelle.strip():
        raise ValueError("Le libellé de la ligne est obligatoire.")
    if not quantite or quantite <= 0:
        raise ValueError("La quantité doit être strictement positive.")
    if not prix_unitaire or prix_unitaire <= 0:
        raise ValueError("Le prix unitaire doit être strictement positif.")
    if analytic_code and not analytic_code_exists(conn, analytic_code):
        raise ValueError(f"Le code analytique « {analytic_code} » n'existe pas dans le plan analytique.")
    conn.execute(
        """INSERT INTO facture_vente_lignes (facture_id, compte_vente, libelle, quantite, prix_unitaire,
                                               analytic_code)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (facture_id, compte_vente, libelle, quantite or 0, prix_unitaire or 0, analytic_code or None),
    )
    conn.commit()


def delete_ligne_facture_vente(conn, ligne_id):
    conn.execute("DELETE FROM facture_vente_lignes WHERE id = ?", (ligne_id,))
    conn.commit()


def list_lignes_facture_vente(conn, facture_id):
    rows = conn.execute(
        "SELECT * FROM facture_vente_lignes WHERE facture_id = ? ORDER BY id", (facture_id,)
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["montant_ht"] = (d["quantite"] or 0) * (d["prix_unitaire"] or 0)
        type_stock, stock_compte, cout_compte = _match_stock_mapping(d["compte_vente"], VENTE_STOCK_MAPPING) or (None, None, None)
        d["type_stock"] = type_stock
        d["stock_compte"] = stock_compte
        result.append(d)
    return result


def compute_facture_totals(conn, facture_id):
    facture = get_facture_vente(conn, facture_id)
    lignes = list_lignes_facture_vente(conn, facture_id)
    total_ht = sum(l["montant_ht"] for l in lignes)
    tva_taux = facture["tva_taux"] if facture else 0
    tva_montant = total_ht * (tva_taux or 0) / 100
    total_ttc = total_ht + tva_montant
    return {"total_ht": total_ht, "tva_taux": tva_taux, "tva_montant": tva_montant, "total_ttc": total_ttc}


def devalider_facture_vente(conn, facture_id):
    """Repasse une facture VALIDÉE en brouillon modifiable, en cas d'erreur
    sur les chiffres constatée après validation : supprime toutes les
    écritures comptables générées par sa validation (débit client, crédit
    ventes, TVA, sorties de stock — repérées par le couple piece=numéro de
    facture / journal='VE', unique à cette facture) puis remet son statut à
    « brouillon » afin que ses lignes redeviennent éditables dans l'onglet
    Facturation. Refuse si l'exercice comptable des écritures est clôturé.
    Retourne le nombre d'écritures supprimées."""
    facture = get_facture_vente(conn, facture_id)
    if not facture:
        raise ValueError("Facture introuvable.")
    if facture["statut"] != "validee":
        raise ValueError("Cette facture n'est pas validée — rien à corriger.")
    exercice_facture = _exercice_of_date(facture["date_facture"])
    if is_exercice_cloture(conn, exercice_facture):
        raise ValueError(
            f"L'exercice {exercice_facture} de cette facture est clôturé : impossible de la corriger."
        )
    ids = [r["id"] for r in conn.execute(
        "SELECT id FROM entries WHERE piece = ? AND journal = 'VE'", (facture["numero"],)
    ).fetchall()]
    deleted, errors = delete_entries_bulk(conn, ids)
    if errors:
        raise ValueError(
            "Impossible de corriger cette facture : " + " ; ".join(errors)
        )
    update_facture_vente(conn, facture_id, statut="brouillon")
    return deleted


def valider_facture_vente(conn, facture_id, exercice=None):
    """Envoie la facture en Saisie : une écriture équilibrée (Débit Client / Crédit
    ventes + TVA), plus une sortie de stock automatique pour chaque ligne liée à un
    compte de marchandises (31) ou de produits finis (36). Retourne la liste des
    avertissements (ex. coût unitaire de stock inconnu)."""
    facture = get_facture_vente(conn, facture_id)
    if not facture:
        raise ValueError("Facture introuvable.")
    if facture["statut"] == "validee":
        raise ValueError("Cette facture est déjà validée.")
    lignes = list_lignes_facture_vente(conn, facture_id)
    if not lignes:
        raise ValueError("La facture ne contient aucune ligne.")
    if not client_exists(conn, facture["client_code"]):
        raise ValueError(f"Le client « {facture['client_code']} » n'existe pas.")

    totals = compute_facture_totals(conn, facture_id)
    date_str = facture["date_facture"]
    piece = facture["numero"]
    warnings = []

    # Débit Client pour le TTC
    client = get_client(conn, facture["client_code"])
    tiers_label = client["raison_sociale"] if client else facture["client_code"]
    add_entry(conn, date_str, piece, "VE", "411000", tiers_label,
              facture["numero"], totals["total_ttc"], 0, client_code=facture["client_code"])

    # Crédit chaque compte de vente pour le HT de la ligne
    for l in lignes:
        add_entry(conn, date_str, piece, "VE", l["compte_vente"], "", l["libelle"],
                  0, l["montant_ht"], client_code=facture["client_code"], quantite=l["quantite"])

    # Crédit TVA facturée
    if totals["tva_montant"]:
        add_entry(conn, date_str, piece, "VE", facture.get("tva_compte") or COMPTE_TVA_VENTES, "",
                  f"TVA {totals['tva_taux']:g}% facture {piece}",
                  0, totals["tva_montant"])

    # Sortie de stock automatique pour les lignes liées aux marchandises/produits finis
    stocks_by_code = {s["code"]: s for s in compute_stocks(conn, exercice=exercice)}
    for l in lignes:
        if not l["type_stock"]:
            continue
        _, stock_compte, cout_compte = _match_stock_mapping(l["compte_vente"], VENTE_STOCK_MAPPING)
        stock = stocks_by_code.get(stock_compte)
        cout_unitaire = stock["cout_unitaire_moyen"] if stock else None
        if cout_unitaire is None:
            warnings.append(
                f"Ligne « {l['libelle']} » : coût unitaire du stock {stock_compte} inconnu — "
                f"aucune sortie de stock comptabilisée pour cette ligne (renseignez un stock initial "
                f"ou des entrées avec quantité dans l'onglet Stocks)."
            )
            continue
        montant_sortie = (l["quantite"] or 0) * cout_unitaire
        if montant_sortie <= 0:
            continue
        add_entry(conn, date_str, piece, "VE", cout_compte, "", f"Sortie stock — {l['libelle']}",
                  montant_sortie, 0)
        add_entry(conn, date_str, piece, "VE", stock_compte, "", f"Sortie stock — {l['libelle']}",
                  0, montant_sortie, quantite=l["quantite"])

    update_facture_vente(conn, facture_id, statut="validee", piece=piece)
    return warnings


# ---------------------------------------------------------------------------
# Factures fournisseurs (achats) — présente une facture d'achat (entête +
# lignes + pied de page), et sa validation envoie automatiquement les
# écritures comptables dans la Saisie : Débit compte(s) d'achat (60x) pour le
# HT de chaque ligne, Crédit Fournisseur (401xxx) pour le net à payer, Crédit
# retenue à la source (44x, paramétrable) le cas échéant. Pour les lignes
# liées à un stock (marchandises 31 ou matières premières 32), une entrée de
# stock automatique est comptabilisée (le stock augmente).
# ---------------------------------------------------------------------------
def create_facture_achat(conn, numero, date_facture, fournisseur_code, entete="", pied_page="",
                          retenue_taux=None, retenue_compte=None):
    if retenue_taux is None:
        retenue_taux = get_setting(conn, "retenue_taux_defaut", RETENUE_TAUX_DEFAUT)
    if retenue_compte is None:
        retenue_compte = get_text_setting(conn, "retenue_compte_defaut", COMPTE_RETENUE_DEFAUT)
    cur = conn.execute(
        """INSERT INTO factures_achat
           (numero, date_facture, fournisseur_code, entete, pied_page, retenue_taux, retenue_compte, statut)
           VALUES (?, ?, ?, ?, ?, ?, ?, 'brouillon')""",
        (numero, date_facture, fournisseur_code, entete, pied_page, retenue_taux, retenue_compte),
    )
    conn.commit()
    return cur.lastrowid


def update_facture_achat(conn, facture_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE factures_achat SET {cols} WHERE id = ?", (*fields.values(), facture_id))
    conn.commit()


def delete_facture_achat(conn, facture_id):
    facture = get_facture_achat(conn, facture_id)
    if facture and facture["statut"] == "validee":
        raise ValueError("Impossible de supprimer une facture déjà validée (écritures envoyées en Saisie).")
    conn.execute("DELETE FROM facture_achat_lignes WHERE facture_id = ?", (facture_id,))
    conn.execute("DELETE FROM factures_achat WHERE id = ?", (facture_id,))
    conn.commit()


def get_facture_achat(conn, facture_id):
    row = conn.execute("SELECT * FROM factures_achat WHERE id = ?", (facture_id,)).fetchone()
    return dict(row) if row else None


def list_factures_achat(conn):
    rows = conn.execute("""
        SELECT f.*, COALESCE(fo.raison_sociale, f.fournisseur_code) AS raison_sociale
        FROM factures_achat f LEFT JOIN fournisseurs fo ON fo.code = f.fournisseur_code
        ORDER BY f.date_facture DESC, f.id DESC
    """).fetchall()
    return [dict(r) for r in rows]


def add_ligne_facture_achat(conn, facture_id, compte_achat, libelle, quantite, prix_unitaire, analytic_code=None):
    conn.execute(
        """INSERT INTO facture_achat_lignes (facture_id, compte_achat, libelle, quantite, prix_unitaire,
                                               analytic_code)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (facture_id, compte_achat, libelle, quantite or 0, prix_unitaire or 0, analytic_code or None),
    )
    conn.commit()


def delete_ligne_facture_achat(conn, ligne_id):
    conn.execute("DELETE FROM facture_achat_lignes WHERE id = ?", (ligne_id,))
    conn.commit()


def list_lignes_facture_achat(conn, facture_id):
    rows = conn.execute(
        "SELECT * FROM facture_achat_lignes WHERE facture_id = ? ORDER BY id", (facture_id,)
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["montant_ht"] = (d["quantite"] or 0) * (d["prix_unitaire"] or 0)
        type_stock, stock_compte, contre_compte = _match_stock_mapping(d["compte_achat"], ACHAT_STOCK_MAPPING) or (None, None, None)
        d["type_stock"] = type_stock
        d["stock_compte"] = stock_compte
        result.append(d)
    return result


def compute_facture_achat_totals(conn, facture_id):
    facture = get_facture_achat(conn, facture_id)
    lignes = list_lignes_facture_achat(conn, facture_id)
    total_ht = sum(l["montant_ht"] for l in lignes)
    retenue_taux = facture["retenue_taux"] if facture else 0
    retenue_montant = total_ht * (retenue_taux or 0) / 100
    net_a_payer = total_ht - retenue_montant
    return {"total_ht": total_ht, "retenue_taux": retenue_taux, "retenue_montant": retenue_montant,
            "net_a_payer": net_a_payer}


def valider_facture_achat(conn, facture_id, exercice=None):
    """Envoie la facture d'achat en Saisie : une écriture équilibrée (Débit achats /
    Crédit fournisseur + retenue), plus une entrée de stock automatique pour chaque
    ligne liée à un compte de marchandises (31) ou de matières premières (32).
    Retourne la liste des avertissements."""
    facture = get_facture_achat(conn, facture_id)
    if not facture:
        raise ValueError("Facture introuvable.")
    if facture["statut"] == "validee":
        raise ValueError("Cette facture est déjà validée.")
    lignes = list_lignes_facture_achat(conn, facture_id)
    if not lignes:
        raise ValueError("La facture ne contient aucune ligne.")
    if not fournisseur_exists(conn, facture["fournisseur_code"]):
        raise ValueError(f"Le fournisseur « {facture['fournisseur_code']} » n'existe pas.")

    totals = compute_facture_achat_totals(conn, facture_id)
    date_str = facture["date_facture"]
    piece = facture["numero"]
    warnings = []

    # Débit chaque compte d'achat pour le HT de la ligne
    for l in lignes:
        add_entry(conn, date_str, piece, "AC", l["compte_achat"], "", l["libelle"],
                  l["montant_ht"], 0, fournisseur_code=facture["fournisseur_code"], quantite=l["quantite"])

    # Crédit Fournisseur pour le net à payer (HT - retenue)
    fournisseur = get_fournisseur(conn, facture["fournisseur_code"])
    tiers_label = fournisseur["raison_sociale"] if fournisseur else facture["fournisseur_code"]
    add_entry(conn, date_str, piece, "AC", "401000", tiers_label,
              facture["numero"], 0, totals["net_a_payer"], fournisseur_code=facture["fournisseur_code"])

    # Crédit retenue fiscale à la source, si applicable
    if totals["retenue_montant"]:
        add_entry(conn, date_str, piece, "AC", facture["retenue_compte"], "",
                  f"Retenue {totals['retenue_taux']:g}% facture {piece}",
                  0, totals["retenue_montant"])

    # Entrée de stock automatique pour les lignes liées aux marchandises/matières premières
    for l in lignes:
        if not l["type_stock"]:
            continue
        _, stock_compte, contre_compte = _match_stock_mapping(l["compte_achat"], ACHAT_STOCK_MAPPING)
        montant_entree = l["montant_ht"]
        if montant_entree <= 0:
            continue
        add_entry(conn, date_str, piece, "AC", stock_compte, "", f"Entrée stock — {l['libelle']}",
                  montant_entree, 0, quantite=l["quantite"])
        add_entry(conn, date_str, piece, "AC", contre_compte, "", f"Entrée stock — {l['libelle']}",
                  0, montant_entree)

    update_facture_achat(conn, facture_id, statut="validee", piece=piece)
    return warnings


def devalider_facture_achat(conn, facture_id):
    """Repasse une facture d'achat VALIDÉE en brouillon modifiable, en cas
    d'erreur sur les chiffres constatée après validation : supprime toutes
    les écritures comptables générées par sa validation (débit achats,
    crédit fournisseur, retenue à la source, entrées de stock — repérées
    par le couple piece=numéro de facture / journal='AC', unique à cette
    facture) puis remet son statut à « brouillon ». Refuse si l'exercice
    comptable des écritures est clôturé. Retourne le nombre d'écritures
    supprimées."""
    facture = get_facture_achat(conn, facture_id)
    if not facture:
        raise ValueError("Facture introuvable.")
    if facture["statut"] != "validee":
        raise ValueError("Cette facture n'est pas validée — rien à corriger.")
    exercice_facture = _exercice_of_date(facture["date_facture"])
    if is_exercice_cloture(conn, exercice_facture):
        raise ValueError(
            f"L'exercice {exercice_facture} de cette facture est clôturé : impossible de la corriger."
        )
    ids = [r["id"] for r in conn.execute(
        "SELECT id FROM entries WHERE piece = ? AND journal = 'AC'", (facture["numero"],)
    ).fetchall()]
    deleted, errors = delete_entries_bulk(conn, ids)
    if errors:
        raise ValueError("Impossible de corriger cette facture : " + " ; ".join(errors))
    update_facture_achat(conn, facture_id, statut="brouillon")
    return deleted


_UNITES_FR = ["", "un", "deux", "trois", "quatre", "cinq", "six", "sept", "huit", "neuf", "dix",
              "onze", "douze", "treize", "quatorze", "quinze", "seize", "dix-sept", "dix-huit", "dix-neuf"]
_DIZAINES_FR = ["", "", "vingt", "trente", "quarante", "cinquante", "soixante", "soixante", "quatre-vingt",
                "quatre-vingt"]


def _trois_chiffres_en_lettres(n):
    """Convertit un nombre de 0 à 999 en toutes lettres françaises."""
    if n == 0:
        return ""
    centaines, reste = divmod(n, 100)
    mots = []
    if centaines:
        mots.append(("cent" if centaines == 1 else _UNITES_FR[centaines] + " cent") + ("s" if centaines > 1 and reste == 0 else ""))
    if reste:
        if reste < 20:
            mots.append(_UNITES_FR[reste])
        else:
            dix, unite = divmod(reste, 10)
            if dix in (7, 9):
                # soixante-dix, quatre-vingt-dix : dizaine précédente + 10-19
                base = _DIZAINES_FR[dix]
                fin = _UNITES_FR[10 + unite]
                mots.append(f"{base}-{fin}" if unite or dix in (7, 9) else base)
            else:
                base = _DIZAINES_FR[dix]
                if unite == 0:
                    mots.append(base + ("s" if dix == 8 else ""))
                elif unite == 1 and dix not in (8,):
                    mots.append(f"{base} et un")
                else:
                    mots.append(f"{base}-{_UNITES_FR[unite]}")
    return " ".join(mots)


def nombre_en_lettres_fr(n):
    """Convertit un entier (francs CFA — pas de centimes) en toutes lettres
    françaises, ex. 500000 -> « cinq cent mille ». Utilisé pour la mention
    légale « Arrêtée la présente facture à la somme de... » sur les
    factures imprimées."""
    n = int(round(n))
    if n == 0:
        return "zéro"
    negatif = n < 0
    n = abs(n)
    tranches = []
    for diviseur, nom_sing, nom_plur in ((10**9, "milliard", "milliards"), (10**6, "million", "millions"),
                                          (10**3, "mille", "mille"), (1, "", "")):
        valeur, n = divmod(n, diviseur)
        if valeur:
            if diviseur == 1:
                tranches.append(_trois_chiffres_en_lettres(valeur))
            elif diviseur == 10**3:
                if valeur == 1:
                    tranches.append("mille")
                else:
                    tranches.append(f"{_trois_chiffres_en_lettres(valeur)} mille")
            else:
                nom = nom_sing if valeur == 1 else nom_plur
                tranches.append(f"{_trois_chiffres_en_lettres(valeur)} {nom}")
    resultat = " ".join(t for t in tranches if t)
    return ("moins " if negatif else "") + resultat


def _html_facture(titre, numero, date_facture, tiers_label, entete, lignes_rows, pied_page, totaux_rows):
    """Construit un document HTML simple, imprimable (Ctrl+P depuis le
    navigateur), commun aux factures de vente et d'achat."""
    lignes_html = "\n".join(
        f"<tr><td>{l[0]}</td><td style='text-align:right'>{l[1]:,.2f}</td>"
        f"<td style='text-align:right'>{l[2]:,.2f}</td><td style='text-align:right'>{l[3]:,.2f}</td></tr>"
        for l in lignes_rows
    )
    totaux_html = "\n".join(
        f"<tr><td colspan='3' style='text-align:right'><b>{label}</b></td>"
        f"<td style='text-align:right'><b>{valeur:,.2f}</b></td></tr>"
        for label, valeur in totaux_rows
    )
    return f"""<!DOCTYPE html>
<html lang="fr"><head><meta charset="utf-8"><title>{titre} {numero}</title>
<style>
  body {{ font-family: Segoe UI, Arial, sans-serif; margin: 40px; color: #222; }}
  h1 {{ font-size: 22px; margin-bottom: 0; }}
  .meta {{ color: #595959; margin-bottom: 20px; }}
  .entete, .pied {{ white-space: pre-line; margin: 16px 0; }}
  table {{ width: 100%; border-collapse: collapse; margin-top: 16px; }}
  th, td {{ border: 1px solid #ccc; padding: 6px 10px; font-size: 14px; }}
  th {{ background: #1F4E78; color: white; text-align: left; }}
  @media print {{ button {{ display: none; }} }}
</style></head>
<body>
<button onclick="window.print()">Imprimer</button>
<h1>{titre} n° {numero}</h1>
<div class="meta">Date : {date_facture} — {tiers_label}</div>
<div class="entete">{entete or ""}</div>
<table>
<tr><th>Libellé</th><th>Quantité</th><th>Prix unitaire</th><th>Montant HT</th></tr>
{lignes_html}
{totaux_html}
</table>
<div class="pied">{pied_page or ""}</div>
</body></html>"""


def _html_facture_pro(conn, titre, numero, date_facture, tiers, lignes_rows, totaux_rows, montant_ttc,
                       entete_libre="", pied_libre=""):
    """Facture professionnelle « type imprimé commercial » : bloc entreprise,
    bloc « Doit : » (client/fournisseur), tableau Réf/Désignation/Qté/PU/
    Montant, bloc de totaux (HT / TVA / TTC), montant en toutes lettres, et
    zone de signature — reprend les informations déjà saisies dans
    ADMIN > Liasse fiscale (dénomination, adresse, IFU, RCCM...) plutôt que
    de les ressaisir.

    `tiers` : dict avec au moins 'nom' ; peut aussi contenir 'adresse',
    'telephone', 'code'.
    `lignes_rows` : liste de tuples (designation, quantite, prix_unitaire, montant_ht).
    `totaux_rows` : liste de tuples (label, valeur) affichés dans le bloc de droite,
    dans l'ordre voulu (ex. Montant H.T, TVA 18%, Montant TTC).
    """
    societe_nom = get_company_value(conn, "societe_nom") or "(Dénomination sociale non renseignée — ADMIN > Liasse fiscale)"
    societe_adresse = get_company_value(conn, "societe_adresse")
    societe_telephone = get_company_value(conn, "societe_telephone")
    societe_ifu = get_company_value(conn, "societe_ifu")
    societe_rccm = get_company_value(conn, "societe_rccm")

    lignes_html = "\n".join(
        f"<tr><td>{i+1}</td><td>{l[0]}</td><td style='text-align:right'>{l[1]:,.2f}</td>"
        f"<td style='text-align:right'>{l[2]:,.2f}</td><td style='text-align:right'>{l[3]:,.2f}</td></tr>"
        for i, l in enumerate(lignes_rows)
    )
    # Complète le tableau avec des lignes vides pour garder une hauteur homogène (style imprimé)
    lignes_vides = "\n".join(
        "<tr><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td><td>&nbsp;</td></tr>"
        for _ in range(max(0, 6 - len(lignes_rows)))
    )
    totaux_html = "\n".join(
        f"<tr><td>{label}</td><td style='text-align:right'>{valeur:,.2f}</td></tr>"
        for label, valeur in totaux_rows
    )
    montant_lettres = nombre_en_lettres_fr(montant_ttc)

    return f"""<!DOCTYPE html>
<html lang="fr"><head><meta charset="utf-8"><title>{titre} {numero}</title>
<style>
  body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 30px; color: #111; font-size: 13px; }}
  .toolbar {{ margin-bottom: 16px; }}
  .cadre-entreprise {{ border: 1px solid #000; padding: 8px 12px; text-align: center; margin-bottom: 10px; }}
  .cadre-entreprise .nom {{ font-weight: bold; font-size: 15px; text-transform: uppercase; }}
  .ligne-ident {{ border: 1px solid #000; border-top: none; padding: 4px 10px; font-size: 11px;
                  display: flex; justify-content: space-between; margin-bottom: 16px; }}
  .entete-facture {{ display: flex; justify-content: space-between; align-items: flex-start; margin-bottom: 14px; }}
  .titre-facture {{ font-weight: bold; font-size: 16px; }}
  .meta-facture div {{ margin: 2px 0; }}
  .cadre-doit {{ border: 1px solid #000; padding: 8px 12px; width: 320px; }}
  .cadre-doit .titre {{ font-weight: bold; text-decoration: underline; margin-bottom: 4px; }}
  table.lignes {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
  table.lignes th, table.lignes td {{ border: 1px solid #000; padding: 5px 8px; font-size: 12px; }}
  table.lignes th {{ background: #eee; text-align: left; }}
  .zone-bas {{ display: flex; justify-content: space-between; margin-top: 4px; }}
  .zone-lettres {{ max-width: 55%; }}
  table.totaux {{ border-collapse: collapse; }}
  table.totaux td {{ border: 1px solid #000; padding: 4px 10px; font-size: 12px; }}
  table.totaux td:first-child {{ font-weight: bold; }}
  table.totaux td:last-child {{ text-align: right; min-width: 110px; }}
  .signature {{ margin-top: 50px; text-align: right; font-size: 11px; }}
  .signature .nom {{ font-weight: bold; }}
  @media print {{ .toolbar {{ display: none; }} }}
</style></head>
<body>
<div class="toolbar"><button onclick="window.print()">🖨️ Imprimer / Enregistrer en PDF</button>
  <span style="margin-left:10px;color:#595959;font-size:12px;">
    Ceci est l'aperçu avant impression — rien n'est encore imprimé. Vérifiez le document, puis cliquez sur
    « Imprimer » ci-dessus (ou Ctrl+P).</span></div>

<div class="cadre-entreprise">
  <div class="nom">{societe_nom}</div>
  <div>{societe_adresse}</div>
</div>
<div class="ligne-ident">
  <span>{"Tél : " + societe_telephone if societe_telephone else ""}</span>
  <span>{"IFU : " + societe_ifu if societe_ifu else ""}</span>
  <span>{"RCCM : " + societe_rccm if societe_rccm else ""}</span>
</div>

<div class="entete-facture">
  <div class="meta-facture">
    <div class="titre-facture">{titre} N° {numero}</div>
    <div>Date : {date_facture}</div>
  </div>
  <div class="cadre-doit">
    <div class="titre">Doit :</div>
    <div>{tiers.get('nom', '')}</div>
    <div>{tiers.get('adresse') or ''}</div>
    <div>{('Tél : ' + tiers['telephone']) if tiers.get('telephone') else ''}</div>
  </div>
</div>

<div class="entete">{entete_libre or ""}</div>

<table class="lignes">
<tr><th style="width:34px">Réf.</th><th>Désignation</th><th style="width:90px">Quantité</th>
<th style="width:110px">P.U. HT</th><th style="width:120px">Montant HT</th></tr>
{lignes_html}
{lignes_vides}
</table>

<div class="zone-bas">
  <div class="zone-lettres">
    <p><b>Arrêtée la présente facture à la somme de :</b><br>{montant_lettres} francs CFA.</p>
    <div class="pied">{pied_libre or ""}</div>
  </div>
  <table class="totaux">
{totaux_html}
  </table>
</div>

<div class="signature">
  <div class="nom">{societe_nom}</div>
  <div>{societe_adresse}</div>
</div>
</body></html>"""


def render_facture_vente_html(conn, facture_id):
    """Facture de vente en HTML — renvoie le contenu (str), sans écrire de
    fichier, pour un usage local (bureau) ET distant (client réseau, qui
    écrira le fichier lui-même sur le poste client)."""
    facture = get_facture_vente(conn, facture_id)
    if not facture:
        raise ValueError("Facture introuvable.")
    lignes = list_lignes_facture_vente(conn, facture_id)
    if not lignes:
        raise ValueError("Cette facture n'a aucune ligne — ajoutez au moins une ligne avant d'imprimer.")
    totals = compute_facture_totals(conn, facture_id)
    client = get_client(conn, facture["client_code"])
    tiers = {"nom": client["raison_sociale"] if client else facture["client_code"],
              "adresse": client["adresse"] if client else None,
              "telephone": client["telephone"] if client else None}
    lignes_rows = [(l["libelle"], l["quantite"], l["prix_unitaire"], l["montant_ht"]) for l in lignes]
    totaux_rows = [("Montant H.T", totals["total_ht"]),
                    (f"T.V.A {totals['tva_taux']:g}%", totals["tva_montant"]),
                    ("Montant TTC", totals["total_ttc"])]
    return _html_facture_pro(conn, "FACTURE", facture["numero"], to_display_date(facture["date_facture"]),
                              tiers, lignes_rows, totaux_rows, totals["total_ttc"], facture["entete"],
                              facture["pied_page"])


def export_facture_vente_html(conn, facture_id, path):
    """Génère la facture de vente en HTML imprimable (bouton « Imprimer »
    intégré, ou Ctrl+P depuis le navigateur) et l'écrit dans `path`."""
    html = render_facture_vente_html(conn, facture_id)
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    return path


def render_facture_achat_html(conn, facture_id):
    """Facture d'achat en HTML — renvoie le contenu (str), sans écrire de fichier."""
    facture = get_facture_achat(conn, facture_id)
    if not facture:
        raise ValueError("Facture introuvable.")
    lignes = list_lignes_facture_achat(conn, facture_id)
    if not lignes:
        raise ValueError("Cette facture n'a aucune ligne — ajoutez au moins une ligne avant d'imprimer.")
    totals = compute_facture_achat_totals(conn, facture_id)
    fournisseur = get_fournisseur(conn, facture["fournisseur_code"])
    tiers = {"nom": fournisseur["raison_sociale"] if fournisseur else facture["fournisseur_code"],
              "adresse": fournisseur["adresse"] if fournisseur else None,
              "telephone": fournisseur["telephone"] if fournisseur else None}
    lignes_rows = [(l["libelle"], l["quantite"], l["prix_unitaire"], l["montant_ht"]) for l in lignes]
    totaux_rows = [("Montant H.T", totals["total_ht"]),
                    (f"Retenue à la source ({totals['retenue_taux']:g}%)", -totals["retenue_montant"]),
                    ("Net à payer", totals["net_a_payer"])]
    return _html_facture_pro(conn, "FACTURE D'ACHAT", facture["numero"], to_display_date(facture["date_facture"]),
                              tiers, lignes_rows, totaux_rows, totals["net_a_payer"], facture["entete"],
                              facture["pied_page"])


def export_facture_achat_html(conn, facture_id, path):
    """Génère la facture d'achat en HTML imprimable (bouton « Imprimer »
    intégré, ou Ctrl+P depuis le navigateur) et l'écrit dans `path`."""
    html = render_facture_achat_html(conn, facture_id)
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    return path


def render_bon_commande_html(conn, facture_id):
    """Bon de commande en HTML — renvoie le contenu (str), sans écrire de fichier."""
    facture = get_facture_achat(conn, facture_id)
    if not facture:
        raise ValueError("Bon de commande introuvable.")
    lignes = list_lignes_facture_achat(conn, facture_id)
    if not lignes:
        raise ValueError("Ce bon de commande n'a aucune ligne — ajoutez au moins une ligne avant d'imprimer.")
    totals = compute_facture_achat_totals(conn, facture_id)
    fournisseur = get_fournisseur(conn, facture["fournisseur_code"])
    tiers = {"nom": fournisseur["raison_sociale"] if fournisseur else facture["fournisseur_code"],
              "adresse": fournisseur["adresse"] if fournisseur else None,
              "telephone": fournisseur["telephone"] if fournisseur else None}
    lignes_rows = [(l["libelle"], l["quantite"], l["prix_unitaire"], l["montant_ht"]) for l in lignes]
    totaux_rows = [("Montant H.T", totals["total_ht"]),
                    (f"Retenue à la source ({totals['retenue_taux']:g}%)", -totals["retenue_montant"]),
                    ("Net estimé", totals["net_a_payer"])]
    entete = facture["entete"] or get_text_setting(conn, "bon_commande_entete_defaut", "")
    pied = facture["pied_page"] or get_text_setting(conn, "bon_commande_pied_defaut", "")
    return _html_facture_pro(conn, "BON DE COMMANDE", facture["numero"], to_display_date(facture["date_facture"]),
                              tiers, lignes_rows, totaux_rows, totals["net_a_payer"], entete, pied)


def export_bon_commande_html(conn, facture_id, path):
    """Génère le BON DE COMMANDE (stade brouillon d'une facture d'achat,
    onglet Factures frs) en HTML imprimable. Reprend l'en-tête/pied de page
    propres à cette commande s'ils sont renseignés, sinon le modèle par
    défaut paramétrable dans ADMIN (« bon_commande_entete_defaut »/
    « bon_commande_pied_defaut »)."""
    html = render_bon_commande_html(conn, facture_id)
    with open(path, "w", encoding="utf-8") as f:
        f.write(html)
    return path


def compute_tft(conn, treso_ouverture=None, exercice=None):
    """treso_ouverture=None : dérivée automatiquement des soldes d'ouverture des
    comptes de trésorerie (521000/531000/570000/585000). Passez une valeur pour
    la forcer manuellement."""
    exercice = exercice or get_current_exercice(conn)
    date_from, date_to = f"{exercice}-01-01", f"{exercice}-12-31"
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    if treso_ouverture is None:
        treso_ouverture = _sum_accounts_cloture(
            [dict(b, solde_cloture=b["solde_ouverture"]) for b in balance], COMPTES_TRESORERIE)
    treso_debit, treso_credit = _sum_accounts(balance, COMPTES_TRESORERIE)
    variation_totale = treso_debit - treso_credit

    def flux(code):
        like_clause = " OR ".join("compte LIKE ?" for _ in COMPTES_TRESORERIE)
        like_params = [f"{p}%" for p in COMPTES_TRESORERIE]
        rows = conn.execute(
            f"SELECT COALESCE(SUM(debit),0) d, COALESCE(SUM(credit),0) c FROM entries "
            f"WHERE ({like_clause}) AND flux_code = ? AND date >= ? AND date <= ?",
            (*like_params, code, date_from, date_to),
        ).fetchone()
        return rows["d"] - rows["c"]

    exploitation = flux("FLUX-EXP")
    investissement = flux("FLUX-INV")
    financement = flux("FLUX-FIN")
    non_classes = variation_totale - (exploitation + investissement + financement)

    cloture = treso_ouverture + variation_totale
    return {
        "ouverture": treso_ouverture,
        "exploitation": exploitation,
        "investissement": investissement,
        "financement": financement,
        "non_classes": non_classes,
        "variation": variation_totale,
        "cloture": cloture,
    }


def compute_tft_officiel(conn, exercice=None):
    """TFT selon la disposition EXACTE du formulaire officiel SYSCOHADA
    (références ZA, FA à FE, ZB...), pour remplissage direct de la feuille
    TFT de la Liasse fiscale. Réutilise les mêmes briques que
    compute_tft_indirect / compute_situation_financiere — donc cohérent
    avec la Balance. Seule la partie confirmée visuellement (ZA, FA-FE) est
    actuellement mappée ; l'investissement et le financement restent à
    positionner une fois les numéros de ligne du modèle officiel confirmés."""
    exercice = exercice or get_current_exercice(conn)
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    cr = compute_liasse_resultat(conn, exercice=exercice)

    treso_ouverture = sum(b["solde_ouverture"] for b in balance if b["classe"] == "5")
    treso_cloture_reelle = sum(b["solde_cloture"] for b in balance if b["classe"] == "5")

    ebe = cr["XD"]
    revenus_financiers = cr["TK"]
    frais_financiers = -cr["RM"]
    fa_cafg = ebe + revenus_financiers + frais_financiers

    def _delta_prefixes(prefixes):
        ouv = sum(b["solde_ouverture"] for b in balance if any(b["code"].startswith(p) for p in prefixes))
        clo = sum(b["solde_cloture"] for b in balance if any(b["code"].startswith(p) for p in prefixes))
        return ouv, clo

    stock_ouv, stock_clo = _delta_prefixes(COMPTES_STOCK_PREFIXES)
    fc_variation_stocks = -(stock_clo - stock_ouv)

    racines_exploit = ["42", "43", "44", "45", "46"]
    creances_ouv = sum(b["solde_ouverture"] for b in balance if account_racine(b["code"]) == RACINE_CLIENTS)
    creances_clo = sum(b["solde_cloture"] for b in balance if account_racine(b["code"]) == RACINE_CLIENTS)
    for r in racines_exploit:
        creances_ouv += sum(b["solde_ouverture"] for b in balance
                             if account_racine(b["code"]) == r and b["solde_ouverture"] > 0)
        creances_clo += sum(b["solde_cloture"] for b in balance
                             if account_racine(b["code"]) == r and b["solde_cloture"] > 0)
    fd_variation_creances = -(creances_clo - creances_ouv)

    dettes_ouv = -sum(b["solde_ouverture"] for b in balance if account_racine(b["code"]) == RACINE_FOURNISSEURS)
    dettes_clo = -sum(b["solde_cloture"] for b in balance if account_racine(b["code"]) == RACINE_FOURNISSEURS)
    for r in racines_exploit:
        dettes_ouv += -sum(b["solde_ouverture"] for b in balance
                            if account_racine(b["code"]) == r and b["solde_ouverture"] < 0)
        dettes_clo += -sum(b["solde_cloture"] for b in balance
                            if account_racine(b["code"]) == r and b["solde_cloture"] < 0)
    fe_variation_passif = dettes_clo - dettes_ouv

    racines_hao = ["47", "48", "49"]
    hao_actif_ouv = sum(b["solde_ouverture"] for b in balance
                         if account_racine(b["code"]) in racines_hao and b["solde_ouverture"] > 0)
    hao_actif_clo = sum(b["solde_cloture"] for b in balance
                         if account_racine(b["code"]) in racines_hao and b["solde_cloture"] > 0)
    hao_passif_ouv = -sum(b["solde_ouverture"] for b in balance
                          if account_racine(b["code"]) in racines_hao and b["solde_ouverture"] < 0)
    hao_passif_clo = -sum(b["solde_cloture"] for b in balance
                          if account_racine(b["code"]) in racines_hao and b["solde_cloture"] < 0)
    fb_variation_hao = -((hao_actif_clo - hao_actif_ouv) - (hao_passif_clo - hao_passif_ouv))

    zb_flux_operationnel = fa_cafg + fb_variation_hao + fc_variation_stocks + fd_variation_creances + fe_variation_passif

    return {
        "ZA": treso_ouverture,
        "FA": fa_cafg, "FB": fb_variation_hao, "FC": fc_variation_stocks,
        "FD": fd_variation_creances, "FE": fe_variation_passif,
        "ZB": zb_flux_operationnel,
        "treso_cloture_reelle": treso_cloture_reelle,
    }


def compute_tft_indirect(conn, exercice=None):
    """TFT selon la méthode indirecte SYSCOHADA (avec CAFG), au même format
    que le modèle officiel : A) trésorerie d'ouverture, détermination de la
    CAFG, variations du BFR (stocks/créances/dettes circulantes), flux
    d'investissement (acquisitions/cessions d'immobilisations), flux de
    financement (capital, subventions, emprunts). Entièrement calculé à
    partir de compute_balance() et compute_liasse_resultat() — donc toujours
    cohérent avec la Balance et le Bilan. La ligne CONTRÔLE compare la
    trésorerie calculée à la trésorerie réelle de la Balance (classe 5) :
    tout écart signale un mouvement de trésorerie non correctement classé."""
    exercice = exercice or get_current_exercice(conn)
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    cr = compute_liasse_resultat(conn, exercice=exercice)

    # ---- Trésorerie (classe 5 entière, comme la Balance) ----
    treso_ouverture = sum(b["solde_ouverture"] for b in balance if b["classe"] == "5")
    treso_cloture_reelle = sum(b["solde_cloture"] for b in balance if b["classe"] == "5")

    # ---- CAFG (à partir des soldes déjà calculés pour le Compte de résultat) ----
    # CAF Exploitation = EBE + Produits de cessions courantes d'immobilisations (754)
    #                   - Valeurs comptables de cessions courantes (654)
    #                   + Transferts de charges d'exploitation (781)
    # puis CAFG = CAF Exploitation + Revenus financiers (77) - Frais financiers (67)
    # — décomposition tirée du rapport financier de référence de l'utilisateur.
    ebe = cr["XD"]  # Excédent brut d'exploitation = Valeur ajoutée - charges de personnel
    produits_cessions_courantes = _sum_range(balance, [(754000, 754999)], classe="7") * -1
    valeurs_comptables_cessions_courantes = _sum_range(balance, [(654000, 654999)], classe="6") * -1
    transferts_charges_exploitation = _sum_range(balance, [(781000, 781999)], classe="7") * -1
    caf_exploitation = ebe + produits_cessions_courantes - valeurs_comptables_cessions_courantes + transferts_charges_exploitation
    revenus_financiers = cr["TK"]      # produits financiers (771, 776)
    frais_financiers = -cr["RM"]       # charges financières (671, 676), en décaissement
    cafg = caf_exploitation + revenus_financiers + frais_financiers

    # ---- Variation du BFR (comparaison ouverture/clôture, cohérente avec le Bilan) ----
    def _delta_racines(prefixes):
        ouverture = sum(b["solde_ouverture"] for b in balance if any(b["code"].startswith(p) for p in prefixes))
        cloture = sum(b["solde_cloture"] for b in balance if any(b["code"].startswith(p) for p in prefixes))
        return ouverture, cloture

    # Stocks : classe 3 ENTIÈRE (pas seulement les 4 comptes maîtres suivis
    # dans l'onglet Stocks) — sinon un sous-compte de stock (33, 37, 38, 39...)
    # disparaît silencieusement de la variation de trésorerie calculée.
    stock_ouv = sum(b["solde_ouverture"] for b in balance if b["classe"] == "3")
    stock_clo = sum(b["solde_cloture"] for b in balance if b["classe"] == "3")
    variation_stocks = -(stock_clo - stock_ouv)  # une hausse de stock consomme de la trésorerie

    # Créances (racines 40 à 45, côté DÉBITEUR uniquement — un compte fournisseur
    # avec avance (409) compte ici, un compte client créditeur (avoir) n'y
    # compte pas, il rejoint les dettes circulantes) — même principe de
    # classement par compte que le Bilan, décomposé ici comme dans le
    # rapport de référence : Créances (40-45 débit) séparées de l'Actif
    # circulant HAO (46-49 débit), qui a sa propre ligne.
    racines_creances = [str(r) for r in range(40, 46)]
    racines_hao = ["46", "47", "48", "49"]

    def _delta_signe(racines, sign):
        ouverture = 0.0
        cloture = 0.0
        for r in racines:
            ouverture += sum(b["solde_ouverture"] for b in balance
                              if account_racine(b["code"]) == r
                              and ((sign == "pos" and b["solde_ouverture"] > 0)
                                   or (sign == "neg" and b["solde_ouverture"] < 0)))
            cloture += sum(b["solde_cloture"] for b in balance
                            if account_racine(b["code"]) == r
                            and ((sign == "pos" and b["solde_cloture"] > 0)
                                 or (sign == "neg" and b["solde_cloture"] < 0)))
        return ouverture, cloture

    creances_ouv, creances_clo = _delta_signe(racines_creances, "pos")
    variation_creances = -(creances_clo - creances_ouv)  # une hausse de créances consomme de la trésorerie

    hao_ouv, hao_clo = _delta_signe(racines_hao, "pos")
    variation_actif_circulant_hao = -(hao_clo - hao_ouv)  # une hausse de créances HAO consomme de la trésorerie

    # Dettes circulantes : racines 40 à 49, côté CRÉDITEUR (toutes, y compris
    # 46-49 — contrairement aux créances/HAO, le crédit n'est pas séparé).
    racines_dettes = [str(r) for r in range(40, 50)]
    dettes_ouv_neg, dettes_clo_neg = _delta_signe(racines_dettes, "neg")
    dettes_ouv, dettes_clo = -dettes_ouv_neg, -dettes_clo_neg
    variation_dettes_circulantes = dettes_clo - dettes_ouv  # une hausse de dettes fournit de la trésorerie

    flux_operationnel = (cafg + variation_stocks + variation_creances + variation_actif_circulant_hao
                          + variation_dettes_circulantes)

    # ---- Flux d'investissement (acquisitions = débit de l'exercice sur les comptes d'immobilisations) ----
    def _debit_classe(prefixes):
        d, c = _sum_accounts(balance, prefixes)
        return d, c

    incorp_debit, incorp_credit = _debit_classe(["20", "21"])
    corp_debit, corp_credit = _debit_classe(["22", "23", "24", "25"])  # +25 : avances/acomptes versés sur immo
    fin_debit, fin_credit = _debit_classe(["26", "27"])
    acquisitions_incorp = -incorp_debit
    acquisitions_corp = -corp_debit
    acquisitions_fin = -fin_debit
    cessions_incorp = incorp_credit  # rare (compte 21 crédité lors d'une cession/sortie)
    cessions_corp = corp_credit
    cessions_fin = fin_credit
    flux_investissement = (acquisitions_incorp + acquisitions_corp + acquisitions_fin
                            + cessions_incorp + cessions_corp + cessions_fin)

    # ---- Flux de financement ----
    # Capital : racine 10 ENTIÈRE (pas seulement 101/104/105) — comme pour le
    # Bilan, une liste de comptes partielle ferait disparaître silencieusement
    # tout sous-compte de capital hors de cette liste.
    capital_debit, capital_credit = _debit_classe(["10"])
    augmentation_capital = capital_credit
    prelevements_capital = -capital_debit
    subv_debit, subv_credit = _debit_classe(["14"])
    subventions_recues = subv_credit
    # Dividendes distribués durant l'exercice (compte 465, Associés — dividendes
    # à payer) — tiré du rapport de référence (=CtaCptSolde("465*")). Un
    # décaissement de dividendes DÉBITE ce compte (le réglant), donc son
    # solde de la période (débit-crédit) représente le montant décaissé.
    dividendes_verses = -_sum_range(balance, [(465000, 465999)], classe="4")
    flux_capitaux_propres = augmentation_capital + subventions_recues + prelevements_capital + dividendes_verses

    emprunts_debit, emprunts_credit = _debit_classe(["16", "17"])
    emprunts_nouveaux = emprunts_credit
    remboursements_emprunts = -emprunts_debit
    flux_capitaux_etrangers = emprunts_nouveaux + remboursements_emprunts

    flux_financement = flux_capitaux_propres + flux_capitaux_etrangers

    variation_treso_nette = flux_operationnel + flux_investissement + flux_financement
    treso_cloture_calculee = treso_ouverture + variation_treso_nette
    ecart = treso_cloture_calculee - treso_cloture_reelle

    return {
        "treso_ouverture": treso_ouverture,
        "ebe": ebe,
        "produits_cessions_courantes": produits_cessions_courantes,
        "valeurs_comptables_cessions_courantes": valeurs_comptables_cessions_courantes,
        "transferts_charges_exploitation": transferts_charges_exploitation,
        "caf_exploitation": caf_exploitation,
        "revenus_financiers": revenus_financiers, "frais_financiers": frais_financiers,
        "cafg": cafg,
        "variation_stocks": variation_stocks, "variation_creances": variation_creances,
        "variation_actif_circulant_hao": variation_actif_circulant_hao,
        "variation_dettes_circulantes": variation_dettes_circulantes,
        "flux_operationnel": flux_operationnel,
        "acquisitions_incorp": acquisitions_incorp, "acquisitions_corp": acquisitions_corp,
        "acquisitions_fin": acquisitions_fin,
        "cessions_incorp": cessions_incorp, "cessions_corp": cessions_corp, "cessions_fin": cessions_fin,
        "flux_investissement": flux_investissement,
        "augmentation_capital": augmentation_capital, "subventions_recues": subventions_recues,
        "prelevements_capital": prelevements_capital, "dividendes_verses": dividendes_verses,
        "flux_capitaux_propres": flux_capitaux_propres,
        "emprunts_nouveaux": emprunts_nouveaux, "remboursements_emprunts": remboursements_emprunts,
        "flux_capitaux_etrangers": flux_capitaux_etrangers,
        "flux_financement": flux_financement,
        "variation_treso_nette": variation_treso_nette,
        "treso_cloture_calculee": treso_cloture_calculee,
        "treso_cloture_reelle": treso_cloture_reelle,
        "ecart": ecart,
    }


def compute_situation_financiere(conn, exercice=None):
    """Situation financière (FR - BFR - TN), présentée selon le modèle
    officiel : capacité d'autofinancement, ratios de rentabilité, puis
    analyse Fonds de Roulement / Besoin en Fonds de Roulement / Trésorerie
    Nette. Entièrement recalculé à partir de compute_bilan(),
    compute_liasse_resultat() et compute_tft_indirect() — donc toujours
    cohérent avec la Balance, le Bilan et le TFT.

    Ressources stables, Fonds de roulement et Besoin en fonds de roulement
    sont calculés à partir de sommes EXHAUSTIVES par classe/racine (jamais
    de liste de comptes partielle) — voir compute_bilan() et le principe
    déjà appliqué pour la Balance/le Bilan : un compte de classe 1 (ou de
    racine 40-49) hors des catégories usuelles ne doit jamais disparaître
    silencieusement du calcul."""
    exercice = exercice or get_current_exercice(conn)
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    b = compute_bilan(conn, exercice=exercice)
    cr = compute_liasse_resultat(conn, exercice=exercice)
    tft = compute_tft_indirect(conn, exercice=exercice)

    resultat_net = cr["XI"]
    resultat_exploitation = cr["XE"]
    cafg = tft["cafg"]
    dividendes_verses = tft["dividendes_verses"]
    autofinancement = cafg + dividendes_verses

    # ---- Ressources stables : classe 1 ENTIÈRE + résultat net (exhaustif,
    # garanti complet — comme le Bilan). Détail Capitaux propres (racines
    # 10-15) / Dettes financières (racines 16-17) / Autres (18-19 et tout
    # reliquat) affiché à titre indicatif, mais qui somme TOUJOURS
    # exactement au total (comme pour le Bilan détaillé).
    ressources_durables_total = -_sum_class(balance, "1")
    capitaux_propres_10_15 = -sum(b2["solde_cloture"] for b2 in balance
                                   if b2["classe"] == "1" and b2["code"][:2] in ("10", "11", "12", "13", "14", "15"))
    dettes_financieres_16_17 = -sum(b2["solde_cloture"] for b2 in balance
                                     if b2["classe"] == "1" and b2["code"][:2] in ("16", "17"))
    autres_ressources_18_19 = ressources_durables_total - capitaux_propres_10_15 - dettes_financieres_16_17

    capitaux_propres_ressources = capitaux_propres_10_15 + resultat_net + autres_ressources_18_19
    dettes_financieres = dettes_financieres_16_17
    ressources_stables = ressources_durables_total + resultat_net
    actifs_immobilises = b["actif"]["Immobilisations nettes"]
    fonds_de_roulement = ressources_stables - actifs_immobilises

    # ---- BFR exploitation : classe 3 (stocks) + racines 40 à 46, CHAQUE
    # COMPTE selon le signe de son propre solde — même méthode que le Bilan
    # (plus de racine 40/41 traitée « en bloc »), conforme au rapport de
    # référence (CtaCptSoldeDébit/Crédit("3*","46*")).
    def _somme_racine(racine, sign=None):
        total = 0.0
        for x in balance:
            if account_racine(x["code"]) != racine:
                continue
            v = x["solde_cloture"]
            if sign == "pos" and v <= 0:
                continue
            if sign == "neg" and v >= 0:
                continue
            total += v
        return total

    racines_exploit = [str(r) for r in range(40, 47)]
    creances_exploit = sum(_somme_racine(r, sign="pos") for r in racines_exploit)
    actif_circulant_exploitation = b["actif"]["Stocks"] + creances_exploit

    dettes_exploit = sum(-_somme_racine(r, sign="neg") for r in racines_exploit)
    passif_circulant_exploitation = dettes_exploit

    besoin_financement_exploitation = actif_circulant_exploitation - passif_circulant_exploitation

    racines_hao = ["47", "48", "49"]
    actif_circulant_hao = sum(_somme_racine(r, sign="pos") for r in racines_hao)
    passif_circulant_hao = sum(-_somme_racine(r, sign="neg") for r in racines_hao)
    besoin_financement_hao = actif_circulant_hao - passif_circulant_hao

    besoin_financement_global = besoin_financement_exploitation + besoin_financement_hao
    tresorerie_nette = fonds_de_roulement - besoin_financement_global

    treso_actif = b["actif"]["Trésorerie actif"]
    treso_passif = b["passif"]["Trésorerie passif"]
    treso_reelle = treso_actif - treso_passif
    controle_ecart = tresorerie_nette - treso_reelle

    rentabilite_economique = (resultat_exploitation / ressources_stables * 100
                               ) if ressources_stables else 0.0
    rentabilite_financiere = (resultat_net / ressources_stables * 100
                               ) if ressources_stables else 0.0

    endettement_financier_brut = dettes_financieres + treso_passif
    endettement_financier_net = endettement_financier_brut - treso_actif

    return {
        "resultat_net_comptable": resultat_net,
        "valeurs_comptables_cessions_courantes": tft["valeurs_comptables_cessions_courantes"],
        "produits_cessions_courantes": tft["produits_cessions_courantes"],
        "transferts_charges_exploitation": tft["transferts_charges_exploitation"],
        "caf_exploitation": tft["caf_exploitation"],
        "ebe": tft["ebe"], "revenus_financiers": tft["revenus_financiers"],
        "frais_financiers": tft["frais_financiers"], "cafg": cafg,
        "dividendes_verses": dividendes_verses, "autofinancement": autofinancement,
        "rentabilite_economique": rentabilite_economique, "rentabilite_financiere": rentabilite_financiere,
        "capitaux_propres_ressources": capitaux_propres_ressources,
        "dettes_financieres": dettes_financieres, "ressources_stables": ressources_stables,
        "actifs_immobilises": actifs_immobilises, "fonds_de_roulement": fonds_de_roulement,
        "actif_circulant_exploitation": actif_circulant_exploitation,
        "passif_circulant_exploitation": passif_circulant_exploitation,
        "besoin_financement_exploitation": besoin_financement_exploitation,
        "actif_circulant_hao": actif_circulant_hao, "passif_circulant_hao": passif_circulant_hao,
        "besoin_financement_hao": besoin_financement_hao,
        "besoin_financement_global": besoin_financement_global,
        "tresorerie_nette": tresorerie_nette, "controle_treso_reelle": treso_reelle,
        "controle_ecart": controle_ecart,
        "flux_operationnel": tft["flux_operationnel"], "flux_investissement": tft["flux_investissement"],
        "flux_financement": tft["flux_financement"], "variation_treso_nette": tft["variation_treso_nette"],
        "endettement_financier_brut": endettement_financier_brut,
        "treso_actif": treso_actif, "endettement_financier_net": endettement_financier_net,
    }


# ---------------------------------------------------------------------------
# Export de la liasse fiscale (.xlsx), mise en page SYSCOHADA système normal
# ---------------------------------------------------------------------------
COMPANY_FIELDS = {
    "societe_nom": "Dénomination sociale",
    "societe_sigle": "Sigle usuel",
    "societe_adresse": "Adresse",
    "societe_telephone": "Téléphone",
    "societe_ifu": "N° IFU du contribuable",
    "societe_rccm": "N° RCCM",
    "societe_teledeclarant": "N° de télédéclarant (NES)",
    "exercice_clos_le": "Exercice clos le (JJ/MM/AAAA)",
}


def get_company_info(conn):
    return {k: conn.execute("SELECT value FROM settings WHERE key = ?", (k,)).fetchone()
            for k in COMPANY_FIELDS}


def get_company_value(conn, key, default=""):
    row = conn.execute("SELECT value FROM settings WHERE key = ?", (key,)).fetchone()
    return row["value"] if row else default


def set_company_value(conn, key, value):
    conn.execute("INSERT OR REPLACE INTO settings (key, value) VALUES (?, ?)", (key, value))
    conn.commit()


def export_liasse_fiscale(conn, path, stock_initial=0.0, treso_ouverture=0.0):
    """Génère un classeur .xlsx : COUVERTURE, BILAN, RESULTAT, TFT
    (mise en page et codes SYSCOHADA système normal)."""
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = openpyxl.Workbook()

    bold = Font(bold=True)
    title_font = Font(bold=True, size=13)
    header_fill = PatternFill("solid", fgColor="D9D9D9")
    thin = Side(style="thin", color="999999")
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    money_fmt = "#,##0"

    def company_row(ws, row=3):
        info = {k: get_company_value(conn, k) for k in COMPANY_FIELDS}
        ws.cell(row=row, column=1, value="Dénomination sociale :")
        ws.cell(row=row, column=3, value=info["societe_nom"])
        ws.cell(row=row + 1, column=1, value="Adresse :")
        ws.cell(row=row + 1, column=3, value=info["societe_adresse"])
        ws.cell(row=row + 2, column=1, value="N° IFU du contribuable :")
        ws.cell(row=row + 2, column=3, value=info["societe_ifu"])
        ws.cell(row=row + 2, column=6, value="Exercice clos le :")
        ws.cell(row=row + 2, column=7, value=info["exercice_clos_le"])
        ws.cell(row=row + 3, column=1, value="N° de télédéclarant (NES) :")
        ws.cell(row=row + 3, column=3, value=info["societe_teledeclarant"])
        for r in range(row, row + 4):
            ws.cell(row=r, column=1).font = bold

    # ---- COUVERTURE ----
    ws = wb.active
    ws.title = "COUVERTURE"
    ws["A1"] = "ÉTATS FINANCIERS — SYSTÈME COMPTABLE OHADA (SYSCOHADA), SYSTÈME NORMAL"
    ws["A1"].font = title_font
    company_row(ws, row=3)
    ws["A9"] = ("Généré automatiquement par l'application Saisie Comptable. Les totaux (AZ, BK, BT, BZ, "
                "CP, DD, DP, DT, DZ) sont calculés directement depuis vos écritures. Le détail par ligne "
                "(AE à AN, CA à CM, DA à DM) est une répartition indicative par plage de comptes — à faire "
                "vérifier par un expert-comptable avant tout dépôt officiel auprès de la DGI.")
    ws["A9"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A9:H9")
    ws.row_dimensions[9].height = 60
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["C"].width = 25

    # ---- BILAN ----
    liasse = compute_liasse_bilan(conn, stock_initial=stock_initial)
    bt = liasse["totaux"]
    ad_net = bt["actif"]["Immobilisations nettes"]
    stocks_net = bt["actif"]["Stocks"]
    creances_net = bt["actif"]["Créances et emplois assimilés"]
    treso_actif_net = bt["actif"]["Trésorerie actif"]
    total_actif = bt["total_actif"]

    ws = wb.create_sheet("BILAN")
    ws["A1"] = "BILAN — SYSTÈME NORMAL"
    ws["A1"].font = title_font
    company_row(ws, row=3)
    headers_row = 8
    ws.cell(row=headers_row, column=1, value="REF").font = bold
    ws.cell(row=headers_row, column=2, value="ACTIF").font = bold
    ws.cell(row=headers_row, column=3, value="BRUT").font = bold
    ws.cell(row=headers_row, column=4, value="AMORT/DEPREC").font = bold
    ws.cell(row=headers_row, column=5, value="NET").font = bold
    ws.cell(row=headers_row, column=7, value="REF").font = bold
    ws.cell(row=headers_row, column=8, value="PASSIF").font = bold
    ws.cell(row=headers_row, column=9, value="NET").font = bold
    for c in range(1, 10):
        ws.cell(row=headers_row, column=c).fill = header_fill

    ad = liasse["actif_detail"]
    actif_lines = [
        ("AE", "Frais de développement et de prospection", ad["AE"]),
        ("AF", "Brevets, licences, logiciels et droits similaires", ad["AF"]),
        ("AG", "Fonds commercial et droit au bail", ad["AG"]),
        ("AH", "Autres immobilisations incorporelles", ad["AH"]),
        ("AJ", "Terrains", ad["AJ"]),
        ("AK", "Bâtiments", ad["AK"]),
        ("AL", "Aménagements, agencements et installations", ad["AL"]),
        ("AM", "Matériel, mobilier et actifs biologiques", ad["AM"]),
        ("AN", "Matériel de transport", ad["AN"]),
        ("AP", "Avances et acomptes versés sur immobilisations", ad["AP"]),
        ("AR", "Titres de participation", ad["AR"]),
        ("AS", "Autres immobilisations financières", ad["AS"]),
    ]
    ac = liasse["actif_circulant_detail"]
    actif_circ_lines = [
        ("BH", "Fournisseurs, avances versées", ac["BH"]),
        ("BI", "Clients", ac["BI"]),
    ]

    pd_ = liasse["passif_detail"]
    passif_lines = [
        ("CA", "Capital", pd_["CA"]),
        ("CD", "Primes liées au capital social", pd_["CD"]),
        ("CF_CG", "Réserves", pd_["CF_CG"]),
        ("CH", "Report à nouveau (+ ou -)", pd_["CH"]),
        ("CJ", "Résultat net de l'exercice", bt["passif"]["Résultat net de l'exercice"]),
        ("CL", "Subventions d'investissement", pd_["CL"]),
        ("CM", "Provisions réglementées", pd_["CM"]),
        ("CP", "TOTAL CAPITAUX PROPRES ET RESSOURCES ASSIMILEES", None),
        ("DA", "Emprunts et dettes financières diverses", pd_["DA"]),
        ("DB", "Dettes de location-acquisition", pd_["DB"]),
        ("DC", "Provisions pour risques et charges", pd_["DC"]),
        ("DD", "TOTAL DETTES FINANCIERES ET RESSOURCES ASSIMILEES", None),
        ("DJ", "Fournisseurs d'exploitation", pd_["DJ"]),
        ("DH", "Clients, avances reçues / Fournisseurs avances (détail)", pd_["DH_avances"]),
        ("DK", "Dettes fiscales et sociales", pd_["DK"]),
        ("DM", "Autres dettes", pd_["DM"]),
    ]

    r = headers_row + 1
    for ref, label, val in actif_lines:
        ws.cell(row=r, column=1, value=ref)
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=5, value=round(val.get("net", val) if isinstance(val, dict) else val))
        ws.cell(row=r, column=5).number_format = money_fmt
        r += 1
    ws.cell(row=r, column=1, value="AZ")
    ws.cell(row=r, column=2, value="TOTAL ACTIF IMMOBILISE").font = bold
    ws.cell(row=r, column=5, value=round(ad_net)).font = bold
    ws.cell(row=r, column=5).number_format = money_fmt
    r += 2
    ws.cell(row=r, column=1, value="BB")
    ws.cell(row=r, column=2, value="STOCKS ET ENCOURS")
    ws.cell(row=r, column=5, value=round(stocks_net)).number_format = money_fmt
    r += 1
    for ref, label, val in actif_circ_lines:
        ws.cell(row=r, column=1, value=ref)
        ws.cell(row=r, column=2, value=label)
        ws.cell(row=r, column=5, value=round(val)).number_format = money_fmt
        r += 1
    ws.cell(row=r, column=1, value="BK")
    ws.cell(row=r, column=2, value="TOTAL ACTIF CIRCULANT").font = bold
    ws.cell(row=r, column=5, value=round(stocks_net + creances_net)).font = bold
    ws.cell(row=r, column=5).number_format = money_fmt
    r += 2
    ws.cell(row=r, column=1, value="BT")
    ws.cell(row=r, column=2, value="TOTAL TRESORERIE-ACTIF").font = bold
    ws.cell(row=r, column=5, value=round(treso_actif_net)).font = bold
    ws.cell(row=r, column=5).number_format = money_fmt
    r += 2
    ws.cell(row=r, column=1, value="BZ")
    ws.cell(row=r, column=2, value="TOTAL GENERAL ACTIF").font = bold
    ws.cell(row=r, column=5, value=round(total_actif)).font = bold
    ws.cell(row=r, column=5).number_format = money_fmt
    last_actif_row = r

    r2 = headers_row + 1
    for ref, label, val in passif_lines:
        ws.cell(row=r2, column=7, value=ref)
        ws.cell(row=r2, column=8, value=label)
        if val is not None:
            ws.cell(row=r2, column=9, value=round(val)).number_format = money_fmt
        else:
            ws.cell(row=r2, column=8).font = bold
        r2 += 1
    total_passif = bt["total_passif"]
    ws.cell(row=r2, column=7, value="DZ")
    ws.cell(row=r2, column=8, value="TOTAL GENERAL PASSIF").font = bold
    ws.cell(row=r2, column=9, value=round(total_passif)).font = bold
    ws.cell(row=r2, column=9).number_format = money_fmt
    r2 += 2
    ws.cell(row=r2, column=7, value="Écart Actif - Passif :")
    ws.cell(row=r2, column=9, value=round(total_actif - total_passif)).number_format = money_fmt

    for col, w in zip("ABCDEFGHI", [6, 40, 14, 14, 14, 3, 6, 40, 16]):
        ws.column_dimensions[col].width = w

    # ---- RESULTAT ----
    cr = compute_liasse_resultat(conn)
    ws = wb.create_sheet("RESULTAT")
    ws["A1"] = "COMPTE DE RÉSULTAT — SYSTÈME NORMAL"
    ws["A1"].font = title_font
    company_row(ws, row=3)
    headers_row = 8
    for c, h in zip((1, 2, 5), ("REF", "LIBELLES", "EXERCICE N")):
        ws.cell(row=headers_row, column=c, value=h).font = bold
        ws.cell(row=headers_row, column=c).fill = header_fill

    resultat_lines = [
        ("TA", "Ventes de marchandises", cr["TA"]),
        ("RA", "Achats de marchandises", -cr["RA"]),
        ("XA", "MARGE COMMERCIALE", cr["XA"]),
        ("TB", "Ventes de produits fabriqués", cr["TB"]),
        ("TC", "Travaux, services vendus", cr["TC"]),
        ("TD", "Produits accessoires", cr["TD"]),
        ("XB", "CHIFFRE D'AFFAIRES", cr["XB"]),
        ("TE", "Production stockée (ou déstockage)", cr["TE"]),
        ("TG", "Subventions d'exploitation", cr["TG"]),
        ("TH", "Autres produits", cr["TH"]),
        ("RC", "Achats de matières premières et fournitures liées", -cr["RC"]),
        ("RE", "Autres achats", -cr["RE"]),
        ("RG", "Transports", -cr["RG"]),
        ("RH", "Services extérieurs", -cr["RH"]),
        ("RI", "Impôts et taxes", -cr["RI"]),
        ("RJ", "Autres charges", -cr["RJ"]),
        ("XC", "VALEUR AJOUTEE", cr["XC"]),
        ("RK", "Charges de personnel", -cr["RK"]),
        ("XD", "EXCEDENT BRUT D'EXPLOITATION", cr["XD"]),
        ("RL", "Dotations aux amortissements, provisions et dépréciations", -cr["RL"]),
        ("XE", "RESULTAT D'EXPLOITATION", cr["XE"]),
        ("TK", "Revenus financiers et assimilés", cr["TK"]),
        ("RM", "Frais financiers et charges assimilées", -cr["RM"]),
        ("XF", "RESULTAT FINANCIER", cr["XF"]),
        ("XG", "RESULTAT DES ACTIVITES ORDINAIRES", cr["XG"]),
        ("XH", "RESULTAT HORS ACTIVITES ORDINAIRES (non tracé)", cr["XH"]),
        ("RQ", "Participation des travailleurs (non tracée)", cr["RQ"]),
        ("RS", "Impôts sur le résultat (non tracé — IS à saisir séparément)", cr["RS"]),
        ("XI", "RESULTAT NET", cr["XI"]),
    ]
    bold_refs = {"XA", "XB", "XC", "XD", "XE", "XF", "XG", "XI"}
    r = headers_row + 1
    for ref, label, val in resultat_lines:
        ws.cell(row=r, column=1, value=ref)
        ws.cell(row=r, column=2, value=label)
        cell = ws.cell(row=r, column=5, value=round(val))
        cell.number_format = money_fmt
        if ref in bold_refs:
            ws.cell(row=r, column=2).font = bold
            cell.font = bold
        r += 1
    for col, w in zip("ABCDE", [6, 55, 3, 3, 16]):
        ws.column_dimensions[col].width = w

    # ---- TFT (simplifié, méthode directe) ----
    tft = compute_tft(conn, treso_ouverture=treso_ouverture)
    ws = wb.create_sheet("TFT")
    ws["A1"] = "TABLEAU DES FLUX DE TRÉSORERIE — méthode directe simplifiée"
    ws["A1"].font = title_font
    ws["A2"] = ("Cette version simplifiée (encaissements/décaissements de trésorerie classés EXP/INV/FIN) "
                "ne correspond PAS exactement au format officiel SYSCOHADA (méthode indirecte avec CAFG). "
                "Elle donne une image de la trésorerie mais doit être retravaillée avec un expert-comptable "
                "pour un dépôt officiel.")
    ws["A2"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 45
    company_row(ws, row=4)
    tft_lines = [
        ("Trésorerie d'ouverture", tft["ouverture"]),
        ("Flux liés aux activités opérationnelles (EXP)", tft["exploitation"]),
        ("Flux liés aux activités d'investissement (INV)", tft["investissement"]),
        ("Flux liés aux activités de financement (FIN)", tft["financement"]),
        ("Flux non classés (à coder)", tft["non_classes"]),
        ("VARIATION NETTE DE TRESORERIE", tft["variation"]),
        ("TRESORERIE DE CLOTURE", tft["cloture"]),
    ]
    r = 10
    for label, val in tft_lines:
        ws.cell(row=r, column=1, value=label)
        cell = ws.cell(row=r, column=3, value=round(val))
        cell.number_format = money_fmt
        r += 1
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["C"].width = 16

    wb.save(path)
    return path


def export_liasse_fiscale_complete(conn, path, stock_initial=0.0):
    """Génère la liasse fiscale COMPLÈTE (mêmes 92 pages, mêmes dimensions que le
    modèle SYSCOHADA système normal fourni) : COUVERTURE/GARDE, BILAN et RESULTAT
    remplis automatiquement depuis vos écritures (soldes de clôture = solde
    d'ouverture + mouvements) ; TFT (officiel, vierge, + un onglet TFT simplifié
    calculé) ; toutes les autres pages (39 notes annexes, ~20 tableaux fiscaux DGI)
    sont conservées avec leur mise en page et leurs dimensions exactes, mais les
    montants qui provenaient du modèle sont effacés (ce ne sont pas vos chiffres)
    pour être complétées manuellement ou par votre expert-comptable."""
    import openpyxl
    from openpyxl.styles import Font

    template_path = os.path.join(_resource_dir(), "etats_financiers_template.xlsx")
    if not os.path.exists(template_path):
        raise FileNotFoundError(
            "Le fichier modèle 'etats_financiers_template.xlsx' est introuvable dans "
            "l'exécutable. Cela signifie qu'il n'a pas été inclus lors de la compilation : "
            "vérifiez que ce fichier est bien présent à la racine du dépôt GitHub (à côté de "
            "main.py) et que .github/workflows/build.yml contient bien la ligne "
            "--add-data \"etats_financiers_template.xlsx;.\", puis relancez le build."
        )
    wb = openpyxl.load_workbook(template_path)
    green = Font(color="FF008000")

    # ---- Supprime les liens externes cassés (source du bandeau Excel
    #      « Impossible d'actualiser... valeurs depuis un classeur lié ») ----
    if getattr(wb, "_external_links", None):
        wb._external_links = []

    # ---- GARDE : identification de l'entité ----
    if "GARDE" in wb.sheetnames:
        g = wb["GARDE"]
        g["D22"] = get_company_value(conn, "societe_nom")
        g["C26"] = get_company_value(conn, "societe_sigle")
        g["C28"] = get_company_value(conn, "societe_adresse")
        g["D30"] = get_company_value(conn, "societe_ifu")
        g["D31"] = get_company_value(conn, "societe_teledeclarant")
        exdate = get_company_value(conn, "exercice_clos_le")
        if exdate:
            parsed = None
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d.%m.%Y"):
                try:
                    parsed = datetime.strptime(exdate.strip(), fmt)
                    break
                except ValueError:
                    continue
            if parsed:
                g["E17"] = parsed
                g["E17"].number_format = "DD/MM/YYYY"
            else:
                g["E17"] = exdate

    # ---- BILAN ----
    liasse = compute_liasse_bilan(conn, stock_initial=stock_initial)
    bt = liasse["totaux"]
    ad = liasse["actif_detail"]
    ac = liasse["actif_circulant_detail"]
    pd_ = liasse["passif_detail"]

    actif_values = {
        "AE": ad["AE"]["net"], "AF": ad["AF"]["net"], "AG": ad["AG"]["net"], "AH": ad["AH"]["net"],
        "AD": ad["AE"]["net"] + ad["AF"]["net"] + ad["AG"]["net"] + ad["AH"]["net"],
        "AJ": ad["AJ"]["net"], "AK": ad["AK"]["net"], "AL": ad["AL"]["net"],
        "AM": ad["AM"]["net"], "AN": ad["AN"]["net"],
        "AI": ad["AJ"]["net"] + ad["AK"]["net"] + ad["AL"]["net"] + ad["AM"]["net"] + ad["AN"]["net"],
        "AP": ad["AP"]["net"], "AR": ad["AR"]["net"], "AS": ad["AS"]["net"],
        "AZ": bt["actif"]["Immobilisations nettes"],
        "BB": bt["actif"]["Stocks"],
        "BH": ac["BH"], "BI": ac["BI"],
        "BK": bt["actif"]["Stocks"] + bt["actif"]["Créances et emplois assimilés"],
        "BT": bt["actif"]["Trésorerie actif"],
        "BZ": bt["total_actif"],
    }
    passif_values = {
        "CA": pd_["CA"], "CD": pd_["CD"], "CF": pd_["CF_CG"], "CH": pd_["CH"],
        "CJ": bt["passif"]["Résultat net de l'exercice"],
        "CL": pd_["CL"], "CM": pd_["CM"],
        "CP": (pd_["CA"] + pd_["CD"] + pd_["CF_CG"] + pd_["CH"]
               + bt["passif"]["Résultat net de l'exercice"] + pd_["CL"] + pd_["CM"]),
        "DA": pd_["DA"], "DB": pd_["DB"], "DC": pd_["DC"],
        "DD": pd_["DA"] + pd_["DB"] + pd_["DC"],
        "DJ": pd_["DJ"], "DH": pd_["DH_avances"], "DK": pd_["DK"], "DM": pd_["DM"],
        "DP": pd_["DJ"] + pd_["DH_avances"] + pd_["DK"] + pd_["DM"],
        "DT": bt["passif"]["Trésorerie passif"],
        "DZ": bt["total_passif"],
    }

    if "BILAN" in wb.sheetnames:
        ws = wb["BILAN"]
        # Efface toutes les valeurs numériques préexistantes du modèle (Brut, Amort,
        # N-1) sur les lignes de données, pour n'y laisser QUE nos propres calculs.
        for row in range(11, 41):
            for col in (6, 7, 8, 9, 13, 14):  # F,G,H,I (actif) / M,N (passif)
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.value = None
        ws["C3"] = get_company_value(conn, "societe_nom")
        ws["C4"] = get_company_value(conn, "societe_adresse")
        ws["C5"] = get_company_value(conn, "societe_ifu")
        for ref, row in {
            "AD": 11, "AE": 12, "AF": 13, "AG": 14, "AH": 15, "AI": 16, "AJ": 17, "AK": 18,
            "AL": 19, "AM": 20, "AN": 21, "AP": 22, "AQ": 23, "AR": 24, "AS": 25, "AZ": 26,
            "BB": 28, "BH": 30, "BI": 31, "BK": 33, "BT": 37, "BZ": 39,
        }.items():
            if ref in actif_values:
                cell = ws.cell(row=row, column=8, value=round(actif_values[ref]))
                cell.font = green
        for ref, row in {
            "CA": 11, "CD": 13, "CF": 15, "CH": 17, "CJ": 18, "CL": 19, "CM": 20, "CP": 21,
            "DA": 22, "DB": 23, "DC": 24, "DD": 25, "DH": 27, "DJ": 29, "DK": 30, "DM": 31,
            "DP": 33, "DT": 36, "DZ": 39,
        }.items():
            if ref in passif_values:
                cell = ws.cell(row=row, column=13, value=round(passif_values[ref]))
                cell.font = green
        ws.cell(row=40, column=13, value="=H39-M39")  # écart de contrôle

    # ---- RESULTAT ----
    cr = compute_liasse_resultat(conn)
    if "RESULTAT" in wb.sheetnames:
        ws = wb["RESULTAT"]
        for row in range(11, 53):
            for col in (9, 10):  # I (exercice N), J (N-1)
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.value = None
        row_map = {"TA": 11, "RA": 12, "XA": 14, "TB": 15, "TC": 16, "TD": 17, "XB": 18,
                   "TE": 19, "TG": 21, "TH": 22, "RC": 24, "RE": 26, "RG": 28, "RH": 29,
                   "RI": 30, "RJ": 31, "XC": 32, "RK": 33, "XD": 34, "RL": 36, "XE": 37,
                   "TK": 38, "RM": 41, "XF": 43, "XG": 44, "XH": 49, "RQ": 50, "RS": 51, "XI": 52}
        sign_negative = {"RA", "RC", "RE", "RG", "RH", "RI", "RJ", "RK", "RL", "RM", "RQ", "RS"}
        for ref, row in row_map.items():
            val = cr.get(ref, 0.0)
            if ref in sign_negative:
                val = -abs(val)
            cell = ws.cell(row=row, column=9, value=round(val))
            cell.font = green

    # ---- TFT : remplit la vraie feuille officielle sur les lignes confirmées
    #      (ZA=10, FA=12, FB=13, FC=14, FD=15, FE=16), + un onglet
    #      supplémentaire avec le calcul complet (méthode indirecte — CAFG),
    #      les mêmes données que l'onglet TFT de l'application ----
    tft_off = compute_tft_officiel(conn)
    tft = compute_tft_indirect(conn)
    if "TFT" in wb.sheetnames:
        ws = wb["TFT"]
        for row in range(10, 42):
            for col in (9, 10):  # I (exercice N), J (N-1)
                cell = ws.cell(row=row, column=col)
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.value = None
        for ref, row in (("ZA", 10), ("FA", 12), ("FB", 13), ("FC", 14), ("FD", 15), ("FE", 16)):
            c = ws.cell(row=row, column=9, value=round(tft_off[ref]))
            c.font = green
        ws["A44"] = (
            "Lignes ZA et FA à FE remplies automatiquement depuis vos écritures (identique à "
            "l'onglet TFT de l'application, section Flux opérationnels). Les lignes d'investissement "
            "et de financement (FF et suivantes) n'ont pas encore de position de cellule confirmée "
            "dans ce modèle — complétez-les manuellement, ou envoyez une capture des lignes "
            "suivantes pour qu'elles soient automatisées aussi. Voir l'onglet « TFT (méthode "
            "indirecte - CAFG) » pour le calcul complet (investissement et financement inclus)."
        )
    ws_tft = wb.create_sheet("TFT (méthode indirecte - CAFG)")
    ws_tft["A1"] = "TABLEAU DE FLUX DE TRÉSORERIE (méthode indirecte — CAFG)"
    ws_tft["A1"].font = Font(bold=True, size=12)
    tft_lignes = [
        ("A - Trésorerie nette au 1er janvier", tft["treso_ouverture"]),
        ("", None),
        ("DÉTERMINATION DE LA CAFG", None),
        ("Excédent Brut d'Exploitation (EBE)", tft["ebe"]),
        ("+ Revenus financiers", tft["revenus_financiers"]),
        ("- Frais financiers", tft["frais_financiers"]),
        ("CAPACITÉ D'AUTOFINANCEMENT GLOBALE (CAFG)", tft["cafg"]),
        ("- Variation des stocks", tft["variation_stocks"]),
        ("- Variation des créances", tft["variation_creances"]),
        ("+ Variation du passif circulant", tft["variation_dettes_circulantes"]),
        ("FLUX DES ACTIVITÉS OPÉRATIONNELLES (A)", tft["flux_operationnel"]),
        ("", None),
        ("FLUX DES ACTIVITÉS D'INVESTISSEMENT", None),
        ("- Acquisitions immobilisations incorporelles", tft["acquisitions_incorp"]),
        ("- Acquisitions immobilisations corporelles", tft["acquisitions_corp"]),
        ("- Acquisitions immobilisations financières", tft["acquisitions_fin"]),
        ("+ Cessions immobilisations incorporelles", tft["cessions_incorp"]),
        ("+ Cessions immobilisations corporelles", tft["cessions_corp"]),
        ("+ Cessions immobilisations financières", tft["cessions_fin"]),
        ("FLUX DES ACTIVITÉS D'INVESTISSEMENT (B)", tft["flux_investissement"]),
        ("", None),
        ("FLUX DES ACTIVITÉS DE FINANCEMENT", None),
        ("+ Augmentation de capital", tft["augmentation_capital"]),
        ("+ Subventions d'investissement reçues", tft["subventions_recues"]),
        ("- Prélèvements sur le capital", tft["prelevements_capital"]),
        ("- Dividendes versés", tft["dividendes_verses"]),
        ("+ Emprunts nouveaux", tft["emprunts_nouveaux"]),
        ("- Remboursements des emprunts", tft["remboursements_emprunts"]),
        ("FLUX DES ACTIVITÉS DE FINANCEMENT (C)", tft["flux_financement"]),
        ("", None),
        ("VARIATION DE LA TRÉSORERIE NETTE (A+B+C)", tft["variation_treso_nette"]),
        ("TRÉSORERIE NETTE CALCULÉE AU 31/12/N", tft["treso_cloture_calculee"]),
        ("CONTRÔLE — Trésorerie réelle (Balance, classe 5)", tft["treso_cloture_reelle"]),
        ("ÉCART", tft["ecart"]),
    ]
    for i, (label, val) in enumerate(tft_lignes):
        ws_tft.cell(row=3 + i, column=1, value=label)
        if val is not None:
            ws_tft.cell(row=3 + i, column=3, value=round(val))
    ws_tft.column_dimensions["A"].width = 55

    # ---- SITUATION FINANCIÈRE (FR-BFR-TN) : mêmes données que l'onglet
    #      correspondant de l'application ----
    sf = compute_situation_financiere(conn)
    ws_sf = wb.create_sheet("SITUATION FIN. (FR-BFR-TN)")
    ws_sf["A1"] = "SITUATION FINANCIÈRE (FR - BFR - TN)"
    ws_sf["A1"].font = Font(bold=True, size=12)
    sf_lignes = [
        ("Résultat net comptable", sf["resultat_net_comptable"]),
        ("EBE", sf["ebe"]), ("+ Revenus financiers", sf["revenus_financiers"]),
        ("- Frais financiers", sf["frais_financiers"]),
        ("CAFG", sf["cafg"]), ("- Dividendes versés", sf["dividendes_verses"]),
        ("AUTOFINANCEMENT", sf["autofinancement"]),
        ("Rentabilité économique (%)", sf["rentabilite_economique"]),
        ("Rentabilité financière (%)", sf["rentabilite_financiere"]),
        ("", None),
        ("Capitaux propres et ressources assimilées", sf["capitaux_propres_ressources"]),
        ("+ Dettes financières", sf["dettes_financieres"]),
        ("= RESSOURCES STABLES", sf["ressources_stables"]),
        ("- Actifs immobilisés", sf["actifs_immobilises"]),
        ("= FONDS DE ROULEMENT (FR)", sf["fonds_de_roulement"]),
        ("", None),
        ("+ Actif circulant d'exploitation", sf["actif_circulant_exploitation"]),
        ("- Passif circulant d'exploitation", sf["passif_circulant_exploitation"]),
        ("= Besoin de financement d'exploitation", sf["besoin_financement_exploitation"]),
        ("+ Actif circulant HAO", sf["actif_circulant_hao"]),
        ("- Passif circulant HAO", sf["passif_circulant_hao"]),
        ("= Besoin de financement HAO", sf["besoin_financement_hao"]),
        ("= BESOIN DE FINANCEMENT GLOBAL (BFR)", sf["besoin_financement_global"]),
        ("", None),
        ("TRÉSORERIE NETTE (FR - BFR)", sf["tresorerie_nette"]),
        ("Contrôle — Trésorerie réelle (Balance)", sf["controle_treso_reelle"]),
        ("Écart", sf["controle_ecart"]),
        ("", None),
        ("+ Flux activités opérationnelles", sf["flux_operationnel"]),
        ("- Flux activités d'investissement", sf["flux_investissement"]),
        ("+ Flux activités de financement", sf["flux_financement"]),
        ("VARIATION DE LA TRÉSORERIE NETTE DE LA PÉRIODE", sf["variation_treso_nette"]),
        ("", None),
        ("Endettement financier brut", sf["endettement_financier_brut"]),
        ("- Trésorerie actif", sf["treso_actif"]),
        ("= ENDETTEMENT FINANCIER NET", sf["endettement_financier_net"]),
    ]
    for i, (label, val) in enumerate(sf_lignes):
        ws_sf.cell(row=3 + i, column=1, value=label)
        if val is not None:
            ws_sf.cell(row=3 + i, column=3, value=round(val, 2))
    ws_sf.column_dimensions["A"].width = 55

    # ---- Toutes les autres pages : structure/dimensions conservées, valeurs
    #      chiffrées (issues du modèle) effacées pour éviter toute confusion ----
    skip = {"GARDE", "BILAN", "RESULTAT", "TFT", "TFT (méthode indirecte - CAFG)",
            "SITUATION FIN. (FR-BFR-TN)"}
    for name in wb.sheetnames:
        if name in skip:
            continue
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.value = None

    # ---- Filet de sécurité : efface tout texte résiduel qui ne serait pas le
    #      nom de VOTRE entité (au cas où le modèle contiendrait encore une
    #      dénomination sociale tierce, pour éviter toute confusion/litige) ----
    my_name = (get_company_value(conn, "societe_nom") or "").strip().upper()
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and "GCM" in cell.value.upper():
                    if not my_name or my_name not in cell.value.upper():
                        cell.value = None

    # ---- NOTE 34 : Fiche de synthèse des principaux indicateurs financiers
    #      (SIG) — mêmes données que l'onglet Compte de résultat de l'app.
    #      Rempli APRÈS le nettoyage général ci-dessus (qui efface d'abord
    #      les anciennes valeurs littérales de l'entité précédente). ----
    if "NOTE 34" in wb.sheetnames:
        ws34 = wb["NOTE 34"]
        note34_map = {"XB": 11, "XA": 12, "XC": 13, "XD": 14, "XE": 15, "XF": 16,
                      "XG": 17, "XH": 18, "XI": 19}
        for ref, row in note34_map.items():
            c = ws34.cell(row=row, column=6, value=round(cr.get(ref, 0.0)))
            c.font = green
        try:
            exercice_n = get_current_exercice(conn)
            exercice_n1 = str(int(exercice_n) - 1)
            if any(e["exercice"] == exercice_n1 for e in list_exercices(conn)):
                cr_n1 = compute_liasse_resultat(conn, exercice=exercice_n1)
                for ref, row in note34_map.items():
                    ws34.cell(row=row, column=7, value=round(cr_n1.get(ref, 0.0)))
        except (ValueError, TypeError):
            pass

    # ---- Uniformise toutes les cellules de type date au format JJ/MM/AAAA ----
    date_format_markers = ("yy", "mm", "dd", "jj", "aaaa")
    for name in wb.sheetnames:
        ws = wb[name]
        for row in ws.iter_rows():
            for cell in row:
                is_date_value = isinstance(cell.value, (datetime, date)) and not isinstance(cell.value, bool)
                fmt = (cell.number_format or "").lower()
                looks_like_date_format = fmt not in ("general", "@") and any(m in fmt for m in date_format_markers)
                if is_date_value or looks_like_date_format:
                    cell.number_format = "DD/MM/YYYY"

    wb.properties.creator = None
    wb.properties.lastModifiedBy = None

    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# Circuit interne Expression de besoin → Bon de commande → Bordereau de
# livraison (menu ENGAGEMENTS-PROJETS) — AUCUN lien avec la comptabilité à
# aucune étape : ce sont des documents de suivi de procédure d'achat interne
# (workflow d'approbation), pas des pièces comptables. La validation de
# chaque étape fait simplement basculer le document dans l'étape suivante,
# en recopiant ses lignes.
# ---------------------------------------------------------------------------

# ---- Expression de besoin ----
def create_expression_besoin(conn, numero, date_demande, demandeur="", service="", entete="", pied_page=""):
    cur = conn.execute(
        """INSERT INTO expressions_besoin (numero, date_demande, demandeur, service, entete, pied_page, statut)
           VALUES (?, ?, ?, ?, ?, ?, 'brouillon')""",
        (numero, date_demande, demandeur, service, entete, pied_page),
    )
    conn.commit()
    return cur.lastrowid


def update_expression_besoin(conn, expression_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE expressions_besoin SET {cols} WHERE id = ?", (*fields.values(), expression_id))
    conn.commit()


def delete_expression_besoin(conn, expression_id):
    conn.execute("DELETE FROM expression_besoin_lignes WHERE expression_id = ?", (expression_id,))
    conn.execute("DELETE FROM expressions_besoin WHERE id = ?", (expression_id,))
    conn.commit()


def get_expression_besoin(conn, expression_id):
    row = conn.execute("SELECT * FROM expressions_besoin WHERE id = ?", (expression_id,)).fetchone()
    return dict(row) if row else None


def list_expressions_besoin(conn):
    rows = conn.execute("SELECT * FROM expressions_besoin ORDER BY date_demande DESC, id DESC").fetchall()
    return [dict(r) for r in rows]


def add_ligne_expression_besoin(conn, expression_id, libelle, quantite, unite=None):
    conn.execute(
        "INSERT INTO expression_besoin_lignes (expression_id, libelle, quantite, unite) VALUES (?, ?, ?, ?)",
        (expression_id, libelle, quantite or 0, unite or None),
    )
    conn.commit()


def delete_ligne_expression_besoin(conn, ligne_id):
    conn.execute("DELETE FROM expression_besoin_lignes WHERE id = ?", (ligne_id,))
    conn.commit()


def list_lignes_expression_besoin(conn, expression_id):
    rows = conn.execute(
        "SELECT * FROM expression_besoin_lignes WHERE expression_id = ? ORDER BY id", (expression_id,)
    ).fetchall()
    return [dict(r) for r in rows]


def valider_expression_besoin(conn, expression_id):
    """Fait basculer l'Expression de besoin en Bon de commande (nouveau
    document, même numéro, lignes recopiées) — AUCUNE écriture comptable.
    L'expression d'origine passe en statut « validee » (verrouillée).
    Retourne l'ID du nouveau bon de commande."""
    exp = get_expression_besoin(conn, expression_id)
    if not exp:
        raise ValueError("Expression de besoin introuvable.")
    if exp["statut"] == "validee":
        raise ValueError("Cette expression de besoin est déjà validée.")
    lignes = list_lignes_expression_besoin(conn, expression_id)
    if not lignes:
        raise ValueError("Ajoutez au moins une ligne avant de valider.")
    bon_id = create_ep_bon_commande(conn, exp["numero"], exp["date_demande"], expression_id=expression_id,
                                     entete=exp["entete"], pied_page=exp["pied_page"])
    for l in lignes:
        add_ligne_ep_bon_commande(conn, bon_id, l["libelle"], l["quantite"], prix_unitaire=0, unite=l["unite"])
    update_expression_besoin(conn, expression_id, statut="validee")
    return bon_id


# ---- Bon de commande (circuit interne, distinct du bon de commande de Factures frs) ----
def create_ep_bon_commande(conn, numero, date_commande, expression_id=None, fournisseur_code="",
                            entete="", pied_page=""):
    cur = conn.execute(
        """INSERT INTO ep_bons_commande (numero, date_commande, expression_id, fournisseur_code, entete,
                                          pied_page, statut)
           VALUES (?, ?, ?, ?, ?, ?, 'brouillon')""",
        (numero, date_commande, expression_id, fournisseur_code, entete, pied_page),
    )
    conn.commit()
    return cur.lastrowid


def update_ep_bon_commande(conn, bon_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE ep_bons_commande SET {cols} WHERE id = ?", (*fields.values(), bon_id))
    conn.commit()


def delete_ep_bon_commande(conn, bon_id):
    conn.execute("DELETE FROM ep_bon_commande_lignes WHERE bon_commande_id = ?", (bon_id,))
    conn.execute("DELETE FROM ep_bons_commande WHERE id = ?", (bon_id,))
    conn.commit()


def get_ep_bon_commande(conn, bon_id):
    row = conn.execute("SELECT * FROM ep_bons_commande WHERE id = ?", (bon_id,)).fetchone()
    return dict(row) if row else None


def list_ep_bons_commande(conn):
    """Liste des bons de commande, avec calcul du retard de paiement : basé
    sur la date de saisie une fois renseignée (paiement enregistré), sinon
    sur la date du jour tant que la date de saisie est vide (retard « en
    cours »), comparée à la date de paiement attendu — même principe que
    list_factures()/list_commandes() (Recouvrement/Achats)."""
    rows = [dict(r) for r in conn.execute(
        "SELECT * FROM ep_bons_commande ORDER BY date_commande DESC, id DESC"
    ).fetchall()]
    today = date.today().strftime("%Y-%m-%d")
    for r in rows:
        if not r.get("date_paiement_attendu"):
            r["statut_paiement"] = ""
            r["depassement_paiement"] = False
            continue
        if r.get("date_saisie"):
            retard = (datetime.strptime(r["date_saisie"], "%Y-%m-%d")
                      - datetime.strptime(r["date_paiement_attendu"], "%Y-%m-%d")).days
            r["statut_paiement"] = f"Saisi (retard {retard} j)" if retard > 0 else "Saisi à temps"
            r["depassement_paiement"] = retard > 0
        elif today > r["date_paiement_attendu"]:
            retard = (datetime.strptime(today, "%Y-%m-%d")
                      - datetime.strptime(r["date_paiement_attendu"], "%Y-%m-%d")).days
            r["statut_paiement"] = f"EN RETARD ({retard} j)"
            r["depassement_paiement"] = True
        else:
            r["statut_paiement"] = "En attente"
            r["depassement_paiement"] = False
    return rows


def add_ligne_ep_bon_commande(conn, bon_id, libelle, quantite, prix_unitaire=0, unite=None,
                               compte_charge=None, analytic_code=None):
    conn.execute(
        """INSERT INTO ep_bon_commande_lignes (bon_commande_id, libelle, quantite, prix_unitaire, unite,
                                                 compte_charge, analytic_code)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (bon_id, libelle, quantite or 0, prix_unitaire or 0, unite or None, compte_charge or None,
         analytic_code or None),
    )
    conn.commit()


def update_ligne_ep_bon_commande(conn, ligne_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE ep_bon_commande_lignes SET {cols} WHERE id = ?", (*fields.values(), ligne_id))
    conn.commit()


def delete_ligne_ep_bon_commande(conn, ligne_id):
    conn.execute("DELETE FROM ep_bon_commande_lignes WHERE id = ?", (ligne_id,))
    conn.commit()


def list_lignes_ep_bon_commande(conn, bon_id):
    rows = conn.execute(
        "SELECT * FROM ep_bon_commande_lignes WHERE bon_commande_id = ? ORDER BY id", (bon_id,)
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["montant_ht"] = (d["quantite"] or 0) * (d["prix_unitaire"] or 0)
        type_stock, _stock_compte, _contre_compte = (
            _match_stock_mapping(d["compte_charge"], ACHAT_STOCK_MAPPING) if d["compte_charge"] else (None, None, None)
        ) or (None, None, None)
        d["type_stock"] = type_stock
        result.append(d)
    return result


def compute_ep_bon_commande_totals(conn, bon_id):
    bon = get_ep_bon_commande(conn, bon_id)
    lignes = list_lignes_ep_bon_commande(conn, bon_id)
    total_ht = sum(l["montant_ht"] for l in lignes)
    retenue_taux = bon["retenue_taux"] if bon else 0
    retenue_montant = total_ht * (retenue_taux or 0) / 100
    net_a_payer = total_ht - retenue_montant
    return {"total_ht": total_ht, "retenue_taux": retenue_taux, "retenue_montant": retenue_montant,
            "net_a_payer": net_a_payer}


def valider_ep_bon_commande(conn, bon_id):
    """Valide le Bon de commande : COMPTABILISE DIRECTEMENT l'achat (Débit
    comptes de charge choisis par ligne avec code analytique, Crédit
    fournisseur, retenue fiscale optionnelle, entrée de stock automatique
    pour les lignes liées à un compte de marchandises/matières premières —
    voir _comptabiliser_lignes_achat(), même moteur que les Règlements) ET
    fait basculer le bon en Bordereau de livraison (lignes recopiées,
    quantité livrée initialisée à la quantité commandée). Un Règlement est
    également créé pour traçabilité, déjà marqué validé (mêmes écritures,
    pas de double comptabilisation). Chaque ligne DOIT avoir un compte de
    charge choisi et un fournisseur doit être renseigné, sous peine de
    refus explicite. Retourne (bordereau_id, reglement_id)."""
    bon = get_ep_bon_commande(conn, bon_id)
    if not bon:
        raise ValueError("Bon de commande introuvable.")
    if bon["statut"] == "validee":
        raise ValueError("Ce bon de commande est déjà validé.")
    lignes = list_lignes_ep_bon_commande(conn, bon_id)
    if not lignes:
        raise ValueError("Ajoutez au moins une ligne avant de valider.")

    fournisseur = get_fournisseur(conn, bon["fournisseur_code"]) if bon["fournisseur_code"] else None
    tiers_label = fournisseur["raison_sociale"] if fournisseur else (bon["fournisseur_code"] or "")
    piece = bon["piece"] or bon["numero"]

    _comptabiliser_lignes_achat(conn, bon["date_commande"], piece, "AC", bon["fournisseur_code"], lignes,
                                 tiers_label, retenue_taux=bon["retenue_taux"], retenue_compte=bon["retenue_compte"])

    bordereau_id = create_bordereau_livraison(conn, bon["numero"], bon["date_commande"], bon_commande_id=bon_id,
                                               entete=bon["entete"], pied_page=bon["pied_page"])
    for l in lignes:
        add_ligne_bordereau_livraison(conn, bordereau_id, l["libelle"], l["quantite"], l["quantite"], unite=l["unite"])

    reglement_id = create_reglement(conn, bon["numero"], bon["date_commande"], bon_commande_id=bon_id,
                                     fournisseur_code=bon["fournisseur_code"] or "",
                                     entete=bon["entete"], pied_page=bon["pied_page"],
                                     retenue_taux=bon["retenue_taux"], retenue_compte=bon["retenue_compte"])
    for l in lignes:
        add_ligne_reglement(conn, reglement_id, compte_charge=l["compte_charge"], libelle=l["libelle"],
                             quantite=l["quantite"], prix_unitaire=l["prix_unitaire"],
                             analytic_code=l.get("analytic_code"))
    # Le Règlement sert ici à la traçabilité (même pièce comptable déjà
    # postée par ce Bon de commande) — marqué validé directement pour ne
    # pas générer une seconde écriture s'il était validé séparément.
    update_reglement(conn, reglement_id, statut="validee", piece=piece)

    update_ep_bon_commande(conn, bon_id, statut="validee", piece=piece)
    return bordereau_id, reglement_id


# ---- Bordereau de livraison ----
def create_bordereau_livraison(conn, numero, date_livraison, bon_commande_id=None, entete="", pied_page=""):
    cur = conn.execute(
        """INSERT INTO bordereaux_livraison (numero, date_livraison, bon_commande_id, entete, pied_page, statut)
           VALUES (?, ?, ?, ?, ?, 'brouillon')""",
        (numero, date_livraison, bon_commande_id, entete, pied_page),
    )
    conn.commit()
    return cur.lastrowid


def update_bordereau_livraison(conn, bordereau_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE bordereaux_livraison SET {cols} WHERE id = ?", (*fields.values(), bordereau_id))
    conn.commit()


def delete_bordereau_livraison(conn, bordereau_id):
    conn.execute("DELETE FROM bordereau_livraison_lignes WHERE bordereau_id = ?", (bordereau_id,))
    conn.execute("DELETE FROM bordereaux_livraison WHERE id = ?", (bordereau_id,))
    conn.commit()


def get_bordereau_livraison(conn, bordereau_id):
    row = conn.execute("SELECT * FROM bordereaux_livraison WHERE id = ?", (bordereau_id,)).fetchone()
    return dict(row) if row else None


def list_bordereaux_livraison(conn):
    rows = conn.execute("SELECT * FROM bordereaux_livraison ORDER BY date_livraison DESC, id DESC").fetchall()
    return [dict(r) for r in rows]


def add_ligne_bordereau_livraison(conn, bordereau_id, libelle, quantite_commandee, quantite_livree, unite=None):
    conn.execute(
        """INSERT INTO bordereau_livraison_lignes (bordereau_id, libelle, quantite_commandee, quantite_livree, unite)
           VALUES (?, ?, ?, ?, ?)""",
        (bordereau_id, libelle, quantite_commandee or 0, quantite_livree or 0, unite or None),
    )
    conn.commit()


def update_ligne_bordereau_livraison(conn, ligne_id, quantite_livree):
    conn.execute("UPDATE bordereau_livraison_lignes SET quantite_livree = ? WHERE id = ?",
                 (quantite_livree or 0, ligne_id))
    conn.commit()


def delete_ligne_bordereau_livraison(conn, ligne_id):
    conn.execute("DELETE FROM bordereau_livraison_lignes WHERE id = ?", (ligne_id,))
    conn.commit()


def list_lignes_bordereau_livraison(conn, bordereau_id):
    rows = conn.execute(
        "SELECT * FROM bordereau_livraison_lignes WHERE bordereau_id = ? ORDER BY id", (bordereau_id,)
    ).fetchall()
    return [dict(r) for r in rows]


def valider_bordereau_livraison(conn, bordereau_id):
    """Dernière étape du circuit : marque simplement le bordereau comme
    validé (réception confirmée) — AUCUNE écriture comptable, fin de
    chaîne."""
    bordereau = get_bordereau_livraison(conn, bordereau_id)
    if not bordereau:
        raise ValueError("Bordereau de livraison introuvable.")
    if bordereau["statut"] == "validee":
        raise ValueError("Ce bordereau est déjà validé.")
    update_bordereau_livraison(conn, bordereau_id, statut="validee")


# ---- Règlements (menu ENGAGEMENTS-PROJETS) ----
# Créés automatiquement à la validation d'un Bon de commande interne (lignes
# recopiées, SANS compte de charge ni code analytique — à choisir ici) ou
# directement. C'EST CET ÉCRAN, ET LUI SEUL, qui comptabilise le circuit
# interne : une fois chaque ligne rattachée à un compte de charge (classe 6),
# avec code analytique et retenue fiscale optionnels, sa validation envoie
# une écriture équilibrée en Saisie — même principe que valider_facture_achat.
def create_reglement(conn, numero, date_reglement, bon_commande_id=None, fournisseur_code="",
                      entete="", pied_page="", retenue_taux=None, retenue_compte=None):
    if retenue_taux is None:
        retenue_taux = get_setting(conn, "retenue_taux_defaut", RETENUE_TAUX_DEFAUT)
    if retenue_compte is None:
        retenue_compte = get_text_setting(conn, "retenue_compte_defaut", COMPTE_RETENUE_DEFAUT)
    cur = conn.execute(
        """INSERT INTO reglements (numero, date_reglement, bon_commande_id, fournisseur_code, entete,
                                    pied_page, retenue_taux, retenue_compte, statut)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, 'brouillon')""",
        (numero, date_reglement, bon_commande_id, fournisseur_code, entete, pied_page, retenue_taux, retenue_compte),
    )
    conn.commit()
    return cur.lastrowid


def update_reglement(conn, reglement_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE reglements SET {cols} WHERE id = ?", (*fields.values(), reglement_id))
    conn.commit()


def delete_reglement(conn, reglement_id):
    reglement = get_reglement(conn, reglement_id)
    if reglement and reglement["statut"] == "validee":
        raise ValueError("Impossible de supprimer un règlement déjà validé — dévalidez-le d'abord.")
    conn.execute("DELETE FROM reglement_lignes WHERE reglement_id = ?", (reglement_id,))
    conn.execute("DELETE FROM reglements WHERE id = ?", (reglement_id,))
    conn.commit()


def get_reglement(conn, reglement_id):
    row = conn.execute("SELECT * FROM reglements WHERE id = ?", (reglement_id,)).fetchone()
    return dict(row) if row else None


def list_reglements(conn):
    rows = conn.execute(
        """SELECT r.*, COALESCE(f.raison_sociale, r.fournisseur_code, '') AS raison_sociale
           FROM reglements r LEFT JOIN fournisseurs f ON f.code = r.fournisseur_code
           ORDER BY r.date_reglement DESC, r.id DESC"""
    ).fetchall()
    return [dict(r) for r in rows]


def add_ligne_reglement(conn, reglement_id, compte_charge, libelle, quantite, prix_unitaire=0, analytic_code=None):
    conn.execute(
        """INSERT INTO reglement_lignes (reglement_id, compte_charge, libelle, quantite, prix_unitaire,
                                          analytic_code)
           VALUES (?, ?, ?, ?, ?, ?)""",
        (reglement_id, compte_charge or None, libelle, quantite or 0, prix_unitaire or 0, analytic_code or None),
    )
    conn.commit()


def update_ligne_reglement(conn, ligne_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE reglement_lignes SET {cols} WHERE id = ?", (*fields.values(), ligne_id))
    conn.commit()


def delete_ligne_reglement(conn, ligne_id):
    conn.execute("DELETE FROM reglement_lignes WHERE id = ?", (ligne_id,))
    conn.commit()


def list_lignes_reglement(conn, reglement_id):
    rows = conn.execute(
        "SELECT * FROM reglement_lignes WHERE reglement_id = ? ORDER BY id", (reglement_id,)
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["montant_ht"] = (d["quantite"] or 0) * (d["prix_unitaire"] or 0)
        type_stock, stock_compte, contre_compte = (
            _match_stock_mapping(d["compte_charge"], ACHAT_STOCK_MAPPING) if d["compte_charge"] else (None, None, None)
        ) or (None, None, None)
        d["type_stock"] = type_stock
        result.append(d)
    return result


def compute_reglement_totals(conn, reglement_id):
    reglement = get_reglement(conn, reglement_id)
    lignes = list_lignes_reglement(conn, reglement_id)
    total_ht = sum(l["montant_ht"] for l in lignes)
    retenue_taux = reglement["retenue_taux"] if reglement else 0
    retenue_montant = total_ht * (retenue_taux or 0) / 100
    net_a_payer = total_ht - retenue_montant
    return {"total_ht": total_ht, "retenue_taux": retenue_taux, "retenue_montant": retenue_montant,
            "net_a_payer": net_a_payer}


def _comptabiliser_lignes_achat(conn, date_str, piece, journal, fournisseur_code, lignes, tiers_label,
                                 retenue_taux=0, retenue_compte=None):
    """Comptabilise un ensemble de lignes d'achat (Débit comptes de charge
    choisis par ligne, avec code analytique / Crédit fournisseur + retenue
    fiscale optionnelle), plus une entrée de stock automatique pour chaque
    ligne liée à un compte de marchandises (31) ou de matières premières
    (32) — logique PARTAGÉE entre la validation d'un Règlement
    (ENGAGEMENTS-PROJETS > Règlements) et la validation directe d'un Bon de
    commande (ENGAGEMENTS-PROJETS > Bon de commande). Chaque ligne DOIT
    avoir un compte de charge choisi, sous peine de refus explicite."""
    sans_compte = [l["libelle"] for l in lignes if not l.get("compte_charge")]
    if sans_compte:
        raise ValueError(
            "Chaque ligne doit avoir un compte débiteur choisi (charge ou immobilisation) avant "
            "validation — manquant pour : " + ", ".join(sans_compte)
        )
    sans_montant = [l["libelle"] for l in lignes if not l.get("montant_ht")]
    if sans_montant:
        raise ValueError(
            "Chaque ligne doit avoir un montant strictement positif (quantité × prix unitaire) avant "
            "validation — vérifiez le prix unitaire (probablement resté à 0, ou saisi dans le mauvais "
            "champ) pour : " + ", ".join(sans_montant)
        )
    if not fournisseur_code or not fournisseur_exists(conn, fournisseur_code):
        raise ValueError(f"Le fournisseur « {fournisseur_code} » n'existe pas ou n'est pas renseigné.")

    total_ht = sum(l["montant_ht"] for l in lignes)
    retenue_montant = total_ht * (retenue_taux or 0) / 100
    net_a_payer = total_ht - retenue_montant

    for l in lignes:
        add_entry(conn, date_str, piece, journal, l["compte_charge"], "", l["libelle"],
                  l["montant_ht"], 0, analytic_code=l.get("analytic_code") or "",
                  fournisseur_code=fournisseur_code, quantite=l.get("quantite") or 0)

    add_entry(conn, date_str, piece, journal, "401000", tiers_label, piece, 0, net_a_payer,
              fournisseur_code=fournisseur_code)

    if retenue_montant:
        add_entry(conn, date_str, piece, journal, retenue_compte, "",
                  f"Retenue {retenue_taux:g}% pièce {piece}", 0, retenue_montant)

    for l in lignes:
        if not l.get("type_stock"):
            continue
        _, stock_compte, contre_compte = _match_stock_mapping(l["compte_charge"], ACHAT_STOCK_MAPPING)
        montant_entree = l["montant_ht"]
        if montant_entree <= 0:
            continue
        add_entry(conn, date_str, piece, journal, stock_compte, "", f"Entrée stock — {l['libelle']}",
                  montant_entree, 0, quantite=l.get("quantite") or 0)
        add_entry(conn, date_str, piece, journal, contre_compte, "", f"Entrée stock — {l['libelle']}",
                  0, montant_entree)

    return {"total_ht": total_ht, "retenue_montant": retenue_montant, "net_a_payer": net_a_payer}


def valider_reglement(conn, reglement_id, exercice=None):
    """Comptabilise le règlement — voir _comptabiliser_lignes_achat().
    Retourne la liste des avertissements."""
    reglement = get_reglement(conn, reglement_id)
    if not reglement:
        raise ValueError("Règlement introuvable.")
    if reglement["statut"] == "validee":
        raise ValueError("Ce règlement est déjà validé.")
    lignes = list_lignes_reglement(conn, reglement_id)
    if not lignes:
        raise ValueError("Le règlement ne contient aucune ligne.")

    date_str = reglement["date_reglement"]
    piece = reglement["numero"]
    fournisseur = get_fournisseur(conn, reglement["fournisseur_code"])
    tiers_label = fournisseur["raison_sociale"] if fournisseur else reglement["fournisseur_code"]

    _comptabiliser_lignes_achat(conn, date_str, piece, "AC", reglement["fournisseur_code"], lignes, tiers_label,
                                 retenue_taux=reglement["retenue_taux"], retenue_compte=reglement["retenue_compte"])

    update_reglement(conn, reglement_id, statut="validee", piece=piece)
    return []


def enregistrer_paiement_reglement(conn, reglement_id, date_paiement, compte_paiement):
    """Enregistre le PAIEMENT bancaire/caisse réel d'un Règlement déjà
    validé (la charge et la dette fournisseur — compte 401000 — ont déjà
    été comptabilisées par valider_reglement()) : comptabilise
    l'encaissement Débit fournisseur (401000, soldant sa dette) / Crédit
    compte banque/caisse choisi, pour le montant NET à payer (après
    retenue). Ne comptabilise le paiement qu'UNE SEULE FOIS par règlement
    (garde-fou `paiement_comptabilise`)."""
    reglement = get_reglement(conn, reglement_id)
    if not reglement:
        raise ValueError("Règlement introuvable.")
    if reglement["statut"] != "validee":
        raise ValueError("Ce règlement doit d'abord être validé (charge comptabilisée) avant d'enregistrer un paiement.")
    if not compte_paiement or not account_exists(conn, compte_paiement):
        raise ValueError(f"Le compte de paiement « {compte_paiement} » n'existe pas.")
    if account_racine(compte_paiement) != "5":
        raise ValueError("Le compte de paiement doit être un compte de trésorerie (classe 5 — banque ou caisse).")
    if reglement["paiement_comptabilise"]:
        raise ValueError("Le paiement de ce règlement a déjà été comptabilisé.")

    totals = compute_reglement_totals(conn, reglement_id)
    fournisseur = get_fournisseur(conn, reglement["fournisseur_code"])
    tiers_label = fournisseur["raison_sociale"] if fournisseur else reglement["fournisseur_code"]
    piece = reglement["piece"] or reglement["numero"]

    _check_exercice_editable(conn, date_paiement)
    add_entry(conn, date_paiement, piece, "BQ", "401000", tiers_label, f"Paiement règlement {piece}",
              totals["net_a_payer"], 0, fournisseur_code=reglement["fournisseur_code"])
    add_entry(conn, date_paiement, piece, "BQ", compte_paiement, tiers_label, f"Paiement règlement {piece}",
              0, totals["net_a_payer"])

    update_reglement(conn, reglement_id, date_paiement=date_paiement, compte_paiement=compte_paiement,
                      paiement_comptabilise=1)
    return totals["net_a_payer"]


def devalider_paiement_reglement(conn, reglement_id):
    """Annule le paiement bancaire déjà comptabilisé d'un règlement (en cas
    d'erreur — mauvais compte banque, mauvais montant...) : supprime les
    écritures du journal 'BQ' pour ce règlement et remet
    `paiement_comptabilise` à 0, SANS toucher à la charge/dette fournisseur
    (déjà validée séparément — voir devalider_reglement pour ça)."""
    reglement = get_reglement(conn, reglement_id)
    if not reglement:
        raise ValueError("Règlement introuvable.")
    if not reglement["paiement_comptabilise"]:
        raise ValueError("Aucun paiement comptabilisé pour ce règlement — rien à annuler.")
    piece = reglement["piece"] or reglement["numero"]
    ids = [r["id"] for r in conn.execute(
        "SELECT id FROM entries WHERE piece = ? AND journal = 'BQ'", (piece,)
    ).fetchall()]
    deleted, errors = delete_entries_bulk(conn, ids)
    if errors:
        raise ValueError("Impossible d'annuler ce paiement : " + " ; ".join(errors))
    update_reglement(conn, reglement_id, date_paiement=None, compte_paiement=None, paiement_comptabilise=0)
    return deleted


def devalider_reglement(conn, reglement_id):
    """Repasse un règlement VALIDÉ en brouillon modifiable, en cas d'erreur
    sur les chiffres constatée après validation : supprime toutes les
    écritures comptables générées par sa validation (piece=numéro,
    journal='AC') puis remet son statut à « brouillon ». Refuse si
    l'exercice comptable est clôturé. Retourne le nombre d'écritures
    supprimées."""
    reglement = get_reglement(conn, reglement_id)
    if not reglement:
        raise ValueError("Règlement introuvable.")
    if reglement["statut"] != "validee":
        raise ValueError("Ce règlement n'est pas validé — rien à corriger.")
    exercice_reglement = _exercice_of_date(reglement["date_reglement"])
    if is_exercice_cloture(conn, exercice_reglement):
        raise ValueError(f"L'exercice {exercice_reglement} de ce règlement est clôturé : impossible de le corriger.")
    if reglement["paiement_comptabilise"]:
        devalider_paiement_reglement(conn, reglement_id)
    ids = [r["id"] for r in conn.execute(
        "SELECT id FROM entries WHERE piece = ? AND journal = 'AC'", (reglement["numero"],)
    ).fetchall()]
    deleted, errors = delete_entries_bulk(conn, ids)
    if errors:
        raise ValueError("Impossible de corriger ce règlement : " + " ; ".join(errors))
    update_reglement(conn, reglement_id, statut="brouillon")
    # Si ce règlement a été créé/comptabilisé directement par la validation
    # d'un Bon de commande (bon_commande_id renseigné), repasse aussi CE bon
    # en brouillon modifiable — sinon il resterait verrouillé « VALIDÉ » sans
    # plus aucune écriture réelle derrière lui (état incohérent).
    if reglement["bon_commande_id"]:
        conn.execute("UPDATE ep_bons_commande SET statut = 'brouillon' WHERE id = ? AND statut = 'validee'",
                     (reglement["bon_commande_id"],))
        conn.commit()
    return deleted


def devalider_ep_bon_commande(conn, bon_id):
    """Repasse un Bon de commande VALIDÉ en brouillon modifiable, en cas
    d'erreur sur les chiffres constatée après validation (ex. prix unitaire
    saisi à 0 par erreur) : supprime les écritures comptables générées par
    sa validation (piece, journal='AC') et repasse en brouillon À LA FOIS
    ce Bon de commande ET le Règlement lié (mêmes écritures, pour rester
    cohérent — voir devalider_reglement, qui fait le symétrique en sens
    inverse). Le Bordereau de livraison déjà créé n'est PAS supprimé (suivi
    de réception indépendant de la comptabilité). Refuse si l'exercice est
    clôturé. Retourne le nombre d'écritures supprimées."""
    bon = get_ep_bon_commande(conn, bon_id)
    if not bon:
        raise ValueError("Bon de commande introuvable.")
    if bon["statut"] != "validee":
        raise ValueError("Ce bon de commande n'est pas validé — rien à corriger.")
    piece = bon["piece"] or bon["numero"]
    exercice_bon = _exercice_of_date(bon["date_commande"])
    if is_exercice_cloture(conn, exercice_bon):
        raise ValueError(f"L'exercice {exercice_bon} de ce bon de commande est clôturé : impossible de le corriger.")
    ids = [r["id"] for r in conn.execute(
        "SELECT id FROM entries WHERE piece = ? AND journal = 'AC'", (piece,)
    ).fetchall()]
    deleted, errors = delete_entries_bulk(conn, ids)
    if errors:
        raise ValueError("Impossible de corriger ce bon de commande : " + " ; ".join(errors))
    update_ep_bon_commande(conn, bon_id, statut="brouillon")
    conn.execute("UPDATE reglements SET statut = 'brouillon' WHERE bon_commande_id = ? AND statut = 'validee'",
                 (bon_id,))
    conn.commit()
    return deleted


# ---------------------------------------------------------------------------
# TRANSPORT (Parc auto, Missions, Pièces de rechange, Réparations) — suivi
# opérationnel, sans lien avec la comptabilité (même principe que le circuit
# Expression de besoin / Bon de commande). « Pièces de rechange » est un
# stock PARTAGÉ, utilisé aussi bien pour les réparations de véhicules
# (menu TRANSPORT) que pour la maintenance générale (menu MAINTENANCE-ÉNERGIE).
# ---------------------------------------------------------------------------

# ---- Parc auto ----
def add_vehicule(conn, immatriculation, marque="", modele="", type_vehicule="",
                  date_acquisition="", chauffeur_affecte="", statut="actif", notes=""):
    cur = conn.execute(
        """INSERT INTO vehicules (immatriculation, marque, modele, type_vehicule, date_acquisition,
                                   chauffeur_affecte, statut, notes)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (immatriculation, marque, modele, type_vehicule, date_acquisition, chauffeur_affecte, statut, notes),
    )
    conn.commit()
    return cur.lastrowid


def update_vehicule(conn, vehicule_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE vehicules SET {cols} WHERE id = ?", (*fields.values(), vehicule_id))
    conn.commit()


def delete_vehicule(conn, vehicule_id):
    conn.execute("DELETE FROM vehicules WHERE id = ?", (vehicule_id,))
    conn.commit()


def get_vehicule(conn, vehicule_id):
    row = conn.execute("SELECT * FROM vehicules WHERE id = ?", (vehicule_id,)).fetchone()
    return dict(row) if row else None


def list_vehicules(conn):
    return [dict(r) for r in conn.execute("SELECT * FROM vehicules ORDER BY immatriculation").fetchall()]


# ---- Missions ----
def add_mission(conn, destination, vehicule_id=None, chauffeur="", motif="", date_depart="", date_retour="",
                 km_depart=None, km_retour=None, statut="en_cours", notes=""):
    cur = conn.execute(
        """INSERT INTO missions (vehicule_id, chauffeur, destination, motif, date_depart, date_retour,
                                  km_depart, km_retour, statut, notes)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (vehicule_id, chauffeur, destination, motif, date_depart, date_retour, km_depart, km_retour, statut, notes),
    )
    conn.commit()
    return cur.lastrowid


def update_mission(conn, mission_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE missions SET {cols} WHERE id = ?", (*fields.values(), mission_id))
    conn.commit()


def delete_mission(conn, mission_id):
    conn.execute("DELETE FROM missions WHERE id = ?", (mission_id,))
    conn.commit()


def list_missions(conn):
    rows = conn.execute(
        """SELECT m.*, COALESCE(v.immatriculation, '') AS immatriculation
           FROM missions m LEFT JOIN vehicules v ON v.id = m.vehicule_id
           ORDER BY COALESCE(m.date_depart, '') DESC, m.id DESC"""
    ).fetchall()
    return [dict(r) for r in rows]


# ---- Pièces de rechange (stock partagé Transport / Maintenance) ----
def add_piece_rechange(conn, designation, code="", quantite_stock=0, unite="", cout_unitaire=0,
                        fournisseur_code="", notes=""):
    cur = conn.execute(
        """INSERT INTO pieces_rechange (code, designation, quantite_stock, unite, cout_unitaire,
                                         fournisseur_code, notes)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (code, designation, quantite_stock or 0, unite, cout_unitaire or 0, fournisseur_code, notes),
    )
    conn.commit()
    return cur.lastrowid


def update_piece_rechange(conn, piece_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE pieces_rechange SET {cols} WHERE id = ?", (*fields.values(), piece_id))
    conn.commit()


def delete_piece_rechange(conn, piece_id):
    conn.execute("DELETE FROM pieces_rechange WHERE id = ?", (piece_id,))
    conn.commit()


def get_piece_rechange(conn, piece_id):
    row = conn.execute("SELECT * FROM pieces_rechange WHERE id = ?", (piece_id,)).fetchone()
    return dict(row) if row else None


def list_pieces_rechange(conn):
    return [dict(r) for r in conn.execute("SELECT * FROM pieces_rechange ORDER BY designation").fetchall()]


# ---- Réparations (véhicule optionnel — une réparation sans véhicule =
# maintenance générale d'équipement, utilisée depuis le menu MAINTENANCE) ----
def create_reparation(conn, description, vehicule_id=None, date_reparation=None, garage="",
                       cout_main_oeuvre=0, statut="en_cours", notes=""):
    date_reparation = date_reparation or date.today().strftime("%Y-%m-%d")
    cur = conn.execute(
        """INSERT INTO reparations (vehicule_id, date_reparation, description, garage, cout_main_oeuvre,
                                     statut, notes)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (vehicule_id, date_reparation, description, garage, cout_main_oeuvre or 0, statut, notes),
    )
    conn.commit()
    return cur.lastrowid


def update_reparation(conn, reparation_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE reparations SET {cols} WHERE id = ?", (*fields.values(), reparation_id))
    conn.commit()


def delete_reparation(conn, reparation_id):
    conn.execute("DELETE FROM reparation_lignes WHERE reparation_id = ?", (reparation_id,))
    conn.execute("DELETE FROM reparations WHERE id = ?", (reparation_id,))
    conn.commit()


def get_reparation(conn, reparation_id):
    row = conn.execute("SELECT * FROM reparations WHERE id = ?", (reparation_id,)).fetchone()
    return dict(row) if row else None


def list_reparations(conn):
    rows = conn.execute(
        """SELECT r.*, COALESCE(v.immatriculation, '') AS immatriculation
           FROM reparations r LEFT JOIN vehicules v ON v.id = r.vehicule_id
           ORDER BY r.date_reparation DESC, r.id DESC"""
    ).fetchall()
    return [dict(r) for r in rows]


def add_ligne_reparation(conn, reparation_id, piece_id, quantite=1):
    """Ajoute une pièce utilisée à une réparation et DÉCRÉMENTE le stock de
    cette pièce (pieces_rechange.quantite_stock) — refuse si le stock est
    insuffisant."""
    piece = get_piece_rechange(conn, piece_id)
    if not piece:
        raise ValueError("Pièce de rechange introuvable.")
    if piece["quantite_stock"] < quantite:
        raise ValueError(
            f"Stock insuffisant pour « {piece['designation']} » : {piece['quantite_stock']:g} disponible(s), "
            f"{quantite:g} demandé(s)."
        )
    conn.execute("INSERT INTO reparation_lignes (reparation_id, piece_id, quantite) VALUES (?, ?, ?)",
                 (reparation_id, piece_id, quantite))
    conn.execute("UPDATE pieces_rechange SET quantite_stock = quantite_stock - ? WHERE id = ?",
                 (quantite, piece_id))
    conn.commit()


def delete_ligne_reparation(conn, ligne_id):
    """Supprime une ligne de réparation et RESTITUE la quantité au stock de
    la pièce (correction d'une erreur de saisie)."""
    row = conn.execute("SELECT * FROM reparation_lignes WHERE id = ?", (ligne_id,)).fetchone()
    if not row:
        return
    conn.execute("UPDATE pieces_rechange SET quantite_stock = quantite_stock + ? WHERE id = ?",
                 (row["quantite"], row["piece_id"]))
    conn.execute("DELETE FROM reparation_lignes WHERE id = ?", (ligne_id,))
    conn.commit()


def list_lignes_reparation(conn, reparation_id):
    rows = conn.execute(
        """SELECT rl.*, p.designation, p.cout_unitaire, p.unite
           FROM reparation_lignes rl JOIN pieces_rechange p ON p.id = rl.piece_id
           WHERE rl.reparation_id = ? ORDER BY rl.id""",
        (reparation_id,),
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["montant"] = d["quantite"] * d["cout_unitaire"]
        result.append(d)
    return result


def compute_cout_total_reparation(conn, reparation_id):
    reparation = get_reparation(conn, reparation_id)
    lignes = list_lignes_reparation(conn, reparation_id)
    cout_pieces = sum(l["montant"] for l in lignes)
    return cout_pieces + (reparation["cout_main_oeuvre"] if reparation else 0)


# ---------------------------------------------------------------------------
# IMMOBILISATIONS — liste des comptes de classe 2 avec fournisseur/prix
# d'achat (fiche), et taux d'amortissement paramétrables par catégorie
# (réutilise IMMO_CATEGORIES, la même catégorisation que le Bilan).
# ---------------------------------------------------------------------------
def set_immobilisation_fiche(conn, compte, fournisseur_code=None, prix_achat=None, date_acquisition=None,
                              base_repartition_quantite=None, base_repartition_unite=None,
                              amortissement_annuel_manuel=None):
    """Crée ou met à jour la fiche (fournisseur, prix d'achat, date, base de
    répartition de l'amortissement) d'un compte d'immobilisation — les
    champs non fournis (None) conservent leur valeur existante plutôt que
    d'être écrasés à vide.

    `base_repartition_quantite`/`base_repartition_unite` : quantité annuelle
    de référence (ex. 5000 tonnes/an, ou 2000 heures/an) permettant de
    calculer le coût d'amortissement PAR UNITÉ D'USAGE de cet équipement —
    utilisée dans les recettes de fabrication (composant « Amortissement
    d'équipement »).

    `amortissement_annuel_manuel` : montant d'amortissement annuel déclaré
    directement (F CFA/an), utilisé à la place de l'amortissement réellement
    comptabilisé (dotations 68x/28x) quand celui-ci n'est pas encore saisi
    dans la Saisie — pratique pour chiffrer un coût de production sans
    attendre que la comptabilité des dotations soit à jour. Si les deux
    sont renseignés, le montant RÉELLEMENT comptabilisé reste prioritaire
    (voir compute_cout_amortissement_unitaire)."""
    existing = conn.execute("SELECT * FROM immobilisations_fiche WHERE compte = ?", (compte,)).fetchone()
    if fournisseur_code is None:
        fournisseur_code = existing["fournisseur_code"] if existing else ""
    if prix_achat is None:
        prix_achat = existing["prix_achat"] if existing else 0
    if date_acquisition is None:
        date_acquisition = existing["date_acquisition"] if existing else ""
    if base_repartition_quantite is None:
        base_repartition_quantite = existing["base_repartition_quantite"] if existing else None
    if base_repartition_unite is None:
        base_repartition_unite = existing["base_repartition_unite"] if existing else ""
    if amortissement_annuel_manuel is None:
        amortissement_annuel_manuel = existing["amortissement_annuel_manuel"] if existing else None
    conn.execute(
        """INSERT OR REPLACE INTO immobilisations_fiche
           (compte, fournisseur_code, prix_achat, date_acquisition,
            base_repartition_quantite, base_repartition_unite, amortissement_annuel_manuel)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (compte, fournisseur_code, prix_achat or 0, date_acquisition,
         base_repartition_quantite, base_repartition_unite, amortissement_annuel_manuel),
    )
    conn.commit()


def get_immobilisation_fiche(conn, compte):
    row = conn.execute("SELECT * FROM immobilisations_fiche WHERE compte = ?", (compte,)).fetchone()
    if row:
        return dict(row)
    return {"compte": compte, "fournisseur_code": "", "prix_achat": 0, "date_acquisition": "",
            "base_repartition_quantite": None, "base_repartition_unite": "",
            "amortissement_annuel_manuel": None}


def categorie_immobilisation(compte):
    """Renvoie le libellé de catégorie IMMO_CATEGORIES correspondant à un
    compte de classe 2 (celle utilisée par le Bilan) — None si hors plage
    (ex. racine 200 isolée)."""
    try:
        code_int = int(compte)
    except (TypeError, ValueError):
        return None
    for label, brut_ranges, _amort_ranges in IMMO_CATEGORIES:
        for lo, hi in brut_ranges:
            if lo <= code_int <= hi:
                return label
    return None


def list_taux_amortissement(conn):
    rows = {r["categorie"]: r["taux_pct"] for r in conn.execute("SELECT * FROM taux_amortissement").fetchall()}
    return [{"categorie": label, "taux_pct": rows.get(label, 0.0)} for label, _b, _a in IMMO_CATEGORIES]


def set_taux_amortissement(conn, categorie, taux_pct):
    conn.execute("INSERT OR REPLACE INTO taux_amortissement (categorie, taux_pct) VALUES (?, ?)",
                 (categorie, taux_pct or 0))
    conn.commit()


def compute_immobilisations_liste(conn, exercice=None):
    """Liste des comptes de classe 2 (immobilisations) AYANT un solde non
    nul dans la Balance, enrichie de leur fiche (fournisseur, prix d'achat)
    et de leur catégorie/taux d'amortissement — avec Valeur Brute (solde du
    compte), Amortissement (solde réel du compte 28x/29x correspondant,
    même méthode exacte que le Bilan — voir IMMO_CATEGORIES) et Valeur
    Nette. `taux_pct` est indicatif (paramétré dans Amortissements) : le
    montant Amortissement affiché reste celui RÉELLEMENT comptabilisé, pas
    une simulation, pour ne jamais diverger de la Balance/du Bilan."""
    exercice = exercice or get_current_exercice(conn)
    balance = compute_balance(conn, only_with_movement=False, exercice=exercice)
    taux_par_categorie = {t["categorie"]: t["taux_pct"] for t in list_taux_amortissement(conn)}

    amort_par_compte = {}
    for label, _brut_ranges, amort_ranges in IMMO_CATEGORIES:
        for b in balance:
            if b["classe"] != "2":
                continue
            code_int = int(b["code"])
            for lo, hi in amort_ranges:
                if lo <= code_int <= hi:
                    amort_par_compte.setdefault(label, 0.0)
                    amort_par_compte[label] += b["solde_cloture"]

    result = []
    for b in balance:
        if b["classe"] != "2" or not b["solde_cloture"]:
            continue
        code_int = int(b["code"])
        if code_int >= 280000:
            continue  # comptes d'amortissement eux-mêmes : pas une immobilisation
        categorie = categorie_immobilisation(b["code"]) or "Non classé"
        fiche = get_immobilisation_fiche(conn, b["code"])
        fournisseur = get_fournisseur(conn, fiche["fournisseur_code"]) if fiche["fournisseur_code"] else None
        brut = b["solde_cloture"]
        amort_categorie = amort_par_compte.get(categorie, 0.0)
        # Répartit l'amortissement RÉEL de la catégorie au prorata du brut de
        # chaque compte de cette catégorie (l'amortissement est comptabilisé
        # par compte 28x global, pas par sous-compte 2x individuel).
        brut_categorie_total = sum(
            x["solde_cloture"] for x in balance
            if x["classe"] == "2" and int(x["code"]) < 280000 and (categorie_immobilisation(x["code"]) or "Non classé") == categorie
        )
        amort = (amort_categorie * (brut / brut_categorie_total)) if brut_categorie_total else 0.0
        result.append({
            "compte": b["code"], "libelle": b["label"], "categorie": categorie,
            "taux_pct": taux_par_categorie.get(categorie, 0.0),
            "fournisseur_code": fiche["fournisseur_code"],
            "fournisseur_nom": fournisseur["raison_sociale"] if fournisseur else "",
            "prix_achat": fiche["prix_achat"], "date_acquisition": fiche["date_acquisition"],
            "valeur_brute": brut, "amortissement": amort, "valeur_nette": brut + amort,
        })
    result.sort(key=lambda l: (l["categorie"], l["compte"]))
    return result


def compute_cout_amortissement_unitaire(conn, compte_immobilisation, exercice=None):
    """Coût d'amortissement PAR UNITÉ D'USAGE (par tonne, par heure...) d'un
    équipement précis, pour l'incorporer au coût de production d'une
    recette de fabrication (composant « Amortissement d'équipement »).

    = amortissement à retenir pour ce compte sur l'exercice
      ÷ base_repartition_quantite (quantité annuelle de référence saisie
      sur la fiche de l'équipement, ex. 5000 tonnes/an ou 2000 heures/an).

    L'amortissement à retenir est, dans l'ordre de priorité :
    1. celui RÉELLEMENT comptabilisé pour ce compte sur l'exercice (même
       valeur que dans l'écran Immobilisations, au prorata de sa catégorie —
       voir compute_immobilisations_liste), s'il est non nul ;
    2. sinon, le montant DÉCLARÉ MANUELLEMENT sur la fiche de l'équipement
       (« Amortissement annuel » — pratique quand les dotations aux
       amortissements ne sont pas encore saisies dans la Saisie).

    Renvoie None si le compte n'a pas de base de répartition renseignée, ou
    si aucun amortissement (réel ou déclaré) n'est disponible — dans les
    deux cas, il faut compléter la fiche de l'équipement dans l'écran
    Immobilisations avant de pouvoir l'utiliser dans une recette."""
    fiche = get_immobilisation_fiche(conn, compte_immobilisation)
    base = fiche.get("base_repartition_quantite")
    if not base:
        return None
    liste = compute_immobilisations_liste(conn, exercice=exercice)
    ligne = next((l for l in liste if l["compte"] == compte_immobilisation), None)
    # « amortissement » est stocké en négatif par convention interne (valeur_nette = brut + amortissement) ;
    # ici on veut un COÛT positif à incorporer dans le coût de production.
    amortissement_reel = abs(ligne["amortissement"]) if ligne and ligne["amortissement"] else 0.0
    if amortissement_reel:
        return amortissement_reel / base
    amortissement_manuel = fiche.get("amortissement_annuel_manuel")
    if amortissement_manuel:
        return amortissement_manuel / base
    return None


IMMOBILISATION_IMPORT_COLUMNS = [
    ("compte", "N° Compte", ["n° compte", "compte", "n compte", "numero compte"]),
    ("fournisseur_code", "Code fournisseur", ["code fournisseur", "fournisseur"]),
    ("prix_achat", "Prix d'achat", ["prix d'achat", "prix achat", "prix"]),
    ("date_acquisition", "Date d'acquisition (JJ/MM/AAAA)",
     ["date d'acquisition (jj/mm/aaaa)", "date d'acquisition", "date acquisition"]),
    ("base_repartition_quantite", "Base de répartition (quantité annuelle)",
     ["base de répartition (quantité annuelle)", "base de repartition", "quantite annuelle"]),
    ("base_repartition_unite", "Unité (tonnes, heures...)", ["unité (tonnes, heures...)", "unite", "unité"]),
    ("amortissement_annuel_manuel", "Amortissement annuel (si pas comptabilisé)",
     ["amortissement annuel (si pas comptabilisé)", "amortissement annuel", "amortissement manuel"]),
]


def export_immobilisations_template(path):
    """Modèle Excel pour importer/mettre à jour en masse les fiches
    d'immobilisations (fournisseur, prix d'achat, date, base de répartition,
    amortissement manuel) — le compte DOIT déjà exister dans le Plan
    comptable et avoir un solde dans la Balance (classe 2) ; ce modèle ne
    crée pas de nouveaux comptes, il complète leur fiche."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Immobilisations"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for i, (_, label, _) in enumerate(IMMOBILISATION_IMPORT_COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    example = ["241101", "FRS-0001", 12000000, "01/01/2026", 5000, "tonnes", 2000000]
    for i, val in enumerate(example, start=1):
        ws.cell(row=2, column=i, value=val)
    for i, w in enumerate([14, 16, 14, 22, 26, 20, 30], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def parse_immobilisations_xlsx(path):
    """Lit un fichier Excel de fiches d'immobilisations et renvoie la liste
    des lignes sous forme de dicts, SANS toucher à la base de données —
    utilisable aussi bien localement (bureau) qu'à distance (le client
    réseau lit le fichier sur son propre poste, puis envoie les lignes au
    serveur via apply_immobilisations_rows)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in header_cells]

    colmap = {}
    for key, _, aliases in IMMOBILISATION_IMPORT_COLUMNS:
        for i, h in enumerate(headers):
            if h in aliases:
                colmap[key] = i
                break

    if "compte" not in colmap:
        raise ValueError(
            "Colonne obligatoire introuvable (« N° Compte »). Utilisez le bouton « Télécharger un modèle »."
        )

    def get(values, key, default=None):
        idx = colmap.get(key)
        if idx is None or idx >= len(values):
            return default
        return values[idx]

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [c.value for c in row]
        if all(v in (None, "") for v in values):
            continue
        compte = str(get(values, "compte") or "").strip()
        date_brute = get(values, "date_acquisition")
        date_str = None
        if isinstance(date_brute, str):
            date_str = date_brute
        elif hasattr(date_brute, "strftime"):
            date_str = date_brute.strftime("%d/%m/%Y")
        rows.append({
            "ligne": row_idx, "compte": compte,
            "fournisseur_code": str(get(values, "fournisseur_code") or "").strip(),
            "prix_achat": get(values, "prix_achat"),
            "date_acquisition": date_str,
            "base_repartition_quantite": get(values, "base_repartition_quantite"),
            "base_repartition_unite": str(get(values, "base_repartition_unite") or "").strip(),
            "amortissement_annuel_manuel": get(values, "amortissement_annuel_manuel"),
        })
    return rows


def apply_immobilisations_rows(conn, rows):
    """Applique en base une liste de lignes déjà lues (voir
    parse_immobilisations_xlsx) : valide chaque compte/fournisseur et met à
    jour la fiche d'immobilisation correspondante — les comptes ou
    fournisseurs introuvables sont ignorés avec un avertissement plutôt que
    de faire échouer tout l'import. Utilisable en RPC par le client réseau."""
    imported, warnings = 0, []
    for r in rows:
        row_idx = r.get("ligne", "?")
        compte = (r.get("compte") or "").strip()
        if not compte:
            warnings.append(f"Ligne {row_idx} : numéro de compte manquant, ligne ignorée.")
            continue
        if not account_exists(conn, compte):
            warnings.append(
                f"Ligne {row_idx} : le compte « {compte} » n'existe pas dans le Plan comptable, ligne ignorée."
            )
            continue
        fournisseur_code = (r.get("fournisseur_code") or "").strip() or None
        if fournisseur_code and not fournisseur_exists(conn, fournisseur_code):
            warnings.append(f"Ligne {row_idx} : le fournisseur « {fournisseur_code} » n'existe pas — laissé vide.")
            fournisseur_code = None
        prix_achat = None
        if r.get("prix_achat") not in (None, ""):
            try:
                prix_achat = float(r["prix_achat"]) or None
            except (TypeError, ValueError):
                warnings.append(f"Ligne {row_idx} : prix d'achat invalide, ignoré.")
        date_acquisition = None
        if r.get("date_acquisition"):
            date_acquisition = to_iso_date(r["date_acquisition"])
            if not date_acquisition:
                warnings.append(f"Ligne {row_idx} : date d'acquisition invalide, ignorée.")
        base_qte = None
        if r.get("base_repartition_quantite") not in (None, ""):
            try:
                base_qte = float(r["base_repartition_quantite"])
            except (TypeError, ValueError):
                warnings.append(f"Ligne {row_idx} : base de répartition invalide, ignorée.")
        base_unite = (r.get("base_repartition_unite") or "").strip() or None
        amort_manuel = None
        if r.get("amortissement_annuel_manuel") not in (None, ""):
            try:
                amort_manuel = float(r["amortissement_annuel_manuel"])
            except (TypeError, ValueError):
                warnings.append(f"Ligne {row_idx} : amortissement annuel invalide, ignoré.")
        set_immobilisation_fiche(
            conn, compte, fournisseur_code=fournisseur_code, prix_achat=prix_achat,
            date_acquisition=date_acquisition, base_repartition_quantite=base_qte,
            base_repartition_unite=base_unite, amortissement_annuel_manuel=amort_manuel,
        )
        imported += 1
    return imported, warnings


def import_immobilisations_from_xlsx(conn, path):
    """Import direct depuis un fichier local (usage bureau uniquement — le
    fichier et la base sont sur la même machine) : lit puis applique en un
    seul appel. Le client réseau utilise séparément
    parse_immobilisations_xlsx (localement) puis apply_immobilisations_rows
    (via RPC), car le fichier Excel n'est pas accessible au serveur."""
    rows = parse_immobilisations_xlsx(path)
    return apply_immobilisations_rows(conn, rows)


# ---------------------------------------------------------------------------
# Balance âgée des créances clients (menu COMMERCE > Recouvrement) — répartit
# le montant des factures NON PAYÉES (table factures_clients) par tranche
# d'ancienneté, selon des seuils (en jours) choisis par l'utilisateur.
# ---------------------------------------------------------------------------
def compute_balance_agee(conn, seuils=(30, 60, 90), date_reference=None):
    """Balance âgée par client : pour chaque client ayant au moins une
    facture non payée (date_paiement_reel vide), répartit le montant de ces
    factures par tranche d'ancienneté (jours écoulés entre la date de
    facture et `date_reference` — aujourd'hui par défaut), selon `seuils`
    (ex. (30,60,90) -> tranches « 0-30 », « 31-60 », « 61-90 », « >90 »).
    Chaque client renvoyé porte aussi le détail facture par facture
    (`factures`), pour l'écran de détail (double-clic)."""
    date_reference = date_reference or date.today().strftime("%Y-%m-%d")
    try:
        ref = datetime.strptime(date_reference, "%Y-%m-%d")
    except ValueError:
        ref = datetime.today()

    rows = conn.execute(
        """SELECT f.*, COALESCE(c.raison_sociale, f.client_code) AS raison_sociale
           FROM factures_clients f LEFT JOIN clients c ON c.code = f.client_code
           WHERE f.date_paiement_reel IS NULL OR f.date_paiement_reel = ''
           ORDER BY f.client_code, f.date_facture"""
    ).fetchall()

    par_client = {}
    for r in rows:
        try:
            d = datetime.strptime(r["date_facture"], "%Y-%m-%d")
            age = (ref - d).days
        except (TypeError, ValueError):
            age = 0
        client_code = r["client_code"]
        if client_code not in par_client:
            par_client[client_code] = {
                "client_code": client_code, "raison_sociale": r["raison_sociale"],
                "tranches": [0.0] * (len(seuils) + 1), "total": 0.0, "factures": [],
            }
        bucket_idx = len(seuils)
        for i, s in enumerate(seuils):
            if age <= s:
                bucket_idx = i
                break
        entry = par_client[client_code]
        entry["tranches"][bucket_idx] += r["montant"]
        entry["total"] += r["montant"]
        entry["factures"].append({
            "id": r["id"], "piece": r["piece"] or "", "libelle": r["libelle"] or "",
            "montant": r["montant"], "date_facture": r["date_facture"], "age_jours": age,
            "tranche": bucket_idx,
        })
    return sorted(par_client.values(), key=lambda c: -c["total"])


def compute_arrete_comptes(conn, date_arrete=None, exercice=None):
    """Tableau de vérification avant clôture (« arrêté de comptes ») —
    rassemble en un seul appel les contrôles habituels de fin de période :
    comptes fournisseurs (soldes anormaux), factures fournisseurs non
    parvenues (livraison reçue sans facture correspondante), balance âgée
    clients, rapprochements bancaires (écart pointé/comptable), impôts et
    charges sociales (soldes 43x/44x/447x), engagements en retard
    (livraison/paiement), et statut de la paie de la dernière période
    saisie. Ne modifie rien — uniquement un état des lieux à une date
    donnée (aujourd'hui par défaut)."""
    date_arrete = date_arrete or date.today().strftime("%Y-%m-%d")
    exercice = exercice or get_current_exercice(conn)

    # 1. Comptes fournisseurs : solde de chaque compte 40x, anomalie si
    #    solde DÉBITEUR (un fournisseur ne devrait normalement pas nous
    #    devoir de l'argent, sauf avance versée).
    balance = compute_balance(conn, only_with_movement=True, exercice=exercice)
    fournisseurs_comptes = [b for b in balance if b["code"].startswith("40")]
    fournisseurs_anomalies = [b for b in fournisseurs_comptes if b["solde_cloture"] > 0.01]

    # 2. Factures fournisseurs non parvenues : commande livrée
    #    (date_livraison_reelle renseignée) mais AUCUNE facture d'achat
    #    n'a jamais été saisie pour ce fournisseur depuis la commande —
    #    signal à vérifier manuellement, pas une détection garantie à 100%
    #    (aucun lien direct commande <-> facture dans le modèle de données).
    commandes = list_commandes(conn)
    factures_achat_par_fournisseur = {}
    for f in list_factures_achat(conn):
        factures_achat_par_fournisseur.setdefault(f["fournisseur_code"], []).append(f["date_facture"])
    factures_non_parvenues = []
    for c in commandes:
        if not c["date_livraison_reelle"]:
            continue
        dates_factures = factures_achat_par_fournisseur.get(c["fournisseur_code"], [])
        a_une_facture_apres = any(d and d >= c["date_commande"] for d in dates_factures)
        if not a_une_facture_apres:
            factures_non_parvenues.append(c)

    # 3. Balance âgée clients (impayés)
    clients_balance_agee = compute_balance_agee(conn, date_reference=date_arrete)

    # 4. Rapprochements bancaires : pour chaque compte 52x, solde
    #    comptable vs total des mouvements pointés — l'écart doit
    #    correspondre aux mouvements non encore retrouvés sur le relevé.
    banques = compute_comptes_prefixe_periode(conn, "52", date_to=date_arrete, exercice=exercice)
    rapprochements = [{
        "compte": b["code"], "libelle": b["label"], "solde_comptable": b["solde_fin_periode"],
    } for b in banques]

    # 5. Impôts et charges sociales : soldes des comptes TVA (443/444),
    #    IUTS/retenues (447), CNSS (43) à la date d'arrêté.
    impots = (compute_comptes_prefixe_periode(conn, "443", date_to=date_arrete, exercice=exercice)
              + compute_comptes_prefixe_periode(conn, "444", date_to=date_arrete, exercice=exercice)
              + compute_comptes_prefixe_periode(conn, "447", date_to=date_arrete, exercice=exercice))
    charges_sociales = compute_comptes_prefixe_periode(conn, "43", date_to=date_arrete, exercice=exercice)

    # 6. Situation des engagements : commandes fournisseurs et factures
    #    clients en dépassement de délai (livraison ou paiement).
    engagements_fournisseurs_retard = [c for c in commandes
                                        if c["depassement_livraison"] or c["depassement_paiement"]]
    factures_clients_retard = [f for f in list_factures(conn) if f["depassement"]]

    # 7. Statut de la paie : les 3 dernières périodes distinctes saisies,
    #    validées ou non.
    periodes = sorted({b["periode"] for b in list_bulletins_paie(conn)}, reverse=True)[:3]
    paie_statuts = [{"periode": p, "validee": est_periode_paie_validee(conn, p)} for p in periodes]

    return {
        "date_arrete": date_arrete, "exercice": exercice,
        "fournisseurs": {"comptes": fournisseurs_comptes, "anomalies": fournisseurs_anomalies},
        "factures_non_parvenues": factures_non_parvenues,
        "clients_balance_agee": clients_balance_agee,
        "rapprochements_bancaires": rapprochements,
        "impots": impots, "charges_sociales": charges_sociales,
        "engagements_fournisseurs_retard": engagements_fournisseurs_retard,
        "factures_clients_retard": factures_clients_retard,
        "paie_statuts": paie_statuts,
    }


# ---------------------------------------------------------------------------
# Synchronisation (menu PARAMÈTRES) — revérifie/répare la structure de
# toutes les tables de la base (utile après une mise à jour du logiciel qui
# a ajouté de nouvelles tables/colonnes à une base existante plus ancienne ;
# ces migrations tournent déjà automatiquement à chaque démarrage via
# _migrate(), ce bouton les rejoue explicitement à la demande et confirme
# l'état de la base).
# ---------------------------------------------------------------------------
def synchroniser_base(conn):
    """Rejoue init_db()/_migrate() sur la connexion en cours (crée toute
    table/colonne manquante, sans jamais toucher aux données existantes) et
    renvoie un petit rapport (nombre de tables, exercice courant, écart du
    Bilan) pour confirmer visuellement que tout est à jour."""
    init_db(conn)
    tables = [r["name"] for r in conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' ORDER BY name"
    ).fetchall()]
    exercice = get_current_exercice(conn)
    try:
        bilan = compute_bilan(conn, exercice=exercice)
        ecart = bilan["ecart"]
    except Exception:
        ecart = None
    return {"nb_tables": len(tables), "tables": tables, "exercice": exercice, "ecart_bilan": ecart}


# ---------------------------------------------------------------------------
# Utilisateurs et niveaux d'accès (menu ADMIN) — gestion des comptes et de
# leur niveau d'accès (paramétrable). L'application ne demande pas encore
# de connexion au démarrage : cet écran pose la base (comptes, mots de
# passe, niveaux) en vue d'un futur contrôle d'accès par écran/action.
# ---------------------------------------------------------------------------
# ---------------------------------------------------------------------------
# STRUCTURE CANONIQUE DES MENUS — source UNIQUE de vérité pour :
# (1) l'écran ADMIN > Niveaux d'accès (case à cocher par sous-menu),
# (2) le filtrage réel des menus affichés selon le niveau d'accès connecté.
# Toute modification de la structure des menus dans main.py (add_top_menu)
# doit être répercutée ici pour rester synchronisée.
# ---------------------------------------------------------------------------
MENU_STRUCTURE = [
    ("SAISIE", [("Saisie des écritures", "saisie"), ("Soldes d'ouverture", "ouverture")]),
    ("COMMERCIAL", [("Clients", "clients"), ("Recouvrement", "recouvrement"), ("Facturation", "facturation"),
                  ("Stocks", "stocks"), ("Marges bénéficiaires", "marges")]),
    ("PRODUCTION", [("Matières premières", "stocks"), ("Fabrication", "production"),
                     ("Produits finis", "stocks")]),
    ("RAPPORTS FINANCIERS", [("Grand livre", "grand_livre"), ("Balance", "balance"),
                             ("Bilan SYSCOHADA", "bilan_syscohada"),
                             ("Compte de résultat (SIG)", "compte_resultat_sig"), ("TFT", "tft"),
                             ("Situation financière", "situation_financiere"),
                             ("Arrêté de comptes", "arrete_comptes")]),
    ("ENGAGEMENTS-PROJETS", [("Fournisseurs", "fournisseurs"), ("Contrats", "contrats"),
                              ("Expression de besoin", "expression_besoin"),
                              ("Bon de commande", "ep_bon_commande"),
                              ("Bordereau de livraison", "bordereau_livraison"),
                              ("Règlements", "reglements")]),
    ("GRH", [("Liste du personnel", "grh_personnel"), ("Time sheet", "grh_time_sheet"), ("KPI", "grh_kpi"),
             ("Tableau de bord GRH", "grh_tableau_bord"), ("HS (hygiène santé)", "grh_hs"),
             ("Paie", "grh_paie")]),
    ("TRESORERIE", [("Trésorerie", "tresorerie")]),
    ("TRANSPORT", [("Parc auto", "transport"), ("Missions", "missions"),
                    ("Pièces de rechange", "pieces_rechange"), ("Réparations", "reparations")]),
    ("IMMOBILISATIONS", [("Immobilisations", "immobilisations"), ("Amortissements", "amortissements")]),
    ("RAPPORTS TECHNIQUES", [("Rapports technique", "rapports_technique")]),
    ("MAINTENANCE-QUALITÉ", [("Énergie", "energie"), ("Maintenance", "maintenance"),
                              ("Pièces de rechange", "pieces_rechange")]),
    ("PARAMÈTRES", [("Exercices comptables (clôture)", "exercices"), ("Plan comptable", "plan_comptable"),
                     ("Plan analytique", "plan_analytique"), ("Plan budgétaire", "plan_budgetaire"),
                     ("Plan bailleurs de fonds", "plan_bailleur"), ("Synchronisation", "synchronisation")]),
    ("ADMIN", [("Taux de TVA", "taux_tva"), ("Taux de retenue à la source", "taux_retenue"),
               ("Modification des factures", "admin_factures"),
               ("Modèle de bon de commande", "admin_modele_bon_commande"),
               ("Niveaux d'accès", "niveaux_acces"), ("Utilisateurs", "utilisateurs"),
               ("Réinitialisation des données", "reinitialisation")]),
]


def get_menus_autorises(conn, niveau_acces):
    """Renvoie l'ensemble des clés de sous-menu (ex. "saisie", "bilan_syscohada")
    autorisées pour un niveau d'accès donné. Le niveau « Administrateur »
    a TOUJOURS accès à tout (garde-fou — même si la table est vide ou mal
    configurée, jamais de risque de verrouillage total de
    l'administration)."""
    if niveau_acces == "Administrateur":
        return {key for _titre, items in MENU_STRUCTURE for _label, key in items}
    rows = conn.execute("SELECT menu_key FROM niveau_acces_menus WHERE niveau_acces = ?", (niveau_acces,))
    return {r["menu_key"] for r in rows}


def set_menus_autorises(conn, niveau_acces, menu_keys):
    """Remplace la liste des sous-menus autorisés pour un niveau d'accès."""
    conn.execute("DELETE FROM niveau_acces_menus WHERE niveau_acces = ?", (niveau_acces,))
    conn.executemany("INSERT OR IGNORE INTO niveau_acces_menus (niveau_acces, menu_key) VALUES (?, ?)",
                      [(niveau_acces, k) for k in menu_keys])
    conn.commit()


def ajouter_niveaux_acces_suggeres_menus(conn):
    """Préconfigure des ensembles de menus adaptés à chaque profil métier
    suggéré (Administrateur/Comptable/Vendeur/Chargé des achats/GRH/
    Trésorier/Usine) — appelée une fois après
    ajouter_niveaux_acces_suggeres(), UNIQUEMENT si aucune configuration de
    menus n'existe encore pour ces niveaux (n'écrase jamais une
    configuration déjà personnalisée par l'utilisateur)."""
    tous = [key for _titre, items in MENU_STRUCTURE for _label, key in items]

    comptable_menus = ["saisie", "ouverture", "grand_livre", "balance", "bilan_syscohada",
                        "compte_resultat_sig", "tft", "situation_financiere", "arrete_comptes",
                        "tresorerie", "exercices", "plan_comptable", "plan_analytique",
                        "plan_budgetaire", "plan_bailleur", "synchronisation"]
    vendeur_menus = ["clients", "recouvrement", "facturation", "stocks", "marges"]
    charge_achats_menus = ["fournisseurs", "contrats", "expression_besoin", "ep_bon_commande",
                            "bordereau_livraison", "reglements"]
    grh_menus = ["grh_personnel", "grh_time_sheet", "grh_kpi", "grh_tableau_bord", "grh_hs", "grh_paie"]
    tresorier_menus = ["tresorerie", "recouvrement", "reglements"]
    usine_menus = ["stocks", "production", "transport", "missions", "pieces_rechange", "reparations",
                   "immobilisations", "amortissements", "energie", "maintenance", "rapports_technique"]

    for niveau, menus in (
        ("Administrateur", tous),
        ("Comptable", comptable_menus),
        ("Vendeur", vendeur_menus),
        ("Chargé des achats", charge_achats_menus),
        ("GRH", grh_menus),
        ("Trésorier", tresorier_menus),
        ("Usine", usine_menus),
    ):
        deja_configure = conn.execute(
            "SELECT 1 FROM niveau_acces_menus WHERE niveau_acces = ? LIMIT 1", (niveau,)
        ).fetchone()
        if not deja_configure:
            set_menus_autorises(conn, niveau, menus)


def list_niveaux_acces(conn):
    return [dict(r) for r in conn.execute("SELECT * FROM niveaux_acces ORDER BY nom").fetchall()]


def add_niveau_acces(conn, nom, description=""):
    conn.execute("INSERT OR REPLACE INTO niveaux_acces (nom, description) VALUES (?, ?)",
                 (nom.strip(), description.strip()))
    conn.commit()


def delete_niveau_acces(conn, nom):
    conn.execute("DELETE FROM niveaux_acces WHERE nom = ?", (nom,))
    conn.commit()


def niveau_acces_exists(conn, nom):
    return conn.execute("SELECT 1 FROM niveaux_acces WHERE nom = ?", (nom,)).fetchone() is not None


NIVEAUX_ACCES_SUGGERES = [
    ("Administrateur", "Accès complet à tous les menus, y compris ADMIN et PARAMÈTRES."),
    ("Comptable", "Saisie, soldes d'ouverture, tous les rapports financiers, trésorerie et paramètres comptables."),
    ("Vendeur", "Clients, recouvrement, facturation, stocks et marges bénéficiaires (menu COMMERCIAL)."),
    ("Chargé des achats", "Fournisseurs, contrats, expressions de besoin, bons de commande, bordereaux de "
                          "livraison et règlements (menu ENGAGEMENTS-PROJETS)."),
    ("GRH", "Personnel, time sheet, KPI, tableau de bord GRH et hygiène santé (menu GRH)."),
    ("Trésorier", "Trésorerie, recouvrement des créances clients et règlements des dettes fournisseurs."),
    ("Usine", "Production, transport, immobilisations, maintenance-qualité et rapports technique."),
]


def ajouter_niveaux_acces_suggeres(conn):
    ajoutes = 0
    for nom, description in NIVEAUX_ACCES_SUGGERES:
        if not niveau_acces_exists(conn, nom):
            add_niveau_acces(conn, nom, description)
            ajoutes += 1
    return ajoutes


def export_niveaux_acces_xlsx(conn, path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Niveaux d'accès"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for i, label in enumerate(["Nom du niveau", "Description"], start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    for r, row in enumerate(conn.execute("SELECT nom, description FROM niveaux_acces ORDER BY nom"), start=2):
        ws.cell(row=r, column=1, value=row["nom"])
        ws.cell(row=r, column=2, value=row["description"])
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 60
    wb.save(path)
    return path


def import_niveaux_acces_xlsx(conn, path):
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = []
    for r in ws.iter_rows(min_row=2):
        values = [c.value for c in r]
        if all(v in (None, "") for v in values):
            continue
        nom = str(values[0] or "").strip()
        description = str(values[1] or "").strip() if len(values) > 1 else ""
        if nom:
            rows.append((nom, description))
    if not rows:
        raise ValueError("Le fichier ne contient aucune ligne valide.")
    conn.execute("DELETE FROM niveaux_acces")
    conn.executemany("INSERT INTO niveaux_acces (nom, description) VALUES (?, ?)", rows)
    conn.commit()
    return len(rows)


def _hash_password(mot_de_passe, sel=None):
    sel = sel or secrets.token_hex(16)
    h = hashlib.sha256((sel + mot_de_passe).encode("utf-8")).hexdigest()
    return h, sel


def add_utilisateur(conn, nom_utilisateur, mot_de_passe, nom_complet="", niveau_acces="Lecture seule", actif=True):
    nom_utilisateur = nom_utilisateur.strip()
    if not nom_utilisateur:
        raise ValueError("Le nom d'utilisateur est obligatoire.")
    if conn.execute("SELECT 1 FROM utilisateurs WHERE nom_utilisateur = ?", (nom_utilisateur,)).fetchone():
        raise ValueError(f"L'utilisateur « {nom_utilisateur} » existe déjà.")
    if not mot_de_passe:
        raise ValueError("Le mot de passe est obligatoire.")
    h, sel = _hash_password(mot_de_passe)
    conn.execute(
        """INSERT INTO utilisateurs (nom_utilisateur, nom_complet, mot_de_passe_hash, sel, niveau_acces, actif,
                                      date_creation)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (nom_utilisateur, nom_complet.strip(), h, sel, niveau_acces, 1 if actif else 0,
         date.today().strftime("%Y-%m-%d")),
    )
    conn.commit()


def update_utilisateur(conn, user_id, nouveau_mot_de_passe=None, **fields):
    """Met à jour un utilisateur. Passer `nouveau_mot_de_passe` pour changer
    le mot de passe (recalcule un nouveau sel) ; les autres champs
    (nom_complet, niveau_acces, actif...) via `fields`."""
    if nouveau_mot_de_passe:
        h, sel = _hash_password(nouveau_mot_de_passe)
        fields["mot_de_passe_hash"] = h
        fields["sel"] = sel
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE utilisateurs SET {cols} WHERE id = ?", (*fields.values(), user_id))
    conn.commit()


def delete_utilisateur(conn, user_id):
    conn.execute("DELETE FROM utilisateurs WHERE id = ?", (user_id,))
    conn.commit()


def list_utilisateurs(conn):
    return [dict(r) for r in conn.execute(
        "SELECT id, nom_utilisateur, nom_complet, niveau_acces, actif, date_creation FROM utilisateurs "
        "ORDER BY nom_utilisateur"
    ).fetchall()]


def verify_password(conn, nom_utilisateur, mot_de_passe):
    """Vérifie un couple utilisateur/mot de passe — utilisable par un futur
    écran de connexion. Renvoie l'utilisateur (sans le hash) si valide,
    None sinon."""
    row = conn.execute("SELECT * FROM utilisateurs WHERE nom_utilisateur = ? AND actif = 1",
                        (nom_utilisateur.strip(),)).fetchone()
    if not row:
        return None
    h, _ = _hash_password(mot_de_passe, sel=row["sel"])
    if h != row["mot_de_passe_hash"]:
        return None
    d = dict(row)
    d.pop("mot_de_passe_hash", None)
    d.pop("sel", None)
    return d


# ---------------------------------------------------------------------------
# Réinitialisation ciblée des données (menu ADMIN) — la Synchronisation ne
# touche JAMAIS aux données (uniquement la structure des tables) ; ceci est
# l'outil explicite et destructif pour vider des catégories de données
# choisies. Chaque catégorie est indépendante des autres — supprimer les
# écritures comptables ne vide PAS automatiquement les soldes d'ouverture
# (table séparée), ni les modules sans lien avec la comptabilité (circuit
# d'engagements, transport...).
# ---------------------------------------------------------------------------
REINIT_CATEGORIES = {
    "entries": "Écritures comptables (Saisie)",
    "opening_balances": "Soldes d'ouverture",
    "immobilisations_fiche": "Fiches immobilisations (fournisseur / prix d'achat)",
    "engagements": "Circuit d'engagements (Expression de besoin, Bon de commande, "
                    "Bordereau de livraison, Règlements)",
    "factures": "Factures (vente, achat, recouvrement client)",
    "transport": "Transport (véhicules, missions, réparations, pièces de rechange)",
}


def reinitialiser_donnees(conn, categories, exercice=None):
    """Vide les catégories de données demandées (voir REINIT_CATEGORIES).
    `exercice=None` (par défaut) supprime TOUTES les années pour les
    catégories concernées par un exercice (écritures, soldes d'ouverture) ;
    passer un exercice précis pour ne vider que cette année-là. Retourne un
    dict {categorie: nb_lignes_supprimees}."""
    rapport = {}
    if "entries" in categories:
        if exercice:
            cur = conn.execute("DELETE FROM entries WHERE substr(date,1,4) = ?", (str(exercice),))
        else:
            cur = conn.execute("DELETE FROM entries")
        rapport["entries"] = cur.rowcount
    if "opening_balances" in categories:
        if exercice:
            cur = conn.execute("DELETE FROM opening_balances WHERE exercice = ?", (str(exercice),))
        else:
            cur = conn.execute("DELETE FROM opening_balances")
        rapport["opening_balances"] = cur.rowcount
    if "immobilisations_fiche" in categories:
        cur = conn.execute("DELETE FROM immobilisations_fiche")
        rapport["immobilisations_fiche"] = cur.rowcount
    if "engagements" in categories:
        n = 0
        for table in ("expression_besoin_lignes", "expressions_besoin", "ep_bon_commande_lignes",
                       "ep_bons_commande", "bordereau_livraison_lignes", "bordereaux_livraison",
                       "reglement_lignes", "reglements"):
            cur = conn.execute(f"DELETE FROM {table}")
            n += cur.rowcount
        rapport["engagements"] = n
    if "factures" in categories:
        n = 0
        for table in ("facture_vente_lignes", "factures_vente", "facture_achat_lignes", "factures_achat",
                       "factures_clients"):
            cur = conn.execute(f"DELETE FROM {table}")
            n += cur.rowcount
        rapport["factures"] = n
    if "transport" in categories:
        n = 0
        for table in ("reparation_lignes", "reparations", "missions", "vehicules", "pieces_rechange"):
            cur = conn.execute(f"DELETE FROM {table}")
            n += cur.rowcount
        rapport["transport"] = n
    conn.commit()
    return rapport


# ---------------------------------------------------------------------------
# GRH (Gestion des Ressources Humaines) — Liste du personnel, Time sheet,
# KPI, Tableau de bord GRH, HS (Hygiène Santé) — suivi opérationnel, sans
# lien avec la comptabilité (même principe que Transport/le circuit
# d'engagements).
# ---------------------------------------------------------------------------

# ---- Liste du personnel ----
def add_personnel(conn, nom, matricule="", prenom="", poste="", service="", date_embauche="",
                   telephone="", email="", salaire_base=0, statut="actif", notes=""):
    if not nom.strip():
        raise ValueError("Le nom est obligatoire.")
    cur = conn.execute(
        """INSERT INTO personnel (matricule, nom, prenom, poste, service, date_embauche, telephone, email,
                                   salaire_base, statut, notes)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (matricule.strip(), nom.strip(), prenom.strip(), poste.strip(), service.strip(), date_embauche,
         telephone.strip(), email.strip(), salaire_base or 0, statut, notes.strip()),
    )
    conn.commit()
    return cur.lastrowid


def update_personnel(conn, personnel_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE personnel SET {cols} WHERE id = ?", (*fields.values(), personnel_id))
    conn.commit()


def delete_personnel(conn, personnel_id):
    conn.execute("DELETE FROM personnel WHERE id = ?", (personnel_id,))
    conn.commit()


def get_personnel(conn, personnel_id):
    row = conn.execute("SELECT * FROM personnel WHERE id = ?", (personnel_id,)).fetchone()
    return dict(row) if row else None


def list_personnel(conn, actifs_only=False):
    q = "SELECT * FROM personnel"
    if actifs_only:
        q += " WHERE statut = 'actif'"
    q += " ORDER BY nom, prenom"
    return [dict(r) for r in conn.execute(q).fetchall()]


# ---------------------------------------------------------------------------
# Paie (GRH > Paie) — reproduit fidèlement les formules du calculateur
# « Paie Burkina » (CNSS, IUTS, exonérations, abattement CADRE/AUTRE...),
# adapté pour lire/écrire dans cette base plutôt que dans un fichier JSON
# séparé, et réutiliser les employés déjà saisis dans GRH > Personnel.
# ---------------------------------------------------------------------------

PAIE_DEFAULT_PARAMS = {
    "taux_cnss_salarie": 0.055,
    "plafond_cnss": 800000,
    "cnss_salariale_plafonnee": 44000,
    "taux_cnss_patronale": 0.16,
    "taux_tpa": 0.03,
    "taux_retenue_obligatoire": 0.01,

    "abattement_cadre": 0.2,
    "abattement_autre": 0.25,

    "taux_plafond_fiscal": 0.08,

    # indemnité: [taux_exonere, plafond_mensuel]
    "exo_logement": [0.2, 75000],
    "exo_fonction": [0.05, 50000],
    "exo_transport": [0.05, 30000],

    # tranches IUTS: [de, a, taux, montant_cumule_anterieur]. a=None -> et plus
    "bareme_iuts": [
        [0, 10000, 0.0, 0],
        [10000, 20000, 0.0, 0],
        [20000, 30000, 0.0, 0],
        [30000, 50000, 0.121, 0],
        [50000, 80000, 0.139, 2420],
        [80000, 120000, 0.157, 6590],
        [120000, 170000, 0.184, 12870],
        [170000, 250000, 0.217, 22070],
        [250000, None, 0.25, 39430],
    ],

    # réduction IUTS selon personnes à charge
    "reduction_charges": {"0": 1.0, "1": 0.92, "2": 0.9, "3": 0.88, "4+": 0.86},
}

PAIE_LIGNE_LABELS = {
    "salaire_base": "Salaire de base", "prime_anciennete": "Prime d'ancienneté",
    "heures_sup": "Heures supplémentaires", "sursalaire": "Sursalaire",
    "gratification": "Gratification", "indemnite_caisse": "Indemnité Caisse",
    "indemnite_logement": "Indemnité Logement", "indemnite_fonction": "Indemnité Fonction",
    "indemnite_transport": "Indemnité Transport", "personnes_a_charge": "Personnes à charge",
    "retenue_pret": "Retenue prêt/avance",
}


def _paie_round(x, ndigits=0):
    """Reproduit ROUND() d'Excel (arrondi arithmétique, pas 'banker's rounding')."""
    if ndigits == 0:
        return math.floor(x + 0.5) if x >= 0 else math.ceil(x - 0.5)
    factor = 10 ** ndigits
    return _paie_round(x * factor) / factor


def _paie_rounddown_to_hundred(x):
    """Reproduit ROUNDDOWN(x, -2) d'Excel : arrondi à la centaine inférieure."""
    return math.floor(x / 100.0) * 100.0


def _paie_charge_key(n):
    n = int(n)
    return "4+" if n >= 4 else str(n)


def _paie_iuts_brut(base_imposable, bareme):
    """Calcule l'IUTS brut par la formule en cascade du barème progressif."""
    x = base_imposable
    for de, a, taux, cumul in bareme:
        if a is None or x < a:
            return (x - de) * taux + cumul
    de, a, taux, cumul = bareme[-1]
    return (x - de) * taux + cumul


def _paie_exoneration_indemnite(taux, plafond, indemnite_versee, salaire_brut):
    """=IF(taux*Brut<=Indem, IF(taux*Brut<=Plafond, taux*Brut, Plafond),
             IF(Indem>=Plafond, Plafond, Indem))"""
    seuil = taux * salaire_brut
    if seuil <= indemnite_versee:
        return seuil if seuil <= plafond else plafond
    return plafond if indemnite_versee >= plafond else indemnite_versee


def get_paie_parametres(conn):
    """Paramètres de paie (taux CNSS, plafonds, barème IUTS...) — stockés en
    JSON dans les settings, modifiables par un administrateur dans
    GRH > Paie > Paramètres. Complète automatiquement les clés manquantes
    avec les valeurs par défaut si le fichier vient d'une version antérieure."""
    raw = get_text_setting(conn, "paie_parametres", "")
    params = copy.deepcopy(PAIE_DEFAULT_PARAMS)
    if raw:
        try:
            saved = json.loads(raw)
            params.update(saved)
        except (ValueError, TypeError):
            pass
    return params


def set_paie_parametres(conn, params):
    set_text_setting(conn, "paie_parametres", json.dumps(params, ensure_ascii=False))


def compute_bulletin_paie(bulletin, params):
    """Calcule un bulletin de paie complet à partir des éléments de gain
    saisis (`bulletin`, un dict) et des paramètres de paie (`params`, voir
    get_paie_parametres) — reproduit exactement les formules du classeur
    Excel de référence (CNSS plafonnée, plafond fiscal, abattement, IUTS à
    9 tranches, réduction pour charges de famille)."""
    F = bulletin.get("salaire_base", 0) or 0
    G = bulletin.get("prime_anciennete", 0) or 0
    H = bulletin.get("heures_sup", 0) or 0
    I = bulletin.get("sursalaire", 0) or 0
    J = bulletin.get("gratification", 0) or 0
    K = bulletin.get("indemnite_caisse", 0) or 0
    L = bulletin.get("indemnite_logement", 0) or 0
    M = bulletin.get("indemnite_fonction", 0) or 0
    N = bulletin.get("indemnite_transport", 0) or 0
    classification = bulletin.get("classification") or "AUTRE"

    O = F + G + H + I + J + K + L + M + N

    if O <= params["plafond_cnss"]:
        P = _paie_round(O * params["taux_cnss_salarie"])
    else:
        P = params["cnss_salariale_plafonnee"]

    Q = params["taux_plafond_fiscal"] * (F + G + H + I)
    R = O - Q if P >= Q else O - P
    R = _paie_round(R)

    base_abattement = F + G + H + I
    taux_abattement = params["abattement_cadre"] if classification == "CADRE" else params["abattement_autre"]
    S = _paie_round(taux_abattement * base_abattement)

    taux_log, plaf_log = params["exo_logement"]
    taux_fct, plaf_fct = params["exo_fonction"]
    taux_trp, plaf_trp = params["exo_transport"]
    T = _paie_exoneration_indemnite(taux_log, plaf_log, L, R)
    U = _paie_exoneration_indemnite(taux_fct, plaf_fct, M, R)
    V = _paie_exoneration_indemnite(taux_trp, plaf_trp, N, R)

    W = S + T + U + V
    X = _paie_rounddown_to_hundred(R - W)
    Y = int(bulletin.get("personnes_a_charge", 0) or 0)
    Z = _paie_iuts_brut(X, params["bareme_iuts"])
    reduction = params["reduction_charges"].get(_paie_charge_key(Y), 1.0)
    AA = _paie_round(Z * reduction)

    AB = O - P - AA
    AC = _paie_round(AB * params["taux_retenue_obligatoire"])
    AD = bulletin.get("retenue_pret", 0) or 0
    AE = AB - AC - AD

    AF = _paie_round(O * params["taux_tpa"])
    AG = _paie_round(O * params["taux_cnss_patronale"])
    AH = AF + AG
    AI = O + AH
    AJ = AG + P
    AK = AF + AA

    return {
        "salaire_base": F, "prime_anciennete": G, "heures_sup": H, "sursalaire": I,
        "gratification": J, "indemnite_caisse": K, "indemnite_logement": L,
        "indemnite_fonction": M, "indemnite_transport": N,
        "remuneration_totale": O, "cnss_salariale": P, "plafond_fiscal": Q, "salaire_brut": R,
        "abattement": S, "exo_logement": T, "exo_fonction": U, "exo_transport": V,
        "total_exonerations": W, "base_imposable": X, "personnes_a_charge": Y,
        "iuts_brut": Z, "iuts_net": AA, "salaire_net": AB, "retenue_obligatoire": AC,
        "retenue_pret": AD, "net_percu": AE, "tpa_patronale": AF, "cnss_patronale": AG,
        "total_charges_patronales": AH, "cout_total_employeur": AI, "cnss_total": AJ,
        "iuts_plus_tpa": AK,
    }


def est_periode_paie_validee(conn, periode):
    return conn.execute(
        "SELECT 1 FROM paie_periodes_validees WHERE periode = ?", (periode,)
    ).fetchone() is not None


def set_bulletin_paie(conn, personnel_id, periode, **champs):
    """Crée ou met à jour le bulletin de paie (éléments de gain saisis, PAS
    les montants calculés — recalculés à la volée par compute_bulletin_paie)
    d'un employé pour une période donnée (« AAAA-MM »)."""
    if not get_personnel(conn, personnel_id):
        raise ValueError(f"Employé ID {personnel_id} introuvable.")
    if not periode or len(periode) != 7 or periode[4] != "-":
        raise ValueError("Période invalide — format attendu : AAAA-MM (ex. 2026-08).")
    if est_periode_paie_validee(conn, periode):
        raise ValueError(
            f"La paie de la période {periode} a déjà été validée (comptabilisée) — "
            f"elle ne peut plus être modifiée."
        )
    existing = conn.execute(
        "SELECT * FROM paie_bulletins WHERE personnel_id = ? AND periode = ?", (personnel_id, periode)
    ).fetchone()
    valeurs = dict(existing) if existing else {}
    valeurs.update({k: v for k, v in champs.items() if v is not None})
    valeurs.setdefault("classification", "AUTRE")
    for champ in ("salaire_base", "prime_anciennete", "heures_sup", "sursalaire", "gratification",
                  "indemnite_caisse", "indemnite_logement", "indemnite_fonction", "indemnite_transport",
                  "personnes_a_charge", "retenue_pret"):
        valeurs.setdefault(champ, 0)
    conn.execute(
        """INSERT INTO paie_bulletins
               (personnel_id, periode, classification, salaire_base, prime_anciennete, heures_sup,
                sursalaire, gratification, indemnite_caisse, indemnite_logement, indemnite_fonction,
                indemnite_transport, personnes_a_charge, retenue_pret, date_saisie)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
           ON CONFLICT(personnel_id, periode) DO UPDATE SET
               classification = excluded.classification, salaire_base = excluded.salaire_base,
               prime_anciennete = excluded.prime_anciennete, heures_sup = excluded.heures_sup,
               sursalaire = excluded.sursalaire, gratification = excluded.gratification,
               indemnite_caisse = excluded.indemnite_caisse, indemnite_logement = excluded.indemnite_logement,
               indemnite_fonction = excluded.indemnite_fonction,
               indemnite_transport = excluded.indemnite_transport,
               personnes_a_charge = excluded.personnes_a_charge, retenue_pret = excluded.retenue_pret,
               date_saisie = excluded.date_saisie""",
        (personnel_id, periode, valeurs["classification"], valeurs["salaire_base"],
         valeurs["prime_anciennete"], valeurs["heures_sup"], valeurs["sursalaire"],
         valeurs["gratification"], valeurs["indemnite_caisse"], valeurs["indemnite_logement"],
         valeurs["indemnite_fonction"], valeurs["indemnite_transport"], valeurs["personnes_a_charge"],
         valeurs["retenue_pret"], date.today().isoformat()),
    )
    conn.commit()


def delete_bulletin_paie(conn, bulletin_id):
    row = conn.execute("SELECT periode FROM paie_bulletins WHERE id = ?", (bulletin_id,)).fetchone()
    if row and est_periode_paie_validee(conn, row["periode"]):
        raise ValueError(
            f"La paie de la période {row['periode']} a déjà été validée (comptabilisée) — "
            f"ce bulletin ne peut plus être supprimé."
        )
    conn.execute("DELETE FROM paie_bulletins WHERE id = ?", (bulletin_id,))
    conn.commit()


def get_bulletin_paie(conn, personnel_id, periode):
    row = conn.execute(
        "SELECT * FROM paie_bulletins WHERE personnel_id = ? AND periode = ?", (personnel_id, periode)
    ).fetchone()
    return dict(row) if row else None


def list_bulletins_paie(conn, periode=None):
    """Bulletins de paie (éléments de gain saisis), avec le nom/matricule de
    l'employé — filtrés sur une période (« AAAA-MM ») si fournie, sinon
    tous. Ne renvoie PAS les montants calculés : utiliser compute_paie_periode
    pour l'état de paie complet (calculé)."""
    q = """SELECT b.*, p.nom, p.prenom, p.matricule, p.poste
           FROM paie_bulletins b JOIN personnel p ON p.id = b.personnel_id"""
    params = []
    if periode:
        q += " WHERE b.periode = ?"
        params.append(periode)
    q += " ORDER BY p.nom, p.prenom"
    return [dict(r) for r in conn.execute(q, params).fetchall()]


def compute_paie_periode(conn, periode):
    """État de paie calculé pour tous les employés ayant un bulletin sur
    cette période — chaque ligne combine les éléments de gain saisis et
    tous les montants calculés (CNSS, IUTS, net perçu, coût employeur...),
    plus les totaux de la période."""
    params = get_paie_parametres(conn)
    bulletins = list_bulletins_paie(conn, periode)
    lignes = []
    totaux = {"net_percu": 0.0, "cnss_total": 0.0, "iuts_net": 0.0,
              "retenue_obligatoire": 0.0, "cout_total_employeur": 0.0}
    for b in bulletins:
        resultat = compute_bulletin_paie(b, params)
        resultat.update({
            "bulletin_id": b["id"], "personnel_id": b["personnel_id"],
            "nom": b["nom"], "prenom": b["prenom"], "matricule": b["matricule"] or "",
            "poste": b["poste"] or "", "classification": b["classification"], "periode": periode,
        })
        lignes.append(resultat)
        for k in totaux:
            totaux[k] += resultat[k]
    return {"periode": periode, "lignes": lignes, "totaux": totaux}


def dupliquer_bulletins_periode(conn, periode_source, periode_cible):
    """Duplique tous les bulletins de paie (éléments de gain saisis, hors
    montants calculés) d'une période vers une autre — pratique en début de
    mois pour repartir des mêmes chiffres que le mois précédent plutôt que
    tout ressaisir. Les bulletins déjà existants sur la période cible pour
    un même employé sont écrasés."""
    if periode_source == periode_cible:
        raise ValueError("La période source et la période cible doivent être différentes.")
    bulletins = list_bulletins_paie(conn, periode_source)
    if not bulletins:
        raise ValueError(f"Aucun bulletin trouvé pour la période {periode_source}.")
    for b in bulletins:
        set_bulletin_paie(
            conn, b["personnel_id"], periode_cible, classification=b["classification"],
            salaire_base=b["salaire_base"], prime_anciennete=b["prime_anciennete"],
            heures_sup=b["heures_sup"], sursalaire=b["sursalaire"], gratification=b["gratification"],
            indemnite_caisse=b["indemnite_caisse"], indemnite_logement=b["indemnite_logement"],
            indemnite_fonction=b["indemnite_fonction"], indemnite_transport=b["indemnite_transport"],
            personnes_a_charge=b["personnes_a_charge"], retenue_pret=b["retenue_pret"],
        )
    return len(bulletins)


PAIE_COMPTES = {
    "salaires": "661100",           # Appointements, salaires — débit (charge)
    "cnss": "431000",               # Sécurité sociale — crédit (salariale + patronale)
    "iuts": "447210",               # État, IUTS — crédit
    "retenue_obligatoire": "447220",  # Retenue obligatoire 1% salaire — crédit
    "avance_pret": "421000",        # Personnel, avances et acomptes — crédit (remboursement de prêt/avance)
    "remunerations_dues": "422000",  # Personnel, rémunérations dues — crédit (net à payer)
    "charges_cnss_patronale": "664100",  # Charges sociales sur rémunérations — débit
    "charges_tpa": "664200",        # Charges sociales (TPA) — débit
    "tpa_a_payer": "442810",        # TPA — crédit
}


def valider_paie_periode(conn, periode, date_str=None, piece=None):
    """Comptabilise la paie d'une période entière (tous les bulletins
    saisis) : pour chaque employé, une écriture regroupant la rémunération
    brute (débit 661100), les retenues salariales (crédit CNSS 431000,
    IUTS 447210, retenue obligatoire 447220, remboursement prêt 421000) et
    le net à payer (crédit 422000) ; puis les charges patronales (débit
    664100/664200, crédit CNSS 431000 / TPA 442810).

    Marque la période comme VALIDÉE : les bulletins ne peuvent plus être
    modifiés ni supprimés ensuite (cohérent avec le fonctionnement des
    factures). Lève une erreur si la période est vide ou déjà validée.

    Retourne (état de paie calculé, pièce comptable utilisée)."""
    if est_periode_paie_validee(conn, periode):
        raise ValueError(f"La paie de la période {periode} a déjà été validée.")
    etat = compute_paie_periode(conn, periode)
    if not etat["lignes"]:
        raise ValueError(f"Aucun bulletin saisi pour la période {periode} — rien à valider.")

    exercice = periode[:4]
    if date_str is None:
        annee, mois = int(periode[:4]), int(periode[5:7])
        dernier_jour = 31 if mois == 12 else (date(annee, mois + 1, 1) - timedelta(days=1)).day
        date_str = f"{annee:04d}-{mois:02d}-{dernier_jour:02d}"
    piece = piece or f"PAIE-{periode}"

    for l in etat["lignes"]:
        libelle = f"Paie {periode} — {l['nom']} {l['prenom'] or ''}".strip()
        O = l["remuneration_totale"]
        if O > 0:
            add_entry(conn, date_str, piece, "OD", PAIE_COMPTES["salaires"], "", libelle, O, 0)
        if l["cnss_salariale"] > 0:
            add_entry(conn, date_str, piece, "OD", PAIE_COMPTES["cnss"], "", libelle, 0, l["cnss_salariale"])
        if l["iuts_net"] > 0:
            add_entry(conn, date_str, piece, "OD", PAIE_COMPTES["iuts"], "", libelle, 0, l["iuts_net"])
        if l["retenue_obligatoire"] > 0:
            add_entry(conn, date_str, piece, "OD", PAIE_COMPTES["retenue_obligatoire"], "", libelle,
                      0, l["retenue_obligatoire"])
        if l["retenue_pret"] > 0:
            add_entry(conn, date_str, piece, "OD", PAIE_COMPTES["avance_pret"], "", libelle,
                      0, l["retenue_pret"])
        if l["net_percu"] > 0:
            add_entry(conn, date_str, piece, "OD", PAIE_COMPTES["remunerations_dues"], "", libelle,
                      0, l["net_percu"])
        # Charges patronales
        if l["cnss_patronale"] > 0:
            add_entry(conn, date_str, piece, "OD", PAIE_COMPTES["charges_cnss_patronale"], "",
                      f"Charges patronales CNSS — {libelle}", l["cnss_patronale"], 0)
            add_entry(conn, date_str, piece, "OD", PAIE_COMPTES["cnss"], "",
                      f"Charges patronales CNSS — {libelle}", 0, l["cnss_patronale"])
        if l["tpa_patronale"] > 0:
            add_entry(conn, date_str, piece, "OD", PAIE_COMPTES["charges_tpa"], "",
                      f"TPA patronale — {libelle}", l["tpa_patronale"], 0)
            add_entry(conn, date_str, piece, "OD", PAIE_COMPTES["tpa_a_payer"], "",
                      f"TPA patronale — {libelle}", 0, l["tpa_patronale"])

    conn.execute(
        "INSERT OR REPLACE INTO paie_periodes_validees (periode, date_validation, piece) VALUES (?, ?, ?)",
        (periode, date.today().isoformat(), piece),
    )
    conn.commit()
    return etat, piece


PAIE_IMPORT_COLUMNS = [
    ("matricule", "Matricule", ["matricule"]),
    ("periode", "Période (AAAA-MM)", ["période (aaaa-mm)", "periode", "période"]),
    ("classification", "Classification (CADRE/AUTRE)", ["classification (cadre/autre)", "classification"]),
    ("salaire_base", "Salaire de base", ["salaire de base", "salaire base"]),
    ("prime_anciennete", "Prime d'ancienneté", ["prime d'ancienneté", "prime anciennete"]),
    ("heures_sup", "Heures supplémentaires", ["heures supplémentaires", "heures sup"]),
    ("sursalaire", "Sursalaire", ["sursalaire"]),
    ("gratification", "Gratification", ["gratification"]),
    ("indemnite_caisse", "Indemnité Caisse", ["indemnité caisse", "indemnite caisse"]),
    ("indemnite_logement", "Indemnité Logement", ["indemnité logement", "indemnite logement"]),
    ("indemnite_fonction", "Indemnité Fonction", ["indemnité fonction", "indemnite fonction"]),
    ("indemnite_transport", "Indemnité Transport", ["indemnité transport", "indemnite transport"]),
    ("personnes_a_charge", "Personnes à charge", ["personnes à charge", "personnes a charge"]),
    ("retenue_pret", "Retenue prêt/avance", ["retenue prêt/avance", "retenue pret"]),
]


def export_paie_bulletins_template(path):
    """Modèle Excel pour importer/mettre à jour en masse les bulletins de
    paie — chaque ligne doit référencer un employé déjà saisi dans
    GRH > Personnel via son matricule."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bulletins de paie"
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for i, (_, label, _) in enumerate(PAIE_IMPORT_COLUMNS, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    example = ["EMP001", "2026-08", "AUTRE", 120000, 5000, 0, 0, 0, 0, 0, 0, 15000, 1, 0]
    for i, val in enumerate(example, start=1):
        ws.cell(row=2, column=i, value=val)
    for i, w in enumerate([14, 16, 20, 14, 16, 18, 12, 14, 16, 18, 18, 18, 16, 16], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def parse_paie_bulletins_xlsx(path):
    """Lit un fichier Excel de bulletins de paie et renvoie la liste des
    lignes sous forme de dicts, SANS toucher à la base de données —
    utilisable localement (bureau) ou à distance (client réseau)."""
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    header_cells = next(ws.iter_rows(min_row=1, max_row=1))
    headers = [str(c.value).strip().lower() if c.value is not None else "" for c in header_cells]

    colmap = {}
    for key, _, aliases in PAIE_IMPORT_COLUMNS:
        for i, h in enumerate(headers):
            if h in aliases:
                colmap[key] = i
                break

    if "matricule" not in colmap:
        raise ValueError(
            "Colonne obligatoire introuvable (« Matricule »). Utilisez le bouton « Télécharger un modèle »."
        )

    def get(values, key, default=None):
        idx = colmap.get(key)
        if idx is None or idx >= len(values):
            return default
        return values[idx]

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [c.value for c in row]
        if all(v in (None, "") for v in values):
            continue
        rows.append({
            "ligne": row_idx,
            "matricule": str(get(values, "matricule") or "").strip(),
            "periode": str(get(values, "periode") or "").strip(),
            "classification": (str(get(values, "classification") or "AUTRE").strip().upper() or "AUTRE"),
            "salaire_base": get(values, "salaire_base"),
            "prime_anciennete": get(values, "prime_anciennete"),
            "heures_sup": get(values, "heures_sup"),
            "sursalaire": get(values, "sursalaire"),
            "gratification": get(values, "gratification"),
            "indemnite_caisse": get(values, "indemnite_caisse"),
            "indemnite_logement": get(values, "indemnite_logement"),
            "indemnite_fonction": get(values, "indemnite_fonction"),
            "indemnite_transport": get(values, "indemnite_transport"),
            "personnes_a_charge": get(values, "personnes_a_charge"),
            "retenue_pret": get(values, "retenue_pret"),
        })
    return rows


def apply_paie_bulletins_rows(conn, rows):
    """Applique en base une liste de bulletins déjà lus (voir
    parse_paie_bulletins_xlsx) : associe chaque ligne à son employé via le
    matricule, valide la période, et enregistre le bulletin — les
    matricules introuvables ou périodes déjà validées sont ignorés avec un
    avertissement plutôt que de faire échouer tout l'import."""
    personnel_par_matricule = {
        (p["matricule"] or "").strip(): p for p in list_personnel(conn) if p["matricule"]
    }
    imported, warnings = 0, []
    for r in rows:
        row_idx = r.get("ligne", "?")
        matricule = (r.get("matricule") or "").strip()
        p = personnel_par_matricule.get(matricule)
        if not p:
            warnings.append(f"Ligne {row_idx} : matricule « {matricule} » introuvable dans GRH > Personnel, "
                             f"ligne ignorée.")
            continue
        periode = (r.get("periode") or "").strip()
        if len(periode) != 7 or periode[4] != "-":
            warnings.append(f"Ligne {row_idx} : période « {periode} » invalide (attendu AAAA-MM), ligne ignorée.")
            continue
        if r.get("classification") not in ("CADRE", "AUTRE"):
            warnings.append(f"Ligne {row_idx} : classification invalide — « AUTRE » utilisée par défaut.")
        try:
            champs = {"classification": r.get("classification") if r.get("classification") in ("CADRE", "AUTRE")
                      else "AUTRE"}
            for champ in ("salaire_base", "prime_anciennete", "heures_sup", "sursalaire", "gratification",
                          "indemnite_caisse", "indemnite_logement", "indemnite_fonction",
                          "indemnite_transport", "retenue_pret"):
                champs[champ] = float(r.get(champ) or 0)
            champs["personnes_a_charge"] = int(float(r.get("personnes_a_charge") or 0))
        except (TypeError, ValueError):
            warnings.append(f"Ligne {row_idx} : valeur numérique invalide, ligne ignorée.")
            continue
        try:
            set_bulletin_paie(conn, p["id"], periode, **champs)
            imported += 1
        except ValueError as exc:
            warnings.append(f"Ligne {row_idx} : {exc}")
    return imported, warnings


def render_bulletin_paie_html(conn, bulletin_id):
    """Bulletin de paie individuel en HTML imprimable (aperçu avant
    impression) — renvoie le contenu (str), sans écrire de fichier, pour un
    usage local (bureau) ET distant (client réseau)."""
    row = conn.execute(
        "SELECT b.*, p.nom, p.prenom, p.matricule, p.poste FROM paie_bulletins b "
        "JOIN personnel p ON p.id = b.personnel_id WHERE b.id = ?", (bulletin_id,)
    ).fetchone()
    if not row:
        raise ValueError("Bulletin introuvable.")
    b = dict(row)
    params = get_paie_parametres(conn)
    r = compute_bulletin_paie(b, params)

    societe_nom = get_company_value(conn, "societe_nom") or "(Dénomination non renseignée — ADMIN > Liasse fiscale)"
    societe_adresse = get_company_value(conn, "societe_adresse")
    societe_telephone = get_company_value(conn, "societe_telephone")

    gains_rows = "\n".join(
        f"<tr><td>{PAIE_LIGNE_LABELS[k]}</td><td style='text-align:right'>{r[k]:,.0f}</td></tr>"
        for k in ("salaire_base", "prime_anciennete", "heures_sup", "sursalaire", "gratification",
                   "indemnite_caisse", "indemnite_logement", "indemnite_fonction", "indemnite_transport")
        if r[k]
    )
    return f"""<!DOCTYPE html>
<html lang="fr"><head><meta charset="utf-8"><title>Bulletin de paie — {b['nom']} {b['prenom'] or ''}</title>
<style>
  body {{ font-family: 'Segoe UI', Arial, sans-serif; margin: 30px; color: #111; font-size: 13px; }}
  .toolbar {{ margin-bottom: 16px; }}
  .cadre-entreprise {{ border: 1px solid #000; padding: 8px 12px; text-align: center; margin-bottom: 10px; }}
  .cadre-entreprise .nom {{ font-weight: bold; font-size: 15px; text-transform: uppercase; }}
  h1 {{ font-size: 16px; text-align: center; margin: 10px 0; }}
  .identite {{ display: flex; justify-content: space-between; border: 1px solid #000; padding: 8px 12px;
               margin-bottom: 12px; }}
  table {{ width: 100%; border-collapse: collapse; margin-bottom: 12px; }}
  th, td {{ border: 1px solid #000; padding: 5px 10px; font-size: 12px; }}
  th {{ background: #eee; text-align: left; }}
  table.totaux td:first-child {{ font-weight: bold; }}
  table.totaux td:last-child {{ text-align: right; }}
  .net {{ font-size: 15px; font-weight: bold; text-align: right; border: 2px solid #000; padding: 8px 12px; }}
  @media print {{ .toolbar {{ display: none; }} }}
</style></head>
<body>
<div class="toolbar"><button onclick="window.print()">🖨️ Imprimer / Enregistrer en PDF</button>
  <span style="margin-left:10px;color:#595959;font-size:12px;">Aperçu avant impression — rien n'est encore imprimé.</span></div>

<div class="cadre-entreprise">
  <div class="nom">{societe_nom}</div>
  <div>{societe_adresse or ''}{' — Tél : ' + societe_telephone if societe_telephone else ''}</div>
</div>
<h1>BULLETIN DE PAIE — {b['periode']}</h1>
<div class="identite">
  <div><b>{b['nom']} {b['prenom'] or ''}</b><br>Matricule : {b['matricule'] or '—'}<br>Poste : {b['poste'] or '—'}</div>
  <div>Classification : {b['classification']}<br>Personnes à charge : {r['personnes_a_charge']}</div>
</div>

<table>
<tr><th>Éléments de gain</th><th style="width:140px">Montant</th></tr>
{gains_rows}
<tr><td><b>Rémunération totale</b></td><td style="text-align:right"><b>{r['remuneration_totale']:,.0f}</b></td></tr>
</table>

<table class="totaux">
<tr><td>CNSS salariale</td><td>{r['cnss_salariale']:,.0f}</td></tr>
<tr><td>Salaire brut</td><td>{r['salaire_brut']:,.0f}</td></tr>
<tr><td>Base imposable</td><td>{r['base_imposable']:,.0f}</td></tr>
<tr><td>IUTS net</td><td>{r['iuts_net']:,.0f}</td></tr>
<tr><td>Salaire net</td><td>{r['salaire_net']:,.0f}</td></tr>
<tr><td>Retenue obligatoire (1%)</td><td>{r['retenue_obligatoire']:,.0f}</td></tr>
<tr><td>Retenue prêt/avance</td><td>{r['retenue_pret']:,.0f}</td></tr>
</table>

<div class="net">NET PERÇU : {r['net_percu']:,.0f} F CFA</div>

<p style="margin-top:30px;font-size:11px;color:#595959;">
Ce bulletin de paie est établi conformément à la législation du travail en vigueur au Burkina Faso.
À conserver sans limitation de durée.</p>
</body></html>"""


# ---- Time sheet ----
def add_time_sheet(conn, personnel_id, date_pointage, heures, activite="", notes=""):
    if not get_personnel(conn, personnel_id):
        raise ValueError("Employé introuvable.")
    if heures is None or heures <= 0:
        raise ValueError("Le nombre d'heures doit être strictement positif.")
    cur = conn.execute(
        "INSERT INTO grh_time_sheet (personnel_id, date_pointage, heures, activite, notes) VALUES (?, ?, ?, ?, ?)",
        (personnel_id, date_pointage, heures, activite.strip(), notes.strip()),
    )
    conn.commit()
    return cur.lastrowid


def update_time_sheet(conn, ts_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE grh_time_sheet SET {cols} WHERE id = ?", (*fields.values(), ts_id))
    conn.commit()


def delete_time_sheet(conn, ts_id):
    conn.execute("DELETE FROM grh_time_sheet WHERE id = ?", (ts_id,))
    conn.commit()


def list_time_sheet(conn, personnel_id=None, date_from=None, date_to=None):
    q = """SELECT t.*, (COALESCE(p.prenom,'') || ' ' || p.nom) AS employe
           FROM grh_time_sheet t JOIN personnel p ON p.id = t.personnel_id WHERE 1=1"""
    params = []
    if personnel_id:
        q += " AND t.personnel_id = ?"
        params.append(personnel_id)
    if date_from:
        q += " AND t.date_pointage >= ?"
        params.append(date_from)
    if date_to:
        q += " AND t.date_pointage <= ?"
        params.append(date_to)
    q += " ORDER BY t.date_pointage DESC, t.id DESC"
    return [dict(r) for r in conn.execute(q, params).fetchall()]


# ---- KPI ----
def add_kpi(conn, indicateur, description="", personnel_id=None, service="", periode="",
            valeur_cible=0, valeur_realisee=0, unite="", statut="en_cours"):
    if not indicateur.strip():
        raise ValueError("Le nom de l'indicateur est obligatoire.")
    cur = conn.execute(
        """INSERT INTO grh_kpi (indicateur, description, personnel_id, service, periode, valeur_cible,
                                 valeur_realisee, unite, statut)
           VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""",
        (indicateur.strip(), description.strip(), personnel_id, service.strip(), periode.strip(),
         valeur_cible or 0, valeur_realisee or 0, unite.strip(), statut),
    )
    conn.commit()
    return cur.lastrowid


def update_kpi(conn, kpi_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE grh_kpi SET {cols} WHERE id = ?", (*fields.values(), kpi_id))
    conn.commit()


def delete_kpi(conn, kpi_id):
    conn.execute("DELETE FROM grh_kpi WHERE id = ?", (kpi_id,))
    conn.commit()


def list_kpi(conn):
    rows = conn.execute(
        """SELECT k.*, CASE WHEN k.personnel_id IS NOT NULL
                            THEN (COALESCE(p.prenom,'') || ' ' || p.nom) ELSE '' END AS employe
           FROM grh_kpi k LEFT JOIN personnel p ON p.id = k.personnel_id
           ORDER BY k.statut, k.indicateur"""
    ).fetchall()
    result = []
    for r in rows:
        d = dict(r)
        d["taux_realisation"] = (d["valeur_realisee"] / d["valeur_cible"] * 100) if d["valeur_cible"] else None
        result.append(d)
    return result


# ---- HS (Hygiène Santé) ----
def add_hs(conn, date_evenement, type_evenement="incident", personnel_id=None, description="",
           gravite="", statut="ouvert", notes=""):
    cur = conn.execute(
        """INSERT INTO grh_hs (personnel_id, date_evenement, type_evenement, description, gravite, statut, notes)
           VALUES (?, ?, ?, ?, ?, ?, ?)""",
        (personnel_id, date_evenement, type_evenement, description.strip(), gravite, statut, notes.strip()),
    )
    conn.commit()
    return cur.lastrowid


def update_hs(conn, hs_id, **fields):
    if not fields:
        return
    cols = ", ".join(f"{k} = ?" for k in fields)
    conn.execute(f"UPDATE grh_hs SET {cols} WHERE id = ?", (*fields.values(), hs_id))
    conn.commit()


def delete_hs(conn, hs_id):
    conn.execute("DELETE FROM grh_hs WHERE id = ?", (hs_id,))
    conn.commit()


def list_hs(conn):
    rows = conn.execute(
        """SELECT h.*, CASE WHEN h.personnel_id IS NOT NULL
                            THEN (COALESCE(p.prenom,'') || ' ' || p.nom) ELSE '' END AS employe
           FROM grh_hs h LEFT JOIN personnel p ON p.id = h.personnel_id
           ORDER BY h.date_evenement DESC, h.id DESC"""
    ).fetchall()
    return [dict(r) for r in rows]


# ---- Tableau de bord GRH ----
def compute_tableau_bord_grh(conn):
    """Synthèse GRH : effectifs, heures pointées (30 derniers jours), KPI
    (en cours / atteints / non atteints), HS (incidents ouverts, par
    gravité) — calculée à la volée à partir des tables ci-dessus."""
    personnel = list_personnel(conn)
    nb_actifs = sum(1 for p in personnel if p["statut"] == "actif")
    nb_total = len(personnel)

    date_limite = (date.today() - timedelta(days=30)).strftime("%Y-%m-%d")
    ts = list_time_sheet(conn, date_from=date_limite)
    total_heures_30j = sum(t["heures"] for t in ts)

    kpis = list_kpi(conn)
    nb_kpi_en_cours = sum(1 for k in kpis if k["statut"] == "en_cours")
    nb_kpi_atteints = sum(1 for k in kpis
                           if k["taux_realisation"] is not None and k["taux_realisation"] >= 100)
    nb_kpi_non_atteints = sum(1 for k in kpis
                               if k["taux_realisation"] is not None and k["taux_realisation"] < 100
                               and k["statut"] != "en_cours")

    hs = list_hs(conn)
    nb_hs_ouverts = sum(1 for h in hs if h["statut"] == "ouvert")
    hs_par_gravite = {}
    for h in hs:
        if h["statut"] == "ouvert":
            g = h["gravite"] or "Non précisée"
            hs_par_gravite[g] = hs_par_gravite.get(g, 0) + 1

    return {
        "nb_personnel_actif": nb_actifs, "nb_personnel_total": nb_total,
        "total_heures_30j": total_heures_30j,
        "nb_kpi_en_cours": nb_kpi_en_cours, "nb_kpi_atteints": nb_kpi_atteints,
        "nb_kpi_non_atteints": nb_kpi_non_atteints, "nb_kpi_total": len(kpis),
        "nb_hs_ouverts": nb_hs_ouverts, "nb_hs_total": len(hs), "hs_par_gravite": hs_par_gravite,
    }


# ---------------------------------------------------------------------------
# TRÉSORERIE (menu TRESORERIE) — banques alignées horizontalement (une
# colonne par compte de trésorerie, classe 5) avec Entrées/Sorties de la
# période, et liste des engagements non encore payés (Règlements validés
# mais dont le paiement bancaire n'a pas encore été comptabilisé — voir
# enregistrer_paiement_reglement()) pour évaluer la capacité à y faire face.
# ---------------------------------------------------------------------------
def compute_tresorerie_banques_horizontal(conn, date_from=None, date_to=None, exercice=None):
    """Une ligne par compte de trésorerie (classe 5), avec Solde début /
    Entrées / Sorties / Solde fin de la période choisie (exercice entier
    par défaut) — pensé pour un affichage banques-en-colonnes. Réutilise
    compute_comptes_prefixe_periode() (même moteur que Impôts/Déclarations
    sociales/Rapprochements bancaires)."""
    lignes = compute_comptes_prefixe_periode(conn, "5", date_from=date_from, date_to=date_to, exercice=exercice)
    total = {
        "solde_debut_periode": sum(l["solde_debut_periode"] for l in lignes),
        "debit_periode": sum(l["debit_periode"] for l in lignes),
        "credit_periode": sum(l["credit_periode"] for l in lignes),
        "solde_fin_periode": sum(l["solde_fin_periode"] for l in lignes),
    }
    return lignes, total


def compute_engagements_a_payer(conn):
    """Liste des Règlements VALIDÉS (charge/dette fournisseur déjà
    comptabilisée) dont le paiement bancaire n'a PAS encore été
    comptabilisé (paiement_comptabilise = 0) — les engagements financiers
    restant à honorer. Compare leur total au solde de trésorerie
    disponible (classe 5) pour évaluer la capacité à y faire face."""
    rows = conn.execute(
        """SELECT r.*, COALESCE(f.raison_sociale, r.fournisseur_code, '') AS raison_sociale
           FROM reglements r LEFT JOIN fournisseurs f ON f.code = r.fournisseur_code
           WHERE r.statut = 'validee' AND r.paiement_comptabilise = 0
           ORDER BY r.date_reglement"""
    ).fetchall()
    engagements = []
    total_engagements = 0.0
    for r in rows:
        totals = compute_reglement_totals(conn, r["id"])
        engagements.append({
            "reglement_id": r["id"], "numero": r["numero"], "date_reglement": r["date_reglement"],
            "fournisseur_code": r["fournisseur_code"], "raison_sociale": r["raison_sociale"],
            "net_a_payer": totals["net_a_payer"],
        })
        total_engagements += totals["net_a_payer"]

    balance = compute_balance(conn, only_with_movement=False)
    treso_disponible = sum(b["solde_cloture"] for b in balance if b["classe"] == "5")

    return {
        "engagements": engagements, "total_engagements": total_engagements,
        "treso_disponible": treso_disponible, "solde_apres_engagements": treso_disponible - total_engagements,
        "peut_faire_face": treso_disponible >= total_engagements,
    }


# ---------------------------------------------------------------------------
# Import Excel — Liste du personnel et Time sheet (GRH), avec modèle
# téléchargeable (mêmes en-têtes que ceux attendus à l'import).
# ---------------------------------------------------------------------------
def export_personnel_template_xlsx(path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Personnel"
    headers = ["Matricule", "Nom", "Prénom", "Poste", "Service", "Date d'embauche (JJ/MM/AAAA)",
               "Téléphone", "Email", "Statut (actif/congé/suspendu/parti)"]
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for i, label in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    ws.append(["EMP-001", "Kone", "Amadou", "Chef chantier", "Production", "15/01/2024",
               "70000000", "akone@exemple.com", "actif"])
    for i, w in enumerate([12, 16, 16, 20, 16, 24, 14, 24, 24], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def import_personnel_xlsx(conn, path):
    """Importe la Liste du personnel depuis un .xlsx (mêmes colonnes que
    export_personnel_template_xlsx). Un matricule déjà existant MET À JOUR
    la fiche ; sinon un nouvel employé est créé. Retourne
    {crees, mis_a_jour, erreurs}."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    existants = {p["matricule"]: p["id"] for p in list_personnel(conn) if p["matricule"]}
    crees, mis_a_jour, erreurs = 0, 0, []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [c.value for c in row]
        if all(v in (None, "") for v in values):
            continue
        matricule = str(values[0] or "").strip()
        nom = str(values[1] or "").strip()
        if not nom:
            erreurs.append(f"Ligne {i} : nom manquant, ignorée.")
            continue
        prenom = str(values[2] or "").strip() if len(values) > 2 else ""
        poste = str(values[3] or "").strip() if len(values) > 3 else ""
        service = str(values[4] or "").strip() if len(values) > 4 else ""
        date_embauche_raw = values[5] if len(values) > 5 else None
        date_embauche = _parse_import_date(date_embauche_raw)
        telephone = str(values[6] or "").strip() if len(values) > 6 else ""
        email = str(values[7] or "").strip() if len(values) > 7 else ""
        statut = str(values[8] or "actif").strip() if len(values) > 8 else "actif"
        try:
            if matricule and matricule in existants:
                update_personnel(conn, existants[matricule], nom=nom, prenom=prenom, poste=poste,
                                  service=service, date_embauche=date_embauche, telephone=telephone,
                                  email=email, statut=statut)
                mis_a_jour += 1
            else:
                new_id = add_personnel(conn, nom, matricule=matricule, prenom=prenom, poste=poste,
                                        service=service, date_embauche=date_embauche or "",
                                        telephone=telephone, email=email, statut=statut)
                if matricule:
                    existants[matricule] = new_id
                crees += 1
        except ValueError as exc:
            erreurs.append(f"Ligne {i} : {exc}")
    return {"crees": crees, "mis_a_jour": mis_a_jour, "erreurs": erreurs}


def export_time_sheet_template_xlsx(path):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Time sheet"
    headers = ["Matricule (doit exister dans la Liste du personnel)", "Date (JJ/MM/AAAA)", "Heures", "Activité"]
    header_font = Font(bold=True, color="FFFFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for i, label in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=label)
        c.font = header_font
        c.fill = header_fill
    ws.append(["EMP-001", "20/08/2026", 8, "Coulage béton"])
    for i, w in enumerate([36, 16, 10, 30], start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    wb.save(path)
    return path


def import_time_sheet_xlsx(conn, path):
    """Importe des pointages Time sheet depuis un .xlsx (mêmes colonnes que
    export_time_sheet_template_xlsx). Chaque ligne DOIT référencer un
    matricule déjà présent dans la Liste du personnel — sinon la ligne est
    ignorée et signalée. Retourne {crees, erreurs}."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    par_matricule = {p["matricule"]: p["id"] for p in list_personnel(conn) if p["matricule"]}
    crees, erreurs = 0, []
    for i, row in enumerate(ws.iter_rows(min_row=2), start=2):
        values = [c.value for c in row]
        if all(v in (None, "") for v in values):
            continue
        matricule = str(values[0] or "").strip()
        if not matricule or matricule not in par_matricule:
            erreurs.append(f"Ligne {i} : matricule « {matricule} » introuvable dans la Liste du personnel, ignorée.")
            continue
        date_str = _parse_import_date(values[1] if len(values) > 1 else None)
        if not date_str:
            erreurs.append(f"Ligne {i} : date manquante ou invalide, ignorée.")
            continue
        try:
            heures = float(values[2]) if len(values) > 2 and values[2] not in (None, "") else 0
        except (TypeError, ValueError):
            erreurs.append(f"Ligne {i} : heures invalides, ignorée.")
            continue
        activite = str(values[3] or "").strip() if len(values) > 3 else ""
        try:
            add_time_sheet(conn, par_matricule[matricule], date_str, heures, activite=activite)
            crees += 1
        except ValueError as exc:
            erreurs.append(f"Ligne {i} : {exc}")
    return {"crees": crees, "erreurs": erreurs}


def _parse_import_date(raw):
    """Convertit une cellule Excel (datetime, ou texte JJ/MM/AAAA ou
    AAAA-MM-JJ) en date ISO AAAA-MM-JJ — chaîne vide si non reconnaissable."""
    if raw in (None, ""):
        return ""
    if isinstance(raw, datetime):
        return raw.strftime("%Y-%m-%d")
    if isinstance(raw, date):
        return raw.strftime("%Y-%m-%d")
    s = str(raw).strip()
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return ""


if __name__ == "__main__":
    # Petit auto-test en ligne de commande (sans Tkinter).
    conn = get_connection(":memory:" if False else "test_core.db")
    print("Comptes chargés :", conn.execute("SELECT COUNT(*) FROM accounts").fetchone()[0])

    add_entry(conn, str(date.today()), "FA-0001", "AC", "601000", "", "Achat marchandises", 1000, 0)
    add_entry(conn, str(date.today()), "FA-0001", "AC", "445200", "", "TVA récupérable", 200, 0)
    add_entry(conn, str(date.today()), "FA-0001", "AC", "401000", "Ets Dupont", "Facture FA-0001", 0, 1200)
    add_entry(conn, str(date.today()), "FV-0001", "VE", "411000", "Société ABC", "Facture FV-0001", 1180, 0)
    add_entry(conn, str(date.today()), "FV-0001", "VE", "701000", "", "Vente marchandises", 0, 1180)

    d, c = totals_debit_credit(conn)
    print("Total débit / crédit :", d, c, "Équilibré :", d == c)

    print("\n--- Balance ---")
    for b in compute_balance(conn):
        print(b)

    print("\n--- Compte de résultat ---")
    cr = compute_compte_resultat(conn)
    print("Résultat net :", cr["resultat_net"])

    print("\n--- Bilan ---")
    bilan = compute_bilan(conn)
    print("Total actif :", bilan["total_actif"], "Total passif :", bilan["total_passif"], "Écart :", bilan["ecart"])

    print("\n--- TFT ---")
    print(compute_tft(conn))

    print("\n--- Grand livre (411000) ---")
    for r in compute_grand_livre(conn, "411000"):
        print(r)

    print("\n--- Stocks ---")
    for s in compute_stocks(conn):
        print(s)

    print("\n--- Production ---")
    print(compute_production(conn))

    conn.close()
    os.remove("test_core.db")
